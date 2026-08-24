"""Spreadsheet builders used by shipping export endpoints."""

import html
import os
import re
from copy import copy
from datetime import datetime
from io import BytesIO

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


PENDING_HEADERS = [
    '序号', '订单日期', '订单号', '产品', '口味', '数量',
    '收件人', '联系方式', '州', '城市', '地址', '邮编',
]

DPD_HEADERS = [
    '客户单号', '转单号', '运输方式', '目的国家', '收件人姓名', '州,省',
    '城市', '联系地址', '收件人电话', '收件人邮箱', '收件人邮编', '重量(KG)',
    '中文品名', '海关报关品名1', '配货信息1', '申报价值1', '申报币种 ',
    '申报品数量1', '申报单重（KG）', '代收货款', '币种', '是否COD', '货物品名',
]

DPD_COLUMN_WIDTHS = [
    18, 12, 12, 10, 22, 16, 16, 34, 18, 30, 14, 12,
    14, 18, 16, 14, 14, 16, 16, 14, 10, 12, 14,
]

FLAVOR_META_KEYS = {
    'pa_flavour', 'pa_flavor', 'flavour', 'flavor', 'pa_taste', 'taste',
    'pa_variant', 'variant', 'pa_smak', 'smak', 'pa_smaki', 'smaki',
}


def _safe_excel_text(value):
    """Normalize customer data before storing it as an explicit string cell."""
    return str(value or '').replace('\r', ' ').replace('\n', ' ').strip()


def _pending_item_flavor(item):
    for meta in item.get('meta_data') or []:
        if not isinstance(meta, dict):
            continue
        key = str(meta.get('key') or '').strip().lower()
        if key in FLAVOR_META_KEYS:
            value = meta.get('display_value') or meta.get('value') or ''
            return html.unescape(_safe_excel_text(value))
    return ''


def prepare_australia_pending_items(line_items):
    """Return product/flavor/quantity rows suitable for the pending AU export."""
    if not isinstance(line_items, list):
        return []

    prepared = []
    for item in line_items:
        if not isinstance(item, dict):
            continue
        flavor = _pending_item_flavor(item)
        parent_name = html.unescape(_safe_excel_text(item.get('parent_name')))
        product = parent_name or html.unescape(_safe_excel_text(item.get('name')))
        if not parent_name and flavor:
            # WooCommerce often appends " - FLAVOR" to the variation name.
            # Keep the base product in 产品 and the variation in 口味.
            product = re.sub(
                rf'\s*[-–—]\s*{re.escape(flavor)}\s*$', '', product,
                flags=re.IGNORECASE,
            ).strip() or product

        quantity = item.get('quantity', '')
        if quantity not in (None, ''):
            try:
                numeric = float(quantity)
                quantity = int(numeric) if numeric.is_integer() else numeric
            except (TypeError, ValueError):
                quantity = _safe_excel_text(quantity)

        prepared.append({
            'product': product,
            'flavor': flavor,
            'quantity': quantity,
        })

    # Keep first-seen product order, while placing all flavors of the same
    # product together so 产品 can be merged exactly like the reference image.
    grouped = {}
    product_order = []
    for item in prepared:
        key = item['product']
        if key not in grouped:
            grouped[key] = []
            product_order.append(key)
        grouped[key].append(item)
    return [item for product in product_order for item in grouped[product]]


def _set_text_cell(cell, value):
    cell.value = _safe_excel_text(value)
    cell.data_type = 's'
    cell.number_format = '@'


def _set_order_date_cell(cell, value):
    raw = _safe_excel_text(value)[:10]
    try:
        cell.value = datetime.strptime(raw, '%Y-%m-%d').date()
        cell.number_format = 'yyyy/m/d'
    except ValueError:
        _set_text_cell(cell, raw)


def build_australia_pending_workbook(rows):
    """Build the AU pending-order workbook using the supplied 12-field header.

    Order-level fields are merged vertically. Product is merged within each
    product group, while flavor and quantity remain one row per line item so no
    picking detail is lost.
    """
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = '澳洲未发货'
    worksheet.sheet_view.showGridLines = False

    thin_gray = Side(style='thin', color='B7B7B7')
    border = Border(left=thin_gray, right=thin_gray, top=thin_gray, bottom=thin_gray)
    center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left = Alignment(horizontal='left', vertical='center', wrap_text=True)

    for column, header in enumerate(PENDING_HEADERS, start=1):
        cell = worksheet.cell(row=1, column=column, value=header)
        cell.font = Font(name='Arial', size=10, bold=False, color='000000')
        cell.fill = PatternFill(fill_type='solid', fgColor='F2F2F2')
        cell.alignment = center
        cell.border = border
    worksheet.row_dimensions[1].height = 24

    row_number = 2
    for order_index, order in enumerate(rows, start=1):
        items = order.get('items') or [{
            'product': '', 'flavor': '', 'quantity': '',
        }]
        order_start = row_number
        product_spans = []
        current_product = None
        product_start = row_number

        for item in items:
            product = _safe_excel_text(item.get('product'))
            if current_product is None:
                current_product = product
                product_start = row_number
            elif product != current_product:
                product_spans.append((product_start, row_number - 1))
                current_product = product
                product_start = row_number

            values = [
                order_index,
                order.get('order_date'),
                _safe_excel_text(order.get('order_number')).lstrip('#'),
                product,
                item.get('flavor'),
                item.get('quantity'),
                order.get('customer_name'),
                order.get('phone'),
                order.get('state'),
                order.get('city'),
                order.get('street_address') or order.get('address'),
                order.get('postcode'),
            ]
            for column, value in enumerate(values, start=1):
                cell = worksheet.cell(row=row_number, column=column)
                cell.border = border
                cell.alignment = left if column == 11 else center
                cell.font = Font(name='Arial', size=10, color='000000')
                if column == 1:
                    cell.value = order_index
                    cell.number_format = '0'
                elif column == 2:
                    _set_order_date_cell(cell, value)
                elif column == 6 and isinstance(value, (int, float)):
                    cell.value = value
                    cell.number_format = '0'
                else:
                    _set_text_cell(cell, value)
            worksheet.row_dimensions[row_number].height = 22
            row_number += 1

        product_spans.append((product_start, row_number - 1))
        order_end = row_number - 1

        if order_end > order_start:
            for column in (1, 2, 3, 7, 8, 9, 10, 11, 12):
                worksheet.merge_cells(
                    start_row=order_start, start_column=column,
                    end_row=order_end, end_column=column,
                )
        for start_row, end_row in product_spans:
            if end_row > start_row:
                worksheet.merge_cells(
                    start_row=start_row, start_column=4,
                    end_row=end_row, end_column=4,
                )

    widths = [8, 13, 17, 30, 30, 10, 18, 18, 12, 18, 44, 12]
    for column, width in enumerate(widths, start=1):
        worksheet.column_dimensions[get_column_letter(column)].width = width

    last_row = max(1, row_number - 1)
    worksheet.freeze_panes = 'A2'
    worksheet.sheet_properties.pageSetUpPr.fitToPage = True
    worksheet.page_setup.orientation = 'landscape'
    worksheet.page_setup.fitToWidth = 1
    worksheet.page_setup.fitToHeight = 0
    worksheet.print_title_rows = '1:1'
    worksheet.print_area = f'A1:L{last_row}'

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output


def build_dpd_shipping_workbook(rows, template_path=None):
    """Build the Polish DPD COD workbook in the partner's A-W layout."""
    if template_path is None:
        template_path = os.path.join(
            os.path.dirname(os.path.abspath(__file__)),
            'assets', 'export_templates', 'dpd_poland_template.xlsx',
        )

    workbook = load_workbook(template_path)
    worksheet = workbook['下单模板']
    actual_headers = [worksheet.cell(1, column).value for column in range(1, 24)]
    if actual_headers != DPD_HEADERS:
        raise ValueError('DPD template headers do not match the expected A-W layout')

    if worksheet.max_row > 1:
        worksheet.delete_rows(2, worksheet.max_row - 1)

    text_columns = {
        1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11,
        13, 14, 15, 17, 21, 22, 23,
    }
    for row_number, item in enumerate(rows, start=2):
        values = [
            item.get('customer_number'),
            '',
            '2581',
            'PL',
            item.get('recipient_name'),
            item.get('province'),
            item.get('city'),
            item.get('address'),
            item.get('phone'),
            item.get('email'),
            item.get('postcode'),
            1,
            'wanju',
            'wanju',
            'wanju',
            1,
            'usd',
            1,
            1,
            float(item.get('cod_amount') or 0),
            'PLN',
            '是',
            'wanju',
        ]
        for column, value in enumerate(values, start=1):
            cell = worksheet.cell(row=row_number, column=column)
            if column in text_columns:
                _set_text_cell(cell, value)
            else:
                cell.value = value
        worksheet.cell(row=row_number, column=20).number_format = '0.00'

    last_row = max(1, len(rows) + 1)
    for column, width in enumerate(DPD_COLUMN_WIDTHS, start=1):
        worksheet.column_dimensions[get_column_letter(column)].width = width
    worksheet.auto_filter.ref = f'A1:W{last_row}'
    worksheet.freeze_panes = 'A2'
    worksheet.sheet_properties.pageSetUpPr.fitToPage = True
    worksheet.page_setup.orientation = 'landscape'
    worksheet.page_setup.fitToWidth = 1
    worksheet.page_setup.fitToHeight = 0
    worksheet.print_title_rows = '1:1'
    worksheet.print_area = f'A1:W{last_row}'

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output


def build_australia_shipping_workbook(rows, template_path=None):
    """Build an xlsx matching the logistics partner's 地址拆分 workbook."""
    if template_path is None:
        template_path = os.path.join(
            os.path.dirname(os.path.abspath(__file__)),
            'assets', 'export_templates', 'australia_shipping_template.xlsx',
        )

    workbook = load_workbook(template_path)
    worksheet = workbook['地址拆分']

    # Row 2 is a sanitized formatting sample retained from the partner's file.
    style_row = []
    for column in range(1, 9):
        source = worksheet.cell(row=2, column=column)
        style_row.append({
            'style': copy(source._style),
            'number_format': source.number_format,
            'alignment': copy(source.alignment),
        })
    if worksheet.max_row > 1:
        worksheet.delete_rows(2, worksheet.max_row - 1)
    if worksheet.max_column > 8:
        worksheet.delete_cols(9, worksheet.max_column - 8)

    for index, item in enumerate(rows, start=1):
        values = [
            index,
            _safe_excel_text(item.get('order_number')),
            _safe_excel_text(item.get('customer_name')),
            _safe_excel_text(item.get('phone')),
            _safe_excel_text(item.get('city')),
            _safe_excel_text(item.get('state')),
            _safe_excel_text(item.get('address')),
            _safe_excel_text(item.get('tracking_number')),
        ]
        row_number = index + 1
        for column, value in enumerate(values, start=1):
            cell = worksheet.cell(row=row_number, column=column, value=value)
            sample = style_row[column - 1]
            cell._style = copy(sample['style'])
            cell.alignment = copy(sample['alignment'])
            if column in (2, 3, 4, 5, 6, 7, 8):
                # Explicit string type protects leading zeroes / long tracking
                # numbers and prevents formula injection without displaying an
                # apostrophe in the logistics partner's sheet.
                cell.data_type = 's'
                cell.number_format = '@'
            else:
                cell.number_format = sample['number_format']

    worksheet.auto_filter.ref = f'A1:H{len(rows) + 1}'
    worksheet.freeze_panes = 'A2'
    worksheet.sheet_properties.pageSetUpPr.fitToPage = True
    worksheet.page_setup.orientation = 'landscape'
    worksheet.page_setup.fitToWidth = 1
    worksheet.page_setup.fitToHeight = 0
    worksheet.print_title_rows = '1:1'
    worksheet.print_area = f'A1:H{len(rows) + 1}'

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output
