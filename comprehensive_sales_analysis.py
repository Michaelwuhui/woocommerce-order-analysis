import sqlite3
import json
import pandas as pd
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

def create_comprehensive_sales_analysis():
    """创建综合销售分析报告，按照用户指定的Excel格式"""
    
    print("正在生成综合销售分析报告...")
    
    # 连接数据库
    conn = sqlite3.connect('woocommerce_orders.db')
    
    # 获取订单数据
    query = """
    SELECT id, date_created, source, line_items, total, status, 
           payment_method, shipping_lines
    FROM orders 
    WHERE line_items IS NOT NULL
    ORDER BY date_created
    """
    
    df = pd.read_sql_query(query, conn)
    conn.close()
    
    print(f"获取到 {len(df)} 条订单数据")
    
    # 解析产品数量
    print("正在解析产品数量和配送状态...")
    
    product_quantities = []
    product_types = []
    
    for index, row in df.iterrows():
        try:
            items = json.loads(row['line_items'])
            total_qty = sum(item.get('quantity', 0) for item in items)
            types_count = len(items)
            product_quantities.append(total_qty)
            product_types.append(types_count)
        except json.JSONDecodeError:
            product_quantities.append(0)
            product_types.append(0)
    
    df['产品数量'] = product_quantities
    df['产品种类数'] = product_types
    
    # 转换日期格式
    df['date_created'] = pd.to_datetime(df['date_created'])
    df['年月'] = df['date_created'].dt.to_period('M')
    
    # 分析配送状态
    df['配送状态'] = df['status'].apply(classify_delivery_status)
    
    # 按月份和网站分组统计
    print("正在生成按月份和网站的统计数据...")
    
    # 创建用户指定格式的数据
    analysis_data = []
    
    # 获取所有月份
    months = sorted(df['年月'].unique())
    
    for month in months:
        month_data = df[df['年月'] == month]
        
        # 按网站分组
        for source in ['https://www.strefajednorazowek.pl', 'https://www.buchmistrz.pl']:
            site_data = month_data[month_data['source'] == source]
            
            if len(site_data) == 0:
                continue
                
            # 计算各项指标
            total_sales = site_data['total'].sum()
            order_count = len(site_data)
            product_quantity = site_data['产品数量'].sum()
            
            # 配送状态统计
            completed_orders = len(site_data[site_data['配送状态'] == '成功签收'])
            shipped_not_delivered = len(site_data[site_data['配送状态'] == '已发货未签收'])
            out_of_stock = len(site_data[site_data['配送状态'] == '缺货'])
            
            # 网站名称简化
            site_name = 'strefajednorazowek.pl' if 'strefajednorazowek' in source else 'buchmistrz.pl'
            
            analysis_data.append({
                '月份': str(month),
                '网站': site_name,
                '总销售': total_sales,
                '订单数量': order_count,
                '产品数量': product_quantity,
                '成功签收': completed_orders,
                '发货未签收': shipped_not_delivered,
                '订单数量_2': order_count,  # 重复列
                '缺货': out_of_stock,
                '订单数量_3': order_count   # 重复列
            })
    
    # 创建DataFrame
    analysis_df = pd.DataFrame(analysis_data)
    
    # 生成文件名
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    filename = f'comprehensive_sales_analysis_{timestamp}.xlsx'
    
    print(f"正在导出综合分析报告到: {filename}")
    
    # 创建Excel文件
    with pd.ExcelWriter(filename, engine='openpyxl') as writer:
        # 创建用户指定格式的表格
        create_user_format_sheet(writer, analysis_df)
        
        # 创建详细数据表
        analysis_df.to_excel(writer, sheet_name='详细数据', index=False)
        
        # 创建月度汇总表
        create_monthly_summary_sheet(writer, df)
        
        # 创建产品分析表
        create_product_analysis_sheet(writer, df)
        
        # 格式化表格
        format_analysis_sheets(writer)
    
    # 显示报告摘要
    print_analysis_summary(analysis_df, df)
    
    return filename

def classify_delivery_status(status):
    """分类配送状态"""
    status_lower = status.lower()
    
    if status_lower in ['completed', 'delivered', 'finished']:
        return '成功签收'
    elif status_lower in ['shipped', 'processing', 'on-hold']:
        return '已发货未签收'
    elif status_lower in ['cancelled', 'refunded', 'failed']:
        return '缺货'
    else:
        return '其他'

def create_user_format_sheet(writer, analysis_df):
    """创建用户指定格式的表格"""
    
    # 创建表格数据，按照用户提供的格式
    table_data = []
    
    # 添加标题行
    table_data.append(['', '', '9月份销售数据', '', '', '', '', '', '', ''])
    table_data.append(['', '总销售', '订单数量', '产品数量', '成功签收', '订单数量', '发货未签收', '订单数量', '缺货', '订单数量'])
    
    # 按月份分组数据
    months = sorted(analysis_df['月份'].unique())
    
    for month in months:
        month_data = analysis_df[analysis_df['月份'] == month]
        
        # 添加月份标题
        table_data.append([f'{month}月份销售数据', '', '', '', '', '', '', '', '', ''])
        table_data.append(['', '总销售', '订单数量', '产品数量', '成功签收', '订单数量', '发货未签收', '订单数量', '缺货', '订单数量'])
        
        # 添加各网站数据
        for _, row in month_data.iterrows():
            table_data.append([
                row['网站'],
                row['总销售'],
                row['订单数量'],
                row['产品数量'],
                row['成功签收'],
                row['订单数量_2'],
                row['发货未签收'],
                row['订单数量_3'],
                row['缺货'],
                row['订单数量']
            ])
        
        # 添加月度汇总
        month_total_sales = month_data['总销售'].sum()
        month_total_orders = month_data['订单数量'].sum()
        month_total_products = month_data['产品数量'].sum()
        month_total_completed = month_data['成功签收'].sum()
        month_total_shipped = month_data['发货未签收'].sum()
        month_total_out_of_stock = month_data['缺货'].sum()
        
        table_data.append([
            '汇总',
            month_total_sales,
            month_total_orders,
            month_total_products,
            month_total_completed,
            month_total_orders,
            month_total_shipped,
            month_total_orders,
            month_total_out_of_stock,
            month_total_orders
        ])
        
        table_data.append(['', '', '', '', '', '', '', '', '', ''])  # 空行分隔
    
    # 写入Excel
    format_df = pd.DataFrame(table_data)
    format_df.to_excel(writer, sheet_name='用户格式报告', index=False, header=False)

def create_monthly_summary_sheet(writer, df):
    """创建月度汇总表"""
    
    monthly_summary = df.groupby(['年月', 'source']).agg({
        'id': 'count',
        '产品数量': 'sum',
        '产品种类数': 'sum',
        'total': 'sum'
    }).rename(columns={'id': '订单数', 'total': '销售额'}).reset_index()
    
    # 添加目标完成度
    monthly_summary['目标完成度(%)'] = (monthly_summary['产品数量'] / 2000 * 100).round(2)
    
    monthly_summary.to_excel(writer, sheet_name='月度汇总', index=False)

def create_product_analysis_sheet(writer, df):
    """创建产品分析表"""
    
    # 产品销量分析
    product_analysis = []
    
    for index, row in df.iterrows():
        try:
            items = json.loads(row['line_items'])
            for item in items:
                product_analysis.append({
                    '订单ID': row['id'],
                    '日期': row['date_created'].strftime('%Y-%m-%d'),
                    '网站': row['source'],
                    '产品名称': item.get('name', '未知'),
                    '产品SKU': item.get('sku', '无SKU'),
                    '数量': item.get('quantity', 0),
                    '单价': item.get('price', 0),
                    '小计': item.get('total', 0)
                })
        except json.JSONDecodeError:
            continue
    
    product_df = pd.DataFrame(product_analysis)
    
    if not product_df.empty:
        product_df.to_excel(writer, sheet_name='产品明细', index=False)
        
        # 产品销量排行
        product_ranking = product_df.groupby(['产品名称', '产品SKU']).agg({
            '数量': 'sum',
            '小计': 'sum',
            '订单ID': 'nunique'
        }).rename(columns={'订单ID': '订单数'}).sort_values('数量', ascending=False).reset_index()
        
        product_ranking.to_excel(writer, sheet_name='产品销量排行', index=False)

def format_analysis_sheets(writer):
    """格式化分析表格"""
    
    def _text_width(s):
        if s is None:
            return 0
        t = str(s)
        w = 0
        for ch in t:
            w += 2 if ord(ch) > 127 else 1
        return w

    # 对所有工作表应用基本格式
    for sheet_name in writer.sheets:
        ws = writer.sheets[sheet_name]
        
        # 全局居中
        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # 自动调整列宽
        for column in ws.columns:
            max_w = 0
            col_letter = column[0].column_letter
            for cell in column:
                try:
                    tw = _text_width(cell.value)
                    if tw > max_w:
                        max_w = tw
                except:
                    pass
            adjusted_width = min(max_w + 4, 100)
            ws.column_dimensions[col_letter].width = adjusted_width

    # 格式化用户格式报告的特殊样式
    if '用户格式报告' in writer.sheets:
        ws = writer.sheets['用户格式报告']
        
        # 设置标题样式
        title_font = Font(size=14, bold=True, color='000080')
        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        
        # 格式化标题行
        for row_num in range(1, ws.max_row + 1):
            cell_value = ws.cell(row=row_num, column=1).value
            if cell_value and '月份销售数据' in str(cell_value):
                for col in range(1, 11):
                    cell = ws.cell(row=row_num, column=col)
                    cell.font = title_font
                    cell.fill = PatternFill(start_color='E6E6FA', end_color='E6E6FA', fill_type='solid')

def print_analysis_summary(analysis_df, df):
    """打印分析摘要"""
    
    print("\n" + "="*60)
    print("综合销售分析报告摘要")
    print("="*60)
    
    total_sales = analysis_df['总销售'].sum()
    total_orders = analysis_df['订单数量'].sum()
    total_products = analysis_df['产品数量'].sum()
    
    print(f"📊 总体统计:")
    print(f"   总销售额: {total_sales:.2f}")
    print(f"   总订单数: {total_orders}")
    print(f"   总产品销量: {total_products} 支")
    print(f"   平均订单金额: {total_sales/total_orders:.2f}")
    print(f"   平均每订单产品数: {total_products/total_orders:.2f} 支")
    
    print(f"\n📈 月度表现:")
    monthly_stats = analysis_df.groupby('月份').agg({
        '总销售': 'sum',
        '订单数量': 'sum',
        '产品数量': 'sum'
    })
    
    for month, stats in monthly_stats.iterrows():
        target_completion = (stats['产品数量'] / 2000) * 100
        print(f"   {month}: {stats['产品数量']} 支产品, {stats['订单数量']} 订单, 销售额 {stats['总销售']:.2f} (目标完成度: {target_completion:.1f}%)")
    
    print(f"\n🏪 网站对比:")
    site_stats = analysis_df.groupby('网站').agg({
        '总销售': 'sum',
        '订单数量': 'sum',
        '产品数量': 'sum'
    })
    
    for site, stats in site_stats.iterrows():
        print(f"   {site}:")
        print(f"     销售额: {stats['总销售']:.2f}")
        print(f"     订单数: {stats['订单数量']}")
        print(f"     产品销量: {stats['产品数量']} 支")
    
    print(f"\n🎯 销售目标分析:")
    avg_monthly_products = monthly_stats['产品数量'].mean()
    print(f"   月度目标: 2000 支")
    print(f"   月均销量: {avg_monthly_products:.0f} 支")
    if avg_monthly_products >= 2000:
        print(f"   ✅ 平均销量已达标")
    else:
        print(f"   ❌ 需要提升 {2000-avg_monthly_products:.0f} 支 ({((2000-avg_monthly_products)/avg_monthly_products*100):.1f}%)")

if __name__ == "__main__":
    filename = create_comprehensive_sales_analysis()
    print(f"\n✅ 综合销售分析报告已完成: {filename}")