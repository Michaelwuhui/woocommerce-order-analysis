"""
WooCommerce Order Analysis Web Dashboard
Flask application with user authentication and data visualization
"""
import sqlite3
import json
import html
from datetime import datetime
from functools import wraps

from flask import Flask, render_template, render_template_string, request, redirect, url_for, flash, jsonify, session, send_file, make_response
from flask_login import LoginManager, UserMixin, login_user, logout_user, login_required, current_user
from oid_utils import woo_post_id  # cross-site-safe WC post id for REST calls
import blocklist  # customer blocklist: auto-cancel COD orders from blacklisted phones
from werkzeug.security import generate_password_hash, check_password_hash
import pandas as pd

app = Flask(__name__)
app.secret_key = 'woocommerce-order-analysis-secret-key-2024'

# Flask-Login setup
login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = 'login'
login_manager.login_message = '请先登录以访问此页面'

# Database configuration
DB_FILE = 'woocommerce_orders.db'

# Simple user storage (in production, use a proper database)
USERS = {
    'admin': {
        'id': '1',
        'username': 'admin',
        'password': generate_password_hash('admin123'),
        'name': '管理员'
    }
}


class User(UserMixin):
    def __init__(self, user_id, username, name, role='user'):
        self.id = user_id
        self.username = username
        self.name = name
        self.role = role
    
    def is_admin(self):
        return self.role == 'admin'
    
    def is_viewer(self):
        """Viewer can see all data but cannot modify"""
        return self.role == 'viewer'
    
    def can_edit(self):
        """Check if user can edit/modify data (not a viewer)"""
        return self.role in ('admin', 'user')
    
    def can_ship(self):
        """Check if user has shipping EDIT permission (operate: ship / confirm /
        mark refused / problem-return / batch / etc.)."""
        if self.role == 'admin':
            return True
        # Check database for can_ship flag
        from flask import current_app
        conn = get_db_connection()
        user = conn.execute('SELECT can_ship FROM users WHERE id = ?', (self.id,)).fetchone()
        conn.close()
        return user and user['can_ship'] == 1

    def can_view_shipping(self):
        """Shipping VIEW permission: can open 发货管理 and see the lists / tracking
        / status / 查物流 / print / export, but cannot operate. Anyone who can ship
        (edit) implicitly can view."""
        if self.can_ship():
            return True
        conn = get_db_connection()
        try:
            user = conn.execute('SELECT can_view_shipping FROM users WHERE id = ?', (self.id,)).fetchone()
            conn.close()
            return bool(user and user['can_view_shipping'] == 1)
        except Exception:
            return False

    def can_view_report(self):
        """Check if user has report viewing permission"""
        if self.role == 'admin':
            return True
        # Check database for can_view_report flag
        from flask import current_app
        conn = get_db_connection()
        try:
            user = conn.execute('SELECT can_view_report FROM users WHERE id = ?', (self.id,)).fetchone()
            conn.close()
            return user and user['can_view_report'] == 1
        except:
            # Column might not exist yet if migration failed or legacy db
            return False

    def can_view_sales_board(self):
        """Check if user has sales board viewing permission (all roles must be explicitly granted)"""
        conn = get_db_connection()
        try:
            user = conn.execute('SELECT can_view_sales_board FROM users WHERE id = ?', (self.id,)).fetchone()
            conn.close()
            return user and user['can_view_sales_board'] == 1
        except:
            return False

    def can_view_own_sales_board(self):
        """Self-only sales board: user may open the board but sees ONLY their own
        sites (sites whose manager == this user's name). Distinct from the full
        can_view_sales_board (whole team). Granted in 用户管理, super-admin only."""
        conn = get_db_connection()
        try:
            user = conn.execute('SELECT can_view_own_sales_board FROM users WHERE id = ?', (self.id,)).fetchone()
            conn.close()
            return bool(user and user['can_view_own_sales_board'] == 1)
        except Exception:
            return False

    def sales_board_own_manager(self):
        """The manager-name this user is scoped to on the self-only board (their
        own name). Returns the name only if they actually have sites under it."""
        try:
            conn = get_db_connection()
            row = conn.execute('SELECT name FROM users WHERE id = ?', (self.id,)).fetchone()
            nm = (row['name'] if row else '') or ''
            has = conn.execute('SELECT 1 FROM sites WHERE manager = ? LIMIT 1', (nm,)).fetchone()
            conn.close()
            return nm if (nm and has) else None
        except Exception:
            return None

    def can_manage_products(self):
        """Check if user can access the multi-site product manager (Layer 1).
        Super admin always passes. Other users must have the flag set explicitly;
        their site-level scope is enforced separately via user_site_permissions."""
        if self.username == 'admin':
            return True
        conn = get_db_connection()
        try:
            user = conn.execute('SELECT can_manage_products FROM users WHERE id = ?', (self.id,)).fetchone()
            conn.close()
            return user and user['can_manage_products'] == 1
        except sqlite3.OperationalError:
            return False

    def can_manage_users(self):
        """Check if user can manage other users and permissions (super admin privilege)"""
        # 'admin' username always has this right as a safety net
        if self.username == 'admin':
            return True
        conn = get_db_connection()
        try:
            user = conn.execute('SELECT can_manage_users FROM users WHERE id = ?', (self.id,)).fetchone()
            conn.close()
            return user and user['can_manage_users'] == 1
        except:
            return False

    def can_view_reconciliation(self):
        """Check if user can access partner reconciliation page (read-only gate).
        Must have can_view_reconciliation flag OR be super admin."""
        if self.username == 'admin':
            return True
        conn = get_db_connection()
        try:
            user = conn.execute('SELECT can_view_reconciliation FROM users WHERE id = ?', (self.id,)).fetchone()
            conn.close()
            return user and user['can_view_reconciliation'] == 1
        except:
            return False

    def can_edit_reconciliation(self):
        """Check if user can edit reconciliation data (write access).
        Must have can_edit_reconciliation flag OR be super admin.
        Partner members (bound in partner_users) typically do NOT have this flag."""
        if self.username == 'admin':
            return True
        conn = get_db_connection()
        try:
            user = conn.execute('SELECT can_edit_reconciliation FROM users WHERE id = ?', (self.id,)).fetchone()
            conn.close()
            return user and user['can_edit_reconciliation'] == 1
        except:
            return False

    def can_view_costs(self):
        """Check if user can view the cost management page (read-only gate).
        Must have can_view_costs flag OR be super admin. Editing costs is
        gated separately by can_edit_costs."""
        if self.username == 'admin':
            return True
        conn = get_db_connection()
        try:
            user = conn.execute('SELECT can_view_costs FROM users WHERE id = ?', (self.id,)).fetchone()
            conn.close()
            return bool(user and user['can_view_costs'] == 1)
        except:
            return False

    def can_edit_costs(self):
        """Check if user can EDIT (add/update/delete) product costs.
        Super admin always passes. Otherwise requires the can_edit_costs flag.
        Implies can_view_costs (no point editing what you can't see — the UI
        respects this, but this method itself doesn't auto-grant view)."""
        if self.username == 'admin':
            return True
        conn = get_db_connection()
        try:
            user = conn.execute('SELECT can_edit_costs FROM users WHERE id = ?', (self.id,)).fetchone()
            conn.close()
            return bool(user and user['can_edit_costs'] == 1)
        except:
            return False

    def can_manage_blocklist(self):
        """Add/remove customer-blocklist entries (which triggers auto-cancel of
        their COD orders) — a privileged action. Admin role always passes;
        other users need the can_manage_blocklist flag, granted per-user in
        用户管理 by the super admin."""
        if self.role == 'admin':
            return True
        conn = get_db_connection()
        try:
            user = conn.execute('SELECT can_manage_blocklist FROM users WHERE id = ?', (self.id,)).fetchone()
            conn.close()
            return bool(user and user['can_manage_blocklist'] == 1)
        except Exception:
            return False

    def get_accessible_partner_ids(self):
        """Return list of partner IDs the user can access.
        Returns None = unrestricted (super admin, or user with permission but no bindings).
        Returns list of IDs = restricted to bound partners.
        Semantics:
          - Super admin: None (all)
          - Has permission + bound to partner(s): only those bound
          - Has permission + no bindings: None (all — internal finance role)
          - No permission: [] (blocked before reaching here)"""
        if self.username == 'admin':
            return None
        conn = get_db_connection()
        try:
            user = conn.execute('SELECT can_view_reconciliation FROM users WHERE id = ?', (self.id,)).fetchone()
            if not user or user['can_view_reconciliation'] != 1:
                conn.close()
                return []
            # Has permission — check if scoped to specific partners
            rows = conn.execute('SELECT partner_id FROM partner_users WHERE user_id = ?', (self.id,)).fetchall()
            conn.close()
            if rows:
                return [r['partner_id'] for r in rows]
            # No bindings = internal finance, sees all
            return None
        except:
            return []


@login_manager.user_loader
def load_user(user_id):
    conn = get_db_connection()
    user_row = conn.execute('SELECT * FROM users WHERE id = ?', (user_id,)).fetchone()
    conn.close()
    if user_row:
        return User(user_row['id'], user_row['username'], user_row['name'], user_row['role'])
    return None


def get_db_connection():
    """Create database connection"""
    conn = sqlite3.connect(DB_FILE)
    conn.row_factory = sqlite3.Row
    return conn


def get_all_managers():
    """Get list of all unique site managers"""
    conn = get_db_connection()
    managers = conn.execute('SELECT DISTINCT manager FROM sites WHERE manager IS NOT NULL AND manager != "" ORDER BY manager').fetchall()
    conn.close()
    return [m['manager'] for m in managers]


def parse_json_field(value):
    """Safely parse JSON field"""
    if not value:
        return {}
    try:
        return json.loads(value)
    except:
        return {}


# ─────────────────────────── Customer risk ─────────────────────────────
# Match repeat freeloaders across email / phone / address. The matching
# keys are deliberately fuzzy-but-conservative: we want to catch a
# customer who tweaks their email but reuses the same phone, but NOT
# false-positive different households sharing one apartment block. See
# _build_risk_index / _assess_customer_risk below.

def _normalize_email(e):
    s = (e or '').strip().lower()
    return s or None


def _is_placeholder_phone(digits):
    """True for obviously-fake numbers people type to bypass a required field:
    all-same-digit (000000000), or a straight ascending/descending run
    (123456789 / 987654321). These must never link unrelated customers."""
    if not digits:
        return True
    if len(set(digits)) <= 1:
        return True
    if len(digits) >= 6:
        diffs = {int(digits[i + 1]) - int(digits[i]) for i in range(len(digits) - 1)}
        if diffs == {1} or diffs == {-1}:
            return True
    return False


def _normalize_phone(p):
    """Keep digits only; collapse to the last 9 digits to fold away the
    PL +48 / AU +61 / AE +971 country codes that customers add inconsistently.
    Returns None if there's no plausible phone number left, or if it's an
    obvious placeholder (so it can't merge / flag unrelated people)."""
    digits = ''.join(c for c in (p or '') if c.isdigit())
    if len(digits) < 7:  # too short to be a real number
        return None
    canonical = digits[-9:] if len(digits) >= 9 else digits
    if _is_placeholder_phone(canonical) or _is_placeholder_phone(digits):
        return None
    return canonical


def _normalize_address(addr_dict):
    """Build a tight match key from address_1 + postcode + city. All
    lowercase, whitespace collapsed. Returns None when too sparse to be
    a useful match (a country code alone isn't a customer)."""
    if not isinstance(addr_dict, dict):
        return None
    import re as _re
    a1 = _re.sub(r'\s+', ' ', (addr_dict.get('address_1') or '').strip().lower())
    pc = (addr_dict.get('postcode') or '').strip().replace(' ', '').lower()
    city = _re.sub(r'\s+', ' ', (addr_dict.get('city') or '').strip().lower())
    parts = [p for p in (a1, pc, city) if p]
    if len(parts) < 2 or not a1:
        return None
    return ' | '.join(parts)


def _addr_for_order(billing, shipping):
    """Pick the more complete of (shipping, billing) for matching."""
    if isinstance(shipping, dict) and shipping.get('address_1'):
        return shipping
    return billing if isinstance(billing, dict) else {}


def _compose_address(addr, sep=', '):
    """Compose a human-readable shipping address from a billing/shipping dict.

    The locality line is "City State Postcode" (e.g. "Brassall QLD 4305") so the
    destination state is always shown. It used to be dropped from every shipping
    view, which made AU/US parcels ambiguous — the same suburb name recurs across
    states, and packers/carriers need the state to route correctly. Polish (DPD/
    InPost) addresses simply have no state, so it falls away cleanly there."""
    if not isinstance(addr, dict):
        return ''
    locality = ' '.join(p for p in (
        (addr.get('city') or '').strip(),
        (addr.get('state') or '').strip(),
        (addr.get('postcode') or '').strip(),
    ) if p)
    parts = [
        (addr.get('address_1') or '').strip(),
        (addr.get('address_2') or '').strip(),
        locality,
        (addr.get('country') or '').strip(),
    ]
    return sep.join(p for p in parts if p)


# AU postcode -> expected state (Australia Post ranges). Checkout's state
# dropdown is customer-chosen and never validated by WC, so e.g. Perth+6000
# can arrive with state=ACT. Gray-zone ranges list every state they can
# legitimately belong to, so only impossible combinations get flagged.
_AU_POSTCODE_STATES = (
    ((200, 299), {'ACT'}), ((800, 999), {'NT'}),
    ((1000, 2599), {'NSW'}), ((2600, 2620), {'ACT', 'NSW'}),
    ((2621, 2899), {'NSW'}), ((2900, 2920), {'ACT', 'NSW'}),
    ((2921, 2999), {'NSW'}), ((3000, 3999), {'VIC'}),
    ((4000, 4999), {'QLD'}), ((5000, 5999), {'SA'}),
    ((6000, 6999), {'WA'}), ((7000, 7999), {'TAS'}),
    ((8000, 8999), {'VIC'}), ((9000, 9999), {'QLD'}),
)


def _au_state_mismatch(addr):
    """Return the expected AU state code(s) ('WA' / 'ACT/NSW') when the chosen
    state cannot contain the postcode; None when consistent or not checkable."""
    if not isinstance(addr, dict) or (addr.get('country') or '').upper() != 'AU':
        return None
    state = (addr.get('state') or '').strip().upper()
    pc_raw = (addr.get('postcode') or '').strip()
    if not state or not pc_raw.isdigit():
        return None
    pc = int(pc_raw)
    if pc == 872:  # tri-state remote zone (NT/SA/WA), never flag
        return None
    for (lo, hi), states in _AU_POSTCODE_STATES:
        if lo <= pc <= hi:
            return None if state in states else '/'.join(sorted(states))
    return None


def _build_risk_index(conn):
    """Scan every order that's been flagged as a problem-return or undelivered
    and group them by normalized email / phone / address. Each bucket lists
    the matching orders so the consumer can dedupe and aggregate.

    Returned shape:
        {
          'problem': {'email': {key: [rec, ...]}, 'phone': {...}, 'addr': {...}},
          'undeliv': {'email': {...}, 'phone': {...}, 'addr': {...}},
        }
    rec = {order_id, number, source, at, loss, type}
    """
    out = {
        'problem': {'email': {}, 'phone': {}, 'addr': {}},
        'undeliv': {'email': {}, 'phone': {}, 'addr': {}},
    }
    try:
        rows = conn.execute("""
            SELECT id, number, billing, shipping, source,
                   is_problem_return, problem_return_type, product_loss_amount, problem_return_at,
                   is_undelivered, shipping_loss_amount, undelivered_at
            FROM orders
            WHERE COALESCE(is_problem_return,0) = 1 OR COALESCE(is_undelivered,0) = 1
        """).fetchall()
    except Exception:
        return out

    for r in rows:
        billing = parse_json_field(r['billing']) or {}
        shipping = parse_json_field(r['shipping']) or {}
        addr_d = _addr_for_order(billing, shipping)
        email = _normalize_email(billing.get('email') or shipping.get('email'))
        phone = _normalize_phone(addr_d.get('phone') or billing.get('phone'))
        addr_key = _normalize_address(addr_d)

        if r['is_problem_return']:
            rec = {
                'order_id': r['id'],
                'number': r['number'],
                'source': r['source'],
                'at': r['problem_return_at'],
                'loss': float(r['product_loss_amount'] or 0),
                'type': r['problem_return_type'] or '',
            }
            if email: out['problem']['email'].setdefault(email, []).append(rec)
            if phone: out['problem']['phone'].setdefault(phone, []).append(rec)
            if addr_key: out['problem']['addr'].setdefault(addr_key, []).append(rec)
        if r['is_undelivered']:
            rec = {
                'order_id': r['id'],
                'number': r['number'],
                'source': r['source'],
                'at': r['undelivered_at'],
                'loss': float(r['shipping_loss_amount'] or 0),
                'type': '',
            }
            if email: out['undeliv']['email'].setdefault(email, []).append(rec)
            if phone: out['undeliv']['phone'].setdefault(phone, []).append(rec)
            if addr_key: out['undeliv']['addr'].setdefault(addr_key, []).append(rec)

    return out


# Threshold for the medium ("repeat refuser") warning. Tunable later.
_RISK_UNDELIVERED_THRESHOLD = 2


def _assess_customer_risk(billing, shipping, idx, current_order_id=None):
    """Look up this customer in the risk index across email/phone/addr and
    return a risk dict, or None if no prior risky history is found.

    The current order itself is excluded by id so a freshly-marked order
    doesn't flag its own self.

    Returns:
        {
          'level': 'high' | 'medium',
          'problem_count': int,            # distinct prior 问题退货 orders
          'problem_loss': float,           # sum of product_loss_amount
          'undeliv_count': int,            # distinct prior 未送达 orders
          'undeliv_loss': float,           # sum of shipping_loss_amount
          'matched_by': ['email', 'phone', 'address'],
          'last_at': '2026-05-18 08:03',   # most recent of either
          'last_number': '7914',
          'types': ['swap', 'short', ...]  # problem-return types seen
        }
    """
    if not idx:
        return None

    billing = billing or {}
    shipping = shipping or {}
    addr_d = _addr_for_order(billing, shipping)
    email = _normalize_email(billing.get('email') or shipping.get('email'))
    phone = _normalize_phone(addr_d.get('phone') or billing.get('phone'))
    addr_key = _normalize_address(addr_d)

    keys = [('email', email), ('phone', phone), ('address', addr_key)]

    p_matched, p_via = _collect_in_indices(idx['problem'], keys, current_order_id)
    u_matched, u_via = _collect_in_indices(idx['undeliv'], keys, current_order_id)

    problem_count = len(p_matched)
    problem_loss = sum(r['loss'] for r in p_matched.values())
    undeliv_count = len(u_matched)
    undeliv_loss = sum(r['loss'] for r in u_matched.values())

    # Severity gating: problem-return at all → high; otherwise medium needs
    # the count threshold so a customer with 1 unlucky no-receipt doesn't
    # trip a yellow flag on every future order.
    level = None
    if problem_count >= 1:
        level = 'high'
    elif undeliv_count >= _RISK_UNDELIVERED_THRESHOLD:
        level = 'medium'

    if not level:
        return None

    # Latest event for display
    all_recs = list(p_matched.values()) + list(u_matched.values())
    latest = max(all_recs, key=lambda r: (r['at'] or ''))
    types = sorted({r['type'] for r in p_matched.values() if r.get('type')})

    return {
        'level': level,
        'problem_count': problem_count,
        'problem_loss': round(problem_loss, 2),
        'undeliv_count': undeliv_count,
        'undeliv_loss': round(undeliv_loss, 2),
        'matched_by': sorted(p_via | u_via),
        'last_at': latest['at'],
        'last_number': latest['number'],
        'types': types,
    }


def _collect_in_indices(category_idx, keys, current_order_id):
    """Helper: walk email/phone/addr sub-indices in `category_idx`, dedupe by
    order_id, track which marker matched."""
    matched = {}
    via = set()
    for marker, k in keys:
        if not k:
            continue
        for rec in category_idx.get(marker, {}).get(k, []):
            if current_order_id is not None and rec['order_id'] == current_order_id:
                continue
            matched[rec['order_id']] = rec
            via.add(marker)
    return matched, via


# A phone / address shared by MORE distinct emails than this is treated as
# non-identifying (placeholder number like 123456789, a shared web form, an
# apartment block) and will NOT merge identities. Email is always trusted.
_IDENTITY_SHARED_KEY_LIMIT = 4


def _resolve_identity_clusters(conn, where_clause, params):
    """Entity resolution over the customer base: union emails that belong to
    the same person, linked by a shared phone number or delivery address.

    Returns a tuple (email_to_cluster, cluster_meta):
      email_to_cluster: {normalized_email: cluster_key}
      cluster_meta:     {cluster_key: {'emails': set, 'phones': set, 'addrs': set,
                                       'matched_by': set(['phone','address'])}}
    The cluster_key is the lexicographically-smallest email in the cluster
    (stable across requests). Emails that share nothing form singleton clusters.

    Over-merge guard: phone/address keys linked to more than
    _IDENTITY_SHARED_KEY_LIMIT distinct emails are dropped before unioning.
    """
    rows = conn.execute(
        f"SELECT billing, shipping FROM orders {where_clause}", params
    ).fetchall()

    email_keys = {}     # email -> {'phones': set, 'addrs': set}
    phone_emails = {}   # phone -> set(emails)
    addr_emails = {}    # addr  -> set(emails)
    for r in rows:
        b = parse_json_field(r['billing']) or {}
        s = parse_json_field(r['shipping']) or {}
        ad = _addr_for_order(b, s)
        e = _normalize_email(b.get('email') or s.get('email'))
        if not e:
            continue
        p = _normalize_phone(ad.get('phone') or b.get('phone'))
        ak = _normalize_address(ad)
        ek = email_keys.setdefault(e, {'phones': set(), 'addrs': set()})
        if p:
            ek['phones'].add(p)
            phone_emails.setdefault(p, set()).add(e)
        if ak:
            ek['addrs'].add(ak)
            addr_emails.setdefault(ak, set()).add(e)

    good_phone = {p for p, es in phone_emails.items() if len(es) <= _IDENTITY_SHARED_KEY_LIMIT}
    good_addr = {a for a, es in addr_emails.items() if len(es) <= _IDENTITY_SHARED_KEY_LIMIT}

    # Union-Find over emails.
    parent = {}

    def find(x):
        parent.setdefault(x, x)
        root = x
        while parent[root] != root:
            root = parent[root]
        while parent[x] != root:  # path compression
            parent[x], x = root, parent[x]
        return root

    def union(a, b):
        ra, rb = find(a), find(b)
        if ra == rb:
            return
        if rb < ra:
            ra, rb = rb, ra
        parent[rb] = ra

    for e in email_keys:
        find(e)

    matched_via = {}  # email -> set of reasons it got merged ('phone'/'address')
    for p, es in phone_emails.items():
        if p not in good_phone or len(es) < 2:
            continue
        es = list(es)
        for other in es[1:]:
            union(es[0], other)
        for e in es:
            matched_via.setdefault(e, set()).add('phone')
    for a, es in addr_emails.items():
        if a not in good_addr or len(es) < 2:
            continue
        es = list(es)
        for other in es[1:]:
            union(es[0], other)
        for e in es:
            matched_via.setdefault(e, set()).add('address')

    email_to_cluster = {e: find(e) for e in email_keys}
    cluster_meta = {}
    for e, ck in email_to_cluster.items():
        m = cluster_meta.setdefault(ck, {'emails': set(), 'phones': set(),
                                         'addrs': set(), 'matched_by': set()})
        m['emails'].add(e)
        m['phones'] |= email_keys[e]['phones']
        m['addrs'] |= email_keys[e]['addrs']
        m['matched_by'] |= matched_via.get(e, set())
    return email_to_cluster, cluster_meta


def extract_flavor_from_meta(item):
    """
    Extract flavor/variation attribute from WooCommerce line_item meta_data.
    Looks for common variation attribute keys like 'flavour', 'flavor', 'pa_flavour', 'pa_flavor', etc.
    Returns the display_value if found, otherwise empty string.
    """
    meta_data = item.get('meta_data', [])
    if not isinstance(meta_data, list):
        return ''
    
    # Common flavor-related keys in WooCommerce (also include Polish 'smak/smaki')
    flavor_keys = ['pa_flavour', 'pa_flavor', 'flavour', 'flavor', 'pa_taste', 'taste', 'pa_variant', 'variant', 'pa_smak', 'smak', 'pa_smaki', 'smaki']
    
    for meta in meta_data:
        if not isinstance(meta, dict):
            continue
        key = meta.get('key', '').lower()
        if key in flavor_keys:
            # Prefer display_value over value for human-readable format
            return meta.get('display_value', '') or meta.get('value', '')
    
    return ''


def extract_puffs_from_meta(item):
    """
    Extract puffs count from WooCommerce line_item meta_data.
    Looks for 'puffs', 'pa_puffs', 'puff_count' keys.
    Returns the numeric puffs value if found, otherwise None.
    """
    import re
    
    meta_data = item.get('meta_data', [])
    if not isinstance(meta_data, list):
        return None
    
    # Common puffs-related keys in WooCommerce (including Polish 'liczba-zaciagniec')
    puffs_keys = ['pa_puffs', 'puffs', 'puff_count', 'pa_puff_count', 'pa_liczba-zaciagniec', 'liczba-zaciagniec']
    
    for meta in meta_data:
        if not isinstance(meta, dict):
            continue
        key = meta.get('key', '').lower()
        if key in puffs_keys:
            value = meta.get('display_value', '') or meta.get('value', '')
            # Extract numeric value from strings like "15000 puffs" or "15000"
            if value:
                match = re.search(r'(\d+)', str(value))
                if match:
                    return int(match.group(1))
    
    return None


def calculate_customer_tier(successful_orders, total_spending, avg_days_between,
                            bad_orders=0, meaningful_orders=0):
    """Calculate customer tier from orders, spending, frequency — then apply a
    refusal/return-rate PENALTY so a serial refuser can't rank high just because
    a few early orders + frequency padded the score.

    bad_orders / meaningful_orders define the refusal rate (see _bad_order_case /
    _meaningful_order_case). The base score (and the revenue numbers it's derived
    from) is unchanged; the penalty only caps/downgrades the resulting tier.
    """
    quality_score = 0
    quality_score += min(successful_orders * 10, 30)  # Max 30 for orders
    quality_score += min(total_spending / 100, 40)    # Max 40 for spending

    # Frequency bonus
    if avg_days_between > 0:
        if avg_days_between < 60:
            quality_score += 30
        elif avg_days_between < 120:
            quality_score += 15

    quality_score = min(quality_score, 100)

    if quality_score >= 80:
        tier = 'vip'
    elif quality_score >= 60:
        tier = 'good'
    elif quality_score >= 40:
        tier = 'normal'
    else:
        tier = 'new'

    # ── Refusal / return-rate penalty ───────────────────────────────────────
    # Denominator = meaningful orders (shipped/attempted/completed; excludes
    # unpaid-cancelled). Fall back to bad+successful if not supplied.
    try:
        bad_orders = int(bad_orders or 0)
    except (TypeError, ValueError):
        bad_orders = 0
    denom = meaningful_orders if meaningful_orders else (bad_orders + (successful_orders or 0))
    bad_rate = (bad_orders / denom) if denom else 0.0

    rank = ['new', 'normal', 'good', 'vip']

    def _cap(t, ceiling):
        return t if rank.index(t) <= rank.index(ceiling) else ceiling

    if bad_orders >= 3 and bad_rate >= 0.5:
        return 'bad'                 # 劣质：惯性拒收/退货
    if bad_rate >= 0.5:
        tier = _cap(tier, 'normal')  # 拒收率≥50%：最高普通
    elif bad_rate >= 0.3:
        tier = _cap(tier, 'good')    # 拒收率≥30%：不给 VIP
    return tier


def get_full_product_name(item):
    """
    Get full product name including variation attributes (like flavor).
    Combines the product name with any variation flavor found in meta_data.
    Returns: (full_name, flavor_only, puffs_from_meta)
    """
    name = item.get('name', '')
    flavor = extract_flavor_from_meta(item)
    puffs = extract_puffs_from_meta(item)
    
    if flavor:
        # If flavor is not already in the name, append it
        if flavor.upper() not in name.upper():
            full_name = f"{name} - {flavor}"
        else:
            full_name = name
    else:
        full_name = name
        flavor = ''
    
    return full_name, flavor, puffs


def get_user_allowed_sources(user_id, is_admin=False, is_viewer=False):
    """
    Get list of source URLs that a user has permission to access.
    Admin and Viewer users can access all sources.
    Returns None if user has no restrictions (can see all), or a list of allowed URLs.
    """
    if is_admin or is_viewer:
        return None  # No restrictions for admin or viewer
    return _resolve_user_sites(user_id)


def _resolve_user_sites(user_id):
    """Site URLs a user is scoped to, resolved live from the 3 permission tables
    (country grants ∪ explicit single-site grants − exclusions). Role-INDEPENDENT —
    callers decide whether a role overrides it (view: admin/viewer see all; edit:
    see get_user_editable_sources). New sites in a granted country auto-inherit."""
    conn = get_db_connection()
    try:
        granted_countries = {r['country'] for r in conn.execute(
            'SELECT country FROM user_country_permissions WHERE user_id = ?', (user_id,)).fetchall()}
        excluded_ids = {r['site_id'] for r in conn.execute(
            'SELECT site_id FROM user_site_exclusions WHERE user_id = ?', (user_id,)).fetchall()}
        explicit_ids = {r['site_id'] for r in conn.execute(
            'SELECT site_id FROM user_site_permissions WHERE user_id = ?', (user_id,)).fetchall()}
        urls = []
        for s in conn.execute('SELECT id, url, country FROM sites').fetchall():
            if s['country'] in granted_countries:
                if s['id'] not in excluded_ids:
                    urls.append(s['url'])      # covered by country grant (future sites included)
            elif s['id'] in explicit_ids:
                urls.append(s['url'])          # explicitly granted single site
        return urls
    finally:
        conn.close()


def get_user_editable_sources(user):
    """Source URLs a user may EDIT (mutate orders) for. None = unrestricted.
      - super admin (username 'admin')            → None (edits everything)
      - sites the user MANAGES (sites.manager == their name) are auto-granted —
        being the 负责人 doubles as an edit grant, no per-site config needed
      - plus any explicitly-granted / country-granted sites (3-table scope)
      - admin with NO scope at all (manages nothing, no grants) → None (finance/
        global admins keep full edit — chosen backward-compat rule)
      - anyone else with no scope                  → empty set (cannot edit)"""
    if getattr(user, 'username', None) == 'admin':
        return None
    sites = set(_resolve_user_sites(user.id))
    # Auto-match: also editable for the sites this user is the 负责人 of. An explicit
    # exclusion still wins (admin can carve a managed site back out if they must).
    uname = (getattr(user, 'name', None) or '').strip()
    if uname:
        conn = get_db_connection()
        try:
            excluded = {r['site_id'] for r in conn.execute(
                'SELECT site_id FROM user_site_exclusions WHERE user_id = ?', (user.id,)).fetchall()}
            sites |= {r['url'] for r in conn.execute(
                'SELECT id, url FROM sites WHERE manager = ?', (uname,)).fetchall()
                if r['id'] not in excluded}
        finally:
            conn.close()
    if not sites and user.is_admin():
        return None
    return sites


def _site_edit_block(sources):
    """Return a 403 JSON response if current_user may NOT edit EVERY url in `sources`,
    else None. Call inside order-mutation endpoints once the order source(s) are known:
        block = _site_edit_block([order['source']])
        if block: return block
    """
    scope = get_user_editable_sources(current_user)
    if scope is None:
        return None
    for src in sources:
        if src not in scope:
            return jsonify({'error': '无权编辑该站点的订单（只能操作已授权的站点）', 'site_forbidden': True}), 403
    return None


def _order_site_edit_block(order_id):
    """403 response if current_user may not edit the given order's site, else None.
    Self-contained (own connection). Returns None for a missing order so the endpoint
    can still emit its own 404."""
    scope = get_user_editable_sources(current_user)
    if scope is None:
        return None
    conn = get_db_connection()
    try:
        row = conn.execute('SELECT source FROM orders WHERE id = ?', (order_id,)).fetchone()
    finally:
        conn.close()
    if row is not None and row['source'] not in scope:
        return jsonify({'error': '无权编辑该站点的订单（只能操作已授权的站点）', 'site_forbidden': True}), 403
    return None


def order_site_editable(f):
    """Decorator: after the role gate, block the request when current_user is not
    allowed to edit the target order's SITE. Resolves order_id from the URL kwarg,
    else from the JSON body's 'order_id'. Place it just above the view function
    (below @editor_required / @shipper_required) so the role check runs first."""
    from functools import wraps
    @wraps(f)
    def wrapper(*args, **kwargs):
        oid = kwargs.get('order_id')
        if oid is None:
            try:
                oid = (request.json or {}).get('order_id')
            except Exception:
                oid = None
        if oid is not None:
            blk = _order_site_edit_block(oid)
            if blk:
                return blk
        return f(*args, **kwargs)
    return wrapper


def build_source_filter_clause(allowed_sources, table_alias=''):
    """
    Build SQL WHERE clause for filtering by allowed sources.
    Returns (clause_string, params_list)
    """
    if allowed_sources is None:
        # No restrictions
        return '', []
    
    if not allowed_sources:
        # Empty list - no access to anything
        return 'AND 1=0', []  # Always false condition
    
    prefix = f'{table_alias}.' if table_alias else ''
    placeholders = ', '.join(['?' for _ in allowed_sources])
    return f'AND {prefix}source IN ({placeholders})', allowed_sources


def get_cny_rate(currency, year_month):
    """
    Get CNY exchange rate for a currency in a specific month.
    Falls back to the most recent rate if not found for the specific month.
    Returns (rate, actual_year_month) or (None, None) if not found.
    """
    if not currency or currency == 'CNY':
        return 1.0, year_month  # CNY to CNY is 1:1
    
    conn = get_db_connection()
    
    # Try exact month first
    rate = conn.execute('''
        SELECT rate_to_cny, year_month FROM exchange_rates
        WHERE currency = ? AND year_month = ?
    ''', (currency, year_month)).fetchone()
    
    if rate:
        conn.close()
        return rate['rate_to_cny'], rate['year_month']
    
    # Fall back to most recent rate before this month
    rate = conn.execute('''
        SELECT rate_to_cny, year_month FROM exchange_rates
        WHERE currency = ? AND year_month <= ?
        ORDER BY year_month DESC
        LIMIT 1
    ''', (currency, year_month)).fetchone()
    
    if rate:
        conn.close()
        return rate['rate_to_cny'], rate['year_month']
    
    # Fall back to any rate for this currency
    rate = conn.execute('''
        SELECT rate_to_cny, year_month FROM exchange_rates
        WHERE currency = ?
        ORDER BY year_month DESC
        LIMIT 1
    ''', (currency,)).fetchone()
    
    conn.close()
    
    if rate:
        return rate['rate_to_cny'], rate['year_month']
    
    return None, None


def convert_to_cny(amount, currency, year_month):
    """
    Convert an amount to CNY using the exchange rate for the given month.
    Returns (cny_amount, rate, actual_month) or (None, None, None) if no rate found.
    """
    if amount is None:
        return None, None, None
    
    rate, actual_month = get_cny_rate(currency, year_month)
    if rate is None:
        return None, None, None
    
    return round(amount * rate, 2), rate, actual_month


def get_all_exchange_rates():
    """Get all exchange rates from database."""
    conn = get_db_connection()
    rates = conn.execute('''
        SELECT id, year_month, currency, rate_to_cny, updated_at
        FROM exchange_rates
        ORDER BY year_month DESC, currency
    ''').fetchall()
    conn.close()
    return [dict(r) for r in rates]


def admin_required(f):
    """Decorator to require admin role"""
    from functools import wraps
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not current_user.is_authenticated or not current_user.is_admin():
            # Return a beautiful access denied page
            return render_template_string('''
<!DOCTYPE html>
<html lang="zh">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>访问受限 - WooCommerce 订单分析</title>
    <link href="https://cdn.jsdelivr.net/npm/bootstrap-icons@1.10.0/font/bootstrap-icons.css" rel="stylesheet">
    <style>
        * { box-sizing: border-box; margin: 0; padding: 0; }
        body {
            min-height: 100vh;
            display: flex;
            align-items: center;
            justify-content: center;
            background: linear-gradient(135deg, #1a1a2e 0%, #16213e 50%, #0f3460 100%);
            font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", Roboto, sans-serif;
            color: #e0e0e0;
        }
        .container {
            text-align: center;
            padding: 60px 40px;
            background: rgba(255, 255, 255, 0.05);
            border-radius: 24px;
            backdrop-filter: blur(10px);
            border: 1px solid rgba(255, 255, 255, 0.1);
            max-width: 480px;
            box-shadow: 0 25px 50px -12px rgba(0, 0, 0, 0.5);
        }
        .icon {
            width: 120px;
            height: 120px;
            margin: 0 auto 30px;
            background: linear-gradient(135deg, rgba(239, 68, 68, 0.2), rgba(239, 68, 68, 0.1));
            border-radius: 50%;
            display: flex;
            align-items: center;
            justify-content: center;
            border: 2px solid rgba(239, 68, 68, 0.3);
        }
        .icon i { font-size: 48px; color: #ef4444; }
        h1 {
            font-size: 28px;
            font-weight: 600;
            margin-bottom: 16px;
            background: linear-gradient(135deg, #f87171, #ef4444);
            -webkit-background-clip: text;
            -webkit-text-fill-color: transparent;
        }
        p { color: #9ca3af; font-size: 16px; line-height: 1.6; margin-bottom: 30px; }
        .user-info {
            background: rgba(0, 0, 0, 0.2);
            padding: 12px 20px;
            border-radius: 12px;
            margin-bottom: 30px;
            font-size: 14px;
        }
        .user-info span { color: #60a5fa; }
        .btn {
            display: inline-flex;
            align-items: center;
            gap: 8px;
            padding: 14px 32px;
            background: linear-gradient(135deg, #3b82f6, #2563eb);
            color: white;
            text-decoration: none;
            border-radius: 12px;
            font-weight: 500;
            transition: all 0.3s ease;
        }
        .btn:hover { transform: translateY(-2px); box-shadow: 0 10px 20px rgba(59, 130, 246, 0.3); }
    </style>
</head>
<body>
    <div class="container">
        <div class="icon"><i class="bi bi-shield-lock"></i></div>
        <h1>访问受限</h1>
        <p>抱歉，您没有权限访问此页面。<br>此功能仅限管理员使用。</p>
        <div class="user-info">当前登录: <span>{{ current_user.name or current_user.username }}</span></div>
        <a href="/" class="btn"><i class="bi bi-house"></i> 返回首页</a>
    </div>
</body>
</html>
            '''), 403
        return f(*args, **kwargs)
    return decorated_function


def user_manager_required(f):
    """Decorator: only users with can_manage_users privilege can access user management"""
    from functools import wraps
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not current_user.is_authenticated:
            return redirect(url_for('login'))
        if not current_user.is_admin() or not current_user.can_manage_users():
            return render_template_string('''
<!DOCTYPE html>
<html lang="zh"><head><meta charset="UTF-8"><title>权限不足</title>
<link href="https://cdn.jsdelivr.net/npm/bootstrap@5.3.0/dist/css/bootstrap.min.css" rel="stylesheet">
<link href="https://cdn.jsdelivr.net/npm/bootstrap-icons@1.10.0/font/bootstrap-icons.css" rel="stylesheet">
<style>
body{background:#0f172a;color:#e2e8f0;display:flex;align-items:center;justify-content:center;min-height:100vh;margin:0;}
.box{background:#1e293b;border:1px solid #334155;border-radius:16px;padding:48px;text-align:center;max-width:440px;}
.icon{font-size:56px;color:#f59e0b;margin-bottom:20px;}
h1{font-size:22px;font-weight:600;margin-bottom:12px;}
p{color:#94a3b8;margin-bottom:24px;}
a{background:#3b82f6;color:#fff;padding:10px 24px;border-radius:8px;text-decoration:none;}
</style></head>
<body><div class="box">
<div class="icon"><i class="bi bi-shield-exclamation"></i></div>
<h1>权限不足</h1>
<p>用户管理功能需要超级管理员授权，请联系 admin 获取权限。</p>
<a href="/"><i class="bi bi-house"></i> 返回首页</a>
</div></body></html>
            '''), 403
        return f(*args, **kwargs)
    return decorated_function


def editor_required(f):
    """Decorator to require edit permission (admin or user, not viewer)"""
    from functools import wraps
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not current_user.is_authenticated or not current_user.can_edit():
            return jsonify({'error': '只读用户无法修改数据', 'readonly': True}), 403
        return f(*args, **kwargs)
    return decorated_function


@app.context_processor
def inject_blocklist_perm():
    """Expose can_manage_blocklist to all templates (controls the block button in
    the customer modal). Per-user permission — see User.can_manage_blocklist."""
    try:
        ok = bool(getattr(current_user, 'is_authenticated', False)
                  and current_user.can_manage_blocklist())
    except Exception:
        ok = False
    return {'can_manage_blocklist': ok}


@app.context_processor
def inject_editable_sources():
    """Expose the current user's EDIT scope to every template so the UI can hide
    order-edit buttons on sites the user can't edit. Emitted as JSON: `null` means
    unrestricted (edit all); otherwise a list of editable source URLs. The backend
    @order_site_editable guard is still the real enforcement — this is just UX."""
    try:
        if getattr(current_user, 'is_authenticated', False):
            scope = get_user_editable_sources(current_user)
        else:
            scope = set()
    except Exception:
        scope = None  # fail open in the UI; backend still guards
    return {'editable_sources_json': json.dumps(None if scope is None else sorted(scope))}


def super_admin_required(f):
    """Decorator: only the built-in super admin (username == 'admin') may access."""
    from functools import wraps
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not current_user.is_authenticated:
            return jsonify({'error': '未登录'}), 401
        if getattr(current_user, 'username', None) != 'admin':
            return jsonify({'error': '该操作仅超级管理员可用'}), 403
        return f(*args, **kwargs)
    return decorated_function


def shipper_required(f):
    """Decorator to require shipping EDIT permission (operate)."""
    from functools import wraps
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not current_user.is_authenticated:
            return redirect(url_for('login'))
        if not current_user.can_ship():
            return jsonify({'error': '无发货权限', 'permission': 'shipping'}), 403
        return f(*args, **kwargs)
    return decorated_function


def shipping_view_required(f):
    """Decorator for READ-ONLY shipping endpoints (page, lists, print, export,
    查物流). Allows shipping VIEW or EDIT. Mutation endpoints keep
    shipper_required (edit only)."""
    from functools import wraps
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not current_user.is_authenticated:
            return redirect(url_for('login'))
        if not current_user.can_view_shipping():
            return jsonify({'error': '无发货查看权限', 'permission': 'shipping_view'}), 403
        return f(*args, **kwargs)
    return decorated_function


def product_manager_required(f):
    """Decorator to require product-manage permission. Super admin passes through."""
    from functools import wraps
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not current_user.is_authenticated:
            return redirect(url_for('login'))
        if not current_user.can_manage_products():
            # If this is an API call, return JSON; otherwise return HTML 403
            if request.path.startswith('/api/'):
                return jsonify({'error': '无产品管理权限', 'permission': 'product_manager'}), 403
            return render_template_string('''
<!DOCTYPE html><html lang="zh"><head><meta charset="UTF-8">
<title>访问受限</title>
<style>body{background:#0f0f23;color:#e0e0e0;font-family:sans-serif;display:flex;align-items:center;justify-content:center;min-height:100vh;margin:0;}
.box{text-align:center;padding:60px;background:rgba(255,255,255,0.05);border-radius:24px;max-width:480px;}
h1{color:#ef4444;margin:0 0 16px;}p{color:#9ca3af;line-height:1.6;}a{color:#a78bfa;}</style></head>
<body><div class="box"><h1>无产品管理权限</h1><p>请联系超级管理员授予「产品管理」权限。</p>
<a href="{{ url_for('dashboard') }}">返回首页</a></div></body></html>
'''), 403
        return f(*args, **kwargs)
    return decorated_function


def costs_view_required(f):
    """Decorator to require cost-view permission. Super admin passes through.
    For API endpoints returns JSON 403; for pages returns a styled HTML 403."""
    from functools import wraps
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not current_user.is_authenticated:
            return redirect(url_for('login'))
        if not current_user.can_view_costs():
            if request.path.startswith('/api/'):
                return jsonify({'error': '无成本管理查看权限', 'permission': 'view_costs'}), 403
            return render_template_string('''
<!DOCTYPE html><html lang="zh"><head><meta charset="UTF-8">
<title>访问受限</title>
<style>body{background:#0f0f23;color:#e0e0e0;font-family:sans-serif;display:flex;align-items:center;justify-content:center;min-height:100vh;margin:0;}
.box{text-align:center;padding:60px;background:rgba(255,255,255,0.05);border-radius:24px;max-width:480px;}
h1{color:#ef4444;margin:0 0 16px;}p{color:#9ca3af;line-height:1.6;}a{color:#a78bfa;}</style></head>
<body><div class="box"><h1>无成本管理查看权限</h1><p>请联系超级管理员授予「成本管理 · 查看」权限。</p>
<a href="{{ url_for('dashboard') }}">返回首页</a></div></body></html>
'''), 403
        return f(*args, **kwargs)
    return decorated_function


def costs_edit_required(f):
    """Decorator to require cost-edit permission (POST/PUT/DELETE on product_costs).
    Super admin passes through. Otherwise requires can_edit_costs flag."""
    from functools import wraps
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not current_user.is_authenticated:
            return jsonify({'error': '未登录'}), 401
        if not current_user.can_edit_costs():
            return jsonify({'error': '无成本管理编辑权限', 'permission': 'edit_costs'}), 403
        return f(*args, **kwargs)
    return decorated_function


def _user_allowed_warehouse_ids(for_view=False):
    """Return the list of warehouse_ids the current user is scoped to for cost
    management. Drives the country/warehouse scope for partner-bound users.

    Two scopes:
      • EDIT scope  (for_view=False, default) — which warehouses' costs the user
        may add/edit/delete. A non-partner non-admin gets [] (manages none).
      • VIEW scope  (for_view=True) — which warehouses' costs the user may SEE.
        Viewing is independently gated by can_view_costs; an internal read-only
        viewer (no partner binding) should see ALL costs even though they edit
        none, so the no-partner case returns None (unrestricted) instead of [].

    Returns:
      None  → unrestricted (super admin, admin role; or for_view + no binding)
      [...] → restricted to these warehouse ids (non-admin with partner bindings)
      []    → no access (edit scope, non-admin without any partner binding)

    Logic:
      • super admin (username 'admin')                → None
      • admin role (e.g. internal finance)            → None
      • non-admin with partner_users binding(s)       → warehouses in the
        countries of all bound partners' bound sites
      • non-admin without partner binding             → None if for_view else []
    """
    if not current_user.is_authenticated:
        return []
    if current_user.username == 'admin':
        return None
    conn = get_db_connection()
    try:
        u = conn.execute('SELECT role FROM users WHERE id = ?', (current_user.id,)).fetchone()
        if u and u['role'] == 'admin':
            return None
        partner_ids = [r['partner_id'] for r in conn.execute(
            'SELECT partner_id FROM partner_users WHERE user_id = ?', (current_user.id,)
        ).fetchall()]
        if not partner_ids:
            return None if for_view else []
        placeholders = ','.join(['?'] * len(partner_ids))
        countries = [r['country'] for r in conn.execute(f'''
            SELECT DISTINCT s.country FROM partner_sites ps
            JOIN sites s ON s.id = ps.site_id
            WHERE ps.partner_id IN ({placeholders})
              AND s.country IS NOT NULL AND s.country != ''
        ''', partner_ids).fetchall()]
        if not countries:
            return []
        cph = ','.join(['?'] * len(countries))
        return [r['id'] for r in conn.execute(
            f'SELECT id FROM warehouses WHERE country IN ({cph})', countries
        ).fetchall()]
    finally:
        conn.close()


def _check_warehouse_scope(warehouse_id):
    """Returns True if current user can manage costs for this warehouse_id."""
    if warehouse_id is None:
        return False
    allowed = _user_allowed_warehouse_ids()
    if allowed is None:
        return True
    try:
        return int(warehouse_id) in allowed
    except (TypeError, ValueError):
        return False


def report_viewer_required(f):
    """Decorator to require report viewing permission"""
    from functools import wraps
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not current_user.is_authenticated:
            return redirect(url_for('login'))
        if not current_user.can_view_report():
            return render_template_string('''
<!DOCTYPE html>
<html lang="zh">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>访问受限 - WooCommerce 订单分析</title>
    <link href="https://cdn.jsdelivr.net/npm/bootstrap-icons@1.10.0/font/bootstrap-icons.css" rel="stylesheet">
    <style>
        * { box-sizing: border-box; margin: 0; padding: 0; }
        body {
            min-height: 100vh;
            display: flex;
            align-items: center;
            justify-content: center;
            background: linear-gradient(135deg, #1a1a2e 0%, #16213e 50%, #0f3460 100%);
            font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", Roboto, sans-serif;
            color: #e0e0e0;
        }
        .container {
            text-align: center;
            padding: 60px 40px;
            background: rgba(255, 255, 255, 0.05);
            border-radius: 24px;
            backdrop-filter: blur(10px);
            border: 1px solid rgba(255, 255, 255, 0.1);
            max-width: 480px;
        }
        .icon {
            width: 120px;
            height: 120px;
            margin: 0 auto 30px;
            background: linear-gradient(135deg, rgba(239, 68, 68, 0.2), rgba(239, 68, 68, 0.1));
            border-radius: 50%;
            display: flex;
            align-items: center;
            justify-content: center;
            border: 2px solid rgba(239, 68, 68, 0.3);
        }
        .icon i { font-size: 48px; color: #ef4444; }
        h1 {
            font-size: 28px;
            font-weight: 600;
            margin-bottom: 16px;
            background: linear-gradient(135deg, #f87171, #ef4444);
            -webkit-background-clip: text;
            -webkit-text-fill-color: transparent;
        }
        p { color: #9ca3af; font-size: 16px; line-height: 1.6; margin-bottom: 30px; }
        .btn {
            display: inline-flex;
            align-items: center;
            gap: 8px;
            padding: 14px 32px;
            background: linear-gradient(135deg, #3b82f6, #2563eb);
            color: white;
            text-decoration: none;
            border-radius: 12px;
            font-weight: 500;
            transition: all 0.3s ease;
        }
        .btn:hover { transform: translateY(-2px); box-shadow: 0 10px 20px rgba(59, 130, 246, 0.3); }
    </style>
</head>
<body>
    <div class="container">
        <div class="icon"><i class="bi bi-file-earmark-lock"></i></div>
        <h1>访问受限</h1>
        <p>您没有查看报告的权限。<br>请联系管理员申请权限。</p>
        <a href="/" class="btn"><i class="bi bi-house"></i> 返回首页</a>
    </div>
</body>
</html>
            '''), 403
        return f(*args, **kwargs)
    return decorated_function


# ============== ROUTES ==============

@app.route('/login', methods=['GET', 'POST'])
def login():
    if current_user.is_authenticated:
        return redirect(url_for('dashboard'))
    
    if request.method == 'POST':
        username = request.form.get('username', '')
        password = request.form.get('password', '')
        
        conn = get_db_connection()
        user_row = conn.execute('SELECT * FROM users WHERE username = ?', (username,)).fetchone()
        conn.close()
        
        if user_row and check_password_hash(user_row['password_hash'], password):
            user = User(user_row['id'], username, user_row['name'], user_row['role'])
            login_user(user)
            flash('登录成功！', 'success')
            next_page = request.args.get('next')
            return redirect(next_page or url_for('dashboard'))
        else:
            flash('用户名或密码错误', 'error')
    
    return render_template('login.html')


@app.route('/logout')
@login_required
def logout():
    logout_user()
    flash('您已成功退出登录', 'info')
    return redirect(url_for('login'))


@app.route('/')
@login_required
def dashboard():
    """Main dashboard with statistics"""
    from datetime import date, timedelta
    conn = get_db_connection()
    
    # Get user's allowed sources for permission filtering
    allowed_sources = get_user_allowed_sources(current_user.id, current_user.is_admin(), current_user.is_viewer())
    source_clause, source_params = build_source_filter_clause(allowed_sources)
    
    # Get filters
    quick_date = request.args.get('quick_date', 'this_month')
    source_filter = request.args.get('source', '')
    manager_filter = request.args.get('manager', '')
    
    # Get all managers
    all_managers = get_all_managers()
    
    # Validate source_filter against allowed sources
    if source_filter and allowed_sources is not None and source_filter not in allowed_sources:
        source_filter = ''  # Reset invalid filter
    
    # Get available sources (filtered by permissions and manager)
    conn = get_db_connection()
    
    # Base source query
    source_query = 'SELECT DISTINCT source FROM orders'
    source_params = []
    source_conditions = []
    
    if allowed_sources is not None:
        if allowed_sources:
            placeholders = ', '.join(['?' for _ in allowed_sources])
            source_conditions.append(f'source IN ({placeholders})')
            source_params.extend(allowed_sources)
        else:
            source_conditions.append('1=0')
            
    if manager_filter:
        # Get sites managed by this manager
        manager_sites = conn.execute('SELECT url FROM sites WHERE manager = ?', (manager_filter,)).fetchall()
        manager_urls = [s['url'] for s in manager_sites]
        if manager_urls:
            placeholders = ', '.join(['?' for _ in manager_urls])
            source_conditions.append(f'source IN ({placeholders})')
            source_params.extend(manager_urls)
        else:
            source_conditions.append('1=0')
            
    if source_conditions:
        source_query += ' WHERE ' + ' AND '.join(source_conditions)
        
    source_query += ' ORDER BY source'
    sources = conn.execute(source_query, source_params).fetchall()
    
    # Process quick date filter
    today = date.today()
    date_from = ''
    date_to = today.isoformat()
    
    if quick_date == 'this_month':
        date_from = today.replace(day=1).isoformat()
    elif quick_date == 'last_month':
        first_of_this_month = today.replace(day=1)
        last_month_end = first_of_this_month - timedelta(days=1)
        last_month_start = last_month_end.replace(day=1)
        date_from = last_month_start.isoformat()
        date_to = last_month_end.isoformat()
    elif quick_date == 'last_quarter':
        date_from = (today - timedelta(days=90)).isoformat()
    elif quick_date == 'half_year':
        date_from = (today - timedelta(days=180)).isoformat()
    elif quick_date == 'one_year':
        date_from = (today - timedelta(days=365)).isoformat()
    elif quick_date == 'all':
        date_from = ''
        date_to = ''
    
    # Build filter conditions with permission check
    conditions = ['1=1']  # Base condition
    params = []
    
    # Add permission filter
    if allowed_sources is not None:
        if allowed_sources:
            placeholders = ', '.join(['?' for _ in allowed_sources])
            conditions.append(f'source IN ({placeholders})')
            params.extend(allowed_sources)
        else:
            conditions.append('1=0')  # No access
            
    # Add manager filter
    if manager_filter:
        # Re-use manager_urls from above or fetch if not done
        if 'manager_urls' not in locals():
            manager_sites = conn.execute('SELECT url FROM sites WHERE manager = ?', (manager_filter,)).fetchall()
            manager_urls = [s['url'] for s in manager_sites]
        
        if manager_urls:
            placeholders = ', '.join(['?' for _ in manager_urls])
            conditions.append(f'source IN ({placeholders})')
            params.extend(manager_urls)
        else:
            conditions.append('1=0')
    
    if date_from and date_to:
        conditions.append(f"date_created >= ? AND date_created <= ?")
        params.extend([date_from, date_to + 'T23:59:59'])
    elif date_from:
        conditions.append(f"date_created >= ?")
        params.append(date_from)
    if source_filter:
        conditions.append("source = ?")
        params.append(source_filter)
    
    date_condition = 'WHERE ' + ' AND '.join(conditions)
    
    # Get overall statistics
    stats = {}
    
    # Total orders
    stats['total_orders'] = conn.execute(f"SELECT COUNT(*) FROM orders {date_condition} AND status NOT IN ('checkout-draft', 'trash')", params).fetchone()[0]
    
    # Cancelled/failed orders count
    stats['cancelled_orders'] = conn.execute(f"SELECT COUNT(*) FROM orders {date_condition} AND status IN ('cancelled', 'failed')", params).fetchone()[0]

    # Undelivered (refused/returned) orders count + shipping loss aggregated by currency.
    # shipping_loss covers BOTH undelivered AND 问题退货 orders (shared column).
    stats['undelivered_orders'] = conn.execute(f"SELECT COUNT(*) FROM orders {date_condition} AND is_undelivered = 1", params).fetchone()[0]
    shipping_loss_raw = conn.execute(f'''
        SELECT currency, SUM(COALESCE(shipping_loss_amount, 0)) as loss
        FROM orders {date_condition} AND (is_undelivered = 1 OR is_problem_return = 1)
        GROUP BY currency
    ''', params).fetchall()
    stats['shipping_loss_by_currency'] = {row['currency'] or 'N/A': float(row['loss'] or 0) for row in shipping_loss_raw}

    # 问题退货 orders count + product (货值) loss aggregated by currency
    stats['problem_return_orders'] = conn.execute(f"SELECT COUNT(*) FROM orders {date_condition} AND is_problem_return = 1", params).fetchone()[0]
    product_loss_raw = conn.execute(f'''
        SELECT currency, SUM(COALESCE(product_loss_amount, 0)) as loss
        FROM orders {date_condition} AND is_problem_return = 1
        GROUP BY currency
    ''', params).fetchall()
    stats['product_loss_by_currency'] = {row['currency'] or 'N/A': float(row['loss'] or 0) for row in product_loss_raw}
    
    # Total revenue by currency - add status filter
    revenue_conditions = conditions.copy()
    revenue_conditions = conditions.copy()
    revenue_conditions.append(_revenue_status_cond())
    revenue_where = 'WHERE ' + ' AND '.join(revenue_conditions)
    
    # Valid orders count (for AOV calculation)
    stats['valid_orders'] = conn.execute(f'SELECT COUNT(*) FROM orders {revenue_where}', params).fetchone()[0]
    
    # Get revenue grouped by currency (total and net = total - shipping − shipping_loss − product_loss)
    # The revenue_where filter already excludes undelivered + problem-return orders
    # (so their totals don't get counted as revenue); their shipping_loss and
    # product_loss are real costs we ate, subtracted here so net matches /monthly.
    revenue_by_currency_raw = conn.execute(f'''
        SELECT currency, SUM(total) as revenue, SUM(shipping_total) as shipping
        FROM orders {revenue_where}
        GROUP BY currency
    ''', params).fetchall()
    stats['total_revenue_by_currency'] = {row['currency']: row['revenue'] or 0 for row in revenue_by_currency_raw}
    stats['net_revenue_by_currency'] = {
        row['currency']: (row['revenue'] or 0) - (row['shipping'] or 0)
                       - stats['shipping_loss_by_currency'].get(row['currency'] or 'N/A', 0)
                       - stats['product_loss_by_currency'].get(row['currency'] or 'N/A', 0)
        for row in revenue_by_currency_raw
    }

    # Also keep a simple total for backward compatibility
    stats['total_revenue'] = sum(stats['total_revenue_by_currency'].values())
    stats['net_revenue'] = sum(stats['net_revenue_by_currency'].values())
    
    # Orders by status with currency
    status_data_raw = conn.execute(f'''
        SELECT status, currency, COUNT(*) as count, SUM(total) as revenue
        FROM orders {date_condition}
        GROUP BY status, currency
        ORDER BY count DESC
    ''', params).fetchall()
    # Group by status, with currency breakdown
    status_dict = {}
    for row in status_data_raw:
        status = row['status']
        if status not in status_dict:
            status_dict[status] = {'status': status, 'count': 0, 'revenue_by_currency': {}}
        status_dict[status]['count'] += row['count']
        currency = row['currency'] or 'N/A'
        if currency not in status_dict[status]['revenue_by_currency']:
            status_dict[status]['revenue_by_currency'][currency] = 0
        status_dict[status]['revenue_by_currency'][currency] += row['revenue'] or 0
    status_data = sorted(status_dict.values(), key=lambda x: x['count'], reverse=True)
    
    # Orders by source with currency
    # Query 1: All orders for total count and total revenue (销售额)
    source_data_raw = conn.execute(f'''
        SELECT source, currency, COUNT(*) as count, SUM(total) as revenue
        FROM orders {date_condition} AND status NOT IN ('checkout-draft', 'trash')
        GROUP BY source, currency
    ''', params).fetchall()
    
    # Query 2: Only successful orders for net revenue (净销售额)
    source_success_conditions = conditions.copy()
    source_success_conditions = conditions.copy()
    source_success_conditions.append(_revenue_status_cond())
    source_success_where = 'WHERE ' + ' AND '.join(source_success_conditions)
    
    source_success_raw = conn.execute(f'''
        SELECT source, currency, SUM(total) as success_revenue, SUM(shipping_total) as success_shipping
        FROM orders {source_success_where}
        GROUP BY source, currency
    ''', params).fetchall()

    # Per-source/per-currency shipping loss from undelivered + 问题退货 orders. Same
    # date filter as the revenue query so the deduction lines up.
    source_loss_raw = conn.execute(f'''
        SELECT source, currency, SUM(COALESCE(shipping_loss_amount, 0)) as loss
        FROM orders {date_condition} AND (is_undelivered = 1 OR is_problem_return = 1)
        GROUP BY source, currency
    ''', params).fetchall()
    source_loss_lookup = {(r['source'], r['currency']): float(r['loss'] or 0) for r in source_loss_raw}

    # Per-source/per-currency 货值损失 from 问题退货 orders.
    source_product_loss_raw = conn.execute(f'''
        SELECT source, currency, SUM(COALESCE(product_loss_amount, 0)) as loss
        FROM orders {date_condition} AND is_problem_return = 1
        GROUP BY source, currency
    ''', params).fetchall()
    source_product_loss_lookup = {(r['source'], r['currency']): float(r['loss'] or 0) for r in source_product_loss_raw}

    # Build success data lookup
    success_lookup = {}
    for row in source_success_raw:
        key = (row['source'], row['currency'])
        success_lookup[key] = {
            'success_revenue': row['success_revenue'] or 0,
            'success_shipping': row['success_shipping'] or 0,
        }

    # Group by source - combine all orders data with success-only data
    source_dict = {}
    for row in source_data_raw:
        source = row['source']
        currency = row['currency']
        if source not in source_dict:
            source_dict[source] = {'source': source, 'count': 0, 'currency': currency, 'revenue': 0, 'net_revenue': 0}
        source_dict[source]['count'] += row['count']
        source_dict[source]['revenue'] += row['revenue'] or 0

        # Net revenue = success revenue − success shipping − this source's
        # shipping_loss (未送达 + 问题退货) − product_loss (问题退货).
        # Mirrors the /monthly view.
        success_data = success_lookup.get((source, currency), {'success_revenue': 0, 'success_shipping': 0})
        loss = source_loss_lookup.get((source, currency), 0)
        p_loss = source_product_loss_lookup.get((source, currency), 0)
        source_dict[source]['net_revenue'] += success_data['success_revenue'] - success_data['success_shipping'] - loss - p_loss
    
    source_data = list(source_dict.values())
    
    # Recent orders - apply permission filter, optionally source filter
    recent_conditions = ['1=1']
    recent_params = []
    
    if allowed_sources is not None:
        if allowed_sources:
            placeholders = ', '.join(['?' for _ in allowed_sources])
            recent_conditions.append(f'source IN ({placeholders})')
            recent_params.extend(allowed_sources)
        else:
            recent_conditions.append('1=0')
    
    if source_filter:
        recent_conditions.append('source = ?')
        recent_params.append(source_filter)

    if manager_filter:
        manager_sites = conn.execute('SELECT url FROM sites WHERE manager = ?', (manager_filter,)).fetchall()
        manager_urls = [s['url'] for s in manager_sites]
        if manager_urls:
            placeholders = ', '.join(['?' for _ in manager_urls])
            recent_conditions.append(f'source IN ({placeholders})')
            recent_params.extend(manager_urls)
        else:
            recent_conditions.append('1=0')
    
    recent_where = 'WHERE ' + ' AND '.join(recent_conditions)
    recent_orders = conn.execute(f'''
        SELECT id, number, status, total, shipping_total, currency, date_created, source, line_items, billing,
               payment_method, is_undelivered
        FROM orders
        {recent_where}
        ORDER BY date_created DESC
        LIMIT 10
    ''', recent_params).fetchall()
    
    # Process recent orders to get product count
    processed_orders = []
    for order in recent_orders:
        order_dict = dict(order)
        items = parse_json_field(order['line_items'])
        order_dict['product_count'] = sum(item.get('quantity', 0) for item in items) if isinstance(items, list) else 0
        
        # Parse customer info from billing
        billing = parse_json_field(order['billing'])
        order_dict['customer_name'] = f"{billing.get('first_name', '')} {billing.get('last_name', '')}".strip() if billing else ''
        order_dict['customer_email'] = billing.get('email', '') if billing else ''
        
        processed_orders.append(order_dict)
    
    # Trend data based on filter - daily for single month, monthly for longer periods
    trend_type = 'daily' if quick_date in ['this_month', 'last_month'] else 'monthly'
    
    # Trend data — historically SUM(total) mixed currencies (e.g. AUD+PLN)
    # producing a meaningless aggregate. Now we GROUP BY (period, currency)
    # and convert each bucket to CNY using the period's per-currency rate,
    # then sum to a single CNY number — same approach /monthly uses.
    success_cond = _revenue_status_cond()
    period_fmt = '%Y-%m-%d' if trend_type == 'daily' else '%Y-%m'
    trend_data_raw = conn.execute(f'''
        SELECT strftime('{period_fmt}', date_created) as period,
               COALESCE(currency, '') as currency,
               COUNT(*) as orders,
               COALESCE(SUM(total), 0) as revenue,
               COALESCE(SUM(CASE WHEN {success_cond} THEN total ELSE 0 END), 0)
                 - COALESCE(SUM(CASE WHEN {success_cond} THEN shipping_total ELSE 0 END), 0)
                 - COALESCE(SUM(CASE WHEN COALESCE(is_undelivered, 0) = 1
                                     THEN COALESCE(shipping_loss_amount, 0) ELSE 0 END), 0) as net,
               SUM(CASE WHEN {success_cond} THEN 1 ELSE 0 END) as success_orders
        FROM orders {date_condition} AND status NOT IN ('checkout-draft', 'trash')
        GROUP BY period, currency
        ORDER BY period
    ''', params).fetchall()

    # Aggregate to CNY per period using each currency's month-specific rate.
    # Per-currency raw amounts are also kept for tooltip drill-down.
    _trend_periods = {}
    _rate_cache = {}
    def _rate_for(curr, period):
        ym = period[:7] if period else None
        key = (curr, ym)
        if key not in _rate_cache:
            r, _ = get_cny_rate(curr, ym) if ym else (None, None)
            _rate_cache[key] = r
        return _rate_cache[key]

    for r in trend_data_raw:
        p = r['period']
        rec = _trend_periods.get(p)
        if rec is None:
            rec = {
                'period': p,
                'orders': 0,
                'success_orders': 0,
                'revenue_cny': 0.0,
                'net_cny': 0.0,
                'by_currency': {},
            }
            _trend_periods[p] = rec
        rec['orders'] += int(r['orders'] or 0)
        rec['success_orders'] += int(r['success_orders'] or 0)
        cur = r['currency'] or '-'
        rate = _rate_for(cur, p)
        rec['by_currency'][cur] = {
            'revenue': float(r['revenue'] or 0),
            'net': float(r['net'] or 0),
            'rate': rate,
        }
        if rate:
            rec['revenue_cny'] += float(r['revenue'] or 0) * rate
            rec['net_cny']     += float(r['net'] or 0) * rate
    trend_data = sorted(_trend_periods.values(), key=lambda d: d['period'])
    for rec in trend_data:
        rec['revenue_cny'] = round(rec['revenue_cny'], 2)
        rec['net_cny']     = round(rec['net_cny'], 2)
        # Back-compat: keep `revenue` field — mixed-currency value reserved
        # ONLY for legacy callers. New chart code MUST use revenue_cny / net_cny.
        rec['revenue'] = round(sum(c['revenue'] for c in rec['by_currency'].values()), 2)
        rec['net']     = round(sum(c['net']     for c in rec['by_currency'].values()), 2)
    
    # Get site managers mapping
    sites = conn.execute('SELECT url, manager FROM sites').fetchall()
    site_managers = {s['url']: s['manager'] or '' for s in sites}
    
    # Calculate CNY totals for revenue
    total_cny = 0
    net_cny = 0
    for currency, amount in stats['total_revenue_by_currency'].items():
        rate, _ = get_cny_rate(currency, date_to[:7] if date_to else (date_from[:7] if date_from else None))
        if rate:
            total_cny += amount * rate
            net_cny += stats['net_revenue_by_currency'].get(currency, 0) * rate
    stats['total_revenue_cny'] = round(total_cny, 2)
    stats['net_revenue_cny'] = round(net_cny, 2)
    
    # Add CNY conversion to source_data
    for source in source_data:
        currency = source.get('currency', 'PLN')
        # Get month from date filter or use current month
        month = date_to[:7] if date_to else (date_from[:7] if date_from else None)
        rate, _ = get_cny_rate(currency, month)
        source['rate_to_cny'] = rate if rate else None
        source['revenue_cny'] = round(source['revenue'] * rate, 2) if rate else None
        source['net_revenue_cny'] = round(source['net_revenue'] * rate, 2) if rate else None
    
    # Add CNY conversion to recent_orders
    for order in processed_orders:
        currency = order.get('currency', 'PLN')
        order_month = order['date_created'][:7] if order.get('date_created') else None
        rate, _ = get_cny_rate(currency, order_month)
        order['rate_to_cny'] = rate if rate else None
        order['total_cny'] = round(float(order['total'] or 0) * rate, 2) if rate else None
        order['net_total'] = float(order['total'] or 0) - float(order.get('shipping_total') or 0)
        order['net_total_cny'] = round(order['net_total'] * rate, 2) if rate else None
    
    # Get customer attributes for the recent orders
    customer_emails = list(set(o['customer_email'] for o in processed_orders if o.get('customer_email')))
    customer_attributes = {}
    
    if customer_emails:
        placeholders = ', '.join(['?' for _ in customer_emails])
        
        # Get manual quality settings
        manual_settings = conn.execute(f'''
            SELECT email, quality_tier FROM customer_settings 
            WHERE email IN ({placeholders})
        ''', customer_emails).fetchall()
        
        # Calculate attributes for each customer
        for email in customer_emails:
            if email not in customer_attributes:
                customer_attributes[email] = {}
            
            # Check manual setting
            manual_tier = 'auto'
            for row in manual_settings:
                if row['email'] == email:
                    manual_tier = row['quality_tier']
                    break
            
            # Get customer stats
            stats_row = conn.execute(f'''
                SELECT
                    COUNT(*) as total_orders,
                    SUM({_success_status_case()}) as successful_orders,
                    {_success_amount_case('total')} as total_spending,
                    SUM({_bad_order_case()}) as bad_orders,
                    SUM({_meaningful_order_case()}) as meaningful_orders,
                    MAX(date_created) as last_order_date,
                    MIN(date_created) as first_order_date
                FROM orders
                WHERE json_extract(billing, '$.email') = ?
            ''', (email,)).fetchone()

            total_orders = stats_row['total_orders'] or 0
            successful_orders = stats_row['successful_orders'] or 0
            total_spending = stats_row['total_spending'] or 0
            
            customer_attributes[email]['order_count'] = total_orders
            customer_attributes[email]['is_new'] = (total_orders <= 1)
            
            tier = manual_tier
            if tier == 'auto':
                avg_days_between = 0
                if total_orders > 1 and stats_row['first_order_date'] and stats_row['last_order_date']:
                    from datetime import datetime
                    try:
                        first_date = datetime.fromisoformat(stats_row['first_order_date'][:19])
                        last_date = datetime.fromisoformat(stats_row['last_order_date'][:19])
                        days_span = (last_date - first_date).days
                        avg_days_between = days_span / (total_orders - 1)
                    except:
                        avg_days_between = 0
                
                tier = calculate_customer_tier(
                    successful_orders, total_spending, avg_days_between,
                    bad_orders=stats_row['bad_orders'] or 0,
                    meaningful_orders=stats_row['meaningful_orders'] or 0)

            if tier == 'vip':
                customer_attributes[email]['quality'] = {'label': 'VIP', 'class': 'text-warning', 'icon': 'star-fill'}
            elif tier == 'good':
                customer_attributes[email]['quality'] = {'label': '优质', 'class': 'text-success', 'icon': 'gem'}
            elif tier == 'normal':
                customer_attributes[email]['quality'] = {'label': '普通', 'class': 'text-primary', 'icon': 'person-check'}
            elif tier == 'new':
                customer_attributes[email]['quality'] = {'label': '新客', 'class': 'text-info', 'icon': 'stars'}
            elif tier == 'bad':
                customer_attributes[email]['quality'] = {'label': '劣质', 'class': 'text-danger', 'icon': 'x-circle'}

    conn.close()
    
    # Calculate overall CNY rate for dashboard
    cny_rate = 0
    if stats.get('net_revenue'):
        cny_rate = round(stats.get('net_revenue_cny', 0) / stats['net_revenue'], 4)
    
    # Get sites with API errors
    conn = get_db_connection()
    api_error_sites = conn.execute("SELECT url FROM sites WHERE api_status = 'error'").fetchall()
    conn.close()
        
    return render_template('dashboard.html',
                         stats=stats,
                         status_data=status_data,
                         source_data=source_data,
                         recent_orders=processed_orders,
                         trend_data=trend_data,
                         trend_type=trend_type,
                         quick_date=quick_date,
                         sources=sources,
                         source_filter=source_filter,
                         manager_filter=manager_filter,
                         all_managers=all_managers,
                         site_managers=site_managers,
                         customer_attributes=customer_attributes,
                         cny_rate=cny_rate,
                         api_error_sites=api_error_sites)


@app.route('/orders')
@login_required
def orders():
    """Order list with filtering and summary statistics"""
    from datetime import date, timedelta
    conn = get_db_connection()
    
    # Get user's allowed sources for permission filtering
    allowed_sources = get_user_allowed_sources(current_user.id, current_user.is_admin(), current_user.is_viewer())
    
    # Get filter parameters
    source_filter = request.args.get('source', '')
    status_filter = request.args.get('status', '')
    date_from = request.args.get('date_from', '')
    date_to = request.args.get('date_to', '')
    search = request.args.get('search', '')
    quick_date = request.args.get('quick_date', '')
    manager_filter = request.args.get('manager', '')
    country_filter = request.args.get('country', '')
    page = int(request.args.get('page', 1))
    per_page = 20
    
    # Default to 'this_month' if no filters are active
    if not any([quick_date, date_from, date_to, search]):
        quick_date = 'this_month'
    
    # Get all managers
    all_managers = get_all_managers()
    
    # Get all countries
    all_countries = conn.execute('SELECT DISTINCT country FROM sites WHERE country IS NOT NULL AND country != "" ORDER BY country').fetchall()
    all_countries = [c['country'] for c in all_countries]
    
    # Validate source_filter against allowed sources
    if source_filter and allowed_sources is not None and source_filter not in allowed_sources:
        source_filter = ''
    
    # Process quick date filter
    today = date.today()
    if quick_date == 'this_month':
        date_from = today.replace(day=1).isoformat()
        date_to = today.isoformat()
    elif quick_date == 'last_month':
        first_of_this_month = today.replace(day=1)
        last_month_end = first_of_this_month - timedelta(days=1)
        last_month_start = last_month_end.replace(day=1)
        date_from = last_month_start.isoformat()
        date_to = last_month_end.isoformat()
    elif quick_date == 'last_quarter':
        date_from = (today - timedelta(days=90)).isoformat()
        date_to = today.isoformat()
    elif quick_date == 'half_year':
        date_from = (today - timedelta(days=180)).isoformat()
        date_to = today.isoformat()
    elif quick_date == 'one_year':
        date_from = (today - timedelta(days=365)).isoformat()
        date_to = today.isoformat()
    
    # Build base conditions with permission filter
    conditions = []
    params = []
    
    # Add permission filter first
    if allowed_sources is not None:
        if allowed_sources:
            placeholders = ', '.join(['?' for _ in allowed_sources])
            conditions.append(f'source IN ({placeholders})')
            params.extend(allowed_sources)
        else:
            conditions.append('1=0')  # No access
            
    # Add manager filter
    if manager_filter:
        manager_sites = conn.execute('SELECT url FROM sites WHERE manager = ?', (manager_filter,)).fetchall()
        manager_urls = [s['url'] for s in manager_sites]
        
        if manager_urls:
            placeholders = ', '.join(['?' for _ in manager_urls])
            conditions.append(f'source IN ({placeholders})')
            params.extend(manager_urls)
        else:
            conditions.append('1=0')
            
    # Add country filter
    if country_filter:
        country_sites = conn.execute('SELECT url FROM sites WHERE country = ?', (country_filter,)).fetchall()
        country_urls = [s['url'] for s in country_sites]
        if country_urls:
            placeholders = ', '.join(['?' for _ in country_urls])
            conditions.append(f'source IN ({placeholders})')
            params.extend(country_urls)
        else:
            conditions.append('1=0')
    
    if source_filter:
        conditions.append('source = ?')
        params.append(source_filter)
    if status_filter == 'undelivered':
        # 'undelivered' is stored on a separate flag column, not in status
        conditions.append('is_undelivered = 1')
    elif status_filter:
        conditions.append('status = ?')
        params.append(status_filter)
    else:
        # Default behavior: exclude draft/trash unless filtered
        conditions.append("status NOT IN ('checkout-draft', 'trash')")
        
    if date_from:
        conditions.append('date_created >= ?')
        params.append(date_from)
    if date_to:
        conditions.append('date_created <= ?')
        params.append(date_to + 'T23:59:59')
    if search:
        like_term = f'%{search}%'
        # When the input is long enough to plausibly be a tracking number, email,
        # or customer detail, also search inside JSON columns (billing/shipping →
        # customer email/name/phone; AST plugin → meta_data; VillaTheme →
        # line_items; generic shipping_lines) and our shipping_logs. Short inputs
        # stay on the fast number/id path to avoid false positives.
        if len(search) >= 6:
            conditions.append('''(
                number LIKE ?
                OR id LIKE ?
                OR billing LIKE ?
                OR shipping LIKE ?
                OR meta_data LIKE ?
                OR line_items LIKE ?
                OR shipping_lines LIKE ?
                OR id IN (SELECT order_id FROM shipping_logs WHERE tracking_number LIKE ?)
            )''')
            params.extend([like_term] * 8)
        else:
            conditions.append('(number LIKE ? OR id LIKE ?)')
            params.extend([like_term, like_term])

    where_clause = ' WHERE ' + ' AND '.join(conditions) if conditions else ''
    
    # Summary statistics by source (with currency).
    # shipping_loss covers both undelivered and problem-return orders because
    # both flows write to the same shipping_loss_amount column.
    stats_query = f'''
        SELECT source, currency,
            COUNT(*) as total_orders,
            SUM(total) as total_amount,
            SUM(shipping_total) as total_shipping,
            SUM(CASE WHEN {_revenue_status_cond()} THEN 1 ELSE 0 END) as success_orders,
            {_revenue_amount_case('total')} as success_amount,
            {_revenue_amount_case('shipping_total')} as success_shipping,
            SUM(CASE WHEN status='failed' THEN 1 ELSE 0 END) as failed_orders,
            SUM(CASE WHEN status='cancelled' THEN 1 ELSE 0 END) as cancelled_orders,
            SUM(CASE WHEN is_undelivered = 1 THEN 1 ELSE 0 END) as undelivered_orders,
            SUM(CASE WHEN COALESCE(is_undelivered,0) = 1 OR COALESCE(is_problem_return,0) = 1
                     THEN COALESCE(shipping_loss_amount, 0) ELSE 0 END) as shipping_loss,
            SUM(CASE WHEN is_problem_return = 1 THEN 1 ELSE 0 END) as problem_return_orders,
            SUM(CASE WHEN is_problem_return = 1 THEN COALESCE(product_loss_amount, 0) ELSE 0 END) as product_loss
        FROM orders {where_clause} GROUP BY source, currency
    '''
    summary_raw = conn.execute(stats_query, params).fetchall()
    summary_stats = [dict(row) for row in summary_raw]
    
    # Get product quantities
    items_query = f'SELECT line_items, status, source FROM orders {where_clause}'
    all_items = conn.execute(items_query, params).fetchall()
    
    source_products = {}
    for row in all_items:
        src = row['source']
        if src not in source_products:
            source_products[src] = {'total': 0, 'success': 0, 'failed': 0, 'cancelled': 0}
        items = parse_json_field(row['line_items'])
        qty = sum(i.get('quantity', 0) for i in items) if isinstance(items, list) else 0
        source_products[src]['total'] += qty
        if row['status'] == 'failed':
            source_products[src]['failed'] += qty
        elif row['status'] == 'cancelled':
            source_products[src]['cancelled'] += qty
        else:
            source_products[src]['success'] += qty
    
    for stat in summary_stats:
        src = stat['source']
        prods = source_products.get(src, {})
        stat['total_products'] = prods.get('total', 0)
        stat['success_products'] = prods.get('success', 0)
        stat['failed_products'] = prods.get('failed', 0)
        stat['cancelled_products'] = prods.get('cancelled', 0)
        # Net = product revenue minus collected shipping fees minus
        # shipping/product losses from undelivered + problem-return orders.
        # Mirrors the /monthly view's formula so the two summaries cross-validate.
        stat['shipping_loss'] = float(stat.get('shipping_loss') or 0)
        stat['product_loss'] = float(stat.get('product_loss') or 0)
        stat['success_net_amount'] = (stat['success_amount'] or 0) - (stat['success_shipping'] or 0) - stat['shipping_loss'] - stat['product_loss']

    totals = {
        'total_orders': sum(s['total_orders'] or 0 for s in summary_stats),
        'total_products': sum(s['total_products'] for s in summary_stats),
        'total_amount': sum(s['total_amount'] or 0 for s in summary_stats),
        'success_orders': sum(s['success_orders'] or 0 for s in summary_stats),
        'success_products': sum(s['success_products'] for s in summary_stats),
        'success_amount': sum(s['success_amount'] or 0 for s in summary_stats),
        'success_net_amount': sum(s['success_net_amount'] for s in summary_stats),
        'failed_orders': sum(s['failed_orders'] or 0 for s in summary_stats),
        'cancelled_orders': sum(s['cancelled_orders'] or 0 for s in summary_stats),
        'undelivered_orders': sum(s.get('undelivered_orders') or 0 for s in summary_stats),
        'shipping_loss': sum(s.get('shipping_loss') or 0 for s in summary_stats),
        'problem_return_orders': sum(s.get('problem_return_orders') or 0 for s in summary_stats),
        'product_loss': sum(s.get('product_loss') or 0 for s in summary_stats),
    }
    # If every row in the summary shares one currency (typically because the
    # user's permission scope is a single country, e.g. 澳洲发货员 only sees AUD
    # sites), the 合计 row can show real currency amounts instead of "(混合货币)".
    # Empty filter result → no currency to pick → fall back to the placeholder.
    currencies_in_view = {s.get('currency') for s in summary_stats if s.get('currency')}
    totals['uniform_currency'] = next(iter(currencies_in_view)) if len(currencies_in_view) == 1 else None
    
    # Get all available months for pagination
    months_query = f'''
        SELECT DISTINCT strftime('%Y-%m', date_created) as month 
        FROM orders {where_clause} 
        ORDER BY month DESC
    '''
    available_months = [row['month'] for row in conn.execute(months_query, params).fetchall()]
    
    # Get current month from page parameter
    current_month = request.args.get('month', '')
    if not current_month:
        # A search targets a specific order/tracking number whose match may live in
        # any month. Defaulting to the latest month would hide a hit in an older one
        # (e.g. searching "105" returns #105 from 2026-05, but only the newest
        # month's orders render, so it silently disappears). When a search is active,
        # show all matches; the month tabs still let the user narrow down afterwards.
        # No search → keep the latest-month default for normal browsing.
        if search:
            current_month = 'all'
        elif available_months:
            current_month = available_months[0]  # Default to most recent month
    
    # Get orders for the selected month (or all if month='all')
    if current_month == 'all':
        # Show all orders matching the filters (no month filter)
        orders_query = f'SELECT * FROM orders {where_clause} ORDER BY date_created DESC'
        orders_data = conn.execute(orders_query, params).fetchall()
    elif current_month:
        month_conditions = conditions.copy() if conditions else []
        month_params = params.copy()
        month_conditions.append("strftime('%Y-%m', date_created) = ?")
        month_params.append(current_month)
        month_where = ' WHERE ' + ' AND '.join(month_conditions)
        
        orders_query = f'SELECT * FROM orders {month_where} ORDER BY date_created DESC'
        orders_data = conn.execute(orders_query, month_params).fetchall()
    else:
        orders_data = []
    
    # Get total count for display
    count_query = f'SELECT COUNT(*) FROM orders {where_clause}'
    total = conn.execute(count_query, params).fetchone()[0]
    
    # Pre-build risk index so every order in this listing gets cheap
    # repeat-freeloader lookups (problem-return + undelivered).
    risk_idx = _build_risk_index(conn)

    processed_orders = []
    # 嘻嘻嘻嘻嘻
    for order in orders_data:
        od = dict(order)
        items = parse_json_field(order['line_items'])
        if isinstance(items, list):
            od['product_count'] = sum(i.get('quantity', 0) for i in items)
            od['products'] = [{
                # WooCommerce stores line-item names HTML-encoded (e.g. "&amp;");
                # decode so Jinja's autoescape doesn't double-encode them.
                'name': html.unescape(i.get('name', '') or ''),
                'quantity': i.get('quantity', 0),
                'total': float(i.get('total', 0))
            } for i in items]
        else:
            od['product_count'] = 0
            od['products'] = []
        billing = parse_json_field(order['billing'])
        shipping = parse_json_field(order['shipping']) if 'shipping' in od else {}
        od['customer_name'] = f"{billing.get('first_name', '')} {billing.get('last_name', '')}".strip()
        od['customer_email'] = billing.get('email', '')

        # Extract DPD address for list display
        meta_data = parse_json_field(order['meta_data'])
        custom_fields = extract_custom_billing_fields(meta_data)

        dpd_address = ''
        if custom_fields.get('dpd_street') or custom_fields.get('dpd_city'):
             parts = [
                f"{custom_fields.get('dpd_street', '')} {custom_fields.get('dpd_house', '')}".strip(),
                custom_fields.get('dpd_zip', ''),
                custom_fields.get('dpd_city', '')
             ]
             dpd_address = ', '.join(filter(None, parts))
        od['dpd_address'] = dpd_address

        # Surface prior-freeloader risk for this order's customer.
        od['customer_risk'] = _assess_customer_risk(billing, shipping, risk_idx, current_order_id=order['id'])

        processed_orders.append(od)
    
    # Get available sources (filtered by permissions and manager)
    conn = get_db_connection()
    
    # Base source query
    source_query = 'SELECT DISTINCT source FROM orders'
    source_params = []
    source_conditions = []
    
    if allowed_sources is not None:
        if allowed_sources:
            placeholders = ', '.join(['?' for _ in allowed_sources])
            source_conditions.append(f'source IN ({placeholders})')
            source_params.extend(allowed_sources)
        else:
            source_conditions.append('1=0')
            
    if manager_filter:
        # Get sites managed by this manager
        manager_sites = conn.execute('SELECT url FROM sites WHERE manager = ?', (manager_filter,)).fetchall()
        manager_urls = [s['url'] for s in manager_sites]
        if manager_urls:
            placeholders = ', '.join(['?' for _ in manager_urls])
            source_conditions.append(f'source IN ({placeholders})')
            source_params.extend(manager_urls)
        else:
            source_conditions.append('1=0')
            
    if source_conditions:
        source_query += ' WHERE ' + ' AND '.join(source_conditions)
        
    source_query += ' ORDER BY source'
    sources = conn.execute(source_query, source_params).fetchall()
    statuses = [dict(s) for s in conn.execute('SELECT DISTINCT status FROM orders').fetchall()]
    # Surface 'undelivered' as a synthetic filter option (it's stored on is_undelivered flag, not status column)
    if not any(s.get('status') == 'undelivered' for s in statuses):
        statuses.append({'status': 'undelivered'})
    
    # Get site managers mapping (url -> manager)
    sites = conn.execute('SELECT url, manager FROM sites').fetchall()
    site_managers = {s['url']: s['manager'] or '' for s in sites}
    
    # Add CNY conversion to summary_stats
    # First, aggregate raw amounts by currency to avoid rounding errors
    raw_totals_by_currency = {}
    for stat in summary_stats:
        currency = stat.get('currency', 'PLN')
        if currency not in raw_totals_by_currency:
            raw_totals_by_currency[currency] = 0
        raw_totals_by_currency[currency] += (stat['success_net_amount'] or 0)
    
    # Calculate total CNY from aggregated raw amounts (sum first, then round)
    totals_cny = 0
    for currency, amount in raw_totals_by_currency.items():
        rate, _ = get_cny_rate(currency, current_month)
        if rate:
            totals_cny += amount * rate
    
    # Now add individual stat CNY values for display
    for stat in summary_stats:
        currency = stat.get('currency', 'PLN')
        # Use current_month for rate lookup
        rate, _ = get_cny_rate(currency, current_month)
        stat['rate_to_cny'] = rate
        if rate:
            stat['total_amount_cny'] = round((stat['total_amount'] or 0) * rate, 2)
            stat['success_amount_cny'] = round((stat['success_amount'] or 0) * rate, 2)
            stat['success_net_amount_cny'] = round((stat['success_net_amount'] or 0) * rate, 2)
        else:
            stat['total_amount_cny'] = None
            stat['success_amount_cny'] = None
            stat['success_net_amount_cny'] = None
    
    totals['success_net_amount_cny'] = round(totals_cny, 2)
    
    # Add CNY conversion to processed_orders
    for order in processed_orders:
        currency = order.get('currency', 'PLN')
        order_month = order['date_created'][:7] if order.get('date_created') else current_month
        rate, _ = get_cny_rate(currency, order_month)
        order['rate_to_cny'] = rate
        order['total_cny'] = round(float(order['total'] or 0) * rate, 2) if rate else None
        order['net_total'] = float(order['total'] or 0) - float(order.get('shipping_total') or 0)
        order['net_total_cny'] = round(order['net_total'] * rate, 2) if rate else None

    # Aggregate column totals over the orders currently rendered (the
    # selected month tab). Mirrors the math in the /monthly view + the
    # filter summary table at the top so all three numbers agree:
    #
    #   订单金额合计  = sum(total) over ALL orders (gross — includes undelivered)
    #   运费合计     = sum(shipping_total) over ALL orders
    #   净额合计     = sum(net_total = total - shipping) ONLY for orders we actually
    #                  collect money on (excludes undelivered — their goods come back
    #                  unsold; we never get the order amount)
    #   运费损失合计 = sum(shipping_loss_amount) for undelivered orders
    #                  (the carrier kept this even though the package returned)
    #   实际净额    = 净额合计 - 运费损失合计
    #
    # Without the undelivered exclusion the per-row "would-be net" of
    # returned packages was inflating 净额合计 — those numbers should
    # never count as revenue, only their shipping cost is a real loss.
    displayed_totals = {
        'order_count': len(processed_orders),
        'undelivered_count': 0,
        'problem_return_count': 0,
        'product_count': 0,
        'by_currency': {},
        'net_cny': 0.0,
        'shipping_loss_cny': 0.0,
        'product_loss_cny': 0.0,
        'final_net_cny': 0.0,  # net minus shipping_loss/product_loss — matches monthly view
    }
    # Mirrors _revenue_status_cond() in SQL — same exclusion list so the
    # order-list totals row matches the filter-summary "净金额" column.
    # Pre-load each site's cod_on_hold_is_shipped flag so we can replicate the
    # SQL's EXISTS check in-memory without hitting the DB per order. The set
    # contains the source URLs where on-hold IS treated as shipped (PL by
    # default; admins can override per site in /settings).
    on_hold_shipped_sources = {
        r['url'] for r in conn.execute(
            "SELECT url FROM sites WHERE cod_on_hold_is_shipped = 1"
        ).fetchall()
    }
    def _is_revenue_order(o):
        st = o.get('status') or ''
        pm = (o.get('payment_method') or 'cod')
        if st in ('failed', 'cancelled', 'checkout-draft', 'trash', 'cheat'):
            return False
        if st == 'pending' and pm != 'cod':
            return False
        if st == 'on-hold' and pm == 'bacs':
            return False
        if st == 'on-hold' and o.get('source') not in on_hold_shipped_sources:
            # AU/AE etc — on-hold here means "received, waiting", not shipped.
            return False
        if o.get('is_undelivered'):
            return False
        if o.get('is_problem_return'):
            return False
        return True

    for o in processed_orders:
        cur = o.get('currency') or 'N/A'
        bucket = displayed_totals['by_currency'].setdefault(cur, {
            'amount': 0.0, 'shipping': 0.0, 'net': 0.0,
            'shipping_loss': 0.0, 'product_loss': 0.0,
            'undelivered_count': 0, 'problem_return_count': 0,
            'order_count': 0,
        })
        bucket['order_count'] += 1
        bucket['amount']   += float(o.get('total') or 0)
        bucket['shipping'] += float(o.get('shipping_total') or 0)
        displayed_totals['product_count'] += int(o.get('product_count') or 0)

        rate = float(o.get('rate_to_cny') or 0)

        if o.get('is_undelivered'):
            # Undelivered: package came back, goods resellable, but carrier
            # kept the shipping fee. Only the shipping_loss counts as a real
            # impact — DON'T add net_total (would imply we earned the
            # product revenue, which we didn't).
            displayed_totals['undelivered_count'] += 1
            bucket['undelivered_count'] += 1
            loss = float(o.get('shipping_loss_amount') or 0)
            bucket['shipping_loss'] += loss
            if rate:
                displayed_totals['shipping_loss_cny'] += loss * rate
            # If the same order also has a problem-return marker, surface
            # its product_loss too (shipping_loss already attributed above).
            if o.get('is_problem_return'):
                displayed_totals['problem_return_count'] += 1
                bucket['problem_return_count'] += 1
                p_loss = float(o.get('product_loss_amount') or 0)
                bucket['product_loss'] += p_loss
                if rate:
                    displayed_totals['product_loss_cny'] += p_loss * rate
        elif o.get('is_problem_return'):
            # Problem return: package came back but contents were wrong/
            # missing/damaged. Both shipping AND product value are lost.
            displayed_totals['problem_return_count'] += 1
            bucket['problem_return_count'] += 1
            s_loss = float(o.get('shipping_loss_amount') or 0)
            p_loss = float(o.get('product_loss_amount') or 0)
            bucket['shipping_loss'] += s_loss
            bucket['product_loss'] += p_loss
            if rate:
                displayed_totals['shipping_loss_cny'] += s_loss * rate
                displayed_totals['product_loss_cny'] += p_loss * rate
        elif _is_revenue_order(o):
            # Successful (or pending delivery from a paid order) — its
            # product revenue counts.
            bucket['net'] += float(o.get('net_total') or 0)
            if o.get('net_total_cny'):
                displayed_totals['net_cny'] += float(o['net_total_cny'])
        # else: failed / cancelled / awaiting-bank-transfer — money never
        # came in, so contributes nothing. Per-row 净额 display still shows
        # the would-be number but it's not summed here.
    displayed_totals['net_cny'] = round(displayed_totals['net_cny'], 2)
    displayed_totals['shipping_loss_cny'] = round(displayed_totals['shipping_loss_cny'], 2)
    displayed_totals['product_loss_cny'] = round(displayed_totals['product_loss_cny'], 2)
    displayed_totals['final_net_cny'] = round(
        displayed_totals['net_cny']
        - displayed_totals['shipping_loss_cny']
        - displayed_totals['product_loss_cny'], 2)

    # Get customer attributes for the displayed orders
    customer_emails = list(set(o['customer_email'] for o in processed_orders if o.get('customer_email')))
    customer_attributes = {}
    
    if customer_emails:
        placeholders = ', '.join(['?' for _ in customer_emails])
        
        # 1. Get manual quality settings
        manual_settings = conn.execute(f'''
            SELECT email, quality_tier FROM customer_settings 
            WHERE email IN ({placeholders})
        ''', customer_emails).fetchall()
        
        # 2. Pre-calculate customer stats for all emails to avoid N+1 query locks
        email_stats = {}
        emails_list = list(customer_emails)
        placeholders_stats = ', '.join(['?' for _ in emails_list])
        
        try:
            stats_results = conn.execute(f'''
                SELECT
                    json_extract(billing, '$.email') as current_email,
                    COUNT(*) as total_orders,
                    SUM({_success_status_case()}) as successful_orders,
                    {_success_amount_case('total')} as total_spending,
                    SUM({_bad_order_case()}) as bad_orders,
                    SUM({_meaningful_order_case()}) as meaningful_orders,
                    MAX(date_created) as last_order_date,
                    MIN(date_created) as first_order_date
                FROM orders
                WHERE json_extract(billing, '$.email') IN ({placeholders_stats})
                GROUP BY current_email
            ''', tuple(emails_list)).fetchall()
            
            for r in stats_results:
                email_stats[r['current_email']] = dict(r)
        except Exception as e:
            # Fallback if there is any issue with the bulk query
            pass

        # 3. Calculate attributes for each customer
        for email in customer_emails:
            if email not in customer_attributes:
                customer_attributes[email] = {}
            
            # Check manual setting
            manual_tier = 'auto'
            for row in manual_settings:
                if row['email'] == email:
                    manual_tier = row['quality_tier']
                    break
            
            # Get customer stats from pre-calculated dictionary
            stats = email_stats.get(email, {})
            
            total_orders = stats.get('total_orders') or 0
            successful_orders = stats.get('successful_orders') or 0
            total_spending = stats.get('total_spending') or 0
            
            # Store order count for display
            customer_attributes[email]['order_count'] = total_orders
            customer_attributes[email]['is_new'] = (total_orders <= 1)
            
            # Determine Tier
            tier = manual_tier
            
            if tier == 'auto':
                # Calculate auto tier using shared logic
                avg_days_between = 0
                if total_orders > 1 and stats['first_order_date'] and stats['last_order_date']:
                    from datetime import datetime
                    try:
                        first_date = datetime.fromisoformat(stats['first_order_date'][:19])
                        last_date = datetime.fromisoformat(stats['last_order_date'][:19])
                        days_span = (last_date - first_date).days
                        avg_days_between = days_span / (total_orders - 1)
                    except:
                        avg_days_between = 0
                
                tier = calculate_customer_tier(
                    successful_orders, total_spending, avg_days_between,
                    bad_orders=stats.get('bad_orders') or 0,
                    meaningful_orders=stats.get('meaningful_orders') or 0)

            # Set attributes based on tier
            if tier == 'vip':
                customer_attributes[email]['quality'] = {'label': 'VIP', 'class': 'text-warning', 'icon': 'star-fill'}
            elif tier == 'good':
                customer_attributes[email]['quality'] = {'label': '优质', 'class': 'text-success', 'icon': 'gem'}
            elif tier == 'normal':
                customer_attributes[email]['quality'] = {'label': '普通', 'class': 'text-primary', 'icon': 'person-check'}
            elif tier == 'new':
                customer_attributes[email]['quality'] = {'label': '新客', 'class': 'text-info', 'icon': 'stars'}
            elif tier == 'bad':
                customer_attributes[email]['quality'] = {'label': '劣质', 'class': 'text-danger', 'icon': 'x-circle'}

    conn.close()
    
    return render_template('orders.html',
                         orders=processed_orders,
                         sources=sources,
                         statuses=statuses,
                         summary_stats=summary_stats,
                         totals=totals,
                         displayed_totals=displayed_totals,
                         site_managers=site_managers,
                         customer_attributes=customer_attributes,
                         all_managers=all_managers,
                         all_countries=all_countries, # 4. 在render_template中返回country_filter和all_countries
                         current_filters={
                             'source': source_filter,
                             'status': status_filter,
                             'date_from': date_from,
                             'date_to': date_to,
                             'search': search,
                             'quick_date': quick_date,
                             'manager': manager_filter,
                             'country': country_filter # 4. 在render_template中返回country_filter和all_countries
                         },
                         available_months=available_months,
                         current_month=current_month,
                         total=total)



@app.route('/monthly')
@login_required
def monthly():
    """Monthly statistics page"""
    conn = get_db_connection()
    
    # Get user's allowed sources for permission filtering
    allowed_sources = get_user_allowed_sources(current_user.id, current_user.is_admin(), current_user.is_viewer())
    
    # Get source filter
    source_filter = request.args.get('source', '')
    manager_filter = request.args.get('manager', '')
    country_filter = request.args.get('country', '')
    start_month = request.args.get('start_month', '')
    end_month = request.args.get('end_month', '')
    
    # Get all managers
    all_managers = get_all_managers()
    
    # Get all countries from sites table
    all_countries = conn.execute('SELECT DISTINCT country FROM sites WHERE country IS NOT NULL AND country != "" ORDER BY country').fetchall()
    all_countries = [c['country'] for c in all_countries]
    
    # Validate source_filter against allowed sources
    if source_filter and allowed_sources is not None and source_filter not in allowed_sources:
        source_filter = ''
    
    # Get available sources (filtered by permissions and manager)
    conn = get_db_connection()
    
    # Base source query
    source_query = 'SELECT DISTINCT source FROM orders'
    source_params = []
    source_conditions = []
    
    if allowed_sources is not None:
        if allowed_sources:
            placeholders = ', '.join(['?' for _ in allowed_sources])
            source_conditions.append(f'source IN ({placeholders})')
            source_params.extend(allowed_sources)
        else:
            source_conditions.append('1=0')
            
    if manager_filter:
        # Get sites managed by this manager
        manager_sites = conn.execute('SELECT url FROM sites WHERE manager = ?', (manager_filter,)).fetchall()
        manager_urls = [s['url'] for s in manager_sites]
        if manager_urls:
            placeholders = ', '.join(['?' for _ in manager_urls])
            source_conditions.append(f'source IN ({placeholders})')
            source_params.extend(manager_urls)
        else:
            source_conditions.append('1=0')
            
    if country_filter:
        # Get sites in this country
        country_sites = conn.execute('SELECT url FROM sites WHERE country = ?', (country_filter,)).fetchall()
        country_urls = [s['url'] for s in country_sites]
        if country_urls:
            placeholders = ', '.join(['?' for _ in country_urls])
            source_conditions.append(f'source IN ({placeholders})')
            source_params.extend(country_urls)
        else:
            source_conditions.append('1=0')
            
    if source_conditions:
        source_query += ' WHERE ' + ' AND '.join(source_conditions)
        
    source_query += ' ORDER BY source'
    all_sources = conn.execute(source_query, source_params).fetchall()
    
    # Build query with permission filter and optional source filter
    conditions = []
    params = []
    
    if allowed_sources is not None:
        if allowed_sources:
            placeholders = ', '.join(['?' for _ in allowed_sources])
            conditions.append(f'source IN ({placeholders})')
            params.extend(allowed_sources)
        else:
            conditions.append('1=0')
            
    # Add manager filter
    if manager_filter:
        # Re-use manager_urls from above or fetch if not done
        if 'manager_urls' not in locals():
            manager_sites = conn.execute('SELECT url FROM sites WHERE manager = ?', (manager_filter,)).fetchall()
            manager_urls = [s['url'] for s in manager_sites]
        
        if manager_urls:
            placeholders = ', '.join(['?' for _ in manager_urls])
            conditions.append(f'source IN ({placeholders})')
            params.extend(manager_urls)
        else:
            conditions.append('1=0')
            
    # Add country filter
    if country_filter:
        if 'country_urls' not in locals():
            country_sites = conn.execute('SELECT url FROM sites WHERE country = ?', (country_filter,)).fetchall()
            country_urls = [s['url'] for s in country_sites]
        
        if country_urls:
            placeholders = ', '.join(['?' for _ in country_urls])
            conditions.append(f'source IN ({placeholders})')
            params.extend(country_urls)
        else:
            conditions.append('1=0')
    
    if source_filter:
        conditions.append('source = ?')
        params.append(source_filter)

    if start_month:
        conditions.append("strftime('%Y-%m', date_created) >= ?")
        params.append(start_month)

    if end_month:
        conditions.append("strftime('%Y-%m', date_created) <= ?")
        params.append(end_month)

    # Exclude checkout drafts and trash — same as orders page
    conditions.append("status NOT IN ('checkout-draft', 'trash')")

    where_clause = 'WHERE ' + ' AND '.join(conditions) if conditions else ''

    query = f'''
        SELECT id, status, date_created, total, shipping_total, source, currency, payment_method, line_items,
               is_undelivered, shipping_loss_amount,
               is_problem_return, product_loss_amount
        FROM orders
        {where_clause}
        ORDER BY date_created DESC
    '''

    df = pd.read_sql_query(query, conn, params=params if params else None)
    # Pre-load on-hold-is-shipped sites so the pandas mask can exclude on-hold
    # orders from AU/AE (and any new countries where on-hold means "received,
    # waiting" rather than "shipped via COD").
    on_hold_shipped_sources = {
        r['url'] for r in conn.execute(
            "SELECT url FROM sites WHERE cod_on_hold_is_shipped = 1"
        ).fetchall()
    }
    conn.close()

    if len(df) == 0:
        return render_template('monthly.html', monthly_stats=[], sources=all_sources, source_filter=source_filter, country_filter=country_filter, all_countries=all_countries, start_month=start_month, end_month=end_month)

    # Calculate product quantities
    def get_product_qty(line_items):
        try:
            items = json.loads(line_items) if line_items else []
            return sum(item.get('quantity', 0) for item in items)
        except:
            return 0

    df['product_qty'] = df['line_items'].apply(get_product_qty)
    df['month'] = pd.to_datetime(df['date_created']).dt.to_period('M')

    # Group by month, source, and currency
    rows = []
    for (month, source), gdf in df.groupby(['month', 'source']):
        # Get the currency for this source (should be the same for all orders from same source)
        currency = gdf['currency'].iloc[0] if len(gdf) > 0 else 'N/A'

        total_orders = len(gdf)
        total_products = gdf['product_qty'].sum()
        total_amount = gdf['total'].sum()

        success_mask = (
            ~gdf['status'].isin(['failed', 'cancelled', 'checkout-draft', 'trash', 'cheat'])
            & ~((gdf['status'] == 'pending') & (gdf['payment_method'].fillna('cod') != 'cod'))
            & ~((gdf['status'] == 'on-hold') & (gdf['payment_method'] == 'bacs'))
            & ~((gdf['status'] == 'on-hold') & ~gdf['source'].isin(on_hold_shipped_sources))
            & (gdf['is_undelivered'].fillna(0) == 0)
            & (gdf['is_problem_return'].fillna(0) == 0)
        )
        success_orders = int(success_mask.sum())
        success_amount = gdf.loc[success_mask, 'total'].sum()
        success_products = gdf.loc[success_mask, 'product_qty'].sum()
        success_shipping = gdf.loc[success_mask, 'shipping_total'].sum()

        failed_mask = gdf['status'] == 'failed'
        failed_orders = int(failed_mask.sum())

        cancelled_mask = gdf['status'] == 'cancelled'
        cancelled_orders = int(cancelled_mask.sum())

        undelivered_mask = gdf['is_undelivered'].fillna(0) == 1
        undelivered_orders = int(undelivered_mask.sum())

        problem_return_mask = gdf['is_problem_return'].fillna(0) == 1
        problem_return_orders = int(problem_return_mask.sum())

        # shipping_loss covers both undelivered AND problem-return orders
        # (same column on either flag); product_loss is problem-return only.
        loss_mask = undelivered_mask | problem_return_mask
        shipping_loss = float(gdf.loc[loss_mask, 'shipping_loss_amount'].fillna(0).sum())
        product_loss = float(gdf.loc[problem_return_mask, 'product_loss_amount'].fillna(0).sum())

        # Net = product revenue (after deducting collected shipping fees)
        # MINUS shipping AND product losses. Without these terms, returned-
        # package and brick-swap losses were silently absorbed.
        success_net_amount = success_amount - success_shipping - shipping_loss - product_loss

        rows.append({
            'month': str(month),
            'source': source,
            'currency': currency,
            'total_orders': total_orders,
            'total_products': int(total_products),
            'total_amount': float(total_amount),
            'success_orders': success_orders,
            'success_products': int(success_products),
            'success_amount': float(success_amount),
            'success_net_amount': float(success_net_amount),
            'failed_orders': failed_orders,
            'cancelled_orders': cancelled_orders,
            'undelivered_orders': undelivered_orders,
            'shipping_loss': shipping_loss,
            'problem_return_orders': problem_return_orders,
            'product_loss': product_loss
        })
    
    # Get sort parameter
    sort_by = request.args.get('sort', 'month')  # default: sort by month
    
    # Sort based on parameter
    if sort_by == 'source':
        rows.sort(key=lambda x: (x['source'], x['month']), reverse=False)
    else:
        rows.sort(key=lambda x: (x['month'], x['source']), reverse=True)
    
    # Get site managers mapping
    conn2 = get_db_connection()
    sites = conn2.execute('SELECT url, manager FROM sites').fetchall()
    site_managers = {s['url']: s['manager'] or '' for s in sites}
    conn2.close()
    
    # Add CNY conversion to each row
    for row in rows:
        currency = row.get('currency', 'PLN')
        month = row.get('month', '')
        rate, _ = get_cny_rate(currency, month)
        row['rate_to_cny'] = rate
        row['success_net_amount_cny'] = round(row['success_net_amount'] * rate, 2) if rate else None
    
    # Calculate monthly aggregates
    monthly_aggregates = {}
    for row in rows:
        m = row['month']
        if m not in monthly_aggregates:
            monthly_aggregates[m] = {'success_net_amount_cny': 0, 'success_orders': 0, 'success_products': 0, 'failed_orders': 0, 'cancelled_orders': 0, 'undelivered_orders': 0, 'problem_return_orders': 0, 'total_orders': 0, 'total_products': 0, 'currency_amounts': {}}
        monthly_aggregates[m]['success_net_amount_cny'] += (row.get('success_net_amount_cny') or 0)
        monthly_aggregates[m]['success_orders'] += row['success_orders']
        monthly_aggregates[m]['success_products'] += row['success_products']
        monthly_aggregates[m]['failed_orders'] += row['failed_orders']
        monthly_aggregates[m]['cancelled_orders'] += row['cancelled_orders']
        monthly_aggregates[m]['undelivered_orders'] += row.get('undelivered_orders', 0)
        monthly_aggregates[m]['problem_return_orders'] += row.get('problem_return_orders', 0)
        monthly_aggregates[m]['total_orders'] += row['total_orders']
        monthly_aggregates[m]['total_products'] += row['total_products']
        # Aggregate amounts by currency (shipping_loss/product_loss tracked here too — same currency as orders)
        currency = row.get('currency', 'PLN')
        if currency not in monthly_aggregates[m]['currency_amounts']:
            monthly_aggregates[m]['currency_amounts'][currency] = {'total_amount': 0, 'success_amount': 0, 'success_net_amount': 0, 'shipping_loss': 0, 'product_loss': 0}
        monthly_aggregates[m]['currency_amounts'][currency]['total_amount'] += row.get('total_amount', 0)
        monthly_aggregates[m]['currency_amounts'][currency]['success_amount'] += row.get('success_amount', 0)
        monthly_aggregates[m]['currency_amounts'][currency]['success_net_amount'] += row.get('success_net_amount', 0)
        monthly_aggregates[m]['currency_amounts'][currency]['shipping_loss'] += row.get('shipping_loss', 0)
        monthly_aggregates[m]['currency_amounts'][currency]['product_loss'] += row.get('product_loss', 0)
        
    return render_template('monthly.html', monthly_stats=rows, monthly_aggregates=monthly_aggregates, sources=all_sources, source_filter=source_filter, manager_filter=manager_filter, country_filter=country_filter, all_managers=all_managers, all_countries=all_countries, sort_by=sort_by, site_managers=site_managers, start_month=start_month, end_month=end_month)


@app.route('/api/monthly/export')
@login_required
def monthly_export():
    """Export monthly statistics to Excel - Admin only"""
    # Check admin permission
    if not current_user.is_admin():
        return jsonify({'error': '只有管理员才能导出数据'}), 403
    
    from io import BytesIO
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.chart import BarChart, LineChart, Reference
    from openpyxl.chart.label import DataLabelList
    
    conn = get_db_connection()
    
    # Get user's allowed sources for permission filtering
    allowed_sources = get_user_allowed_sources(current_user.id, current_user.is_admin(), current_user.is_viewer())
    
    # Get filters
    source_filter = request.args.get('source', '')
    manager_filter = request.args.get('manager', '')
    manager_filter = request.args.get('manager', '')
    country_filter = request.args.get('country', '')
    start_month = request.args.get('start_month', '')
    end_month = request.args.get('end_month', '')
    
    # Build query conditions
    conditions = []
    params = []
    
    if allowed_sources is not None:
        if allowed_sources:
            placeholders = ', '.join(['?' for _ in allowed_sources])
            conditions.append(f'source IN ({placeholders})')
            params.extend(allowed_sources)
        else:
            conditions.append('1=0')
    
    if manager_filter:
        manager_sites = conn.execute('SELECT url FROM sites WHERE manager = ?', (manager_filter,)).fetchall()
        manager_urls = [s['url'] for s in manager_sites]
        if manager_urls:
            placeholders = ', '.join(['?' for _ in manager_urls])
            conditions.append(f'source IN ({placeholders})')
            params.extend(manager_urls)
        else:
            conditions.append('1=0')
    
    if country_filter:
        country_sites = conn.execute('SELECT url FROM sites WHERE country = ?', (country_filter,)).fetchall()
        country_urls = [s['url'] for s in country_sites]
        if country_urls:
            placeholders = ', '.join(['?' for _ in country_urls])
            conditions.append(f'source IN ({placeholders})')
            params.extend(country_urls)
        else:
            conditions.append('1=0')
    
    if source_filter:
        conditions.append('source = ?')
        params.append(source_filter)
        
    if start_month:
        conditions.append("strftime('%Y-%m', date_created) >= ?")
        params.append(start_month)

    if end_month:
        conditions.append("strftime('%Y-%m', date_created) <= ?")
        params.append(end_month)

    # Exclude checkout drafts and trash — same as orders page
    conditions.append("status NOT IN ('checkout-draft', 'trash')")

    where_clause = 'WHERE ' + ' AND '.join(conditions) if conditions else ''

    query = f'''
        SELECT id, status, date_created, total, shipping_total, source, currency, payment_method, line_items,
               is_undelivered, shipping_loss_amount,
               is_problem_return, product_loss_amount
        FROM orders
        {where_clause}
        ORDER BY date_created DESC
    '''

    df = pd.read_sql_query(query, conn, params=params if params else None)

    # Get site managers + on-hold-is-shipped lookup in the same pass
    sites = conn.execute('SELECT url, manager, cod_on_hold_is_shipped FROM sites').fetchall()
    site_managers = {s['url']: s['manager'] or '' for s in sites}
    on_hold_shipped_sources = {s['url'] for s in sites if s['cod_on_hold_is_shipped'] == 1}
    conn.close()

    if len(df) == 0:
        return jsonify({'error': '没有数据可导出'}), 400

    # Calculate product quantities
    def get_product_qty(line_items):
        try:
            items = json.loads(line_items) if line_items else []
            return sum(item.get('quantity', 0) for item in items)
        except:
            return 0

    df['product_qty'] = df['line_items'].apply(get_product_qty)
    df['month'] = pd.to_datetime(df['date_created']).dt.to_period('M')

    # Group by month, source
    rows = []
    for (month, source), gdf in df.groupby(['month', 'source']):
        currency = gdf['currency'].iloc[0] if len(gdf) > 0 else 'N/A'

        total_orders = len(gdf)
        total_products = gdf['product_qty'].sum()
        total_amount = gdf['total'].sum()

        success_mask = (
            ~gdf['status'].isin(['failed', 'cancelled', 'checkout-draft', 'trash', 'cheat'])
            & ~((gdf['status'] == 'pending') & (gdf['payment_method'].fillna('cod') != 'cod'))
            & ~((gdf['status'] == 'on-hold') & (gdf['payment_method'] == 'bacs'))
            & ~((gdf['status'] == 'on-hold') & ~gdf['source'].isin(on_hold_shipped_sources))
            & (gdf['is_undelivered'].fillna(0) == 0)
            & (gdf['is_problem_return'].fillna(0) == 0)
        )
        success_orders = int(success_mask.sum())
        success_amount = gdf.loc[success_mask, 'total'].sum()
        success_products = gdf.loc[success_mask, 'product_qty'].sum()
        success_shipping = gdf.loc[success_mask, 'shipping_total'].sum()

        failed_orders = int((gdf['status'] == 'failed').sum())
        cancelled_orders = int((gdf['status'] == 'cancelled').sum())

        undelivered_mask = gdf['is_undelivered'].fillna(0) == 1
        undelivered_orders = int(undelivered_mask.sum())

        problem_return_mask = gdf['is_problem_return'].fillna(0) == 1
        problem_return_orders = int(problem_return_mask.sum())

        loss_mask = undelivered_mask | problem_return_mask
        shipping_loss = float(gdf.loc[loss_mask, 'shipping_loss_amount'].fillna(0).sum())
        product_loss = float(gdf.loc[problem_return_mask, 'product_loss_amount'].fillna(0).sum())

        # Net = product revenue minus collected shipping fees minus
        # shipping AND product losses (kept consistent with the /monthly view).
        success_net_amount = success_amount - success_shipping - shipping_loss - product_loss

        # Get CNY rate
        rate, _ = get_cny_rate(currency, str(month))
        success_net_amount_cny = round(success_net_amount * rate, 2) if rate else 0
        shipping_loss_cny = round(shipping_loss * rate, 2) if rate else 0
        product_loss_cny = round(product_loss * rate, 2) if rate else 0

        rows.append({
            '月份': str(month),
            '网站': source.replace('https://www.', '').replace('https://', ''),
            '负责人': site_managers.get(source, ''),
            '货币': currency,
            '总订单数': total_orders,
            '总产品数': int(total_products),
            '总金额': round(float(total_amount), 2),
            '成功订单数': success_orders,
            '成功产品数': int(success_products),
            '成功金额': round(float(success_amount), 2),
            '成功净金额': round(float(success_net_amount), 2),
            '失败订单': failed_orders,
            '取消订单': cancelled_orders,
            '未送达订单': undelivered_orders,
            '运费损失': round(shipping_loss, 2),
            '运费损失(CNY)': shipping_loss_cny,
            '问题退货订单': problem_return_orders,
            '货值损失': round(product_loss, 2),
            '货值损失(CNY)': product_loss_cny,
            '汇率': round(rate, 4) if rate else 0,
            '净金额(CNY)': round(success_net_amount_cny, 2),
            '_total_amount': float(total_amount),
            '_success_amount': float(success_amount),
            '_success_net_amount': float(success_net_amount),
            '_net_amount_cny': success_net_amount_cny,
            '_shipping_loss': shipping_loss,
            '_shipping_loss_cny': shipping_loss_cny,
            '_product_loss': product_loss,
            '_product_loss_cny': product_loss_cny
        })
    
    # Sort by month descending
    rows.sort(key=lambda x: x['月份'], reverse=True)
    
    # Calculate monthly totals with all fields
    monthly_totals = {}
    for row in rows:
        m = row['月份']
        if m not in monthly_totals:
            monthly_totals[m] = {
                'site_count': 0,
                'total_orders': 0,
                'total_products': 0,
                'total_amount': 0,
                'success_orders': 0,
                'success_products': 0,
                'success_amount': 0,
                'success_net_amount': 0,
                'failed_orders': 0,
                'cancelled_orders': 0,
                'undelivered_orders': 0,
                'shipping_loss': 0,
                'shipping_loss_cny': 0,
                'problem_return_orders': 0,
                'product_loss': 0,
                'product_loss_cny': 0,
                'net_amount_cny': 0
            }
        monthly_totals[m]['site_count'] += 1
        monthly_totals[m]['total_orders'] += row['总订单数']
        monthly_totals[m]['total_products'] += row['总产品数']
        monthly_totals[m]['total_amount'] += row['_total_amount']
        monthly_totals[m]['success_orders'] += row['成功订单数']
        monthly_totals[m]['success_products'] += row['成功产品数']
        monthly_totals[m]['success_amount'] += row['_success_amount']
        monthly_totals[m]['success_net_amount'] += row['_success_net_amount']
        monthly_totals[m]['failed_orders'] += row['失败订单']
        monthly_totals[m]['cancelled_orders'] += row['取消订单']
        monthly_totals[m]['undelivered_orders'] += row.get('未送达订单', 0)
        monthly_totals[m]['shipping_loss'] += row.get('_shipping_loss', 0)
        monthly_totals[m]['shipping_loss_cny'] += row.get('_shipping_loss_cny', 0)
        monthly_totals[m]['problem_return_orders'] += row.get('问题退货订单', 0)
        monthly_totals[m]['product_loss'] += row.get('_product_loss', 0)
        monthly_totals[m]['product_loss_cny'] += row.get('_product_loss_cny', 0)
        monthly_totals[m]['net_amount_cny'] += row['_net_amount_cny']
    
    # Create Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.title = '月度统计'
    
    # Determine header currency based on country filter
    country_currency_map = {
        'PL': 'PLN',
        'AU': 'AUD',
        'AE': 'AED',
        'DE': 'EUR',
        'FR': 'EUR',
        'UK': 'GBP',
        'US': 'USD',
        'CN': 'CNY'
    }
    header_currency = country_currency_map.get(country_filter, '混合货币') if country_filter else '混合货币'
    
    # Write headers (exclude hidden fields starting with _)
    headers = [k for k in rows[0].keys() if not k.startswith('_')] if rows else []
    
    # Modify header names to include currency
    header_display_map = {
        '总金额': f'总金额({header_currency})',
        '成功金额': f'成功金额({header_currency})',
        '成功净金额': f'成功净金额({header_currency})'
    }
    
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    
    for col, header in enumerate(headers, 1):
        display_header = header_display_map.get(header, header)
        cell = ws.cell(row=1, column=col, value=display_header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Write data (exclude hidden fields) with center alignment
    center_align = Alignment(horizontal='center', vertical='center')
    for row_idx, row_data in enumerate(rows, 2):
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=row_data[header])
            cell.alignment = center_align
    
    # Add empty row
    summary_start_row = len(rows) + 3
    
    # Determine summary currency based on country filter
    country_currency_map = {
        'PL': 'PLN',
        'AU': 'AUD',
        'AE': 'AED',
        'DE': 'EUR',
        'FR': 'EUR',
        'UK': 'GBP',
        'US': 'USD',
        'CN': 'CNY'
    }
    summary_currency = country_currency_map.get(country_filter, '混合货币') if country_filter else '混合货币'
    
    # Add monthly summary with all columns
    summary_font = Font(bold=True, color='006400')
    for idx, (month, totals) in enumerate(sorted(monthly_totals.items(), reverse=True)):
        row_num = summary_start_row + idx
        # Column A: 月份收入总计
        cell = ws.cell(row=row_num, column=1, value=f'{month}月收入总计')
        cell.font = summary_font
        cell.alignment = center_align
        # Column B: 站点数量
        cell = ws.cell(row=row_num, column=2, value=f'{totals["site_count"]}个站点')
        cell.font = summary_font
        cell.alignment = center_align
        # Column C: 负责人 - 留空
        ws.cell(row=row_num, column=3, value='').alignment = center_align
        # Column D: 货币 - 显示汇总货币
        cell = ws.cell(row=row_num, column=4, value=summary_currency)
        cell.font = summary_font
        cell.alignment = center_align
        # Column E: 总订单数
        cell = ws.cell(row=row_num, column=5, value=totals['total_orders'])
        cell.font = summary_font
        cell.alignment = center_align
        # Column F: 总产品数
        cell = ws.cell(row=row_num, column=6, value=totals['total_products'])
        cell.font = summary_font
        cell.alignment = center_align
        # Column G: 总金额
        cell = ws.cell(row=row_num, column=7, value=round(totals['total_amount'], 2))
        cell.font = summary_font
        cell.alignment = center_align
        # Column H: 成功订单数
        cell = ws.cell(row=row_num, column=8, value=totals['success_orders'])
        cell.font = summary_font
        cell.alignment = center_align
        # Column I: 成功产品数
        cell = ws.cell(row=row_num, column=9, value=totals['success_products'])
        cell.font = summary_font
        cell.alignment = center_align
        # Column J: 成功金额
        cell = ws.cell(row=row_num, column=10, value=round(totals['success_amount'], 2))
        cell.font = summary_font
        cell.alignment = center_align
        # Column K: 成功净金额
        cell = ws.cell(row=row_num, column=11, value=round(totals['success_net_amount'], 2))
        cell.font = summary_font
        cell.alignment = center_align
        # Column L: 失败订单
        cell = ws.cell(row=row_num, column=12, value=totals['failed_orders'])
        cell.font = summary_font
        cell.alignment = center_align
        # Column M: 取消订单
        cell = ws.cell(row=row_num, column=13, value=totals['cancelled_orders'])
        cell.font = summary_font
        cell.alignment = center_align
        # Column N: 未送达订单
        cell = ws.cell(row=row_num, column=14, value=totals['undelivered_orders'])
        cell.font = summary_font
        cell.alignment = center_align
        # Column O: 运费损失
        cell = ws.cell(row=row_num, column=15, value=round(totals['shipping_loss'], 2))
        cell.font = summary_font
        cell.alignment = center_align
        # Column P: 运费损失(CNY)
        cell = ws.cell(row=row_num, column=16, value=round(totals['shipping_loss_cny'], 2))
        cell.font = summary_font
        cell.alignment = center_align
        # Column Q: 问题退货订单
        cell = ws.cell(row=row_num, column=17, value=totals['problem_return_orders'])
        cell.font = summary_font
        cell.alignment = center_align
        # Column R: 货值损失
        cell = ws.cell(row=row_num, column=18, value=round(totals['product_loss'], 2))
        cell.font = summary_font
        cell.alignment = center_align
        # Column S: 货值损失(CNY)
        cell = ws.cell(row=row_num, column=19, value=round(totals['product_loss_cny'], 2))
        cell.font = summary_font
        cell.alignment = center_align
        # Column T: 汇率 - 留空
        ws.cell(row=row_num, column=20, value='').alignment = center_align
        # Column U: 净金额(CNY)
        cell = ws.cell(row=row_num, column=21, value=f'¥{round(totals["net_amount_cny"], 2)}')
        cell.font = summary_font
        cell.alignment = center_align
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        col_letter = column[0].column_letter
        for cell in column:
            try:
                if cell.value:
                    cell_len = len(str(cell.value))
                    # Chinese characters count as 2
                    for ch in str(cell.value):
                        if ord(ch) > 127:
                            cell_len += 1
                    if cell_len > max_length:
                        max_length = cell_len
            except:
                pass
        ws.column_dimensions[col_letter].width = min(max_length + 2, 50)
    
    # Freeze first row
    ws.freeze_panes = 'A2'
    
    # Create charts based on monthly summary data
    if monthly_totals:
        # Sort months chronologically for charts
        sorted_months = sorted(monthly_totals.keys())
        
        # Create a new sheet for chart data
        chart_data_start_row = summary_start_row + len(monthly_totals) + 3
        
        # Write chart data headers
        ws.cell(row=chart_data_start_row, column=1, value='月份').font = header_font
        ws.cell(row=chart_data_start_row, column=1).fill = header_fill
        ws.cell(row=chart_data_start_row, column=1).alignment = center_align
        ws.cell(row=chart_data_start_row, column=2, value='净金额(CNY)').font = header_font
        ws.cell(row=chart_data_start_row, column=2).fill = header_fill
        ws.cell(row=chart_data_start_row, column=2).alignment = center_align
        ws.cell(row=chart_data_start_row, column=3, value='订单数').font = header_font
        ws.cell(row=chart_data_start_row, column=3).fill = header_fill
        ws.cell(row=chart_data_start_row, column=3).alignment = center_align
        
        # Write chart data (chronological order)
        for idx, month in enumerate(sorted_months):
            row_num = chart_data_start_row + 1 + idx
            ws.cell(row=row_num, column=1, value=month).alignment = center_align
            ws.cell(row=row_num, column=2, value=round(monthly_totals[month]['net_amount_cny'], 2)).alignment = center_align
            ws.cell(row=row_num, column=3, value=monthly_totals[month]['success_orders']).alignment = center_align
        
        chart_data_end_row = chart_data_start_row + len(sorted_months)
        
        # Create Bar Chart for Net Amount
        bar_chart = BarChart()
        bar_chart.type = "col"
        bar_chart.grouping = "clustered"
        bar_chart.title = "月度净金额趋势"
        bar_chart.y_axis.title = "净金额 (CNY)"
        bar_chart.x_axis.title = "月份"
        bar_chart.style = 12
        bar_chart.width = 18
        bar_chart.height = 10
        
        data = Reference(ws, min_col=2, min_row=chart_data_start_row, max_row=chart_data_end_row, max_col=2)
        cats = Reference(ws, min_col=1, min_row=chart_data_start_row + 1, max_row=chart_data_end_row)
        bar_chart.add_data(data, titles_from_data=True)
        bar_chart.set_categories(cats)
        bar_chart.shape = 4
        
        # Add data labels
        bar_chart.dataLabels = DataLabelList()
        bar_chart.dataLabels.showVal = True
        
        # Place bar chart
        chart_row = chart_data_end_row + 2
        ws.add_chart(bar_chart, f"A{chart_row}")
        
        # Create Line Chart for Orders
        line_chart = LineChart()
        line_chart.title = "月度订单数趋势"
        line_chart.y_axis.title = "订单数"
        line_chart.x_axis.title = "月份"
        line_chart.style = 13
        line_chart.width = 18
        line_chart.height = 10
        
        order_data = Reference(ws, min_col=3, min_row=chart_data_start_row, max_row=chart_data_end_row, max_col=3)
        line_chart.add_data(order_data, titles_from_data=True)
        line_chart.set_categories(cats)
        
        # Add data labels
        line_chart.dataLabels = DataLabelList()
        line_chart.dataLabels.showVal = True
        
        # Place line chart next to bar chart
        ws.add_chart(line_chart, f"K{chart_row}")
    
    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    # Generate filename
    timestamp = datetime.now().strftime('%Y%m%d%H%M')
    country_name = country_filter if country_filter else '全部'
    filename = f'月度统计汇总（{country_name}）{timestamp}.xlsx'
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )

@app.route('/cancelled-analysis')
@login_required
def cancelled_analysis():
    """Cancelled/Failed Order Analysis Page"""
    conn = get_db_connection()
    
    # Get user's allowed sources for permission filtering
    allowed_sources = get_user_allowed_sources(current_user.id, current_user.is_admin(), current_user.is_viewer())
    
    # Get filters
    source_filter = request.args.get('source', '')
    manager_filter = request.args.get('manager', '')
    country_filter = request.args.get('country', '')
    quick_date = request.args.get('quick_date', 'this_month')
    month_filter = request.args.get('month', '')
    
    # Get all countries for filter
    all_countries = conn.execute('SELECT DISTINCT country FROM sites WHERE country IS NOT NULL AND country != "" ORDER BY country').fetchall()
    all_countries = [c['country'] for c in all_countries]
    
    # Calculate date range based on quick_date or month
    from datetime import timedelta
    today = datetime.now().date()
    date_from = ''
    date_to = today.isoformat()
    
    # If month filter is specified, it overrides quick_date
    if month_filter:
        # Month filter takes priority - set quick_date to empty to show no button is active
        quick_date = ''
        date_from = month_filter + '-01'
        # Calculate last day of month
        import calendar
        year, month = int(month_filter[:4]), int(month_filter[5:7])
        last_day = calendar.monthrange(year, month)[1]
        date_to = f"{month_filter}-{last_day:02d}"
    elif quick_date == 'this_month':
        date_from = today.replace(day=1).isoformat()
    elif quick_date == 'last_month':
        last_month_end = today.replace(day=1) - timedelta(days=1)
        date_from = last_month_end.replace(day=1).isoformat()
        date_to = last_month_end.isoformat()
    elif quick_date == 'last_quarter':
        date_from = (today - timedelta(days=90)).isoformat()
    elif quick_date == 'half_year':
        date_from = (today - timedelta(days=180)).isoformat()
    elif quick_date == 'one_year':
        date_from = (today - timedelta(days=365)).isoformat()
    elif quick_date == 'all':
        date_from = ''
        date_to = ''
    
    # Get all managers
    all_managers_rows = get_all_managers()
    # If get_all_managers returns a list of strings, use it directly. Assuming it does based on other usages.
    all_managers = all_managers_rows 

    # Get site managers mapping for display
    site_managers_rows = conn.execute('SELECT url, manager FROM sites').fetchall()
    site_managers = {row['url']: row['manager'] for row in site_managers_rows if row['manager']}
    
    # Validate source_filter against allowed sources
    if source_filter and allowed_sources is not None and source_filter not in allowed_sources:
        source_filter = ''
    
    # Get available sources
    source_query = 'SELECT DISTINCT source FROM orders'
    source_params = []
    source_conditions = []
    
    if allowed_sources is not None:
        if allowed_sources:
            placeholders = ', '.join(['?' for _ in allowed_sources])
            source_conditions.append(f'source IN ({placeholders})')
            source_params.extend(allowed_sources)
        else:
            source_conditions.append('1=0')
            
    if manager_filter:
        manager_sites = conn.execute('SELECT url FROM sites WHERE manager = ?', (manager_filter,)).fetchall()
        manager_urls = [s['url'] for s in manager_sites]
        if manager_urls:
            placeholders = ', '.join(['?' for _ in manager_urls])
            source_conditions.append(f'source IN ({placeholders})')
            source_params.extend(manager_urls)
        else:
            source_conditions.append('1=0')
            
    if country_filter:
        country_sites = conn.execute('SELECT url FROM sites WHERE country = ?', (country_filter,)).fetchall()
        country_urls = [s['url'] for s in country_sites]
        if country_urls:
            placeholders = ', '.join(['?' for _ in country_urls])
            source_conditions.append(f'source IN ({placeholders})')
            source_params.extend(country_urls)
        else:
            source_conditions.append('1=0')

    if source_conditions:
        source_query += ' WHERE ' + ' AND '.join(source_conditions)
        
    source_query += ' ORDER BY source'
    all_sources = conn.execute(source_query, source_params).fetchall()
    
    # Build query conditions for problem orders: cancelled / failed / undelivered / 问题退货
    conditions = ["(status IN ('cancelled', 'failed') OR is_undelivered = 1 OR is_problem_return = 1)"]
    params = []
    
    if allowed_sources is not None:
        if allowed_sources:
            placeholders = ', '.join(['?' for _ in allowed_sources])
            conditions.append(f'source IN ({placeholders})')
            params.extend(allowed_sources)
        else:
            conditions.append('1=0')
    
    manager_urls = []
    if manager_filter:
        manager_sites = conn.execute('SELECT url FROM sites WHERE manager = ?', (manager_filter,)).fetchall()
        manager_urls = [s['url'] for s in manager_sites]
        
        if manager_urls:
            placeholders = ', '.join(['?' for _ in manager_urls])
            conditions.append(f'source IN ({placeholders})')
            params.extend(manager_urls)
        else:
            conditions.append('1=0')
            
    if country_filter:
        country_sites = conn.execute('SELECT url FROM sites WHERE country = ?', (country_filter,)).fetchall()
        country_urls = [s['url'] for s in country_sites]
        if country_urls:
            placeholders = ', '.join(['?' for _ in country_urls])
            conditions.append(f'source IN ({placeholders})')
            params.extend(country_urls)
        else:
            conditions.append('1=0')
    
    if source_filter:
        conditions.append('source = ?')
        params.append(source_filter)
    
    # Add date range filter
    if date_from:
        conditions.append('date_created >= ?')
        params.append(date_from)
    if date_to:
        conditions.append('date_created <= ?')
        params.append(date_to + 'T23:59:59')
    
    where_clause = 'WHERE ' + ' AND '.join(conditions)
    
    # Query problem orders (cancelled / failed / undelivered / 问题退货) with date information
    query = f'''
        SELECT id, number, status, total, shipping_total, currency, date_created, date_modified, source,
               line_items, billing, payment_method,
               is_undelivered, shipping_loss_amount,
               is_problem_return, problem_return_type, product_loss_amount
        FROM orders
        {where_clause}
        ORDER BY date_created DESC
    '''

    orders_data = conn.execute(query, params).fetchall()

    # Get available months (also covers undelivered / 问题退货)
    months_query = '''
        SELECT DISTINCT strftime('%Y-%m', date_created) as month
        FROM orders
        WHERE status IN ('cancelled', 'failed') OR is_undelivered = 1 OR is_problem_return = 1
        ORDER BY month DESC
    '''
    available_months = conn.execute(months_query).fetchall()
    available_months = [m['month'] for m in available_months if m['month']]

    # Process data for stats
    total_cancelled = 0
    total_amount = 0
    cancellation_rate = 0
    orders_list = []

    timing_stats = {
        'within_1_day': {'count': 0, 'amount': 0},
        '1_to_7_days': {'count': 0, 'amount': 0},
        'over_7_days': {'count': 0, 'amount': 0}
    }

    status_stats = {
        'cancelled': {'count': 0, 'amount': 0},
        'failed': {'count': 0, 'amount': 0},
        'undelivered': {'count': 0, 'amount': 0, 'shipping_loss': 0},
        'problem_return': {'count': 0, 'amount': 0, 'shipping_loss': 0, 'product_loss': 0}
    }

    source_stats = {}
    customer_type_stats = {
        'new': {'count': 0, 'amount': 0},
        'repeat': {'count': 0, 'amount': 0}
    }
    
    # Helper to check customer history
    customer_order_counts = {}
    if orders_data:
        # Get customer order counts for all customers in this batch
        # This is a simplified approach - ideally check against all historical orders
        emails = []
        for o in orders_data:
            if o['billing']:
                billing = parse_json_field(o['billing'])
                if billing and billing.get('email'):
                    emails.append(billing.get('email'))
        
        if emails:
            placeholders = ', '.join(['?' for _ in emails])
            # Check historical success orders
            history_query = f'''
                SELECT json_extract(billing, '$.email') as email, COUNT(*) as count
                FROM orders 
                WHERE status = 'completed' 
                AND json_extract(billing, '$.email') IN ({placeholders})
                GROUP BY email
            '''
            history_counts = conn.execute(history_query, emails).fetchall()
            customer_order_counts = {r['email']: r['count'] for r in history_counts}

    from datetime import datetime as dt
    
    for order in orders_data:
        od = dict(order)
        status = order['status']
        total = float(order['total'] or 0)
        source = order['source']
        
        # Parse dates
        date_created = order['date_created']
        date_modified = order['date_modified']
        
        days_to_cancel = 0
        if date_created and date_modified:
            try:
                created = dt.fromisoformat(date_created[:19])
                modified = dt.fromisoformat(date_modified[:19])
                days_to_cancel = (modified - created).days
            except:
                days_to_cancel = 0
        
        od['days_to_cancel'] = days_to_cancel
        
        # Timing classification
        if days_to_cancel <= 1:
            timing_stats['within_1_day']['count'] += 1
            timing_stats['within_1_day']['amount'] += total
            od['timing_category'] = '1天内'
        elif days_to_cancel <= 7:
            timing_stats['1_to_7_days']['count'] += 1
            timing_stats['1_to_7_days']['amount'] += total
            od['timing_category'] = '1-7天'
        else:
            timing_stats['over_7_days']['count'] += 1
            timing_stats['over_7_days']['amount'] += total
            od['timing_category'] = '7天+'
        
        # Status stats — undelivered (delivery refused/returned) and 问题退货
        # (swap/short/damaged contents) are their own buckets, even if the
        # underlying status is e.g. 'completed' before the order came back.
        is_undelivered = bool(order['is_undelivered']) if 'is_undelivered' in order.keys() else False
        is_problem_return = bool(order['is_problem_return']) if 'is_problem_return' in order.keys() else False
        loss_amount = float(order['shipping_loss_amount'] or 0) if 'shipping_loss_amount' in order.keys() else 0.0
        product_loss = float(order['product_loss_amount'] or 0) if 'product_loss_amount' in order.keys() else 0.0
        od['is_undelivered'] = is_undelivered
        od['is_problem_return'] = is_problem_return
        od['shipping_loss_amount'] = loss_amount
        od['product_loss_amount'] = product_loss

        if is_problem_return:
            status_stats['problem_return']['count'] += 1
            status_stats['problem_return']['amount'] += total
            status_stats['problem_return']['shipping_loss'] += loss_amount
            status_stats['problem_return']['product_loss'] += product_loss
        elif is_undelivered:
            status_stats['undelivered']['count'] += 1
            status_stats['undelivered']['amount'] += total
            status_stats['undelivered']['shipping_loss'] += loss_amount
        elif status in status_stats:
            status_stats[status]['count'] += 1
            status_stats[status]['amount'] += total

        # Source stats
        if source not in source_stats:
            source_stats[source] = {'count': 0, 'amount': 0, 'cancelled': 0, 'failed': 0,
                                    'undelivered': 0, 'shipping_loss': 0,
                                    'problem_return': 0, 'product_loss': 0}
        source_stats[source]['count'] += 1
        source_stats[source]['amount'] += total
        if is_problem_return:
            source_stats[source]['problem_return'] += 1
            source_stats[source]['shipping_loss'] += loss_amount
            source_stats[source]['product_loss'] += product_loss
        elif is_undelivered:
            source_stats[source]['undelivered'] += 1
            source_stats[source]['shipping_loss'] += loss_amount
        elif status == 'cancelled':
            source_stats[source]['cancelled'] += 1
        else:
            source_stats[source]['failed'] += 1
        
        # Customer type classification
        billing = parse_json_field(order['billing'])
        email = billing.get('email', '') if billing else ''
        od['customer_name'] = f"{billing.get('first_name', '')} {billing.get('last_name', '')}".strip() if billing else ''
        od['customer_email'] = email
        
        if email:
            order_count = customer_order_counts.get(email, 0)
            if order_count <= 1:
                customer_type_stats['new']['count'] += 1
                customer_type_stats['new']['amount'] += total
                od['customer_type'] = '新客'
            else:
                customer_type_stats['repeat']['count'] += 1
                customer_type_stats['repeat']['amount'] += total
                od['customer_type'] = '老客'
        else:
            od['customer_type'] = '未知'
        
        # Parse line items for product count
        items = parse_json_field(order['line_items'])
        od['product_count'] = sum(i.get('quantity', 0) for i in items) if isinstance(items, list) else 0
        
        # Calculate net amount and CNY conversion
        shipping = float(order['shipping_total'] or 0)
        od['shipping_total'] = shipping
        od['net_total'] = total - shipping
        
        # Get CNY rate
        order_month = order['date_created'][:7] if order['date_created'] else None
        currency = order['currency']
        rate, _ = get_cny_rate(currency, order_month)
        od['rate_to_cny'] = rate
        od['total_cny'] = round(total * rate, 2) if rate else None
        od['net_total_cny'] = round(od['net_total'] * rate, 2) if rate else None
        
        orders_list.append(od)
    
    # Get total orders for cancellation rate
    total_query_conditions = []
    total_params = []
    
    if allowed_sources is not None:
        if allowed_sources:
            placeholders = ', '.join(['?' for _ in allowed_sources])
            total_query_conditions.append(f'source IN ({placeholders})')
            total_params.extend(allowed_sources)
            
    if manager_filter and manager_urls:
        placeholders = ', '.join(['?' for _ in manager_urls])
        total_query_conditions.append(f'source IN ({placeholders})')
        total_params.extend(manager_urls)

    if country_filter and country_urls:
        placeholders = ', '.join(['?' for _ in country_urls])
        total_query_conditions.append(f'source IN ({placeholders})')
        total_params.extend(country_urls)
    
    if source_filter:
        total_query_conditions.append('source = ?')
        total_params.append(source_filter)
    
    if date_from:
        total_query_conditions.append('date_created >= ?')
        total_params.append(date_from)
    if date_to:
        total_query_conditions.append('date_created <= ?')
        total_params.append(date_to + 'T23:59:59')
    
    total_where = 'WHERE ' + ' AND '.join(total_query_conditions) if total_query_conditions else ''
    total_orders_count = conn.execute(f'SELECT COUNT(*) FROM orders {total_where}', total_params).fetchone()[0]
    
    # Calculate totals
    total_cancelled = len(orders_list)
    total_amount = sum(float(o['total'] or 0) for o in orders_list)
    cancellation_rate = (total_cancelled / total_orders_count * 100) if total_orders_count > 0 else 0
    
    # Group amounts by currency
    currency_amounts = {}
    for o in orders_list:
        currency = o.get('currency', 'USD')
        amount = float(o.get('total') or 0)
        if currency not in currency_amounts:
            currency_amounts[currency] = 0
        currency_amounts[currency] += amount
    
    # Calculate total CNY amount
    total_cny = 0
    currency_stats = []
    for currency, amount in sorted(currency_amounts.items(), key=lambda x: x[1], reverse=True):
        rate, _ = get_cny_rate(currency, None)  # Use current rate
        cny_amount = round(amount * rate, 2) if rate else 0
        total_cny += cny_amount
        currency_stats.append({
            'currency': currency,
            'amount': round(amount, 2),
            'cny_amount': cny_amount,
            'rate': rate
        })
    
    # Convert source_stats to list for template
    source_stats_list = []
    for source, stats in source_stats.items():
        source_stats_list.append({
            'source': source,
            'manager': site_managers.get(source, ''),
            'count': stats['count'],
            'amount': stats['amount'],
            'cancelled': stats['cancelled'],
            'failed': stats['failed'],
            'undelivered': stats.get('undelivered', 0),
            'shipping_loss': stats.get('shipping_loss', 0)
        })
    source_stats_list.sort(key=lambda x: x['count'], reverse=True)
    
    # Get source display mode for user
    source_display_mode = 'full'
    if current_user.is_authenticated:
        user_pref = conn.execute('''
            SELECT preference_value FROM user_preferences 
            WHERE user_id = ? AND preference_key = 'source_display_mode'
        ''', (current_user.id,)).fetchone()
        if user_pref:
            source_display_mode = user_pref['preference_value']
    
    conn.close()
    
    return render_template('cancelled.html',
        orders=orders_list,
        total_cancelled=total_cancelled,
        total_amount=total_amount,
        total_orders=total_orders_count,
        cancellation_rate=round(cancellation_rate, 2),
        timing_stats=timing_stats,
        status_stats=status_stats,
        source_stats=source_stats_list,
        customer_type_stats=customer_type_stats,
        source_filter=source_filter,
        manager_filter=manager_filter,
        country_filter=country_filter,
        sources=all_sources, # pass all available sources for filter
        all_managers=all_managers,
        site_managers=site_managers,
        all_countries=all_countries,
        quick_date=quick_date,
        available_months=available_months,
        current_month=month_filter,
        source_display_mode=source_display_mode,
        currency_stats=currency_stats,
        total_cny=total_cny
    )
    """Cancelled/Failed Order Analysis Page"""
    conn = get_db_connection()
    # Get user's allowed sources for permission filtering
    allowed_sources = get_user_allowed_sources(current_user.id, current_user.is_admin(), current_user.is_viewer())
    
    # Get filters
    source_filter = request.args.get('source', '')
    manager_filter = request.args.get('manager', '')
    country_filter = request.args.get('country', '')
    quick_date = request.args.get('quick_date', 'this_month')
    month_filter = request.args.get('month', '')
    
    # Get all countries for filter
    all_countries = conn.execute('SELECT DISTINCT country FROM sites WHERE country IS NOT NULL AND country != "" ORDER BY country').fetchall()
    all_countries = [c['country'] for c in all_countries]
    
    # Calculate date range based on quick_date or month
    from datetime import timedelta
    today = datetime.now().date()
    date_from = ''
    date_to = today.isoformat()
    
    # If month filter is specified, it overrides quick_date
    if month_filter:
        # Month filter takes priority - set quick_date to empty to show no button is active
        quick_date = ''
        date_from = month_filter + '-01'
        # Calculate last day of month
        import calendar
        year, month = int(month_filter[:4]), int(month_filter[5:7])
        last_day = calendar.monthrange(year, month)[1]
        date_to = f"{month_filter}-{last_day:02d}"
    elif quick_date == 'this_month':
        date_from = today.replace(day=1).isoformat()
    elif quick_date == 'last_month':
        last_month_end = today.replace(day=1) - timedelta(days=1)
        date_from = last_month_end.replace(day=1).isoformat()
        date_to = last_month_end.isoformat()
    elif quick_date == 'last_quarter':
        date_from = (today - timedelta(days=90)).isoformat()
    elif quick_date == 'half_year':
        date_from = (today - timedelta(days=180)).isoformat()
    elif quick_date == 'one_year':
        date_from = (today - timedelta(days=365)).isoformat()
    elif quick_date == 'all':
        date_from = ''
        date_to = ''
    
    # Get all managers
    all_managers = get_all_managers()
    
    # Validate source_filter against allowed sources
    if source_filter and allowed_sources is not None and source_filter not in allowed_sources:
        source_filter = ''
    
    # Get available sources
    source_query = 'SELECT DISTINCT source FROM orders'
    source_params = []
    source_conditions = []
    
    if allowed_sources is not None:
        if allowed_sources:
            placeholders = ', '.join(['?' for _ in allowed_sources])
            source_conditions.append(f'source IN ({placeholders})')
            source_params.extend(allowed_sources)
        else:
            source_conditions.append('1=0')
            
    if manager_filter:
        manager_sites = conn.execute('SELECT url FROM sites WHERE manager = ?', (manager_filter,)).fetchall()
        manager_urls = [s['url'] for s in manager_sites]
        if manager_urls:
            placeholders = ', '.join(['?' for _ in manager_urls])
            source_conditions.append(f'source IN ({placeholders})')
            source_params.extend(manager_urls)
        else:
            source_conditions.append('1=0')
            
    if source_conditions:
        source_query += ' WHERE ' + ' AND '.join(source_conditions)
        
    source_query += ' ORDER BY source'
    all_sources = conn.execute(source_query, source_params).fetchall()
    
    # Build query conditions for problem orders: cancelled / failed / undelivered / 问题退货
    conditions = ["(status IN ('cancelled', 'failed') OR is_undelivered = 1 OR is_problem_return = 1)"]
    params = []
    
    if allowed_sources is not None:
        if allowed_sources:
            placeholders = ', '.join(['?' for _ in allowed_sources])
            conditions.append(f'source IN ({placeholders})')
            params.extend(allowed_sources)
        else:
            conditions.append('1=0')
    
    manager_urls = []
    if manager_filter:
        manager_sites = conn.execute('SELECT url FROM sites WHERE manager = ?', (manager_filter,)).fetchall()
        manager_urls = [s['url'] for s in manager_sites]
        
        if manager_urls:
            placeholders = ', '.join(['?' for _ in manager_urls])
            conditions.append(f'source IN ({placeholders})')
            params.extend(manager_urls)
        else:
            conditions.append('1=0')
            
    if country_filter:
        country_sites = conn.execute('SELECT url FROM sites WHERE country = ?', (country_filter,)).fetchall()
        country_urls = [s['url'] for s in country_sites]
        if country_urls:
            placeholders = ', '.join(['?' for _ in country_urls])
            conditions.append(f'source IN ({placeholders})')
            params.extend(country_urls)
        else:
            conditions.append('1=0')
    
    if source_filter:
        conditions.append('source = ?')
        params.append(source_filter)
    
    # Add date range filter
    if date_from:
        conditions.append('date_created >= ?')
        params.append(date_from)
    if date_to:
        conditions.append('date_created <= ?')
        params.append(date_to + 'T23:59:59')
    
    where_clause = 'WHERE ' + ' AND '.join(conditions)
    
    # Query cancelled/failed orders with date information
    query = f'''
        SELECT id, number, status, total, shipping_total, currency, date_created, date_modified, source, line_items, billing
        FROM orders
        {where_clause}
        ORDER BY date_created DESC
    '''
    
    orders_data = conn.execute(query, params).fetchall()
    
    # Get available months
    months_query = '''
        SELECT DISTINCT strftime('%Y-%m', date_created) as month 
        FROM orders 
        WHERE status IN ('cancelled', 'failed')
        ORDER BY month DESC
    '''
    available_months = [row['month'] for row in conn.execute(months_query).fetchall()]
    
    # Get site managers mapping
    sites = conn.execute('SELECT url, manager FROM sites').fetchall()
    site_managers = {s['url']: s['manager'] or '' for s in sites}
    
    # Calculate statistics
    from datetime import datetime as dt
    
    # By timing
    timing_stats = {
        'within_1_day': {'count': 0, 'amount': 0},
        '1_to_7_days': {'count': 0, 'amount': 0},
        'over_7_days': {'count': 0, 'amount': 0}
    }
    
    # By status
    status_stats = {'cancelled': {'count': 0, 'amount': 0}, 'failed': {'count': 0, 'amount': 0}}
    
    # By source
    source_stats = {}
    
    # By customer type
    customer_type_stats = {'new': {'count': 0, 'amount': 0}, 'repeat': {'count': 0, 'amount': 0}}
    
    # Get customer order counts for customer type classification
    customer_emails = set()
    orders_list = []
    
    for order in orders_data:
        billing = parse_json_field(order['billing'])
        email = billing.get('email', '') if billing else ''
        if email:
            customer_emails.add(email)
    
    # Get order counts for all customers
    customer_order_counts = {}
    if customer_emails:
        emails_list = list(customer_emails)
        placeholders_stats = ', '.join(['?' for _ in emails_list])
        try:
            counts = conn.execute(f'''
                SELECT json_extract(billing, '$.email') as email, COUNT(*) as cnt 
                FROM orders 
                WHERE json_extract(billing, '$.email') IN ({placeholders_stats})
                GROUP BY email
            ''', tuple(emails_list)).fetchall()
            for r in counts:
                customer_order_counts[r['email']] = r['cnt']
        except Exception as e:
            pass
        
        # Ensure all existing emails have some default value if not returned
        for email in emails_list:
            if email not in customer_order_counts:
                customer_order_counts[email] = 0
    
    for order in orders_data:
        od = dict(order)
        status = order['status']
        total = float(order['total'] or 0)
        source = order['source']
        
        # Parse dates
        date_created = order['date_created']
        date_modified = order['date_modified']
        
        days_to_cancel = 0
        if date_created and date_modified:
            try:
                created = dt.fromisoformat(date_created[:19])
                modified = dt.fromisoformat(date_modified[:19])
                days_to_cancel = (modified - created).days
            except:
                days_to_cancel = 0
        
        od['days_to_cancel'] = days_to_cancel
        
        # Timing classification
        if days_to_cancel <= 1:
            timing_stats['within_1_day']['count'] += 1
            timing_stats['within_1_day']['amount'] += total
            od['timing_category'] = '1天内'
        elif days_to_cancel <= 7:
            timing_stats['1_to_7_days']['count'] += 1
            timing_stats['1_to_7_days']['amount'] += total
            od['timing_category'] = '1-7天'
        else:
            timing_stats['over_7_days']['count'] += 1
            timing_stats['over_7_days']['amount'] += total
            od['timing_category'] = '7天+'
        
        # Status stats — undelivered (delivery refused/returned) and 问题退货
        # (swap/short/damaged contents) are their own buckets, even if the
        # underlying status is e.g. 'completed' before the order came back.
        is_undelivered = bool(order['is_undelivered']) if 'is_undelivered' in order.keys() else False
        is_problem_return = bool(order['is_problem_return']) if 'is_problem_return' in order.keys() else False
        loss_amount = float(order['shipping_loss_amount'] or 0) if 'shipping_loss_amount' in order.keys() else 0.0
        product_loss = float(order['product_loss_amount'] or 0) if 'product_loss_amount' in order.keys() else 0.0
        od['is_undelivered'] = is_undelivered
        od['is_problem_return'] = is_problem_return
        od['shipping_loss_amount'] = loss_amount
        od['product_loss_amount'] = product_loss

        if is_problem_return:
            status_stats['problem_return']['count'] += 1
            status_stats['problem_return']['amount'] += total
            status_stats['problem_return']['shipping_loss'] += loss_amount
            status_stats['problem_return']['product_loss'] += product_loss
        elif is_undelivered:
            status_stats['undelivered']['count'] += 1
            status_stats['undelivered']['amount'] += total
            status_stats['undelivered']['shipping_loss'] += loss_amount
        elif status in status_stats:
            status_stats[status]['count'] += 1
            status_stats[status]['amount'] += total

        # Source stats
        if source not in source_stats:
            source_stats[source] = {'count': 0, 'amount': 0, 'cancelled': 0, 'failed': 0,
                                    'undelivered': 0, 'shipping_loss': 0,
                                    'problem_return': 0, 'product_loss': 0}
        source_stats[source]['count'] += 1
        source_stats[source]['amount'] += total
        if is_problem_return:
            source_stats[source]['problem_return'] += 1
            source_stats[source]['shipping_loss'] += loss_amount
            source_stats[source]['product_loss'] += product_loss
        elif is_undelivered:
            source_stats[source]['undelivered'] += 1
            source_stats[source]['shipping_loss'] += loss_amount
        elif status == 'cancelled':
            source_stats[source]['cancelled'] += 1
        else:
            source_stats[source]['failed'] += 1
        
        # Customer type classification
        billing = parse_json_field(order['billing'])
        email = billing.get('email', '') if billing else ''
        od['customer_name'] = f"{billing.get('first_name', '')} {billing.get('last_name', '')}".strip() if billing else ''
        od['customer_email'] = email
        
        if email:
            order_count = customer_order_counts.get(email, 0)
            if order_count <= 1:
                customer_type_stats['new']['count'] += 1
                customer_type_stats['new']['amount'] += total
                od['customer_type'] = '新客'
            else:
                customer_type_stats['repeat']['count'] += 1
                customer_type_stats['repeat']['amount'] += total
                od['customer_type'] = '老客'
        else:
            od['customer_type'] = '未知'
        
        # Parse line items for product count
        items = parse_json_field(order['line_items'])
        od['product_count'] = sum(i.get('quantity', 0) for i in items) if isinstance(items, list) else 0
        
        # Calculate net amount and CNY conversion
        shipping = float(order['shipping_total'] or 0)
        od['shipping_total'] = shipping
        od['net_total'] = total - shipping
        
        # Get CNY rate
        order_month = order['date_created'][:7] if order['date_created'] else None
        currency = order['currency']
        rate, _ = get_cny_rate(currency, order_month)
        od['rate_to_cny'] = rate
        od['total_cny'] = round(total * rate, 2) if rate else None
        od['net_total_cny'] = round(od['net_total'] * rate, 2) if rate else None
        
        orders_list.append(od)
    
    # Get total orders for cancellation rate
    total_query_conditions = []
    total_params = []
    
    if allowed_sources is not None:
        if allowed_sources:
            placeholders = ', '.join(['?' for _ in allowed_sources])
            total_query_conditions.append(f'source IN ({placeholders})')
            total_params.extend(allowed_sources)
            
    if manager_filter and manager_urls:
        placeholders = ', '.join(['?' for _ in manager_urls])
        total_query_conditions.append(f'source IN ({placeholders})')
        total_params.extend(manager_urls)
    
    if source_filter:
        total_query_conditions.append('source = ?')
        total_params.append(source_filter)
    
    if date_from:
        total_query_conditions.append('date_created >= ?')
        total_params.append(date_from)
    if date_to:
        total_query_conditions.append('date_created <= ?')
        total_params.append(date_to + 'T23:59:59')
    
    total_where = 'WHERE ' + ' AND '.join(total_query_conditions) if total_query_conditions else ''
    total_orders_count = conn.execute(f'SELECT COUNT(*) FROM orders {total_where}', total_params).fetchone()[0]
    
    # Calculate totals
    total_cancelled = len(orders_list)
    total_amount = sum(float(o['total'] or 0) for o in orders_list)
    cancellation_rate = (total_cancelled / total_orders_count * 100) if total_orders_count > 0 else 0
    
    # Group amounts by currency
    currency_amounts = {}
    for o in orders_list:
        currency = o.get('currency', 'USD')
        amount = float(o.get('total') or 0)
        if currency not in currency_amounts:
            currency_amounts[currency] = 0
        currency_amounts[currency] += amount
    
    # Calculate total CNY amount
    total_cny = 0
    currency_stats = []
    for currency, amount in sorted(currency_amounts.items(), key=lambda x: x[1], reverse=True):
        rate, _ = get_cny_rate(currency, None)  # Use current rate
        cny_amount = round(amount * rate, 2) if rate else 0
        total_cny += cny_amount
        currency_stats.append({
            'currency': currency,
            'amount': round(amount, 2),
            'cny_amount': cny_amount,
            'rate': rate
        })
    
    # Convert source_stats to list for template
    source_stats_list = []
    for source, stats in source_stats.items():
        source_stats_list.append({
            'source': source,
            'manager': site_managers.get(source, ''),
            'count': stats['count'],
            'amount': stats['amount'],
            'cancelled': stats['cancelled'],
            'failed': stats['failed'],
            'undelivered': stats.get('undelivered', 0),
            'shipping_loss': stats.get('shipping_loss', 0)
        })
    source_stats_list.sort(key=lambda x: x['count'], reverse=True)
    
    # Get source display mode for user
    source_display_mode = 'full'
    if current_user.is_authenticated:
        user_pref = conn.execute('''
            SELECT preference_value FROM user_preferences 
            WHERE user_id = ? AND preference_key = 'source_display_mode'
        ''', (current_user.id,)).fetchone()
        if user_pref:
            source_display_mode = user_pref['preference_value']
    
    conn.close()
    
    return render_template('cancelled.html',
        orders=orders_list,
        total_cancelled=total_cancelled,
        total_amount=total_amount,
        total_orders=total_orders_count,
        cancellation_rate=round(cancellation_rate, 2),
        timing_stats=timing_stats,
        status_stats=status_stats,
        source_stats=source_stats_list,
        customer_type_stats=customer_type_stats,
        source_filter=source_filter,
        manager_filter=manager_filter,
        country_filter=country_filter,
        sources=all_sources, # pass all available sources for filter
        all_managers=all_managers,
        site_managers=site_managers,
        all_countries=all_countries,
        quick_date=quick_date,
        available_months=available_months,
        current_month=month_filter,
        source_display_mode=source_display_mode,
        currency_stats=currency_stats,
        total_cny=total_cny
    )


@app.route('/customers')
@login_required
def customers():
    from datetime import date, timedelta
    conn = get_db_connection()

    # Get user's allowed sources for permission filtering
    allowed_sources = get_user_allowed_sources(current_user.id, current_user.is_admin(), current_user.is_viewer())

    # Get source filter
    source_filter = request.args.get('source', '')
    manager_filter = request.args.get('manager', '')
    country_filter = request.args.get('country', '')

    # Period filter (period口径: every number reflects only orders in range).
    # Default is 'all' so the page keeps showing the lifetime customer base
    # until the user explicitly scopes to a month.
    quick_date = request.args.get('quick_date', 'all')
    date_from = request.args.get('date_from', '')
    date_to = request.args.get('date_to', '')
    # Explicit dates win; otherwise derive the range from the quick preset.
    if not date_from and not date_to:
        today = date.today()
        if quick_date == 'this_month':
            date_from = today.replace(day=1).isoformat()
            date_to = today.isoformat()
        elif quick_date == 'last_month':
            first_of_this = today.replace(day=1)
            last_month_end = first_of_this - timedelta(days=1)
            date_from = last_month_end.replace(day=1).isoformat()
            date_to = last_month_end.isoformat()
        elif quick_date == 'last_quarter':
            date_from = (today - timedelta(days=90)).isoformat()
            date_to = today.isoformat()
        elif quick_date == 'half_year':
            date_from = (today - timedelta(days=180)).isoformat()
            date_to = today.isoformat()
        elif quick_date == 'one_year':
            date_from = (today - timedelta(days=365)).isoformat()
            date_to = today.isoformat()
        # 'all' (or anything else) → no date bounds

    # Build the shared date-range SQL fragment (applied to BOTH the customer
    # aggregation query and the identity-resolution scan so they stay in sync).
    date_conditions = []
    date_params = []
    if date_from:
        date_conditions.append('date_created >= ?')
        date_params.append(date_from)
    if date_to:
        date_conditions.append('date_created <= ?')
        date_params.append(date_to + 'T23:59:59')

    # All countries for the filter dropdown
    all_countries = [c['country'] for c in conn.execute(
        "SELECT DISTINCT country FROM sites WHERE country IS NOT NULL AND country != '' ORDER BY country"
    ).fetchall()]

    # Get all managers
    all_managers = get_all_managers()
    
    # Validate source_filter against allowed sources
    if source_filter and allowed_sources is not None and source_filter not in allowed_sources:
        source_filter = ''
    
    # Get available sources (filtered by permissions and manager)
    conn = get_db_connection()
    
    # Base source query
    source_query = 'SELECT DISTINCT source FROM orders'
    source_params = []
    source_conditions = []
    
    if allowed_sources is not None:
        if allowed_sources:
            placeholders = ', '.join(['?' for _ in allowed_sources])
            source_conditions.append(f'source IN ({placeholders})')
            source_params.extend(allowed_sources)
        else:
            source_conditions.append('1=0')
            
    if manager_filter:
        # Get sites managed by this manager
        manager_sites = conn.execute('SELECT url FROM sites WHERE manager = ?', (manager_filter,)).fetchall()
        manager_urls = [s['url'] for s in manager_sites]
        if manager_urls:
            placeholders = ', '.join(['?' for _ in manager_urls])
            source_conditions.append(f'source IN ({placeholders})')
            source_params.extend(manager_urls)
        else:
            source_conditions.append('1=0')
            
    if source_conditions:
        source_query += ' WHERE ' + ' AND '.join(source_conditions)
        
    source_query += ' ORDER BY source'
    all_sources = conn.execute(source_query, source_params).fetchall()
    
    # Build query with permission filter and optional source filter
    base_condition = "json_extract(billing, '$.email') IS NOT NULL AND json_extract(billing, '$.email') != ''"
    conditions = [base_condition]
    params = []
    
    if allowed_sources is not None:
        if allowed_sources:
            placeholders = ', '.join(['?' for _ in allowed_sources])
            conditions.append(f'source IN ({placeholders})')
            params.extend(allowed_sources)
        else:
            conditions.append('1=0')
            
    # Add manager filter
    if manager_filter:
        # Re-use manager_urls from above or fetch if not done
        if 'manager_urls' not in locals():
            manager_sites = conn.execute('SELECT url FROM sites WHERE manager = ?', (manager_filter,)).fetchall()
            manager_urls = [s['url'] for s in manager_sites]
        
        if manager_urls:
            placeholders = ', '.join(['?' for _ in manager_urls])
            conditions.append(f'source IN ({placeholders})')
            params.extend(manager_urls)
        else:
            conditions.append('1=0')
    
    # Add country filter (resolve to that country's site URLs)
    if country_filter:
        country_sites = conn.execute('SELECT url FROM sites WHERE country = ?', (country_filter,)).fetchall()
        country_urls = [s['url'] for s in country_sites]
        if country_urls:
            placeholders = ', '.join(['?' for _ in country_urls])
            conditions.append(f'source IN ({placeholders})')
            params.extend(country_urls)
        else:
            conditions.append('1=0')

    if source_filter:
        conditions.append('source = ?')
        params.append(source_filter)

    # Period口径: scope every aggregate to the selected date range.
    conditions.extend(date_conditions)
    params.extend(date_params)

    where_clause = 'WHERE ' + ' AND '.join(conditions)

    query = f'''
        SELECT
            json_extract(billing, '$.email') as email,
            json_extract(billing, '$.first_name') || ' ' || json_extract(billing, '$.last_name') as name,
            json_extract(billing, '$.phone') as phone,
            COUNT(*) as total_orders,
            SUM({_success_status_case()}) as successful_orders,
            {_success_amount_case('total')} as total_spent,
            SUM(CASE WHEN COALESCE(is_undelivered, 0) = 1 THEN 1 ELSE 0 END) as undelivered_orders,
            SUM(CASE WHEN COALESCE(is_undelivered,0) = 1 OR COALESCE(is_problem_return,0) = 1
                     THEN COALESCE(shipping_loss_amount, 0) ELSE 0 END) as shipping_loss_total,
            SUM(CASE WHEN COALESCE(is_problem_return, 0) = 1 THEN 1 ELSE 0 END) as problem_return_orders,
            SUM(CASE WHEN COALESCE(is_problem_return, 0) = 1 THEN COALESCE(product_loss_amount, 0) ELSE 0 END) as product_loss_total,
            MAX(date_created) as last_order_date,
            MIN(date_created) as first_order_date,
            GROUP_CONCAT(DISTINCT source) as sources,
            GROUP_CONCAT(DISTINCT currency) as currencies
        FROM orders
        {where_clause}
        GROUP BY email
        ORDER BY total_spent DESC
    '''
    
    customers_data = conn.execute(query, params).fetchall()

    # Site→manager map used to label each customer's sites in the 多站下单 column.
    site_managers_global = {s['url']: (s['manager'] or '')
                            for s in conn.execute('SELECT url, manager FROM sites').fetchall()}

    # Identity resolution: union emails that share a phone/address into one
    # person so someone spreading orders across multiple emails collapses into a
    # single row. Uses the SAME filter as the customer query above so the scope
    # matches (permissions / manager / source filters).
    email_to_cluster, cluster_meta = _resolve_identity_clusters(conn, where_clause, params)

    conn.close()

    total_revenue = 0
    repeat_customers = 0
    new_customers_month = 0
    new_customers_last_month = 0
    # Repeat-rate numerators/denominators for the three switchable口径:
    #   headcount = repeat_customers / total_customers
    #   by_orders = repeat_success_orders / total_success_orders
    #   by_revenue = repeat_spending / total_revenue
    total_success_orders = 0
    repeat_success_orders = 0
    repeat_spending = 0.0

    import datetime
    now = datetime.datetime.now()
    thirty_days_ago = (now - datetime.timedelta(days=30)).strftime('%Y-%m-%d')
    sixty_days_ago = (now - datetime.timedelta(days=60)).strftime('%Y-%m-%d')
    ninety_days_ago = (now - datetime.timedelta(days=90)).strftime('%Y-%m-%d')

    tier_counts = {'VIP': 0, '优质': 0, '普通': 0, '新客': 0}

    # Aggregate loss bookkeeping so the page can headline "who's costing us money".
    loss_stats = {
        'customers_with_loss': 0,        # distinct identities with any shipping/product loss
        'undelivered_customers': 0,      # identities with ≥1 undelivered order
        'problem_return_customers': 0,   # identities with ≥1 problem-return order
        'shipping_loss_total': 0.0,      # summed across identities (native currency, see note)
        'product_loss_total': 0.0,
        # Per-currency breakdown so the overview strip can label amounts with
        # their real currency instead of an ambiguous bare number. Single-country
        # users get one bucket; admins viewing all countries get several.
        'by_currency': {},               # {currency: {'shipping': x, 'product': y}}
    }

    # ── Phase A: fold the per-email SQL rows into identity clusters ──
    # Each SQL row is one raw email (GROUP BY email). We sum its numbers into the
    # cluster that _resolve_identity_clusters assigned its normalized email to.
    # Guest orders with no email never merge — each gets a unique singleton key.
    identities = {}
    _singleton_seq = 0
    for row in customers_data:
        c = dict(row)
        raw_email = c.get('email')
        norm = _normalize_email(raw_email)
        ck = email_to_cluster.get(norm) if norm else None
        if not ck:
            _singleton_seq += 1
            ck = f"__noemail_{_singleton_seq}"

        idn = identities.get(ck)
        if idn is None:
            idn = identities[ck] = {
                'cluster_key': ck,
                'emails_by_spend': [], 'names_by_spend': [], 'phones': set(),
                'total_orders': 0, 'successful_orders': 0, 'total_spent': 0.0,
                'undelivered_orders': 0, 'shipping_loss_total': 0.0,
                'problem_return_orders': 0, 'product_loss_total': 0.0,
                'sources': set(), 'currencies': set(),
                'first_order_date': None, 'last_order_date': None,
            }
        spend = float(c.get('total_spent') or 0)
        if raw_email:
            idn['emails_by_spend'].append((spend, raw_email))
        nm = (c.get('name') or '').strip()
        if nm:
            idn['names_by_spend'].append((spend, nm))
        if c.get('phone'):
            idn['phones'].add(str(c['phone']).strip())
        idn['total_orders'] += int(c.get('total_orders') or 0)
        idn['successful_orders'] += int(c.get('successful_orders') or 0)
        idn['total_spent'] += spend
        idn['undelivered_orders'] += int(c.get('undelivered_orders') or 0)
        idn['shipping_loss_total'] += float(c.get('shipping_loss_total') or 0)
        idn['problem_return_orders'] += int(c.get('problem_return_orders') or 0)
        idn['product_loss_total'] += float(c.get('product_loss_total') or 0)
        for s in (c.get('sources') or '').split(','):
            s = s.strip()
            if s:
                idn['sources'].add(s)
        for cur in (c.get('currencies') or '').split(','):
            cur = cur.strip()
            if cur:
                idn['currencies'].add(cur)
        fod, lod = c.get('first_order_date'), c.get('last_order_date')
        if fod:
            idn['first_order_date'] = fod if idn['first_order_date'] is None else min(idn['first_order_date'], fod)
        if lod:
            idn['last_order_date'] = lod if idn['last_order_date'] is None else max(idn['last_order_date'], lod)

    # ── Phase C: derive per-identity display fields + accumulate page stats ──
    customers_list = []
    for ck, idn in identities.items():
        c = {}
        idn['emails_by_spend'].sort(reverse=True)
        idn['names_by_spend'].sort(reverse=True)
        # Representative name + primary email = the highest-spending within the
        # merged identity (so the "main" account leads the display).
        primary_email = idn['emails_by_spend'][0][1] if idn['emails_by_spend'] else ''
        c['email'] = primary_email
        c['name'] = idn['names_by_spend'][0][1] if idn['names_by_spend'] else (primary_email or 'Unknown')
        c['phone'] = next(iter(idn['phones'])) if idn['phones'] else ''

        # Merged-identity surfacing: how many distinct emails/phones rolled up.
        distinct_emails = sorted({e for _, e in idn['emails_by_spend']}, key=str.lower)
        c['identity_emails'] = distinct_emails
        c['identity_email_count'] = len(distinct_emails)
        c['identity_phone_count'] = len(idn['phones'])
        c['identity_matched_by'] = sorted(cluster_meta.get(ck, {}).get('matched_by', set()))

        c['total_orders'] = idn['total_orders']
        c['successful_orders'] = idn['successful_orders']
        c['total_spent'] = round(idn['total_spent'], 2)
        c['undelivered_orders'] = idn['undelivered_orders']
        c['shipping_loss_total'] = round(idn['shipping_loss_total'], 2)
        c['problem_return_orders'] = idn['problem_return_orders']
        c['product_loss_total'] = round(idn['product_loss_total'], 2)
        c['total_loss'] = round(c['shipping_loss_total'] + c['product_loss_total'], 2)
        c['refusal_rate'] = round(c['undelivered_orders'] / c['total_orders'] * 100, 1) if c['total_orders'] else 0
        c['first_order_date'] = idn['first_order_date']
        c['last_order_date'] = idn['last_order_date']
        cur_str = sorted(idn['currencies'])[0] if idn['currencies'] else 'N/A'
        c['currency'] = cur_str
        c['currencies'] = ','.join(sorted(idn['currencies']))

        total_revenue += c['total_spent']

        if c['total_loss'] > 0:
            loss_stats['customers_with_loss'] += 1
        if c['undelivered_orders'] > 0:
            loss_stats['undelivered_customers'] += 1
        if c['problem_return_orders'] > 0:
            loss_stats['problem_return_customers'] += 1
        loss_stats['shipping_loss_total'] += c['shipping_loss_total']
        loss_stats['product_loss_total'] += c['product_loss_total']
        if c['shipping_loss_total'] or c['product_loss_total']:
            bucket = loss_stats['by_currency'].setdefault(cur_str, {'shipping': 0.0, 'product': 0.0})
            bucket['shipping'] += c['shipping_loss_total']
            bucket['product'] += c['product_loss_total']

        total_success_orders += c['successful_orders']
        if c['successful_orders'] > 1:
            repeat_customers += 1
            repeat_success_orders += c['successful_orders']
            repeat_spending += c['total_spent']
        if c['first_order_date'] and c['first_order_date'] >= thirty_days_ago:
            new_customers_month += 1
        elif c['first_order_date'] and c['first_order_date'] >= sixty_days_ago:
            new_customers_last_month += 1

        # Calculate Tier (same logic as API), now over the merged identity.
        avg_days = 0
        if c['successful_orders'] > 1 and c['first_order_date'] and c['last_order_date']:
            first = datetime.datetime.fromisoformat(c['first_order_date'])
            last = datetime.datetime.fromisoformat(c['last_order_date'])
            days = (last - first).days or 1
            avg_days = days / c['successful_orders']

        score = min(c['successful_orders'] * 10, 30) + min(c['total_spent'] / 100, 40)
        score += 30 if avg_days > 0 and avg_days < 60 else (15 if avg_days < 120 else 0)
        score = min(score, 100)

        if score >= 80: tier = 'VIP'
        elif score >= 60: tier = '优质'
        elif score >= 40: tier = '普通'
        else: tier = '新客'

        c['tier'] = tier
        tier_counts[tier] += 1

        # Smart Actions
        actions = []
        if tier == 'VIP':
            actions.append({'type': 'success', 'icon': 'gift', 'text': '专属礼遇'})
            actions.append({'type': 'primary', 'icon': 'people', 'text': '邀请入群'})
        elif c['last_order_date'] and c['last_order_date'] < ninety_days_ago:
            actions.append({'type': 'warning', 'icon': 'ticket-perforated', 'text': '召回优惠券'})
        elif c['successful_orders'] == 1 and c['first_order_date'] and c['first_order_date'] >= thirty_days_ago:
            actions.append({'type': 'info', 'icon': 'book', 'text': '欢迎指南'})
            actions.append({'type': 'info', 'icon': 'bag-plus', 'text': '关联推荐'})
        elif c['successful_orders'] > 3:
            actions.append({'type': 'primary', 'icon': 'arrow-repeat', 'text': '订阅服务'})
        c['actions'] = actions

        # Sources across the WHOLE identity (drives the 多站下单 column).
        sources = sorted(idn['sources'])
        c['source'] = sources[0] if sources else 'Unknown'
        c['all_sources'] = sources
        c['site_count'] = len(sources)
        c['site_list'] = [{
            'url': s,
            'short': s.replace('https://www.', '').replace('https://', ''),
            'manager': site_managers_global.get(s, ''),
        } for s in sources]

        customers_list.append(c)

    # Identity count drives all the per-customer rate stats below.
    total_customers = len(customers_list)
    # Initial server-side order: highest spenders first (DataTables re-sorts client-side).
    customers_list.sort(key=lambda x: x['total_spent'], reverse=True)

    # Calculate growth rate
    if new_customers_last_month > 0:
        growth_rate = ((new_customers_month - new_customers_last_month) / new_customers_last_month) * 100
    else:
        growth_rate = 100 if new_customers_month > 0 else 0

    loss_stats['shipping_loss_total'] = round(loss_stats['shipping_loss_total'], 2)
    loss_stats['product_loss_total'] = round(loss_stats['product_loss_total'], 2)
    loss_stats['combined_loss_total'] = round(
        loss_stats['shipping_loss_total'] + loss_stats['product_loss_total'], 2)
    # Round per-currency buckets and expose a sorted list for the template,
    # plus a single uniform_currency when there's exactly one in view.
    for cur, b in loss_stats['by_currency'].items():
        b['shipping'] = round(b['shipping'], 2)
        b['product'] = round(b['product'], 2)
    loss_stats['currency_rows'] = [
        {'currency': cur, 'shipping': b['shipping'], 'product': b['product']}
        for cur, b in sorted(loss_stats['by_currency'].items(),
                             key=lambda kv: -(kv[1]['shipping'] + kv[1]['product']))
    ]
    loss_stats['uniform_currency'] = (
        loss_stats['currency_rows'][0]['currency']
        if len(loss_stats['currency_rows']) == 1 else None
    )

    stats = {
        'total_customers': total_customers,
        'avg_ltv': total_revenue / total_customers if total_customers > 0 else 0,
        'repeat_rate': (repeat_customers / total_customers * 100) if total_customers > 0 else 0,
        # Weighted repeat-rate variants (switchable on the KPI card).
        'repeat_rate_orders': (repeat_success_orders / total_success_orders * 100) if total_success_orders > 0 else 0,
        'repeat_rate_revenue': (repeat_spending / total_revenue * 100) if total_revenue > 0 else 0,
        'repeat_customers': repeat_customers,
        'repeat_success_orders': repeat_success_orders,
        'total_success_orders': total_success_orders,
        'repeat_spending': round(repeat_spending, 2),
        'new_customer_rate': (new_customers_month / total_customers * 100) if total_customers > 0 else 0,
        'new_customers_month': new_customers_month,
        'new_customers_last_month': new_customers_last_month,
        'growth_rate': growth_rate,
        'tier_counts': tier_counts,
        'loss': loss_stats,
    }
    
    # Get site managers mapping
    conn2 = get_db_connection()
    sites = conn2.execute('SELECT url, manager FROM sites').fetchall()
    site_managers = {s['url']: s['manager'] or '' for s in sites}
    conn2.close()
    
    current_filters = {
        'source': source_filter,
        'manager': manager_filter,
        'country': country_filter,
        'quick_date': quick_date,
        'date_from': date_from,
        'date_to': date_to,
    }
    period_active = bool(date_from or date_to)
    return render_template('customers.html', customers=customers_list, stats=stats, sources=all_sources,
                           source_filter=source_filter, manager_filter=manager_filter, all_managers=all_managers,
                           site_managers=site_managers, all_countries=all_countries,
                           current_filters=current_filters, period_active=period_active)


@app.route('/api/customers/loss-list')
@login_required
def loss_customers_list():
    """货值损失客户名单: every customer with a 问题退货 (is_problem_return) order —
    调包/少件/损坏 — deduped by identity (email/phone/address) with loss totals.
    Scoped to the user's allowed sources. These are the customers the high-risk
    badge flags on re-order in 待发货/已发货."""
    conn = get_db_connection()
    allowed = get_user_allowed_sources(current_user.id, current_user.is_admin(), current_user.is_viewer())
    where = "WHERE COALESCE(is_problem_return, 0) = 1"
    params = []
    if allowed is not None:
        if not allowed:
            conn.close()
            return jsonify([])
        ph = ','.join(['?'] * len(allowed))
        where += f" AND source IN ({ph})"
        params = list(allowed)
    rows = conn.execute(f"""
        SELECT id, number, billing, shipping, source, problem_return_type,
               product_loss_amount, problem_return_at, currency
        FROM orders {where}
        ORDER BY problem_return_at DESC
    """, params).fetchall()
    conn.close()

    TYPE_MAP = {'swap': '调包/假货', 'short': '少件/空盒', 'damaged': '损坏', 'other': '其它'}
    customers = {}
    for r in rows:
        billing = parse_json_field(r['billing']) or {}
        shipping = parse_json_field(r['shipping']) or {}
        addr_d = _addr_for_order(billing, shipping)
        email = _normalize_email(billing.get('email') or shipping.get('email'))
        phone = _normalize_phone(addr_d.get('phone') or billing.get('phone'))
        addr_key = _normalize_address(addr_d)
        key = email or phone or addr_key or f"o{r['id']}"
        name = (f"{addr_d.get('first_name', '')} {addr_d.get('last_name', '')}".strip()
                or f"{billing.get('first_name', '')} {billing.get('last_name', '')}".strip())
        disp_addr = _compose_address(addr_d)
        c = customers.get(key)
        if not c:
            c = {'name': name, 'email': billing.get('email') or '',
                 'phone': addr_d.get('phone') or billing.get('phone') or '',
                 'address': disp_addr, 'count': 0, 'loss': 0.0, 'types': set(),
                 'orders': [], 'last_at': '', 'currency': ''}
            customers[key] = c
        c['count'] += 1
        c['loss'] += float(r['product_loss_amount'] or 0)
        if r['problem_return_type']:
            c['types'].add(r['problem_return_type'])
        if not c['currency'] and r['currency']:
            c['currency'] = r['currency']
        c['orders'].append({'number': r['number'],
                            'at': (r['problem_return_at'] or '')[:16].replace('T', ' '),
                            'loss': float(r['product_loss_amount'] or 0),
                            'source': (r['source'] or '').replace('https://www.', '').replace('https://', '')})
        if (r['problem_return_at'] or '') > c['last_at']:
            c['last_at'] = r['problem_return_at'] or ''
    out = []
    for c in customers.values():
        c['types'] = [TYPE_MAP.get(t, t) for t in sorted(c['types'])]
        c['loss'] = round(c['loss'], 2)
        c['last_at'] = (c['last_at'] or '')[:16].replace('T', ' ')
        out.append(c)
    out.sort(key=lambda x: (-x['loss'], -x['count']))
    return jsonify(out)


@app.route('/api/chart-data')
@login_required
def chart_data():
    """API endpoint for chart data"""
    conn = get_db_connection()
    
    chart_type = request.args.get('type', 'monthly')
    
    if chart_type == 'monthly':
        data = conn.execute('''
            SELECT strftime('%Y-%m', date_created) as month,
                   source,
                   COUNT(*) as orders,
                   SUM(total) as revenue
            FROM orders
            WHERE date_created >= date('now', '-12 months')
            GROUP BY month, source
            ORDER BY month
        ''').fetchall()
        
        result = {}
        for row in data:
            month = row['month']
            if month not in result:
                result[month] = {}
            result[month][row['source']] = {
                'orders': row['orders'],
                'revenue': row['revenue']
            }
        
        conn.close()
        return jsonify(result)
    
    elif chart_type == 'status':
        # Surface 'undelivered' as its own slice — covers the underlying status
        data = conn.execute('''
            SELECT
                CASE WHEN COALESCE(is_undelivered, 0) = 1 THEN 'undelivered' ELSE status END AS status,
                COUNT(*) as count
            FROM orders
            GROUP BY CASE WHEN COALESCE(is_undelivered, 0) = 1 THEN 'undelivered' ELSE status END
        ''').fetchall()
        
        conn.close()
        return jsonify([dict(row) for row in data])
    
    conn.close()
    return jsonify({})


@app.route('/api/order/<order_id>')
@login_required
def get_order_details(order_id):
    """API endpoint to get order details"""
    conn = get_db_connection()
    
    order = conn.execute('''
        SELECT * FROM orders WHERE id = ?
    ''', (order_id,)).fetchone()
    
    if not order:
        return jsonify({'error': 'Order not found'}), 404
    
    order_dict = dict(order)
    
    # Parse JSON fields
    order_dict['billing'] = parse_json_field(order['billing'])
    order_dict['shipping'] = parse_json_field(order['shipping'])
    order_dict['line_items'] = parse_json_field(order['line_items'])
    order_dict['shipping_lines'] = parse_json_field(order['shipping_lines'])
    order_dict['meta_data'] = parse_json_field(order['meta_data'])
    order_dict['fee_lines'] = parse_json_field(order['fee_lines'])
    order_dict['coupon_lines'] = parse_json_field(order['coupon_lines'])
    order_dict['coupon_lines'] = parse_json_field(order['coupon_lines'])
    order_dict['refunds'] = parse_json_field(order['refunds'])
    
    # Calculate customer total spending
    if order_dict['billing'] and order_dict['billing'].get('email'):
        email = order_dict['billing']['email']
        # 统计已完成订单（历史消费）
        customer_stats = conn.execute(f'''
            SELECT COUNT(*) as count, SUM(total) as total
            FROM orders
            WHERE json_extract(billing, '$.email') = ? AND {_success_status_cond()}
        ''', (email,)).fetchone()
        
        # 按状态分类统计所有订单 — 未送达视为独立分桶（覆盖底层 status）
        status_stats = conn.execute('''
            SELECT
                CASE WHEN COALESCE(is_undelivered, 0) = 1 THEN 'undelivered' ELSE status END AS status,
                COUNT(*) as count,
                SUM(total) as total
            FROM orders
            WHERE json_extract(billing, '$.email') = ?
            GROUP BY CASE WHEN COALESCE(is_undelivered, 0) = 1 THEN 'undelivered' ELSE status END
        ''', (email,)).fetchall()
        
        status_breakdown = {}
        for row in status_stats:
            status_breakdown[row['status']] = {
                'count': row['count'],
                'total': float(row['total'] or 0)
            }
        
        order_dict['customer_stats'] = {
            'total_orders': customer_stats['count'] if customer_stats else 0,
            'total_spent': float(customer_stats['total'] or 0) if customer_stats else 0,
            'status_breakdown': status_breakdown
        }
    
    # Get site credentials for API calls
    site_row = conn.execute('SELECT manager, url, consumer_key, consumer_secret FROM sites WHERE url = ?', (order_dict.get('source', ''),)).fetchone()
    order_dict['site_manager'] = site_row['manager'] if site_row and site_row['manager'] else ''
    
    # Fetch Order Notes from WooCommerce API.
    # Use orders.id (WC post ID); order['number'] can be a customer-facing
    # sequential number that the REST API rejects with 404.
    order_dict['order_notes'] = []
    if site_row and site_row['consumer_key'] and site_row['consumer_secret']:
        try:
            import requests as req
            api_url = f"{site_row['url']}/wp-json/wc/v3/orders/{woo_post_id(order['id'])}/notes"
            response = req.get(
                api_url,
                auth=(site_row['consumer_key'], site_row['consumer_secret']),
                timeout=5
            )
            if response.status_code == 200:
                order_dict['order_notes'] = response.json()
            else:
                print(f"Warning: Failed to fetch notes for order {order['id']}: {response.status_code} {response.text}")
        except Exception as e:
            print(f"Error fetching notes for order {order['id']}: {e}")

    # Get local shipping log if exists (for manually shipped orders not yet synced)
    shipping_log = conn.execute('SELECT tracking_number, carrier_slug, shipped_at FROM shipping_logs WHERE order_id = ?', (order_id,)).fetchone()
    if shipping_log:
        order_dict['shipping_log'] = dict(shipping_log)

    # Resolve who marked the order as undelivered (for audit display)
    if order_dict.get('is_undelivered') and order_dict.get('undelivered_by'):
        marker = conn.execute('SELECT name FROM users WHERE id = ?', (order_dict['undelivered_by'],)).fetchone()
        order_dict['undelivered_by_name'] = marker['name'] if marker else None

    # Resolve who marked the order as a problem return (for audit display)
    if order_dict.get('is_problem_return') and order_dict.get('problem_return_by'):
        pr_marker = conn.execute('SELECT name FROM users WHERE id = ?', (order_dict['problem_return_by'],)).fetchone()
        order_dict['problem_return_by_name'] = pr_marker['name'] if pr_marker else None

    if order_dict.get('warehouse_id'):
        wh = conn.execute('SELECT name FROM warehouses WHERE id=?', (order_dict['warehouse_id'],)).fetchone()
        order_dict['warehouse_name'] = wh['name'] if wh else None

    conn.close()

    return jsonify(order_dict)


@app.route('/api/customer/<email>')
@login_required
def get_customer_details(email):
    """API endpoint to get customer analysis data"""
    from urllib.parse import unquote
    email = unquote(email)
    
    conn = get_db_connection()

    # Resolve this customer's full IDENTITY (emails linked by a shared phone /
    # address), scoped to the sites this user may see, so the detail view
    # aggregates the same merged identity the customer list shows. Callers pass
    # just one email; we expand it to every email in the identity here.
    allowed_sources = get_user_allowed_sources(current_user.id, current_user.is_admin(), current_user.is_viewer())
    scope_conditions = ["status NOT IN ('checkout-draft', 'trash')"]
    scope_params = []
    if allowed_sources is not None:
        if allowed_sources:
            ph = ','.join(['?'] * len(allowed_sources))
            scope_conditions.append(f'source IN ({ph})')
            scope_params.extend(allowed_sources)
        else:
            scope_conditions.append('1=0')
    scope_where = 'WHERE ' + ' AND '.join(scope_conditions)

    norm_email = _normalize_email(email)
    email_to_cluster, cluster_meta = _resolve_identity_clusters(conn, scope_where, scope_params)
    cluster_key = email_to_cluster.get(norm_email) if norm_email else None
    if cluster_key and cluster_key in cluster_meta:
        identity_emails = sorted(cluster_meta[cluster_key]['emails'])
        identity_matched_by = sorted(cluster_meta[cluster_key]['matched_by'])
    else:
        identity_emails = [norm_email] if norm_email else []
        identity_matched_by = []

    # Get all orders across EVERY email in the identity (case-insensitive).
    if identity_emails:
        ident_ph = ','.join(['?'] * len(identity_emails))
        order_sql = f'''
            SELECT id, number, status, total, currency, shipping_total, date_created, source, line_items, billing,
                   payment_method, is_undelivered, shipping_loss_amount,
                   is_problem_return, delivery_confirmed, undelivered_note, carrier_status
            FROM orders
            WHERE lower(trim(json_extract(billing, '$.email'))) IN ({ident_ph})
              AND {' AND '.join(scope_conditions)}
            ORDER BY date_created DESC
        '''
        orders = conn.execute(order_sql, identity_emails + scope_params).fetchall()
    else:
        orders = conn.execute('''
            SELECT id, number, status, total, currency, shipping_total, date_created, source, line_items, billing,
                   payment_method, is_undelivered, shipping_loss_amount,
                   is_problem_return, delivery_confirmed, undelivered_note, carrier_status
            FROM orders
            WHERE json_extract(billing, '$.email') = ? AND status NOT IN ('checkout-draft', 'trash')
            ORDER BY date_created DESC
        ''', (email,)).fetchall()

    # On-hold is "shipped" only for sites flagged so (PL by default).
    on_hold_shipped_sources = {
        r['url'] for r in conn.execute(
            "SELECT url FROM sites WHERE cod_on_hold_is_shipped = 1"
        ).fetchall()
    }

    conn.close()

    if not orders:
        return jsonify({'error': 'Customer not found'}), 404

    # Process orders
    order_list = []
    all_products = {}
    site_spending = {}
    spending_by_currency = {}  # Track spending by currency
    total_spending = 0
    successful_orders = 0
    failed_orders = 0
    cancelled_orders = 0
    undelivered_orders = 0
    bad_orders = 0          # refused/returned/problem/COD-failed — drives the tier penalty
    meaningful_orders = 0   # shipped/attempted/completed (excl. unpaid-cancelled) — tier denominator
    shipping_loss_total = 0.0
    dates = []
    
    # Initialize customer details
    customer_phone = ''
    
    for order in orders:
        order_dict = dict(order)
        billing = parse_json_field(order['billing'])
        line_items = parse_json_field(order['line_items'])
        
        # Get customer name
        customer_name = f"{billing.get('first_name', '')} {billing.get('last_name', '')}".strip()
        
        # Only set if we haven't found a phone yet (processing from newest to oldest)
        if not customer_phone:
             customer_phone = billing.get('phone') or ''
        
        # Calculate order products
        order_products = []
        order_qty = 0
        for item in (line_items or []):
            # Get full product name including flavor from meta_data
            full_name, flavor, meta_puffs = get_full_product_name(item)
            product_name = full_name or item.get('name', 'Unknown')
            qty = item.get('quantity', 1)
            total = float(item.get('total', 0))
            order_products.append({
                'name': product_name,
                'quantity': qty,
                'total': total
            })
            order_qty += qty
            
            # Aggregate products
            if product_name in all_products:
                all_products[product_name]['quantity'] += qty
                all_products[product_name]['total'] += total
            else:
                all_products[product_name] = {'quantity': qty, 'total': total}
        
        # Site spending
        source = order['source'] or 'Unknown'
        if source in site_spending:
            site_spending[source]['orders'] += 1
            site_spending[source]['amount'] += float(order['total'] or 0)
        else:
            site_spending[source] = {'orders': 1, 'amount': float(order['total'] or 0)}
        
        # Status counting
        status = order['status']
        currency = order['currency'] or 'N/A'
        payment_method = order['payment_method'] or ''
        order_source = order['source'] or ''
        is_undelivered = bool(order['is_undelivered']) if 'is_undelivered' in order.keys() else False
        on_hold_is_success = (status == 'on-hold'
                              and payment_method != 'bacs'
                              and order_source in on_hold_shipped_sources)
        if is_undelivered:
            undelivered_orders += 1
            shipping_loss_total += float(order['shipping_loss_amount'] or 0) if 'shipping_loss_amount' in order.keys() else 0
        elif status in ['completed', 'shipped', 'delivered', 'partial-shipped'] or on_hold_is_success or (status == 'processing' and payment_method and payment_method != 'cod'):
            successful_orders += 1
            order_total = float(order['total'] or 0)
            total_spending += order_total
            # Track by currency
            if currency not in spending_by_currency:
                spending_by_currency[currency] = 0
            spending_by_currency[currency] += order_total
        elif status == 'failed':
            failed_orders += 1
        elif status == 'cancelled':
            cancelled_orders += 1

        # Customer-quality penalty inputs (tier-only; not revenue):
        # bad = refused/returned (manual) | carrier-returned | problem return | COD failed.
        carrier_st = (order['carrier_status'] if 'carrier_status' in order.keys() else None)
        is_problem = bool(order['is_problem_return']) if 'is_problem_return' in order.keys() else False
        if is_undelivered or carrier_st == 'returned' or is_problem \
                or (status == 'failed' and (payment_method or 'cod') == 'cod'):
            bad_orders += 1
        if status not in ('cancelled', 'checkout-draft', 'trash'):
            meaningful_orders += 1

        # Dates
        if order['date_created']:
            dates.append(order['date_created'][:10])
        
        order_list.append({
            'id': order['id'],
            'number': order['number'],
            'status': status,
            'total': float(order['total'] or 0),
            'currency': currency,
            'date_created': order['date_created'],
            'source': source.replace('https://www.', ''),
            'product_count': order_qty,
            'products': order_products,
            'email': (billing.get('email') or '').strip(),
            # Outcome / special status — so a suspicious customer's refused orders
            # are visible at a glance (these live in local flags, NOT the WC status).
            'is_undelivered': is_undelivered,
            'is_problem_return': bool(order['is_problem_return']) if 'is_problem_return' in order.keys() else False,
            'delivery_confirmed': bool(order['delivery_confirmed']) if 'delivery_confirmed' in order.keys() else False,
            'undelivered_note': (order['undelivered_note'] or '') if 'undelivered_note' in order.keys() else '',
            'shipping_loss_amount': float(order['shipping_loss_amount'] or 0) if 'shipping_loss_amount' in order.keys() else 0,
            'carrier_status': (order['carrier_status'] if 'carrier_status' in order.keys() else None),
        })
    
    # Calculate frequency
    unique_dates = sorted(set(dates))
    if len(unique_dates) >= 2:
        from datetime import datetime
        first_date = datetime.fromisoformat(unique_dates[0])
        last_date = datetime.fromisoformat(unique_dates[-1])
        days_span = (last_date - first_date).days or 1
        avg_days_between = days_span / len(unique_dates)
    else:
        avg_days_between = 0
    
    # Sort products by quantity
    sorted_products = sorted(all_products.items(), key=lambda x: x[1]['quantity'], reverse=True)
    
    # Customer quality score (simple algorithm)
    quality_score = 0
    quality_score += min(successful_orders * 10, 30)  # Max 30 for orders
    quality_score += min(total_spending / 100, 40)    # Max 40 for spending
    quality_score += 30 if avg_days_between > 0 and avg_days_between < 60 else (15 if avg_days_between < 120 else 0)  # Frequency bonus
    quality_score = min(quality_score, 100)
    
    # Check for manual override — honor a tier set on ANY email in the identity.
    conn = get_db_connection()
    if identity_emails:
        ms_ph = ','.join(['?'] * len(identity_emails))
        manual_setting = conn.execute(
            f"SELECT quality_tier FROM customer_settings "
            f"WHERE lower(trim(email)) IN ({ms_ph}) AND COALESCE(quality_tier,'auto') != 'auto' LIMIT 1",
            identity_emails
        ).fetchone()
    else:
        manual_setting = conn.execute('SELECT quality_tier FROM customer_settings WHERE email = ?', (email,)).fetchone()
    conn.close()
    manual_tier = manual_setting['quality_tier'] if manual_setting else 'auto'
    
    # Determine customer tier
    if manual_tier != 'auto':
        if manual_tier == 'vip':
            customer_tier = {'level': 'VIP', 'color': '#f59e0b', 'icon': 'star-fill', 'manual': True}
        elif manual_tier == 'good':
            customer_tier = {'level': '优质', 'color': '#10b981', 'icon': 'gem', 'manual': True}
        elif manual_tier == 'normal':
            customer_tier = {'level': '普通', 'color': '#3b82f6', 'icon': 'person-check', 'manual': True}
        elif manual_tier == 'new':
            customer_tier = {'level': '新客', 'color': '#0dcaf0', 'icon': 'stars', 'manual': True}
        elif manual_tier == 'bad':
            customer_tier = {'level': '劣质', 'color': '#ef4444', 'icon': 'x-circle', 'manual': True}
        else:
            # Fallback to auto if unknown
            manual_tier = 'auto'
            
    if manual_tier == 'auto':
        tier = calculate_customer_tier(successful_orders, total_spending, avg_days_between,
                                       bad_orders=bad_orders, meaningful_orders=meaningful_orders)
        if tier == 'vip':
            customer_tier = {'level': 'VIP', 'color': '#f59e0b', 'icon': 'star-fill', 'manual': False}
        elif tier == 'good':
            customer_tier = {'level': '优质', 'color': '#10b981', 'icon': 'gem', 'manual': False}
        elif tier == 'normal':
            customer_tier = {'level': '普通', 'color': '#3b82f6', 'icon': 'person-check', 'manual': False}
        elif tier == 'bad':
            customer_tier = {'level': '劣质', 'color': '#ef4444', 'icon': 'x-circle', 'manual': False}
        else:
            customer_tier = {'level': '新客', 'color': '#0dcaf0', 'icon': 'stars', 'manual': False}

    customer_tier['value'] = manual_tier
    
    # Calculate CNY total for customer spending
    from datetime import datetime
    current_month = datetime.now().strftime('%Y-%m')
    spending_cny = 0
    for currency, amount in spending_by_currency.items():
        rate, _ = get_cny_rate(currency, current_month)
        if rate:
            spending_cny += amount * rate

    # Blocklist status (auto-cancel COD), matched by normalized phone.
    norm_ph = _normalize_phone(customer_phone)
    is_blocked = False
    block_info = None
    if norm_ph:
        conn_b = get_db_connection()
        brow = conn_b.execute(
            "SELECT * FROM blocked_customers WHERE phone = ?", (norm_ph,)
        ).fetchone()
        conn_b.close()
        if brow:
            is_blocked = True
            block_info = {
                'phone': brow['phone'],
                'reason': brow['reason'],
                'auto_cancel': bool(brow['auto_cancel']),
                'cancelled_count': brow['cancelled_count'] or 0,
                'created_by': brow['created_by'],
                'created_at': brow['created_at'],
                'last_cancelled_at': brow['last_cancelled_at'],
            }

    result = {
        'email': email,
        'name': customer_name,
        'phone': customer_phone,
        # Identity surfacing: which emails were merged + why.
        'identity_emails': identity_emails,
        'identity_email_count': len(identity_emails),
        'identity_matched_by': identity_matched_by,
        'total_orders': len(order_list),
        'successful_orders': successful_orders,
        'failed_orders': failed_orders,
        'cancelled_orders': cancelled_orders,
        'undelivered_orders': undelivered_orders,
        'shipping_loss_total': round(shipping_loss_total, 2),
        'refusal_rate': round(undelivered_orders / len(order_list) * 100, 1) if order_list else 0,
        'total_spending': total_spending,
        'spending_by_currency': spending_by_currency,  # New: currency breakdown
        'spending_cny': round(spending_cny, 2),  # New: CNY total
        'avg_order_value': total_spending / successful_orders if successful_orders > 0 else 0,
        'first_order_date': unique_dates[0] if unique_dates else None,
        'last_order_date': unique_dates[-1] if unique_dates else None,
        'avg_days_between_orders': round(avg_days_between, 1),
        'quality_score': round(quality_score),
        'customer_tier': customer_tier,
        'is_blocked': is_blocked,
        'block_info': block_info,
        'phone_normalized': norm_ph,
        'site_spending': [{'site': k.replace('https://www.', ''), 'orders': v['orders'], 'amount': v['amount']} for k, v in site_spending.items()],
        'top_products': [{'name': k, 'quantity': v['quantity'], 'total': v['total']} for k, v in sorted_products[:10]],
        'orders': order_list[:20]  # Limit to 20 most recent
    }
    
    return jsonify(result)


@app.route('/api/customer/quality', methods=['POST'])
@login_required
@editor_required
def update_customer_quality():
    """Update customer quality tier manually"""
    data = request.json
    email = data.get('email')
    quality = data.get('quality')
    
    if not email or not quality:
        return jsonify({'success': False, 'error': 'Missing required fields'}), 400
        
    conn = get_db_connection()
    try:
        conn.execute('''
            INSERT INTO customer_settings (email, quality_tier, updated_at) 
            VALUES (?, ?, datetime('now'))
            ON CONFLICT(email) DO UPDATE SET 
                quality_tier = excluded.quality_tier,
                updated_at = excluded.updated_at
        ''', (email, quality))
        conn.commit()
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500
    finally:
        conn.close()


@app.route('/api/customer/block', methods=['POST'])
@login_required
def block_customer():
    """Add a customer to the auto-cancel blocklist, keyed by normalized phone.
    Their PRE-SHIPMENT COD orders get auto-cancelled by the hourly enforcer;
    prepaid orders are never touched. Who may do this is configurable in
    用户管理 (per-user can_manage_blocklist permission)."""
    if not current_user.can_manage_blocklist():
        return jsonify({'success': False, 'error': '你没有管理拦截名单的权限'}), 403
    data = request.json or {}
    phone_raw = (data.get('phone') or '').strip()
    email = (data.get('email') or '').strip()
    name = (data.get('name') or '').strip()
    reason = (data.get('reason') or '').strip() or '惯性COD拒收'
    auto_cancel = 0 if data.get('auto_cancel') is False else 1

    norm = _normalize_phone(phone_raw)
    if not norm:
        return jsonify({'success': False, 'error': '该客户没有有效手机号，无法按手机号拦截'}), 400

    conn = get_db_connection()
    try:
        conn.execute('''
            INSERT INTO blocked_customers
                (phone, raw_phone, name, email, reason, auto_cancel, created_by, created_at)
            VALUES (?, ?, ?, ?, ?, ?, ?, datetime('now'))
            ON CONFLICT(phone) DO UPDATE SET
                raw_phone = excluded.raw_phone,
                name = excluded.name,
                email = excluded.email,
                reason = excluded.reason,
                auto_cancel = excluded.auto_cancel
        ''', (norm, phone_raw, name, email, reason, auto_cancel, current_user.name))
        conn.commit()
        return jsonify({'success': True, 'phone': norm})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500
    finally:
        conn.close()


@app.route('/api/customer/unblock', methods=['POST'])
@login_required
def unblock_customer():
    """Remove a customer from the auto-cancel blocklist."""
    if not current_user.can_manage_blocklist():
        return jsonify({'success': False, 'error': '你没有管理拦截名单的权限'}), 403
    data = request.json or {}
    phone_raw = (data.get('phone') or '').strip()
    norm = _normalize_phone(phone_raw) or phone_raw  # accept raw or already-normalized
    if not norm:
        return jsonify({'success': False, 'error': '缺少手机号'}), 400
    conn = get_db_connection()
    try:
        conn.execute("DELETE FROM blocked_customers WHERE phone = ?", (norm,))
        conn.commit()
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500
    finally:
        conn.close()


@app.route('/api/blocklist/enforce', methods=['POST'])
@login_required
@admin_required
def blocklist_enforce_now():
    """Run blocklist enforcement on demand. Body {dry_run: true} previews only."""
    data = request.get_json(silent=True) or {}
    dry = bool(data.get('dry_run', False))
    conn = get_db_connection()
    try:
        summary = blocklist.enforce(conn, dry_run=dry, actor=f'manual:{current_user.name}')
        return jsonify({'success': True, 'summary': summary})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500
    finally:
        conn.close()


# Status translation — COD mode (default)
STATUS_LABELS = {
    'pending': '待处理',
    'processing': '待发货',
    'on-hold': '已发货',
    'shipped': '已发货',
    'completed': '已完成',
    'cancelled': '已取消',
    'refunded': '已退款',
    'failed': '失败',
    'delivered': '已送达',
    'partial-shipped': '部分发货',
    'checkout-draft': '草稿',
    'undelivered': '未送达',
}

# Status translation — Online payment mode (Stripe, credit card, etc.)
STATUS_LABELS_ONLINE = {
    'pending': '待支付',
    'processing': '已支付·待发货',
    'on-hold': '已发货',
    'shipped': '已发货',
    'completed': '已完成',
    'cancelled': '已取消',
    'refunded': '已退款',
    'failed': '支付失败',
    'delivered': '已送达',
    'partial-shipped': '部分发货',
    'checkout-draft': '草稿',
    'undelivered': '未送达',
}

# Status translation — Bank transfer mode (bacs)
# Key difference: on-hold = awaiting bank transfer, NOT shipped
STATUS_LABELS_BACS = {
    'pending': '待处理',
    'on-hold': '待转账确认',
    'processing': '已确认·待发货',
    'shipped': '已发货',
    'completed': '已完成',
    'cancelled': '已取消',
    'refunded': '已退款',
    'failed': '失败',
    'delivered': '已送达',
    'partial-shipped': '部分发货',
    'checkout-draft': '草稿',
    'undelivered': '未送达',
}


def get_status_label(status, payment_method=None):
    """Get Chinese status label, context-aware based on payment method.
    Three modes: COD, bacs (bank transfer), online (Stripe etc.)."""
    if payment_method == 'bacs':
        return STATUS_LABELS_BACS.get(status, STATUS_LABELS.get(status, status))
    elif payment_method and payment_method != 'cod':
        return STATUS_LABELS_ONLINE.get(status, STATUS_LABELS.get(status, status))
    return STATUS_LABELS.get(status, status)


def is_cod(payment_method):
    """Check if a payment method is COD (cash on delivery).
    Treats None/empty as COD (most PL sites default to COD)."""
    return not payment_method or payment_method == 'cod'


# ========== Reusable SQL fragments for order success classification ==========
# These account for the difference between COD and online payment modes.
#
# COD success:    on-hold + completed (current behavior preserved)
# Online success: processing + on-hold + completed (processing = already paid)
# Never success:  pending (online), failed, cancelled, checkout-draft, trash, cheat

def _on_hold_is_shipped_clause(prefix=''):
    """SQL fragment used by the status helpers below.
    True when the order's site has cod_on_hold_is_shipped=1 — i.e. on-hold
    is the local convention for "shipped" (PL COD workflow). Without this
    gate, on-hold orders from AU/AE would wrongly count as success."""
    p = f'{prefix}.' if prefix else ''
    return (f"EXISTS (SELECT 1 FROM sites _s WHERE _s.url = {p}source "
            f"AND _s.cod_on_hold_is_shipped = 1)")


def _success_status_case(prefix='', as_name='is_success'):
    """SQL CASE expression: 1 if order is 'successful', 0 otherwise.
    Three payment modes:
      COD:   on-hold(shipped) + shipped/delivered/partial-shipped + completed
      bacs:  on-hold is NOT success (awaiting bank transfer) — only after processing
      Online(Stripe): processing(paid) + shipped/delivered/partial-shipped + completed
    The on-hold→success rule additionally requires that the order's site has
    cod_on_hold_is_shipped=1 (set per-country in /settings). PL sites have it
    on by default; AU/AE have it off, so on-hold there is never a success.
    Undelivered (refused/returned) and 问题退货 (swap/short/damaged) are never
    a success regardless of payment mode."""
    p = f'{prefix}.' if prefix else ''
    on_hold_clause = _on_hold_is_shipped_clause(prefix)
    return f"""(CASE WHEN ({p}status IN ('completed','shipped','delivered','partial-shipped')
                    OR ({p}status = 'on-hold' AND COALESCE({p}payment_method,'cod') != 'bacs'
                        AND {on_hold_clause})
                    OR ({p}status = 'processing' AND COALESCE({p}payment_method,'cod') != 'cod'))
                    AND COALESCE({p}is_undelivered, 0) = 0
                    AND COALESCE({p}is_problem_return, 0) = 0
               THEN 1 ELSE 0 END)"""


def _bad_order_case(prefix='', as_name='is_bad'):
    """SQL CASE: 1 if the order is a 'bad' outcome for customer-quality scoring.
    Bad = refused/returned (manual flag), carrier-detected return, problem return
    (swap/short/damaged), or a COD order that failed (refused at the door).
    This is ONLY used by the customer-tier penalty — it does NOT touch revenue or
    reconciliation (those keep using _success_/_revenue_ helpers unchanged)."""
    p = f'{prefix}.' if prefix else ''
    return f"""(CASE WHEN COALESCE({p}is_undelivered, 0) = 1
                    OR {p}carrier_status = 'returned'
                    OR COALESCE({p}is_problem_return, 0) = 1
                    OR ({p}status = 'failed' AND COALESCE({p}payment_method,'cod') = 'cod')
               THEN 1 ELSE 0 END)"""


def _meaningful_order_case(prefix=''):
    """SQL CASE: 1 if the order counts toward the customer-quality denominator —
    i.e. a real fulfillment attempt/outcome. Excludes unpaid-cancelled (abandoned)
    and drafts/trash, which would otherwise dilute the refusal rate."""
    p = f'{prefix}.' if prefix else ''
    return (f"(CASE WHEN {p}status NOT IN ('cancelled','checkout-draft','trash') "
            f"THEN 1 ELSE 0 END)")


def _success_status_cond(prefix=''):
    """SQL boolean condition: true if order is 'successful'.
    Same logic as _success_status_case but for WHERE clauses."""
    p = f'{prefix}.' if prefix else ''
    on_hold_clause = _on_hold_is_shipped_clause(prefix)
    return f"""(({p}status IN ('completed','shipped','delivered','partial-shipped')
        OR ({p}status = 'on-hold' AND COALESCE({p}payment_method,'cod') != 'bacs'
            AND {on_hold_clause})
        OR ({p}status = 'processing' AND COALESCE({p}payment_method,'cod') != 'cod'))
        AND COALESCE({p}is_undelivered, 0) = 0
        AND COALESCE({p}is_problem_return, 0) = 0)"""


def _revenue_status_cond(prefix=''):
    """SQL boolean condition: true if order should count toward revenue.
    Excludes: failed/cancelled/draft/trash/cheat, refunded (fully refunded —
    money was returned, so no net revenue), online-pending, bacs-on-hold,
    on-hold orders from sites that don't treat on-hold as shipped (AU/AE),
    undelivered (package returned, no money collected), and 问题退货 (contents
    were wrong/missing/damaged, customer never paid or money was refunded)."""
    p = f'{prefix}.' if prefix else ''
    on_hold_clause = _on_hold_is_shipped_clause(prefix)
    return f"""({p}status NOT IN ('failed','cancelled','checkout-draft','trash','cheat','refunded')
        AND NOT ({p}status = 'pending' AND COALESCE({p}payment_method,'cod') != 'cod')
        AND NOT ({p}status = 'on-hold' AND {p}payment_method = 'bacs')
        AND NOT ({p}status = 'on-hold' AND NOT {on_hold_clause})
        AND COALESCE({p}is_undelivered, 0) = 0
        AND COALESCE({p}is_problem_return, 0) = 0)"""


def _active_status_cond(prefix=''):
    """SQL boolean condition: like _revenue_status_cond but a narrower exclusion list
    (only filters failed/cancelled/refunded, not draft/trash/cheat). Used in legacy
    trend / chart queries where draft/trash never carried real data anyway. Kept
    distinct so behavior matches the previous inline strings exactly — do NOT
    consolidate without checking each callsite's data semantics."""
    p = f'{prefix}.' if prefix else ''
    on_hold_clause = _on_hold_is_shipped_clause(prefix)
    return f"""({p}status NOT IN ('failed','cancelled','refunded')
        AND NOT ({p}status = 'pending' AND COALESCE({p}payment_method,'cod') != 'cod')
        AND NOT ({p}status = 'on-hold' AND {p}payment_method = 'bacs')
        AND NOT ({p}status = 'on-hold' AND NOT {on_hold_clause})
        AND COALESCE({p}is_undelivered, 0) = 0
        AND COALESCE({p}is_problem_return, 0) = 0)"""


def _success_amount_case(col, prefix=''):
    """SQL fragment: SUM(CASE WHEN <successful> THEN <col> ELSE 0 END).
    Use for summing amounts/quantities of orders that count as 'successful'."""
    p = f'{prefix}.' if prefix else ''
    return f"SUM(CASE WHEN {_success_status_cond(prefix)} THEN {p}{col} ELSE 0 END)"


def _revenue_amount_case(col, prefix=''):
    """SQL fragment: SUM(CASE WHEN <counts toward revenue> THEN <col> ELSE 0 END)."""
    p = f'{prefix}.' if prefix else ''
    return f"SUM(CASE WHEN {_revenue_status_cond(prefix)} THEN {p}{col} ELSE 0 END)"

# Currency symbols mapping
CURRENCY_SYMBOLS = {
    'PLN': 'zł',      # Polish Złoty
    'AUD': 'A$',      # Australian Dollar
    'AED': 'د.إ',     # UAE Dirham
    'USD': '$',       # US Dollar
    'EUR': '€',       # Euro
    'GBP': '£',       # British Pound
}


def format_amount_with_currency(amount, currency):
    """Format amount with currency code, e.g., '145.00 PLN'"""
    try:
        return f"{float(amount):,.2f} {currency}"
    except:
        return f"{amount} {currency}"


def get_currency_symbol(currency):
    """Get symbol for a currency code"""
    return CURRENCY_SYMBOLS.get(currency, currency)


@app.template_filter('status_label')
def status_label_filter(status, payment_method=None, is_undelivered=0):
    """Jinja filter: {{ order.status|status_label(order.payment_method, order.is_undelivered) }}.
    is_undelivered=1 wins over the underlying status (kept on a separate column to
    survive WooCommerce sync overwrites — see init_undelivered_columns)."""
    if is_undelivered:
        return '未送达'
    return get_status_label(status, payment_method)


@app.template_filter('status_class')
def status_class_filter(status, payment_method=None, is_undelivered=0):
    """Jinja filter: returns CSS class for status badge, payment-mode-aware.
    bacs + on-hold uses pending style (yellow = awaiting payment)."""
    if is_undelivered:
        return 'status-undelivered'
    if status == 'on-hold' and payment_method == 'bacs':
        return 'status-pending'
    return f'status-{status}'


@app.template_filter('format_currency')
def format_currency_filter(value, currency=None):
    """Format currency value, optionally with currency code"""
    try:
        formatted = f"{float(value):,.2f}"
        if currency:
            return f"{formatted} {currency}"
        return formatted
    except:
        return value


@app.template_filter('format_date')
def format_date_filter(value):
    try:
        dt = datetime.fromisoformat(value.replace('Z', '+00:00'))
        return dt.strftime('%Y-%m-%d %H:%M')
    except:
        return value



# -----------------------------
# Data Synchronization Features
# -----------------------------

def init_sites_table():
    """Initialize sites table"""
    conn = get_db_connection()
    conn.execute('''
        CREATE TABLE IF NOT EXISTS sites (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            url TEXT NOT NULL,
            consumer_key TEXT NOT NULL,
            consumer_secret TEXT NOT NULL,
            manager TEXT,
            last_sync TEXT,
            api_status TEXT DEFAULT 'unknown',
            last_api_error TEXT
        )
    ''')
    # Add manager column if not exists (for existing databases)
    try:
        conn.execute('ALTER TABLE sites ADD COLUMN manager TEXT')
    except:
        pass  # Column already exists
    # Add api_status column if not exists
    try:
        conn.execute("ALTER TABLE sites ADD COLUMN api_status TEXT DEFAULT 'unknown'")
    except:
        pass  # Column already exists
    # Add last_api_error column if not exists
    try:
        conn.execute('ALTER TABLE sites ADD COLUMN last_api_error TEXT')
    except:
        pass  # Column already exists
    # Add tracking_api_status column for woo-tracking REST API plugin availability
    try:
        conn.execute('ALTER TABLE sites ADD COLUMN tracking_api_status TEXT')
    except:
        pass  # Column already exists
    # Add country column for site categorization by country
    try:
        conn.execute('ALTER TABLE sites ADD COLUMN country TEXT')
    except:
        pass  # Column already exists
    # Add product_master_id for WooMultistore-managed sites
    # (NULL = standalone site, otherwise points to product_masters.id)
    try:
        conn.execute('ALTER TABLE sites ADD COLUMN product_master_id INTEGER')
    except:
        pass  # Column already exists
    # Whether `status=on-hold` counts as "shipped/success" for this site.
    # In Poland the COD workflow puts shipped COD orders into on-hold until the
    # carrier collects payment, so on-hold is real revenue. In AU/AE there is no
    # COD; on-hold there means "received, waiting for something" — never shipped.
    # Initial value is set per-country below (PL→1, others→0); admins can
    # toggle individual sites on the settings page afterward.
    try:
        conn.execute('ALTER TABLE sites ADD COLUMN cod_on_hold_is_shipped INTEGER')
    except:
        pass  # Column already exists
    # Backfill ONLY rows that still have NULL (i.e. never touched by an admin).
    # Once an admin toggles a row to 0 or 1, this WHERE clause skips it on the
    # next boot. New sites added later also get classified here on first boot
    # after their country is set.
    conn.execute("""
        UPDATE sites SET cod_on_hold_is_shipped = 1
        WHERE cod_on_hold_is_shipped IS NULL AND country = 'PL'
    """)
    conn.execute("""
        UPDATE sites SET cod_on_hold_is_shipped = 0
        WHERE cod_on_hold_is_shipped IS NULL AND country IS NOT NULL AND country != 'PL'
    """)
    conn.commit()
    conn.close()


def init_product_masters_table():
    """Initialize product_masters table — stores WooCommerce master site
    credentials used for product CRUD on multistore-managed sites. These are
    intentionally separate from `sites` so they never appear in any order /
    customer / analytics view (which all query `sites`)."""
    conn = get_db_connection()
    conn.execute('''
        CREATE TABLE IF NOT EXISTS product_masters (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            label TEXT NOT NULL,
            url TEXT NOT NULL,
            consumer_key TEXT NOT NULL,
            consumer_secret TEXT NOT NULL,
            api_status TEXT DEFAULT 'unknown',
            last_api_error TEXT,
            last_tested_at TEXT,
            created_at TEXT DEFAULT CURRENT_TIMESTAMP,
            updated_at TEXT DEFAULT CURRENT_TIMESTAMP
        )
    ''')
    try:
        conn.execute('CREATE INDEX IF NOT EXISTS idx_product_masters_url ON product_masters(url)')
    except sqlite3.OperationalError:
        pass
    conn.commit()
    conn.close()


def init_sync_logs_table():
    """Initialize sync_logs table for storing synchronization history"""
    conn = get_db_connection()
    conn.execute('''
        CREATE TABLE IF NOT EXISTS sync_logs (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            site_id INTEGER,
            site_url TEXT,
            status TEXT,
            message TEXT,
            new_orders INTEGER DEFAULT 0,
            updated_orders INTEGER DEFAULT 0,
            sync_time TEXT,
            duration_seconds INTEGER
        )
    ''')
    conn.commit()
    conn.close()


def init_settings_table():
    """Initialize settings table for storing app configuration"""
    conn = get_db_connection()
    conn.execute('''
        CREATE TABLE IF NOT EXISTS settings (
            key TEXT PRIMARY KEY,
            value TEXT
        )
    ''')
    # Insert default autosync settings if not exist
    conn.execute('''
        INSERT OR IGNORE INTO settings (key, value) VALUES ('autosync_enabled', 'false')
    ''')
    conn.execute('''
        INSERT OR IGNORE INTO settings (key, value) VALUES ('autosync_interval', '900')
    ''')
    # Note: source_display_mode is now per-user, stored in user_preferences table
    conn.commit()
    conn.close()


def init_user_preferences_table():
    """Initialize user_preferences table for storing per-user settings"""
    conn = get_db_connection()
    conn.execute('''
        CREATE TABLE IF NOT EXISTS user_preferences (
            user_id INTEGER,
            preference_key TEXT,
            preference_value TEXT,
            PRIMARY KEY (user_id, preference_key),
            FOREIGN KEY (user_id) REFERENCES users(id) ON DELETE CASCADE
        )
    ''')
    conn.commit()
    conn.close()


@app.context_processor
def inject_settings():
    """Inject user-specific settings into all templates"""
    # Default value
    source_display_mode = 'full'
    
    # Try to get user-specific preference if logged in
    if current_user.is_authenticated:
        conn = get_db_connection()
        pref = conn.execute('''
            SELECT preference_value FROM user_preferences 
            WHERE user_id = ? AND preference_key = 'source_display_mode'
        ''', (current_user.id,)).fetchone()
        conn.close()
        if pref:
            source_display_mode = pref['preference_value']
    
    return dict(source_display_mode=source_display_mode)


def save_sync_log(site_id, site_url, status, message, new_orders=0, updated_orders=0, duration_seconds=0):
    """Save a sync log entry to the database"""
    conn = get_db_connection()
    conn.execute('''
        INSERT INTO sync_logs (site_id, site_url, status, message, new_orders, updated_orders, sync_time, duration_seconds)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?)
    ''', (site_id, site_url, status, message, new_orders, updated_orders, datetime.now().strftime('%Y-%m-%d %H:%M:%S'), duration_seconds))
    conn.commit()
    conn.close()

def init_users_table():
    """Initialize users table"""
    conn = get_db_connection()
    conn.execute('''
        CREATE TABLE IF NOT EXISTS users (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            username TEXT UNIQUE NOT NULL,
            password_hash TEXT NOT NULL,
            name TEXT,
            role TEXT DEFAULT 'user',
            created_at TEXT DEFAULT CURRENT_TIMESTAMP
        )
    ''')
    conn.execute('''
        CREATE TABLE IF NOT EXISTS user_site_permissions (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER,
            site_id INTEGER,
            UNIQUE(user_id, site_id)
        )
    ''')
    # Country-level grant: user can see ALL sites in this country, including any
    # added later (auto-inherit). Resolved live against sites.country, so new
    # sites need no per-user click. See get_user_allowed_sources.
    conn.execute('''
        CREATE TABLE IF NOT EXISTS user_country_permissions (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER,
            country TEXT,
            UNIQUE(user_id, country)
        )
    ''')
    # Exclusions carved out of a country grant — a site in a granted country the
    # user should NOT see. (user_site_permissions stays the explicit GRANT list
    # for sites in non-granted countries.)
    conn.execute('''
        CREATE TABLE IF NOT EXISTS user_site_exclusions (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER,
            site_id INTEGER,
            UNIQUE(user_id, site_id)
        )
    ''')
    conn.commit()
    
    # Migrate default admin user if not exists
    existing = conn.execute('SELECT id FROM users WHERE username = ?', ('admin',)).fetchone()
    if not existing:
        conn.execute('''
            INSERT INTO users (username, password_hash, name, role)
            VALUES (?, ?, ?, ?)
        ''', ('admin', generate_password_hash('admin123'), '管理员', 'admin'))
        conn.commit()
    
    conn.close()


def init_undelivered_columns():
    """Add columns to orders table for tracking undelivered (refused/returned) orders.
    Kept on a separate flag column instead of overloading status so the WC-sourced
    status field can keep flowing through sync untouched. The WC sync paths
    (sync_utils.save_orders_to_db and 1.wooorders_sqlite.py) use UPSERT that only
    updates WC-managed columns, so these local flags survive every refresh."""
    conn = get_db_connection()
    for ddl in (
        'ALTER TABLE orders ADD COLUMN is_undelivered INTEGER DEFAULT 0',
        'ALTER TABLE orders ADD COLUMN shipping_loss_amount REAL DEFAULT 0',
        'ALTER TABLE orders ADD COLUMN undelivered_at TEXT',
        'ALTER TABLE orders ADD COLUMN undelivered_by INTEGER',
        'ALTER TABLE orders ADD COLUMN undelivered_note TEXT',
    ):
        try:
            conn.execute(ddl)
        except sqlite3.OperationalError:
            pass  # Column already exists
    try:
        conn.execute('CREATE INDEX IF NOT EXISTS idx_orders_is_undelivered ON orders(is_undelivered)')
    except sqlite3.OperationalError:
        pass
    conn.commit()
    conn.close()


def init_problem_return_columns():
    """Add columns to orders for tracking problem returns — packages that came
    back with wrong/missing/damaged contents (e.g. the brick-swap scam).
    Distinct from is_undelivered: there the package never reached the customer
    and only the shipping fee is lost; here the goods themselves are lost."""
    conn = get_db_connection()
    for ddl in (
        'ALTER TABLE orders ADD COLUMN is_problem_return INTEGER DEFAULT 0',
        'ALTER TABLE orders ADD COLUMN problem_return_type TEXT',
        'ALTER TABLE orders ADD COLUMN product_loss_amount REAL DEFAULT 0',
        'ALTER TABLE orders ADD COLUMN problem_return_at TEXT',
        'ALTER TABLE orders ADD COLUMN problem_return_by INTEGER',
        'ALTER TABLE orders ADD COLUMN problem_return_note TEXT',
        'ALTER TABLE orders ADD COLUMN problem_return_evidence TEXT',
    ):
        try:
            conn.execute(ddl)
        except sqlite3.OperationalError:
            pass  # Column already exists
    try:
        conn.execute('CREATE INDEX IF NOT EXISTS idx_orders_is_problem_return ON orders(is_problem_return)')
    except sqlite3.OperationalError:
        pass
    conn.commit()
    conn.close()


def init_shipping_tables():
    """Initialize shipping-related tables"""
    conn = get_db_connection()

    # Add can_ship column to users table if not exists
    try:
        conn.execute('ALTER TABLE users ADD COLUMN can_ship INTEGER DEFAULT 0')
        conn.commit()
    except:
        pass  # Column already exists

    # Add can_view_shipping — view-only access to 发货管理 (see lists/status but
    # cannot operate). Anyone with can_ship implicitly has view.
    try:
        conn.execute('ALTER TABLE users ADD COLUMN can_view_shipping INTEGER DEFAULT 0')
        conn.commit()
    except:
        pass  # Column already exists

    # Shipping logs table
    conn.execute('''
        CREATE TABLE IF NOT EXISTS shipping_logs (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            order_id INTEGER NOT NULL,
            woo_order_id INTEGER NOT NULL,
            source TEXT NOT NULL,
            tracking_number TEXT NOT NULL,
            carrier_slug TEXT,
            shipped_by INTEGER,
            shipped_at TEXT DEFAULT CURRENT_TIMESTAMP,
            completed_at TEXT,
            status TEXT DEFAULT 'shipped'
        )
    ''')

    # Re-shipment (补发货) support. A reship appends a NEW shipping_logs row that
    # carries the reason while the ORIGINAL parcel row is kept for history.
    # is_reship distinguishes it from a normal/split parcel so the shipped list
    # can render the prior tracking as superseded rather than as a 分批 parcel.
    for ddl in (
        'ALTER TABLE shipping_logs ADD COLUMN reship_reason TEXT',
        'ALTER TABLE shipping_logs ADD COLUMN is_reship INTEGER DEFAULT 0',
    ):
        try:
            conn.execute(ddl)
        except sqlite3.OperationalError:
            pass  # Column already exists

    # Shipping carriers table
    conn.execute('''
        CREATE TABLE IF NOT EXISTS shipping_carriers (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            slug TEXT NOT NULL UNIQUE,
            name TEXT NOT NULL,
            tracking_url TEXT,
            is_active INTEGER DEFAULT 1
        )
    ''')
    conn.commit()
    
    # Insert default carriers if table is empty
    existing = conn.execute('SELECT COUNT(*) FROM shipping_carriers').fetchone()[0]
    if existing == 0:
        default_carriers = [
            ('inpost', 'InPost', 'https://inpost.pl/sledzenie-przesylek?number={tracking}'),
            ('dpd', 'DPD', 'https://www.dpd.com.pl/tracking?q={tracking}'),
        ]
        conn.executemany('INSERT INTO shipping_carriers (slug, name, tracking_url) VALUES (?, ?, ?)', default_carriers)
        conn.commit()

    # Backfill carriers for orders shipped from China to overseas markets.
    # Idempotent: INSERT OR IGNORE only adds the row when slug is new, so
    # repeat boots don't reset URLs that an admin may have edited later.
    conn.execute('''
        INSERT OR IGNORE INTO shipping_carriers (slug, name, tracking_url)
        VALUES ('ems', 'EMS', 'https://t.17track.net/en#nums={tracking}')
    ''')
    conn.commit()

    # Default the AU scheduled carrier-tracking toggle ON. INSERT OR IGNORE means
    # any value the operator later sets in 系统设置 is preserved across reboots.
    try:
        conn.execute("INSERT OR IGNORE INTO settings (key, value) VALUES ('auto_track_au', '1')")
        conn.commit()
    except Exception:
        pass

    conn.close()


def init_product_tables():
    """Initialize product analysis tables (brands, series, product_mappings)"""
    conn = get_db_connection()
    
    # Brands table
    conn.execute('''
        CREATE TABLE IF NOT EXISTS brands (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL UNIQUE,
            aliases TEXT,
            created_at TEXT DEFAULT CURRENT_TIMESTAMP
        )
    ''')
    
    # Series table
    conn.execute('''
        CREATE TABLE IF NOT EXISTS series (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            brand_id INTEGER,
            name TEXT NOT NULL,
            aliases TEXT,
            UNIQUE(brand_id, name),
            FOREIGN KEY (brand_id) REFERENCES brands(id)
        )
    ''')
    
    # Product mappings table
    conn.execute('''
        CREATE TABLE IF NOT EXISTS product_mappings (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            raw_name TEXT NOT NULL,
            brand_id INTEGER,
            series_id INTEGER,
            puff_count INTEGER,
            flavor TEXT,
            is_manual INTEGER DEFAULT 0,
            source TEXT,
            created_at TEXT DEFAULT CURRENT_TIMESTAMP,
            UNIQUE(raw_name, source),
            FOREIGN KEY (brand_id) REFERENCES brands(id),
            FOREIGN KEY (series_id) REFERENCES series(id)
        )
    ''')
    
    conn.commit()
    
    # Insert default brands if table is empty
    existing = conn.execute('SELECT COUNT(*) FROM brands').fetchone()[0]
    if existing == 0:
        default_brands = [
            ('IGET', '["iget", "I-GET"]'),
            ('FUMO', '["Fumo"]'),
            ('Crystal Blind', '["CRYSTAL BLIND", "Crystal-Blind"]'),
            ('POD SALT', '["Pod Salt", "PODSALT"]'),
            ('Waka', '["WAKA"]'),
            ('UMIN', '["Umin"]'),
            ('Isgo', '["ISGO", "Isgo bar"]'),
            ('FIZZY', '["Fizzy", "FIZZY TWINS"]'),
            ('FUMOT', '["Fumot"]'),
            ('Esin', '["ESIN", "Esin Vape"]'),
            ('XCOR', '["Xcor", "X-COR"]'),
            ('ELF BAR', '["Elf Bar", "ELFBAR"]'),
            ('Lost Mary', '["LOST MARY", "LostMary"]'),
            ('Geek Bar', '["GEEK BAR", "GeekBar"]'),
        ]
        conn.executemany('INSERT OR IGNORE INTO brands (name, aliases) VALUES (?, ?)', default_brands)
        conn.commit()
    
    conn.close()


import re


def normalize_flavor(flavor):
    """
    Normalize flavor name for consistent aggregation.
    Handles variations like: "Love 66", "love-66", "LOVE_66" -> "LOVE 66"
    """
    if not flavor:
        return ''
    # Convert to uppercase
    s = flavor.upper()
    # Replace common separators (hyphens, underscores) with space
    s = re.sub(r'[-_]+', ' ', s)
    # Remove extra whitespace
    s = ' '.join(s.split())
    return s


def normalize_raw_name(name):
    """Canonical form of a product raw_name used as the key for product_mappings.

    WooCommerce stores names with HTML entities (e.g. `&#8222;`); some UI paths
    save the same product as already-decoded chars (e.g. `„`). Both should map
    to the same row, so we always store and lookup the *decoded* form.
    """
    if not name:
        return name
    try:
        return html.unescape(str(name))
    except Exception:
        return str(name)


def parse_product_name(name, brands_cache=None, series_cache=None):
    """
    Parse product name to extract brand, series, puff count, and flavor.
    
    Examples:
    - "IGET ONE 12000 puffs - Mixed Berries" → brand: IGET, puffs: 12000, flavor: Mixed Berries
    - "Crystal Blind 25000 Puffs" → brand: Crystal Blind, puffs: 25000
    - "FUMO king 6000 puffs Disposable Vape 20mg" → brand: FUMO, series: king, puffs: 6000
    """
    if not name:
        return {'brand': None, 'series': None, 'puffs': None, 'flavor': None, 'normalized': None}
    
    result = {'brand': None, 'series': None, 'puffs': None, 'flavor': None, 'normalized': None}
    
    # 1. Extract puff count
    # Handle optional "+" and Polish term "zaciągnięć"
    # Support spaced/dotted/comma'd thousands: "50 000", "50.000", "50,000"
    puffs_end = None  # end index of the puffs marker in `name`, used by flavor fallback below
    puff_match = re.search(r'\b(\d{1,3}(?:[\s.,]\d{3})+|\d+)\+?\s*(?:puffs?|zaciągnięć)', name, re.IGNORECASE)
    if puff_match:
        result['puffs'] = int(re.sub(r'[\s.,]', '', puff_match.group(1)))
        puffs_end = puff_match.end()
    else:
        # Fallback: Look for "Number Disposable" pattern e.g. "9000 Disposable"
        disposable_match = re.search(r'\b(\d{1,3}(?:[\s.,]\d{3})+|\d+)\s*Disposable', name, re.IGNORECASE)
        if disposable_match and int(re.sub(r'[\s.,]', '', disposable_match.group(1))) >= 100:
            result['puffs'] = int(re.sub(r'[\s.,]', '', disposable_match.group(1)))
            puffs_end = disposable_match.end()
    
    # 2. Get brands from cache or database
    if brands_cache is None:
        conn = get_db_connection()
        brands_rows = conn.execute('SELECT id, name, aliases FROM brands').fetchall()
        conn.close()
        brands_cache = []
        for row in brands_rows:
            brand_name = row['name']
            aliases = []
            if row['aliases']:
                try:
                    aliases = json.loads(row['aliases'])
                except:
                    pass
            brands_cache.append({
                'id': row['id'],
                'name': brand_name,
                'aliases': aliases,
                'patterns': [brand_name.upper()] + [a.upper() for a in aliases]
            })
    
    # 3. Match brand (earliest position first, then longest match)
    name_upper = name.upper()
    matched_brand = None
    matched_pos = len(name_upper)
    matched_len = 0

    for brand in brands_cache:
        for pattern in brand['patterns']:
            pos = name_upper.find(pattern)
            if pos >= 0 and (pos < matched_pos or (pos == matched_pos and len(pattern) > matched_len)):
                matched_brand = brand
                matched_pos = pos
                matched_len = len(pattern)
    
    if matched_brand:
        result['brand'] = matched_brand['name']
        result['brand_id'] = matched_brand['id']

    # 3b. Match series (if brand found and series_cache available)
    if matched_brand and series_cache is None:
        try:
            conn2 = get_db_connection()
            s_rows = conn2.execute('SELECT id, brand_id, name FROM series').fetchall()
            conn2.close()
            series_cache = [{'id': r['id'], 'brand_id': r['brand_id'], 'name': r['name']} for r in s_rows]
        except:
            series_cache = []
    if matched_brand and series_cache:
        matched_series = None
        matched_series_len = 0
        for s in series_cache:
            if s['brand_id'] == matched_brand['id']:
                if s['name'].upper() in name_upper and len(s['name']) > matched_series_len:
                    matched_series = s
                    matched_series_len = len(s['name'])
        if matched_series:
            result['series'] = matched_series['name']
            result['series_id'] = matched_series['id']

    # 4. Extract flavor (usually after separator)
    flavor = None
    for sep in [' - ', ' – ', ' | ', ' / ']:
        if sep in name:
            parts = name.split(sep)
            if len(parts) > 1:
                flavor = parts[-1].strip()
                # Don't treat product type/specs as flavor
                if any(x in flavor.lower() for x in ['puff', 'disposable', 'vape', 'mg', 'ml']):
                    flavor = None
                break

    # 4b. Fallback: if no separator-based flavor was found but we know where the
    # puffs marker ends, treat the remaining text as a flavor candidate. Handles
    # formats like "Merrymi Blade 30000 Puffs Aperol" where the flavor is just
    # appended without a separator.
    if not flavor and puffs_end is not None:
        tail = name[puffs_end:].strip()
        # Drop leading separators/punctuation (keep periods inside flavors like "Mr. Blue")
        tail = re.sub(r'^[\s\-–|/,:;]+', '', tail).strip()
        if tail:
            # Strip standalone product-spec words (English + Polish equivalents)
            cleaned = re.sub(
                r'\b(disposable|vape|jednorazowy|e-?papieros|puffs?)\b',
                ' ',
                tail,
                flags=re.IGNORECASE,
            )
            # Strip standalone unit values like "20mg", "5ml"
            cleaned = re.sub(r'\b\d+\s*(mg|ml)\b', ' ', cleaned, flags=re.IGNORECASE)
            # Collapse separator residue (but preserve periods)
            cleaned = re.sub(r'[\-–|/,:;]+', ' ', cleaned)
            cleaned = ' '.join(cleaned.split()).strip()
            if cleaned:
                flavor = cleaned

    result['flavor'] = flavor
    
    # 5. Build normalized name
    parts = []
    if result['brand']:
        parts.append(result['brand'])
    if result['puffs']:
        parts.append(f"{result['puffs']} Puffs")
    if result['flavor']:
        parts.append(result['flavor'])
    
    result['normalized'] = ' - '.join(parts) if parts else name
    
    return result


def init_sales_board_tables():
    """Initialize sales board related tables"""
    conn = get_db_connection()

    # Add can_view_sales_board column to users table
    try:
        conn.execute('ALTER TABLE users ADD COLUMN can_view_sales_board INTEGER DEFAULT 0')
        conn.commit()
    except:
        pass  # Column already exists

    # Add can_view_own_sales_board — self-only board (sees just the user's own
    # sites). Distinct from the full can_view_sales_board.
    try:
        conn.execute('ALTER TABLE users ADD COLUMN can_view_own_sales_board INTEGER DEFAULT 0')
        conn.commit()
    except:
        pass  # Column already exists

    # Add can_manage_users column — super admin privilege
    try:
        conn.execute('ALTER TABLE users ADD COLUMN can_manage_users INTEGER DEFAULT 0')
        conn.commit()
    except:
        pass  # Column already exists

    # Ensure the 'admin' username always has can_manage_users = 1
    try:
        conn.execute("UPDATE users SET can_manage_users = 1 WHERE username = 'admin'")
        conn.commit()
    except:
        pass

    # Add can_manage_products column — gates access to /product-manager. Site-level
    # scoping reuses user_site_permissions (the same table that filters orders).
    try:
        conn.execute('ALTER TABLE users ADD COLUMN can_manage_products INTEGER DEFAULT 0')
        conn.commit()
    except:
        pass  # Column already exists

    # Sales targets table - monthly targets per manager
    conn.execute('''
        CREATE TABLE IF NOT EXISTS sales_targets (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            year_month TEXT NOT NULL,
            manager TEXT NOT NULL,
            monthly_target REAL DEFAULT 0,
            weekly_targets TEXT DEFAULT '{}',
            base_salary REAL DEFAULT 7000,
            commission_rate REAL DEFAULT 0.05,
            notes TEXT DEFAULT '',
            created_at TEXT DEFAULT CURRENT_TIMESTAMP,
            updated_at TEXT DEFAULT CURRENT_TIMESTAMP,
            UNIQUE(year_month, manager)
        )
    ''')

    # No-commission brands table
    conn.execute('''
        CREATE TABLE IF NOT EXISTS no_commission_brands (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            brand_name TEXT NOT NULL UNIQUE,
            created_at TEXT DEFAULT CURRENT_TIMESTAMP
        )
    ''')

    # Insert default no-commission brands
    default_brands = ['L8', 'Esin', 'UR1']
    for brand in default_brands:
        try:
            conn.execute('INSERT OR IGNORE INTO no_commission_brands (brand_name) VALUES (?)', (brand,))
        except:
            pass

    conn.commit()
    conn.close()


def init_sales_groups_tables():
    """Initialize sales groups related tables"""
    conn = get_db_connection()

    # Sales groups table
    conn.execute('''
        CREATE TABLE IF NOT EXISTS sales_groups (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL,
            leader_manager TEXT NOT NULL,
            bonus_rate REAL DEFAULT 0.02,
            created_at TEXT DEFAULT CURRENT_TIMESTAMP,
            updated_at TEXT DEFAULT CURRENT_TIMESTAMP
        )
    ''')

    # Sales group members table
    conn.execute('''
        CREATE TABLE IF NOT EXISTS sales_group_members (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            group_id INTEGER NOT NULL,
            manager TEXT NOT NULL,
            created_at TEXT DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (group_id) REFERENCES sales_groups(id) ON DELETE CASCADE,
            UNIQUE(group_id, manager)
        )
    ''')

    # Sales-board-only exchange rate overrides (does not affect rest of the system)
    conn.execute('''
        CREATE TABLE IF NOT EXISTS sales_board_exchange_rates (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            year_month TEXT NOT NULL,
            currency TEXT NOT NULL,
            rate_to_cny REAL NOT NULL,
            updated_at TEXT DEFAULT CURRENT_TIMESTAMP,
            updated_by TEXT DEFAULT '',
            UNIQUE(year_month, currency)
        )
    ''')

    # Sales board export history
    conn.execute('''
        CREATE TABLE IF NOT EXISTS sales_board_exports (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            year_month TEXT NOT NULL,
            filename TEXT NOT NULL,
            file_path TEXT NOT NULL,
            file_size INTEGER DEFAULT 0,
            hide_leader INTEGER DEFAULT 0,
            created_by TEXT DEFAULT '',
            created_at TEXT DEFAULT CURRENT_TIMESTAMP
        )
    ''')

    conn.commit()
    conn.close()


def init_partner_reconciliation_tables():
    """Initialize partner reconciliation related tables"""
    conn = get_db_connection()

    # Partners table
    conn.execute('''
        CREATE TABLE IF NOT EXISTS partners (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL,
            code TEXT UNIQUE,
            description TEXT,
            cost_ratio REAL DEFAULT 0.50,
            partner_profit_ratio REAL DEFAULT 0.25,
            our_profit_ratio REAL DEFAULT 0.25,
            currency TEXT DEFAULT 'PLN',
            created_at TEXT DEFAULT CURRENT_TIMESTAMP,
            updated_at TEXT DEFAULT CURRENT_TIMESTAMP
        )
    ''')

    # Partner-site bindings
    conn.execute('''
        CREATE TABLE IF NOT EXISTS partner_sites (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            partner_id INTEGER NOT NULL,
            site_id INTEGER NOT NULL,
            created_at TEXT DEFAULT CURRENT_TIMESTAMP,
            UNIQUE(partner_id, site_id),
            FOREIGN KEY (partner_id) REFERENCES partners(id) ON DELETE CASCADE
        )
    ''')

    # Partner-user bindings (which users can see which partner's data)
    conn.execute('''
        CREATE TABLE IF NOT EXISTS partner_users (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            partner_id INTEGER NOT NULL,
            user_id INTEGER NOT NULL,
            created_at TEXT DEFAULT CURRENT_TIMESTAMP,
            UNIQUE(partner_id, user_id),
            FOREIGN KEY (partner_id) REFERENCES partners(id) ON DELETE CASCADE
        )
    ''')

    # Reconciliation statements (monthly)
    conn.execute('''
        CREATE TABLE IF NOT EXISTS reconciliation_statements (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            partner_id INTEGER NOT NULL,
            period_year INTEGER NOT NULL,
            period_month INTEGER NOT NULL,
            total_orders INTEGER DEFAULT 0,
            total_gross_pln REAL DEFAULT 0,
            total_net_pln REAL DEFAULT 0,
            cost_amount_pln REAL DEFAULT 0,
            partner_profit_pln REAL DEFAULT 0,
            our_receivable_pln REAL DEFAULT 0,
            exchange_rate_cny REAL,
            our_receivable_cny REAL,
            status TEXT DEFAULT 'draft',
            is_manual INTEGER DEFAULT 0,
            confirmed_at TEXT,
            confirmed_by INTEGER,
            locked_at TEXT,
            notes TEXT,
            created_at TEXT DEFAULT CURRENT_TIMESTAMP,
            updated_at TEXT DEFAULT CURRENT_TIMESTAMP,
            UNIQUE(partner_id, period_year, period_month),
            FOREIGN KEY (partner_id) REFERENCES partners(id) ON DELETE CASCADE
        )
    ''')

    # Partner receipts (money received from partners)
    conn.execute('''
        CREATE TABLE IF NOT EXISTS partner_receipts (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            partner_id INTEGER NOT NULL,
            statement_id INTEGER,
            receipt_date TEXT NOT NULL,
            amount_pln REAL NOT NULL,
            exchange_rate_cny REAL,
            amount_cny REAL,
            payment_method TEXT,
            reference_no TEXT,
            receipt_url TEXT,
            notes TEXT,
            created_by INTEGER,
            created_at TEXT DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (partner_id) REFERENCES partners(id) ON DELETE CASCADE,
            FOREIGN KEY (statement_id) REFERENCES reconciliation_statements(id) ON DELETE SET NULL
        )
    ''')

    # Add can_view_reconciliation column to users table
    try:
        conn.execute('ALTER TABLE users ADD COLUMN can_view_reconciliation INTEGER DEFAULT 0')
    except sqlite3.OperationalError:
        pass  # column already exists

    # Add can_edit_reconciliation column (write access — separate from view)
    try:
        conn.execute('ALTER TABLE users ADD COLUMN can_edit_reconciliation INTEGER DEFAULT 0')
        # Migration: users who currently have view access and are NOT partner members
        # were effectively "internal editors" — preserve their edit ability.
        conn.execute('''
            UPDATE users SET can_edit_reconciliation = 1
            WHERE can_view_reconciliation = 1
              AND id NOT IN (SELECT DISTINCT user_id FROM partner_users)
        ''')
    except sqlite3.OperationalError:
        pass  # column already exists

    # Add can_view_costs column — gates access to /product-costs (cost management).
    # Default 0; super admin (username='admin') always passes regardless of flag.
    try:
        conn.execute('ALTER TABLE users ADD COLUMN can_view_costs INTEGER DEFAULT 0')
    except sqlite3.OperationalError:
        pass  # column already exists

    # Add can_edit_costs column — gates write access to product_costs (POST/PUT/DELETE).
    # Separated from can_view_costs so we can have read-only viewers (e.g. partners
    # who need to verify costs) AND distinct editors. Super admin always passes.
    try:
        conn.execute('ALTER TABLE users ADD COLUMN can_edit_costs INTEGER DEFAULT 0')
        # Migration: previously, only admin-role users could write costs (the old
        # @admin_required decorator). Preserve that behavior — auto-grant
        # can_edit_costs to existing admin-role users so nothing breaks.
        conn.execute('UPDATE users SET can_edit_costs = 1 WHERE role = ?', ('admin',))
    except sqlite3.OperationalError:
        pass  # column already exists

    # Add can_manage_blocklist column — gates who may add/remove customer
    # blocklist entries (auto-cancel COD). Admin role passes implicitly; granted
    # per-user to others in 用户管理 (super admin only). Default 0.
    try:
        conn.execute('ALTER TABLE users ADD COLUMN can_manage_blocklist INTEGER DEFAULT 0')
        conn.execute('UPDATE users SET can_manage_blocklist = 1 WHERE role = ?', ('admin',))
    except sqlite3.OperationalError:
        pass  # column already exists

    # ---------- P2: AUDIT LOG ----------
    # Every state change / edit / dispute / confirm / receipt-attach lands here.
    # Append-only; never delete rows. Powers the timeline shown in the
    # statement detail modal so partners and admins can see what changed,
    # when, by whom, and (for disputes) why.
    conn.execute('''
        CREATE TABLE IF NOT EXISTS reconciliation_audit_log (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            statement_id INTEGER NOT NULL,
            action TEXT NOT NULL,
            actor_id INTEGER,
            actor_username TEXT,
            actor_role TEXT,
            field TEXT,
            old_value TEXT,
            new_value TEXT,
            note TEXT,
            ip TEXT,
            created_at TEXT DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (statement_id) REFERENCES reconciliation_statements(id) ON DELETE CASCADE
        )
    ''')
    conn.execute('CREATE INDEX IF NOT EXISTS idx_audit_stmt ON reconciliation_audit_log(statement_id, created_at DESC)')

    # ---------- P2: ORDER SNAPSHOT ----------
    # When a statement is generated, we freeze the list of order IDs that
    # contributed. If the underlying order is later edited (status / total /
    # is_undelivered), the saved statement totals stay correct AND we can
    # show a diff "this order was X at generation, is Y now".
    conn.execute('''
        CREATE TABLE IF NOT EXISTS reconciliation_statement_orders (
            statement_id INTEGER NOT NULL,
            order_id TEXT NOT NULL,
            order_number TEXT,
            status_at_gen TEXT,
            total_at_gen REAL,
            shipping_at_gen REAL,
            shipping_loss_at_gen REAL,
            is_undelivered_at_gen INTEGER,
            currency_at_gen TEXT,
            date_created TEXT,
            PRIMARY KEY (statement_id, order_id),
            FOREIGN KEY (statement_id) REFERENCES reconciliation_statements(id) ON DELETE CASCADE
        )
    ''')
    conn.execute('CREATE INDEX IF NOT EXISTS idx_stmt_orders_stmt ON reconciliation_statement_orders(statement_id)')

    # ---------- P2: extra columns on reconciliation_statements ----------
    # confirmed_name: typed by partner when confirming (e-signature)
    # confirmed_ip:   recorded for audit
    # actual_cost_pln_snapshot: real product-cost number frozen at gen time
    for col_def in [
        ('confirmed_name', 'TEXT'),
        ('confirmed_ip', 'TEXT'),
        ('actual_cost_pln_snapshot', 'REAL'),
        # Settlement basis this statement was generated on:
        #   'contract' = 约定毛利 (净销售 × 固定比例)
        #   'actual'   = 实际毛利 (合伙人先收回真实进价，剩余毛利按 pp:op 分)
        # Default 'contract' so all existing statements keep their meaning.
        ('calc_mode', "TEXT DEFAULT 'contract'"),
    ]:
        try:
            conn.execute(f"ALTER TABLE reconciliation_statements ADD COLUMN {col_def[0]} {col_def[1]}")
        except sqlite3.OperationalError:
            pass  # column already exists

    # Seed default partner: 金谷金毅（波兰）
    existing = conn.execute("SELECT id FROM partners WHERE code = 'poland_jin'").fetchone()
    if not existing:
        cursor = conn.execute('''
            INSERT INTO partners (name, code, description, cost_ratio, partner_profit_ratio, our_profit_ratio, currency)
            VALUES (?, ?, ?, ?, ?, ?, ?)
        ''', ('金谷金毅（波兰）', 'poland_jin', '波兰线下店铺合伙人，负责囤货、发货、售后，COD收款', 0.50, 0.25, 0.25, 'PLN'))
        partner_id = cursor.lastrowid
        # Auto-bind all PL sites
        pl_sites = conn.execute("SELECT id FROM sites WHERE country = 'PL'").fetchall()
        for site in pl_sites:
            conn.execute('INSERT OR IGNORE INTO partner_sites (partner_id, site_id) VALUES (?, ?)',
                        (partner_id, site['id']))

    conn.commit()
    conn.close()


def init_product_costs_tables():
    """Initialize product costs and profit settings tables"""
    conn = get_db_connection()

    conn.execute('''
        CREATE TABLE IF NOT EXISTS product_costs (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            brand_id INTEGER NOT NULL,
            series_id INTEGER,
            puff_count INTEGER,
            flavor TEXT,
            country TEXT NOT NULL DEFAULT 'PL',
            cost_price REAL NOT NULL,
            cost_currency TEXT NOT NULL DEFAULT 'PLN',
            effective_date TEXT NOT NULL DEFAULT '2024-01-01',
            notes TEXT,
            created_at TEXT DEFAULT CURRENT_TIMESTAMP,
            updated_at TEXT DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (brand_id) REFERENCES brands(id),
            FOREIGN KEY (series_id) REFERENCES series(id)
        )
    ''')

    conn.execute('''
        CREATE TABLE IF NOT EXISTS sales_board_profit_settings (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            year_month TEXT NOT NULL UNIQUE,
            profit_mode TEXT NOT NULL DEFAULT 'percentage',
            profit_percentage REAL NOT NULL DEFAULT 50.0,
            updated_at TEXT DEFAULT CURRENT_TIMESTAMP
        )
    ''')

    conn.execute('''
        CREATE UNIQUE INDEX IF NOT EXISTS idx_product_costs_match
        ON product_costs (brand_id, COALESCE(series_id, 0), COALESCE(puff_count, 0), COALESCE(flavor, ''), country)
    ''')

    conn.commit()
    conn.close()


def init_warehouses():
    """Initialize warehouses table, add warehouse_id to product_costs and orders, migrate data."""
    conn = get_db_connection()

    conn.execute('''
        CREATE TABLE IF NOT EXISTS warehouses (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL,
            code TEXT NOT NULL UNIQUE,
            country TEXT NOT NULL,
            default_currency TEXT NOT NULL DEFAULT 'PLN',
            is_active INTEGER DEFAULT 1,
            notes TEXT,
            created_at TEXT DEFAULT CURRENT_TIMESTAMP,
            updated_at TEXT DEFAULT CURRENT_TIMESTAMP
        )
    ''')

    # Seed default warehouses from existing site countries (only if table is empty)
    existing = conn.execute('SELECT COUNT(*) FROM warehouses').fetchone()[0]
    if existing == 0:
        country_defaults = {
            'PL': ('波兰仓库', 'PLN'),
            'AU': ('澳洲仓库', 'AUD'),
            'AE': ('阿联酋仓库', 'AED'),
        }
        countries = conn.execute("SELECT DISTINCT country FROM sites WHERE country IS NOT NULL AND country != ''").fetchall()
        for row in countries:
            c = row['country']
            name, cur = country_defaults.get(c, (c + ' 仓库', 'PLN'))
            try:
                conn.execute('INSERT INTO warehouses (name, code, country, default_currency) VALUES (?, ?, ?, ?)',
                             (name, c, c, cur))
            except:
                pass

    # Add warehouse_id to product_costs
    try:
        conn.execute('ALTER TABLE product_costs ADD COLUMN warehouse_id INTEGER')
    except:
        pass
    # Backfill product_costs.warehouse_id from country
    conn.execute('''
        UPDATE product_costs SET warehouse_id = (
            SELECT w.id FROM warehouses w WHERE w.code = product_costs.country
        ) WHERE warehouse_id IS NULL
    ''')

    # Rebuild index to include effective_date — same product can have multiple
    # cost prices over time (e.g. Feb 28元, Mar 30元). Must keep ALL versions
    # so historical orders use the cost effective on their order date.
    # Unique constraint now requires effective_date to be different.
    conn.execute('DROP INDEX IF EXISTS idx_product_costs_match')
    conn.execute('''
        CREATE UNIQUE INDEX IF NOT EXISTS idx_product_costs_match
        ON product_costs (brand_id, COALESCE(series_id, 0), COALESCE(puff_count, 0), COALESCE(flavor, ''), COALESCE(warehouse_id, 0), effective_date)
    ''')
    # Lookup index: scan-by-product + effective_date DESC for "cost at date" queries
    conn.execute('''
        CREATE INDEX IF NOT EXISTS idx_product_costs_lookup
        ON product_costs (brand_id, warehouse_id, effective_date DESC)
    ''')

    # Add warehouse_id to orders
    try:
        conn.execute('ALTER TABLE orders ADD COLUMN warehouse_id INTEGER')
    except:
        pass
    # Backfill orders.warehouse_id from site country -> first warehouse in that country.
    # Earlier code joined on `w.code` (which holds the Chinese warehouse name) instead
    # of `w.country`, so the backfill never matched and ~38% of orders ended up with
    # NULL warehouse_id — which downstream cost lookups treat as "unmapped".
    conn.execute('''
        UPDATE orders SET warehouse_id = (
            SELECT MIN(w.id) FROM warehouses w
            JOIN sites s ON s.country = w.country
            WHERE s.url = orders.source
        ) WHERE warehouse_id IS NULL
    ''')
    try:
        conn.execute('CREATE INDEX IF NOT EXISTS idx_orders_warehouse_id ON orders(warehouse_id)')
    except:
        pass

    # Add country_percentages column to profit settings if missing
    try:
        conn.execute('ALTER TABLE sales_board_profit_settings ADD COLUMN country_percentages TEXT DEFAULT "{}"')
    except:
        pass

    conn.commit()
    conn.close()


def init_blocklist_tables():
    """Customer blocklist — auto-cancel COD orders from blacklisted phones.
    Keyed by normalized phone (last 9 digits): emails rotate, the phone is the
    stable identity. Enforcement lives in blocklist.py (run by the hourly cron).
    """
    conn = get_db_connection()
    conn.execute('''
        CREATE TABLE IF NOT EXISTS blocked_customers (
            phone TEXT PRIMARY KEY,
            raw_phone TEXT,
            name TEXT,
            email TEXT,
            reason TEXT,
            auto_cancel INTEGER DEFAULT 1,
            created_by TEXT,
            created_at TEXT DEFAULT CURRENT_TIMESTAMP,
            cancelled_count INTEGER DEFAULT 0,
            last_cancelled_at TEXT
        )
    ''')
    conn.execute('''
        CREATE TABLE IF NOT EXISTS blocked_cancel_log (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            order_id TEXT,
            order_number TEXT,
            phone TEXT,
            source TEXT,
            total REAL,
            currency TEXT,
            old_status TEXT,
            result TEXT,
            detail TEXT,
            actor TEXT,
            created_at TEXT DEFAULT CURRENT_TIMESTAMP
        )
    ''')
    # Global kill-switch. Default ON is safe: enforcement is a no-op until a phone
    # is actually added to blocked_customers.
    conn.execute("INSERT OR IGNORE INTO settings (key, value) VALUES ('blocklist_auto_cancel_enabled', '1')")
    # NOTE: who may *manage* the blocklist is a per-user permission
    # (users.can_manage_blocklist, granted in 用户管理), not a global setting.
    conn.commit()
    conn.close()


# Initialize tables on startup
with app.app_context():
    init_sites_table()
    init_product_masters_table()
    init_sync_logs_table()
    init_settings_table()
    init_users_table()
    init_shipping_tables()
    init_undelivered_columns()
    init_problem_return_columns()
    init_product_tables()
    init_user_preferences_table()
    init_sales_board_tables()
    init_sales_groups_tables()
    init_partner_reconciliation_tables()
    init_product_costs_tables()
    init_warehouses()
    init_blocklist_tables()

@app.route('/settings')
@login_required
@admin_required
def settings():
    """Settings page for site management - Admin only"""
    conn = get_db_connection()
    sites = conn.execute('SELECT * FROM sites').fetchall()

    # Get exchange rates
    exchange_rates = conn.execute('''
        SELECT id, year_month, currency, rate_to_cny, updated_at
        FROM exchange_rates
        ORDER BY year_month DESC, currency
    ''').fetchall()

    # Get distinct currencies from orders
    currencies = conn.execute('SELECT DISTINCT currency FROM orders WHERE currency IS NOT NULL').fetchall()
    currency_list = [c['currency'] for c in currencies if c['currency']]

    # Get product masters (for the multistore dropdown + management section)
    pm_rows = conn.execute('''
        SELECT id, label, url, consumer_key, api_status, last_api_error,
               last_tested_at, created_at, updated_at
        FROM product_masters ORDER BY label
    ''').fetchall()
    # Build a quick id→label/url map for chip display on each site row
    product_masters_list = []
    masters_lookup = {}
    for r in pm_rows:
        ref_count = conn.execute(
            'SELECT COUNT(*) FROM sites WHERE product_master_id = ?', (r['id'],)
        ).fetchone()[0]
        d = dict(r)
        d['ref_count'] = ref_count
        d['consumer_key_short'] = (d['consumer_key'] or '')[:10] + '...' if d['consumer_key'] else ''
        product_masters_list.append(d)
        masters_lookup[r['id']] = {
            'label': r['label'],
            'url': r['url'],
            'host': r['url'].replace('https://www.', '').replace('https://', '').replace('http://', '')
        }

    conn.close()
    return render_template('settings.html',
                          sites=sites,
                          exchange_rates=exchange_rates,
                          currencies=currency_list,
                          product_masters=product_masters_list,
                          masters_lookup=masters_lookup)


@app.route('/api/exchange-rates', methods=['GET', 'POST'])
@login_required
@admin_required
def exchange_rates_api():
    """API for managing exchange rates"""
    conn = get_db_connection()
    
    if request.method == 'GET':
        rates = conn.execute('''
            SELECT id, year_month, currency, rate_to_cny, updated_at
            FROM exchange_rates
            ORDER BY year_month DESC, currency
        ''').fetchall()
        conn.close()
        return jsonify([dict(r) for r in rates])
    
    elif request.method == 'POST':
        data = request.get_json()
        year_month = data.get('year_month', '')
        currency = data.get('currency', '').upper()
        rate_to_cny = data.get('rate_to_cny')
        
        if not year_month or not currency or rate_to_cny is None:
            conn.close()
            return jsonify({'error': '请填写所有字段'}), 400
        
        try:
            rate_to_cny = float(rate_to_cny)
            if rate_to_cny <= 0:
                raise ValueError("Rate must be positive")
        except (ValueError, TypeError):
            conn.close()
            return jsonify({'error': '汇率必须是正数'}), 400
        
        try:
            conn.execute('''
                INSERT OR REPLACE INTO exchange_rates (year_month, currency, rate_to_cny, updated_at)
                VALUES (?, ?, ?, datetime('now'))
            ''', (year_month, currency, rate_to_cny))
            conn.commit()
            conn.close()
            return jsonify({'success': True, 'message': '汇率保存成功'})
        except Exception as e:
            conn.close()
            return jsonify({'error': str(e)}), 500


@app.route('/api/exchange-rates/<int:rate_id>', methods=['DELETE'])
@login_required
@admin_required
def delete_exchange_rate(rate_id):
    """Delete an exchange rate"""
    conn = get_db_connection()
    conn.execute('DELETE FROM exchange_rates WHERE id = ?', (rate_id,))
    conn.commit()
    conn.close()
    return jsonify({'success': True, 'message': '汇率已删除'})


@app.route('/api/sites/import-from-script', methods=['POST'])
@login_required
@admin_required
def import_sites_from_script():
    """从 1.wooorders_sqlite.py 的硬编码配置导入站点到数据库（管理员专用：同样会批量建站）"""
    import ast
    import re
    
    script_path = '1.wooorders_sqlite.py'
    
    try:
        with open(script_path, 'r', encoding='utf-8') as f:
            content = f.read()
        
        # 提取 HARDCODED_SITES 列表
        pattern = r'HARDCODED_SITES\s*=\s*(\[[\s\S]*?\n\])'
        match = re.search(pattern, content)
        
        if not match:
            # 尝试旧格式 sites = [...]
            pattern = r'^sites\s*=\s*(\[[\s\S]*?\n\])'
            match = re.search(pattern, content, re.MULTILINE)
        
        if not match:
            return jsonify({'error': '无法在脚本中找到站点配置'}), 400
        
        # 解析 Python 列表
        sites_str = match.group(1)
        sites_list = ast.literal_eval(sites_str)
        
        conn = get_db_connection()
        imported = 0
        skipped = 0
        
        for site in sites_list:
            url = site.get('url', '')
            ck = site.get('ck', '')
            cs = site.get('cs', '')
            
            if not all([url, ck, cs]):
                continue
            
            # 检查是否已存在
            existing = conn.execute('SELECT id FROM sites WHERE url = ?', (url,)).fetchone()
            if existing:
                skipped += 1
                continue
            
            conn.execute('INSERT INTO sites (url, consumer_key, consumer_secret) VALUES (?, ?, ?)',
                        (url, ck, cs))
            imported += 1
        
        conn.commit()
        conn.close()
        
        return jsonify({
            'success': True, 
            'imported': imported, 
            'skipped': skipped,
            'message': f'成功导入 {imported} 个站点，跳过 {skipped} 个已存在的站点'
        })
        
    except FileNotFoundError:
        return jsonify({'error': f'找不到脚本文件: {script_path}'}), 404
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# ============================================================================
# Product Masters (WooMultistore master sites for product CRUD)
# ============================================================================

def get_product_api_endpoint(conn, site_row):
    """Resolve the (url, consumer_key, consumer_secret) tuple to use for product
    CRUD on a given site row. For WooMultistore-managed sites this routes to
    the configured master; standalone sites route to themselves.

    Always pass a fresh `conn` — closes only the caller's connection."""
    master_id = site_row['product_master_id'] if 'product_master_id' in site_row.keys() else None
    if master_id:
        master = conn.execute(
            'SELECT url, consumer_key, consumer_secret FROM product_masters WHERE id = ?',
            (master_id,)
        ).fetchone()
        if master:
            return (master['url'], master['consumer_key'], master['consumer_secret'])
    return (site_row['url'], site_row['consumer_key'], site_row['consumer_secret'])


@app.route('/api/product-masters', methods=['GET'])
@login_required
@admin_required
def list_product_masters():
    """List all configured product masters (for the dropdown + management UI)."""
    conn = get_db_connection()
    rows = conn.execute('''
        SELECT id, label, url, consumer_key,
               api_status, last_api_error, last_tested_at, created_at, updated_at
        FROM product_masters ORDER BY label
    ''').fetchall()
    # Annotate with reference count (how many sites use this master)
    result = []
    for r in rows:
        ref_count = conn.execute(
            'SELECT COUNT(*) FROM sites WHERE product_master_id = ?',
            (r['id'],)
        ).fetchone()[0]
        d = dict(r)
        d['ref_count'] = ref_count
        # Mask the key in list output (only show first 10 chars)
        d['consumer_key'] = (d['consumer_key'] or '')[:10] + '...' if d['consumer_key'] else ''
        result.append(d)
    conn.close()
    return jsonify(result)


@app.route('/api/product-masters', methods=['POST'])
@login_required
@admin_required
def create_product_master():
    """Create a new product master."""
    data = request.get_json(silent=True) or {}
    label = (data.get('label') or '').strip()
    url = (data.get('url') or '').strip().rstrip('/')
    ck = (data.get('consumer_key') or '').strip()
    cs = (data.get('consumer_secret') or '').strip()

    if not all([label, url, ck, cs]):
        return jsonify({'error': '请填写完整：别名、URL、Consumer Key、Consumer Secret'}), 400
    if not url.startswith(('http://', 'https://')):
        return jsonify({'error': 'URL 必须以 http:// 或 https:// 开头'}), 400

    conn = get_db_connection()
    try:
        cur = conn.execute('''
            INSERT INTO product_masters (label, url, consumer_key, consumer_secret)
            VALUES (?, ?, ?, ?)
        ''', (label, url, ck, cs))
        conn.commit()
        new_id = cur.lastrowid
    except Exception as e:
        conn.close()
        return jsonify({'error': f'创建失败: {e}'}), 500
    conn.close()
    return jsonify({'success': True, 'id': new_id, 'label': label, 'url': url})


@app.route('/api/product-masters/<int:master_id>', methods=['PUT'])
@login_required
@admin_required
def update_product_master(master_id):
    """Update a product master. consumer_secret is optional — empty means keep current."""
    data = request.get_json(silent=True) or {}
    label = (data.get('label') or '').strip()
    url = (data.get('url') or '').strip().rstrip('/')
    ck = (data.get('consumer_key') or '').strip()
    cs = (data.get('consumer_secret') or '').strip()  # empty = unchanged

    if not all([label, url, ck]):
        return jsonify({'error': '请填写：别名、URL、Consumer Key'}), 400
    if not url.startswith(('http://', 'https://')):
        return jsonify({'error': 'URL 必须以 http:// 或 https:// 开头'}), 400

    conn = get_db_connection()
    existing = conn.execute('SELECT id FROM product_masters WHERE id = ?', (master_id,)).fetchone()
    if not existing:
        conn.close()
        return jsonify({'error': 'Master 站点不存在'}), 404
    try:
        if cs:
            conn.execute('''
                UPDATE product_masters
                SET label = ?, url = ?, consumer_key = ?, consumer_secret = ?,
                    updated_at = datetime('now')
                WHERE id = ?
            ''', (label, url, ck, cs, master_id))
        else:
            conn.execute('''
                UPDATE product_masters
                SET label = ?, url = ?, consumer_key = ?,
                    updated_at = datetime('now')
                WHERE id = ?
            ''', (label, url, ck, master_id))
        conn.commit()
    except Exception as e:
        conn.close()
        return jsonify({'error': f'更新失败: {e}'}), 500
    conn.close()
    return jsonify({'success': True})


@app.route('/api/product-masters/<int:master_id>', methods=['DELETE'])
@login_required
@admin_required
def delete_product_master(master_id):
    """Delete a product master — only if no site references it."""
    conn = get_db_connection()
    ref_count = conn.execute(
        'SELECT COUNT(*) FROM sites WHERE product_master_id = ?',
        (master_id,)
    ).fetchone()[0]
    if ref_count > 0:
        ref_sites = conn.execute(
            'SELECT url FROM sites WHERE product_master_id = ? LIMIT 5',
            (master_id,)
        ).fetchall()
        site_list = '、'.join(s['url'].replace('https://www.', '').replace('https://', '') for s in ref_sites)
        conn.close()
        return jsonify({
            'error': f'无法删除：还有 {ref_count} 个站点引用此 master（{site_list}）。请先在那些站点的设置里取消勾选。'
        }), 409
    try:
        conn.execute('DELETE FROM product_masters WHERE id = ?', (master_id,))
        conn.commit()
    except Exception as e:
        conn.close()
        return jsonify({'error': f'删除失败: {e}'}), 500
    conn.close()
    return jsonify({'success': True})


@app.route('/api/product-masters/<int:master_id>/test', methods=['POST'])
@login_required
@admin_required
def test_product_master(master_id):
    """Test the master's WC REST API connection (read + write capability check)."""
    import requests as req
    conn = get_db_connection()
    m = conn.execute(
        'SELECT id, url, consumer_key, consumer_secret FROM product_masters WHERE id = ?',
        (master_id,)
    ).fetchone()
    if not m:
        conn.close()
        return jsonify({'error': 'Master 站点不存在'}), 404

    api_url = f"{m['url']}/wp-json/wc/v3/products?per_page=1"
    try:
        resp = req.get(api_url, auth=(m['consumer_key'], m['consumer_secret']),
                       timeout=15,
                       headers={'User-Agent': 'WooCommerce API Client-Python/3.0.0',
                                'Accept': 'application/json'})
    except req.RequestException as e:
        conn.execute('''
            UPDATE product_masters
            SET api_status = 'error', last_api_error = ?, last_tested_at = datetime('now')
            WHERE id = ?
        ''', (f'网络错误: {e}', master_id))
        conn.commit()
        conn.close()
        return jsonify({'status': 'error', 'message': f'连接失败: {e}'}), 200

    body = resp.text or ''
    if resp.status_code != 200:
        msg = f'HTTP {resp.status_code}'
        try:
            j = resp.json()
            msg = f"{msg}: {j.get('message', body[:200])}"
        except Exception:
            msg = f'{msg}: {body[:200]}'
        conn.execute('''
            UPDATE product_masters
            SET api_status = 'error', last_api_error = ?, last_tested_at = datetime('now')
            WHERE id = ?
        ''', (msg, master_id))
        conn.commit()
        conn.close()
        return jsonify({'status': 'error', 'message': msg}), 200

    # 200 OK — read works. Note: this doesn't verify write perm; we'll surface
    # write errors at first edit. Keeping the test light avoids creating junk products.
    conn.execute('''
        UPDATE product_masters
        SET api_status = 'ok', last_api_error = NULL, last_tested_at = datetime('now')
        WHERE id = ?
    ''', (master_id,))
    conn.commit()
    conn.close()
    return jsonify({'status': 'ok', 'message': '连接成功（已验证读权限；写权限将在首次编辑时验证）'}), 200


# ============================================================================
# Product Manager — Layer 1: stock + price batch editing across sites
# ============================================================================

# Whitelist of fields editable through this UI. WC product schema has many more,
# but we deliberately restrict to avoid accidental clobbering of unrelated fields.
PRODUCT_EDIT_WHITELIST = {
    'manage_stock', 'stock_quantity', 'stock_status',
    'regular_price', 'sale_price',
}


def _resolve_site_for_product_edit(conn, site_id):
    """Return (site_row, api_url, ck, cs) or raise ValueError with user-friendly message.
    Routes multistore-managed sites to their master automatically.

    Permission rule: only the site's named manager (sites.manager == users.name)
    can edit its products. Super admin bypasses. This is stricter than
    user_site_permissions (used elsewhere for read access) — viewers and other
    users cannot edit even if they have read permission for the site."""
    site = conn.execute(
        'SELECT id, url, consumer_key, consumer_secret, product_master_id, manager FROM sites WHERE id = ?',
        (site_id,)
    ).fetchone()
    if not site:
        raise ValueError(f'站点 {site_id} 不存在')

    # Site-level permission check: name must match site manager (admin bypasses)
    if not current_user.is_admin():
        user_name = (current_user.name or '').strip()
        site_manager = (site['manager'] or '').strip()
        if not user_name or not site_manager or user_name != site_manager:
            raise ValueError(
                f'你没有该站点（{site["url"]}）的产品管理权限，'
                f'仅站点负责人可在此管理产品。'
            )

    api_url, ck, cs = get_product_api_endpoint(conn, site)
    if not (api_url and ck and cs):
        raise ValueError('站点未配置完整的 WC REST API 凭据')
    return site, api_url, ck, cs


def _parse_wc_response(resp):
    """Parse a WC REST API response defensively. Returns (data, error_message).

    Handles all the ways the call can go wrong without exceptions:
      - non-2xx status: extracts WC's error message (or HTML preview if not JSON)
      - 200 OK with HTML body: surfaces a hint about CF/cache/PHP-error
      - empty / malformed JSON: returns None data with explanation

    Either `data` is the parsed JSON (could be {} or [] for empty), or
    `error_message` is a Chinese string ready to put in the API response.
    Both will not be set; one will be None."""
    body_preview = ((resp.text or '')[:300]).replace('\n', ' ')
    looks_like_html = body_preview.lstrip().lower().startswith(('<!doctype', '<html', '<?xml'))

    if resp.status_code not in (200, 201):
        try:
            j = resp.json()
            msg = j.get('message') or body_preview or '(无响应内容)'
        except Exception:
            msg = body_preview or '(无响应内容)'
        return None, f'WC API 错误 HTTP {resp.status_code}: {msg}'

    try:
        return resp.json(), None
    except Exception:
        if looks_like_html:
            hint = (
                f'WC API 返回了 HTML 而不是 JSON（HTTP {resp.status_code}）。'
                f'常见原因：CloudFlare 拦截、缓存插件命中、PHP 致命错误页。'
                f'响应开头: {body_preview}'
            )
        else:
            hint = f'WC API 响应解析失败（HTTP {resp.status_code}）：{body_preview}'
        return None, hint


@app.route('/api/product-manager/products')
@login_required
@product_manager_required
def product_manager_list():
    """List products from a single site via WC REST API (live, no caching).
    For multistore-managed sites, fetches from the configured master.

    Query params:
      site_id    (required) — id of a site row in `sites` table
      search     (optional) — full-text search (WC's `?search=`)
      status     (optional) — publish/draft/private/any (default: any)
      page       (optional) — pagination, default 1
      per_page   (optional) — default 50, max 100
    """
    import requests as req

    try:
        site_id = int(request.args.get('site_id', '0'))
    except (TypeError, ValueError):
        return jsonify({'error': 'site_id 必须是整数'}), 400
    if not site_id:
        return jsonify({'error': '请指定 site_id'}), 400

    search = request.args.get('search', '').strip()
    status = request.args.get('status', 'any').strip() or 'any'
    try:
        page = max(1, int(request.args.get('page', '1')))
        per_page = min(100, max(1, int(request.args.get('per_page', '50'))))
    except (TypeError, ValueError):
        page = 1
        per_page = 50

    conn = get_db_connection()
    try:
        site, api_url, ck, cs = _resolve_site_for_product_edit(conn, site_id)
    except ValueError as e:
        conn.close()
        return jsonify({'error': str(e)}), 400

    # Surface the routing decision for the UI ("写到 master.vapego.pl" hint)
    routing_info = {
        'site_url': site['url'],
        'effective_url': api_url,
        'is_routed_to_master': bool(site['product_master_id']),
    }
    if site['product_master_id']:
        m = conn.execute(
            'SELECT label FROM product_masters WHERE id = ?',
            (site['product_master_id'],)
        ).fetchone()
        routing_info['master_label'] = m['label'] if m else None
    conn.close()

    params = {'page': page, 'per_page': per_page, 'status': status}
    if search:
        params['search'] = search

    try:
        resp = req.get(
            f'{api_url}/wp-json/wc/v3/products',
            auth=(ck, cs),
            params=params,
            timeout=60,  # WC list endpoints can be slow on big catalogs
            headers={'User-Agent': 'WooCommerce API Client-Python/3.0.0',
                     'Accept': 'application/json'},
        )
    except req.RequestException as e:
        return jsonify({'error': f'连接 WC API 失败: {e}', 'routing': routing_info}), 502

    products, err = _parse_wc_response(resp)
    if err:
        app.logger.warning(f'GET products from {api_url} failed: {err}')
        return jsonify({'error': err, 'routing': routing_info}), 502
    products = products or []
    # Strip down to the fields we actually need (smaller payload, less surface area)
    slim = []
    for p in products:
        slim.append({
            'id': p.get('id'),
            'name': p.get('name'),
            'sku': p.get('sku', ''),
            'type': p.get('type', 'simple'),
            'status': p.get('status'),
            'manage_stock': bool(p.get('manage_stock', False)),
            'stock_quantity': p.get('stock_quantity'),
            'stock_status': p.get('stock_status', 'instock'),
            'regular_price': p.get('regular_price', ''),
            'sale_price': p.get('sale_price', ''),
            'price': p.get('price', ''),  # display-only effective price
            'permalink': p.get('permalink', ''),
            'image': (p.get('images') or [{}])[0].get('src', '') if p.get('images') else '',
            'variations_count': len(p.get('variations') or []),
        })

    total = int(resp.headers.get('X-WP-Total', len(slim)))
    total_pages = int(resp.headers.get('X-WP-TotalPages', 1))
    return jsonify({
        'products': slim,
        'page': page,
        'per_page': per_page,
        'total': total,
        'total_pages': total_pages,
        'routing': routing_info,
    })


def _build_product_update_payload(data):
    """Build a WC API PUT body from raw user input. Returns (payload, error_msg).
    Filters out non-whitelisted fields and normalizes types.

    Empty string values are treated as 'clear this field' (matches WC API behavior)
    EXCEPT for stock_quantity, where empty means 'leave unchanged' (skip the field)
    so toggling manage_stock without retyping the qty doesn't reset stock to 0."""
    payload = {}

    if 'manage_stock' in data:
        payload['manage_stock'] = bool(data['manage_stock'])

    # stock_quantity — process independently of whether manage_stock is in this
    # payload. The frontend uses dirty-tracking and may send stock_quantity alone
    # (manage_stock unchanged on the WC side stays as-is). Empty/None means skip.
    if 'stock_quantity' in data:
        v = data['stock_quantity']
        if v != '' and v is not None:
            try:
                qty = int(v)
            except (TypeError, ValueError):
                return None, '库存数量必须是整数'
            if qty < 0:
                return None, '库存数量不能为负数'
            payload['stock_quantity'] = qty

    # stock_status — only set if non-empty
    if 'stock_status' in data and data['stock_status']:
        valid = {'instock', 'outofstock', 'onbackorder'}
        if data['stock_status'] not in valid:
            return None, f'无效的库存状态: {data["stock_status"]}'
        payload['stock_status'] = data['stock_status']

    # Prices — WC accepts strings. Empty string explicitly clears the field on WC.
    # Treat None/empty as clear (= valid input), only reject actually malformed numbers.
    for k in ('regular_price', 'sale_price'):
        if k in data:
            v = data[k]
            if v is None or v == '':
                payload[k] = ''
                continue
            try:
                f = float(v)
            except (TypeError, ValueError):
                return None, f'价格格式错误: {k}={v}'
            if f < 0:
                return None, f'价格不能为负数: {k}={v}'
            payload[k] = str(v).strip()

    # Drop empty payload — would be a no-op write
    if not payload:
        return None, '没有可更新的字段'
    # Drop fields not in whitelist (defense in depth)
    payload = {k: v for k, v in payload.items() if k in PRODUCT_EDIT_WHITELIST}
    return payload, None


@app.route('/api/product-manager/variations/<int:site_id>/<int:parent_id>')
@login_required
@product_manager_required
def product_manager_list_variations(site_id, parent_id):
    """List all variations of a variable product (paginated under the hood;
    most variable products have <100 variations so we just fetch all)."""
    import requests as req

    conn = get_db_connection()
    try:
        site, api_url, ck, cs = _resolve_site_for_product_edit(conn, site_id)
    except ValueError as e:
        conn.close()
        return jsonify({'error': str(e)}), 400
    conn.close()

    # WC variations max 100 per page; fetch all pages until empty
    all_variations = []
    page = 1
    while True:
        try:
            resp = req.get(
                f'{api_url}/wp-json/wc/v3/products/{parent_id}/variations',
                auth=(ck, cs),
                params={'page': page, 'per_page': 100},
                timeout=60,
                headers={'User-Agent': 'WooCommerce API Client-Python/3.0.0',
                         'Accept': 'application/json'},
            )
        except req.RequestException as e:
            return jsonify({'error': f'连接 WC API 失败: {e}'}), 502

        batch, err = _parse_wc_response(resp)
        if err:
            app.logger.warning(f'GET variations of {parent_id} from {api_url} failed: {err}')
            return jsonify({'error': err}), 502
        batch = batch or []
        if not batch:
            break
        all_variations.extend(batch)
        if len(batch) < 100:
            break
        page += 1
        if page > 10:  # safety bound — 1000 variations is already absurd
            break

    slim = []
    for v in all_variations:
        # WC returns attributes as list of {id, name, option}
        attr_summary = ' / '.join(
            f"{a.get('name', '')}: {a.get('option', '')}"
            for a in (v.get('attributes') or [])
            if a.get('option')
        )
        slim.append({
            'id': v.get('id'),
            'parent_id': parent_id,
            'sku': v.get('sku', ''),
            'attributes_summary': attr_summary,
            'manage_stock': bool(v.get('manage_stock', False)),
            'stock_quantity': v.get('stock_quantity'),
            'stock_status': v.get('stock_status', 'instock'),
            'regular_price': v.get('regular_price', ''),
            'sale_price': v.get('sale_price', ''),
            'price': v.get('price', ''),
            'image': (v.get('image') or {}).get('src', ''),
        })
    return jsonify({'variations': slim, 'parent_id': parent_id})


@app.route('/api/product-manager/products/<int:site_id>/<int:parent_id>/variations/<int:variation_id>', methods=['PUT'])
@login_required
@product_manager_required
def product_manager_update_variation(site_id, parent_id, variation_id):
    """Update a single variation. Routes to master if applicable."""
    import requests as req

    data = request.get_json(silent=True) or {}
    payload, err = _build_product_update_payload(data)
    if err:
        return jsonify({'success': False, 'error': err}), 400

    conn = get_db_connection()
    try:
        site, api_url, ck, cs = _resolve_site_for_product_edit(conn, site_id)
    except ValueError as e:
        conn.close()
        return jsonify({'success': False, 'error': str(e)}), 400
    conn.close()

    try:
        resp = req.put(
            f'{api_url}/wp-json/wc/v3/products/{parent_id}/variations/{variation_id}',
            auth=(ck, cs),
            json=payload,
            timeout=90,  # WooMultistore can take 30-60s to sync to child stores
            headers={'User-Agent': 'WooCommerce API Client-Python/3.0.0',
                     'Content-Type': 'application/json',
                     'Accept': 'application/json'},
        )
    except req.RequestException as e:
        return jsonify({'success': False, 'error': f'连接失败: {e}'}), 502

    v, err = _parse_wc_response(resp)
    if err:
        app.logger.warning(f'PUT variation {parent_id}/{variation_id} on {api_url} failed: {err}')
        return jsonify({'success': False, 'error': err}), 502

    v = v or {}
    return jsonify({
        'success': True,
        'variation': {
            'id': v.get('id'),
            'manage_stock': bool(v.get('manage_stock', False)),
            'stock_quantity': v.get('stock_quantity'),
            'stock_status': v.get('stock_status'),
            'regular_price': v.get('regular_price', ''),
            'sale_price': v.get('sale_price', ''),
            'price': v.get('price', ''),
        },
    })


@app.route('/api/product-manager/products/<int:site_id>/<int:product_id>', methods=['PUT'])
@login_required
@product_manager_required
def product_manager_update(site_id, product_id):
    """Update a single product on a specific site. Routes to master if applicable."""
    import requests as req

    data = request.get_json(silent=True) or {}
    payload, err = _build_product_update_payload(data)
    if err:
        return jsonify({'success': False, 'error': err}), 400

    conn = get_db_connection()
    try:
        site, api_url, ck, cs = _resolve_site_for_product_edit(conn, site_id)
    except ValueError as e:
        conn.close()
        return jsonify({'success': False, 'error': str(e)}), 400

    routing_label = None
    if site['product_master_id']:
        m = conn.execute(
            'SELECT label FROM product_masters WHERE id = ?',
            (site['product_master_id'],)
        ).fetchone()
        routing_label = m['label'] if m else None
    conn.close()

    try:
        resp = req.put(
            f'{api_url}/wp-json/wc/v3/products/{product_id}',
            auth=(ck, cs),
            json=payload,
            timeout=90,  # WooMultistore can take 30-60s to sync to child stores
            headers={'User-Agent': 'WooCommerce API Client-Python/3.0.0',
                     'Content-Type': 'application/json',
                     'Accept': 'application/json'},
        )
    except req.RequestException as e:
        return jsonify({'success': False, 'error': f'连接失败: {e}'}), 502

    p, err = _parse_wc_response(resp)
    if err:
        app.logger.warning(f'PUT product {product_id} on {api_url} failed: {err}')
        return jsonify({'success': False, 'error': err}), 502

    p = p or {}
    return jsonify({
        'success': True,
        'product': {
            'id': p.get('id'),
            'manage_stock': bool(p.get('manage_stock', False)),
            'stock_quantity': p.get('stock_quantity'),
            'stock_status': p.get('stock_status'),
            'regular_price': p.get('regular_price', ''),
            'sale_price': p.get('sale_price', ''),
            'price': p.get('price', ''),
        },
        'routing_label': routing_label,
    })


@app.route('/api/product-manager/bulk', methods=['POST'])
@login_required
@product_manager_required
def product_manager_bulk():
    """Bulk update products on a single site.

    Request body:
      {
        "site_id": 123,
        "items": [
          { "product_id": 456, "manage_stock": false, "stock_status": "outofstock" },
          { "product_id": 789, "regular_price": "199.00", "sale_price": "159.00" },
          ...
        ]
      }

    Returns per-item success/failure so partial failures don't lose state.
    """
    import requests as req

    data = request.get_json(silent=True) or {}
    try:
        site_id = int(data.get('site_id', 0))
    except (TypeError, ValueError):
        return jsonify({'error': 'site_id 必须是整数'}), 400
    items = data.get('items') or []
    if not site_id or not isinstance(items, list) or not items:
        return jsonify({'error': '请提供 site_id 和非空 items 列表'}), 400
    if len(items) > 200:
        return jsonify({'error': '单次最多处理 200 个产品，请分批'}), 400

    conn = get_db_connection()
    try:
        site, api_url, ck, cs = _resolve_site_for_product_edit(conn, site_id)
    except ValueError as e:
        conn.close()
        return jsonify({'error': str(e)}), 400
    conn.close()

    headers = {
        'User-Agent': 'WooCommerce API Client-Python/3.0.0',
        'Content-Type': 'application/json',
        'Accept': 'application/json',
    }

    results = {'success': [], 'failed': []}
    for item in items:
        pid = item.get('product_id')
        parent_id = item.get('parent_id')  # if present → this is a variation
        if not pid:
            results['failed'].append({'product_id': None, 'error': '缺少 product_id'})
            continue

        payload, err = _build_product_update_payload(item)
        if err:
            results['failed'].append({'product_id': pid, 'parent_id': parent_id, 'error': err})
            continue

        # Route to variation endpoint if parent_id is given, otherwise to product endpoint
        if parent_id:
            url = f'{api_url}/wp-json/wc/v3/products/{parent_id}/variations/{pid}'
        else:
            url = f'{api_url}/wp-json/wc/v3/products/{pid}'

        try:
            resp = req.put(url, auth=(ck, cs), json=payload, timeout=90, headers=headers)
        except req.RequestException as e:
            results['failed'].append({
                'product_id': pid, 'parent_id': parent_id,
                'error': f'连接失败: {e}'
            })
            continue

        p, err = _parse_wc_response(resp)
        if err:
            results['failed'].append({
                'product_id': pid, 'parent_id': parent_id,
                'error': err,
            })
            continue
        p = p or {}
        results['success'].append({
            'product_id': pid,
            'parent_id': parent_id,
            'manage_stock': bool(p.get('manage_stock', False)),
            'stock_quantity': p.get('stock_quantity'),
            'stock_status': p.get('stock_status'),
            'regular_price': p.get('regular_price', ''),
            'sale_price': p.get('sale_price', ''),
        })

    return jsonify({
        'success_count': len(results['success']),
        'failed_count': len(results['failed']),
        'results': results,
    })


# ----------------------------------------------------------------------------
# Layer 2: product cloning across sites
# ----------------------------------------------------------------------------

_WC_HEADERS = {
    'User-Agent': 'WooCommerce API Client-Python/3.0.0',
    'Content-Type': 'application/json',
    'Accept': 'application/json',
}


def _resolve_taxonomy_on_target(api_url, ck, cs, taxonomy, source_terms):
    """For each source category/tag, look up the corresponding term on the
    target site by slug. Returns (target_terms_list, warnings_list).

    `taxonomy` = 'categories' | 'tags'. Source terms are WC's typical
    [{id, name, slug}] shape. We don't auto-create missing terms — we just
    warn so the user knows to set them manually on the target."""
    import requests as req

    target_terms = []
    warnings = []
    label = '分类' if taxonomy == 'categories' else '标签'
    for src in source_terms or []:
        slug = (src.get('slug') or '').strip()
        name = src.get('name') or slug or '?'
        if not slug:
            warnings.append(f'源 {label} "{name}" 缺少 slug，已跳过')
            continue
        try:
            resp = req.get(
                f'{api_url}/wp-json/wc/v3/products/{taxonomy}',
                auth=(ck, cs),
                params={'slug': slug, 'per_page': 1},
                timeout=20,
                headers={'User-Agent': _WC_HEADERS['User-Agent'], 'Accept': 'application/json'},
            )
            terms, err = _parse_wc_response(resp)
            if err or not terms:
                warnings.append(f'目标站不存在 {label} "{name}" (slug={slug})，已跳过')
                continue
            target_terms.append({'id': terms[0].get('id')})
        except Exception as e:
            warnings.append(f'查询 {label} "{name}" 失败: {e}')
    return target_terms, warnings


def _clone_variations(src_url, src_ck, src_cs, tgt_url, tgt_ck, tgt_cs,
                      src_parent_id, tgt_parent_id, options):
    """Clone all variations of a variable product. Returns dict with success/failed lists."""
    import requests as req

    # Fetch all source variations (paginated)
    all_variations = []
    page = 1
    while True:
        try:
            resp = req.get(
                f'{src_url}/wp-json/wc/v3/products/{src_parent_id}/variations',
                auth=(src_ck, src_cs),
                params={'page': page, 'per_page': 100},
                timeout=60,
                headers={'User-Agent': _WC_HEADERS['User-Agent'], 'Accept': 'application/json'},
            )
            batch, err = _parse_wc_response(resp)
            if err or not batch:
                break
            all_variations.extend(batch)
            if len(batch) < 100:
                break
            page += 1
            if page > 10:
                break
        except Exception:
            break

    success = []
    failed = []
    for v in all_variations:
        payload = {
            'regular_price': v.get('regular_price', ''),
            'sale_price': v.get('sale_price', ''),
            'manage_stock': bool(v.get('manage_stock', False)),
            'stock_status': v.get('stock_status', 'instock'),
            # Map attributes by name (WC will match against parent's local attrs)
            'attributes': [
                {'name': a.get('name'), 'option': a.get('option')}
                for a in (v.get('attributes') or []) if a.get('name')
            ],
        }
        if v.get('manage_stock') and v.get('stock_quantity') is not None:
            payload['stock_quantity'] = v['stock_quantity']
        if v.get('weight'):
            payload['weight'] = v['weight']
        if v.get('dimensions'):
            payload['dimensions'] = v['dimensions']
        sku = (v.get('sku') or '').strip()
        if sku:
            payload['sku'] = sku
        if options.get('include_images') and v.get('image') and v['image'].get('src'):
            payload['image'] = {'src': v['image'].get('src')}

        try:
            resp = req.post(
                f'{tgt_url}/wp-json/wc/v3/products/{tgt_parent_id}/variations',
                auth=(tgt_ck, tgt_cs), json=payload, timeout=60, headers=_WC_HEADERS,
            )
            new_v, err = _parse_wc_response(resp)
            if err:
                # SKU collision? Retry without sku as a soft fallback
                if sku and 'sku' in err.lower():
                    retry_payload = dict(payload)
                    retry_payload.pop('sku', None)
                    try:
                        resp2 = req.post(
                            f'{tgt_url}/wp-json/wc/v3/products/{tgt_parent_id}/variations',
                            auth=(tgt_ck, tgt_cs), json=retry_payload, timeout=60, headers=_WC_HEADERS,
                        )
                        new_v, err2 = _parse_wc_response(resp2)
                        if err2:
                            failed.append({'src_id': v.get('id'), 'sku': sku, 'error': err2})
                            continue
                        success.append({'src_id': v.get('id'), 'tgt_id': new_v.get('id'), 'note': 'SKU 冲突，已跳过 SKU'})
                        continue
                    except Exception as e:
                        failed.append({'src_id': v.get('id'), 'sku': sku, 'error': str(e)})
                        continue
                failed.append({'src_id': v.get('id'), 'sku': sku, 'error': err})
                continue
            success.append({'src_id': v.get('id'), 'tgt_id': new_v.get('id')})
        except Exception as e:
            failed.append({'src_id': v.get('id'), 'sku': sku, 'error': str(e)})

    return {'success': success, 'failed': failed, 'total_source': len(all_variations)}


def _extract_source_image_urls(html, src_host):
    """Find every image URL in `html` that is hosted on the source site.

    WooCommerce only sideloads the gallery (`images` array) when creating a
    product — it copies the description/short_description HTML verbatim, so any
    <img> embedded in the copy keeps pointing at the source domain. This finds
    those source-owned image URLs so the caller can re-host them.

    Returns (ordered_keys, groups) where each key maps to
      {'submit': <absolute https url to fetch>, 'tokens': {raw strings in html}}
    Grouping by a scheme-stripped key means an image written as both
    https://, http:// and //host/… forms is fetched once but every textual
    variant gets rewritten. External (non-source) images are left untouched."""
    import re
    from urllib.parse import urlparse

    if not html:
        return [], {}

    def _bare(host):
        h = (host or '').lower().strip()
        return h[4:] if h.startswith('www.') else h

    src_bare = _bare(urlparse(src_host if '//' in src_host else '//' + src_host).netloc or src_host)
    token_re = re.compile(
        r"""(?:https?:)?//[^/"'\s)]+/[^"'\s)]+?\.(?:jpe?g|png|gif|webp|avif|svg|bmp)(?:\?[^"'\s)]*)?""",
        re.IGNORECASE,
    )

    order, groups = [], {}
    for tok in token_re.findall(html):
        absu = ('https:' + tok) if tok.startswith('//') else tok
        if _bare(urlparse(absu).netloc) != src_bare:
            continue  # external image — not ours to re-host
        key = re.sub(r'^https?:', '', absu)  # //host/path?query
        if key not in groups:
            groups[key] = {'submit': absu.replace('&amp;', '&'), 'tokens': set()}
            order.append(key)
        groups[key]['tokens'].add(tok)
    return order, groups


def _migrate_inline_images_after_clone(src, new_data, src_url, tgt_url,
                                       tgt_ck, tgt_cs, new_id, warnings):
    """Re-host description/short_description images that point at the source
    site into the target's media library, then rewrite the product HTML.

    WooCommerce never sideloads inline content images (only the `images`
    gallery), and WC consumer keys can't authenticate WP's /wp/v2/media upload
    endpoint — so we reuse WC's own sideloading: append the content images to
    the product's `images` array (PUT #1) to make WC fetch them, read back the
    new URLs, then PUT #2 to save the rewritten HTML and reset the gallery to
    the original images only (so inline images don't pollute the gallery).

    Best-effort: any failure leaves the product intact (description still
    points at the source) and records a warning. Never raises."""
    import re
    import requests as req
    from urllib.parse import urlparse

    orig_desc = src.get('description', '') or ''
    orig_short = src.get('short_description', '') or ''

    order, groups = _extract_source_image_urls(
        orig_desc + '\n' + orig_short, urlparse(src_url).netloc)
    if not order:
        return  # no inline source images — nothing to do

    MAX_IMG = 100
    truncated = len(order) > MAX_IMG
    if truncated:
        order = order[:MAX_IMG]

    content_urls = [groups[k]['submit'] for k in order]

    # Gallery created by the initial POST — keep these by id across both PUTs.
    gallery_ids = [im.get('id') for im in (new_data.get('images') or []) if im.get('id')]
    gcount = len(gallery_ids)
    put_url = f'{tgt_url}/wp-json/wc/v3/products/{new_id}'

    # PUT #1 — append content images so WC sideloads them into the media library.
    put1_images = [{'id': gid} for gid in gallery_ids] + [{'src': u} for u in content_urls]
    try:
        resp = req.put(put_url, auth=(tgt_ck, tgt_cs), json={'images': put1_images},
                       timeout=120, headers=_WC_HEADERS)
    except Exception as e:
        warnings.append(f'文案内图片迁移失败（描述仍指向源站）：{e}')
        return
    data1, err = _parse_wc_response(resp)
    if err:
        warnings.append(f'文案内图片迁移失败（描述仍指向源站）：{err}')
        return

    resp_imgs = data1.get('images') or []
    new_for_content = resp_imgs[gcount:]

    # Map each source URL -> new target URL. Index alignment is reliable (WC
    # preserves submission order); fall back to filename matching if counts drift.
    url_map = {}
    if len(new_for_content) == len(content_urls):
        for k, im in zip(order, new_for_content):
            if im.get('src'):
                url_map[k] = im['src']
    else:
        def _basekey(u):
            name = urlparse(u).path.rsplit('/', 1)[-1]
            name = re.sub(r'\.(?:jpe?g|png|gif|webp|avif|svg|bmp)$', '', name, flags=re.I)
            return re.sub(r'-\d+x\d+$', '', name).lower()
        new_by_base = {}
        for im in resp_imgs:
            if im.get('src'):
                new_by_base.setdefault(_basekey(im['src']), im['src'])
        for k in order:
            ns = new_by_base.get(_basekey(groups[k]['submit']))
            if ns:
                url_map[k] = ns

    if not url_map:
        warnings.append('文案内图片迁移：未匹配到新图片地址，描述仍指向源站')
        return

    def _rewrite(html):
        for k, ns in url_map.items():
            for tok in groups[k]['tokens']:
                html = html.replace(tok, ns)
        return html

    # PUT #2 — save rewritten HTML and restore the gallery to the original
    # images only (drops the inline images from the gallery / featured image;
    # they remain in the media library, referenced by the description).
    put2 = {
        'description': _rewrite(orig_desc),
        'short_description': _rewrite(orig_short),
        'images': [{'id': gid} for gid in gallery_ids],
    }
    try:
        resp = req.put(put_url, auth=(tgt_ck, tgt_cs), json=put2,
                       timeout=120, headers=_WC_HEADERS)
    except Exception as e:
        warnings.append(f'文案内图片已上传但描述回写失败：{e}')
        return
    _, err = _parse_wc_response(resp)
    if err:
        warnings.append(f'文案内图片已上传但描述回写失败：{err}')
        return

    msg = f'文案内图片迁移：{len(url_map)}/{len(order)} 张已重新托管到目标站'
    if truncated:
        msg += f'（图片超过 {MAX_IMG} 张，仅处理前 {MAX_IMG} 张）'
    warnings.append(msg)


def _clone_one_product(src_url, src_ck, src_cs, tgt_url, tgt_ck, tgt_cs,
                       source_product_id, options):
    """Clone a single product. Returns dict with new_id + warnings, or error key."""
    import requests as req

    # 1. Fetch full source product
    try:
        resp = req.get(
            f'{src_url}/wp-json/wc/v3/products/{source_product_id}',
            auth=(src_ck, src_cs), timeout=60,
            headers={'User-Agent': _WC_HEADERS['User-Agent'], 'Accept': 'application/json'},
        )
        src, err = _parse_wc_response(resp)
        if err:
            return {'error': f'读取源产品失败: {err}'}
    except Exception as e:
        return {'error': f'读取源产品失败: {e}'}

    if not src:
        return {'error': '源产品不存在'}

    warnings = []

    # 2. Build base payload
    payload = {
        'name': src.get('name', ''),
        'type': src.get('type', 'simple'),
        'status': options.get('status_on_target', 'draft'),
        'description': src.get('description', ''),
        'short_description': src.get('short_description', ''),
        'regular_price': src.get('regular_price', ''),
        'sale_price': src.get('sale_price', ''),
        'manage_stock': bool(src.get('manage_stock', False)),
        'stock_status': src.get('stock_status', 'instock'),
        'tax_status': src.get('tax_status', 'taxable'),
        'tax_class': src.get('tax_class', ''),
        'featured': bool(src.get('featured', False)),
        'catalog_visibility': src.get('catalog_visibility', 'visible'),
        'weight': src.get('weight', ''),
        'dimensions': src.get('dimensions') or {},
    }
    if src.get('manage_stock') and src.get('stock_quantity') is not None:
        payload['stock_quantity'] = src['stock_quantity']

    # SKU
    src_sku = (src.get('sku') or '').strip()
    if src_sku:
        payload['sku'] = src_sku

    # Images — WC fetches these from URL on POST
    if options.get('include_images'):
        imgs = src.get('images') or []
        payload['images'] = [
            {'src': i.get('src'), 'name': i.get('name', ''), 'alt': i.get('alt', '')}
            for i in imgs if i.get('src')
        ]

    # Categories / tags — resolved by slug on target
    target_cats, cat_warnings = _resolve_taxonomy_on_target(
        tgt_url, tgt_ck, tgt_cs, 'categories', src.get('categories') or [])
    if target_cats:
        payload['categories'] = target_cats
    warnings.extend(cat_warnings)

    target_tags, tag_warnings = _resolve_taxonomy_on_target(
        tgt_url, tgt_ck, tgt_cs, 'tags', src.get('tags') or [])
    if target_tags:
        payload['tags'] = target_tags
    warnings.extend(tag_warnings)

    # Attributes (variation axes for variable; spec attrs for simple)
    attrs_in = src.get('attributes') or []
    attrs_out = []
    for a in attrs_in:
        attrs_out.append({
            'id': 0,  # local attribute on target (avoids global pa_xxx ID mismatch)
            'name': a.get('name'),
            'position': a.get('position', 0),
            'visible': a.get('visible', True),
            'variation': a.get('variation', src.get('type') == 'variable'),
            'options': a.get('options') or [],
        })
    if attrs_out:
        payload['attributes'] = attrs_out

    # 3. POST to target — handle SKU collision with -COPY suffix
    def _post_create(p):
        return req.post(
            f'{tgt_url}/wp-json/wc/v3/products',
            auth=(tgt_ck, tgt_cs), json=p, timeout=90, headers=_WC_HEADERS,
        )

    try:
        resp = _post_create(payload)
        new_data, err = _parse_wc_response(resp)
    except Exception as e:
        return {'error': f'创建目标产品失败: {e}'}

    if err and src_sku and ('sku' in err.lower() or 'unique' in err.lower()):
        # Retry with a -COPY suffix
        retry_payload = dict(payload)
        retry_payload['sku'] = src_sku + '-COPY'
        try:
            resp = _post_create(retry_payload)
            new_data, err = _parse_wc_response(resp)
            if not err:
                warnings.append(f'SKU "{src_sku}" 在目标站冲突，已改用 "{retry_payload["sku"]}"')
                payload = retry_payload
        except Exception as e:
            return {'error': f'重试创建失败: {e}'}

    if err:
        return {'error': f'创建目标产品失败: {err}'}
    if not new_data or not new_data.get('id'):
        return {'error': '创建目标产品成功但未返回新 ID'}

    new_id = new_data['id']

    # 4. Variations (if variable + option enabled)
    if options.get('include_variations') and src.get('type') == 'variable':
        var_results = _clone_variations(
            src_url, src_ck, src_cs, tgt_url, tgt_ck, tgt_cs,
            source_product_id, new_id, options,
        )
        if var_results['failed']:
            warnings.append(f'变体克隆：{len(var_results["success"])} 成功 / {len(var_results["failed"])} 失败')
        else:
            warnings.append(f'变体克隆：{len(var_results["success"])} 个成功')

    # 5. Inline content images — WC sideloads only the gallery (`images`), not
    #    images embedded in description/short_description HTML, so re-host those
    #    too and rewrite the copy. Best-effort: never fails the clone.
    if options.get('include_images'):
        try:
            _migrate_inline_images_after_clone(
                src, new_data, src_url, tgt_url, tgt_ck, tgt_cs, new_id, warnings)
        except Exception as e:
            warnings.append(f'文案内图片迁移异常（描述仍指向源站）：{e}')

    return {
        'new_id': new_id,
        'name': new_data.get('name'),
        'sku': new_data.get('sku', ''),
        'permalink': new_data.get('permalink', ''),
        'warnings': warnings,
    }


@app.route('/api/product-manager/clone', methods=['POST'])
@login_required
@product_manager_required
def product_manager_clone():
    """Clone products from source site to target site.

    Request body:
      {
        "source_site_id": 16,
        "target_site_id": 1,
        "product_ids": [9946, 9950, ...],
        "include_variations": true,
        "include_images": true,
        "status_on_target": "draft"
      }
    """
    data = request.get_json(silent=True) or {}
    try:
        source_site_id = int(data.get('source_site_id', 0))
        target_site_id = int(data.get('target_site_id', 0))
    except (TypeError, ValueError):
        return jsonify({'error': 'site_id 必须是整数'}), 400

    product_ids = data.get('product_ids') or []
    if not (source_site_id and target_site_id and isinstance(product_ids, list) and product_ids):
        return jsonify({'error': '请提供 source_site_id / target_site_id / product_ids'}), 400
    if source_site_id == target_site_id:
        return jsonify({'error': '源站点和目标站点不能相同'}), 400
    if len(product_ids) > 50:
        return jsonify({'error': '单次最多克隆 50 个产品，请分批操作'}), 400

    options = {
        'include_variations': bool(data.get('include_variations', True)),
        'include_images': bool(data.get('include_images', True)),
        'status_on_target': data.get('status_on_target') or 'draft',
    }
    if options['status_on_target'] not in ('draft', 'pending', 'private', 'publish'):
        options['status_on_target'] = 'draft'

    conn = get_db_connection()
    try:
        src_site, src_url, src_ck, src_cs = _resolve_site_for_product_edit(conn, source_site_id)
        tgt_site, tgt_url, tgt_ck, tgt_cs = _resolve_site_for_product_edit(conn, target_site_id)
    except ValueError as e:
        conn.close()
        return jsonify({'error': str(e)}), 400
    conn.close()

    results = {'success': [], 'failed': []}
    for pid in product_ids:
        try:
            r = _clone_one_product(src_url, src_ck, src_cs, tgt_url, tgt_ck, tgt_cs, pid, options)
        except Exception as e:
            results['failed'].append({'product_id': pid, 'error': f'未知错误: {e}'})
            continue
        if r.get('error'):
            results['failed'].append({'product_id': pid, 'error': r['error']})
        else:
            results['success'].append({
                'source_id': pid,
                'target_id': r['new_id'],
                'name': r.get('name'),
                'sku': r.get('sku'),
                'permalink': r.get('permalink'),
                'warnings': r.get('warnings') or [],
            })

    return jsonify({
        'success_count': len(results['success']),
        'failed_count': len(results['failed']),
        'results': results,
        'target_url': tgt_url,
    })


@app.route('/product-manager')
@login_required
@product_manager_required
def product_manager():
    """Product manager page — Layer 1 of multi-site product management.
    Site list is scoped to user's permissions: super admin sees all; everyone
    else sees only sites where they are the named manager (sites.manager ==
    users.name). This matches the edit-permission rule enforced by
    _resolve_site_for_product_edit."""
    conn = get_db_connection()
    sites = conn.execute('''
        SELECT id, url, manager, country, product_master_id
        FROM sites
        WHERE consumer_key IS NOT NULL AND consumer_key != ''
          AND consumer_secret IS NOT NULL AND consumer_secret != ''
        ORDER BY country, manager, url
    ''').fetchall()

    # Non-admin: restrict to sites where the user is the named manager.
    if not current_user.is_admin():
        user_name = (current_user.name or '').strip()
        sites = [s for s in sites if user_name and (s['manager'] or '').strip() == user_name]

    # Build display info, including which master (if any) drives writes for each site
    masters_lookup = {}
    pm_rows = conn.execute('SELECT id, label, url FROM product_masters').fetchall()
    for r in pm_rows:
        masters_lookup[r['id']] = {
            'label': r['label'],
            'host': r['url'].replace('https://www.', '').replace('https://', '').replace('http://', ''),
        }
    conn.close()
    return render_template('product_manager.html', sites=sites, masters_lookup=masters_lookup)


@app.route('/api/sites', methods=['POST'])
@login_required
@admin_required
def add_site():
    """Add a new WooCommerce site. Admin-only: a site stores read/WRITE API
    credentials and config (country/master/manager) that feeds revenue &
    reconciliation — the UI is admin-gated, so the API must be too."""
    data = request.json
    url = data.get('url')
    ck = data.get('consumer_key')
    cs = data.get('consumer_secret')
    manager = data.get('manager', '')
    mask_id = data.get('mask_id', '')
    country = data.get('country', '')
    # product_master_id: optional. Empty string / None / 0 / "" → NULL (standalone site)
    raw_pm = data.get('product_master_id')
    product_master_id = int(raw_pm) if raw_pm not in (None, '', 0, '0') else None

    if not all([url, ck, cs]):
        return jsonify({'error': 'Missing required fields'}), 400

    # Remove trailing slash from URL
    if url.endswith('/'):
        url = url[:-1]

    conn = get_db_connection()
    try:
        conn.execute('''INSERT INTO sites
            (url, consumer_key, consumer_secret, manager, mask_id, country, product_master_id)
            VALUES (?, ?, ?, ?, ?, ?, ?)''',
                     (url, ck, cs, manager, mask_id, country, product_master_id))
        conn.commit()
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'error': str(e)}), 500
    finally:
        conn.close()



@app.route('/api/sites/<int:site_id>', methods=['PUT'])
@login_required
@admin_required
def update_site(site_id):
    """Update a WooCommerce site. Admin-only (edits API credentials/config)."""
    data = request.json
    url = data.get('url')
    ck = data.get('consumer_key')
    cs = data.get('consumer_secret')
    manager = data.get('manager', '')
    mask_id = data.get('mask_id', '')
    country = data.get('country', '')
    # product_master_id: optional. Empty string / None / 0 / "" → NULL (un-link / standalone)
    raw_pm = data.get('product_master_id')
    product_master_id = int(raw_pm) if raw_pm not in (None, '', 0, '0') else None
    # Whether on-hold means "shipped" for this site (PL COD convention).
    # Defaults to current DB value when the client doesn't send the flag, so
    # legacy callers (e.g. the import-from-script path) keep their setting.
    cod_on_hold_raw = data.get('cod_on_hold_is_shipped', None)

    if not all([url, ck, cs]):
        return jsonify({'error': 'Missing required fields'}), 400

    # Remove trailing slash from URL
    if url.endswith('/'):
        url = url[:-1]

    conn = get_db_connection()
    try:
        if cod_on_hold_raw is None:
            conn.execute('''UPDATE sites
                SET url = ?, consumer_key = ?, consumer_secret = ?,
                    manager = ?, mask_id = ?, country = ?, product_master_id = ?
                WHERE id = ?''',
                         (url, ck, cs, manager, mask_id, country, product_master_id, site_id))
        else:
            cod_on_hold_val = 1 if cod_on_hold_raw in (1, '1', True, 'true', 'on') else 0
            conn.execute('''UPDATE sites
                SET url = ?, consumer_key = ?, consumer_secret = ?,
                    manager = ?, mask_id = ?, country = ?, product_master_id = ?,
                    cod_on_hold_is_shipped = ?
                WHERE id = ?''',
                         (url, ck, cs, manager, mask_id, country, product_master_id,
                          cod_on_hold_val, site_id))
        conn.commit()
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'error': str(e)}), 500
    finally:
        conn.close()

@app.route('/api/sites/<int:site_id>', methods=['DELETE'])
@login_required
@admin_required
def delete_site(site_id):
    """Delete a WooCommerce site. Admin-only."""
    print(f"Received delete request for site_id: {site_id}") # Debug log
    conn = get_db_connection()
    try:
        conn.execute('DELETE FROM sites WHERE id = ?', (site_id,))
        conn.commit()
        print(f"Successfully deleted site_id: {site_id}") # Debug log
        return jsonify({'success': True})
    except Exception as e:
        print(f"Error deleting site: {e}") # Debug log
        return jsonify({'error': str(e)}), 500
    finally:
        conn.close()

# Global sync status storage
# Format: {site_id: {'status': 'idle'|'running'|'success'|'error', 'progress': 0, 'message': '', 'logs': []}}
SYNC_STATUS = {}

@app.route('/api/sync/status/<int:site_id>')
@login_required
def get_sync_status(site_id):
    """Get synchronization status for a site.

    Returns the in-memory SYNC_STATUS entry plus a derived `stale_seconds`
    (seconds since last heartbeat) so the frontend can tell a healthy
    long-running sync apart from a zombie one whose worker is gone.
    Workers don't share memory, so a poll hitting a different worker than
    the one running the sync will see no entry at all — also surfaced as
    a fresh 'unknown' status to the client.
    """
    import time as _time
    entry = SYNC_STATUS.get(site_id)
    if entry is None:
        # This worker has no record. Either the sync never ran here, or it
        # ran in another worker that's since been recycled (HUP / OOM).
        return jsonify({
            'status': 'unknown',
            'message': '',
            'logs': [],
            'stale_seconds': None,
        })

    out = dict(entry)
    updated_at = out.get('updated_at')
    if updated_at:
        out['stale_seconds'] = round(_time.time() - updated_at, 1)
    else:
        out['stale_seconds'] = None
    return jsonify(out)

@app.route('/api/sync', methods=['POST'])
@login_required
def sync_data():
    """Trigger data synchronization"""
    import sync_utils
    import threading
    
    site_id = request.json.get('site_id')
    if not site_id:
        return jsonify({'error': 'Missing site_id'}), 400
        
    site_id = int(site_id)
    
    # Initialize status
    SYNC_STATUS[site_id] = {
        'status': 'running',
        'message': 'Starting synchronization...',
        'logs': [f"[{datetime.now().strftime('%H:%M:%S')}] Job started"]
    }
    
    def run_sync(app_context, site_id):
        with app_context:
            sync_start_time = datetime.now()
            try:
                conn = get_db_connection()
                site = conn.execute('SELECT * FROM sites WHERE id = ?', (site_id,)).fetchone()
                conn.close()
                
                if not site:
                    SYNC_STATUS[site_id]['status'] = 'error'
                    SYNC_STATUS[site_id]['message'] = 'Site not found'
                    return

                def progress_callback(msg):
                    timestamp = datetime.now().strftime('%H:%M:%S')
                    log_entry = f"[{timestamp}] {msg}"
                    SYNC_STATUS[site_id]['message'] = msg
                    SYNC_STATUS[site_id]['logs'].append(log_entry)
                    print(log_entry) # Keep console logging for debug

                result = sync_utils.sync_site(
                    site['url'], 
                    site['consumer_key'], 
                    site['consumer_secret'],
                    progress_callback
                )
                
                duration = int((datetime.now() - sync_start_time).total_seconds())
                
                if result['status'] == 'success':
                    # Update last sync time
                    conn = get_db_connection()
                    conn.execute('UPDATE sites SET last_sync = ? WHERE id = ?', 
                                 (datetime.now().strftime('%Y-%m-%d %H:%M:%S'), site_id))
                    conn.commit()
                    conn.close()
                    
                    SYNC_STATUS[site_id]['status'] = 'success'
                    SYNC_STATUS[site_id]['message'] = 'Synchronization completed successfully'
                    SYNC_STATUS[site_id]['logs'].append(f"[{datetime.now().strftime('%H:%M:%S')}] Finished: New {result['new_orders']}, Updated {result['updated_orders']}")
                    
                    # Save sync log to database
                    save_sync_log(site_id, site['url'], 'success', 
                                  f"New: {result['new_orders']}, Updated: {result['updated_orders']}", 
                                  result['new_orders'], result['updated_orders'], duration)
                else:
                    SYNC_STATUS[site_id]['status'] = 'error'
                    SYNC_STATUS[site_id]['message'] = result.get('message', 'Unknown error')
                    SYNC_STATUS[site_id]['logs'].append(f"[{datetime.now().strftime('%H:%M:%S')}] Error: {result.get('message')}")
                    
                    # Save error log to database
                    save_sync_log(site_id, site['url'], 'error', result.get('message', 'Unknown error'), 0, 0, duration)
                    
            except Exception as e:
                duration = int((datetime.now() - sync_start_time).total_seconds())
                SYNC_STATUS[site_id]['status'] = 'error'
                SYNC_STATUS[site_id]['message'] = str(e)
                SYNC_STATUS[site_id]['logs'].append(f"[{datetime.now().strftime('%H:%M:%S')}] Critical Error: {str(e)}")
                
                # Save error log to database
                try:
                    save_sync_log(site_id, '', 'error', str(e), 0, 0, duration)
                except:
                    pass

    # Run sync in background thread
    thread = threading.Thread(target=run_sync, args=(app.app_context(), site_id))
    thread.start()
    
    return jsonify({'success': True, 'message': 'Synchronization started'})


@app.route('/api/sync/deep/<int:site_id>', methods=['POST'])
@login_required
def deep_sync_site(site_id):
    """Trigger deep sync for a single site using 1.wooorders_sqlite.py"""
    import subprocess
    import threading
    
    # Use unique ID for deep sync status (site_id + 100000)
    status_id = site_id + 100000
    
    SYNC_STATUS[status_id] = {
        'status': 'running',
        'message': '正在启动单站点深度同步...',
        'logs': [f"[{datetime.now().strftime('%H:%M:%S')}] Deep sync started for site {site_id}"]
    }
    
    def run_deep_sync(app_context, site_id, status_id):
        with app_context:
            try:
                conn = get_db_connection()
                site = conn.execute('SELECT * FROM sites WHERE id = ?', (site_id,)).fetchone()
                conn.close()
                
                if not site:
                    SYNC_STATUS[status_id]['status'] = 'error'
                    SYNC_STATUS[status_id]['message'] = 'Site not found'
                    return
                
                site_url = site['url']
                SYNC_STATUS[status_id]['message'] = f'正在深度同步 {site_url}...'
                SYNC_STATUS[status_id]['logs'].append(f"[{datetime.now().strftime('%H:%M:%S')}] Processing {site_url}")
                
                # Import sync functions directly
                from woocommerce import API
                import time
                import random
                
                # Create WooCommerce API client
                wcapi = API(
                    url=site_url,
                    consumer_key=site['consumer_key'],
                    consumer_secret=site['consumer_secret'],
                    version="wc/v3",
                    timeout=30
                )
                
                # Fetch all orders (full resync)
                orders = []
                page = 1
                per_page = 50
                
                SYNC_STATUS[status_id]['message'] = f'正在获取所有订单...'
                
                while True:
                    try:
                        time.sleep(random.uniform(0.5, 1))
                        response = wcapi.get("orders", params={
                            "per_page": per_page,
                            "page": page,
                            "expand": "line_items,shipping_lines,tax_lines,fee_lines,coupon_lines,refunds"
                        })
                        
                        if response.status_code in (401, 403):
                            # API authentication/authorization error
                            error_msg = f"API认证失败 (HTTP {response.status_code})"
                            try:
                                error_data = response.json()
                                if 'message' in error_data:
                                    error_msg = f"{error_msg}: {error_data['message']}"
                            except:
                                pass
                            
                            SYNC_STATUS[status_id]['logs'].append(f"[{datetime.now().strftime('%H:%M:%S')}] ❌ {error_msg}")
                            SYNC_STATUS[status_id]['status'] = 'error'
                            SYNC_STATUS[status_id]['message'] = error_msg
                            
                            # Update site API status in database
                            conn = get_db_connection()
                            conn.execute('UPDATE sites SET api_status = ?, last_api_error = ? WHERE id = ?', 
                                         ('error', error_msg, site_id))
                            conn.commit()
                            conn.close()
                            return
                        
                        if response.status_code != 200:
                            SYNC_STATUS[status_id]['logs'].append(f"[{datetime.now().strftime('%H:%M:%S')}] HTTP {response.status_code}")
                            break
                        
                        data = response.json()
                        if not data:
                            break
                        
                        for order in data:
                            order['source'] = site_url
                        
                        # Save orders using sync_utils
                        import sync_utils
                        sync_utils.save_orders_to_db(data)
                        orders.extend(data)
                        
                        SYNC_STATUS[status_id]['message'] = f'已获取 {len(orders)} 个订单 (页 {page})'
                        SYNC_STATUS[status_id]['logs'].append(f"[{datetime.now().strftime('%H:%M:%S')}] Page {page}: {len(data)} orders")
                        
                        page += 1
                        
                    except Exception as e:
                        SYNC_STATUS[status_id]['logs'].append(f"[{datetime.now().strftime('%H:%M:%S')}] Error: {str(e)}")
                        break
                
                # Update last sync time and API status (success)
                conn = get_db_connection()
                conn.execute('UPDATE sites SET last_sync = ?, api_status = ?, last_api_error = NULL WHERE id = ?', 
                             (datetime.now().strftime('%Y-%m-%d %H:%M:%S'), 'ok', site_id))
                conn.commit()
                conn.close()
                
                SYNC_STATUS[status_id]['status'] = 'success'
                SYNC_STATUS[status_id]['message'] = f'深度同步完成，共 {len(orders)} 个订单'
                SYNC_STATUS[status_id]['logs'].append(f"[{datetime.now().strftime('%H:%M:%S')}] Completed: {len(orders)} total orders")
                
            except Exception as e:
                SYNC_STATUS[status_id]['status'] = 'error'
                SYNC_STATUS[status_id]['message'] = str(e)
                SYNC_STATUS[status_id]['logs'].append(f"[{datetime.now().strftime('%H:%M:%S')}] Critical Error: {str(e)}")
    
    thread = threading.Thread(target=run_deep_sync, args=(app.app_context(), site_id, status_id))
    thread.start()
    
    return jsonify({'success': True, 'sync_id': status_id, 'message': 'Deep sync started'})


@app.route('/api/site/<int:site_id>/check', methods=['POST'])
@login_required
def check_site_api(site_id):
    """Check API connectivity for a site"""
    from woocommerce import API
    
    conn = get_db_connection()
    site = conn.execute('SELECT * FROM sites WHERE id = ?', (site_id,)).fetchone()
    
    if not site:
        conn.close()
        return jsonify({'success': False, 'error': 'Site not found'}), 404
    
    try:
        wcapi = API(
            url=site['url'],
            consumer_key=site['consumer_key'],
            consumer_secret=site['consumer_secret'],
            version="wc/v3",
            timeout=15,
            user_agent="WooCommerce API Client-Python/3.0.0" # Critical for WAF bypass
        )
        
        read_status = 'unknown'
        write_status = 'unknown'
        error_msg = None
        
        # 1. Test Read Permission
        try:
            response = wcapi.get("orders", params={"per_page": 1})
            if response.status_code == 200:
                read_status = 'ok'
                
                # 2. Test Write Permission (Only if read is OK)
                try:
                    orders = response.json()
                    if orders and len(orders) > 0:
                        test_order_id = orders[0]['id']
                        
                        # Try to add test note
                        test_note_response = wcapi.post(
                            f"orders/{test_order_id}/notes",
                            data={
                                "note": "[API权限测试] 此消息用于验证写权限，将立即删除",
                                "customer_note": False
                            }
                        )
                        
                        if test_note_response.status_code in (200, 201):
                            write_status = 'ok'
                            # Cleanup
                            try:
                                note_id = test_note_response.json().get('id')
                                if note_id:
                                    wcapi.delete(f"orders/{test_order_id}/notes/{note_id}")
                            except:
                                pass
                        elif test_note_response.status_code in (401, 403):
                            write_status = 'error'
                            error_msg = "写权限被拒绝"
                        else:
                            write_status = 'error'
                            error_msg = f"写入失败 HTTP {test_note_response.status_code}"
                    else:
                        write_status = 'unknown' # No order to test on
                except Exception as e:
                    write_status = 'error'
                    error_msg = f"写测试出错: {str(e)}"
            elif response.status_code in (401, 403):
                read_status = 'error'
                write_status = 'unknown'
                error_msg = "读权限被拒绝"
            else:
                read_status = 'error'
                write_status = 'unknown'
                error_msg = f"连接异常 HTTP {response.status_code}"
                
        except Exception as e:
            read_status = 'error'
            write_status = 'unknown'
            error_msg = f"连接失败: {str(e)}"

        # Update DB
        conn.execute('''
            UPDATE sites 
            SET api_read_status = ?, api_write_status = ?, last_api_error = ?
            WHERE id = ?
        ''', (read_status, write_status, error_msg, site_id))
        conn.commit()
        conn.close()

        status = 'ok' if (read_status == 'ok' and write_status == 'ok') else 'error'
        
        return jsonify({
            'success': True, 
            'status': status,
            'message': error_msg or 'API连接正常',
            'read': read_status,
            'write': write_status
        })

    except Exception as e:
        return jsonify({'success': False, 'error': f"系统错误: {str(e)}"}), 500


@app.route('/api/site/<int:site_id>/check-tracking-api', methods=['POST'])
@login_required
def check_tracking_api(site_id):
    """Check if woo-tracking REST API plugin is available on the site"""
    import requests as req
    
    conn = get_db_connection()
    site = conn.execute('SELECT * FROM sites WHERE id = ?', (site_id,)).fetchone()
    
    if not site:
        conn.close()
        return jsonify({'success': False, 'error': '站点不存在'}), 404
    
    tracking_api_url = f"{site['url']}/wp-json/woo-tracking/v1/carriers"
    
    # Headers to simulate official WooCommerce API client
    headers = {
        "User-Agent": "WooCommerce API Client-Python/3.0.0",
        "Content-Type": "application/json",
        "Accept": "application/json"
    }
    
    try:
        response = req.get(
            tracking_api_url,
            auth=(site['consumer_key'], site['consumer_secret']),
            timeout=15,
            headers=headers
        )
        
        if response.status_code == 200:
            data = response.json()
            carriers = data.get('carriers', [])
            tracking_status = 'ok'
            message = f"物流API可用，支持 {len(carriers)} 个物流商"
        elif response.status_code == 404:
            tracking_status = 'not_installed'
            message = "物流追踪插件未安装"
        elif response.status_code in (401, 403):
            tracking_status = 'auth_error'
            message = "物流API权限不足"
        else:
            tracking_status = 'error'
            message = f"物流API异常 HTTP {response.status_code}"
        
        # Update database
        conn.execute('UPDATE sites SET tracking_api_status = ? WHERE id = ?', (tracking_status, site_id))
        conn.commit()
        conn.close()
        
        return jsonify({
            'success': True,
            'tracking_api_status': tracking_status,
            'message': message,
            'carriers': data.get('carriers', []) if response.status_code == 200 else []
        })
        
    except req.exceptions.Timeout:
        conn.execute('UPDATE sites SET tracking_api_status = ? WHERE id = ?', ('timeout', site_id))
        conn.commit()
        conn.close()
        return jsonify({'success': False, 'tracking_api_status': 'timeout', 'error': '连接超时'})
    except Exception as e:
        conn.close()
        return jsonify({'success': False, 'error': f"检测失败: {str(e)}"}), 500


@app.route('/api/site/<int:site_id>/email-logs')
@login_required
def get_site_email_logs(site_id):
    """Get email logs from WordPress site via woo-tracking REST API"""
    import requests as req
    
    conn = get_db_connection()
    site = conn.execute('SELECT * FROM sites WHERE id = ?', (site_id,)).fetchone()
    
    if not site:
        conn.close()
        return jsonify({'success': False, 'error': '站点不存在'}), 404
    
    conn.close()
    
    page = request.args.get('page', 1, type=int)
    per_page = request.args.get('per_page', 20, type=int)
    
    email_logs_url = f"{site['url']}/wp-json/woo-tracking/v1/email-logs"
    
    headers = {
        "User-Agent": "WooCommerce API Client-Python/3.0.0",
        "Content-Type": "application/json",
        "Accept": "application/json"
    }
    
    try:
        response = req.get(
            email_logs_url,
            params={'page': page, 'per_page': per_page},
            auth=(site['consumer_key'], site['consumer_secret']),
            timeout=30,
            headers=headers
        )
        
        if response.status_code == 200:
            return jsonify({'success': True, **response.json()})
        elif response.status_code == 404:
            return jsonify({'success': False, 'error': '邮件日志API不可用', 'logs': []})
        else:
            return jsonify({'success': False, 'error': f'API错误 {response.status_code}', 'logs': []})
            
    except Exception as e:
        return jsonify({'success': False, 'error': f'请求失败: {str(e)}', 'logs': []})


@app.route('/api/site/<int:site_id>/email-stats')
@login_required
def get_site_email_stats(site_id):
    """Get email statistics from WordPress site via woo-tracking REST API"""
    import requests as req
    
    conn = get_db_connection()
    site = conn.execute('SELECT * FROM sites WHERE id = ?', (site_id,)).fetchone()
    
    if not site:
        conn.close()
        return jsonify({'success': False, 'error': '站点不存在'}), 404
    
    conn.close()
    
    email_stats_url = f"{site['url']}/wp-json/woo-tracking/v1/email-stats"
    
    headers = {
        "User-Agent": "WooCommerce API Client-Python/3.0.0",
        "Content-Type": "application/json",
        "Accept": "application/json"
    }
    
    try:
        response = req.get(
            email_stats_url,
            auth=(site['consumer_key'], site['consumer_secret']),
            timeout=15,
            headers=headers
        )
        
        if response.status_code == 200:
            return jsonify({'success': True, **response.json()})
        elif response.status_code == 404:
            return jsonify({'success': False, 'error': '邮件统计API不可用'})
        else:
            return jsonify({'success': False, 'error': f'API错误 {response.status_code}'})
            
    except Exception as e:
        return jsonify({'success': False, 'error': f'请求失败: {str(e)}'})


# Global status dict for API check progress
CHECK_STATUS = {}

@app.route('/api/sites/check-all', methods=['POST'])
@login_required
def check_all_sites_api():
    """Check API connectivity for all sites - runs in background thread to avoid timeout"""
    import threading
    
    # Use a unique ID for this check operation
    CHECK_ALL_ID = 888888
    
    CHECK_STATUS[CHECK_ALL_ID] = {
        'status': 'running',
        'message': '正在启动API检测...',
        'results': [],
        'progress': 0,
        'total': 0
    }
    
    def run_check_all(app_context, status_id):
        from woocommerce import API
        
        with app_context:
            try:
                conn = get_db_connection()
                sites = conn.execute('SELECT * FROM sites').fetchall()
                
                total_sites = len(sites)
                CHECK_STATUS[status_id]['total'] = total_sites
                results = []
                
                for index, site in enumerate(sites):
                    site_id = site['id']
                    site_url = site['url']
                    read_status = 'unknown'
                    write_status = 'unknown'
                    error_msg = None
                    
                    CHECK_STATUS[status_id]['message'] = f'正在检测 {site_url} ({index+1}/{total_sites})...'
                    CHECK_STATUS[status_id]['progress'] = index + 1
                    
                    try:
                        wcapi = API(
                            url=site_url,
                            consumer_key=site['consumer_key'],
                            consumer_secret=site['consumer_secret'],
                            version="wc/v3",
                            timeout=15,
                            user_agent="WooCommerce API Client-Python/3.0.0"
                        )
                        
                        # Test READ permission
                        try:
                            response = wcapi.get("orders", params={"per_page": 1})
                            
                            if response.status_code == 200:
                                read_status = 'ok'
                            elif response.status_code in (401, 403):
                                read_status = 'error'
                                error_msg = f"读权限认证失败 (HTTP {response.status_code})"
                                try:
                                    error_data = response.json()
                                    if 'message' in error_data:
                                        error_msg = f"{error_msg}: {error_data['message']}"
                                except:
                                    pass
                            else:
                                read_status = 'error'
                                error_msg = f"读取失败 HTTP {response.status_code}"
                        except Exception as e:
                            read_status = 'error'
                            error_msg = f"读权限测试失败: {str(e)}"
                        
                        # Test WRITE permission (only if read is ok)
                        if read_status == 'ok':
                            try:
                                orders_response = wcapi.get("orders", params={"per_page": 1, "status": "any"})
                                
                                if orders_response.status_code == 200:
                                    orders = orders_response.json()
                                    
                                    if orders and len(orders) > 0:
                                        test_order_id = orders[0]['id']
                                        
                                        test_note_response = wcapi.post(
                                            f"orders/{test_order_id}/notes",
                                            data={
                                                "note": "[API权限测试] 此消息用于验证写权限，将立即删除",
                                                "customer_note": False
                                            }
                                        )
                                        
                                        if test_note_response.status_code in (200, 201):
                                            write_status = 'ok'
                                            try:
                                                note_id = test_note_response.json().get('id')
                                                if note_id:
                                                    wcapi.delete(f"orders/{test_order_id}/notes/{note_id}")
                                            except:
                                                pass
                                        elif test_note_response.status_code in (401, 403):
                                            write_status = 'error'
                                            if not error_msg:
                                                error_msg = "写权限被拒绝"
                                        else:
                                            write_status = 'error'
                                            if not error_msg:
                                                error_msg = f"写入测试失败 HTTP {test_note_response.status_code}"
                                    else:
                                        write_status = 'unknown'
                                        if not error_msg:
                                            error_msg = "无订单可测试写权限"
                                else:
                                    write_status = 'unknown'
                            except Exception as e:
                                write_status = 'error'
                                if not error_msg:
                                    error_msg = f"写权限测试失败: {str(e)}"
                        else:
                            write_status = 'unknown'
                            
                    except Exception as e:
                        read_status = 'error'
                        write_status = 'unknown'
                        error_msg = f"连接失败: {str(e)}"
                    
                    # Update database
                    conn.execute('''
                        UPDATE sites 
                        SET api_read_status = ?, api_write_status = ?, last_api_error = ?
                        WHERE id = ?
                    ''', (read_status, write_status, error_msg, site_id))
                    conn.commit()
                    
                    results.append({
                        'site_id': site_id,
                        'url': site_url,
                        'read': read_status,
                        'write': write_status,
                        'message': error_msg
                    })
                    
                    CHECK_STATUS[status_id]['results'] = results
                
                conn.close()
                
                # Calculate final stats
                ok_count = sum(1 for r in results if r['read'] == 'ok' and r['write'] == 'ok')
                error_count = sum(1 for r in results if r['read'] == 'error' or r['write'] == 'error')
                
                CHECK_STATUS[status_id]['status'] = 'success'
                CHECK_STATUS[status_id]['message'] = f'检测完成！正常: {ok_count} 个，异常: {error_count} 个'
                
            except Exception as e:
                CHECK_STATUS[status_id]['status'] = 'error'
                CHECK_STATUS[status_id]['message'] = f'检测失败: {str(e)}'
    
    thread = threading.Thread(target=run_check_all, args=(app.app_context(), CHECK_ALL_ID))
    thread.start()
    
    return jsonify({'success': True, 'check_id': CHECK_ALL_ID, 'message': 'API检测已启动'})


@app.route('/api/sites/check-status/<int:check_id>')
@login_required
def get_check_status(check_id):
    """Get the status of an ongoing API check operation"""
    if check_id not in CHECK_STATUS:
        return jsonify({'status': 'unknown', 'message': '检测任务不存在'})
    
    return jsonify(CHECK_STATUS[check_id])

@app.route('/api/sync/clean/<int:site_id>', methods=['POST'])
@login_required
def clean_sync_site(site_id):
    """Clean deleted orders for a single site"""
    import threading
    
    # Use unique ID for clean sync status (site_id + 200000)
    status_id = site_id + 200000
    
    SYNC_STATUS[status_id] = {
        'status': 'running',
        'message': '正在启动清理同步...',
        'logs': [f"[{datetime.now().strftime('%H:%M:%S')}] Clean sync started for site {site_id}"]
    }
    
    def run_clean_sync(app_context, site_id, status_id):
        with app_context:
            try:
                conn = get_db_connection()
                site = conn.execute('SELECT * FROM sites WHERE id = ?', (site_id,)).fetchone()
                conn.close()
                
                if not site:
                    SYNC_STATUS[status_id]['status'] = 'error'
                    SYNC_STATUS[status_id]['message'] = 'Site not found'
                    return
                
                # Strip whitespace to handle potential database issues
                site_url = site['url'].strip()
                SYNC_STATUS[status_id]['message'] = f'正在获取远程订单ID...'
                SYNC_STATUS[status_id]['logs'].append(f"[{datetime.now().strftime('%H:%M:%S')}] Fetching remote order IDs for {site_url}")
                
                from woocommerce import API
                import time
                import random
                
                wcapi = API(
                    url=site_url,
                    consumer_key=site['consumer_key'],
                    consumer_secret=site['consumer_secret'],
                    version="wc/v3",
                    timeout=30,
                    user_agent="WooCommerce API Client-Python/3.0.0"
                )
                
                # Fetch all remote order IDs
                remote_ids = set()
                page = 1
                per_page = 100
                
                while True:
                    try:
                        time.sleep(random.uniform(0.5, 1))
                        response = wcapi.get("orders", params={
                            "per_page": per_page,
                            "page": page,
                            "_fields": "id"
                        })
                        
                        if response.status_code != 200:
                            SYNC_STATUS[status_id]['logs'].append(f"[{datetime.now().strftime('%H:%M:%S')}] HTTP Error {response.status_code} on page {page}")
                            break
                        
                        data = response.json()
                        if not data:
                            break
                        
                        for order in data:
                            remote_ids.add(str(order['id']))
                        
                        SYNC_STATUS[status_id]['message'] = f'已获取 {len(remote_ids)} 个远程订单ID (页 {page})'
                        page += 1
                        
                    except Exception as e:
                        SYNC_STATUS[status_id]['logs'].append(f"[{datetime.now().strftime('%H:%M:%S')}] Error: {str(e)}")
                        break
                
                SYNC_STATUS[status_id]['logs'].append(f"[{datetime.now().strftime('%H:%M:%S')}] Found {len(remote_ids)} remote orders")
                
                # Always clean checkout-draft orders first (they are not returned by API)
                conn = get_db_connection()
                draft_orders = conn.execute("SELECT id FROM orders WHERE source = ? AND status = 'checkout-draft'", (site_url,)).fetchall()
                draft_ids = set(str(o['id']) for o in draft_orders)
                conn.close()
                
                draft_deleted = 0
                if draft_ids:
                    SYNC_STATUS[status_id]['logs'].append(f"[{datetime.now().strftime('%H:%M:%S')}] Found {len(draft_ids)} checkout-draft orders to clean")
                    try:
                        conn = get_db_connection()
                        placeholders = ','.join(['?' for _ in draft_ids])
                        conn.execute(f"DELETE FROM orders WHERE source = ? AND id IN ({placeholders})", 
                                     [site_url] + list(draft_ids))
                        conn.commit()
                        draft_deleted = conn.total_changes
                        conn.close()
                        SYNC_STATUS[status_id]['logs'].append(f"[{datetime.now().strftime('%H:%M:%S')}] Deleted {draft_deleted} checkout-draft orders")
                    except Exception as e:
                        SYNC_STATUS[status_id]['logs'].append(f"[{datetime.now().strftime('%H:%M:%S')}] Error deleting drafts: {str(e)}")
                
                if not remote_ids:
                    if draft_deleted > 0:
                        SYNC_STATUS[status_id]['status'] = 'success'
                        SYNC_STATUS[status_id]['message'] = f'清理完成，删除了 {draft_deleted} 个草稿订单'
                    else:
                        SYNC_STATUS[status_id]['status'] = 'error'
                        SYNC_STATUS[status_id]['message'] = '未获取到远程订单ID，跳过删除以避免误删'
                    return
                
                # Get local order IDs - use trimmed URL and handle exact match
                conn = get_db_connection()
                local_orders = conn.execute('SELECT id FROM orders WHERE source = ?', (site_url,)).fetchall()
                conn.close()
                
                local_ids = set(str(o['id']) for o in local_orders)
                orphaned_ids = local_ids - remote_ids
                
                SYNC_STATUS[status_id]['logs'].append(f"[{datetime.now().strftime('%H:%M:%S')}] Local (source='{site_url}'): {len(local_ids)}, Orphaned: {len(orphaned_ids)}")
                
                if not orphaned_ids:
                    SYNC_STATUS[status_id]['status'] = 'success'
                    SYNC_STATUS[status_id]['message'] = '没有需要清理的订单'
                    return
                
                # 通过受保护的归档逻辑处理孤儿单（P0-a：先归档留底 + 大批量回滚保护，取代直接物理删除）
                SYNC_STATUS[status_id]['message'] = f'正在处理 {len(orphaned_ids)} 个疑似已删除订单...'
                try:
                    removed = _get_woosync().archive_orphaned_orders(site_url, remote_ids)
                    if removed == 0 and len(orphaned_ids) > 0:
                        SYNC_STATUS[status_id]['status'] = 'success'
                        SYNC_STATUS[status_id]['message'] = f'检测到 {len(orphaned_ids)} 个孤儿单、疑似站点回滚，已跳过删除并记录告警（数据已保留，未丢失）'
                        SYNC_STATUS[status_id]['logs'].append(f"[{datetime.now().strftime('%H:%M:%S')}] Suspected rollback: skipped delete, kept {len(orphaned_ids)} orders, alert logged")
                    else:
                        SYNC_STATUS[status_id]['status'] = 'success'
                        SYNC_STATUS[status_id]['message'] = f'清理完成，已归档并移除 {removed} 个订单'
                        SYNC_STATUS[status_id]['logs'].append(f"[{datetime.now().strftime('%H:%M:%S')}] Archived & removed {removed} orphaned orders")
                except Exception as e:
                    SYNC_STATUS[status_id]['status'] = 'error'
                    SYNC_STATUS[status_id]['message'] = f'清理失败: {str(e)}'
                    SYNC_STATUS[status_id]['logs'].append(f"[{datetime.now().strftime('%H:%M:%S')}] Error: {str(e)}")
                
            except Exception as e:
                SYNC_STATUS[status_id]['status'] = 'error'
                SYNC_STATUS[status_id]['message'] = str(e)
                SYNC_STATUS[status_id]['logs'].append(f"[{datetime.now().strftime('%H:%M:%S')}] Critical Error: {str(e)}")
    
    thread = threading.Thread(target=run_clean_sync, args=(app.app_context(), site_id, status_id))
    thread.start()
    
    return jsonify({'success': True, 'sync_id': status_id, 'message': 'Clean sync started'})


@app.route('/api/sync/all', methods=['POST'])
@login_required
def sync_all_data():
    """Trigger data synchronization for ALL sites"""
    import sync_utils
    import threading
    import time as _time

    # Use a special ID for "all sites" sync status
    ALL_SITES_ID = 999999

    # `updated_at` is a wall-clock heartbeat the frontend uses to detect a
    # zombie sync — if the gunicorn worker hosting this thread is killed
    # (e.g. by a deploy HUP) the dict goes with it, and a fresh worker has
    # no record of this sync_id at all. The frontend polling would then sit
    # forever on the last cached logs. Bumping this on every progress tick
    # lets the UI flag "stale" when updated_at falls too far behind now.
    SYNC_STATUS[ALL_SITES_ID] = {
        'status': 'running',
        'message': 'Starting global synchronization...',
        'logs': [f"[{datetime.now().strftime('%H:%M:%S')}] Global sync job started"],
        'updated_at': _time.time(),
    }

    def _touch():
        SYNC_STATUS[ALL_SITES_ID]['updated_at'] = _time.time()

    def run_sync_all(app_context):
        with app_context:
            try:
                conn = get_db_connection()
                sites = conn.execute('SELECT * FROM sites').fetchall()
                conn.close()

                if not sites:
                    SYNC_STATUS[ALL_SITES_ID]['status'] = 'error'
                    SYNC_STATUS[ALL_SITES_ID]['message'] = 'No sites found to sync'
                    _touch()
                    return

                total_sites = len(sites)
                SYNC_STATUS[ALL_SITES_ID]['logs'].append(f"[{datetime.now().strftime('%H:%M:%S')}] Found {total_sites} sites to sync")
                _touch()

                for index, site in enumerate(sites):
                    site_id = site['id']
                    site_url = site['url']
                    current_step = index + 1

                    msg = f"Syncing site {current_step}/{total_sites}: {site_url}"
                    SYNC_STATUS[ALL_SITES_ID]['message'] = msg
                    SYNC_STATUS[ALL_SITES_ID]['logs'].append(f"[{datetime.now().strftime('%H:%M:%S')}] --- Starting {site_url} ---")
                    _touch()

                    def progress_callback(msg):
                        timestamp = datetime.now().strftime('%H:%M:%S')
                        # Prefix log with site info
                        log_entry = f"[{timestamp}] [{site_url}] {msg}"
                        SYNC_STATUS[ALL_SITES_ID]['logs'].append(log_entry)
                        # Only update main message if it's significant, otherwise keep "Syncing site X/Y"
                        # Actually, let's update the message to show detail
                        SYNC_STATUS[ALL_SITES_ID]['message'] = f"[{current_step}/{total_sites}] {site_url}: {msg}"
                        _touch()

                    result = sync_utils.sync_site(
                        site['url'],
                        site['consumer_key'],
                        site['consumer_secret'],
                        progress_callback
                    )

                    if result['status'] == 'success':
                        # Update last sync time
                        conn = get_db_connection()
                        conn.execute('UPDATE sites SET last_sync = ? WHERE id = ?',
                                     (datetime.now().strftime('%Y-%m-%d %H:%M:%S'), site_id))
                        conn.commit()
                        conn.close()
                        SYNC_STATUS[ALL_SITES_ID]['logs'].append(f"[{datetime.now().strftime('%H:%M:%S')}] {site_url} Completed")
                    else:
                        SYNC_STATUS[ALL_SITES_ID]['logs'].append(f"[{datetime.now().strftime('%H:%M:%S')}] {site_url} Failed: {result.get('message')}")
                    _touch()

                SYNC_STATUS[ALL_SITES_ID]['status'] = 'success'
                SYNC_STATUS[ALL_SITES_ID]['message'] = 'All sites synchronized successfully'
                SYNC_STATUS[ALL_SITES_ID]['logs'].append(f"[{datetime.now().strftime('%H:%M:%S')}] Global sync finished")
                _touch()

            except Exception as e:
                SYNC_STATUS[ALL_SITES_ID]['status'] = 'error'
                SYNC_STATUS[ALL_SITES_ID]['message'] = str(e)
                SYNC_STATUS[ALL_SITES_ID]['logs'].append(f"[{datetime.now().strftime('%H:%M:%S')}] Critical Error: {str(e)}")
                _touch()

    # Run sync in background thread
    thread = threading.Thread(target=run_sync_all, args=(app.app_context(),))
    thread.start()
    
    return jsonify({'success': True, 'message': 'Global synchronization started', 'sync_id': ALL_SITES_ID})


# =====================================================================
# 数据备份与灾备后台管理 API（P0-a 归档/回滚告警 + P0-b 一致性热备到 R2）
# 对应脚本 backup_db.py / 1.wooorders_sqlite.py；管理员专属
# =====================================================================
_BK_DIR = '/www/backups/woo-orders'
_BK_OFFSITE_CFG = '/www/wwwroot/woo-analysis/backup_offsite.json'
_BK_SCRIPT = '/www/wwwroot/woo-analysis/backup_db.py'
_BK_VENV_PY = '/www/wwwroot/woo-analysis/venv/bin/python'
_BK_NAME_RE = re.compile(r'^woocommerce_orders_\d{8}_\d{6}\.db\.gz$')


def _bk_local_list():
    import os, glob
    files = []
    if os.path.isdir(_BK_DIR):
        for p in glob.glob(os.path.join(_BK_DIR, 'woocommerce_orders_*.db.gz')):
            try:
                st = os.stat(p)
                files.append({'name': os.path.basename(p), 'size': st.st_size, 'mtime': st.st_mtime})
            except OSError:
                pass
    files.sort(key=lambda f: f['mtime'], reverse=True)
    return files


def _bk_offsite_cfg():
    try:
        with open(_BK_OFFSITE_CFG, encoding='utf-8') as f:
            return json.load(f)
    except Exception:
        return {}


@app.route('/api/settings/backup/status', methods=['GET'])
@login_required
@admin_required
def backup_status():
    """备份状态总览：本地份数/最近一次、异地配置、定时任务、归档订单数、未确认回滚告警"""
    import time
    files = _bk_local_list()
    latest = files[0] if files else None
    cfg = _bk_offsite_cfg()
    mode = cfg.get('mode', 'none')
    bucket = cfg.get('s3', {}).get('bucket') if mode == 's3' else None

    cron_on = False
    try:
        import subprocess
        r = subprocess.run(['/usr/bin/crontab', '-l'], capture_output=True, text=True)
        cron_on = 'backup_db.py' in (r.stdout or '')
    except Exception:
        pass

    conn = get_db_connection()
    try:
        archived = conn.execute('SELECT COUNT(*) FROM orders_archive').fetchone()[0]
    except Exception:
        archived = None
    alerts = []
    try:
        rows = conn.execute(
            "SELECT id, created_at, site_url, alert_type, detail FROM sync_alerts "
            "WHERE COALESCE(acknowledged,0)=0 ORDER BY created_at DESC LIMIT 50"
        ).fetchall()
        alerts = [dict(r) for r in rows]
    except Exception:
        pass
    conn.close()

    def _fmt(ts):
        try:
            return time.strftime('%Y-%m-%d %H:%M:%S', time.localtime(ts))
        except Exception:
            return None

    return jsonify({
        'local': {
            'count': len(files),
            'total_size': sum(f['size'] for f in files),
            'latest_name': latest['name'] if latest else None,
            'latest_time': _fmt(latest['mtime']) if latest else None,
            'latest_size': latest['size'] if latest else None,
            'dir': _BK_DIR,
        },
        'offsite': {'mode': mode, 'bucket': bucket},
        'cron_hourly': cron_on,
        'archived_orders': archived,
        'alerts': alerts,
    })


@app.route('/api/settings/backup/run', methods=['POST'])
@login_required
@admin_required
def backup_run():
    """立即执行一次热备（含异地上传），返回日志尾部"""
    import subprocess
    try:
        proc = subprocess.run(
            [_BK_VENV_PY, _BK_SCRIPT],
            cwd='/www/wwwroot/woo-analysis',
            capture_output=True, text=True, timeout=180
        )
        out = ((proc.stdout or '') + (proc.stderr or '')).strip()
        tail = '\n'.join(out.splitlines()[-12:])
        return jsonify({'success': proc.returncode == 0, 'output': tail})
    except subprocess.TimeoutExpired:
        return jsonify({'success': False, 'error': '备份超时（>180s）'}), 500
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/settings/backup/list', methods=['GET'])
@login_required
@admin_required
def backup_list():
    """本地备份文件列表（最多 50 份，最新在前）"""
    import time
    files = _bk_local_list()[:50]
    for f in files:
        f['time'] = time.strftime('%Y-%m-%d %H:%M:%S', time.localtime(f['mtime']))
    return jsonify({'files': files})


@app.route('/api/settings/backup/download', methods=['GET'])
@login_required
@admin_required
def backup_download():
    """下载一份本地备份（严格校验文件名，防路径穿越）"""
    import os
    name = request.args.get('name', '')
    if not _BK_NAME_RE.match(name):
        return jsonify({'error': '非法文件名'}), 400
    path = os.path.join(_BK_DIR, name)
    if not os.path.isfile(path):
        return jsonify({'error': '文件不存在'}), 404
    return send_file(path, as_attachment=True)


@app.route('/api/settings/backup/offsite-check', methods=['POST'])
@login_required
@admin_required
def backup_offsite_check():
    """按需列举异地 R2 桶内对象，确认异地链路正常"""
    cfg = _bk_offsite_cfg()
    if cfg.get('mode') != 's3':
        return jsonify({'success': False, 'error': '异地未配置为 S3/R2'}), 400
    s3 = cfg.get('s3', {})
    try:
        import boto3
        from botocore.config import Config as _Cfg
        client = boto3.client(
            's3', endpoint_url=s3.get('endpoint_url'),
            aws_access_key_id=s3['access_key_id'],
            aws_secret_access_key=s3['secret_access_key'],
            region_name=s3.get('region', 'auto'),
            config=_Cfg(signature_version='s3v4'),
        )
        resp = client.list_objects_v2(Bucket=s3['bucket'], Prefix=s3.get('prefix', ''))
        objs = sorted(resp.get('Contents', []), key=lambda o: o['LastModified'])
        latest = objs[-1] if objs else None
        return jsonify({
            'success': True,
            'bucket': s3['bucket'],
            'count': len(objs),
            'latest_size': latest['Size'] if latest else None,
            'latest_time': latest['LastModified'].astimezone().strftime('%Y-%m-%d %H:%M:%S') if latest else None,
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/settings/backup/alerts/ack', methods=['POST'])
@login_required
@admin_required
def backup_alert_ack():
    """确认（忽略）回滚告警：传 {id:N} 或 {all:true}"""
    data = request.json or {}
    conn = get_db_connection()
    try:
        if data.get('all'):
            conn.execute('UPDATE sync_alerts SET acknowledged=1 WHERE COALESCE(acknowledged,0)=0')
        elif data.get('id') is not None:
            conn.execute('UPDATE sync_alerts SET acknowledged=1 WHERE id=?', (int(data['id']),))
        else:
            conn.close()
            return jsonify({'success': False, 'error': '缺少 id'}), 400
        conn.commit()
    except Exception as e:
        conn.close()
        return jsonify({'success': False, 'error': str(e)}), 500
    conn.close()
    return jsonify({'success': True})


# 惰性加载并缓存 cron 同步模块（文件名带数字，无法普通 import）；
# 复用其 archive_orphaned_orders / fetch_all_remote_order_ids，让 UI 与 cron 对孤儿单的处理保持一致。
_WOOSYNC = None


def _get_woosync():
    global _WOOSYNC
    if _WOOSYNC is None:
        import importlib.util
        spec = importlib.util.spec_from_file_location("woosync", "/www/wwwroot/woo-analysis/1.wooorders_sqlite.py")
        m = importlib.util.module_from_spec(spec)
        spec.loader.exec_module(m)
        _WOOSYNC = m
    return _WOOSYNC


@app.route('/api/settings/backup/site-diff', methods=['POST'])
@login_required
@admin_required
def backup_site_diff():
    """只读：对比某站点'镜像有、线上没有'的订单（疑似该站回滚丢失的订单）。不修改任何数据。"""
    data = request.json or {}
    site_id = data.get('site_id')
    conn = get_db_connection()
    site = conn.execute('SELECT id, url, consumer_key, consumer_secret FROM sites WHERE id = ?', (site_id,)).fetchone()
    if not site:
        conn.close()
        return jsonify({'success': False, 'error': '站点不存在'}), 404
    site_url = (site['url'] or '').strip()

    # 拉取该站线上现有订单ID（与 clean 同步同一套判定）
    try:
        from woocommerce import API
        wcapi = API(url=site_url, consumer_key=site['consumer_key'], consumer_secret=site['consumer_secret'],
                    version="wc/v3", timeout=30, user_agent="WooCommerce API Client-Python/3.0.0")
        remote_ids = _get_woosync().fetch_all_remote_order_ids(wcapi, site_url)
    except Exception as e:
        conn.close()
        return jsonify({'success': False, 'error': f'获取线上订单失败：{e}'}), 502

    if not remote_ids:
        conn.close()
        return jsonify({'success': False, 'error': '未能获取该站线上订单（API 异常或确实无单），为避免误判已中止对比'}), 502

    rows = conn.execute(
        "SELECT id, number, status, date_created, total, currency, billing FROM orders "
        "WHERE source = ? ORDER BY date_created DESC", (site_url,)
    ).fetchall()
    conn.close()

    missing = []
    for r in rows:
        if str(r['id']) not in remote_ids:
            cust = ''
            try:
                b = json.loads(r['billing']) if r['billing'] else {}
                cust = ((b.get('first_name') or '') + ' ' + (b.get('last_name') or '')).strip()
            except Exception:
                pass
            missing.append({
                'id': r['id'], 'number': r['number'], 'status': r['status'],
                'date_created': r['date_created'], 'total': r['total'],
                'currency': r['currency'], 'customer': cust,
            })

    return jsonify({
        'success': True,
        'site_url': site_url,
        'mirror_count': len(rows),
        'live_count': len(remote_ids),
        'missing_count': len(missing),
        'missing': missing[:500],
    })


@app.route('/api/settings/autosync', methods=['GET'])
@login_required
def get_autosync_status():
    """Get autosync status from database and verify cron status"""
    conn = get_db_connection()
    enabled_row = conn.execute("SELECT value FROM settings WHERE key = 'autosync_enabled'").fetchone()
    interval_row = conn.execute("SELECT value FROM settings WHERE key = 'autosync_interval'").fetchone()
    conn.close()
    
    enabled = enabled_row['value'] == 'true' if enabled_row else False
    interval = int(interval_row['value']) if interval_row else 900
    
    # Verify cron job exists if enabled
    if enabled:
        import subprocess
        try:
            result = subprocess.run(['/usr/bin/crontab', '-l'], capture_output=True, text=True)
            if 'auto_sync.py' not in result.stdout:
                enabled = False  # Cron was removed externally
        except:
            pass
    
    return jsonify({
        'enabled': enabled,
        'interval': interval
    })

@app.route('/api/settings/autosync', methods=['POST'])
@login_required
def set_autosync_status():
    """Set autosync status, save to database and manage cron job"""
    import subprocess
    
    data = request.json
    conn = get_db_connection()
    
    # Get current values
    enabled_row = conn.execute("SELECT value FROM settings WHERE key = 'autosync_enabled'").fetchone()
    interval_row = conn.execute("SELECT value FROM settings WHERE key = 'autosync_interval'").fetchone()
    
    current_enabled = enabled_row['value'] == 'true' if enabled_row else False
    current_interval = int(interval_row['value']) if interval_row else 900
    
    new_enabled = bool(data.get('enabled', current_enabled))
    new_interval = current_interval
    
    if 'interval' in data:
        interval = int(data['interval'])
        if 300 <= interval <= 86400:  # 5 min to 24 hours
            new_interval = interval
    
    # Save to database
    conn.execute("INSERT OR REPLACE INTO settings (key, value) VALUES ('autosync_enabled', ?)", 
                 ('true' if new_enabled else 'false',))
    conn.execute("INSERT OR REPLACE INTO settings (key, value) VALUES ('autosync_interval', ?)", 
                 (str(new_interval),))
    conn.commit()
    conn.close()
    
    # Manage cron job
    try:
        # Get existing crontab
        result = subprocess.run(['/usr/bin/crontab', '-l'], capture_output=True, text=True)
        existing_crontab = result.stdout if result.returncode == 0 else ''
        
        # Remove any existing auto_sync.py entries
        lines = [line for line in existing_crontab.split('\n') 
                 if line.strip() and 'auto_sync.py' not in line]
        
        if new_enabled:
            # Convert interval (seconds) to cron expression
            interval_mins = new_interval // 60
            
            if interval_mins < 60:
                # Every X minutes
                cron_schedule = f"*/{interval_mins} * * * *"
            elif interval_mins < 1440:
                # Every X hours
                hours = interval_mins // 60
                cron_schedule = f"0 */{hours} * * *"
            else:
                # Once a day at 6am
                cron_schedule = "0 6 * * *"
            
            # Add new cron job
            script_path = '/www/wwwroot/woo-analysis'
            cron_line = f"{cron_schedule} cd {script_path} && {script_path}/venv/bin/python auto_sync.py >> {script_path}/auto_sync.log 2>&1"
            lines.append(cron_line)
        
        # Write new crontab
        new_crontab = '\n'.join(lines) + '\n' if lines else ''
        process = subprocess.Popen(['/usr/bin/crontab', '-'], stdin=subprocess.PIPE, text=True)
        process.communicate(input=new_crontab)
        
    except Exception as e:
        return jsonify({'success': False, 'error': str(e), 'enabled': new_enabled, 'interval': new_interval}), 500
    
    return jsonify({'success': True, 'enabled': new_enabled, 'interval': new_interval})


@app.route('/api/sync/logs')
@login_required
def get_sync_logs():
    """Get synchronization logs"""
    site_id = request.args.get('site_id', '')
    limit = int(request.args.get('limit', 50))
    
    conn = get_db_connection()
    
    if site_id:
        logs = conn.execute('''
            SELECT * FROM sync_logs 
            WHERE site_id = ? 
            ORDER BY sync_time DESC 
            LIMIT ?
        ''', (site_id, limit)).fetchall()
    else:
        logs = conn.execute('''
            SELECT * FROM sync_logs 
            ORDER BY sync_time DESC 
            LIMIT ?
        ''', (limit,)).fetchall()
    
    conn.close()
    
    return jsonify([dict(row) for row in logs])


@app.route('/api/sync/file-logs')
@login_required
def get_sync_file_logs():
    """读取同步日志文件的最近内容"""
    import os
    log_type = request.args.get('type', 'auto')
    lines_count = min(int(request.args.get('lines', 200)), 500)
    
    log_files = {
        'auto': '/www/wwwroot/woo-analysis/auto_sync.log',
        'deep': '/www/wwwroot/woo-analysis/deep_sync.log',
        'clean': '/www/wwwroot/woo-analysis/clean_sync.log',
    }
    
    log_file = log_files.get(log_type)
    if not log_file or not os.path.exists(log_file):
        return jsonify({'lines': [], 'total_lines': 0, 'file_size': 0})
    
    file_size = os.path.getsize(log_file)
    
    # 从文件末尾读取指定行数
    try:
        import subprocess
        result = subprocess.run(
            ['tail', '-n', str(lines_count), log_file],
            capture_output=True, text=True, timeout=5
        )
        content_lines = result.stdout.strip().split('\n') if result.stdout.strip() else []
        
        # 统计总行数
        wc_result = subprocess.run(
            ['wc', '-l', log_file],
            capture_output=True, text=True, timeout=5
        )
        total_lines = int(wc_result.stdout.strip().split()[0]) if wc_result.stdout.strip() else 0
        
    except Exception:
        content_lines = []
        total_lines = 0
    
    return jsonify({
        'lines': content_lines,
        'total_lines': total_lines,
        'file_size': file_size,
        'file_size_mb': round(file_size / 1024 / 1024, 2)
    })


@app.route('/api/sync/dashboard')
@login_required
def get_sync_dashboard():
    """获取全局同步状态摘要"""
    import subprocess, os
    
    conn = get_db_connection()
    
    # 1. 获取自动同步设置
    autosync_enabled_row = conn.execute("SELECT value FROM settings WHERE key = 'autosync_enabled'").fetchone()
    autosync_interval_row = conn.execute("SELECT value FROM settings WHERE key = 'autosync_interval'").fetchone()
    
    autosync_enabled = autosync_enabled_row['value'] == 'true' if autosync_enabled_row else False
    autosync_interval = int(autosync_interval_row['value']) if autosync_interval_row else 900
    
    # 2. 获取 Crontab 配置
    cron_info = {}
    try:
        result = subprocess.run(['/usr/bin/crontab', '-l'], capture_output=True, text=True)
        crontab = result.stdout if result.returncode == 0 else ''
        for line in crontab.split('\n'):
            if 'auto_sync.py' in line and not line.startswith('#'):
                parts = line.split()
                if len(parts) >= 5:
                    cron_info['auto_sync'] = ' '.join(parts[:5])
            elif '1.wooorders_sqlite.py' in line and '--clean' not in line and not line.startswith('#'):
                parts = line.split()
                if len(parts) >= 5:
                    cron_info['deep_sync'] = ' '.join(parts[:5])
            elif '1.wooorders_sqlite.py' in line and '--clean' in line and not line.startswith('#'):
                parts = line.split()
                if len(parts) >= 5:
                    cron_info['clean_sync'] = ' '.join(parts[:5])
    except Exception:
        pass
    
    # 3. 获取最近同步记录统计
    stats_24h = conn.execute('''
        SELECT 
            COUNT(*) as total,
            SUM(CASE WHEN status = 'success' THEN 1 ELSE 0 END) as success,
            SUM(CASE WHEN status != 'success' THEN 1 ELSE 0 END) as failed,
            SUM(new_orders) as new_orders,
            SUM(updated_orders) as updated_orders
        FROM sync_logs 
        WHERE sync_time >= datetime('now', '-24 hours')
    ''').fetchone()
    
    last_sync = conn.execute('''
        SELECT sync_time, status, message FROM sync_logs 
        ORDER BY sync_time DESC LIMIT 1
    ''').fetchone()
    
    # 4. 日志文件大小
    log_sizes = {}
    for name, path in [('auto', 'auto_sync.log'), ('deep', 'deep_sync.log'), ('clean', 'clean_sync.log')]:
        full_path = f'/www/wwwroot/woo-analysis/{path}'
        if os.path.exists(full_path):
            size = os.path.getsize(full_path)
            log_sizes[name] = round(size / 1024 / 1024, 2)
        else:
            log_sizes[name] = 0
    
    conn.close()
    
    return jsonify({
        'autosync': {
            'enabled': autosync_enabled,
            'interval': autosync_interval,
            'cron_schedule': cron_info.get('auto_sync'),
        },
        'deep_sync': {
            'cron_schedule': cron_info.get('deep_sync'),
        },
        'clean_sync': {
            'cron_schedule': cron_info.get('clean_sync'),
        },
        'stats_24h': dict(stats_24h) if stats_24h else {},
        'last_sync': dict(last_sync) if last_sync else None,
        'log_sizes': log_sizes
    })


@app.route('/api/sync/summary')
@login_required
def get_sync_summary():
    """Get sync summary for all sites"""
    conn = get_db_connection()
    
    # Get all sites with their latest sync info
    sites = conn.execute('SELECT * FROM sites').fetchall()
    
    summary = []
    for site in sites:
        # Get latest log for this site
        latest_log = conn.execute('''
            SELECT * FROM sync_logs 
            WHERE site_id = ? 
            ORDER BY sync_time DESC 
            LIMIT 1
        ''', (site['id'],)).fetchone()
        
        # Get stats for last 7 days
        stats = conn.execute('''
            SELECT 
                COUNT(*) as total_syncs,
                SUM(CASE WHEN status = 'success' THEN 1 ELSE 0 END) as success_count,
                SUM(new_orders) as total_new_orders,
                SUM(updated_orders) as total_updated_orders,
                AVG(duration_seconds) as avg_duration
            FROM sync_logs 
            WHERE site_id = ? AND sync_time >= datetime('now', '-7 days')
        ''', (site['id'],)).fetchone()
        
        summary.append({
            'site_id': site['id'],
            'url': site['url'],
            'last_sync': site['last_sync'],
            'latest_log': dict(latest_log) if latest_log else None,
            'stats_7_days': dict(stats) if stats else None
        })
    
    conn.close()
    
    return jsonify(summary)


@app.route('/api/sync/deep', methods=['POST'])
@login_required
def trigger_deep_sync():
    """Trigger TRUE deep sync — fetch every order page from each site, no date filter."""
    import subprocess
    import threading

    DEEP_SYNC_ID = 888888

    SYNC_STATUS[DEEP_SYNC_ID] = {
        'status': 'running',
        'message': '正在启动深度同步...',
        'logs': [f"[{datetime.now().strftime('%H:%M:%S')}] Deep sync job started"]
    }

    def run_deep_sync(app_context):
        with app_context:
            try:
                # Use the standalone full-resync script (no date filter, all pages, every site)
                script_path = '/www/wwwroot/woo-analysis/full_resync_all.py'
                venv_python = '/www/wwwroot/woo-analysis/venv/bin/python'

                SYNC_STATUS[DEEP_SYNC_ID]['message'] = '正在执行全量深度同步...'
                SYNC_STATUS[DEEP_SYNC_ID]['logs'].append(f"[{datetime.now().strftime('%H:%M:%S')}] Running {script_path}")

                result = subprocess.run(
                    [venv_python, '-u', script_path],
                    cwd='/www/wwwroot/woo-analysis',
                    capture_output=True,
                    text=True,
                    timeout=7200  # 2 hour timeout for full resync of all sites
                )
                
                if result.returncode == 0:
                    SYNC_STATUS[DEEP_SYNC_ID]['status'] = 'success'
                    SYNC_STATUS[DEEP_SYNC_ID]['message'] = '深度同步完成'
                    SYNC_STATUS[DEEP_SYNC_ID]['logs'].append(f"[{datetime.now().strftime('%H:%M:%S')}] Completed successfully")
                    # Add last few lines of output
                    output_lines = result.stdout.strip().split('\n')[-10:]
                    for line in output_lines:
                        SYNC_STATUS[DEEP_SYNC_ID]['logs'].append(line)
                else:
                    SYNC_STATUS[DEEP_SYNC_ID]['status'] = 'error'
                    SYNC_STATUS[DEEP_SYNC_ID]['message'] = '深度同步失败'
                    SYNC_STATUS[DEEP_SYNC_ID]['logs'].append(f"[{datetime.now().strftime('%H:%M:%S')}] Error: {result.stderr[:500]}")
                    
            except subprocess.TimeoutExpired:
                SYNC_STATUS[DEEP_SYNC_ID]['status'] = 'error'
                SYNC_STATUS[DEEP_SYNC_ID]['message'] = '深度同步超时'
            except Exception as e:
                SYNC_STATUS[DEEP_SYNC_ID]['status'] = 'error'
                SYNC_STATUS[DEEP_SYNC_ID]['message'] = str(e)
    
    thread = threading.Thread(target=run_deep_sync, args=(app.app_context(),))
    thread.start()
    
    return jsonify({'success': True, 'sync_id': DEEP_SYNC_ID, 'message': 'Deep sync started'})


@app.route('/api/sync/clean/all', methods=['POST'])
@login_required
def clean_all_sites():
    """Clean deleted orders from all sites"""
    import threading
    
    CLEAN_ALL_ID = 999999
    
    SYNC_STATUS[CLEAN_ALL_ID] = {
        'status': 'running',
        'message': '正在启动全站点清理同步...',
        'logs': [f"[{datetime.now().strftime('%H:%M:%S')}] Clean all sites started"]
    }
    
    def run_clean_all(app_context):
        with app_context:
            try:
                from woocommerce import API
                import time
                import random
                
                conn = get_db_connection()
                sites = conn.execute('SELECT * FROM sites').fetchall()
                conn.close()
                
                total_deleted = 0
                
                for i, site in enumerate(sites):
                    site_url = site['url'].strip() # Ensure no whitespace
                    SYNC_STATUS[CLEAN_ALL_ID]['message'] = f'正在处理 {site_url} ({i+1}/{len(sites)})...'
                    SYNC_STATUS[CLEAN_ALL_ID]['logs'].append(f"[{datetime.now().strftime('%H:%M:%S')}] Processing {site_url}")
                    
                    try:
                        wcapi = API(
                            url=site_url,
                            consumer_key=site['consumer_key'],
                            consumer_secret=site['consumer_secret'],
                            version="wc/v3",
                            timeout=30,
                            user_agent="WooCommerce API Client-Python/3.0.0" # Masquerade as official client
                        )
                        
                        # Fetch all remote order IDs
                        remote_ids = set()
                        page = 1
                        per_page = 100
                        
                        while True:
                            try:
                                time.sleep(random.uniform(0.5, 1))
                                response = wcapi.get("orders", params={
                                    "per_page": per_page,
                                    "page": page,
                                    "_fields": "id"
                                })
                                
                                if response.status_code != 200:
                                    break
                                
                                data = response.json()
                                if not data:
                                    break
                                
                                for order in data:
                                    remote_ids.add(str(order['id']))
                                
                                page += 1
                                
                            except Exception as e:
                                SYNC_STATUS[CLEAN_ALL_ID]['logs'].append(f"[{datetime.now().strftime('%H:%M:%S')}] Error fetching: {str(e)[:50]}")
                                break
                        
                        SYNC_STATUS[CLEAN_ALL_ID]['logs'].append(f"[{datetime.now().strftime('%H:%M:%S')}] Found {len(remote_ids)} remote orders")
                        
                        if not remote_ids:
                            SYNC_STATUS[CLEAN_ALL_ID]['logs'].append(f"[{datetime.now().strftime('%H:%M:%S')}] Skipping {site_url} - no remote orders")
                            continue
                        
                        # Get local order IDs
                        conn = get_db_connection()
                        local_orders = conn.execute('SELECT id FROM orders WHERE source = ?', (site_url,)).fetchall()
                        conn.close()
                        
                        local_ids = set(str(o['id']) for o in local_orders)
                        orphaned_ids = local_ids - remote_ids
                        
                        if orphaned_ids:
                            # 受保护的归档逻辑（P0-a）：大批量孤儿单（疑似回滚）会被跳过并告警，而非删除
                            removed = _get_woosync().archive_orphaned_orders(site_url, remote_ids)
                            if removed == 0:
                                SYNC_STATUS[CLEAN_ALL_ID]['logs'].append(f"[{datetime.now().strftime('%H:%M:%S')}] {site_url}: {len(orphaned_ids)} orphans, suspected rollback -> skipped & alerted (orders kept)")
                            else:
                                total_deleted += removed
                                SYNC_STATUS[CLEAN_ALL_ID]['logs'].append(f"[{datetime.now().strftime('%H:%M:%S')}] Archived & removed {removed} orphaned orders from {site_url}")
                        else:
                            SYNC_STATUS[CLEAN_ALL_ID]['logs'].append(f"[{datetime.now().strftime('%H:%M:%S')}] No orphaned orders in {site_url}")
                            
                        # Clean up draft/trash orders regardless of remote status
                        conn = get_db_connection()
                        draft_orders = conn.execute("SELECT id FROM orders WHERE source = ? AND status IN ('checkout-draft', 'trash')", (site_url,)).fetchall()
                        draft_ids = [str(o['id']) for o in draft_orders]
                        
                        if draft_ids:
                            placeholders = ','.join(['?' for _ in draft_ids])
                            conn.execute(f"DELETE FROM orders WHERE source = ? AND id IN ({placeholders})", 
                                         [site_url] + list(draft_ids))
                            conn.commit()
                            total_deleted += len(draft_ids)
                            SYNC_STATUS[CLEAN_ALL_ID]['logs'].append(f"[{datetime.now().strftime('%H:%M:%S')}] Deleted {len(draft_ids)} draft/trash orders from {site_url}")
                        conn.close()
                            
                    except Exception as e:
                        SYNC_STATUS[CLEAN_ALL_ID]['logs'].append(f"[{datetime.now().strftime('%H:%M:%S')}] Error: {str(e)[:100]}")
                
                SYNC_STATUS[CLEAN_ALL_ID]['status'] = 'success'
                SYNC_STATUS[CLEAN_ALL_ID]['message'] = f'清理完成，共删除 {total_deleted} 个订单'
                SYNC_STATUS[CLEAN_ALL_ID]['logs'].append(f"[{datetime.now().strftime('%H:%M:%S')}] Completed: {total_deleted} total deleted")
                
            except Exception as e:
                SYNC_STATUS[CLEAN_ALL_ID]['status'] = 'error'
                SYNC_STATUS[CLEAN_ALL_ID]['message'] = str(e)
                SYNC_STATUS[CLEAN_ALL_ID]['logs'].append(f"[{datetime.now().strftime('%H:%M:%S')}] Critical Error: {str(e)}")
    
    thread = threading.Thread(target=run_clean_all, args=(app.app_context(),))
    thread.start()
    
    return jsonify({'success': True, 'sync_id': CLEAN_ALL_ID, 'message': 'Clean all started'})


@app.route('/api/cron/status')
@login_required
def get_cron_status():
    """Get current cron job status for deep sync"""
    import subprocess
    
    try:
        result = subprocess.run(['/usr/bin/crontab', '-l'], capture_output=True, text=True)
        crontab_content = result.stdout
        
        # Check if our deep sync job exists
        job_pattern = '1.wooorders_sqlite.py'
        has_job = job_pattern in crontab_content
        
        # Extract schedule if exists
        schedule = None
        if has_job:
            for line in crontab_content.split('\n'):
                if job_pattern in line and not line.startswith('#'):
                    parts = line.split()
                    if len(parts) >= 5:
                        schedule = ' '.join(parts[:5])
                    break
        
        return jsonify({
            'enabled': has_job,
            'schedule': schedule,
            'raw': crontab_content if has_job else None
        })
    except Exception as e:
        return jsonify({'enabled': False, 'error': str(e)})


@app.route('/api/cron/setup', methods=['POST'])
@login_required
def setup_cron():
    """Setup cron job for deep sync"""
    import subprocess
    
    data = request.json
    hour = int(data.get('hour', 3))  # Default 3 AM
    minute = int(data.get('minute', 0))
    
    if not (0 <= hour <= 23 and 0 <= minute <= 59):
        return jsonify({'error': 'Invalid time'}), 400
    
    try:
        # Get existing crontab
        result = subprocess.run(['/usr/bin/crontab', '-l'], capture_output=True, text=True)
        existing_crontab = result.stdout if result.returncode == 0 else ''
        
        # Remove existing deep sync job if any
        lines = [line for line in existing_crontab.split('\n') 
                 if '1.wooorders_sqlite.py' not in line and line.strip()]
        
        # Add new job
        new_job = f"{minute} {hour} * * * cd /www/wwwroot/woo-analysis && /www/wwwroot/woo-analysis/venv/bin/python 1.wooorders_sqlite.py >> /www/wwwroot/woo-analysis/deep_sync.log 2>&1"
        lines.append(new_job)
        
        # Write new crontab
        new_crontab = '\n'.join(lines) + '\n'
        process = subprocess.Popen(['/usr/bin/crontab', '-'], stdin=subprocess.PIPE, text=True)
        process.communicate(input=new_crontab)
        
        if process.returncode == 0:
            return jsonify({'success': True, 'message': f'Cron job set for {hour:02d}:{minute:02d}'})
        else:
            return jsonify({'error': 'Failed to set crontab'}), 500
            
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/cron/remove', methods=['DELETE'])
@login_required
def remove_cron():
    """Remove cron job for deep sync"""
    import subprocess
    
    try:
        result = subprocess.run(['/usr/bin/crontab', '-l'], capture_output=True, text=True)
        existing_crontab = result.stdout if result.returncode == 0 else ''
        
        # Remove deep sync job
        lines = [line for line in existing_crontab.split('\n') 
                 if '1.wooorders_sqlite.py' not in line and line.strip()]
        
        new_crontab = '\n'.join(lines) + '\n' if lines else ''
        process = subprocess.Popen(['/usr/bin/crontab', '-'], stdin=subprocess.PIPE, text=True)
        process.communicate(input=new_crontab)
        
        return jsonify({'success': True, 'message': 'Cron job removed'})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/cron/clean/status')
@login_required
def get_clean_cron_status():
    """Get status of clean sync cron job"""
    import subprocess
    
    try:
        result = subprocess.run(['/usr/bin/crontab', '-l'], capture_output=True, text=True)
        if result.returncode != 0:
            return jsonify({'enabled': False})
        
        crontab = result.stdout
        # Look for clean sync job (python 1.wooorders_sqlite.py --clean or curl clean/all)
        for line in crontab.split('\n'):
            if '--clean' in line or 'sync/clean/all' in line:
                # Parse cron schedule
                parts = line.split()
                if len(parts) >= 5:
                    minute = parts[0]
                    hour = parts[1]
                    day_of_week = parts[4]
                    days = ['周日', '周一', '周二', '周三', '周四', '周五', '周六']
                    day_name = days[int(day_of_week)] if day_of_week.isdigit() and 0 <= int(day_of_week) <= 6 else day_of_week
                    return jsonify({
                        'enabled': True,
                        'schedule': f'{day_name} {hour}:{minute.zfill(2)}',
                        'hour': int(hour) if hour.isdigit() else 4,
                        'day': int(day_of_week) if day_of_week.isdigit() else 0
                    })
        
        return jsonify({'enabled': False})
    except Exception as e:
        return jsonify({'enabled': False, 'error': str(e)})


@app.route('/api/cron/clean/setup', methods=['POST'])
@login_required
def setup_clean_cron():
    """Setup cron job for clean sync"""
    import subprocess
    
    data = request.json
    hour = int(data.get('hour', 4))  # Default 4 AM
    day = int(data.get('day', 0))    # Default Sunday
    minute = int(data.get('minute', 0))
    
    if not (0 <= hour <= 23 and 0 <= minute <= 59 and 0 <= day <= 6):
        return jsonify({'error': 'Invalid time or day'}), 400
    
    try:
        # Get existing crontab
        result = subprocess.run(['/usr/bin/crontab', '-l'], capture_output=True, text=True)
        existing_crontab = result.stdout if result.returncode == 0 else ''
        
        # Remove existing clean sync job if any
        lines = [line for line in existing_crontab.split('\n') 
                 if '--clean' not in line and 'sync/clean/all' not in line and line.strip()]
        
        # Add new job (runs weekly on specified day)
        new_job = f"{minute} {hour} * * {day} cd /www/wwwroot/woo-analysis && /www/wwwroot/woo-analysis/venv/bin/python 1.wooorders_sqlite.py --clean >> /www/wwwroot/woo-analysis/clean_sync.log 2>&1"
        lines.append(new_job)
        
        # Write new crontab
        new_crontab = '\n'.join(lines) + '\n'
        process = subprocess.Popen(['/usr/bin/crontab', '-'], stdin=subprocess.PIPE, text=True)
        process.communicate(input=new_crontab)
        
        days = ['周日', '周一', '周二', '周三', '周四', '周五', '周六']
        if process.returncode == 0:
            return jsonify({'success': True, 'message': f'Clean sync cron set for {days[day]} {hour:02d}:{minute:02d}'})
        else:
            return jsonify({'error': 'Failed to set crontab'}), 500
            
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/cron/clean/remove', methods=['DELETE'])
@login_required
def remove_clean_cron():
    """Remove cron job for clean sync"""
    import subprocess
    
    try:
        result = subprocess.run(['/usr/bin/crontab', '-l'], capture_output=True, text=True)
        existing_crontab = result.stdout if result.returncode == 0 else ''
        
        # Remove clean sync job
        lines = [line for line in existing_crontab.split('\n') 
                 if '--clean' not in line and 'sync/clean/all' not in line and line.strip()]
        
        new_crontab = '\n'.join(lines) + '\n' if lines else ''
        process = subprocess.Popen(['/usr/bin/crontab', '-'], stdin=subprocess.PIPE, text=True)
        process.communicate(input=new_crontab)
        
        return jsonify({'success': True, 'message': 'Clean sync cron job removed'})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


# ============== USER MANAGEMENT API ==============


@app.route('/users')
@login_required
@user_manager_required
def users_page():
    """User management page"""
    return render_template('users.html', is_super_admin=(current_user.username == 'admin'))


@app.route('/api/users')
@login_required
@user_manager_required
def get_users():
    """Get all users"""
    conn = get_db_connection()
    # We try progressively older schemas — newest column set first, then fall
    # back if columns are missing. The init_*_tables() functions always add
    # them on startup, so in practice the first SELECT succeeds.
    SCHEMAS = [
        # 0: latest — incl. 库存权限 can_view_inventory / can_manage_inventory
        'SELECT id, username, name, role, can_ship, can_view_report, can_view_sales_board, can_manage_users, can_view_reconciliation, can_edit_reconciliation, can_manage_products, can_view_costs, can_edit_costs, can_view_own_sales_board, can_view_shipping, can_manage_blocklist, can_view_inventory, can_manage_inventory, created_at FROM users',
        # 0b: 库存列尚未迁移时回退(其余同上)
        'SELECT id, username, name, role, can_ship, can_view_report, can_view_sales_board, can_manage_users, can_view_reconciliation, can_edit_reconciliation, can_manage_products, can_view_costs, can_edit_costs, can_view_own_sales_board, can_view_shipping, can_manage_blocklist, 0 as can_view_inventory, 0 as can_manage_inventory, created_at FROM users',
        # 1: missing can_edit_costs
        'SELECT id, username, name, role, can_ship, can_view_report, can_view_sales_board, can_manage_users, can_view_reconciliation, can_edit_reconciliation, can_manage_products, can_view_costs, 0 as can_edit_costs, created_at FROM users',
        # 2: missing can_view_costs
        'SELECT id, username, name, role, can_ship, can_view_report, can_view_sales_board, can_manage_users, can_view_reconciliation, can_edit_reconciliation, can_manage_products, 0 as can_view_costs, 0 as can_edit_costs, created_at FROM users',
        # 3: missing can_manage_products
        'SELECT id, username, name, role, can_ship, can_view_report, can_view_sales_board, can_manage_users, can_view_reconciliation, can_edit_reconciliation, 0 as can_manage_products, 0 as can_view_costs, 0 as can_edit_costs, created_at FROM users',
        # 4: missing can_edit_reconciliation
        'SELECT id, username, name, role, can_ship, can_view_report, can_view_sales_board, can_manage_users, can_view_reconciliation, 0 as can_edit_reconciliation, 0 as can_manage_products, 0 as can_view_costs, 0 as can_edit_costs, created_at FROM users',
        # 5: missing can_view_reconciliation
        'SELECT id, username, name, role, can_ship, can_view_report, can_view_sales_board, can_manage_users, 0 as can_view_reconciliation, 0 as can_edit_reconciliation, 0 as can_manage_products, 0 as can_view_costs, 0 as can_edit_costs, created_at FROM users',
        # 6: missing can_manage_users
        'SELECT id, username, name, role, can_ship, can_view_report, can_view_sales_board, 0 as can_manage_users, 0 as can_view_reconciliation, 0 as can_edit_reconciliation, 0 as can_manage_products, 0 as can_view_costs, 0 as can_edit_costs, created_at FROM users',
        # 7: oldest — only can_ship
        'SELECT id, username, name, role, can_ship, 0 as can_view_report, 0 as can_view_sales_board, 0 as can_manage_users, 0 as can_view_reconciliation, 0 as can_edit_reconciliation, 0 as can_manage_products, 0 as can_view_costs, 0 as can_edit_costs, created_at FROM users',
    ]
    users = None
    for sql in SCHEMAS:
        try:
            users = conn.execute(sql).fetchall()
            break
        except sqlite3.OperationalError:
            continue
    if users is None:
        users = []
    # Per-user site count — sites are linked to a person via sites.manager == name.
    # Drives the 销售看板(仅本人站点) toggle: a user with 0 sites can't be granted it.
    site_counts = {r['manager']: r['c'] for r in conn.execute(
        "SELECT manager, COUNT(*) c FROM sites WHERE manager IS NOT NULL AND manager != '' GROUP BY manager"
    ).fetchall()}
    conn.close()
    is_super_admin = (current_user.username == 'admin')
    result = []
    for row in users:
        u = dict(row)
        u.setdefault('can_view_own_sales_board', 0)
        u.setdefault('can_view_shipping', 0)
        u.setdefault('can_manage_blocklist', 0)
        u.setdefault('can_view_inventory', 0)
        u.setdefault('can_manage_inventory', 0)
        u['site_count'] = site_counts.get(u.get('name') or '', 0)
        # Mark users that the current operator cannot modify
        u['is_super_admin'] = (u['username'] == 'admin')
        u['is_protected'] = u['is_super_admin'] or (u['role'] == 'admin' and not is_super_admin)
        result.append(u)
    return jsonify(result)


@app.route('/api/users', methods=['POST'])
@login_required
@user_manager_required
def add_user():
    """Add a new user"""
    data = request.json
    username = data.get('username', '').strip()
    password = data.get('password', '')
    name = data.get('name', '').strip()
    role = data.get('role', 'user')

    if not username or not password:
        return jsonify({'error': 'Username and password required'}), 400

    if role not in ['admin', 'user', 'viewer']:
        role = 'user'

    # Only super admin can create admin-role users
    if role == 'admin' and current_user.username != 'admin':
        return jsonify({'error': '无权创建管理员账户，请联系超级管理员'}), 403

    conn = get_db_connection()
    try:
        conn.execute('''
            INSERT INTO users (username, password_hash, name, role)
            VALUES (?, ?, ?, ?)
        ''', (username, generate_password_hash(password), name, role))
        conn.commit()
        return jsonify({'success': True})
    except sqlite3.IntegrityError:
        return jsonify({'error': 'Username already exists'}), 400
    finally:
        conn.close()


@app.route('/api/users/<int:user_id>', methods=['PUT'])
@login_required
@user_manager_required
def update_user(user_id):
    """Update a user"""
    data = request.json
    name = data.get('name', '').strip()
    role = data.get('role', 'user')
    password = data.get('password', '')
    can_ship = data.get('can_ship', 0)
    can_view_report = data.get('can_view_report', 0)
    can_view_sales_board = data.get('can_view_sales_board', 0)
    can_view_own_sales_board = data.get('can_view_own_sales_board', 0)
    can_manage_products = 1 if data.get('can_manage_products') else 0
    # Shipping perm pair (operator-level, like can_ship): "ship" implies "view".
    can_view_shipping_val = 1 if (data.get('can_view_shipping') or can_ship) else 0

    if role not in ['admin', 'user', 'viewer']:
        role = 'user'

    conn = get_db_connection()
    try:
        # Look up the target user to enforce hierarchy
        target_user = conn.execute('SELECT username, role, name FROM users WHERE id = ?', (user_id,)).fetchone()
        if not target_user:
            return jsonify({'error': '用户不存在'}), 404

        is_super_admin = (current_user.username == 'admin')
        target_is_super_admin = (target_user['username'] == 'admin')
        target_is_admin = (target_user['role'] == 'admin')

        # Rule 1: Super admin account can ONLY be modified by itself
        if target_is_super_admin and not is_super_admin:
            return jsonify({'error': '无权修改超级管理员账户'}), 403

        # Rule 2: Non-super-admin cannot modify admin-role users
        if target_is_admin and not is_super_admin:
            return jsonify({'error': '无权修改管理员账户，请联系超级管理员'}), 403

        # Rule 3: Non-super-admin cannot promote users to admin role
        if role == 'admin' and not is_super_admin:
            return jsonify({'error': '无权设置管理员角色，请联系超级管理员'}), 403

        # Rule 4: Super admin's role cannot be downgraded
        if target_is_super_admin and role != 'admin':
            return jsonify({'error': '超级管理员角色不可更改'}), 403

        # Only the 'admin' superuser can grant/revoke can_manage_users,
        # can_view_reconciliation, can_edit_reconciliation, can_view_costs, can_edit_costs
        if is_super_admin:
            can_manage_users_val = 1 if data.get('can_manage_users') else 0
            can_view_reconciliation_val = 1 if data.get('can_view_reconciliation') else 0
            can_edit_reconciliation_val = 1 if data.get('can_edit_reconciliation') else 0
            can_view_costs_val = 1 if data.get('can_view_costs') else 0
            can_edit_costs_val = 1 if data.get('can_edit_costs') else 0
            # Edit permission requires view permission (can't edit what you can't see)
            if can_edit_reconciliation_val and not can_view_reconciliation_val:
                can_view_reconciliation_val = 1
            if can_edit_costs_val and not can_view_costs_val:
                can_view_costs_val = 1
            # 销售看板（仅本人站点）: only grantable when the user actually has sites
            # (sites.manager == their name). Server-side guard behind the UI gray-out.
            can_view_own_sales_board_val = 1 if data.get('can_view_own_sales_board') else 0
            if can_view_own_sales_board_val and not conn.execute(
                    'SELECT 1 FROM sites WHERE manager = ? LIMIT 1', (name,)).fetchone():
                can_view_own_sales_board_val = 0
            can_manage_blocklist_val = 1 if data.get('can_manage_blocklist') else 0
            # 库存权限:操作含查看(不能操作看不到的东西)
            can_view_inventory_val = 1 if data.get('can_view_inventory') else 0
            can_manage_inventory_val = 1 if data.get('can_manage_inventory') else 0
            if can_manage_inventory_val and not can_view_inventory_val:
                can_view_inventory_val = 1
            if password:
                conn.execute('UPDATE users SET name=?, role=?, can_ship=?, can_view_shipping=?, can_view_report=?, can_view_sales_board=?, can_view_own_sales_board=?, can_manage_users=?, can_view_reconciliation=?, can_edit_reconciliation=?, can_manage_products=?, can_view_costs=?, can_edit_costs=?, can_manage_blocklist=?, can_view_inventory=?, can_manage_inventory=?, password_hash=? WHERE id=?',
                            (name, role, can_ship, can_view_shipping_val, can_view_report, can_view_sales_board, can_view_own_sales_board_val, can_manage_users_val, can_view_reconciliation_val, can_edit_reconciliation_val, can_manage_products, can_view_costs_val, can_edit_costs_val, can_manage_blocklist_val, can_view_inventory_val, can_manage_inventory_val, generate_password_hash(password), user_id))
            else:
                conn.execute('UPDATE users SET name=?, role=?, can_ship=?, can_view_shipping=?, can_view_report=?, can_view_sales_board=?, can_view_own_sales_board=?, can_manage_users=?, can_view_reconciliation=?, can_edit_reconciliation=?, can_manage_products=?, can_view_costs=?, can_edit_costs=?, can_manage_blocklist=?, can_view_inventory=?, can_manage_inventory=? WHERE id=?',
                            (name, role, can_ship, can_view_shipping_val, can_view_report, can_view_sales_board, can_view_own_sales_board_val, can_manage_users_val, can_view_reconciliation_val, can_edit_reconciliation_val, can_manage_products, can_view_costs_val, can_edit_costs_val, can_manage_blocklist_val, can_view_inventory_val, can_manage_inventory_val, user_id))
        else:
            # Non-superadmin: cannot change can_manage_users, can_view_sales_board, or set role to admin
            # but CAN grant can_manage_products (a regular operator-level permission)
            if password:
                conn.execute('UPDATE users SET name=?, role=?, can_ship=?, can_view_shipping=?, can_view_report=?, can_manage_products=?, password_hash=? WHERE id=?',
                            (name, role, can_ship, can_view_shipping_val, can_view_report, can_manage_products, generate_password_hash(password), user_id))
            else:
                conn.execute('UPDATE users SET name=?, role=?, can_ship=?, can_view_shipping=?, can_view_report=?, can_manage_products=? WHERE id=?',
                            (name, role, can_ship, can_view_shipping_val, can_view_report, can_manage_products, user_id))
        conn.commit()
        return jsonify({'success': True})
    finally:
        conn.close()


@app.route('/api/users/<int:user_id>', methods=['DELETE'])
@login_required
@user_manager_required
def delete_user(user_id):
    """Delete a user"""
    # Prevent deleting yourself
    if str(user_id) == str(current_user.id):
        return jsonify({'error': '不能删除自己的账户'}), 400

    conn = get_db_connection()
    try:
        target_user = conn.execute('SELECT username, role FROM users WHERE id = ?', (user_id,)).fetchone()
        if not target_user:
            return jsonify({'error': '用户不存在'}), 404

        # Super admin account can never be deleted
        if target_user['username'] == 'admin':
            return jsonify({'error': '超级管理员账户不可删除'}), 403

        # Non-super-admin cannot delete admin-role users
        if target_user['role'] == 'admin' and current_user.username != 'admin':
            return jsonify({'error': '无权删除管理员账户，请联系超级管理员'}), 403

        conn.execute('DELETE FROM user_site_permissions WHERE user_id = ?', (user_id,))
        conn.execute('DELETE FROM users WHERE id = ?', (user_id,))
        conn.commit()
        return jsonify({'success': True})
    finally:
        conn.close()


@app.route('/api/users/change-password', methods=['POST'])
@login_required
def change_own_password():
    """Change current user's password"""
    data = request.json
    old_password = data.get('old_password', '')
    new_password = data.get('new_password', '')
    
    if not old_password or not new_password:
        return jsonify({'error': 'Both old and new password required'}), 400
    
    conn = get_db_connection()
    user = conn.execute('SELECT * FROM users WHERE id = ?', (current_user.id,)).fetchone()
    
    if not check_password_hash(user['password_hash'], old_password):
        conn.close()
        return jsonify({'error': 'Incorrect old password'}), 400
    
    conn.execute('UPDATE users SET password_hash = ? WHERE id = ?',
                (generate_password_hash(new_password), current_user.id))
    conn.commit()
    conn.close()
    
    return jsonify({'success': True, 'message': 'Password changed successfully'})


@app.route('/api/users/<int:user_id>/permissions')
@login_required
@user_manager_required
def get_user_permissions(user_id):
    """Get user's site permissions (site-level grants + country grants + exclusions)."""
    conn = get_db_connection()
    explicit = [r['site_id'] for r in conn.execute(
        'SELECT site_id FROM user_site_permissions WHERE user_id = ?', (user_id,)).fetchall()]
    countries = [r['country'] for r in conn.execute(
        'SELECT country FROM user_country_permissions WHERE user_id = ?', (user_id,)).fetchall()]
    exclusions = [r['site_id'] for r in conn.execute(
        'SELECT site_id FROM user_site_exclusions WHERE user_id = ?', (user_id,)).fetchall()]
    sites = conn.execute('SELECT id, url, country, manager FROM sites ORDER BY country, url').fetchall()
    urow = conn.execute('SELECT name FROM users WHERE id = ?', (user_id,)).fetchone()
    conn.close()

    return jsonify({
        'allowed_sites': explicit,          # explicitly-granted single sites (non-granted countries)
        'granted_countries': countries,     # fully-granted countries (auto-inherit new sites)
        'excluded_sites': exclusions,       # sites carved out of a country grant
        'user_name': urow['name'] if urow else '',   # to flag sites this user MANAGES (auto-editable)
        'all_sites': [dict(s) for s in sites]
    })


@app.route('/api/users/<int:user_id>/permissions', methods=['PUT'])
@login_required
@user_manager_required
def update_user_permissions(user_id):
    """Update user's site permissions: explicit single-site grants, country grants
    (auto-inherit future sites), and per-site exclusions carved out of a grant."""
    data = request.json
    site_ids = data.get('site_ids', []) or []
    country_grants = data.get('country_grants', []) or []
    site_exclusions = data.get('site_exclusions', []) or []

    conn = get_db_connection()
    try:
        # Replace all three sets atomically
        conn.execute('DELETE FROM user_site_permissions WHERE user_id = ?', (user_id,))
        conn.execute('DELETE FROM user_country_permissions WHERE user_id = ?', (user_id,))
        conn.execute('DELETE FROM user_site_exclusions WHERE user_id = ?', (user_id,))
        for sid in site_ids:
            conn.execute('INSERT OR IGNORE INTO user_site_permissions (user_id, site_id) VALUES (?, ?)', (user_id, sid))
        for c in country_grants:
            conn.execute('INSERT OR IGNORE INTO user_country_permissions (user_id, country) VALUES (?, ?)', (user_id, c))
        for sid in site_exclusions:
            conn.execute('INSERT OR IGNORE INTO user_site_exclusions (user_id, site_id) VALUES (?, ?)', (user_id, sid))
        conn.commit()
    finally:
        conn.close()

    return jsonify({'success': True})


# ============== PARTNER RECONCILIATION ==============

def reconciliation_viewer_required(f):
    """Decorator: access partner reconciliation page (super admin, internal viewer, or partner member)"""
    from functools import wraps
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not current_user.is_authenticated:
            return redirect(url_for('login'))
        if not current_user.can_view_reconciliation():
            return render_template_string('''
<!DOCTYPE html>
<html lang="zh"><head><meta charset="UTF-8"><title>权限不足</title>
<link href="https://cdn.jsdelivr.net/npm/bootstrap@5.3.0/dist/css/bootstrap.min.css" rel="stylesheet">
<link href="https://cdn.jsdelivr.net/npm/bootstrap-icons@1.10.0/font/bootstrap-icons.css" rel="stylesheet">
<style>
body{background:#0f172a;color:#e2e8f0;display:flex;align-items:center;justify-content:center;min-height:100vh;margin:0;}
.box{background:#1e293b;border:1px solid #334155;border-radius:16px;padding:48px;text-align:center;max-width:440px;}
.icon{font-size:56px;color:#f59e0b;margin-bottom:20px;}
h1{font-size:22px;font-weight:600;margin-bottom:12px;}
p{color:#94a3b8;margin-bottom:24px;}
a{background:#3b82f6;color:#fff;padding:10px 24px;border-radius:8px;text-decoration:none;}
</style></head>
<body><div class="box">
<div class="icon"><i class="bi bi-shield-exclamation"></i></div>
<h1>权限不足</h1>
<p>合伙人对账功能需要授权，请联系超级管理员开通。</p>
<a href="/"><i class="bi bi-house"></i> 返回首页</a>
</div></body></html>
            '''), 403
        return f(*args, **kwargs)
    return decorated_function


def reconciliation_api_required(f):
    """Decorator: same check but returns JSON 403 for APIs"""
    from functools import wraps
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not current_user.is_authenticated:
            return jsonify({'error': '未登录'}), 401
        if not current_user.can_view_reconciliation():
            return jsonify({'error': '无对账权限'}), 403
        return f(*args, **kwargs)
    return decorated_function


def _is_reconciliation_admin():
    """Check if current user can manage (not just view) reconciliation data.
    Requires can_edit_reconciliation flag (or being super admin).
    View-only users (like partner members 金谷/金毅) will return False here."""
    return current_user.can_edit_reconciliation()


# ============================================================================
# TIME-AWARE PRODUCT COST LOOKUP
# ============================================================================
# product_costs has effective_date — costs change over time. Same product can
# have 2/28 entry @ 28 PLN and 3/15 entry @ 30 PLN. When valuing an order made
# on 3/20, we need the 30 PLN row (latest effective_date <= order_date).

def _build_dated_cost_index(conn=None, only_warehouse_ids=None):
    """Load all product_costs rows into a structure that supports
    'cost at order date' lookups.

    Returns a dict mapping
      key = (brand_id, series_id_or_0, puff_count_or_0, flavor_or_blank, warehouse_id)
      value = list of {effective_date, price, currency} sorted by effective_date DESC

    Plus a coarse-match index keyed by (warehouse_id, brand_id, series_id, puff_count)
    for the 'brand+puffs+series' fallback (drops flavor).

    Plus an even coarser brand-only fallback by (warehouse_id, brand_id).

    Caller passes a connection if reusing one, or None to open a fresh one.
    """
    own_conn = conn is None
    if own_conn:
        conn = get_db_connection()
    try:
        sql = '''SELECT brand_id, series_id, puff_count, flavor, warehouse_id,
                        cost_price, cost_currency, effective_date
                 FROM product_costs'''
        params = []
        if only_warehouse_ids:
            placeholders = ','.join(['?'] * len(only_warehouse_ids))
            sql += f' WHERE warehouse_id IN ({placeholders})'
            params = list(only_warehouse_ids)
        sql += ' ORDER BY effective_date DESC, id DESC'
        rows = conn.execute(sql, params).fetchall()
    finally:
        if own_conn:
            conn.close()

    exact = {}            # (b, s, p, f, w) -> [{date, price, currency}, ...]
    by_bspw = {}          # (w, b, s, p) -> [...]
    by_bpw = {}           # (w, b, p) -> [...]    (drop series + flavor)
    by_bw = {}            # (w, b) -> [...]       (brand only)
    for r in rows:
        b = r['brand_id']
        s = r['series_id']
        p = r['puff_count']
        f = (r['flavor'] or '').strip() or None
        w = r['warehouse_id']
        entry = {
            'effective_date': r['effective_date'] or '0000-00-00',
            'price': float(r['cost_price'] or 0),
            'currency': r['cost_currency'] or 'PLN',
        }
        exact.setdefault((b, s, p, f, w), []).append(entry)
        by_bspw.setdefault((w, b, s, p), []).append(entry)
        by_bpw.setdefault((w, b, p), []).append(entry)
        by_bw.setdefault((w, b), []).append(entry)
    return {'exact': exact, 'bspw': by_bspw, 'bpw': by_bpw, 'bw': by_bw}


def _cost_at_date(idx, brand_id, series_id, puff_count, flavor, warehouse_id, order_date):
    """Look up cost effective on a given order_date. Returns dict
    {price, currency, effective_date, match_level} or None.

    Tries match levels in order:
      1. exact: brand+series+puffs+flavor+warehouse
      2. bspw : brand+series+puffs+warehouse (drop flavor)
      3. bpw  : brand+puffs+warehouse        (drop series)
      4. bw   : brand+warehouse              (any puffs)

    Within a level, picks the entry with the largest effective_date <= order_date.
    """
    if not brand_id or not warehouse_id:
        return None
    flavor_key = (flavor or '').strip() or None

    candidates = [
        ('exact', idx['exact'].get((brand_id, series_id, puff_count, flavor_key, warehouse_id))),
        ('bspw',  idx['bspw'].get((warehouse_id, brand_id, series_id, puff_count))),
        ('bpw',   idx['bpw'].get((warehouse_id, brand_id, puff_count))),
        ('bw',    idx['bw'].get((warehouse_id, brand_id))),
    ]
    od = order_date or '9999-99-99'  # if no date, treat as "now" — pick latest cost
    for level, entries in candidates:
        if not entries:
            continue
        # entries already sorted by effective_date DESC; pick first <= od
        for e in entries:
            if e['effective_date'] <= od:
                return {
                    'price': e['price'],
                    'currency': e['currency'],
                    'effective_date': e['effective_date'],
                    'match_level': level,
                }
        # If all entries are AFTER order_date (e.g. cost was first set 2026-04
        # but order is from 2025-12), fall back to oldest known cost as a best
        # effort and tag it accordingly so the UI can warn.
        oldest = entries[-1]
        return {
            'price': oldest['price'],
            'currency': oldest['currency'],
            'effective_date': oldest['effective_date'],
            'match_level': level + '_future',  # cost record is newer than order
        }
    return None


def _resolve_product_to_brand(product_name, source, brands_cache, product_mappings_cache):
    """Resolve a product line item to (brand_id, series_id, puff_count, flavor).

    First tries product_mappings (manual override), then falls back to
    parse_product_name. Returns None for missing fields if unresolvable."""
    pn_key = normalize_raw_name(product_name)
    pm = (product_mappings_cache.get((pn_key, source))
          or product_mappings_cache.get((pn_key, ''))
          or product_mappings_cache.get((pn_key, None)))
    if pm:
        return pm['brand_id'], pm['series_id'], pm['puff_count'], pm['flavor']
    parsed = parse_product_name(product_name, brands_cache)
    b_id = None
    if parsed.get('brand'):
        for bc in brands_cache:
            if bc['name'].upper() == parsed['brand'].upper():
                b_id = bc['id']
                break
    return b_id, None, parsed.get('puffs'), parsed.get('flavor')


def _calc_partner_net_sales(partner_id, year, month):
    """Aggregate net sales for all sites bound to partner in given month.
    Uses the partner's configured currency to filter orders.
    Net amount = success_amount − success_shipping − undelivered_shipping_loss
    (matches /monthly + /orders summary).
    Field names use _pln suffix for historical reasons but values are in partner's native currency."""
    conn = get_db_connection()
    # Get partner currency
    p = conn.execute('SELECT currency FROM partners WHERE id = ?', (partner_id,)).fetchone()
    currency = (p['currency'] if p and p['currency'] else 'PLN')

    # Get bound site URLs
    sites = conn.execute('''
        SELECT s.url FROM partner_sites ps
        JOIN sites s ON s.id = ps.site_id
        WHERE ps.partner_id = ?
    ''', (partner_id,)).fetchall()
    if not sites:
        conn.close()
        return {'total_orders': 0, 'total_gross_pln': 0, 'total_net_pln': 0,
                'shipping_loss': 0, 'undelivered_orders': 0, 'currency': currency}

    site_urls = [s['url'] for s in sites]
    placeholders = ','.join(['?'] * len(site_urls))
    month_str = f'{year:04d}-{month:02d}'

    # Aggregate from orders table — filtered by partner's currency.
    # Same "success" filter as /monthly. Net subtracts shipping_loss for
    # undelivered orders so the partner's settlement reflects real revenue.
    success_cond = _revenue_status_cond()
    query = f'''
        SELECT
            COUNT(*) as total_orders,
            COALESCE(SUM(CASE WHEN {success_cond} THEN total ELSE 0 END), 0) as gross,
            COALESCE(SUM(CASE WHEN {success_cond} THEN total ELSE 0 END), 0)
              - COALESCE(SUM(CASE WHEN {success_cond} THEN shipping_total ELSE 0 END), 0)
              - COALESCE(SUM(CASE WHEN COALESCE(is_undelivered, 0) = 1
                                   THEN COALESCE(shipping_loss_amount, 0) ELSE 0 END), 0) as net,
            COUNT(CASE WHEN {success_cond} THEN 1 END) as success_orders,
            COALESCE(SUM(CASE WHEN COALESCE(is_undelivered, 0) = 1
                              THEN COALESCE(shipping_loss_amount, 0) ELSE 0 END), 0) as shipping_loss,
            SUM(CASE WHEN COALESCE(is_undelivered, 0) = 1 THEN 1 ELSE 0 END) as undelivered_orders
        FROM orders
        WHERE source IN ({placeholders})
        AND currency = ?
        AND strftime('%Y-%m', date_created) = ?
    '''
    params = site_urls + [currency, month_str]
    row = conn.execute(query, params).fetchone()
    conn.close()
    return {
        'total_orders': row['success_orders'] or 0,
        'total_gross_pln': round(row['gross'] or 0, 2),
        'total_net_pln': round(row['net'] or 0, 2),
        'shipping_loss': round(row['shipping_loss'] or 0, 2),
        'undelivered_orders': row['undelivered_orders'] or 0,
        'currency': currency,
    }


def _calc_partner_recon_detail(partner_id, year, month, site_filter=None, manager_filter=None):
    """Comprehensive monthly aggregation for the partner reconciliation drill-down.

    Returns the same numbers as _calc_partner_net_sales (so existing callers stay
    consistent) plus per-status breakdown, per-site breakdown, per-product
    breakdown using DATE-AWARE actual product costs. Everything is in the
    partner's native currency unless explicitly suffixed.

    Optional filters (used by 产品明细 tab to slice by site or manager):
      site_filter    — single site URL; only that site's orders contribute
      manager_filter — manager name; only sites with that manager contribute
    Both are AND-applied. None means no filter.
    """
    conn = get_db_connection()
    try:
        partner = conn.execute(
            'SELECT id, name, currency, cost_ratio, partner_profit_ratio, our_profit_ratio FROM partners WHERE id = ?',
            (partner_id,)
        ).fetchone()
        if not partner:
            return None
        currency = partner['currency'] or 'PLN'

        # Bound site URLs (and ids -> for warehouse mapping)
        sites = conn.execute('''
            SELECT s.id, s.url, s.country, s.manager
            FROM partner_sites ps
            JOIN sites s ON s.id = ps.site_id
            WHERE ps.partner_id = ?
        ''', (partner_id,)).fetchall()
        # Apply optional site / manager filters BEFORE building site_urls.
        # Both checks done on the partner-bound subset, so users can never
        # see data for sites outside their partner scope.
        if site_filter:
            sites = [s for s in sites if s['url'] == site_filter]
        if manager_filter:
            sites = [s for s in sites if (s['manager'] or '') == manager_filter]
        if not sites:
            return _empty_recon_detail(partner, currency, year, month)

        site_urls = [s['url'] for s in sites]
        site_meta = {s['url']: {'country': s['country'] or '', 'manager': s['manager'] or '', 'id': s['id']} for s in sites}
        placeholders = ','.join(['?'] * len(site_urls))
        month_str = f'{year:04d}-{month:02d}'

        # Which of these partner sites treat on-hold as "shipped" (PL by default).
        on_hold_shipped_sources = {
            r['url'] for r in conn.execute(
                f"SELECT url FROM sites WHERE cod_on_hold_is_shipped = 1 AND url IN ({placeholders})",
                site_urls,
            ).fetchall()
        }

        # Pull every order in scope (partner sites + currency + month). We keep
        # ALL statuses so we can break them out (success / failed / cancelled /
        # undelivered / pending). Loaders downstream filter for revenue.
        all_orders = conn.execute(f'''
            SELECT id, number, status, payment_method, total, shipping_total, currency,
                   line_items, source, date_created, warehouse_id,
                   is_undelivered, shipping_loss_amount,
                   billing
            FROM orders
            WHERE source IN ({placeholders})
            AND currency = ?
            AND strftime('%Y-%m', date_created) = ?
            ORDER BY date_created DESC
        ''', site_urls + [currency, month_str]).fetchall()

        # Build cost index (warehouses scoped to partner sites' countries — narrows the load)
        partner_countries = {s['country'] for s in sites if s['country']}
        wh_scope = []
        if partner_countries:
            qmarks = ','.join(['?'] * len(partner_countries))
            wh_scope = [r['id'] for r in conn.execute(
                f'SELECT id FROM warehouses WHERE country IN ({qmarks})', list(partner_countries)
            ).fetchall()]
        cost_idx = _build_dated_cost_index(conn=conn, only_warehouse_ids=wh_scope or None)

        # Country -> default warehouse id list (for orders missing warehouse_id)
        country_default_wh_ids = {}
        for w in conn.execute('SELECT id, country FROM warehouses ORDER BY country, id').fetchall():
            country_default_wh_ids.setdefault(w['country'], []).append(w['id'])

        # Brands + product_mappings cache for line-item resolution
        brands_rows = conn.execute('SELECT id, name, aliases FROM brands').fetchall()
        brands_cache = []
        brand_names_by_id = {}
        for row in brands_rows:
            brand_name = row['name']
            brand_names_by_id[row['id']] = brand_name
            try:
                aliases = json.loads(row['aliases']) if row['aliases'] else []
            except Exception:
                aliases = []
            brands_cache.append({
                'id': row['id'], 'name': brand_name, 'aliases': aliases,
                'patterns': [brand_name.upper()] + [a.upper() for a in aliases]
            })
        pm_rows = conn.execute('SELECT raw_name, source, brand_id, series_id, puff_count, flavor FROM product_mappings').fetchall()
        product_mappings_cache = {}
        for pm in pm_rows:
            product_mappings_cache[(normalize_raw_name(pm['raw_name']), pm['source'])] = pm

    finally:
        conn.close()

    # ---- Aggregate ----
    by_status = {  # status_label -> {orders, gross, shipping, net}
        'success':     {'orders': 0, 'gross': 0.0, 'shipping': 0.0, 'net': 0.0},
        'failed':      {'orders': 0, 'gross': 0.0, 'shipping': 0.0, 'net': 0.0},
        'cancelled':   {'orders': 0, 'gross': 0.0, 'shipping': 0.0, 'net': 0.0},
        'undelivered': {'orders': 0, 'gross': 0.0, 'shipping': 0.0, 'net': 0.0, 'shipping_loss': 0.0},
        'pending':     {'orders': 0, 'gross': 0.0, 'shipping': 0.0, 'net': 0.0},
    }
    by_site = {}      # site_url -> {orders, success_orders, gross, shipping, net, shipping_loss, undelivered}
    by_product = {}   # key -> { brand, series, puffs, flavor, qty, revenue, cost, has_cost, items_count, mapped }

    total_gross = 0.0       # successful gross only
    total_shipping = 0.0    # successful shipping only
    total_shipping_loss = 0.0
    total_actual_cost = 0.0
    unmapped_revenue = 0.0
    unmapped_qty = 0

    success_count = 0
    failed_count = 0
    cancelled_count = 0
    undelivered_count = 0
    pending_count = 0
    total_count = len(all_orders)

    for o in all_orders:
        status = o['status'] or ''
        pm_method = o['payment_method']
        is_undel = bool(o['is_undelivered'])
        gross = float(o['total'] or 0)
        ship = float(o['shipping_total'] or 0)
        loss = float(o['shipping_loss_amount'] or 0)

        # Classification
        is_failed = status == 'failed'
        is_cancelled = status == 'cancelled'
        # 'success' uses the same condition the SQL helper does — see _revenue_status_cond.
        # on-hold without the site's "treated as shipped" flag (AU/AE etc) is
        # also bucketed into pending here so it doesn't inflate net sales.
        is_pending_unsuccessful = (
            (status == 'pending' and (pm_method or 'cod') != 'cod')
            or (status == 'on-hold' and pm_method == 'bacs')
            or (status == 'on-hold' and (o['source'] or '') not in on_hold_shipped_sources)
        )
        # 'refunded' = fully refunded order: money was returned, so it produced
        # no net revenue and must NOT count toward net sales (kept in sync with
        # _revenue_status_cond). Mostly AU today; will grow once online card
        # payments (with refunds) roll out site-wide. Like draft/trash/cheat it
        # then lands in no status bucket — intentional, not part of net.
        is_success = (
            status not in ('failed', 'cancelled', 'checkout-draft', 'trash', 'cheat', 'refunded')
            and not is_pending_unsuccessful
            and not is_undel
        )

        bucket_key = None
        if is_undel:
            bucket_key = 'undelivered'
            undelivered_count += 1
            by_status[bucket_key]['shipping_loss'] += loss
            total_shipping_loss += loss
        elif is_failed:
            bucket_key = 'failed'
            failed_count += 1
        elif is_cancelled:
            bucket_key = 'cancelled'
            cancelled_count += 1
        elif is_pending_unsuccessful:
            bucket_key = 'pending'
            pending_count += 1
        elif is_success:
            bucket_key = 'success'
            success_count += 1

        if bucket_key:
            b = by_status[bucket_key]
            b['orders'] += 1
            b['gross'] += gross
            b['shipping'] += ship
            b['net'] += (gross - ship)

        # Site bucket
        url = o['source']
        if url not in by_site:
            by_site[url] = {
                'site_url': url,
                'country': site_meta.get(url, {}).get('country', ''),
                'manager': site_meta.get(url, {}).get('manager', ''),
                'orders': 0, 'success_orders': 0, 'undelivered_orders': 0,
                'gross': 0.0, 'shipping': 0.0, 'net': 0.0, 'shipping_loss': 0.0,
            }
        s = by_site[url]
        s['orders'] += 1
        if is_success:
            s['success_orders'] += 1
            s['gross'] += gross
            s['shipping'] += ship
            s['net'] += (gross - ship)
            total_gross += gross
            total_shipping += ship
        if is_undel:
            s['undelivered_orders'] += 1
            s['shipping_loss'] += loss

        # Per-product cost & sales — only for SUCCESSFUL orders (those that
        # actually generate revenue and consume inventory).
        if not is_success:
            continue
        items = parse_json_field(o['line_items'])
        if not isinstance(items, list):
            continue
        order_date = (o['date_created'] or '')[:10]
        order_country = site_meta.get(url, {}).get('country') or 'PL'
        order_wh_id = o['warehouse_id']

        # SHIPPING ALLOCATION (per item): split this order's shipping_total
        # across line items by quantity. Industry standard for product-margin
        # analysis: shipping correlates with package size, which correlates
        # with item count. So a 25 PLN shipping order with 3 units → 8.33/unit.
        # The product detail tab uses this to compute "true margin" (excluding
        # shipping cost), helping the partner make product-selection decisions.
        order_total_qty = 0
        for it in items:
            order_total_qty += int(it.get('quantity', 0) or 0)
        order_ship_total = float(o['shipping_total'] or 0)
        per_unit_shipping = (order_ship_total / order_total_qty) if order_total_qty > 0 else 0.0

        for item in items:
            qty = int(item.get('quantity', 0) or 0)
            if qty <= 0:
                continue
            raw_name = item.get('name', '') or ''
            item_total = float(item.get('total', 0) or 0)

            b_id, s_id, p_cnt, flav = _resolve_product_to_brand(
                raw_name, url, brands_cache, product_mappings_cache
            )

            # Lookup cost effective on this order's date. Try the order's own
            # warehouse first, THEN fall back to the order's country default
            # warehouse(s). The fallback makes costing robust to orders carrying a
            # wrong/cross-country warehouse_id (e.g. a PL order mis-tagged to the
            # AU 仓库): without it the lookup misses the cost that IS configured
            # for the order's country, wrongly flagging the line as 未匹配 — and
            # disagreeing with the 未匹配明细 drill-down (api_recon_unmapped_products),
            # which already falls back to the country default warehouses.
            cost_entry = None
            if b_id:
                country_wh = country_default_wh_ids.get(order_country, [])
                effective_wh_ids = ([order_wh_id] if order_wh_id else []) + [w for w in country_wh if w != order_wh_id]
                for ewh in effective_wh_ids:
                    cost_entry = _cost_at_date(cost_idx, b_id, s_id, p_cnt, flav, ewh, order_date)
                    if cost_entry:
                        break

            # Determine cost in partner's currency.
            # POLICY (rolled back from v24/v25): if a product is unmapped
            # (not in product_costs) or currency conversion fails, the line
            # contributes 0 to actual_cost. This intentionally inflates the
            # displayed margin% so the team is forced to fill in real costs
            # — i.e. it's a forcing function for cost data quality, NOT an
            # accidental mis-estimate. The "未匹配产品" warning continues to
            # flag the affected rows.
            line_cost_partner_curr = None
            if cost_entry:
                if cost_entry['currency'] == currency:
                    line_cost_partner_curr = cost_entry['price'] * qty
                else:
                    rate_cost_cny, _ = _lookup_partner_rate(cost_entry['currency'], year, month)
                    rate_partner_cny, _ = _lookup_partner_rate(currency, year, month)
                    if rate_cost_cny and rate_partner_cny:
                        line_cost_partner_curr = cost_entry['price'] * qty * (rate_cost_cny / rate_partner_cny)
            if line_cost_partner_curr is None:
                # Unmapped — track as warning, do NOT add to total_actual_cost
                unmapped_revenue += item_total
                unmapped_qty += qty
            else:
                total_actual_cost += line_cost_partner_curr

            # Build product key — group by mapped (brand, series, puffs, flavor),
            # falling back to the raw_name if we can't resolve.
            if b_id:
                pkey = ('mapped', b_id, s_id or 0, p_cnt or 0, (flav or '').lower())
            else:
                pkey = ('raw', raw_name)

            if pkey not in by_product:
                if pkey[0] == 'mapped':
                    name_parts = [brand_names_by_id.get(b_id, '')]
                    if p_cnt:
                        name_parts.append(f"{p_cnt}口")
                    if flav:
                        name_parts.append(flav)
                    label = ' / '.join(p for p in name_parts if p) or raw_name
                else:
                    label = raw_name
                by_product[pkey] = {
                    'label': label,
                    'brand_id': b_id,
                    'brand': brand_names_by_id.get(b_id, '') if b_id else None,
                    'series_id': s_id,
                    'puff_count': p_cnt,
                    'flavor': flav,
                    'qty': 0,
                    'revenue': 0.0,
                    'cost': 0.0,
                    'cost_mapped': 0.0,      # portion from real product_costs
                    'cost_estimated': 0.0,   # portion from 50% fallback
                    'qty_mapped': 0,
                    'qty_estimated': 0,
                    'allocated_shipping': 0.0,  # shipping share allocated by qty
                    'has_cost': False,
                    'cost_unit': None,
                    'cost_currency': None,
                    'cost_effective_date': None,
                    'cost_match_level': None,
                    'line_count': 0,
                    'mapped': bool(b_id),
                }
            row = by_product[pkey]
            row['qty'] += qty
            row['revenue'] += item_total
            row['line_count'] += 1
            row['allocated_shipping'] += per_unit_shipping * qty
            if line_cost_partner_curr is not None:
                # Real cost from product_costs
                row['cost'] += line_cost_partner_curr
                row['cost_mapped'] += line_cost_partner_curr
                row['qty_mapped'] += qty
                row['has_cost'] = True
                row['cost_unit'] = cost_entry['price']
                row['cost_currency'] = cost_entry['currency']
                row['cost_effective_date'] = cost_entry['effective_date']
                row['cost_match_level'] = cost_entry['match_level']
            else:
                # Unmapped — count qty in the "estimated" bucket so the UI can
                # flag the row, but do NOT add to row['cost']. Keeps margin%
                # artificially inflated as a forcing function.
                row['qty_estimated'] += qty

    # Normalize / sort
    products_list = sorted(by_product.values(), key=lambda r: (-r['revenue']))
    sites_list = sorted(by_site.values(), key=lambda r: (-r['net']))
    for p in products_list:
        p['cost'] = round(p['cost'], 2)
        p['cost_mapped'] = round(p.get('cost_mapped', 0), 2)
        p['cost_estimated'] = round(p.get('cost_estimated', 0), 2)
        p['revenue'] = round(p['revenue'], 2)
        p['allocated_shipping'] = round(p.get('allocated_shipping', 0), 2)
        p['shipping_unit'] = round(p['allocated_shipping'] / p['qty'], 4) if p['qty'] > 0 else 0
        # 毛利 = revenue − cost (cost includes both real and 50%-fallback estimated).
        # has_cost is now always True (estimated cost is still a cost).
        p['margin'] = round(p['revenue'] - p['cost'], 2)
        p['margin_pct'] = round(p['margin'] / p['revenue'] * 100, 2) if p['revenue'] > 0 else None
        # Aliases — kept for legacy frontend that may still reference them.
        p['net_profit'] = p['margin']
        p['net_margin_pct'] = p['margin_pct']
        # Shipping ratio: informational. High ratio = product often in small orders.
        p['shipping_ratio_pct'] = round(p['allocated_shipping'] / p['revenue'] * 100, 2) if p['revenue'] > 0 else None
        # Estimation ratio — what % of this row's revenue used the 50% fallback
        p['estimated_ratio_pct'] = round(p['qty_estimated'] / p['qty'] * 100, 1) if p['qty'] > 0 else 0
        # is_fully_estimated → entire row is from fallback; UI flags it specially
        p['is_fully_estimated'] = (p['qty_mapped'] == 0 and p['qty_estimated'] > 0)
        p['is_partially_estimated'] = (p['qty_mapped'] > 0 and p['qty_estimated'] > 0)
    for s in sites_list:
        for k in ('gross', 'shipping', 'net', 'shipping_loss'):
            s[k] = round(s[k], 2)
    for k, v in by_status.items():
        for fk in ('gross', 'shipping', 'net'):
            v[fk] = round(v[fk], 2)
        if 'shipping_loss' in v:
            v['shipping_loss'] = round(v['shipping_loss'], 2)

    total_net = total_gross - total_shipping - total_shipping_loss
    # NULL means "unset" → use contract default; an explicit 0 is a real ratio
    # and must be honored (0 is falsy, so `or DEFAULT` would wrongly override it).
    cost_ratio = float(partner['cost_ratio']) if partner['cost_ratio'] is not None else 0.5
    pp_ratio = float(partner['partner_profit_ratio']) if partner['partner_profit_ratio'] is not None else 0.25
    op_ratio = float(partner['our_profit_ratio']) if partner['our_profit_ratio'] is not None else 0.25

    # Actual margin = net − actual cost (50%-fallback already inside total_actual_cost)
    actual_margin = total_net - total_actual_cost
    actual_margin_pct = round(actual_margin / total_net * 100, 2) if total_net > 0 else None
    # Sum mapped vs estimated cost from product rows for transparency in UI
    total_cost_mapped = round(sum(p.get('cost_mapped', 0) or 0 for p in products_list), 2)
    total_cost_estimated = round(sum(p.get('cost_estimated', 0) or 0 for p in products_list), 2)
    estimated_ratio_pct = round(total_cost_estimated / total_actual_cost * 100, 1) if total_actual_cost > 0 else None

    result = {
        'partner_id': partner['id'],
        'partner_name': partner['name'],
        'period_year': year,
        'period_month': month,
        'currency': currency,
        # Counts
        'total_count': total_count,
        'success_orders': success_count,
        'failed_orders': failed_count,
        'cancelled_orders': cancelled_count,
        'undelivered_orders': undelivered_count,
        'pending_orders': pending_count,
        # Top-line numbers (in partner currency, *_pln suffix kept for legacy compat)
        'total_orders': success_count,                  # legacy: success only
        'total_gross_pln': round(total_gross, 2),
        'total_shipping_pln': round(total_shipping, 2),
        'shipping_loss': round(total_shipping_loss, 2),
        'total_net_pln': round(total_net, 2),
        # Cost (two ways). Contract-basis cost coverage is computed on the
        # SUCCESSFUL net (净销售 + 未送达损失) — a shipping loss never shrinks the
        # partner's product-cost coverage; the loss is instead borne 50/50 by the
        # two profit parties below (kept in sync with _compute_statement_split).
        'cost_amount_pln': round((total_net + total_shipping_loss) * cost_ratio, 2),
        'actual_cost_pln': round(total_actual_cost, 2),              # real product_costs lookup (incl. 50% fallback)
        'cost_mapped_pln': total_cost_mapped,                        # portion from real product_costs
        'cost_estimated_pln': total_cost_estimated,                  # portion from 50% fallback
        'estimated_ratio_pct': estimated_ratio_pct,                  # cost_estimated / actual_cost × 100
        'cost_unmapped_revenue_pln': round(unmapped_revenue, 2),
        'cost_unmapped_qty': unmapped_qty,
        # Actual margin (key new field — for 实际毛利 / 实际毛利率 cards)
        'actual_margin_pln': round(actual_margin, 2),
        'actual_margin_pct': actual_margin_pct,
        # Profit breakdown (contract basis): split the successful net by the
        # ratios, then each profit party bears half the undelivered loss.
        'partner_profit_pln': round((total_net + total_shipping_loss) * pp_ratio - total_shipping_loss / 2, 2),
        'our_receivable_pln': round((total_net + total_shipping_loss) * op_ratio - total_shipping_loss / 2, 2),
        # Ratios echoed for UI
        'cost_ratio': cost_ratio,
        'partner_profit_ratio': pp_ratio,
        'our_profit_ratio': op_ratio,
        # Drill-down breakdowns
        'by_status': by_status,
        'by_site': sites_list,
        'by_product': products_list,
    }
    _enrich_statement_cny(result, currency)
    return result


def _empty_recon_detail(partner, currency, year, month):
    """Empty skeleton when partner has no sites bound yet."""
    return {
        'partner_id': partner['id'],
        'partner_name': partner['name'],
        'period_year': year,
        'period_month': month,
        'currency': currency,
        'total_count': 0,
        'success_orders': 0,
        'failed_orders': 0,
        'cancelled_orders': 0,
        'undelivered_orders': 0,
        'pending_orders': 0,
        'total_orders': 0,
        'total_gross_pln': 0,
        'total_shipping_pln': 0,
        'shipping_loss': 0,
        'total_net_pln': 0,
        'cost_amount_pln': 0,
        'actual_cost_pln': 0,
        'cost_mapped_pln': 0,
        'cost_estimated_pln': 0,
        'estimated_ratio_pct': None,
        'cost_unmapped_revenue_pln': 0,
        'cost_unmapped_qty': 0,
        'actual_margin_pln': 0,
        'actual_margin_pct': None,
        'partner_profit_pln': 0,
        'our_receivable_pln': 0,
        'cost_ratio': float(partner['cost_ratio']) if partner['cost_ratio'] is not None else 0.5,
        'partner_profit_ratio': float(partner['partner_profit_ratio']) if partner['partner_profit_ratio'] is not None else 0.25,
        'our_profit_ratio': float(partner['our_profit_ratio']) if partner['our_profit_ratio'] is not None else 0.25,
        'by_status': {},
        'by_site': [],
        'by_product': [],
    }


# ============================================================================
# P2: AUDIT LOG + ORDER SNAPSHOT helpers
# ============================================================================
def _audit_log(statement_id, action, field=None, old=None, new=None, note=None, conn=None):
    """Write one entry to reconciliation_audit_log.

    Always idempotent — caller passes whatever `action` makes sense
    ('create', 'regenerate', 'edit_notes', 'edit_rate', 'lock', 'unlock',
     'confirm', 'dispute', 'resolve_dispute', 'settle',
     'attach_receipt', 'detach_receipt', 'edit_receipt', 'delete_receipt').
    Captures current_user + IP automatically.

    Caller may pass an existing connection to coalesce with the same
    transaction; otherwise we open a fresh one.
    """
    own_conn = conn is None
    if own_conn:
        conn = get_db_connection()
    try:
        actor_id = current_user.id if current_user.is_authenticated else None
        actor_username = current_user.username if current_user.is_authenticated else None
        actor_role = current_user.role if (current_user.is_authenticated and hasattr(current_user, 'role')) else None
        ip = None
        try:
            if request:
                ip = request.headers.get('X-Forwarded-For', request.remote_addr)
                if ip and ',' in ip:
                    ip = ip.split(',')[0].strip()
        except Exception:
            pass
        conn.execute('''
            INSERT INTO reconciliation_audit_log
                (statement_id, action, actor_id, actor_username, actor_role,
                 field, old_value, new_value, note, ip)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        ''', (
            statement_id, action, actor_id, actor_username, actor_role,
            field,
            None if old is None else str(old)[:500],
            None if new is None else str(new)[:500],
            (note or '')[:1000] if note else None,
            ip,
        ))
        if own_conn:
            conn.commit()
    finally:
        if own_conn:
            conn.close()


def _snapshot_statement_orders(statement_id, partner_id, year, month, conn):
    """Freeze the list of orders that contributed to this statement.

    Stores a snapshot of (status, total, shipping, shipping_loss, ...) so the
    saved statement remains stable even if the underlying orders are later
    edited. Replaces any previous snapshot for the same statement (re-generate
    case). Caller must pass an open connection — we don't commit.
    """
    # Use partner's currency + bound sites to identify the orders
    p = conn.execute('SELECT currency FROM partners WHERE id = ?', (partner_id,)).fetchone()
    currency = (p['currency'] if p else 'PLN') or 'PLN'
    sites = conn.execute('''SELECT s.url FROM partner_sites ps
        JOIN sites s ON s.id = ps.site_id WHERE ps.partner_id = ?''', (partner_id,)).fetchall()
    if not sites:
        return 0
    site_urls = [s['url'] for s in sites]
    placeholders = ','.join(['?'] * len(site_urls))
    month_str = f'{year:04d}-{month:02d}'
    rows = conn.execute(f'''
        SELECT id, number, status, total, shipping_total, shipping_loss_amount,
               is_undelivered, currency, date_created
        FROM orders
        WHERE source IN ({placeholders})
          AND currency = ?
          AND strftime('%Y-%m', date_created) = ?
    ''', site_urls + [currency, month_str]).fetchall()

    # Wipe + replace
    conn.execute('DELETE FROM reconciliation_statement_orders WHERE statement_id = ?', (statement_id,))
    for r in rows:
        conn.execute('''INSERT INTO reconciliation_statement_orders
            (statement_id, order_id, order_number, status_at_gen, total_at_gen,
             shipping_at_gen, shipping_loss_at_gen, is_undelivered_at_gen,
             currency_at_gen, date_created)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)''',
            (statement_id, r['id'], r['number'], r['status'],
             float(r['total'] or 0), float(r['shipping_total'] or 0),
             float(r['shipping_loss_amount'] or 0),
             1 if r['is_undelivered'] else 0,
             r['currency'], r['date_created']))
    return len(rows)


def _lookup_partner_rate(currency, year, month):
    """Look up CNY rate for a currency in a specific month from exchange_rates table.
    Returns (rate, actual_year_month) or (None, None) if no rate configured."""
    if not currency or currency == 'CNY':
        return 1.0, f'{year:04d}-{month:02d}'
    ym = f'{year:04d}-{month:02d}'
    return get_cny_rate(currency, ym)


def _enrich_statement_cny(stmt_dict, partner_currency='PLN'):
    """Attach CNY-converted fields to a statement dict.
    Uses saved exchange_rate_cny if present, otherwise looks up from exchange_rates table."""
    rate = stmt_dict.get('exchange_rate_cny')
    rate_source = 'saved'
    if not rate:
        looked_up, actual_ym = _lookup_partner_rate(
            partner_currency,
            stmt_dict.get('period_year') or 0,
            stmt_dict.get('period_month') or 0
        )
        rate = looked_up
        rate_source = f'system({actual_ym})' if looked_up else 'none'

    stmt_dict['effective_rate_cny'] = rate
    stmt_dict['rate_source'] = rate_source

    if rate:
        for pln_field, cny_field in [
            ('total_gross_pln', 'total_gross_cny'),
            ('total_net_pln', 'total_net_cny'),
            ('cost_amount_pln', 'cost_amount_cny'),
            ('actual_cost_pln', 'actual_cost_cny'),
            ('cost_mapped_pln', 'cost_mapped_cny'),
            ('cost_estimated_pln', 'cost_estimated_cny'),
            ('actual_margin_pln', 'actual_margin_cny'),
            ('partner_profit_pln', 'partner_profit_cny'),
            ('our_receivable_pln', 'our_receivable_cny'),
        ]:
            v = stmt_dict.get(pln_field)
            if v is not None:
                stmt_dict[cny_field] = round(v * rate, 2)
    return stmt_dict


# ---------- Partner CRUD ----------

@app.route('/partner-reconciliation')
@login_required
@reconciliation_viewer_required
def partner_reconciliation_page():
    """Partner reconciliation main page"""
    is_recon_admin = _is_reconciliation_admin()
    import time
    resp = make_response(render_template('partner_reconciliation.html',
                          is_recon_admin=is_recon_admin,
                          is_super_admin=(current_user.username == 'admin'),
                          cache_bust=int(time.time())))
    resp.headers['Cache-Control'] = 'no-store, no-cache, must-revalidate, max-age=0'
    resp.headers['Pragma'] = 'no-cache'
    resp.headers['Expires'] = '0'
    return resp


@app.route('/api/partners')
@login_required
@reconciliation_api_required
def api_list_partners():
    """List partners the current user can access"""
    allowed_ids = current_user.get_accessible_partner_ids()
    conn = get_db_connection()
    if allowed_ids is None:
        partners = conn.execute('SELECT * FROM partners ORDER BY id').fetchall()
    elif len(allowed_ids) == 0:
        conn.close()
        return jsonify([])
    else:
        placeholders = ','.join(['?'] * len(allowed_ids))
        partners = conn.execute(f'SELECT * FROM partners WHERE id IN ({placeholders}) ORDER BY id', allowed_ids).fetchall()
    conn.close()
    return jsonify([dict(p) for p in partners])


@app.route('/api/partners', methods=['POST'])
@login_required
@reconciliation_api_required
def api_create_partner():
    """Create a new partner (admin only)"""
    if not _is_reconciliation_admin():
        return jsonify({'error': '无权创建合伙人'}), 403
    data = request.json
    name = (data.get('name') or '').strip()
    code = (data.get('code') or '').strip()
    if not name:
        return jsonify({'error': '合伙人名称必填'}), 400
    conn = get_db_connection()
    try:
        cursor = conn.execute('''
            INSERT INTO partners (name, code, description, cost_ratio, partner_profit_ratio, our_profit_ratio, currency)
            VALUES (?, ?, ?, ?, ?, ?, ?)
        ''', (name, code or None, data.get('description', ''),
              float(data.get('cost_ratio', 0.5)),
              float(data.get('partner_profit_ratio', 0.25)),
              float(data.get('our_profit_ratio', 0.25)),
              data.get('currency', 'PLN')))
        conn.commit()
        return jsonify({'success': True, 'id': cursor.lastrowid})
    except sqlite3.IntegrityError as e:
        return jsonify({'error': f'创建失败: {e}'}), 400
    finally:
        conn.close()


@app.route('/api/partners/<int:partner_id>', methods=['PUT'])
@login_required
@reconciliation_api_required
def api_update_partner(partner_id):
    """Update partner info and ratios (admin only)"""
    if not _is_reconciliation_admin():
        return jsonify({'error': '无权修改合伙人'}), 403
    data = request.json
    conn = get_db_connection()
    try:
        conn.execute('''
            UPDATE partners SET name=?, description=?, cost_ratio=?, partner_profit_ratio=?, our_profit_ratio=?,
                currency=?, updated_at=CURRENT_TIMESTAMP
            WHERE id=?
        ''', ((data.get('name') or '').strip(),
              data.get('description', ''),
              float(data.get('cost_ratio', 0.5)),
              float(data.get('partner_profit_ratio', 0.25)),
              float(data.get('our_profit_ratio', 0.25)),
              data.get('currency', 'PLN'),
              partner_id))
        conn.commit()
        return jsonify({'success': True})
    finally:
        conn.close()


@app.route('/api/partners/<int:partner_id>', methods=['DELETE'])
@login_required
@reconciliation_api_required
def api_delete_partner(partner_id):
    """Delete partner (super admin only)"""
    if current_user.username != 'admin':
        return jsonify({'error': '只有超级管理员可以删除合伙人'}), 403
    conn = get_db_connection()
    try:
        conn.execute('DELETE FROM partners WHERE id = ?', (partner_id,))
        conn.commit()
        return jsonify({'success': True})
    finally:
        conn.close()


@app.route('/api/partners/<int:partner_id>/sites')
@login_required
@reconciliation_api_required
def api_get_partner_sites(partner_id):
    """Get sites bound to a partner + all available sites"""
    conn = get_db_connection()
    bound = conn.execute('SELECT site_id FROM partner_sites WHERE partner_id = ?', (partner_id,)).fetchall()
    all_sites = conn.execute('SELECT id, url, country, manager FROM sites ORDER BY country, url').fetchall()
    conn.close()
    return jsonify({
        'bound_site_ids': [b['site_id'] for b in bound],
        'all_sites': [dict(s) for s in all_sites]
    })


@app.route('/api/partners/<int:partner_id>/sites', methods=['PUT'])
@login_required
@reconciliation_api_required
def api_update_partner_sites(partner_id):
    """Update sites bound to a partner (admin only)"""
    if not _is_reconciliation_admin():
        return jsonify({'error': '无权修改站点绑定'}), 403
    data = request.json
    site_ids = data.get('site_ids', [])
    conn = get_db_connection()
    try:
        conn.execute('DELETE FROM partner_sites WHERE partner_id = ?', (partner_id,))
        for sid in site_ids:
            conn.execute('INSERT OR IGNORE INTO partner_sites (partner_id, site_id) VALUES (?, ?)',
                        (partner_id, sid))
        conn.commit()
        return jsonify({'success': True})
    finally:
        conn.close()


@app.route('/api/partners/<int:partner_id>/users')
@login_required
@reconciliation_api_required
def api_get_partner_users(partner_id):
    """Get users bound to a partner + eligible users (those with can_view_reconciliation).
    Super admin 'admin' is excluded — it already sees everything."""
    conn = get_db_connection()
    bound = conn.execute('SELECT user_id FROM partner_users WHERE partner_id = ?', (partner_id,)).fetchall()
    try:
        eligible_users = conn.execute('''
            SELECT id, username, name, role
            FROM users
            WHERE can_view_reconciliation = 1 AND username != 'admin'
            ORDER BY id
        ''').fetchall()
    except sqlite3.OperationalError:
        eligible_users = []
    conn.close()
    return jsonify({
        'bound_user_ids': [b['user_id'] for b in bound],
        'all_users': [dict(u) for u in eligible_users]
    })


@app.route('/api/partners/<int:partner_id>/users', methods=['PUT'])
@login_required
@reconciliation_api_required
def api_update_partner_users(partner_id):
    """Update users bound to a partner (super admin only)"""
    if current_user.username != 'admin':
        return jsonify({'error': '只有超级管理员可以绑定用户'}), 403
    data = request.json
    user_ids = data.get('user_ids', [])
    conn = get_db_connection()
    try:
        conn.execute('DELETE FROM partner_users WHERE partner_id = ?', (partner_id,))
        for uid in user_ids:
            conn.execute('INSERT OR IGNORE INTO partner_users (partner_id, user_id) VALUES (?, ?)',
                        (partner_id, uid))
        conn.commit()
        return jsonify({'success': True})
    finally:
        conn.close()


# ---------- Reconciliation Statements ----------

def _check_partner_access(partner_id):
    """Return True if current user can access this partner's data."""
    allowed = current_user.get_accessible_partner_ids()
    if allowed is None:
        return True
    return partner_id in allowed


@app.route('/api/reconciliation/statements')
@login_required
@reconciliation_api_required
def api_list_statements():
    """List statements, filtered by access + optional partner_id/year"""
    partner_id = request.args.get('partner_id', type=int)
    year = request.args.get('year', type=int)
    allowed_ids = current_user.get_accessible_partner_ids()

    conn = get_db_connection()
    query = 'SELECT * FROM reconciliation_statements WHERE 1=1'
    params = []
    if partner_id:
        if allowed_ids is not None and partner_id not in allowed_ids:
            conn.close()
            return jsonify({'error': '无权查看此合伙人'}), 403
        query += ' AND partner_id = ?'
        params.append(partner_id)
    elif allowed_ids is not None:
        if not allowed_ids:
            conn.close()
            return jsonify([])
        placeholders = ','.join(['?'] * len(allowed_ids))
        query += f' AND partner_id IN ({placeholders})'
        params.extend(allowed_ids)
    if year:
        query += ' AND period_year = ?'
        params.append(year)
    query += ' ORDER BY period_year DESC, period_month DESC, partner_id'
    rows = conn.execute(query, params).fetchall()
    # Load partner currencies for enrichment
    partner_curr = {p['id']: p['currency'] for p in conn.execute('SELECT id, currency FROM partners').fetchall()}
    conn.close()
    result = []
    for r in rows:
        d = dict(r)
        d['currency'] = partner_curr.get(d['partner_id'], 'PLN')
        _enrich_statement_cny(d, d['currency'])
        result.append(d)
    return jsonify(result)


@app.route('/api/reconciliation/statements/preview')
@login_required
@reconciliation_api_required
def api_preview_statement():
    """Preview a statement without saving — full drill-down detail.

    Returns the same top-level numbers as before (total_orders / gross / net /
    cost / partner_profit / our_receivable) for backward compat, plus new
    fields: actual_cost_pln, by_status, by_site, by_product. Drill-down UI
    can render all sections from this single response.
    """
    partner_id = request.args.get('partner_id', type=int)
    year = request.args.get('year', type=int)
    month = request.args.get('month', type=int)
    if not all([partner_id, year, month]):
        return jsonify({'error': '缺少参数'}), 400
    if not _check_partner_access(partner_id):
        return jsonify({'error': '无权查看此合伙人'}), 403

    detail = _calc_partner_recon_detail(partner_id, year, month)
    if detail is None:
        return jsonify({'error': '合伙人不存在'}), 404
    return jsonify(detail)


@app.route('/api/reconciliation/orders')
@login_required
@reconciliation_api_required
def api_recon_orders():
    """Order list filtered by partner (their bound sites + currency + period).

    Replicates the data-shape of /orders for embedding inside the partner
    reconciliation page, INCLUDING the page-totals summary row used by the
    "本页合计" header. When a year+month filter is set, returns ALL orders
    for that month (capped at 5000) — no internal pagination, matching the
    /orders page behavior where one month = one page.
    """
    partner_id = request.args.get('partner_id', type=int)
    year = request.args.get('year', type=int)
    month = request.args.get('month', type=int)
    status_filter = (request.args.get('status') or '').strip()
    source_filter = (request.args.get('source') or '').strip()
    search = (request.args.get('search') or '').strip()
    # Pagination — only kicks in when a month filter is NOT set. With a month
    # filter, we dump the entire month in one response (5000-row safety cap).
    requested_page = max(1, request.args.get('page', 1, type=int))
    requested_per_page = min(5000, max(20, request.args.get('per_page', 200, type=int)))
    use_pagination = not (year and month)

    if not partner_id:
        return jsonify({'error': '缺少 partner_id 参数'}), 400
    if not _check_partner_access(partner_id):
        return jsonify({'error': '无权查看此合伙人'}), 403

    conn = get_db_connection()
    try:
        partner = conn.execute('SELECT id, currency FROM partners WHERE id = ?', (partner_id,)).fetchone()
        if not partner:
            return jsonify({'error': '合伙人不存在'}), 404
        currency = partner['currency'] or 'PLN'

        sites = conn.execute('''
            SELECT s.url FROM partner_sites ps
            JOIN sites s ON s.id = ps.site_id
            WHERE ps.partner_id = ?
        ''', (partner_id,)).fetchall()
        if not sites:
            return jsonify({'orders': [], 'total': 0, 'page': requested_page,
                            'per_page': requested_per_page, 'currency': currency,
                            'displayed_totals': _empty_displayed_totals(currency)})
        site_urls = [s['url'] for s in sites]
        if source_filter:
            if source_filter not in site_urls:
                return jsonify({'error': '该站点不在合伙人范围内'}), 403
            allowed_urls = [source_filter]
        else:
            allowed_urls = site_urls

        placeholders = ','.join(['?'] * len(allowed_urls))
        conditions = [f'source IN ({placeholders})', 'currency = ?']
        params = list(allowed_urls) + [currency]
        if year and month:
            conditions.append("strftime('%Y-%m', date_created) = ?")
            params.append(f'{year:04d}-{month:02d}')
        elif year:
            conditions.append("strftime('%Y', date_created) = ?")
            params.append(str(year))
        if status_filter:
            if status_filter == 'undelivered':
                conditions.append('COALESCE(is_undelivered, 0) = 1')
            elif status_filter == 'success':
                conditions.append(_revenue_status_cond())
            else:
                conditions.append('status = ?')
                params.append(status_filter)
        if search:
            conditions.append("(number LIKE ? OR id LIKE ? OR billing LIKE ?)")
            like = f'%{search}%'
            params.extend([like, like, like])
        where_sql = ' AND '.join(conditions)

        total = conn.execute(f'SELECT COUNT(*) FROM orders WHERE {where_sql}', params).fetchone()[0]

        # Fetch rows — paginated only if no month filter
        if use_pagination:
            offset = (requested_page - 1) * requested_per_page
            limit_sql = f'LIMIT {requested_per_page} OFFSET {offset}'
            page_returned = requested_page
            per_page_returned = requested_per_page
        else:
            # All-in-one-page mode: cap at 5000 rows for safety
            limit_sql = 'LIMIT 5000'
            page_returned = 1
            per_page_returned = max(total, 1)

        rows = conn.execute(f'''
            SELECT id, number, status, payment_method, currency, date_created,
                   total, shipping_total, line_items, source, billing,
                   is_undelivered, shipping_loss_amount
            FROM orders WHERE {where_sql}
            ORDER BY date_created DESC
            {limit_sql}
        ''', params).fetchall()

        site_managers = {s['url']: s['manager'] for s in conn.execute('SELECT url, manager FROM sites').fetchall()}
        site_country_map = {s['url']: s['country'] for s in conn.execute('SELECT url, country FROM sites').fetchall()}
        on_hold_shipped_sources = {
            r['url'] for r in conn.execute(
                "SELECT url FROM sites WHERE cod_on_hold_is_shipped = 1"
            ).fetchall()
        }

        # Lookup CNY rate for the period (used per-order net_cny conversion).
        # Fall back to month-of-order if no period rate, so multi-month dumps
        # still get correct conversions.
        period_rate = None
        if year and month:
            period_rate, _ = _lookup_partner_rate(currency, year, month)

        # ---------- per-order COST machinery ----------
        # Build the date-aware cost index + brand/product-mappings caches once
        # per request. Then for each order's line items, look up the cost that
        # was effective on the order date and sum to a per-order cost.
        # Cost is in partner's native currency (we convert from cost_currency
        # via CNY when they differ).
        cost_idx = _build_dated_cost_index(conn=conn)
        country_default_wh_ids = {}
        for w in conn.execute('SELECT id, country FROM warehouses ORDER BY country, id').fetchall():
            country_default_wh_ids.setdefault(w['country'], []).append(w['id'])
        brands_rows = conn.execute('SELECT id, name, aliases FROM brands').fetchall()
        brands_cache = []
        for row in brands_rows:
            try:
                aliases = json.loads(row['aliases']) if row['aliases'] else []
            except Exception:
                aliases = []
            brands_cache.append({
                'id': row['id'], 'name': row['name'], 'aliases': aliases,
                'patterns': [row['name'].upper()] + [a.upper() for a in aliases]
            })
        pm_rows = conn.execute('SELECT raw_name, source, brand_id, series_id, puff_count, flavor FROM product_mappings').fetchall()
        product_mappings_cache = {}
        for pm in pm_rows:
            product_mappings_cache[(normalize_raw_name(pm['raw_name']), pm['source'])] = pm

        # Aggregator for "本页合计" — mirrors the orders.html displayed_totals.
        totals = {
            'order_count': 0,
            'undelivered_count': 0,
            'product_count': 0,
            'currency': currency,
            'amount': 0.0,
            'shipping': 0.0,
            'net': 0.0,                # gross-shipping for revenue-status orders only
            'shipping_loss': 0.0,      # loss from undelivered orders
            # ── new cost / margin aggregates ──
            'cost': 0.0,               # total actual product cost (revenue orders only)
            'margin': 0.0,             # net − cost
            'unmapped_qty': 0,         # qty of line items without cost mapping
            'rate_to_cny': period_rate,
            'net_cny': 0.0,
            'shipping_loss_cny': 0.0,
            'final_net_cny': 0.0,
            'cost_cny': 0.0,
            'margin_cny': 0.0,
        }

        def _is_revenue(status, pm, is_undel, src=None):
            if status in ('failed', 'cancelled', 'checkout-draft', 'trash', 'cheat'):
                return False
            if status == 'pending' and (pm or 'cod') != 'cod':
                return False
            if status == 'on-hold' and pm == 'bacs':
                return False
            if status == 'on-hold' and src not in on_hold_shipped_sources:
                # AU/AE etc — on-hold is "received, waiting", never shipped.
                return False
            if is_undel:
                return False
            return True

        # Per-request cache for warehouse-lookup-by-source (avoids repeat DB hits).
        # Each order needs to know which warehouse(s) to look up costs against.
        order_warehouse_cache = {}  # source_url -> [warehouse_ids]

        # Cache for partner-currency conversion of cost-currency (avoids
        # repeat _lookup_partner_rate calls for the same currency in a request)
        _cost_rate_cache = {}
        def _convert_cost_to_partner(price, cost_currency, year, month):
            """Convert a single-unit cost price from cost_currency to partner currency."""
            if cost_currency == currency:
                return price
            cache_key = (cost_currency, year, month)
            if cache_key not in _cost_rate_cache:
                rcost, _ = _lookup_partner_rate(cost_currency, year, month)
                rpart, _ = _lookup_partner_rate(currency, year, month)
                _cost_rate_cache[cache_key] = (rcost, rpart)
            rcost, rpart = _cost_rate_cache[cache_key]
            if rcost and rpart:
                return price * (rcost / rpart)
            return None  # rate unavailable — caller treats as unmapped

        result_orders = []
        for r in rows:
            items = parse_json_field(r['line_items']) or []
            products = []
            qty_total = 0
            for it in items if isinstance(items, list) else []:
                qty = int(it.get('quantity', 0) or 0)
                qty_total += qty
                products.append({
                    'name': it.get('name', ''),
                    'quantity': qty,
                    'total': float(it.get('total', 0) or 0),
                })
            billing = parse_json_field(r['billing']) or {}
            customer_name = ''
            customer_email = ''
            if isinstance(billing, dict):
                customer_name = ((billing.get('first_name') or '') + ' ' + (billing.get('last_name') or '')).strip()
                customer_email = billing.get('email', '') or ''

            shipping_loss = float(r['shipping_loss_amount'] or 0)
            is_undel = bool(r['is_undelivered'])
            gross = float(r['total'] or 0)
            ship = float(r['shipping_total'] or 0)
            net = max(0, gross - ship)
            order_rate = period_rate
            if order_rate is None:
                # multi-month dump: look up the order's own month
                if r['date_created']:
                    try:
                        y, m = r['date_created'][:7].split('-')
                        order_rate, _ = _lookup_partner_rate(currency, int(y), int(m))
                    except Exception:
                        order_rate = None
            net_cny = round(net * order_rate, 2) if (order_rate and net) else None

            # ── COST CALCULATION (per-order, line-item level, date-aware) ──
            # When a line item has no real cost in product_costs, fall back to
            # 50% of revenue. Same logic as _calc_partner_recon_detail so the
            # 订单明细 tab and 产品明细 tab agree on totals.
            UNMAPPED_FALLBACK_RATIO = 0.5
            order_is_revenue = _is_revenue(r['status'] or '', r['payment_method'], is_undel, r['source'])
            order_cost = 0.0
            order_unmapped_qty = 0
            cost_eligible = order_is_revenue and not is_undel  # only revenue orders consume inventory
            if cost_eligible and isinstance(items, list):
                order_date = (r['date_created'] or '')[:10]
                order_country = site_country_map.get(r['source'], 'PL') or 'PL'
                # Get year/month from order date for cost-currency conversion
                try:
                    od_y, od_m = (r['date_created'] or '0000-00')[:7].split('-')
                    od_y, od_m = int(od_y), int(od_m)
                except Exception:
                    od_y, od_m = (year or 0), (month or 0)
                effective_wh_ids = country_default_wh_ids.get(order_country, [])
                for it in items:
                    qty = int(it.get('quantity', 0) or 0)
                    if qty <= 0:
                        continue
                    raw_name = it.get('name', '') or ''
                    line_total = float(it.get('total', 0) or 0)
                    b_id, s_id, p_cnt, flav = _resolve_product_to_brand(
                        raw_name, r['source'], brands_cache, product_mappings_cache
                    )
                    cost_entry = None
                    if b_id and effective_wh_ids:
                        for ewh in effective_wh_ids:
                            cost_entry = _cost_at_date(cost_idx, b_id, s_id, p_cnt, flav, ewh, order_date)
                            if cost_entry:
                                break
                    line_cost = None
                    if cost_entry:
                        unit_in_partner = _convert_cost_to_partner(
                            cost_entry['price'], cost_entry['currency'], od_y, od_m
                        )
                        if unit_in_partner is not None:
                            line_cost = unit_in_partner * qty
                    if line_cost is None:
                        # POLICY: unmapped lines contribute 0 to order_cost.
                        # Margin% will be inflated for these orders, which is
                        # the intended forcing function to drive cost data
                        # quality. The unmapped warning still flags them.
                        order_unmapped_qty += qty
                    else:
                        order_cost += line_cost

            order_cost = round(order_cost, 2)
            order_margin = round(net - order_cost, 2) if cost_eligible else None
            order_margin_pct = None
            if cost_eligible and net > 0:
                order_margin_pct = round((net - order_cost) / net * 100, 1)
            order_cost_cny = round(order_cost * order_rate, 2) if (order_rate and cost_eligible) else None
            order_margin_cny = round(order_margin * order_rate, 2) if (order_rate and order_margin is not None) else None

            result_orders.append({
                'id': r['id'],
                'number': r['number'],
                'date_created': r['date_created'],
                'status': r['status'],
                'payment_method': r['payment_method'],
                'currency': r['currency'],
                'source': r['source'],
                'manager': site_managers.get(r['source'], ''),
                'total': gross,
                'shipping_total': ship,
                'net_total': net,
                'net_total_cny': net_cny,
                'rate_to_cny': order_rate,
                'product_count': qty_total,
                'products': products,
                'customer_name': customer_name,
                'customer_email': customer_email,
                'is_undelivered': is_undel,
                'shipping_loss_amount': shipping_loss,
                # ── cost/margin fields (margin% may be inflated for orders
                #     with unmapped products — by design, see policy comment) ──
                'cost_eligible': cost_eligible,
                'cost': order_cost if cost_eligible else None,
                'cost_cny': order_cost_cny,
                'margin': order_margin,
                'margin_cny': order_margin_cny,
                'margin_pct': order_margin_pct,
                'unmapped_qty': order_unmapped_qty,
            })

            # Aggregate totals row
            totals['order_count'] += 1
            totals['product_count'] += qty_total
            totals['amount'] += gross
            totals['shipping'] += ship
            if is_undel:
                totals['undelivered_count'] += 1
                totals['shipping_loss'] += shipping_loss
                if order_rate:
                    totals['shipping_loss_cny'] += shipping_loss * order_rate
            elif order_is_revenue:
                totals['net'] += net
                if net_cny is not None:
                    totals['net_cny'] += net_cny
                totals['cost'] += order_cost
                totals['margin'] += (net - order_cost)
                if order_rate:
                    totals['cost_cny'] += order_cost * order_rate
                    totals['margin_cny'] += (net - order_cost) * order_rate
                totals['unmapped_qty'] += order_unmapped_qty

        # Round totals for display
        for k in ('amount', 'shipping', 'net', 'shipping_loss',
                  'cost', 'margin', 'net_cny', 'shipping_loss_cny',
                  'cost_cny', 'margin_cny'):
            totals[k] = round(totals[k], 2)
        totals['final_net_cny'] = round(totals['net_cny'] - totals['shipping_loss_cny'], 2)
        totals['margin_pct'] = round(totals['margin'] / totals['net'] * 100, 1) if totals['net'] > 0 else None

        return jsonify({
            'orders': result_orders,
            'total': total,
            'page': page_returned,
            'per_page': per_page_returned,
            'use_pagination': use_pagination,
            'currency': currency,
            'partner_id': partner_id,
            'displayed_totals': totals,
        })
    finally:
        conn.close()


def _empty_displayed_totals(currency):
    return {
        'order_count': 0, 'undelivered_count': 0, 'product_count': 0,
        'currency': currency,
        'amount': 0.0, 'shipping': 0.0, 'net': 0.0, 'shipping_loss': 0.0,
        'cost': 0.0, 'margin': 0.0, 'unmapped_qty': 0,
        'rate_to_cny': None, 'net_cny': 0.0, 'shipping_loss_cny': 0.0, 'final_net_cny': 0.0,
        'cost_cny': 0.0, 'margin_cny': 0.0, 'margin_pct': None,
    }


@app.route('/api/reconciliation/summary-stats')
@login_required
@reconciliation_api_required
def api_recon_summary_stats():
    """Per-site summary statistics for the partner — replicates the
    "筛选结果统计" table from /orders, scoped to partner-bound sites."""
    partner_id = request.args.get('partner_id', type=int)
    year = request.args.get('year', type=int)
    month = request.args.get('month', type=int)
    if not partner_id:
        return jsonify({'error': '缺少 partner_id'}), 400
    if not _check_partner_access(partner_id):
        return jsonify({'error': '无权查看此合伙人'}), 403

    conn = get_db_connection()
    try:
        partner = conn.execute('SELECT id, currency FROM partners WHERE id = ?', (partner_id,)).fetchone()
        if not partner:
            return jsonify({'error': '合伙人不存在'}), 404
        currency = partner['currency'] or 'PLN'

        sites = conn.execute('''
            SELECT s.url, s.country, s.manager
            FROM partner_sites ps
            JOIN sites s ON s.id = ps.site_id
            WHERE ps.partner_id = ?
        ''', (partner_id,)).fetchall()
        if not sites:
            return jsonify({'rows': [], 'totals': {}, 'currency': currency})

        site_urls = [s['url'] for s in sites]
        site_meta = {s['url']: {'country': s['country'], 'manager': s['manager']} for s in sites}
        placeholders = ','.join(['?'] * len(site_urls))

        params = list(site_urls) + [currency]
        date_cond = ''
        if year and month:
            date_cond = " AND strftime('%Y-%m', date_created) = ?"
            params.append(f'{year:04d}-{month:02d}')
        elif year:
            date_cond = " AND strftime('%Y', date_created) = ?"
            params.append(str(year))

        # Per-site aggregation
        success_cond = _revenue_status_cond()
        rows = conn.execute(f'''
            SELECT
                source,
                COUNT(*) as total_orders,
                COALESCE(SUM(total), 0) as total_amount,
                SUM(CASE WHEN {success_cond} THEN 1 ELSE 0 END) as success_orders,
                COALESCE(SUM(CASE WHEN {success_cond} THEN total ELSE 0 END), 0) as success_amount,
                COALESCE(SUM(CASE WHEN {success_cond} THEN shipping_total ELSE 0 END), 0) as success_shipping,
                SUM(CASE WHEN status = 'failed' THEN 1 ELSE 0 END) as failed_orders,
                SUM(CASE WHEN status = 'cancelled' THEN 1 ELSE 0 END) as cancelled_orders,
                SUM(CASE WHEN COALESCE(is_undelivered, 0) = 1 THEN 1 ELSE 0 END) as undelivered_orders,
                COALESCE(SUM(CASE WHEN COALESCE(is_undelivered, 0) = 1
                                  THEN COALESCE(shipping_loss_amount, 0) ELSE 0 END), 0) as shipping_loss
            FROM orders
            WHERE source IN ({placeholders}) AND currency = ? {date_cond}
            GROUP BY source
            ORDER BY success_amount DESC
        ''', params).fetchall()

        rate, _ = _lookup_partner_rate(currency, year or 0, month or 0)

        result_rows = []
        totals = {
            'total_orders': 0, 'total_amount': 0, 'success_orders': 0,
            'success_amount': 0, 'success_shipping': 0, 'success_net': 0,
            'failed_orders': 0, 'cancelled_orders': 0,
            'undelivered_orders': 0, 'shipping_loss': 0,
            'success_net_cny': 0,
        }
        for r in rows:
            success_net = float(r['success_amount']) - float(r['success_shipping']) - float(r['shipping_loss'])
            success_net_cny = success_net * rate if rate else None
            row = {
                'source': r['source'],
                'country': site_meta.get(r['source'], {}).get('country', ''),
                'manager': site_meta.get(r['source'], {}).get('manager', ''),
                'currency': currency,
                'total_orders': r['total_orders'],
                'total_amount': float(r['total_amount'] or 0),
                'success_orders': r['success_orders'],
                'success_amount': float(r['success_amount'] or 0),
                'success_shipping': float(r['success_shipping'] or 0),
                'success_net': round(success_net, 2),
                'failed_orders': r['failed_orders'],
                'cancelled_orders': r['cancelled_orders'],
                'undelivered_orders': r['undelivered_orders'],
                'shipping_loss': float(r['shipping_loss'] or 0),
                'rate_to_cny': rate,
                'success_net_cny': round(success_net_cny, 2) if success_net_cny is not None else None,
            }
            result_rows.append(row)
            totals['total_orders'] += row['total_orders']
            totals['total_amount'] += row['total_amount']
            totals['success_orders'] += row['success_orders']
            totals['success_amount'] += row['success_amount']
            totals['success_shipping'] += row['success_shipping']
            totals['success_net'] += row['success_net']
            totals['failed_orders'] += row['failed_orders']
            totals['cancelled_orders'] += row['cancelled_orders']
            totals['undelivered_orders'] += row['undelivered_orders']
            totals['shipping_loss'] += row['shipping_loss']
            if success_net_cny is not None:
                totals['success_net_cny'] += success_net_cny
        for k in totals:
            totals[k] = round(totals[k], 2) if isinstance(totals[k], float) else totals[k]

        return jsonify({
            'rows': result_rows,
            'totals': totals,
            'currency': currency,
            'rate_to_cny': rate,
        })
    finally:
        conn.close()


@app.route('/api/reconciliation/products')
@login_required
@reconciliation_api_required
def api_recon_products():
    """Per-product breakdown for the period: qty, revenue, cost (date-aware), margin.
    Optional ?source=<site_url> and ?manager=<name> narrow the scope to a
    single site or a single team member (for selection-decision analysis)."""
    partner_id = request.args.get('partner_id', type=int)
    year = request.args.get('year', type=int)
    month = request.args.get('month', type=int)
    source_filter = (request.args.get('source') or '').strip() or None
    manager_filter = (request.args.get('manager') or '').strip() or None
    if not all([partner_id, year, month]):
        return jsonify({'error': '缺少 partner_id/year/month'}), 400
    if not _check_partner_access(partner_id):
        return jsonify({'error': '无权查看此合伙人'}), 403

    detail = _calc_partner_recon_detail(partner_id, year, month,
                                        site_filter=source_filter,
                                        manager_filter=manager_filter)
    if detail is None:
        return jsonify({'error': '合伙人不存在'}), 404
    # Sum allocated shipping across all products — should equal the sum of
    # shipping_total for the period's successful orders. Useful as a reference
    # so partners can sanity-check the allocation.
    products = detail['by_product']
    total_allocated_shipping = round(sum(p.get('allocated_shipping', 0) or 0 for p in products), 2)
    total_cost_mapped = round(sum(p.get('cost_mapped', 0) or 0 for p in products), 2)
    total_cost_estimated = round(sum(p.get('cost_estimated', 0) or 0 for p in products), 2)
    return jsonify({
        'partner_id': partner_id,
        'period_year': year,
        'period_month': month,
        'currency': detail['currency'],
        'products': products,
        'totals': {
            'revenue': detail['total_gross_pln'] - detail['total_shipping_pln'],
            'actual_cost': detail['actual_cost_pln'],
            'cost_mapped': total_cost_mapped,        # real product_costs portion
            'cost_estimated': total_cost_estimated,  # 50% fallback portion
            'allocated_shipping': total_allocated_shipping,
            'unmapped_revenue': detail['cost_unmapped_revenue_pln'],
            'unmapped_qty': detail['cost_unmapped_qty'],
        },
    })


@app.route('/api/reconciliation/unmapped-products')
@login_required
@reconciliation_api_required
def api_recon_unmapped_products():
    """Aggregate unmapped products (no cost row matching brand+series+puffs at
    order date) for the partner's bound sites in the given month.

    Returns one row per (brand, puffs, country, flavor) group, with raw
    product names underneath. Mirrors /api/sales-board/unmapped's shape so
    the frontend modal stays familiar.
    """
    partner_id = request.args.get('partner_id', type=int)
    year = request.args.get('year', type=int)
    month = request.args.get('month', type=int)
    if not all([partner_id, year, month]):
        return jsonify({'error': '缺少 partner_id/year/month'}), 400
    if not _check_partner_access(partner_id):
        return jsonify({'error': '无权查看此合伙人'}), 403

    conn = get_db_connection()
    try:
        partner = conn.execute('SELECT id, currency FROM partners WHERE id = ?', (partner_id,)).fetchone()
        if not partner:
            return jsonify({'error': '合伙人不存在'}), 404
        currency = partner['currency'] or 'PLN'

        sites = conn.execute('''
            SELECT s.url, s.country, s.manager
            FROM partner_sites ps
            JOIN sites s ON s.id = ps.site_id
            WHERE ps.partner_id = ?
        ''', (partner_id,)).fetchall()
        if not sites:
            return jsonify({'partner_id': partner_id, 'currency': currency, 'unmapped': []})
        site_urls = [s['url'] for s in sites]
        site_country_map = {s['url']: s['country'] or 'PL' for s in sites}
        site_manager_map = {s['url']: s['manager'] or '' for s in sites}

        placeholders = ','.join(['?'] * len(site_urls))
        month_str = f'{year:04d}-{month:02d}'

        # Pull line_items from this period's revenue orders only
        orders = conn.execute(f'''
            SELECT id, line_items, source, currency, warehouse_id, date_created
            FROM orders
            WHERE source IN ({placeholders})
            AND currency = ?
            AND strftime('%Y-%m', date_created) = ?
            AND {_revenue_status_cond()}
        ''', site_urls + [currency, month_str]).fetchall()

        # Build cost index + brand/product-mappings caches
        cost_idx = _build_dated_cost_index(conn=conn)
        country_default_wh_ids = {}
        for w in conn.execute('SELECT id, country FROM warehouses ORDER BY country, id').fetchall():
            country_default_wh_ids.setdefault(w['country'], []).append(w['id'])
        brands_rows = conn.execute('SELECT id, name, aliases FROM brands').fetchall()
        brands_cache = []
        for row in brands_rows:
            try:
                aliases = json.loads(row['aliases']) if row['aliases'] else []
            except Exception:
                aliases = []
            brands_cache.append({
                'id': row['id'], 'name': row['name'], 'aliases': aliases,
                'patterns': [row['name'].upper()] + [a.upper() for a in aliases]
            })
        pm_rows = conn.execute('SELECT raw_name, source, brand_id, series_id, puff_count, flavor FROM product_mappings').fetchall()
        product_mappings_cache = {}
        for pm in pm_rows:
            product_mappings_cache[(normalize_raw_name(pm['raw_name']), pm['source'])] = pm

        # CNY rate for the period — for revenue conversion
        rate_cny, _ = _lookup_partner_rate(currency, year, month)

        # Group unmapped by (brand, puffs, country, flavor)
        unmapped = {}  # key -> {brand, puffs, country, flavor, qty, revenue, revenue_cny, managers, products{name -> {qty, revenue}}}
        for o in orders:
            items = parse_json_field(o['line_items'])
            if not isinstance(items, list):
                continue
            order_country = site_country_map.get(o['source'], 'PL') or 'PL'
            order_manager = site_manager_map.get(o['source'], '')
            order_date = (o['date_created'] or '')[:10]
            effective_wh_ids = country_default_wh_ids.get(order_country, [])

            for item in items:
                qty = int(item.get('quantity', 0) or 0)
                if qty <= 0:
                    continue
                raw_name = item.get('name', '') or ''
                item_total = float(item.get('total', 0) or 0)
                b_id, s_id, p_cnt, flav = _resolve_product_to_brand(
                    raw_name, o['source'], brands_cache, product_mappings_cache
                )
                cost_entry = None
                if b_id and effective_wh_ids:
                    for ewh in effective_wh_ids:
                        cost_entry = _cost_at_date(cost_idx, b_id, s_id, p_cnt, flav, ewh, order_date)
                        if cost_entry:
                            break
                if cost_entry:
                    continue  # mapped — skip

                # ── unmapped ──
                brand_label = None
                if b_id:
                    for bc in brands_cache:
                        if bc['id'] == b_id:
                            brand_label = bc['name']
                            break
                brand_label = brand_label or 'Unknown'
                key = (brand_label, p_cnt or 0, order_country, (flav or '').lower())
                if key not in unmapped:
                    unmapped[key] = {
                        'brand': brand_label, 'puffs': p_cnt, 'flavor': flav,
                        'country': order_country,
                        'qty': 0, 'revenue': 0.0, 'revenue_cny': 0.0,
                        'managers': set(),
                        'products': {},
                    }
                u = unmapped[key]
                u['qty'] += qty
                u['revenue'] += item_total
                if rate_cny:
                    u['revenue_cny'] += item_total * rate_cny
                if order_manager:
                    u['managers'].add(order_manager)
                # raw product names with qty / revenue
                if raw_name not in u['products']:
                    u['products'][raw_name] = {'name': raw_name, 'qty': 0, 'revenue': 0.0, 'revenue_cny': 0.0}
                p = u['products'][raw_name]
                p['qty'] += qty
                p['revenue'] += item_total
                if rate_cny:
                    p['revenue_cny'] += item_total * rate_cny

        # Flatten to list, sort by revenue desc
        result = []
        for u in unmapped.values():
            u['managers'] = sorted(u['managers'])
            u['products'] = sorted(u['products'].values(), key=lambda p: -p['revenue'])
            for p in u['products']:
                p['revenue'] = round(p['revenue'], 2)
                p['revenue_cny'] = round(p['revenue_cny'], 2)
            u['revenue'] = round(u['revenue'], 2)
            u['revenue_cny'] = round(u['revenue_cny'], 2)
            result.append(u)
        result.sort(key=lambda u: -u['revenue'])

        return jsonify({
            'partner_id': partner_id,
            'currency': currency,
            'rate_to_cny': rate_cny,
            'period': month_str,
            'unmapped': result,
        })
    finally:
        conn.close()


# ============================================================================
# DASHBOARD APIs (P1)
# ============================================================================
# These power the "数据看板" tab. Each is per-partner and reuses
# _calc_partner_recon_detail / _calc_partner_net_sales for the heavy lifting.

@app.route('/api/reconciliation/dashboard/monthly-trend')
@login_required
@reconciliation_api_required
def api_recon_dashboard_monthly_trend():
    """Per-partner trend across the last N months.

    Returns one row per month with: net sales, contract cost, actual cost
    (date-aware), our receivable, money received, success/failed/cancelled/
    undelivered counts, shipping_loss. The frontend renders this as a
    multi-line chart.
    """
    partner_id = request.args.get('partner_id', type=int)
    months = max(3, min(24, request.args.get('months', 12, type=int)))
    if not partner_id:
        return jsonify({'error': '缺少 partner_id'}), 400
    if not _check_partner_access(partner_id):
        return jsonify({'error': '无权查看此合伙人'}), 403

    conn = get_db_connection()
    try:
        partner = conn.execute('SELECT id, currency, cost_ratio, partner_profit_ratio, our_profit_ratio FROM partners WHERE id = ?', (partner_id,)).fetchone()
        if not partner:
            return jsonify({'error': '合伙人不存在'}), 404
        currency = partner['currency'] or 'PLN'

        sites = conn.execute('''SELECT s.url FROM partner_sites ps
            JOIN sites s ON s.id = ps.site_id WHERE ps.partner_id = ?''', (partner_id,)).fetchall()
        if not sites:
            return jsonify({'partner_id': partner_id, 'currency': currency, 'months': []})
        site_urls = [s['url'] for s in sites]
        placeholders = ','.join(['?'] * len(site_urls))

        # Build month list [today_back_to N months ago]
        from datetime import date
        today = date.today()
        month_list = []
        y, m = today.year, today.month
        for _ in range(months):
            month_list.append((y, m))
            m -= 1
            if m == 0:
                m = 12
                y -= 1
        month_list.reverse()  # oldest first

        # Single grouped SQL — one trip to DB instead of N queries
        # (returns gross / shipping / shipping_loss / counts per (year, month))
        success_cond = _revenue_status_cond()
        ym_min = f'{month_list[0][0]:04d}-{month_list[0][1]:02d}'
        ym_max = f'{month_list[-1][0]:04d}-{month_list[-1][1]:02d}'
        rows = conn.execute(f'''
            SELECT
                strftime('%Y-%m', date_created) AS ym,
                COUNT(*) AS total_count,
                SUM(CASE WHEN {success_cond} THEN 1 ELSE 0 END) AS success_count,
                SUM(CASE WHEN status='failed' THEN 1 ELSE 0 END) AS failed_count,
                SUM(CASE WHEN status='cancelled' THEN 1 ELSE 0 END) AS cancelled_count,
                SUM(CASE WHEN COALESCE(is_undelivered, 0) = 1 THEN 1 ELSE 0 END) AS undel_count,
                COALESCE(SUM(CASE WHEN {success_cond} THEN total ELSE 0 END), 0) AS gross,
                COALESCE(SUM(CASE WHEN {success_cond} THEN shipping_total ELSE 0 END), 0) AS ship,
                COALESCE(SUM(CASE WHEN COALESCE(is_undelivered, 0) = 1
                                  THEN COALESCE(shipping_loss_amount, 0) ELSE 0 END), 0) AS loss
            FROM orders
            WHERE source IN ({placeholders})
              AND currency = ?
              AND strftime('%Y-%m', date_created) BETWEEN ? AND ?
            GROUP BY ym
            ORDER BY ym
        ''', site_urls + [currency, ym_min, ym_max]).fetchall()
        agg = {r['ym']: r for r in rows}

        # Receipts per month (sum of partner_receipts.amount_pln by year-month)
        rcpt_rows = conn.execute(f'''
            SELECT strftime('%Y-%m', receipt_date) AS ym,
                   COALESCE(SUM(amount_pln), 0) AS received
            FROM partner_receipts
            WHERE partner_id = ?
              AND strftime('%Y-%m', receipt_date) BETWEEN ? AND ?
            GROUP BY ym
        ''', (partner_id, ym_min, ym_max)).fetchall()
        rcpt_map = {r['ym']: float(r['received'] or 0) for r in rcpt_rows}

        cr  = float(partner['cost_ratio']) if partner['cost_ratio'] is not None else 0.5
        opr = float(partner['our_profit_ratio']) if partner['our_profit_ratio'] is not None else 0.25

        # Build full series — fill 0s for empty months. Actual cost computation
        # is expensive (per-line product lookup) so we only call the detail
        # function for months that actually have orders.
        series = []
        for (yy, mm) in month_list:
            ym = f'{yy:04d}-{mm:02d}'
            r = agg.get(ym)
            if r is None or (r['success_count'] or 0) == 0:
                series.append({
                    'year_month': ym, 'year': yy, 'month': mm,
                    'success_orders': 0, 'failed_orders': 0,
                    'cancelled_orders': 0, 'undelivered_orders': 0,
                    'gross': 0, 'shipping': 0, 'shipping_loss': 0,
                    'net': 0, 'contract_cost': 0, 'actual_cost': 0,
                    'our_receivable': 0, 'received': round(rcpt_map.get(ym, 0), 2),
                    'undelivered_rate': 0.0, 'failed_rate': 0.0,
                })
                continue
            gross = float(r['gross'] or 0)
            ship = float(r['ship'] or 0)
            loss = float(r['loss'] or 0)
            net = gross - ship - loss
            total = int(r['total_count'] or 0)
            undel_rate = (int(r['undel_count'] or 0) / total * 100) if total > 0 else 0
            failed_rate = (int(r['failed_count'] or 0) / total * 100) if total > 0 else 0
            # actual_cost_pln from full detail calc (date-aware)
            detail = _calc_partner_recon_detail(partner_id, yy, mm)
            actual_cost = detail['actual_cost_pln'] if detail else 0
            series.append({
                'year_month': ym, 'year': yy, 'month': mm,
                'success_orders': int(r['success_count'] or 0),
                'failed_orders': int(r['failed_count'] or 0),
                'cancelled_orders': int(r['cancelled_count'] or 0),
                'undelivered_orders': int(r['undel_count'] or 0),
                'gross': round(gross, 2),
                'shipping': round(ship, 2),
                'shipping_loss': round(loss, 2),
                'net': round(net, 2),
                # Contract basis: cost coverage on successful net (net+loss), our
                # receivable also bears half the undelivered loss (kept in sync
                # with _compute_statement_split / _calc_partner_recon_detail).
                'contract_cost': round((net + loss) * cr, 2),
                'actual_cost': round(actual_cost, 2),
                'our_receivable': round((net + loss) * opr - loss / 2, 2),
                'received': round(rcpt_map.get(ym, 0), 2),
                'undelivered_rate': round(undel_rate, 2),
                'failed_rate': round(failed_rate, 2),
            })

        return jsonify({
            'partner_id': partner_id,
            'currency': currency,
            'months': series,
        })
    finally:
        conn.close()


@app.route('/api/reconciliation/dashboard/aging')
@login_required
@reconciliation_api_required
def api_recon_dashboard_aging():
    """Aging analysis for unpaid statements.

    Each statement that has outstanding balance (our_receivable - received > 0)
    is bucketed by days since the period end:
        0-30, 30-60, 60-90, 90+ days.
    Frontend renders as a bar chart + drilldown table.
    """
    partner_id = request.args.get('partner_id', type=int)
    if not partner_id:
        return jsonify({'error': '缺少 partner_id'}), 400
    if not _check_partner_access(partner_id):
        return jsonify({'error': '无权查看此合伙人'}), 403

    conn = get_db_connection()
    try:
        partner = conn.execute('SELECT id, name, currency FROM partners WHERE id = ?', (partner_id,)).fetchone()
        if not partner:
            return jsonify({'error': '合伙人不存在'}), 404
        currency = partner['currency'] or 'PLN'

        # Each statement + sum of its receipts
        rows = conn.execute('''
            SELECT s.id, s.period_year, s.period_month, s.our_receivable_pln,
                   s.exchange_rate_cny, s.status, s.created_at,
                   COALESCE(SUM(pr.amount_pln), 0) AS received
            FROM reconciliation_statements s
            LEFT JOIN partner_receipts pr ON pr.statement_id = s.id
            WHERE s.partner_id = ?
            GROUP BY s.id
            ORDER BY s.period_year, s.period_month
        ''', (partner_id,)).fetchall()

        from datetime import date
        today = date.today()
        buckets = {
            '0-30':  {'count': 0, 'amount': 0.0, 'amount_cny': 0.0, 'statements': []},
            '30-60': {'count': 0, 'amount': 0.0, 'amount_cny': 0.0, 'statements': []},
            '60-90': {'count': 0, 'amount': 0.0, 'amount_cny': 0.0, 'statements': []},
            '90+':   {'count': 0, 'amount': 0.0, 'amount_cny': 0.0, 'statements': []},
        }
        total_outstanding = 0.0
        total_outstanding_cny = 0.0

        for r in rows:
            owed = float(r['our_receivable_pln'] or 0) - float(r['received'] or 0)
            if owed <= 0.01:  # already settled
                continue
            # Period end date = last day of the month
            import calendar
            last_day = calendar.monthrange(r['period_year'], r['period_month'])[1]
            period_end = date(r['period_year'], r['period_month'], last_day)
            age_days = (today - period_end).days
            if age_days < 0:
                # Period not over yet — bucket as 0-30 (or skip; we keep it for visibility)
                age_days = 0
            if age_days <= 30:
                bk = '0-30'
            elif age_days <= 60:
                bk = '30-60'
            elif age_days <= 90:
                bk = '60-90'
            else:
                bk = '90+'

            rate = r['exchange_rate_cny']
            if not rate:
                rate, _ = _lookup_partner_rate(currency, r['period_year'], r['period_month'])
            owed_cny = owed * rate if rate else 0

            buckets[bk]['count'] += 1
            buckets[bk]['amount'] += owed
            buckets[bk]['amount_cny'] += owed_cny
            buckets[bk]['statements'].append({
                'id': r['id'],
                'period': f"{r['period_year']}-{r['period_month']:02d}",
                'our_receivable': float(r['our_receivable_pln'] or 0),
                'received': float(r['received'] or 0),
                'outstanding': round(owed, 2),
                'outstanding_cny': round(owed_cny, 2) if rate else None,
                'age_days': age_days,
                'status': r['status'],
            })
            total_outstanding += owed
            total_outstanding_cny += owed_cny

        for bk in buckets.values():
            bk['amount'] = round(bk['amount'], 2)
            bk['amount_cny'] = round(bk['amount_cny'], 2)

        return jsonify({
            'partner_id': partner_id,
            'currency': currency,
            'today': today.isoformat(),
            'buckets': buckets,
            'total_outstanding': round(total_outstanding, 2),
            'total_outstanding_cny': round(total_outstanding_cny, 2),
        })
    finally:
        conn.close()


@app.route('/api/reconciliation/dashboard/comparison')
@login_required
@reconciliation_api_required
def api_recon_dashboard_comparison():
    """Current month vs previous month vs same-month-last-year.

    Returns side-by-side metrics for trend-arrow display in the dashboard.
    """
    partner_id = request.args.get('partner_id', type=int)
    year = request.args.get('year', type=int)
    month = request.args.get('month', type=int)
    if not all([partner_id, year, month]):
        return jsonify({'error': '缺少 partner_id/year/month'}), 400
    if not _check_partner_access(partner_id):
        return jsonify({'error': '无权查看此合伙人'}), 403

    def shift_month(y, m, delta):
        idx = (y * 12 + (m - 1)) + delta
        return idx // 12, (idx % 12) + 1

    cur_y, cur_m = year, month
    prev_y, prev_m = shift_month(cur_y, cur_m, -1)
    yoy_y, yoy_m = cur_y - 1, cur_m

    cur = _calc_partner_recon_detail(partner_id, cur_y, cur_m)
    if cur is None:
        return jsonify({'error': '合伙人不存在'}), 404
    prev = _calc_partner_recon_detail(partner_id, prev_y, prev_m) or {}
    yoy = _calc_partner_recon_detail(partner_id, yoy_y, yoy_m) or {}

    def metric_block(d):
        return {
            'success_orders': d.get('success_orders', 0),
            'undelivered_orders': d.get('undelivered_orders', 0),
            'total_gross_pln': d.get('total_gross_pln', 0),
            'total_net_pln':   d.get('total_net_pln', 0),
            'actual_cost_pln': d.get('actual_cost_pln', 0),
            'our_receivable_pln': d.get('our_receivable_pln', 0),
        }

    def diff(a, b):
        if not b:
            return {'abs': a, 'pct': None}
        return {'abs': round(a - b, 2), 'pct': round((a - b) / b * 100, 1) if b else None}

    return jsonify({
        'partner_id': partner_id,
        'currency': cur['currency'],
        'current':  {'period': f'{cur_y}-{cur_m:02d}',  'metrics': metric_block(cur)},
        'previous': {'period': f'{prev_y}-{prev_m:02d}', 'metrics': metric_block(prev)},
        'yoy':      {'period': f'{yoy_y}-{yoy_m:02d}',  'metrics': metric_block(yoy)},
        'mom_diff': {  # current vs previous-month
            k: diff(metric_block(cur)[k], metric_block(prev).get(k, 0))
            for k in metric_block(cur).keys()
        },
        'yoy_diff': {  # current vs same-month-last-year
            k: diff(metric_block(cur)[k], metric_block(yoy).get(k, 0))
            for k in metric_block(cur).keys()
        },
    })


def _compute_statement_split(net, actual_cost, cost_ratio, pp_ratio, op_ratio, mode='contract', shipping_loss=0):
    """Return (cost_amount, partner_profit, our_receivable) for a statement.

    The undelivered shipping loss (未送达运费损失) is a SHARED cost borne 50/50 by
    the partner and us in BOTH modes — it is never a product cost, so it never
    shrinks the partner's cost coverage.

    'contract' (约定毛利): split the successful net (净销售 + 未送达损失 = 总销售 −
        运费成本) by the fixed ratios, then charge each side half the loss:
          cost    = 成功净额 × cost_ratio
          partner = 成功净额 × pp_ratio − 未送达损失 ÷ 2
          ours    = 成功净额 × op_ratio − 未送达损失 ÷ 2
        With no undelivered loss this reduces to the plain 净销售 × ratio split.
    'actual' (实际毛利): the partner first recovers their real product cost, then
        the remaining margin (净销售 − 实际成本) is split between partner and us in
        proportion to (pp_ratio : op_ratio). The loss already sits inside 净销售,
        so it is split pp:op (= 50/50 when pp==op) automatically — no separate
        handling needed. Falls back to 50/50 when pp+op == 0.

    cost + partner + ours always reconciles to `net` (the money actually
    distributed) when the ratios sum to 1.

    None ratios mean "unset" → contract defaults (0.5 / 0.25 / 0.25); a real 0 is kept.
    """
    net = net or 0
    loss = shipping_loss or 0
    cr = 0.5 if cost_ratio is None else float(cost_ratio)
    pp = 0.25 if pp_ratio is None else float(pp_ratio)
    op = 0.25 if op_ratio is None else float(op_ratio)
    if mode == 'actual':
        cost = round(actual_cost or 0, 2)
        remainder = net - cost
        denom = pp + op
        pp_share, op_share = (pp / denom, op / denom) if denom > 0 else (0.5, 0.5)
        return cost, round(remainder * pp_share, 2), round(remainder * op_share, 2)
    # contract: carve the undelivered loss out, split the successful net by the
    # ratios, then split the loss 50/50 across the two profit parties.
    succ_net = net + loss
    half = loss / 2.0
    return round(succ_net * cr, 2), round(succ_net * pp - half, 2), round(succ_net * op - half, 2)


@app.route('/api/reconciliation/statements/generate', methods=['POST'])
@login_required
@reconciliation_api_required
def api_generate_statement():
    """Generate (or refresh) a monthly statement from order data"""
    if not _is_reconciliation_admin():
        return jsonify({'error': '无权生成对账单'}), 403
    data = request.json
    partner_id = int(data.get('partner_id'))
    year = int(data.get('year'))
    month = int(data.get('month'))

    conn = get_db_connection()
    partner = conn.execute('SELECT * FROM partners WHERE id = ?', (partner_id,)).fetchone()
    if not partner:
        conn.close()
        return jsonify({'error': '合伙人不存在'}), 404

    # Check if already locked/settled
    existing = conn.execute('''SELECT id, status FROM reconciliation_statements
        WHERE partner_id=? AND period_year=? AND period_month=?''',
        (partner_id, year, month)).fetchone()
    if existing and existing['status'] in ('locked', 'settled'):
        conn.close()
        return jsonify({'error': '该月份对账单已锁定，无法重新生成'}), 400

    # Use the full detail aggregator so we also capture actual_cost_pln_snapshot
    detail = _calc_partner_recon_detail(partner_id, year, month)
    if detail is None:
        conn.close()
        return jsonify({'error': '合伙人数据计算失败'}), 500
    net = detail['total_net_pln']
    actual_cost_snapshot = detail['actual_cost_pln']

    # Guard against generating an EMPTY statement. If the chosen period has no
    # successful (revenue-generating) orders, total_orders/net are 0 and every
    # split below comes out all-zeros. Silently saving that produces a blank
    # "bill" that reports success but shows nothing — exactly the
    # "按照实际毛利生成的账单是空的" report. This happens when a period with no
    # orders is picked (the year dropdown offers future / pre-launch months),
    # or while order data is briefly unavailable. Refuse instead of writing the
    # empty row (and, on regenerate, refuse so we never clobber a good
    # statement with zeros). A genuinely zero-activity month never needs an
    # auto-generated statement; 手工录入 covers that rare case.
    if (detail.get('total_orders') or 0) == 0 and (net or 0) == 0:
        conn.close()
        return jsonify({'error': f'{year}-{month:02d} 期间没有可对账的成功订单（净销售额为 0），未生成对账单。请确认所选合伙人与月份正确、且订单已同步后重试。'}), 400

    mode = data.get('mode') or 'contract'
    if mode not in ('contract', 'actual'):
        mode = 'contract'
    cost, p_profit, our_recv = _compute_statement_split(
        net, actual_cost_snapshot, partner['cost_ratio'],
        partner['partner_profit_ratio'], partner['our_profit_ratio'], mode,
        shipping_loss=detail.get('shipping_loss') or 0)
    # In 实际毛利 mode the margin is only trustworthy if every product has a real
    # cost; unmapped lines fall back to a 50% estimate (would inflate margin).
    unmapped_qty = detail.get('cost_unmapped_qty') or 0

    # Auto-lookup exchange rate from system settings
    rate, _ = _lookup_partner_rate(partner['currency'], year, month)
    our_recv_cny = round(our_recv * rate, 2) if rate else None

    try:
        is_regenerate = bool(existing)
        if existing:
            conn.execute('''UPDATE reconciliation_statements
                SET total_orders=?, total_gross_pln=?, total_net_pln=?,
                    cost_amount_pln=?, partner_profit_pln=?, our_receivable_pln=?,
                    exchange_rate_cny=?, our_receivable_cny=?,
                    actual_cost_pln_snapshot=?, calc_mode=?,
                    status='generated', updated_at=CURRENT_TIMESTAMP
                WHERE id=?''',
                (detail['total_orders'], detail['total_gross_pln'], net,
                 cost, p_profit, our_recv, rate, our_recv_cny,
                 actual_cost_snapshot, mode, existing['id']))
            stmt_id = existing['id']
        else:
            cursor = conn.execute('''INSERT INTO reconciliation_statements
                (partner_id, period_year, period_month, total_orders, total_gross_pln, total_net_pln,
                 cost_amount_pln, partner_profit_pln, our_receivable_pln,
                 exchange_rate_cny, our_receivable_cny, actual_cost_pln_snapshot, calc_mode, status)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 'generated')''',
                (partner_id, year, month, detail['total_orders'], detail['total_gross_pln'], net,
                 cost, p_profit, our_recv, rate, our_recv_cny,
                 actual_cost_snapshot, mode))
            stmt_id = cursor.lastrowid

        # P2: snapshot the orders that contributed to this statement
        snapshot_count = _snapshot_statement_orders(stmt_id, partner_id, year, month, conn)

        # P2: audit log
        mode_label = '实际毛利' if mode == 'actual' else '约定毛利'
        _audit_log(stmt_id, 'regenerate' if is_regenerate else 'create',
                   note=f'口径={mode_label}, 净销售={net} {partner["currency"] or "PLN"}, 实际成本={actual_cost_snapshot}, 快照订单 {snapshot_count} 条', conn=conn)

        conn.commit()
        return jsonify({'success': True, 'id': stmt_id, 'exchange_rate_cny': rate,
                        'snapshot_orders': snapshot_count, 'actual_cost_pln': actual_cost_snapshot,
                        'calc_mode': mode, 'unmapped_qty': unmapped_qty})
    finally:
        conn.close()


@app.route('/api/reconciliation/statements/manual', methods=['POST'])
@login_required
@reconciliation_api_required
def api_manual_statement():
    """Manually create a statement (for historical data before 2026)"""
    if not _is_reconciliation_admin():
        return jsonify({'error': '无权录入对账单'}), 403
    data = request.json
    partner_id = int(data.get('partner_id'))
    year = int(data.get('year'))
    month = int(data.get('month'))
    net = float(data.get('total_net_pln', 0))

    conn = get_db_connection()
    partner = conn.execute('SELECT * FROM partners WHERE id = ?', (partner_id,)).fetchone()
    if not partner:
        conn.close()
        return jsonify({'error': '合伙人不存在'}), 404

    # Manual entries are historical (no order data → no real cost) → contract basis.
    cost, p_profit, our_recv = _compute_statement_split(
        net, 0, partner['cost_ratio'],
        partner['partner_profit_ratio'], partner['our_profit_ratio'], 'contract')

    # Auto-lookup exchange rate from system settings
    rate, _ = _lookup_partner_rate(partner['currency'], year, month)
    our_recv_cny = round(our_recv * rate, 2) if rate else None

    try:
        conn.execute('''INSERT OR REPLACE INTO reconciliation_statements
            (partner_id, period_year, period_month, total_orders, total_gross_pln, total_net_pln,
             cost_amount_pln, partner_profit_pln, our_receivable_pln,
             exchange_rate_cny, our_receivable_cny, calc_mode, status, is_manual, notes)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 'contract', 'generated', 1, ?)''',
            (partner_id, year, month,
             int(data.get('total_orders', 0)),
             float(data.get('total_gross_pln', net)),
             net, cost, p_profit, our_recv,
             rate, our_recv_cny,
             data.get('notes', '手工录入历史数据')))
        conn.commit()
        return jsonify({'success': True, 'exchange_rate_cny': rate})
    finally:
        conn.close()


@app.route('/api/reconciliation/statements/<int:stmt_id>', methods=['GET'])
@login_required
@reconciliation_api_required
def api_get_statement(stmt_id):
    """Get statement detail"""
    conn = get_db_connection()
    stmt = conn.execute('''
        SELECT s.*, p.name as partner_name, p.currency
        FROM reconciliation_statements s
        JOIN partners p ON s.partner_id = p.id
        WHERE s.id = ?
    ''', (stmt_id,)).fetchone()
    if not stmt:
        conn.close()
        return jsonify({'error': '对账单不存在'}), 404
    if not _check_partner_access(stmt['partner_id']):
        conn.close()
        return jsonify({'error': '无权查看'}), 403
    receipts = conn.execute('''
        SELECT * FROM partner_receipts WHERE statement_id = ? ORDER BY receipt_date DESC
    ''', (stmt_id,)).fetchall()
    conn.close()
    result = dict(stmt)
    _enrich_statement_cny(result, stmt['currency'] or 'PLN')
    rate = result.get('effective_rate_cny')
    result['receipts'] = [dict(r) for r in receipts]
    result['total_received_pln'] = round(sum(r['amount_pln'] or 0 for r in receipts), 2)
    result['outstanding_pln'] = round((stmt['our_receivable_pln'] or 0) - result['total_received_pln'], 2)
    if rate:
        result['total_received_cny'] = round(result['total_received_pln'] * rate, 2)
        result['outstanding_cny'] = round(result['outstanding_pln'] * rate, 2)

    # Attach live drill-down detail (status / site / product breakdown) so the
    # detail modal can show how the saved totals were composed. The saved
    # *_pln fields above are the locked snapshot — drill-down is recomputed
    # from current data; if orders changed since generation the live numbers
    # may differ, which is intentional (user wants to see what's there now).
    if not stmt['is_manual']:
        try:
            live_detail = _calc_partner_recon_detail(
                stmt['partner_id'], stmt['period_year'], stmt['period_month']
            )
            if live_detail:
                result['live_detail'] = {
                    'success_orders': live_detail['success_orders'],
                    'failed_orders': live_detail['failed_orders'],
                    'cancelled_orders': live_detail['cancelled_orders'],
                    'undelivered_orders': live_detail['undelivered_orders'],
                    'pending_orders': live_detail['pending_orders'],
                    'shipping_loss': live_detail['shipping_loss'],
                    'actual_cost_pln': live_detail['actual_cost_pln'],
                    # Contract-basis cost reference (净销售 × cost_ratio, loss carved
                    # out) — needed so the 约定 vs 实际 差额 in the detail modal compares
                    # against the CONTRACT cost, not the statement's stored cost
                    # (which equals the actual cost on an 实际毛利 statement → diff 0).
                    'cost_amount_pln': live_detail['cost_amount_pln'],
                    'cost_unmapped_revenue_pln': live_detail['cost_unmapped_revenue_pln'],
                    'cost_unmapped_qty': live_detail['cost_unmapped_qty'],
                    'by_status': live_detail['by_status'],
                    'by_site': live_detail['by_site'],
                    'by_product': live_detail['by_product'],
                }
        except Exception as e:
            app.logger.warning(f"live_detail failed for stmt {stmt_id}: {e}")
    return jsonify(result)


@app.route('/api/reconciliation/statements/<int:stmt_id>/export', methods=['GET'])
@login_required
@reconciliation_api_required
def api_export_statement(stmt_id):
    """Export one reconciliation statement as a styled multi-sheet .xlsx.

    Sheets: 对账单汇总 / 按站点 / 按产品 / 收款记录. Read-only — access mirrors
    the detail view (via _check_partner_access), so a partner can export their
    own statement. Top-line amounts come from the saved snapshot; the
    by-site / by-product drill-down is recomputed live, exactly like the
    detail modal. Manual statements have no live drill-down → only the
    summary (+ receipts) sheet is produced.
    """
    from io import BytesIO
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    conn = get_db_connection()
    stmt = conn.execute('''
        SELECT s.*, p.name AS partner_name, p.currency
        FROM reconciliation_statements s
        JOIN partners p ON s.partner_id = p.id
        WHERE s.id = ?
    ''', (stmt_id,)).fetchone()
    if not stmt:
        conn.close()
        return jsonify({'error': '对账单不存在'}), 404
    if not _check_partner_access(stmt['partner_id']):
        conn.close()
        return jsonify({'error': '无权查看'}), 403
    receipts = conn.execute(
        'SELECT * FROM partner_receipts WHERE statement_id = ? ORDER BY receipt_date DESC',
        (stmt_id,)
    ).fetchall()
    conn.close()

    data = dict(stmt)
    currency = data.get('currency') or 'PLN'
    _enrich_statement_cny(data, currency)
    rate = data.get('effective_rate_cny')
    total_received = round(sum((r['amount_pln'] or 0) for r in receipts), 2)
    outstanding = round((data.get('our_receivable_pln') or 0) - total_received, 2)
    received_cny = round(total_received * rate, 2) if rate else None
    outstanding_cny = round(outstanding * rate, 2) if rate else None

    # Live drill-down (status / site / product) — only for system-generated stmts
    live = None
    if not data.get('is_manual'):
        try:
            live = _calc_partner_recon_detail(
                data['partner_id'], data['period_year'], data['period_month'])
        except Exception as e:
            app.logger.warning(f"export: live detail failed for stmt {stmt_id}: {e}")

    STATUS_LABELS = {
        'draft': '草稿', 'generated': '已生成', 'disputed': '合伙人有异议',
        'confirmed': '合伙人已确认', 'locked': '已锁定', 'settled': '已结清',
    }

    # ---- shared styles ----
    FONT = '微软雅黑'
    title_font = Font(name=FONT, size=15, bold=True, color='FFFFFF')
    title_fill = PatternFill('solid', fgColor='1E3A8A')
    hdr_font   = Font(name=FONT, size=11, bold=True, color='FFFFFF')
    hdr_fill   = PatternFill('solid', fgColor='374151')
    key_font   = Font(name=FONT, size=11, bold=True)
    key_fill   = PatternFill('solid', fgColor='F3F4F6')
    val_font   = Font(name=FONT, size=11)
    total_font = Font(name=FONT, size=11, bold=True, color='1D4ED8')
    total_fill = PatternFill('solid', fgColor='EFF6FF')
    MONEY = '#,##0.00'
    thin = Side(style='thin', color='D1D5DB')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    C = Alignment(horizontal='center', vertical='center')
    R = Alignment(horizontal='right', vertical='center')
    L = Alignment(horizontal='left', vertical='center', wrap_text=True)

    def put(ws, r, c, value=None, font=None, fill=None, align=None, fmt=None, bd=True):
        cell = ws.cell(row=r, column=c, value=value)
        if font: cell.font = font
        if fill: cell.fill = fill
        if align: cell.alignment = align
        if fmt: cell.number_format = fmt
        if bd: cell.border = border
        return cell

    def disp_w(s):
        # CJK chars take ~2 cells of width
        return sum(2 if ord(ch) > 0x2E7F else 1 for ch in str(s))

    def write_table(ws, columns, rows, money_cols=(), int_cols=(), total_last=False):
        ws.sheet_view.showGridLines = False
        for ci, h in enumerate(columns, 1):
            put(ws, 1, ci, h, font=hdr_font, fill=hdr_fill, align=C)
        n = len(rows)
        for ri, rowvals in enumerate(rows, 2):
            is_tot = total_last and ri == n + 1
            for ci, v in enumerate(rowvals, 1):
                num = ci in money_cols or ci in int_cols
                cell = put(ws, ri, ci, v,
                           font=(total_font if is_tot else val_font),
                           fill=(total_fill if is_tot else None),
                           align=(R if num else L))
                if ci in money_cols and isinstance(v, (int, float)):
                    cell.number_format = MONEY
                elif ci in int_cols and isinstance(v, (int, float)):
                    cell.number_format = '#,##0'
        for ci in range(1, len(columns) + 1):
            vals = [columns[ci - 1]] + [r[ci - 1] for r in rows]
            w = max((disp_w(x) for x in vals), default=10)
            ws.column_dimensions[get_column_letter(ci)].width = min(46, max(10, w + 3))
        ws.freeze_panes = 'A2'

    wb = Workbook()

    # ===================== Sheet 1: 对账单汇总 =====================
    ws = wb.active
    ws.title = '对账单汇总'
    ws.sheet_view.showGridLines = False
    for i, w in enumerate([26, 18, 18, 18], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.merge_cells('A1:D1')
    put(ws, 1, 1, '合伙人对账单', font=title_font, fill=title_fill, align=C, bd=False)
    for c in range(2, 5):
        put(ws, 1, c, None, fill=title_fill, bd=False)
    ws.row_dimensions[1].height = 30

    rate_txt = f'1 {currency} = {rate} CNY' if rate else '未配置'
    confirm_txt = '—'
    if data.get('confirmed_name') or data.get('confirmed_at'):
        nm = data.get('confirmed_name') or ''
        at = data.get('confirmed_at')
        confirm_txt = (f'{nm}（{at}）' if at else nm) or '—'
    mode_label = '实际毛利' if (data.get('calc_mode') == 'actual') else '约定毛利'
    info_pairs = [
        ('合伙人', data.get('partner_name') or '', '期间', f"{data['period_year']}-{data['period_month']:02d}"),
        ('货币', currency, '汇率', rate_txt),
        ('计算口径', mode_label, '类型', '手工录入' if data.get('is_manual') else '系统生成'),
        ('状态', STATUS_LABELS.get(data.get('status'), data.get('status') or ''),
         '生成时间', str(data.get('created_at') or '')),
        ('确认签字', confirm_txt, '', ''),
    ]
    r = 2
    for k1, v1, k2, v2 in info_pairs:
        put(ws, r, 1, k1, font=key_font, fill=key_fill, align=L)
        put(ws, r, 2, v1, font=val_font, align=L)
        put(ws, r, 3, k2, font=key_font, fill=key_fill, align=L)
        put(ws, r, 4, v2, font=val_font, align=L)
        r += 1

    r += 1  # spacer
    put(ws, r, 1, '项目', font=hdr_font, fill=hdr_fill, align=C)
    put(ws, r, 2, f'金额（{currency}）', font=hdr_font, fill=hdr_fill, align=C)
    put(ws, r, 3, '金额（CNY）', font=hdr_font, fill=hdr_fill, align=C)
    put(ws, r, 4, None, fill=hdr_fill)
    ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=4)
    r += 1

    net = data.get('total_net_pln') or 0

    def pctlabel(amount):
        return f'（{(amount or 0) / net * 100:.0f}%）' if net else ''

    summary_rows = []
    summary_rows.append(('订单数', data.get('total_orders') or 0, None, False, True))
    summary_rows.append(('总销售额', data.get('total_gross_pln'), data.get('total_gross_cny'), False, False))
    if live:
        ss = (live.get('by_status', {}) or {}).get('success', {}) or {}
        ship_succ = ss.get('shipping', 0) or 0
        loss = live.get('shipping_loss') or 0
        summary_rows.append(('减：运费成本（成功订单）', -ship_succ, round(-ship_succ * rate, 2) if rate else None, False, False))
        summary_rows.append(('减：运费损失（未送达）', -loss, round(-loss * rate, 2) if rate else None, False, False))
    summary_rows.append(('净销售额', data.get('total_net_pln'), data.get('total_net_cny'), True, False))
    if (data.get('calc_mode') or 'contract') == 'actual':
        # 实际毛利口径: cost leg = real product cost; profit/receivable are shares
        # of the actual margin (净销售 − 实际成本).
        cost_amt = data.get('cost_amount_pln') or 0
        margin = (net or 0) - cost_amt

        def margin_pct(amount):
            return f'（实际毛利的 {(amount or 0) / margin * 100:.0f}%）' if margin else ''

        summary_rows.append(('实际产品成本（合伙人留）', data.get('cost_amount_pln'), data.get('cost_amount_cny'), False, False))
        summary_rows.append(('实际毛利（净销售−实际成本）', round(margin, 2), round(margin * rate, 2) if rate else None, True, False))
        summary_rows.append(('合伙人利润' + margin_pct(data.get('partner_profit_pln')), data.get('partner_profit_pln'), data.get('partner_profit_cny'), False, False))
        summary_rows.append(('我方应收' + margin_pct(data.get('our_receivable_pln')), data.get('our_receivable_pln'), data.get('our_receivable_cny'), True, False))
    else:
        summary_rows.append(('约定成本' + pctlabel(data.get('cost_amount_pln')), data.get('cost_amount_pln'), data.get('cost_amount_cny'), False, False))
        summary_rows.append(('合伙人利润' + pctlabel(data.get('partner_profit_pln')), data.get('partner_profit_pln'), data.get('partner_profit_cny'), False, False))
        summary_rows.append(('我方应收' + pctlabel(data.get('our_receivable_pln')), data.get('our_receivable_pln'), data.get('our_receivable_cny'), True, False))
        if live and live.get('actual_cost_pln') is not None:
            ac = live.get('actual_cost_pln') or 0
            summary_rows.append(('实际产品成本（参考）', ac, round(ac * rate, 2) if rate else None, False, False))
    summary_rows.append(('已收', total_received, received_cny, False, False))
    summary_rows.append(('未收', outstanding, outstanding_cny, True, False))

    for label, pln, cny, is_total, is_count in summary_rows:
        fl = total_fill if is_total else None
        vf = total_font if is_total else val_font
        put(ws, r, 1, label, font=(total_font if is_total else key_font), fill=fl, align=L)
        put(ws, r, 2, pln, font=vf, fill=fl, align=R, fmt=('#,##0' if is_count else MONEY))
        put(ws, r, 3, (cny if cny is not None else None), font=vf, fill=fl, align=R, fmt=(None if is_count else MONEY))
        put(ws, r, 4, None, fill=fl)
        ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=4)
        r += 1

    if data.get('notes'):
        r += 1
        put(ws, r, 1, '备注', font=key_font, fill=key_fill, align=L)
        put(ws, r, 2, data.get('notes'), font=val_font, align=L)
        put(ws, r, 3, None)
        put(ws, r, 4, None)
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4)
        ws.row_dimensions[r].height = 40

    # ===================== Sheet 2: 按站点 =====================
    if live and live.get('by_site'):
        site_cols = ['站点', '国家', '负责人', '订单数', '成功单', '未送达单',
                     f'总销售（{currency}）', '运费', '净销售', '运费损失']
        site_rows = []
        for s in live['by_site']:
            site_rows.append([
                s.get('site_url', ''), s.get('country', ''), s.get('manager', ''),
                s.get('orders', 0), s.get('success_orders', 0), s.get('undelivered_orders', 0),
                round(s.get('gross', 0), 2), round(s.get('shipping', 0), 2),
                round(s.get('net', 0), 2), round(s.get('shipping_loss', 0), 2),
            ])
        if site_rows:
            site_rows.append([
                '合计', '', '',
                sum(x[3] for x in site_rows), sum(x[4] for x in site_rows), sum(x[5] for x in site_rows),
                round(sum(x[6] for x in site_rows), 2), round(sum(x[7] for x in site_rows), 2),
                round(sum(x[8] for x in site_rows), 2), round(sum(x[9] for x in site_rows), 2),
            ])
        write_table(wb.create_sheet('按站点'), site_cols, site_rows,
                    money_cols=(7, 8, 9, 10), int_cols=(4, 5, 6), total_last=True)

    # ===================== Sheet 3: 按产品 =====================
    if live and live.get('by_product'):
        prod_cols = ['产品', '品牌', '口数', '口味', '销量',
                     f'销售额（{currency}）', '成本', '毛利', '毛利率%', '有成本']
        prod_rows = []
        for p in live['by_product']:
            mp = p.get('margin_pct')
            prod_rows.append([
                p.get('label', ''), p.get('brand') or '', p.get('puff_count') or '', p.get('flavor') or '',
                p.get('qty', 0), round(p.get('revenue', 0), 2), round(p.get('cost', 0), 2),
                round(p.get('margin', 0), 2), (round(mp, 1) if mp is not None else ''),
                ('是' if p.get('has_cost') else '否'),
            ])
        write_table(wb.create_sheet('按产品'), prod_cols, prod_rows,
                    money_cols=(6, 7, 8), int_cols=(3, 5))

    # ===================== Sheet 4: 收款记录 =====================
    if receipts:
        rcpt_cols = ['收款日期', f'金额（{currency}）', '汇率', '金额（CNY）', '收款方式', '参考号', '备注']
        rcpt_rows = []
        for rc in receipts:
            rcpt_rows.append([
                rc['receipt_date'] or '', round(rc['amount_pln'] or 0, 2),
                rc['exchange_rate_cny'] or '',
                (round(rc['amount_cny'], 2) if rc['amount_cny'] is not None else ''),
                rc['payment_method'] or '', rc['reference_no'] or '', rc['notes'] or '',
            ])
        write_table(wb.create_sheet('收款记录'), rcpt_cols, rcpt_rows, money_cols=(2, 4))

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    filename = f"对账单_{data.get('partner_name') or ''}_{data['period_year']}-{data['period_month']:02d}.xlsx"
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )


@app.route('/api/reconciliation/statements/<int:stmt_id>', methods=['PUT'])
@login_required
@reconciliation_api_required
def api_update_statement(stmt_id):
    """Update statement (exchange rate, notes, status). Every change is logged
    to reconciliation_audit_log for traceability."""
    data = request.json
    conn = get_db_connection()
    stmt = conn.execute('SELECT * FROM reconciliation_statements WHERE id = ?', (stmt_id,)).fetchone()
    if not stmt:
        conn.close()
        return jsonify({'error': '对账单不存在'}), 404
    if not _check_partner_access(stmt['partner_id']):
        conn.close()
        return jsonify({'error': '无权修改'}), 403

    is_admin = _is_reconciliation_admin()
    try:
        # ----- Partner-only actions (confirm) -----
        if not is_admin:
            if data.get('action') == 'confirm' and stmt['status'] == 'generated':
                # P2: e-signature — typed name required, must match partner display name
                typed_name = (data.get('confirm_name') or '').strip()
                if not typed_name:
                    return jsonify({'error': '请输入您的姓名以确认对账单'}), 400
                # Verify against the partner's name (relaxed: substring match — partner names
                # often have parentheses / regions; we accept any substring of length >= 2)
                partner = conn.execute('SELECT name FROM partners WHERE id = ?', (stmt['partner_id'],)).fetchone()
                ref_name = (partner['name'] if partner else '').strip()
                if len(typed_name) < 2:
                    return jsonify({'error': '姓名至少 2 个字符'}), 400
                if typed_name not in ref_name and ref_name not in typed_name:
                    return jsonify({'error': f'姓名不匹配（请输入合伙人名称中的姓名部分，参考：{ref_name}）'}), 400
                ip = request.headers.get('X-Forwarded-For', request.remote_addr) or ''
                if ip and ',' in ip:
                    ip = ip.split(',')[0].strip()
                conn.execute('''UPDATE reconciliation_statements
                    SET status='confirmed', confirmed_at=CURRENT_TIMESTAMP, confirmed_by=?,
                        confirmed_name=?, confirmed_ip=?, updated_at=CURRENT_TIMESTAMP
                    WHERE id=?''', (current_user.id, typed_name, ip, stmt_id))
                _audit_log(stmt_id, 'confirm', field='status', old='generated', new='confirmed',
                           note=f'电子确认：{typed_name} ({ip})', conn=conn)
                conn.commit()
                return jsonify({'success': True})
            return jsonify({'error': '无权执行此操作'}), 403

        # ----- Admin actions: rate / notes / lock / unlock / settle / confirm -----
        if stmt['status'] in ('locked', 'settled') and data.get('action') != 'unlock':
            return jsonify({'error': '对账单已锁定'}), 400

        exchange_rate = data.get('exchange_rate_cny')
        notes = data.get('notes')
        action = data.get('action')

        updates = []
        params = []
        audit_entries = []  # list of (action, field, old, new, note) tuples

        if exchange_rate is not None:
            rate = float(exchange_rate)
            old_rate = stmt['exchange_rate_cny']
            if old_rate != rate:  # only log if actually changed
                updates.append('exchange_rate_cny=?')
                params.append(rate)
                updates.append('our_receivable_cny=?')
                params.append(round((stmt['our_receivable_pln'] or 0) * rate, 2))
                audit_entries.append(('edit_rate', 'exchange_rate_cny', old_rate, rate, None))
        if notes is not None and (notes or '') != (stmt['notes'] or ''):
            updates.append('notes=?')
            params.append(notes)
            audit_entries.append(('edit_notes', 'notes', stmt['notes'], notes, None))
        if action == 'lock':
            updates.append("status='locked'")
            updates.append('locked_at=CURRENT_TIMESTAMP')
            audit_entries.append(('lock', 'status', stmt['status'], 'locked', None))
        elif action == 'unlock':
            updates.append("status='generated'")
            updates.append('locked_at=NULL')
            audit_entries.append(('unlock', 'status', stmt['status'], 'generated', None))
        elif action == 'settle':
            updates.append("status='settled'")
            audit_entries.append(('settle', 'status', stmt['status'], 'settled', None))
        elif action == 'confirm':
            updates.append("status='confirmed'")
            updates.append('confirmed_at=CURRENT_TIMESTAMP')
            updates.append('confirmed_by=?')
            params.append(current_user.id)
            audit_entries.append(('confirm', 'status', stmt['status'], 'confirmed', '管理员代确认'))

        if updates:
            updates.append('updated_at=CURRENT_TIMESTAMP')
            params.append(stmt_id)
            conn.execute(f'UPDATE reconciliation_statements SET {", ".join(updates)} WHERE id=?', params)
            for (a, f, o, n, note) in audit_entries:
                _audit_log(stmt_id, a, field=f, old=o, new=n, note=note, conn=conn)
            conn.commit()
        return jsonify({'success': True})
    finally:
        conn.close()


# ============================================================================
# P2: DISPUTE WORKFLOW
# ============================================================================
# Status transitions powered by these endpoints:
#   generated → disputed  (partner clicks "提出异议", types reason)
#   disputed  → generated (admin clicks "已处理", types resolution note,
#                          optionally regenerates from current order data)
# Both transitions write a normal audit_log entry with the message in `note`,
# so the timeline shows the full conversation.

@app.route('/api/reconciliation/statements/<int:stmt_id>/dispute', methods=['POST'])
@login_required
@reconciliation_api_required
def api_dispute_statement(stmt_id):
    """Partner raises a dispute on a statement.

    Caller can be partner OR admin. Body must include `note` (the dispute
    reason, ≥ 5 chars). Status flips to 'disputed' if currently 'generated'
    or 'confirmed'. Locked/settled statements cannot be disputed (would need
    admin to unlock first).
    """
    data = request.json or {}
    note = (data.get('note') or '').strip()
    if len(note) < 5:
        return jsonify({'error': '请说明异议内容（至少 5 个字符）'}), 400

    conn = get_db_connection()
    stmt = conn.execute('SELECT * FROM reconciliation_statements WHERE id = ?', (stmt_id,)).fetchone()
    if not stmt:
        conn.close()
        return jsonify({'error': '对账单不存在'}), 404
    if not _check_partner_access(stmt['partner_id']):
        conn.close()
        return jsonify({'error': '无权操作'}), 403
    if stmt['status'] in ('locked', 'settled'):
        conn.close()
        return jsonify({'error': '已锁定/已结清的对账单无法提出异议，请联系管理员先解锁'}), 400
    try:
        old_status = stmt['status']
        conn.execute('''UPDATE reconciliation_statements
            SET status='disputed', updated_at=CURRENT_TIMESTAMP WHERE id=?''', (stmt_id,))
        _audit_log(stmt_id, 'dispute', field='status', old=old_status, new='disputed',
                   note=note, conn=conn)
        conn.commit()
        return jsonify({'success': True, 'new_status': 'disputed'})
    finally:
        conn.close()


@app.route('/api/reconciliation/statements/<int:stmt_id>/resolve-dispute', methods=['POST'])
@login_required
@reconciliation_api_required
def api_resolve_dispute(stmt_id):
    """Admin resolves a dispute. Optional `regenerate=true` re-runs the
    aggregation from current order data (re-snapshots).
    """
    if not _is_reconciliation_admin():
        return jsonify({'error': '无权处理异议'}), 403
    data = request.json or {}
    note = (data.get('note') or '').strip()
    if len(note) < 3:
        return jsonify({'error': '请说明处理说明（至少 3 个字符）'}), 400
    do_regen = bool(data.get('regenerate'))

    conn = get_db_connection()
    stmt = conn.execute('SELECT * FROM reconciliation_statements WHERE id = ?', (stmt_id,)).fetchone()
    if not stmt:
        conn.close()
        return jsonify({'error': '对账单不存在'}), 404
    if stmt['status'] != 'disputed':
        conn.close()
        return jsonify({'error': '对账单当前不是异议中状态'}), 400
    try:
        partner_id = stmt['partner_id']
        year = stmt['period_year']
        month = stmt['period_month']
        partner = conn.execute('SELECT * FROM partners WHERE id = ?', (partner_id,)).fetchone()

        if do_regen:
            # Recalculate from current order data and update the statement totals
            detail = _calc_partner_recon_detail(partner_id, year, month)
            if detail is None:
                return jsonify({'error': '重生成失败：无法计算'}), 500
            net = detail['total_net_pln']
            actual_cost_snapshot = detail['actual_cost_pln']
            # Re-run in the SAME basis the statement was originally generated on.
            mode = stmt['calc_mode'] or 'contract'
            cost, p_profit, our_recv = _compute_statement_split(
                net, actual_cost_snapshot, partner['cost_ratio'],
                partner['partner_profit_ratio'], partner['our_profit_ratio'], mode,
                shipping_loss=detail.get('shipping_loss') or 0)
            rate, _ = _lookup_partner_rate(partner['currency'], year, month)
            our_recv_cny = round(our_recv * rate, 2) if rate else None
            conn.execute('''UPDATE reconciliation_statements
                SET status='generated', total_orders=?, total_gross_pln=?, total_net_pln=?,
                    cost_amount_pln=?, partner_profit_pln=?, our_receivable_pln=?,
                    exchange_rate_cny=?, our_receivable_cny=?, actual_cost_pln_snapshot=?,
                    updated_at=CURRENT_TIMESTAMP
                WHERE id=?''',
                (detail['total_orders'], detail['total_gross_pln'], net,
                 cost, p_profit, our_recv, rate, our_recv_cny, actual_cost_snapshot,
                 stmt_id))
            snapshot_count = _snapshot_statement_orders(stmt_id, partner_id, year, month, conn)
            _audit_log(stmt_id, 'resolve_dispute', field='status', old='disputed', new='generated',
                       note=f'已处理并重新生成：{note} (新净销售={net}, 快照 {snapshot_count} 单)', conn=conn)
        else:
            conn.execute('''UPDATE reconciliation_statements
                SET status='generated', updated_at=CURRENT_TIMESTAMP WHERE id=?''', (stmt_id,))
            _audit_log(stmt_id, 'resolve_dispute', field='status', old='disputed', new='generated',
                       note=f'已处理：{note}', conn=conn)
        conn.commit()
        return jsonify({'success': True, 'new_status': 'generated'})
    finally:
        conn.close()


@app.route('/api/reconciliation/statements/<int:stmt_id>/audit-log')
@login_required
@reconciliation_api_required
def api_get_statement_audit_log(stmt_id):
    """Return the timeline of changes for a statement."""
    conn = get_db_connection()
    stmt = conn.execute('SELECT partner_id FROM reconciliation_statements WHERE id = ?', (stmt_id,)).fetchone()
    if not stmt:
        conn.close()
        return jsonify({'error': '对账单不存在'}), 404
    if not _check_partner_access(stmt['partner_id']):
        conn.close()
        return jsonify({'error': '无权查看'}), 403
    rows = conn.execute('''
        SELECT id, action, actor_id, actor_username, actor_role,
               field, old_value, new_value, note, ip, created_at
        FROM reconciliation_audit_log
        WHERE statement_id = ?
        ORDER BY created_at DESC, id DESC
    ''', (stmt_id,)).fetchall()
    conn.close()
    return jsonify([dict(r) for r in rows])


@app.route('/api/reconciliation/statements/<int:stmt_id>/snapshot-orders')
@login_required
@reconciliation_api_required
def api_get_statement_snapshot_orders(stmt_id):
    """Return the frozen order list (with optional comparison to current state)."""
    conn = get_db_connection()
    stmt = conn.execute('SELECT partner_id FROM reconciliation_statements WHERE id = ?', (stmt_id,)).fetchone()
    if not stmt:
        conn.close()
        return jsonify({'error': '对账单不存在'}), 404
    if not _check_partner_access(stmt['partner_id']):
        conn.close()
        return jsonify({'error': '无权查看'}), 403
    snap = conn.execute('''
        SELECT * FROM reconciliation_statement_orders
        WHERE statement_id = ? ORDER BY date_created
    ''', (stmt_id,)).fetchall()
    out = []
    drift_count = 0
    for s in snap:
        d = dict(s)
        cur = conn.execute('''SELECT status, total, shipping_total,
                                     COALESCE(shipping_loss_amount,0) AS sla,
                                     COALESCE(is_undelivered, 0) AS undel
                              FROM orders WHERE id = ?''', (s['order_id'],)).fetchone()
        if cur:
            d['status_now'] = cur['status']
            d['total_now'] = cur['total']
            d['shipping_now'] = cur['shipping_total']
            d['shipping_loss_now'] = cur['sla']
            d['is_undelivered_now'] = bool(cur['undel'])
            d['drifted'] = (
                (cur['status'] != s['status_at_gen']) or
                (abs(float(cur['total'] or 0) - float(s['total_at_gen'] or 0)) > 0.01) or
                (abs(float(cur['shipping_total'] or 0) - float(s['shipping_at_gen'] or 0)) > 0.01) or
                (bool(cur['undel']) != bool(s['is_undelivered_at_gen']))
            )
        else:
            d['drifted'] = True  # order deleted from DB
            d['status_now'] = '_DELETED_'
        if d.get('drifted'):
            drift_count += 1
        out.append(d)
    conn.close()
    return jsonify({'orders': out, 'count': len(out), 'drift_count': drift_count})


@app.route('/api/reconciliation/statements/<int:stmt_id>', methods=['DELETE'])
@login_required
@reconciliation_api_required
def api_delete_statement(stmt_id):
    """Delete a statement (admin only, not if locked/settled)"""
    if not _is_reconciliation_admin():
        return jsonify({'error': '无权删除'}), 403
    conn = get_db_connection()
    stmt = conn.execute('SELECT status FROM reconciliation_statements WHERE id = ?', (stmt_id,)).fetchone()
    if not stmt:
        conn.close()
        return jsonify({'error': '对账单不存在'}), 404
    if stmt['status'] in ('locked', 'settled'):
        conn.close()
        return jsonify({'error': '已锁定的对账单不能删除'}), 400
    try:
        # Delete the frozen order snapshot first, then the statement itself.
        # Same transaction -> atomic. (The schema declares ON DELETE CASCADE, but
        # get_db_connection() doesn't enable PRAGMA foreign_keys, so we delete
        # children explicitly to avoid orphan rows piling up in *_statement_orders.)
        # Note: reconciliation_audit_log is intentionally retained as history.
        conn.execute('DELETE FROM reconciliation_statement_orders WHERE statement_id = ?', (stmt_id,))
        conn.execute('DELETE FROM reconciliation_statements WHERE id = ?', (stmt_id,))
        conn.commit()
        return jsonify({'success': True})
    finally:
        conn.close()


# ---------- Partner Receipts (money received from partner) ----------

@app.route('/api/reconciliation/receipts')
@login_required
@reconciliation_api_required
def api_list_receipts():
    """List receipts, filtered by access + optional partner_id"""
    partner_id = request.args.get('partner_id', type=int)
    allowed_ids = current_user.get_accessible_partner_ids()

    conn = get_db_connection()
    query = '''SELECT r.*, p.name as partner_name, p.currency as currency, s.period_year, s.period_month
               FROM partner_receipts r
               JOIN partners p ON r.partner_id = p.id
               LEFT JOIN reconciliation_statements s ON r.statement_id = s.id
               WHERE 1=1'''
    params = []
    if partner_id:
        if allowed_ids is not None and partner_id not in allowed_ids:
            conn.close()
            return jsonify({'error': '无权查看'}), 403
        query += ' AND r.partner_id = ?'
        params.append(partner_id)
    elif allowed_ids is not None:
        if not allowed_ids:
            conn.close()
            return jsonify([])
        placeholders = ','.join(['?'] * len(allowed_ids))
        query += f' AND r.partner_id IN ({placeholders})'
        params.extend(allowed_ids)
    query += ' ORDER BY r.receipt_date DESC, r.id DESC'
    rows = conn.execute(query, params).fetchall()
    conn.close()
    return jsonify([dict(r) for r in rows])


@app.route('/api/reconciliation/receipts', methods=['POST'])
@login_required
@reconciliation_api_required
def api_create_receipt():
    """Record a new receipt from partner"""
    if not _is_reconciliation_admin():
        return jsonify({'error': '无权录入收款'}), 403
    data = request.json
    partner_id = int(data.get('partner_id'))
    if not _check_partner_access(partner_id):
        return jsonify({'error': '无权'}), 403

    amount_pln = float(data.get('amount_pln', 0))
    rate = data.get('exchange_rate_cny')
    rate_val = float(rate) if rate else None

    # Auto-lookup rate from system if not provided
    if rate_val is None and data.get('receipt_date'):
        conn2 = get_db_connection()
        p = conn2.execute('SELECT currency FROM partners WHERE id = ?', (partner_id,)).fetchone()
        conn2.close()
        if p:
            try:
                y, m = data['receipt_date'].split('-')[:2]
                rate_val, _ = _lookup_partner_rate(p['currency'], int(y), int(m))
            except Exception:
                pass

    amount_cny = round(amount_pln * rate_val, 2) if rate_val else None

    conn = get_db_connection()
    try:
        cursor = conn.execute('''INSERT INTO partner_receipts
            (partner_id, statement_id, receipt_date, amount_pln, exchange_rate_cny, amount_cny,
             payment_method, reference_no, receipt_url, notes, created_by)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)''',
            (partner_id,
             data.get('statement_id') or None,
             data.get('receipt_date'),
             amount_pln, rate_val, amount_cny,
             data.get('payment_method', ''),
             data.get('reference_no', ''),
             data.get('receipt_url', ''),
             data.get('notes', ''),
             current_user.id))
        receipt_id = cursor.lastrowid
        # P2 audit: if linked to a statement, record on its timeline
        stmt_id = data.get('statement_id')
        if stmt_id:
            curr = data.get('currency') or 'PLN'
            _audit_log(int(stmt_id), 'attach_receipt', field='receipt',
                       new=f'#{receipt_id}',
                       note=f'新增收款 {amount_pln} (rate {rate_val}, ref {data.get("reference_no", "-")})', conn=conn)
        conn.commit()
        return jsonify({'success': True, 'id': receipt_id})
    finally:
        conn.close()


@app.route('/api/reconciliation/receipts/<int:receipt_id>', methods=['PUT'])
@login_required
@reconciliation_api_required
def api_update_receipt(receipt_id):
    """Update a receipt"""
    if not _is_reconciliation_admin():
        return jsonify({'error': '无权修改'}), 403
    data = request.json
    conn = get_db_connection()
    receipt = conn.execute('SELECT * FROM partner_receipts WHERE id = ?', (receipt_id,)).fetchone()
    if not receipt:
        conn.close()
        return jsonify({'error': '收款记录不存在'}), 404
    if not _check_partner_access(receipt['partner_id']):
        conn.close()
        return jsonify({'error': '无权'}), 403

    amount_pln = float(data.get('amount_pln', receipt['amount_pln']))
    rate = data.get('exchange_rate_cny')
    rate_val = float(rate) if rate else None
    amount_cny = round(amount_pln * rate_val, 2) if rate_val else None

    try:
        old_stmt_id = receipt['statement_id']
        new_stmt_id = data.get('statement_id') or None
        old_amount = receipt['amount_pln']
        conn.execute('''UPDATE partner_receipts SET
            statement_id=?, receipt_date=?, amount_pln=?, exchange_rate_cny=?, amount_cny=?,
            payment_method=?, reference_no=?, receipt_url=?, notes=?
            WHERE id=?''',
            (new_stmt_id,
             data.get('receipt_date', receipt['receipt_date']),
             amount_pln, rate_val, amount_cny,
             data.get('payment_method', ''),
             data.get('reference_no', ''),
             data.get('receipt_url', ''),
             data.get('notes', ''),
             receipt_id))
        # P2 audit: receipt edits show up on linked statements' timeline
        if old_stmt_id:
            _audit_log(int(old_stmt_id), 'edit_receipt', field='receipt',
                       old=f'#{receipt_id} {old_amount}', new=f'#{receipt_id} {amount_pln}',
                       note='修改收款记录', conn=conn)
        if new_stmt_id and new_stmt_id != old_stmt_id:
            _audit_log(int(new_stmt_id), 'attach_receipt', field='receipt',
                       new=f'#{receipt_id}', note=f'关联收款 {amount_pln}', conn=conn)
        if old_stmt_id and old_stmt_id != new_stmt_id:
            _audit_log(int(old_stmt_id), 'detach_receipt', field='receipt',
                       old=f'#{receipt_id}', note='移除关联', conn=conn)
        conn.commit()
        return jsonify({'success': True})
    finally:
        conn.close()


@app.route('/api/reconciliation/receipts/<int:receipt_id>', methods=['DELETE'])
@login_required
@reconciliation_api_required
def api_delete_receipt(receipt_id):
    """Delete a receipt (admin only)"""
    if not _is_reconciliation_admin():
        return jsonify({'error': '无权删除'}), 403
    conn = get_db_connection()
    try:
        # P2 audit: log deletion on linked statement before deleting
        receipt = conn.execute('SELECT statement_id, amount_pln FROM partner_receipts WHERE id = ?', (receipt_id,)).fetchone()
        if receipt and receipt['statement_id']:
            _audit_log(int(receipt['statement_id']), 'delete_receipt', field='receipt',
                       old=f'#{receipt_id} {receipt["amount_pln"]}',
                       note='删除收款记录', conn=conn)
        conn.execute('DELETE FROM partner_receipts WHERE id = ?', (receipt_id,))
        conn.commit()
        return jsonify({'success': True})
    finally:
        conn.close()


@app.route('/api/reconciliation/rate')
@login_required
@reconciliation_api_required
def api_get_reconciliation_rate():
    """Lookup system exchange rate for a currency+year-month. Used by frontend to prefill rates."""
    currency = request.args.get('currency', 'PLN')
    year = request.args.get('year', type=int)
    month = request.args.get('month', type=int)
    if not year or not month:
        return jsonify({'error': '缺少参数'}), 400
    rate, actual_ym = _lookup_partner_rate(currency, year, month)
    return jsonify({
        'currency': currency,
        'year': year,
        'month': month,
        'rate': rate,
        'source_year_month': actual_ym if rate else None
    })


@app.route('/api/reconciliation/overview')
@login_required
@reconciliation_api_required
def api_reconciliation_overview():
    """Overview: total receivable, total received, outstanding, per partner"""
    allowed_ids = current_user.get_accessible_partner_ids()
    conn = get_db_connection()

    # Get partners
    if allowed_ids is None:
        partners = conn.execute('SELECT * FROM partners ORDER BY id').fetchall()
    elif len(allowed_ids) == 0:
        conn.close()
        return jsonify({'partners': [], 'totals': {}})
    else:
        placeholders = ','.join(['?'] * len(allowed_ids))
        partners = conn.execute(f'SELECT * FROM partners WHERE id IN ({placeholders}) ORDER BY id', allowed_ids).fetchall()

    result = []
    total_receivable = 0
    total_received = 0
    total_receivable_cny = 0
    total_received_cny = 0

    for p in partners:
        # Sum PLN amounts
        recv_row = conn.execute('''SELECT COALESCE(SUM(our_receivable_pln), 0) as total
            FROM reconciliation_statements WHERE partner_id = ?''', (p['id'],)).fetchone()
        paid_row = conn.execute('''SELECT COALESCE(SUM(amount_pln), 0) as total
            FROM partner_receipts WHERE partner_id = ?''', (p['id'],)).fetchone()
        receivable = round(recv_row['total'] or 0, 2)
        received = round(paid_row['total'] or 0, 2)

        # Sum CNY amounts using each statement's historical rate
        stmt_rows = conn.execute('''SELECT period_year, period_month, our_receivable_pln, exchange_rate_cny, our_receivable_cny
            FROM reconciliation_statements WHERE partner_id = ?''', (p['id'],)).fetchall()
        p_receivable_cny = 0
        for sr in stmt_rows:
            r_cny = sr['our_receivable_cny']
            if r_cny is None and sr['our_receivable_pln']:
                rate = sr['exchange_rate_cny']
                if not rate:
                    rate, _ = _lookup_partner_rate(p['currency'], sr['period_year'], sr['period_month'])
                if rate:
                    r_cny = (sr['our_receivable_pln'] or 0) * rate
            p_receivable_cny += (r_cny or 0)

        # Sum received CNY from receipts
        receipt_rows = conn.execute('''SELECT amount_pln, exchange_rate_cny, amount_cny, receipt_date
            FROM partner_receipts WHERE partner_id = ?''', (p['id'],)).fetchall()
        p_received_cny = 0
        for rr in receipt_rows:
            a_cny = rr['amount_cny']
            if a_cny is None and rr['amount_pln']:
                rate = rr['exchange_rate_cny']
                if not rate and rr['receipt_date']:
                    try:
                        y, m = rr['receipt_date'].split('-')[:2]
                        rate, _ = _lookup_partner_rate(p['currency'], int(y), int(m))
                    except:
                        rate = None
                if rate:
                    a_cny = (rr['amount_pln'] or 0) * rate
            p_received_cny += (a_cny or 0)

        p_receivable_cny = round(p_receivable_cny, 2)
        p_received_cny = round(p_received_cny, 2)

        total_receivable += receivable
        total_received += received
        total_receivable_cny += p_receivable_cny
        total_received_cny += p_received_cny

        result.append({
            'id': p['id'],
            'name': p['name'],
            'currency': p['currency'] or 'PLN',
            'total_receivable_pln': receivable,
            'total_received_pln': received,
            'outstanding_pln': round(receivable - received, 2),
            'total_receivable_cny': p_receivable_cny,
            'total_received_cny': p_received_cny,
            'outstanding_cny': round(p_receivable_cny - p_received_cny, 2)
        })
    conn.close()
    return jsonify({
        'partners': result,
        'totals': {
            'total_receivable_pln': round(total_receivable, 2),
            'total_received_pln': round(total_received, 2),
            'outstanding_pln': round(total_receivable - total_received, 2),
            'total_receivable_cny': round(total_receivable_cny, 2),
            'total_received_cny': round(total_received_cny, 2),
            'outstanding_cny': round(total_receivable_cny - total_received_cny, 2)
        }
    })


# ============== PRODUCT ANALYSIS ==============

@app.route('/products')
@login_required
def products():
    """Product analysis page"""
    from datetime import date, timedelta
    conn = get_db_connection()
    
    # Get user's allowed sources for permission filtering
    allowed_sources = get_user_allowed_sources(current_user.id, current_user.is_admin(), current_user.is_viewer())
    
    # Get filter parameters
    source_filter = request.args.get('source', '')
    brand_filter = request.args.get('brand', '')
    puff_filter = request.args.get('puffs', '')
    quick_date = request.args.get('quick_date', 'this_month')
    manager_filter = request.args.get('manager', '')
    country_filter = request.args.get('country', '')
    
    # Get all managers
    all_managers = get_all_managers()

    # Get all countries
    all_countries = conn.execute('SELECT DISTINCT country FROM sites WHERE country IS NOT NULL AND country != "" ORDER BY country').fetchall()
    all_countries = [c['country'] for c in all_countries]
    
    # Validate source_filter against allowed sources
    if source_filter and allowed_sources is not None and source_filter not in allowed_sources:
        source_filter = ''
    
    # Process quick date filter
    today = date.today()
    date_from = ''
    date_to = today.isoformat()
    
    if quick_date == 'last_week':
        # Last 7 days
        date_from = (today - timedelta(days=7)).isoformat()
    elif quick_date == 'last_15_days':
        # Last 15 days
        date_from = (today - timedelta(days=15)).isoformat()
    elif quick_date == 'this_month':
        date_from = today.replace(day=1).isoformat()
    elif quick_date == 'last_month':
        first_of_this_month = today.replace(day=1)
        last_month_end = first_of_this_month - timedelta(days=1)
        last_month_start = last_month_end.replace(day=1)
        date_from = last_month_start.isoformat()
        date_to = last_month_end.isoformat()
    elif quick_date == 'last_quarter':
        date_from = (today - timedelta(days=90)).isoformat()
    elif quick_date == 'half_year':
        date_from = (today - timedelta(days=180)).isoformat()
    elif quick_date == 'one_year':
        date_from = (today - timedelta(days=365)).isoformat()
    elif quick_date == 'all':
        date_from = ''
        date_to = ''
    
    # Build filter conditions with permission check
    conditions = [_revenue_status_cond()]
    params = []
    
    # Add permission filter
    if allowed_sources is not None:
        if allowed_sources:
            placeholders = ', '.join(['?' for _ in allowed_sources])
            conditions.append(f'source IN ({placeholders})')
            params.extend(allowed_sources)
        else:
            conditions.append('1=0')
            
    # Add manager filter
    if manager_filter:
        # Re-use manager_urls from above or fetch if not done
        if 'manager_urls' not in locals():
            manager_sites = conn.execute('SELECT url FROM sites WHERE manager = ?', (manager_filter,)).fetchall()
            manager_urls = [s['url'] for s in manager_sites]
        
        if manager_urls:
            placeholders = ', '.join(['?' for _ in manager_urls])
            conditions.append(f'source IN ({placeholders})')
            params.extend(manager_urls)
        else:
            conditions.append('1=0')

    # Add country filter
    if country_filter:
        country_sites = conn.execute('SELECT url FROM sites WHERE country = ?', (country_filter,)).fetchall()
        country_urls = [s['url'] for s in country_sites]
        if country_urls:
            placeholders = ', '.join(['?' for _ in country_urls])
            conditions.append(f'source IN ({placeholders})')
            params.extend(country_urls)
        else:
            conditions.append('1=0')
    
    if date_from and date_to:
        conditions.append("date_created >= ? AND date_created <= ?")
        params.extend([date_from, date_to + 'T23:59:59'])
    elif date_from:
        conditions.append("date_created >= ?")
        params.append(date_from)
    
    if source_filter:
        conditions.append("source = ?")
        params.append(source_filter)
    
    where_clause = 'WHERE ' + ' AND '.join(conditions)
    
    # Get all orders with line items (including currency and shipping)
    orders = conn.execute(f'''
        SELECT id, line_items, source, currency, total, shipping_total, date_created
        FROM orders {where_clause}
    ''', params).fetchall()
    
    # Get brands list for parsing and filtering
    brands = conn.execute('SELECT id, name, aliases FROM brands ORDER BY name').fetchall()
    brands_list = [dict(b) for b in brands]
    
    # Build brands cache for parsing
    brands_cache = []
    for row in brands:
        brand_name = row['name']
        aliases = []
        if row['aliases']:
            try:
                aliases = json.loads(row['aliases'])
            except:
                pass
        brands_cache.append({
            'id': row['id'],
            'name': brand_name,
            'aliases': aliases,
            'patterns': [brand_name.upper()] + [a.upper() for a in aliases]
        })

    # Load series cache
    _series_rows = conn.execute('SELECT id, brand_id, name FROM series').fetchall()
    series_cache = [{'id': r['id'], 'brand_id': r['brand_id'], 'name': r['name']} for r in _series_rows]
    series_names_map = {r['id']: r['name'] for r in _series_rows}

    # Load manual product mappings
    mappings_rows = conn.execute('''
        SELECT pm.raw_name, pm.puff_count, pm.flavor, pm.series_id, b.name as brand_name
        FROM product_mappings pm
        LEFT JOIN brands b ON pm.brand_id = b.id
        WHERE pm.is_manual = 1
    ''').fetchall()
    manual_mappings = {}
    for m in mappings_rows:
        # Index by canonical form so HTML-entity vs decoded variants both match
        key = normalize_raw_name(m['raw_name'])
        manual_mappings[key] = {
            'brand': m['brand_name'],
            'puffs': m['puff_count'],
            'flavor': m['flavor'],
            'series': series_names_map.get(m['series_id'], '')
        }

    # Aggregate product data
    product_stats = {}
    brand_stats = {}
    puff_stats = {}
    unknown_products_map = {}
    
    for order in orders:
        items = parse_json_field(order['line_items'])
        if not isinstance(items, list):
            continue
        
        # Calculate order items sum for shipping pro-rating and discount calculation
        order_items_sum = sum(float(i.get('total', 0)) for i in items)
        order_shipping = float(order['shipping_total'] or 0)
        order_total = float(order['total'] or 0)
        
        # Calculate order-level discount ratio
        # order_total = items_sum - discount + shipping
        # So: discount = items_sum + shipping - order_total
        # discount_ratio = (order_total - shipping) / items_sum (actual revenue / item totals)
        expected_items_value = order_total - order_shipping
        if order_items_sum > 0:
            discount_ratio = expected_items_value / order_items_sum
        else:
            discount_ratio = 1.0
        
        for item in items:
            product_name = item.get('name', '')
            quantity = item.get('quantity', 0)
            item_total = float(item.get('total', 0))
            
            # Apply order-level discount to get actual item revenue
            total = item_total * discount_ratio
            
            # Pro-rate shipping based on discounted total
            item_shipping = 0
            if order_items_sum > 0:
                item_shipping = order_shipping * (item_total / order_items_sum)
            
            gross_total = total + item_shipping
            
            # Extract flavor from WooCommerce variation meta_data
            meta_flavor = extract_flavor_from_meta(item)
            
            # Check for manual mapping first (using full name with flavor).
            # Canonicalize the lookup key so saved HTML-entity / decoded variants
            # both hit the same mapping row.
            full_name, _, meta_puffs = get_full_product_name(item)
            full_name_key = normalize_raw_name(full_name)
            product_name_key = normalize_raw_name(product_name)
            series = ''
            if full_name_key in manual_mappings:
                mapping = manual_mappings[full_name_key]
                brand = mapping.get('brand') or 'Unknown'
                puffs = mapping.get('puffs') or meta_puffs
                flavor = mapping.get('flavor') or meta_flavor or ''
                series = mapping.get('series') or ''
            elif product_name_key in manual_mappings:
                mapping = manual_mappings[product_name_key]
                brand = mapping.get('brand') or 'Unknown'
                puffs = mapping.get('puffs') or meta_puffs
                flavor = mapping.get('flavor') or meta_flavor or ''
                series = mapping.get('series') or ''
            else:
                parsed = parse_product_name(product_name, brands_cache, series_cache)
                brand = parsed.get('brand') or 'Unknown'
                puffs = meta_puffs or parsed.get('puffs')
                flavor = meta_flavor or parsed.get('flavor') or ''
                series = parsed.get('series') or ''
            
            # Apply brand filter
            if brand_filter and brand != brand_filter:
                continue
            
            # Apply puff filter
            if puff_filter and puffs and str(puffs) != puff_filter:
                continue
            
            # Product level stats (brand + puffs + flavor)
            # Normalize flavor for key AND display to ensure consistent aggregation
            # This handles variations like "Love 66", "love-66", "LOVE_66" -> "LOVE 66"
            flavor_normalized = normalize_flavor(flavor)
            if not flavor_normalized:
                flavor_normalized = 'NO FLAVOR'
                flavor_display = ''
            else:
                flavor_display = flavor_normalized  # Use normalized for display too
            product_key = f"{brand}|{series}|{puffs or 'N/A'}|{flavor_normalized}"
            if product_key not in product_stats:
                product_stats[product_key] = {
                    'name': product_name,
                    'brand': brand,
                    'series': series,
                    'puffs': puffs,
                    'flavor': flavor_display,
                    'quantity': 0,
                    'revenue_by_currency': {},
                    'gross_revenue_by_currency': {},
                    'order_count': 0
                }
            product_stats[product_key]['quantity'] += quantity
            # Track revenue by currency
            currency = order['currency'] or 'N/A'
            if currency not in product_stats[product_key]['revenue_by_currency']:
                product_stats[product_key]['revenue_by_currency'][currency] = 0
                product_stats[product_key]['gross_revenue_by_currency'][currency] = 0
            product_stats[product_key]['revenue_by_currency'][currency] += total
            product_stats[product_key]['gross_revenue_by_currency'][currency] += gross_total
            product_stats[product_key]['order_count'] += 1
            
            # Brand level stats
            if brand not in brand_stats:
                brand_stats[brand] = {'quantity': 0, 'revenue_by_currency': {}, 'gross_revenue_by_currency': {}, 'order_count': 0}
            brand_stats[brand]['quantity'] += quantity
            if currency not in brand_stats[brand]['revenue_by_currency']:
                brand_stats[brand]['revenue_by_currency'][currency] = 0
                brand_stats[brand]['gross_revenue_by_currency'][currency] = 0
            brand_stats[brand]['revenue_by_currency'][currency] += total
            brand_stats[brand]['gross_revenue_by_currency'][currency] += gross_total
            brand_stats[brand]['order_count'] += 1
            
            # Puff level stats
            puff_key = str(puffs) if puffs else 'Unknown'
            if puff_key not in puff_stats:
                puff_stats[puff_key] = {'quantity': 0, 'revenue_by_currency': {}, 'gross_revenue_by_currency': {}}
            puff_stats[puff_key]['quantity'] += quantity
            if currency not in puff_stats[puff_key]['revenue_by_currency']:
                puff_stats[puff_key]['revenue_by_currency'][currency] = 0
                puff_stats[puff_key]['gross_revenue_by_currency'][currency] = 0
            puff_stats[puff_key]['revenue_by_currency'][currency] += total
            puff_stats[puff_key]['gross_revenue_by_currency'][currency] += gross_total
            
            # Collect unknown products for manual mapping
            if puffs is None or brand == 'Unknown':
                if product_name not in unknown_products_map:
                    unknown_products_map[product_name] = {
                        'name': product_name,
                        'brand': brand,
                        'puffs': puffs or 'Unknown',
                        'count': 0
                    }
                unknown_products_map[product_name]['count'] += 1
    
    # Sort products by quantity (top sellers)
    top_n = int(request.args.get('top_n', 50))
    top_products = sorted(product_stats.values(), key=lambda x: x['quantity'], reverse=True)[:top_n]
    
    # Sort brands by quantity
    brand_ranking = sorted(
        [{'name': k, **v} for k, v in brand_stats.items()],
        key=lambda x: x['quantity'],
        reverse=True
    )
    
    # Sort puffs by quantity
    puff_ranking = sorted(
        [{'puffs': k, **v} for k, v in puff_stats.items()],
        key=lambda x: x['quantity'],
        reverse=True
    )
    
    # Get available sources (filtered by permissions and manager)
    conn = get_db_connection()
    
    # Base source query
    source_query = 'SELECT DISTINCT source FROM orders'
    source_params = []
    source_conditions = []
    
    if allowed_sources is not None:
        if allowed_sources:
            placeholders = ', '.join(['?' for _ in allowed_sources])
            source_conditions.append(f'source IN ({placeholders})')
            source_params.extend(allowed_sources)
        else:
            source_conditions.append('1=0')
            
    if manager_filter:
        # Get sites managed by this manager
        manager_sites = conn.execute('SELECT url FROM sites WHERE manager = ?', (manager_filter,)).fetchall()
        manager_urls = [s['url'] for s in manager_sites]
        if manager_urls:
            placeholders = ', '.join(['?' for _ in manager_urls])
            source_conditions.append(f'source IN ({placeholders})')
            source_params.extend(manager_urls)
        else:
            source_conditions.append('1=0')
            
    if source_conditions:
        source_query += ' WHERE ' + ' AND '.join(source_conditions)
        
    source_query += ' ORDER BY source'
    sources = conn.execute(source_query, source_params).fetchall()
    
    # Get available puff counts
    puff_options = sorted([p for p in puff_stats.keys() if p != 'Unknown'], key=lambda x: int(x) if x.isdigit() else 0)
    
    # Get site managers mapping
    sites = conn.execute('SELECT url, manager FROM sites').fetchall()
    site_managers = {s['url']: s['manager'] or '' for s in sites}
    
    # Calculate totals with revenue by currency
    total_revenue_by_currency = {}
    total_gross_revenue_by_currency = {}
    for p in product_stats.values():
        for currency, amount in p.get('revenue_by_currency', {}).items():
            if currency not in total_revenue_by_currency:
                total_revenue_by_currency[currency] = 0
            total_revenue_by_currency[currency] += amount

        for currency, amount in p.get('gross_revenue_by_currency', {}).items():
            if currency not in total_gross_revenue_by_currency:
                total_gross_revenue_by_currency[currency] = 0
            total_gross_revenue_by_currency[currency] += amount

    # Shipping_loss for the same date/source/manager/country window. The
    # main product query uses _revenue_status_cond() which already excludes
    # undelivered orders, so loss isn't reflected in product revenue. Pull
    # it separately and deduct from the net card so the value matches the
    # /monthly + dashboard "实际净额" figure.
    loss_conditions = list(conditions)
    # Replace _revenue_status_cond() with is_undelivered=1 for this query
    loss_conditions[0] = 'is_undelivered = 1'
    loss_where = 'WHERE ' + ' AND '.join(loss_conditions)
    shipping_loss_by_currency = {}
    for r in conn.execute(
        f"SELECT currency, SUM(COALESCE(shipping_loss_amount, 0)) AS loss FROM orders {loss_where} GROUP BY currency",
        params,
    ).fetchall():
        cur = r['currency'] or 'N/A'
        shipping_loss_by_currency[cur] = float(r['loss'] or 0)

    # Subtract the loss from each currency's product revenue so the card
    # represents real net (gross product revenue − shipping_loss).
    for cur, loss in shipping_loss_by_currency.items():
        if cur in total_revenue_by_currency and loss > 0:
            total_revenue_by_currency[cur] -= loss

    # Calculate CNY total for products
    total_revenue_cny = 0
    total_gross_revenue_cny = 0
    from datetime import datetime
    current_month = datetime.now().strftime('%Y-%m')

    for currency, amount in total_revenue_by_currency.items():
        rate, _ = get_cny_rate(currency, current_month)
        if rate:
            total_revenue_cny += amount * rate

    for currency, amount in total_gross_revenue_by_currency.items():
        rate, _ = get_cny_rate(currency, current_month)
        if rate:
            total_gross_revenue_cny += amount * rate

    totals = {
        'total_quantity': sum(p['quantity'] for p in product_stats.values()),
        'total_revenue_by_currency': total_revenue_by_currency,
        'total_gross_revenue_by_currency': total_gross_revenue_by_currency,
        'total_revenue_cny': round(total_revenue_cny, 2),
        'total_gross_revenue_cny': round(total_gross_revenue_cny, 2),
        'shipping_loss_by_currency': {c: round(v, 2) for c, v in shipping_loss_by_currency.items() if v > 0},
        'brand_count': len(brand_stats),
        'product_count': len(product_stats)
    }
    
    # Calculate weekly trend data by flavor
    # TOP 10 flavors are determined by the page's date filter
    # But the chart always shows the last 8 weeks of trend data for those flavors
    
    # Step 1: Get TOP 10 flavors directly from top_products (same data as the product table)
    # This ensures 100% consistency between the trend buttons and the product table
    flavor_totals = {}  # {flavor: quantity}
    for p in top_products:  # Use same data source as the product table
        flavor = p.get('flavor') or ''
        if not flavor:
            continue
        if flavor not in flavor_totals:
            flavor_totals[flavor] = 0
        flavor_totals[flavor] += p['quantity']
    
    # Sort by quantity and take TOP 10
    flavor_ranking = sorted(flavor_totals.items(), key=lambda x: x[1], reverse=True)[:10]
    top_flavors = [f[0] for f in flavor_ranking]
    top_flavor_qtys = {f[0]: f[1] for f in flavor_ranking}
    # For matching in 8-week data (case-insensitive)
    top_flavors_normalized = {f.upper().strip(): f for f in top_flavors}
    
    # Step 2: Query 8 weeks of data for the weekly trend chart
    eight_weeks_ago = (today - timedelta(weeks=8)).isoformat()
    
    # Build conditions for 8-week query (with same permission filtering, but NOT date filter)
    trend_conditions = [_active_status_cond(), "date_created >= ?"]
    trend_params = [eight_weeks_ago]
    
    # Add permission filter
    if allowed_sources is not None:
        if allowed_sources:
            placeholders = ', '.join(['?' for _ in allowed_sources])
            trend_conditions.append(f'source IN ({placeholders})')
            trend_params.extend(allowed_sources)
        else:
            trend_conditions.append('1=0')
    
    # Add manager filter if set
    if manager_filter:
        manager_sites = conn.execute('SELECT url FROM sites WHERE manager = ?', (manager_filter,)).fetchall()
        manager_urls_for_trend = [s['url'] for s in manager_sites]
        if manager_urls_for_trend:
            placeholders = ', '.join(['?' for _ in manager_urls_for_trend])
            trend_conditions.append(f'source IN ({placeholders})')
            trend_params.extend(manager_urls_for_trend)
        else:
            trend_conditions.append('1=0')
    
    # Add source filter if set
    if source_filter:
        trend_conditions.append("source = ?")
        trend_params.append(source_filter)
    
    trend_where_clause = 'WHERE ' + ' AND '.join(trend_conditions)
    
    trend_orders = conn.execute(f'''
        SELECT id, line_items, source, date_created
        FROM orders {trend_where_clause}
    ''', trend_params).fetchall()
    
    weekly_flavor_data = {}  # {week_key: {flavor: quantity}}
    
    for order in trend_orders:
        items = parse_json_field(order['line_items'])
        if not isinstance(items, list):
            continue
        
        # Get week number from order date
        order_date_str = order['date_created']
        if order_date_str:
            try:
                order_date = datetime.strptime(order_date_str[:10], '%Y-%m-%d')
                # Use ISO week number with year
                year, week_num, _ = order_date.isocalendar()
                week_key = f"{year}-W{week_num:02d}"
                week_start = order_date - timedelta(days=order_date.weekday())
                week_label = week_start.strftime('%m/%d')
                
                if week_key not in weekly_flavor_data:
                    weekly_flavor_data[week_key] = {'label': week_label, 'flavors': {}}
                
                # Sum quantities by flavor for this week (only for top_flavors from page filter)
                for item in items:
                    quantity = item.get('quantity', 0)
                    # Get flavor from item (check manual mappings first)
                    product_name = item.get('name', '')
                    full_name, flavor_only, _ = get_full_product_name(item)
                    
                    # Check manual mappings (canonicalize for entity-safe lookup)
                    full_name_key = normalize_raw_name(full_name)
                    if full_name_key in manual_mappings and manual_mappings[full_name_key].get('flavor'):
                        flavor = manual_mappings[full_name_key]['flavor']
                    elif flavor_only:
                        flavor = flavor_only
                    else:
                        flavor = '未知口味'
                    
                    # Only aggregate for top flavors (determined by page filter)
                    # Use normalized matching to handle case differences
                    normalized_flavor = flavor.upper().strip()
                    if normalized_flavor in top_flavors_normalized:
                        # Use the display name from top_flavors for consistency
                        display_flavor = top_flavors_normalized[normalized_flavor]
                        if display_flavor not in weekly_flavor_data[week_key]['flavors']:
                            weekly_flavor_data[week_key]['flavors'][display_flavor] = 0
                        weekly_flavor_data[week_key]['flavors'][display_flavor] += quantity
            except:
                pass
    
    # Sort weeks and take last 8 weeks
    sorted_weeks = sorted(weekly_flavor_data.keys())[-8:]
    
    # Build structured data for chart
    weekly_trend_data = {
        'weeks': [weekly_flavor_data[w]['label'] for w in sorted_weeks],
        'flavors': top_flavors,
        'datasets': []
    }
    
    # Create dataset for each flavor (preserving page filter ranking order)
    # Include the page-filter-period total for correct frontend sorting
    for flavor in top_flavors:
        flavor_data = []
        for week_key in sorted_weeks:
            qty = weekly_flavor_data.get(week_key, {}).get('flavors', {}).get(flavor, 0)
            flavor_data.append(qty)
        weekly_trend_data['datasets'].append({
            'flavor': flavor,
            'data': flavor_data,
            'pageTotal': top_flavor_qtys.get(flavor, 0)  # Total from page filter period
        })
    
    # Calculate weekly trend data by PRODUCT (TOP 10 products)
    # Get top 10 products for the trend
    top_10_products = top_products[:10]
    
    # Build product keys for matching
    top_product_keys = {}  # {normalized_key: {brand, puffs, flavor, label}}
    for p in top_10_products:
        brand = p.get('brand') or 'Unknown'
        puffs = p.get('puffs') or 'N/A'
        flavor = normalize_flavor(p.get('flavor') or '')
        key = f"{brand}|{puffs}|{flavor}"
        label = f"{brand} {puffs} {p.get('flavor', '')[:15]}"  # Short label for display
        top_product_keys[key] = {'brand': brand, 'puffs': puffs, 'flavor': flavor, 'label': label, 'pageTotal': p['quantity']}
    
    # Aggregate weekly data for these products from trend_orders
    weekly_product_data = {}  # {week_key: {product_key: quantity}}
    for order in trend_orders:
        items = parse_json_field(order['line_items'])
        if not isinstance(items, list):
            continue
        
        order_date_str = order['date_created']
        if order_date_str:
            try:
                order_date = datetime.strptime(order_date_str[:10], '%Y-%m-%d')
                year, week_num, _ = order_date.isocalendar()
                week_key = f"{year}-W{week_num:02d}"
                week_start = order_date - timedelta(days=order_date.weekday())
                week_label = week_start.strftime('%m/%d')
                
                if week_key not in weekly_product_data:
                    weekly_product_data[week_key] = {'label': week_label, 'products': {}}
                
                for item in items:
                    quantity = item.get('quantity', 0)
                    product_name = item.get('name', '')
                    
                    # Get brand/puffs/flavor using the same logic as main aggregation
                    meta_flavor = extract_flavor_from_meta(item)
                    full_name, _, meta_puffs = get_full_product_name(item)
                    full_name_key = normalize_raw_name(full_name)
                    product_name_key = normalize_raw_name(product_name)

                    if full_name_key in manual_mappings:
                        mapping = manual_mappings[full_name_key]
                        brand = mapping.get('brand') or 'Unknown'
                        puffs = mapping.get('puffs') or meta_puffs
                        flavor = mapping.get('flavor') or meta_flavor or ''
                    elif product_name_key in manual_mappings:
                        mapping = manual_mappings[product_name_key]
                        brand = mapping.get('brand') or 'Unknown'
                        puffs = mapping.get('puffs') or meta_puffs
                        flavor = mapping.get('flavor') or meta_flavor or ''
                    else:
                        parsed = parse_product_name(product_name, brands_cache)
                        brand = parsed.get('brand') or 'Unknown'
                        puffs = meta_puffs or parsed.get('puffs')
                        flavor = meta_flavor or parsed.get('flavor') or ''
                    
                    # Normalize and create key
                    flavor_norm = normalize_flavor(flavor)
                    product_key = f"{brand}|{puffs or 'N/A'}|{flavor_norm}"
                    
                    # Only aggregate for top 10 products
                    if product_key in top_product_keys:
                        if product_key not in weekly_product_data[week_key]['products']:
                            weekly_product_data[week_key]['products'][product_key] = 0
                        weekly_product_data[week_key]['products'][product_key] += quantity
            except:
                pass
    
    # Build product trend chart data
    product_trend_data = {
        'weeks': [weekly_product_data.get(w, {}).get('label', '') for w in sorted_weeks],
        'products': [top_product_keys[k]['label'] for k in top_product_keys],
        'datasets': []
    }
    
    for key, info in top_product_keys.items():
        prod_data = []
        for week_key in sorted_weeks:
            qty = weekly_product_data.get(week_key, {}).get('products', {}).get(key, 0)
            prod_data.append(qty)
        product_trend_data['datasets'].append({
            'label': info['label'],
            'data': prod_data,
            'pageTotal': info['pageTotal']
        })
    
    conn.close()
    
    return render_template('products.html',
                          top_products=top_products,
                          brand_ranking=brand_ranking,
                          puff_ranking=puff_ranking,
                          totals=totals,
                          brands=brands_list,
                          sources=sources,
                          puff_options=puff_options,
                          site_managers=site_managers,
                          weekly_trend=weekly_trend_data,
                          product_trend=product_trend_data,
                          current_filters={
                              'source': source_filter,
                              'brand': brand_filter,
                              'puffs': puff_filter,
                              'quick_date': quick_date,
                              'manager': manager_filter,
                              'country': country_filter,
                              'top_n': top_n
                          },
                          all_managers=all_managers,
                          all_countries=all_countries)


@app.route('/api/brands')
@login_required
def get_brands():
    """Get all brands"""
    conn = get_db_connection()
    brands = conn.execute('SELECT * FROM brands ORDER BY name').fetchall()
    conn.close()
    
    result = []
    for b in brands:
        aliases = []
        if b['aliases']:
            try:
                aliases = json.loads(b['aliases'])
            except:
                pass
        result.append({
            'id': b['id'],
            'name': b['name'],
            'aliases': aliases,
            'created_at': b['created_at']
        })
    
    return jsonify(result)


@app.route('/api/brands', methods=['POST'])
@login_required
@admin_required
def add_brand():
    """Add a new brand"""
    data = request.json
    name = data.get('name', '').strip()
    aliases = data.get('aliases', [])
    
    if not name:
        return jsonify({'error': 'Brand name is required'}), 400
    
    conn = get_db_connection()
    existing = conn.execute('SELECT id, aliases FROM brands WHERE name = ?', (name,)).fetchone()
    if existing:
        existing_aliases = json.loads(existing['aliases']) if existing['aliases'] else []
        for a in aliases:
            if a and a not in existing_aliases:
                existing_aliases.append(a)
        conn.execute('UPDATE brands SET aliases = ? WHERE id = ?',
                    (json.dumps(existing_aliases) if existing_aliases else None, existing['id']))
        conn.commit()
        conn.close()
        return jsonify({'success': True, 'id': existing['id'], 'merged': True})
    try:
        conn.execute('INSERT INTO brands (name, aliases) VALUES (?, ?)',
                    (name, json.dumps(aliases) if aliases else None))
        conn.commit()
        brand_id = conn.execute('SELECT last_insert_rowid()').fetchone()[0]
        conn.close()
        return jsonify({'success': True, 'id': brand_id})
    except sqlite3.IntegrityError:
        conn.close()
        return jsonify({'error': 'Brand already exists'}), 400


@app.route('/api/brands/<int:brand_id>', methods=['PUT'])
@login_required
@admin_required
def update_brand(brand_id):
    """Update a brand"""
    data = request.json
    name = data.get('name', '').strip()
    aliases = data.get('aliases', [])
    
    if not name:
        return jsonify({'error': 'Brand name is required'}), 400
    
    conn = get_db_connection()
    conn.execute('UPDATE brands SET name = ?, aliases = ? WHERE id = ?',
                (name, json.dumps(aliases) if aliases else None, brand_id))
    conn.commit()
    conn.close()
    
    return jsonify({'success': True})


@app.route('/api/brands/<int:brand_id>', methods=['DELETE'])
@login_required
@admin_required
def delete_brand(brand_id):
    """Delete a brand"""
    conn = get_db_connection()
    conn.execute('DELETE FROM brands WHERE id = ?', (brand_id,))
    conn.execute('UPDATE product_mappings SET brand_id = NULL WHERE brand_id = ?', (brand_id,))
    conn.commit()
    conn.close()
    
    return jsonify({'success': True})


@app.route('/api/series')
@login_required
def get_series():
    """Get series, optionally filtered by brand_id"""
    brand_id = request.args.get('brand_id')
    conn = get_db_connection()
    try:
        if brand_id:
            rows = conn.execute('SELECT id, brand_id, name FROM series WHERE brand_id = ? ORDER BY name', (brand_id,)).fetchall()
        else:
            rows = conn.execute('SELECT id, brand_id, name FROM series ORDER BY brand_id, name').fetchall()
        return jsonify([dict(r) for r in rows])
    finally:
        conn.close()


@app.route('/api/series', methods=['POST'])
@login_required
@admin_required
def add_series():
    data = request.json
    name = (data.get('name') or '').strip()
    brand_id = data.get('brand_id')
    if not name or not brand_id:
        return jsonify({'error': 'name and brand_id are required'}), 400
    conn = get_db_connection()
    try:
        existing = conn.execute('SELECT id FROM series WHERE brand_id = ? AND name = ?', (brand_id, name)).fetchone()
        if existing:
            return jsonify({'success': True, 'id': existing['id'], 'existed': True})
        conn.execute('INSERT INTO series (brand_id, name) VALUES (?, ?)', (brand_id, name))
        conn.commit()
        new_id = conn.execute('SELECT last_insert_rowid()').fetchone()[0]
        return jsonify({'success': True, 'id': new_id})
    finally:
        conn.close()


@app.route('/api/products/stats')
@login_required
def get_product_stats():
    """Get product statistics API"""
    from datetime import date, timedelta
    
    conn = get_db_connection()
    allowed_sources = get_user_allowed_sources(current_user.id, current_user.is_admin(), current_user.is_viewer())
    
    # Get parameters
    source = request.args.get('source', '')
    days = int(request.args.get('days', 30))
    
    # Build conditions
    conditions = [_active_status_cond()]
    params = []
    
    if allowed_sources is not None:
        if allowed_sources:
            placeholders = ', '.join(['?' for _ in allowed_sources])
            conditions.append(f'source IN ({placeholders})')
            params.extend(allowed_sources)
        else:
            conditions.append('1=0')
    
    if source:
        conditions.append('source = ?')
        params.append(source)
    
    date_from = (date.today() - timedelta(days=days)).isoformat()
    conditions.append('date_created >= ?')
    params.append(date_from)
    
    where_clause = 'WHERE ' + ' AND '.join(conditions)
    
    orders = conn.execute(f'''
        SELECT line_items FROM orders {where_clause}
    ''', params).fetchall()
    
    # Get brands cache
    brands_rows = conn.execute('SELECT id, name, aliases FROM brands').fetchall()
    brands_cache = []
    for row in brands_rows:
        brand_name = row['name']
        aliases = []
        if row['aliases']:
            try:
                aliases = json.loads(row['aliases'])
            except:
                pass
        brands_cache.append({
            'id': row['id'],
            'name': brand_name,
            'aliases': aliases,
            'patterns': [brand_name.upper()] + [a.upper() for a in aliases]
        })
    
    conn.close()
    
    # Aggregate
    stats = {}
    for order in orders:
        items = parse_json_field(order['line_items'])
        if not isinstance(items, list):
            continue
        for item in items:
            name = item.get('name', '')
            qty = item.get('quantity', 0)
            total = float(item.get('total', 0))
            
            # Extract flavor from meta_data
            meta_flavor = extract_flavor_from_meta(item)
            
            parsed = parse_product_name(name, brands_cache)
            # Use meta_flavor if available, otherwise use parsed flavor
            flavor = meta_flavor or parsed.get('flavor') or ''
            
            # Create key with flavor for proper aggregation
            base_key = parsed.get('normalized') or name
            key = f"{base_key} - {flavor}" if flavor and flavor.upper() not in base_key.upper() else base_key
            
            if key not in stats:
                stats[key] = {
                    'name': key,
                    'brand': parsed.get('brand'),
                    'puffs': parsed.get('puffs'),
                    'flavor': flavor,
                    'quantity': 0,
                    'revenue': 0
                }
            stats[key]['quantity'] += qty
            stats[key]['revenue'] += total
    
    result = sorted(stats.values(), key=lambda x: x['quantity'], reverse=True)[:100]
    return jsonify(result)


@app.route('/api/products/unknown')
@login_required
def get_unknown_products():
    """Get products that could not be mapped to any brand"""
    from datetime import date, timedelta
    
    conn = get_db_connection()
    
    # Get parameters
    days = int(request.args.get('days', 90))
    limit = int(request.args.get('limit', 50))
    
    # Get orders from last N days
    date_from = (date.today() - timedelta(days=days)).isoformat()
    
    orders = conn.execute(f'''
        SELECT line_items, source FROM orders
        WHERE date_created >= ? AND {_active_status_cond()}
    ''', (date_from,)).fetchall()
    
    # Get brands cache
    brands_rows = conn.execute('SELECT id, name, aliases FROM brands').fetchall()
    brands_cache = []
    for row in brands_rows:
        brand_name = row['name']
        aliases = []
        if row['aliases']:
            try:
                aliases = json.loads(row['aliases'])
            except:
                pass
        brands_cache.append({
            'id': row['id'],
            'name': brand_name,
            'aliases': aliases,
            'patterns': [brand_name.upper()] + [a.upper() for a in aliases]
        })
    
    conn.close()
    
    # Find unknown products
    unknown_products = {}
    
    for order in orders:
        items = parse_json_field(order['line_items'])
        if not isinstance(items, list):
            continue
        
        for item in items:
            name = item.get('name', '')
            if not name:
                continue
            
            quantity = item.get('quantity', 0)
            
            # Get full name with flavor from meta_data
            full_name, meta_flavor, meta_puffs = get_full_product_name(item)
            
            # Parse and check if brand is found
            parsed = parse_product_name(name, brands_cache)
            
            if parsed.get('brand') is None:
                # This product is unknown
                # Use a simplified key (remove flavor for grouping)
                key = name.split(' - ')[0].strip() if ' - ' in name else name
                
                if key not in unknown_products:
                    unknown_products[key] = {
                        'name': key,
                        'sample_full_name': full_name,  # Include flavor in sample
                        'puffs': parsed.get('puffs'),
                        'quantity': 0,
                        'sources': set()
                    }
                unknown_products[key]['quantity'] += quantity
                unknown_products[key]['sources'].add(order['source'])
    
    # Convert sets to lists and sort by quantity
    result = []
    for key, product in unknown_products.items():
        product['sources'] = list(product['sources'])
        result.append(product)
    
    result = sorted(result, key=lambda x: x['quantity'], reverse=True)[:limit]
    
    return jsonify(result)


@app.route('/api/products/mapping', methods=['GET', 'POST'])
@login_required
def product_mapping():
    """Get or save product mapping"""
    conn = get_db_connection()
    
    if request.method == 'GET':
        # Get mapping for a specific product name
        raw_name = request.args.get('name', '')
        if not raw_name:
            conn.close()
            return jsonify({'error': 'Name is required'}), 400

        # Match against either the literal form OR the HTML-entity-decoded form
        # (legacy rows may have either encoding; new rows are always decoded).
        candidates = {raw_name, normalize_raw_name(raw_name)}
        placeholders = ','.join(['?' for _ in candidates])
        mapping = conn.execute(f'''
            SELECT pm.*, b.name as brand_name
            FROM product_mappings pm
            LEFT JOIN brands b ON pm.brand_id = b.id
            WHERE pm.raw_name IN ({placeholders})
            ORDER BY pm.is_manual DESC, pm.id DESC LIMIT 1
        ''', tuple(candidates)).fetchone()
        
        conn.close()
        
        if mapping:
            return jsonify({
                'id': mapping['id'],
                'raw_name': mapping['raw_name'],
                'brand_id': mapping['brand_id'],
                'brand_name': mapping['brand_name'],
                'series_id': mapping['series_id'],
                'puff_count': mapping['puff_count'],
                'flavor': mapping['flavor'],
                'is_manual': mapping['is_manual']
            })
        else:
            return jsonify({'exists': False})
    
    else:  # POST - Save mapping
        if not current_user.can_edit():
            conn.close()
            return jsonify({'success': False, 'error': '只读用户无法修改数据', 'readonly': True}), 403

        data = request.get_json()
        raw_name = data.get('raw_name', '').strip()

        if not raw_name:
            conn.close()
            return jsonify({'success': False, 'error': 'Product name is required'})

        # Canonicalize: always store the HTML-entity-decoded form so the same
        # product saved from different UIs ends up as a single row.
        raw_name = normalize_raw_name(raw_name)
        
        brand_id = data.get('brand_id')
        series_id = data.get('series_id')
        puff_count = data.get('puff_count')
        flavor = data.get('flavor')

        # Convert to proper types
        if brand_id and str(brand_id).strip():
            try:
                brand_id = int(brand_id)
            except ValueError:
                brand_id = None
        else:
            brand_id = None

        if series_id and str(series_id).strip():
            try:
                series_id = int(series_id)
            except ValueError:
                series_id = None
        else:
            series_id = None

        if puff_count and str(puff_count).strip():
            try:
                puff_count = int(puff_count)
            except ValueError:
                puff_count = None
        else:
            puff_count = None

        if flavor:
            flavor = str(flavor).strip()
        else:
            flavor = ''

        try:
            # Find any pre-existing mapping for either the canonical form OR
            # legacy HTML-entity-encoded variants. We collapse both onto the
            # canonical row so future lookups find a single source of truth.
            cur = conn.execute(
                'SELECT id, raw_name FROM product_mappings WHERE raw_name = ? OR raw_name = ?',
                (raw_name, html.escape(raw_name, quote=False))
            ).fetchall()

            if cur:
                # Update the first one to the canonical form, delete the rest.
                primary_id = cur[0]['id']
                conn.execute('''
                    UPDATE product_mappings
                    SET raw_name = ?, brand_id = ?, series_id = ?, puff_count = ?, flavor = ?, is_manual = 1
                    WHERE id = ?
                ''', (raw_name, brand_id, series_id, puff_count, flavor, primary_id))
                for row in cur[1:]:
                    conn.execute('DELETE FROM product_mappings WHERE id = ?', (row['id'],))
            else:
                # Insert new
                conn.execute('''
                    INSERT INTO product_mappings (raw_name, brand_id, series_id, puff_count, flavor, is_manual)
                    VALUES (?, ?, ?, ?, ?, 1)
                ''', (raw_name, brand_id, series_id, puff_count, flavor))

            conn.commit()
            conn.close()
            return jsonify({'success': True})
        except Exception as e:
            conn.close()
            return jsonify({'success': False, 'error': str(e)})


@app.route('/api/products/samples')
@login_required
def get_product_samples():
    """Get sample orders containing a specific product (supports brand+puffs+flavor matching)"""
    product_name = request.args.get('name', '')
    brand_filter = request.args.get('brand', '')
    puffs_filter = request.args.get('puffs', '')
    flavor_filter = request.args.get('flavor', '')
    
    if not product_name:
        return jsonify({'error': 'Product name is required'}), 400
    
    conn = get_db_connection()
    
    # Get all site managers first (before closing connection)
    sites = conn.execute('SELECT url, manager FROM sites').fetchall()
    site_managers = {s['url']: s['manager'] or '' for s in sites}
    
    # Get brands cache for parsing
    brands_rows = conn.execute('SELECT id, name, aliases FROM brands').fetchall()
    brands_cache = []
    for row in brands_rows:
        brand_name = row['name']
        aliases = []
        if row['aliases']:
            try:
                aliases = json.loads(row['aliases'])
            except:
                pass
        brands_cache.append({
            'id': row['id'],
            'name': brand_name,
            'aliases': aliases,
            'patterns': [brand_name.upper()] + [a.upper() for a in aliases]
        })
    
    # Load manual product mappings
    mappings_rows = conn.execute('''
        SELECT pm.raw_name, pm.puff_count, pm.flavor, b.name as brand_name
        FROM product_mappings pm
        LEFT JOIN brands b ON pm.brand_id = b.id
        WHERE pm.is_manual = 1
    ''').fetchall()
    manual_mappings = {}
    for m in mappings_rows:
        manual_mappings[normalize_raw_name(m['raw_name'])] = {
            'brand': m['brand_name'],
            'puffs': m['puff_count'],
            'flavor': m['flavor']
        }

    # Search for orders - increase limit to find more sources
    orders = conn.execute(f'''
        SELECT id, number, source, date_created, line_items
        FROM orders
        WHERE {_active_status_cond()}
        ORDER BY date_created DESC
        LIMIT 2000
    ''').fetchall()
    
    conn.close()
    
    # Find orders with matching product (by brand+puffs+flavor or exact name)
    results = []
    sources_map = {}
    use_combination_match = brand_filter or puffs_filter or flavor_filter
    
    for order in orders:
        items = parse_json_field(order['line_items'])
        if not isinstance(items, list):
            continue
        
        for item in items:
            item_name = item.get('name', '')
            if not item_name:
                continue
            
            matched = False
            
            if use_combination_match:
                # Match by brand+puffs+flavor combination
                # First, parse the product to get its brand/puffs/flavor
                meta_flavor = extract_flavor_from_meta(item)
                full_name, _, meta_puffs = get_full_product_name(item)
                full_name_key = normalize_raw_name(full_name)
                item_name_key = normalize_raw_name(item_name)

                # Check manual mapping first
                if full_name_key in manual_mappings:
                    mapping = manual_mappings[full_name_key]
                    item_brand = mapping.get('brand') or 'Unknown'
                    item_puffs = str(mapping.get('puffs') or meta_puffs or '')
                    item_flavor = mapping.get('flavor') or meta_flavor or ''
                elif item_name_key in manual_mappings:
                    mapping = manual_mappings[item_name_key]
                    item_brand = mapping.get('brand') or 'Unknown'
                    item_puffs = str(mapping.get('puffs') or meta_puffs or '')
                    item_flavor = mapping.get('flavor') or meta_flavor or ''
                else:
                    # Parse automatically
                    parsed = parse_product_name(item_name, brands_cache)
                    item_brand = parsed.get('brand') or 'Unknown'
                    item_puffs = str(meta_puffs or parsed.get('puffs') or '')
                    item_flavor = meta_flavor or parsed.get('flavor') or ''
                
                # Check if matches the filter criteria
                brand_match = not brand_filter or item_brand.upper() == brand_filter.upper()
                puffs_match = not puffs_filter or item_puffs == puffs_filter
                # Flavor matching: if filter provided, do case-insensitive contains;
                # if filter is empty, only match items that ALSO have no recognized flavor
                # (otherwise an "unknown flavor" bucket would pull in every flavor variant)
                if flavor_filter:
                    flavor_match = bool(item_flavor) and flavor_filter.upper() in item_flavor.upper()
                else:
                    flavor_match = not item_flavor

                matched = brand_match and puffs_match and flavor_match
            else:
                # Fallback to exact name match
                matched = item_name == product_name
            
            if matched:
                # Extract domain from source for display
                source = order['source']
                source_display = source.replace('https://www.', '').replace('https://', '').split('/')[0]
                
                # Get manager from pre-loaded site_managers
                manager_name = site_managers.get(source, '')
                
                # Always collect all sources
                sources_map[source_display] = manager_name
                
                # Only add to results if under limit
                if len(results) < 10:
                    results.append({
                        'order_number': order['number'],
                        'source': source_display,
                        'manager': manager_name,
                        'date': order['date_created'][:10] if order['date_created'] else ''
                    })
                break  # One match per order is enough
    
    sources_list = [{'site': k, 'manager': v} for k, v in sources_map.items()]

    parsed = parse_product_name(product_name, brands_cache)
    parsed_brand = parsed.get('brand') or ''
    parsed_puffs = parsed.get('puffs') or ''
    parsed_flavor = parsed.get('flavor') or ''
    product_name_key = normalize_raw_name(product_name)
    if product_name_key in manual_mappings:
        m = manual_mappings[product_name_key]
        if m.get('brand'): parsed_brand = m['brand']
        if m.get('puffs'): parsed_puffs = m['puffs']
        if m.get('flavor'): parsed_flavor = m['flavor']

    return jsonify({
        'sources': sources_list,
        'orders': results,
        'total_found': len(results),
        'parsed_brand': parsed_brand,
        'parsed_puffs': parsed_puffs,
        'parsed_flavor': parsed_flavor
    })



@app.route('/api/settings', methods=['GET', 'POST'])
@login_required
def settings_api():
    """API endpoint for general settings (admin only for POST)"""
    conn = get_db_connection()
    
    if request.method == 'GET':
        settings = conn.execute("SELECT key, value FROM settings").fetchall()
        conn.close()
        result = {row['key']: row['value'] for row in settings}
        
        # Also include user's display preference
        if current_user.is_authenticated:
            pref = get_db_connection()
            user_pref = pref.execute('''
                SELECT preference_value FROM user_preferences 
                WHERE user_id = ? AND preference_key = 'source_display_mode'
            ''', (current_user.id,)).fetchone()
            pref.close()
            result['source_display_mode'] = user_pref['preference_value'] if user_pref else 'full'
        
        return jsonify(result)
        
    elif request.method == 'POST':
        data = request.json
        try:
            need_update_cron = False
            for key, value in data.items():
                # source_display_mode is now per-user
                if key == 'source_display_mode':
                    # Save to user_preferences instead
                    conn.execute('''
                        INSERT OR REPLACE INTO user_preferences (user_id, preference_key, preference_value)
                        VALUES (?, 'source_display_mode', ?)
                    ''', (current_user.id, str(value)))
                elif key in ['autosync_enabled', 'autosync_interval']:
                    # These remain global settings (admin only)
                    conn.execute("INSERT OR REPLACE INTO settings (key, value) VALUES (?, ?)", (key, str(value)))
                    need_update_cron = True
                elif key in ['big_order_qty_threshold', 'big_order_amount_threshold']:
                    # Big-order alert thresholds (global)
                    conn.execute("INSERT OR REPLACE INTO settings (key, value) VALUES (?, ?)", (key, str(value)))
                elif key == 'auto_track_au':
                    # AU scheduled carrier-tracking toggle ('1'/'0'), read by the
                    # nightly resolve_outcomes.py cron job.
                    conn.execute("INSERT OR REPLACE INTO settings (key, value) VALUES (?, ?)", (key, str(value)))
                elif key == 'blocklist_auto_cancel_enabled':
                    # Blocklist master switch. Admin-only (governs an automated
                    # order-cancelling action).
                    if not current_user.is_admin():
                        continue
                    v = '1' if str(value).strip().lower() in ('1', 'true', 'yes', 'on') else '0'
                    conn.execute("INSERT OR REPLACE INTO settings (key, value) VALUES (?, ?)", (key, v))
                elif key == 'auto_confirm_delivered_enabled':
                    # Auto-confirm master switch (待确认结局 → 已签收). Admin-only
                    # (governs an automated action that completes orders + fires
                    # WC 'completed' emails). Read by the hourly auto_sync.py cron
                    # via auto_confirm.is_enabled().
                    if not current_user.is_admin():
                        continue
                    v = '1' if str(value).strip().lower() in ('1', 'true', 'yes', 'on') else '0'
                    conn.execute("INSERT OR REPLACE INTO settings (key, value) VALUES (?, ?)", (key, v))
                    if v == '1':
                        # Forward-only: stamp the effective-start (same datetime('now')
                        # format as carrier_status_at) so ONLY deliveries detected after
                        # turning on get auto-confirmed — never the existing backlog.
                        conn.execute("INSERT OR REPLACE INTO settings (key, value) VALUES ('auto_confirm_delivered_since', datetime('now'))")
            conn.commit()
            
            # 如果自动同步设置发生变更，同步更新 Crontab 定时任务
            if need_update_cron:
                try:
                    import subprocess
                    # 重新读取最新的数据库设置
                    enabled_row = conn.execute("SELECT value FROM settings WHERE key = 'autosync_enabled'").fetchone()
                    interval_row = conn.execute("SELECT value FROM settings WHERE key = 'autosync_interval'").fetchone()
                    
                    sync_enabled = enabled_row['value'] == 'true' if enabled_row else False
                    sync_interval = int(interval_row['value']) if interval_row else 900
                    
                    # 获取现有 crontab
                    result = subprocess.run(['/usr/bin/crontab', '-l'], capture_output=True, text=True)
                    existing_crontab = result.stdout if result.returncode == 0 else ''
                    
                    # 移除已有的 auto_sync.py 条目
                    lines = [line for line in existing_crontab.split('\n') 
                             if line.strip() and 'auto_sync.py' not in line]
                    
                    if sync_enabled:
                        # 将间隔（秒）转换为 cron 表达式
                        interval_mins = sync_interval // 60
                        
                        if interval_mins < 60:
                            cron_schedule = f"*/{interval_mins} * * * *"
                        elif interval_mins < 1440:
                            hours = interval_mins // 60
                            cron_schedule = f"0 */{hours} * * *"
                        else:
                            cron_schedule = "0 6 * * *"
                        
                        script_path = '/www/wwwroot/woo-analysis'
                        cron_line = f"{cron_schedule} cd {script_path} && {script_path}/venv/bin/python auto_sync.py >> {script_path}/auto_sync.log 2>&1"
                        lines.append(cron_line)
                    
                    # 写入新的 crontab
                    new_crontab = '\n'.join(lines) + '\n' if lines else ''
                    process = subprocess.Popen(['/usr/bin/crontab', '-'], stdin=subprocess.PIPE, text=True)
                    process.communicate(input=new_crontab)
                except Exception as cron_err:
                    # cron 更新失败不影响设置保存的成功返回，但记录日志
                    app.logger.error(f"更新 crontab 失败: {str(cron_err)}")
            
            conn.close()
            return jsonify({'success': True})
        except Exception as e:
            conn.close()
            return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/user/preferences', methods=['GET', 'POST'])
@login_required
def user_preferences_api():
    """API endpoint for user-specific preferences"""
    conn = get_db_connection()
    
    if request.method == 'GET':
        prefs = conn.execute('''
            SELECT preference_key, preference_value FROM user_preferences
            WHERE user_id = ?
        ''', (current_user.id,)).fetchall()
        conn.close()
        return jsonify({row['preference_key']: row['preference_value'] for row in prefs})
        
    elif request.method == 'POST':
        data = request.json
        try:
            for key, value in data.items():
                # Allow specific preference keys only
                if key in ['source_display_mode']:
                    conn.execute('''
                        INSERT OR REPLACE INTO user_preferences (user_id, preference_key, preference_value)
                        VALUES (?, ?, ?)
                    ''', (current_user.id, key, str(value)))
            conn.commit()
            conn.close()
            return jsonify({'success': True})
        except Exception as e:
            conn.close()
            return jsonify({'success': False, 'error': str(e)}), 500


# ============== SHIPPING MANAGEMENT ==============

@app.route('/shipping')
@login_required
@shipping_view_required
def shipping():
    """Shipping management page"""
    conn = get_db_connection()
    
    # Get carriers
    carriers = conn.execute('SELECT * FROM shipping_carriers WHERE is_active = 1').fetchall()
    
    # Get all managers for filter
    managers = get_all_managers()
    
    # Get sites for filter
    allowed_sources = get_user_allowed_sources(current_user.id, current_user.is_admin(), current_user.is_viewer())
    if allowed_sources is None:
        sites = conn.execute('SELECT id, url, manager FROM sites').fetchall()
    elif not allowed_sources:
        sites = []
    else:
        placeholders = ','.join(['?' for _ in allowed_sources])
        sites = conn.execute(f'SELECT id, url, manager FROM sites WHERE url IN ({placeholders})', allowed_sources).fetchall()
    
    # Get all countries
    all_countries = conn.execute('SELECT DISTINCT country FROM sites WHERE country IS NOT NULL AND country != "" ORDER BY country').fetchall()
    all_countries = [c['country'] for c in all_countries]
    
    # Auto-confirm switch state (待确认结局 → 已签收), for the toolbar toggle.
    _ac_row = conn.execute("SELECT value FROM settings WHERE key='auto_confirm_delivered_enabled'").fetchone()
    auto_confirm_enabled = bool(_ac_row) and str(_ac_row['value']).strip().lower() in ('1', 'true', 'yes', 'on')
    _ac_since_row = conn.execute("SELECT value FROM settings WHERE key='auto_confirm_delivered_since'").fetchone()
    auto_confirm_since = _ac_since_row['value'] if _ac_since_row and _ac_since_row['value'] else ''

    auto_confirm_stats = {'candidates': 0, 'backlog': 0, 'returned': 0, 'since': auto_confirm_since}
    outcome_scope_sql = ''
    outcome_scope_params = []
    if allowed_sources is None:
        pass
    elif not allowed_sources:
        outcome_scope_sql = ' AND 1=0'
    else:
        placeholders = ','.join(['?' for _ in allowed_sources])
        outcome_scope_sql = f' AND o.source IN ({placeholders})'
        outcome_scope_params.extend(allowed_sources)
    outcome_base = """
        FROM orders o
        LEFT JOIN sites s ON o.source = s.url
        LEFT JOIN shipping_logs sl ON sl.id = (
            SELECT id FROM shipping_logs WHERE order_id = o.id ORDER BY id DESC LIMIT 1
        )
        WHERE o.status IN ('on-hold', 'shipped', 'partial-shipped')
          AND o.payment_method = 'cod'
          AND COALESCE(o.is_undelivered, 0) = 0
          AND COALESCE(o.is_problem_return, 0) = 0
          AND COALESCE(o.delivery_confirmed, 0) = 0
    """
    pending_age_case = (
        "CASE COALESCE(s.country, '') "
        "WHEN 'PL' THEN 1 "
        "WHEN 'AU' THEN 14 "
        "ELSE 7 END"
    )
    shipped_expr = "datetime(replace(substr(COALESCE(sl.shipped_at, o.date_modified, o.date_created), 1, 19), 'T', ' '))"
    ready_sql = f" AND {shipped_expr} <= datetime('now', '-' || {pending_age_case} || ' days')"
    if auto_confirm_since:
        auto_confirm_stats['candidates'] = conn.execute(
            f"SELECT COUNT(*) {outcome_base} AND o.carrier_status = 'delivered' "
            f"AND o.carrier_status_at IS NOT NULL{ready_sql} "
            f"{outcome_scope_sql}",
            outcome_scope_params
        ).fetchone()[0]
    else:
        auto_confirm_stats['backlog'] = conn.execute(
            f"SELECT COUNT(*) {outcome_base} AND o.carrier_status = 'delivered'{ready_sql}{outcome_scope_sql}",
            outcome_scope_params
        ).fetchone()[0]
    auto_confirm_stats['returned'] = conn.execute(
        f"SELECT COUNT(*) {outcome_base} AND o.carrier_status = 'returned'{ready_sql}{outcome_scope_sql}",
        outcome_scope_params
    ).fetchone()[0]

    conn.close()
    return render_template('shipping.html', carriers=carriers, managers=managers, sites=sites,
                           all_countries=all_countries, is_super_admin=(current_user.username == 'admin'),
                           has_au_access=_has_au_access(), auto_confirm_enabled=auto_confirm_enabled,
                           auto_confirm_stats=auto_confirm_stats)


def _has_au_access():
    """True if the user may see the AU shipping sheet: the super admin, or anyone
    holding an explicit Australian-site permission. NOTE: viewer/admin role-based
    all-access does NOT count here — this gate is intentionally stricter than the
    rest of the system so only AU-relevant staff (and the owner) see this sheet."""
    if getattr(current_user, 'username', None) == 'admin':
        return True
    conn = get_db_connection()
    try:
        row = conn.execute('''SELECT 1 FROM user_site_permissions p
            JOIN sites s ON p.site_id = s.id
            WHERE p.user_id = ? AND s.country = 'AU' LIMIT 1''', (current_user.id,)).fetchone()
    finally:
        conn.close()
    return bool(row)


@app.route('/au-orders')
@login_required
def au_orders_sheet():
    """Clean, light spreadsheet-style view of ALL Australian-market orders, laid
    out like the team's old Tencent sheet (是否发货 / 运单号 / 发货日期 / 订单号 /
    下单时间 / 站点 / 产品 / 总金额 / 支付方式 / 运输方式 / 客户 / 电话 / 地址 /
    客户备注 / 内部备注). Read-only — the AU team prefers a spreadsheet over the
    full system UI. Access is gated to AU-site-permission holders (+ super admin)."""
    from datetime import datetime as _dt
    if not _has_au_access():
        return render_template_string(
            '<div style="font-family:-apple-system,Microsoft YaHei,sans-serif;padding:48px;color:#555;text-align:center;">'
            '<h2 style="color:#333;">无权访问</h2>'
            '<p>「澳洲发货表」仅限拥有澳洲站点权限的账号查看。</p></div>'), 403
    conn = get_db_connection()
    au_sites = [r['url'] for r in conn.execute("SELECT url FROM sites WHERE country='AU'").fetchall()]
    allowed = get_user_allowed_sources(current_user.id, current_user.is_admin(), current_user.is_viewer())
    if allowed is not None:
        au_sites = [u for u in au_sites if u in allowed]
    if not au_sites:
        conn.close()
        return render_template('au_orders.html', rows=[], count=0,
                               generated_at=_dt.now().strftime('%Y-%m-%d %H:%M'))
    ph = ','.join(['?'] * len(au_sites))
    orders = conn.execute(f"""
        SELECT o.id, o.number, o.status, o.total, o.currency, o.date_created, o.date_modified,
               o.source, o.billing, o.shipping, o.line_items, o.shipping_lines, o.meta_data,
               o.customer_note, o.payment_method,
               n.note AS internal_note
        FROM orders o
        LEFT JOIN (
            SELECT order_id, note, date_created FROM order_notes
            WHERE customer_note = 0 AND added_by_user = 1
            GROUP BY order_id HAVING date_created = MAX(date_created)
        ) n ON o.id = n.order_id
        WHERE o.source IN ({ph})
          AND o.status NOT IN ('checkout-draft', 'trash')
        ORDER BY o.date_created DESC
    """, au_sites).fetchall()

    # All parcels per order (split shipment / 分批发货) from our local shipping_logs.
    # This is the immediate source of truth for orders shipped via this system;
    # the order's AST meta is merged in below to also cover tracking added
    # directly in WP-admin (and post-sync state).
    logs_map = {}
    _oids = [o['id'] for o in orders]
    if _oids:
        ph2 = ','.join(['?'] * len(_oids))
        for r in conn.execute(f"""
            SELECT sl.order_id, sl.tracking_number, sl.carrier_slug, sl.shipped_at,
                   sc.name AS carrier_name
            FROM shipping_logs sl
            LEFT JOIN shipping_carriers sc ON sc.slug = sl.carrier_slug
            WHERE sl.order_id IN ({ph2}) ORDER BY sl.id""", _oids).fetchall():
            logs_map.setdefault(r['order_id'], []).append({
                'tracking_number': (r['tracking_number'] or '').strip(),
                'carrier_slug': r['carrier_slug'] or '',
                'shipped_at': r['shipped_at'] or '',
            })

    brands_rows = conn.execute('SELECT id, name, aliases FROM brands').fetchall()
    conn.close()
    # Brand cache for compact product labels (built once, reused per line item).
    brands_cache = []
    for br in brands_rows:
        try:
            aliases = json.loads(br['aliases']) if br['aliases'] else []
        except Exception:
            aliases = []
        if not isinstance(aliases, list):
            aliases = []
        brands_cache.append({'id': br['id'], 'name': br['name'], 'aliases': aliases,
                             'patterns': [(br['name'] or '').upper()] + [str(a).upper() for a in aliases]})

    def pay_label(pm):
        p = (pm or '').lower()
        if 'stripe' in p:
            return '条纹支付'
        if p == 'bacs':
            return '银行转账'
        if 'cod' in p:
            return '货到付款'
        if 'paypal' in p:
            return 'PayPal'
        if p == 'custom_gateway':
            return '在线支付'
        return pm or ''

    def product_cell(it):
        # Display = 品牌 + 口数 + 口味（完整，方便看清是哪个口味）。
        # color_key = 品牌 + 口数 ONLY —— 发货员按"同品牌同口数"分色拣货，不按口味
        # （口味太多、按口味上色太花、记不住）。
        nm = it.get('name', '') or ''
        try:
            p = parse_product_name(nm, brands_cache, series_cache=[])
            brand = p.get('brand') or ''
            puffs = str(p['puffs']) if p.get('puffs') else ''
            flavor = p.get('flavor') or ''
        except Exception:
            brand = puffs = flavor = ''
        # Robust puff fallback: the strict parser misses glued numbers like
        # "INGOT9000"; grab the first 4-6 digit run so the same physical product
        # groups under one color regardless of name formatting.
        if not puffs:
            mm = re.search(r'\d{4,6}', nm)
            if mm:
                puffs = mm.group()
        display = ' - '.join(x for x in (brand, puffs, flavor) if x).strip() or nm
        key = ' '.join(x for x in (brand, puffs) if x).strip() or display
        return {'name': display, 'key': key, 'brand': brand, 'qty': it.get('quantity', 1)}

    def extract_trackings(meta_list):
        # Return ALL tracking records from the order meta. AST stores a LIST
        # (_wc_shipment_tracking_items) — for a split shipment that's multiple
        # records, so we iterate the whole list, not just the first. Falls back
        # to the generic single _tracking_number. Each = (number, provider, date).
        out = []
        for m in (meta_list or []):
            if isinstance(m, dict) and m.get('key') == '_wc_shipment_tracking_items':
                v = m.get('value') or []
                if isinstance(v, list):
                    for rec in v:
                        if isinstance(rec, dict) and str(rec.get('tracking_number', '')).strip():
                            out.append((str(rec['tracking_number']).strip(),
                                        str(rec.get('tracking_provider', '') or ''),
                                        str(rec.get('date_shipped', '') or '')))
                if out:
                    return out
        for m in (meta_list or []):
            if isinstance(m, dict) and m.get('key') == '_tracking_number' and str(m.get('value', '')).strip():
                prov = ''
                for mm in (meta_list or []):
                    if isinstance(mm, dict) and mm.get('key') == '_tracking_provider':
                        prov = str(mm.get('value', '') or '')
                return [(str(m['value']).strip(), prov, '')]
        return []

    def carrier_label(provider, method_title):
        # Show the ACTUAL carrier (from the tracking provider) over the checkout
        # shipping method — e.g. provider 'ems' → EMS even when the order's
        # shipping method is 'Australia Post'.
        p = (provider or '').lower().replace('_', '-')
        m = {'ems': 'EMS', 'australia-post': 'Australia Post', 'auspost': 'Australia Post',
             'inpost-paczkomaty': 'InPost', 'inpost': 'InPost', 'dpd': 'DPD', 'dpd-pl': 'DPD',
             'china-post': 'EMS/中国邮政', 'gls': 'GLS', 'poczta-polska': 'Poczta Polska'}
        if p in m:
            return m[p]
        if provider:
            return provider.upper() if len(provider) <= 4 else provider
        return method_title or ''

    SHIPPED = ('completed', 'shipped', 'partial-shipped')
    rows = []
    for o in orders:
        billing = parse_json_field(o['billing'])
        shipping_info = parse_json_field(o['shipping'])
        line_items = parse_json_field(o['line_items'])
        shipping_lines = parse_json_field(o['shipping_lines'])
        meta_list = parse_json_field(o['meta_data'])
        meta_tracks = extract_trackings(meta_list)
        logs = logs_map.get(o['id'], [])
        # Merge parcels: our shipping_logs first (immediate, ordered #1..N), then
        # any AST-meta records not already present (WP-admin entries / post-sync).
        trackings, seen = [], set()
        for lg in logs:
            tn = lg['tracking_number']
            if tn and tn.lower() not in seen:
                seen.add(tn.lower())
                trackings.append({'number': tn, 'carrier': carrier_label(lg['carrier_slug'], '')})
        for tn, prov, _ds in meta_tracks:
            if tn and tn.lower() not in seen:
                seen.add(tn.lower())
                trackings.append({'number': tn, 'carrier': carrier_label(prov, '')})
        addr = shipping_info if shipping_info and shipping_info.get('address_1') else billing
        tracking = trackings[0]['number'] if trackings else ''
        is_shipped = bool(trackings) or (o['status'] in SHIPPED)
        # Hide orders cancelled before ever shipping — typically the customer
        # never paid and WooCommerce auto-cancelled on timeout. They just clutter
        # the shipping sheet. If goods actually went out (has a tracking number),
        # keep the row so dispatched stock stays visible.
        if o['status'] == 'cancelled' and not is_shipped:
            continue
        # 发货日期: earliest local parcel date, else AST date_shipped, else date_modified.
        ship_date = ''
        log_dates = [lg['shipped_at'][:10] for lg in logs if lg.get('shipped_at')]
        if log_dates:
            ship_date = min(log_dates)
        if not ship_date:
            ast_ts = meta_tracks[0][2] if meta_tracks else ''
            if ast_ts and ast_ts.isdigit():
                try:
                    ship_date = _dt.fromtimestamp(int(ast_ts)).strftime('%Y-%m-%d')
                except Exception:
                    pass
        if not ship_date and is_shipped:
            ship_date = (o['date_modified'] or '')[:10]
        name = f"{addr.get('first_name', '')} {addr.get('last_name', '')}".strip() \
            or f"{billing.get('first_name', '')} {billing.get('last_name', '')}".strip()
        rows.append({
            'shipped': is_shipped,
            'tracking': tracking,
            'trackings': trackings,
            'ship_date': ship_date,
            'number': o['number'],
            'order_time': (o['date_created'] or '').replace('T', ' ')[:16],
            'source': (o['source'] or '').replace('https://www.', '').replace('https://', ''),
            'products': [product_cell(it) for it in (line_items or [])],
            'total': f"{float(o['total'] or 0):.2f}",
            'currency': o['currency'] or 'AUD',
            'payment': pay_label(o['payment_method']),
            'shipping_method': (trackings[0]['carrier'] if trackings
                                else carrier_label('', shipping_lines[0].get('method_title', '') if shipping_lines else '')),
            'customer': name,
            'phone': addr.get('phone') or billing.get('phone', ''),
            'address': _compose_address(addr),
            'customer_note': o['customer_note'] or '',
            'internal_note': o['internal_note'] or '',
        })

    # Color-code products by 品牌+口数 (NOT flavor): same brand+puff → same color,
    # so the picker grabs the right box. Specific brands have a fixed, memorable
    # color (per request); every other brand+puff key gets a stable auto hue spaced
    # by the golden angle (137.5°) for good separation.
    BRAND_CHIP = {
        'alibarbar': ('#f6e7a0', '#7d5e10'),   # 金色
        'umin':      ('#d6ebfb', '#1f5d8c'),   # 淡蓝色
    }
    keys = sorted({p['key'] for r in rows for p in r['products']})
    hue_map = {k: round((i * 137.508) % 360) for i, k in enumerate(keys)}
    for r in rows:
        for p in r['products']:
            ov = BRAND_CHIP.get((p.get('brand') or '').lower())
            if ov:
                p['bg'], p['fg'] = ov
            else:
                hue = hue_map.get(p['key'], 0)
                p['bg'], p['fg'] = f'hsl({hue},70%,91%)', f'hsl({hue},52%,30%)'

    return render_template('au_orders.html', rows=rows, count=len(rows),
                           generated_at=_dt.now().strftime('%Y-%m-%d %H:%M'))


def extract_custom_billing_fields(meta_data):
    """Extract custom billing fields from meta_data"""
    custom_fields = {
        'customer_inpost_id': '',
        'customer_social': '',
        'dpd_street': '',
        'dpd_house': '',
        'dpd_zip': '',
        'dpd_city': ''
    }
    
    if not meta_data:
        return custom_fields
        
    for meta in meta_data:
        if isinstance(meta, dict):
            key = meta.get('key')
            value = meta.get('value')
            
            if key == '_billing_inpost':
                custom_fields['customer_inpost_id'] = value
            elif key == '_billing_social':
                custom_fields['customer_social'] = value
            elif key == '_billing_adres_dpd':
                custom_fields['dpd_street'] = value
            elif key == '_billing_numer_domu':
                custom_fields['dpd_house'] = value
            elif key == '_billing_kod_pocztowy':
                custom_fields['dpd_zip'] = value
            elif key == '_billing_miejscowosc':
                custom_fields['dpd_city'] = value
                
    return custom_fields


def get_big_order_thresholds(conn):
    """Read big-order alert thresholds from settings (with defaults).

    Returns (qty_threshold, amount_threshold) as floats. A threshold of 0
    disables that dimension. Defaults: qty>=10 units, amount>=1500.
    """
    rows = conn.execute(
        "SELECT key, value FROM settings "
        "WHERE key IN ('big_order_qty_threshold', 'big_order_amount_threshold')"
    ).fetchall()
    vals = {r['key']: r['value'] for r in rows}
    try:
        qty = float(vals.get('big_order_qty_threshold') or 10)
    except (TypeError, ValueError):
        qty = 10
    try:
        amount = float(vals.get('big_order_amount_threshold') or 1500)
    except (TypeError, ValueError):
        amount = 1500
    return qty, amount


def evaluate_big_order(product_count, total, qty_threshold, amount_threshold):
    """Return (is_big, reasons[]) for an order given the thresholds.

    Amount is compared against the order total in its own currency — a rough
    trigger for a shipping-floor prompt, not an accounting figure.
    """
    reasons = []
    if qty_threshold and product_count >= qty_threshold:
        reasons.append(f"数量 {product_count} ≥ {int(qty_threshold)}")
    if amount_threshold and total >= amount_threshold:
        t = int(total) if float(total).is_integer() else round(total, 2)
        reasons.append(f"金额 {t} ≥ {int(amount_threshold)}")
    return (len(reasons) > 0, reasons)


@app.route('/api/shipping/pending')
@login_required
@shipping_view_required
def get_pending_orders():
    """Get orders pending shipment (status=processing)"""
    conn = get_db_connection()
    qty_threshold, amount_threshold = get_big_order_thresholds(conn)

    # Get filter parameters
    source_filter = request.args.get('source', '')
    manager_filter = request.args.get('manager', '')
    country_filter = request.args.get('country', '')
    
    query = '''
        SELECT o.id, o.number, o.status, o.total, o.currency, o.date_created,
               o.source, o.billing, o.shipping, o.line_items, o.meta_data, o.shipping_total, o.shipping_lines,
               o.customer_note, o.warehouse_id,
               s.manager,
               w.name as warehouse_name,
               n.note as latest_note, n.date_created as latest_note_date, n.author as latest_note_author
        FROM orders o
        LEFT JOIN sites s ON o.source = s.url
        LEFT JOIN warehouses w ON o.warehouse_id = w.id
        LEFT JOIN (
            SELECT order_id, note, date_created, author
            FROM order_notes
            WHERE customer_note = 0
            GROUP BY order_id
            HAVING date_created = MAX(date_created)
        ) n ON o.id = n.order_id
        WHERE o.status IN ('processing', 'offline')
    '''
    params = []
    
    # Apply source filter
    allowed_sources = get_user_allowed_sources(current_user.id, current_user.is_admin(), current_user.is_viewer())
    if allowed_sources is not None:
        if not allowed_sources:
            # User has restricted access but valid sources list is empty -> no access
            query += ' AND 1=0'
        else:
            placeholders = ','.join(['?' for _ in allowed_sources])
            query += f' AND o.source IN ({placeholders})'
            params.extend(allowed_sources)
    
    if source_filter:
        query += ' AND o.source = ?'
        params.append(source_filter)
    
    if manager_filter:
        query += ' AND s.manager = ?'
        params.append(manager_filter)
        
    if country_filter:
        query += ' AND s.country = ?'
        params.append(country_filter)
    
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    search = request.args.get('search')
    
    if start_date:
        query += ' AND o.date_created >= ?'
        params.append(start_date + ' 00:00:00')
    
    if end_date:
        query += ' AND o.date_created <= ?'
        params.append(end_date + ' 23:59:59')
    
    if search:
        search_term = f'%{search}%'
        query += ' AND (o.number LIKE ? OR o.billing LIKE ? OR o.shipping LIKE ?)'
        params.extend([search_term, search_term, search_term])
    
    query += ' ORDER BY o.date_created DESC'
    
    orders = conn.execute(query, params).fetchall()

    # Parcels already shipped for these orders (split shipment / 分批发货).
    # shipping_logs is local-only and untouched by sync, so a partial order
    # keeps its 'processing' status and stays in this queue; the parcel rows
    # tell us how many packages already went out.
    order_ids = [o['id'] for o in orders]
    parcels_map = {}
    if order_ids:
        ph = ','.join(['?'] * len(order_ids))
        for r in conn.execute(
            f'''SELECT sl.order_id, sl.tracking_number, sl.carrier_slug, sl.shipped_at,
                       sc.name AS carrier_name
                FROM shipping_logs sl
                LEFT JOIN shipping_carriers sc ON sc.slug = sl.carrier_slug
                WHERE sl.order_id IN ({ph}) ORDER BY sl.id''', order_ids).fetchall():
            parcels_map.setdefault(r['order_id'], []).append({
                'tracking_number': r['tracking_number'],
                'carrier_slug': r['carrier_slug'],
                'carrier_name': r['carrier_name'] or r['carrier_slug'],
                'shipped_at': r['shipped_at'],
            })

    # Build the risk index once (small scan over flagged orders only) and
    # reuse it for every row in this listing. Closing conn AFTER the build.
    risk_idx = _build_risk_index(conn)
    conn.close()

    result = []
    for order in orders:
        billing = parse_json_field(order['billing'])
        shipping_info = parse_json_field(order['shipping'])
        line_items = parse_json_field(order['line_items'])
        shipping_lines = parse_json_field(order['shipping_lines'])
        shipping_method = shipping_lines[0].get('method_title', 'Unknown') if shipping_lines and len(shipping_lines) > 0 else ''

        # Get shipping address (prefer shipping, fallback to billing)
        addr = shipping_info if shipping_info and shipping_info.get('address_1') else billing
        meta_data = parse_json_field(order['meta_data'])
        custom_fields = extract_custom_billing_fields(meta_data)

        # Calculate customer address (Standard)
        customer_address = _compose_address(addr)

        # DPD Fallback: If standard address is empty but DPD fields exist
        if not addr.get('address_1') and (custom_fields.get('dpd_street') or custom_fields.get('dpd_city')):
            dpd_parts = [
                f"{custom_fields.get('dpd_street', '')} {custom_fields.get('dpd_house', '')}".strip(),
                custom_fields.get('dpd_zip', ''),
                custom_fields.get('dpd_city', '')
            ]
            customer_address = ', '.join(filter(None, dpd_parts))

        product_count = sum(item.get('quantity', 1) for item in (line_items or []))
        order_total = float(order['total'] or 0)
        is_big_order, big_order_reasons = evaluate_big_order(
            product_count, order_total, qty_threshold, amount_threshold)
        customer_risk = _assess_customer_risk(billing, shipping_info, risk_idx, current_order_id=order['id'])

        result.append({
            'id': order['id'],
            'number': order['number'],
            'total': order_total,
            'currency': order['currency'],
            'date_created': order['date_created'],
            'source': order['source'].replace('https://www.', '').replace('https://', ''),
            'manager': order['manager'] or '',
            'customer_name': f"{addr.get('first_name', '')} {addr.get('last_name', '')}".strip(),
            'customer_email': billing.get('email', ''),
            'customer_phone': addr.get('phone') or billing.get('phone', ''),
            'customer_address': customer_address,
            'state_mismatch': _au_state_mismatch(addr),
            'customer_inpost_id': custom_fields['customer_inpost_id'],
            'customer_social': custom_fields['customer_social'],
            'products': [{'name': item.get('name', ''), 'quantity': item.get('quantity', 1), 'total': float(item.get('total', 0))} for item in (line_items or [])],
            'shipping_total': float(order['shipping_total'] or 0),
            'shipping_method': shipping_method,
            'product_count': product_count,
            'is_big_order': is_big_order,
            'big_order_reasons': big_order_reasons,
            'customer_note': order['customer_note'] or '',
            'warehouse_id': order['warehouse_id'],
            'warehouse_name': order['warehouse_name'] or '',
            'customer_risk': customer_risk,
            'parcels': parcels_map.get(order['id'], []),
            'parcels_shipped': len(parcels_map.get(order['id'], [])),
            'latest_note': order['latest_note'] or '',
            'latest_note_date': order['latest_note_date'] or '',
            'latest_note_author': order['latest_note_author'] or ''
        })

    return jsonify(result)


@app.route('/api/shipping/pending-outcome')
@login_required
@shipping_view_required
def get_pending_outcome_orders():
    """「待确认结局」queue: COD orders that were shipped (on-hold/shipped) some
    days ago but whose fate was never recorded — neither 已签收 nor 拒收 nor
    问题退货. These are the orders starving the repeat-freeloader risk index:
    every refusal hiding here is a risk signal that never got fed back.

    Oldest first, so the shipper works the backlog down. The risk badge is
    attached so a refusal from a known bad customer is obvious at a glance."""
    from datetime import timedelta
    conn = get_db_connection()

    explicit_days = 'days' in request.args
    try:
        days = int(request.args.get('days')) if explicit_days else None
    except (TypeError, ValueError):
        days = None
    if days is not None:
        days = max(0, min(days, 3650))
        cutoff = (datetime.now() - timedelta(days=days)).strftime('%Y-%m-%d %H:%M:%S')

    source_filter = request.args.get('source', '')
    country_filter = request.args.get('country', '')

    query = '''
        SELECT o.id, o.number, o.status, o.total, o.currency, o.date_created,
               o.source, o.billing, o.shipping, o.line_items, o.meta_data,
               o.shipping_total, o.customer_note, o.warehouse_id,
               o.carrier_status, o.carrier_status_at,
               s.manager,
               w.name AS warehouse_name,
               sl.tracking_number, sl.carrier_slug, sl.shipped_at,
               n.note AS latest_note, n.date_created AS latest_note_date, n.author AS latest_note_author
        FROM orders o
        LEFT JOIN sites s ON o.source = s.url
        LEFT JOIN warehouses w ON o.warehouse_id = w.id
        LEFT JOIN shipping_logs sl ON sl.id = (
            SELECT id FROM shipping_logs WHERE order_id = o.id ORDER BY id DESC LIMIT 1
        )
        LEFT JOIN (
            SELECT order_id, note, date_created, author
            FROM order_notes
            WHERE customer_note = 0
            GROUP BY order_id
            HAVING date_created = MAX(date_created)
        ) n ON o.id = n.order_id
        WHERE o.status IN ('on-hold', 'shipped', 'partial-shipped')
          AND o.payment_method = 'cod'
          AND COALESCE(o.is_undelivered, 0) = 0
          AND COALESCE(o.is_problem_return, 0) = 0
          AND COALESCE(o.delivery_confirmed, 0) = 0
    '''
    params = []
    if days is not None:
        query += """
          AND datetime(replace(substr(COALESCE(sl.shipped_at, o.date_modified, o.date_created), 1, 19), 'T', ' ')) <= ?
        """
        params.append(cutoff)
    else:
        query += """
          AND datetime(replace(substr(COALESCE(sl.shipped_at, o.date_modified, o.date_created), 1, 19), 'T', ' ')) <= datetime(
              'now',
              '-' || CASE COALESCE(s.country, '')
                       WHEN 'PL' THEN 1
                       WHEN 'AU' THEN 14
                       ELSE 7
                     END || ' days'
          )
        """

    allowed_sources = get_user_allowed_sources(current_user.id, current_user.is_admin(), current_user.is_viewer())
    if allowed_sources is not None:
        if not allowed_sources:
            query += ' AND 1=0'
        else:
            placeholders = ','.join(['?' for _ in allowed_sources])
            query += f' AND o.source IN ({placeholders})'
            params.extend(allowed_sources)

    if source_filter:
        query += ' AND o.source = ?'
        params.append(source_filter)
    if country_filter:
        query += ' AND s.country = ?'
        params.append(country_filter)

    # Surface the most actionable rows first: detected returns (need a 拒收
    # decision), then attention, then carrier-confirmed deliveries (ready to
    # batch-approve), then everything still in transit / unchecked — each group
    # oldest-first so the backlog drains.
    query += """ ORDER BY CASE COALESCE(o.carrier_status,'')
                            WHEN 'returned' THEN 0
                            WHEN 'attention' THEN 1
                            WHEN 'delivered' THEN 2
                            WHEN 'in_transit' THEN 3
                            ELSE 4 END, o.date_created ASC"""

    orders = conn.execute(query, params).fetchall()
    risk_idx = _build_risk_index(conn)
    conn.close()

    now = datetime.now()
    result = []
    for order in orders:
        billing = parse_json_field(order['billing']) or {}
        shipping_info = parse_json_field(order['shipping']) or {}
        line_items = parse_json_field(order['line_items'])
        addr = shipping_info if shipping_info and shipping_info.get('address_1') else billing
        meta_data = parse_json_field(order['meta_data'])
        custom_fields = extract_custom_billing_fields(meta_data)

        customer_address = _compose_address(addr)
        if not addr.get('address_1') and (custom_fields.get('dpd_street') or custom_fields.get('dpd_city')):
            dpd_parts = [f"{custom_fields.get('dpd_street', '')} {custom_fields.get('dpd_house', '')}".strip(),
                         custom_fields.get('dpd_zip', ''), custom_fields.get('dpd_city', '')]
            customer_address = ', '.join(filter(None, dpd_parts))

        days_pending = None
        try:
            d = datetime.fromisoformat((order['date_created'] or '')[:19])
            days_pending = (now - d).days
        except (ValueError, TypeError):
            pass

        result.append({
            'id': order['id'],
            'number': order['number'],
            'status': order['status'],
            'total': float(order['total'] or 0),
            'currency': order['currency'],
            'date_created': order['date_created'],
            'days_pending': days_pending,
            'carrier_status': order['carrier_status'] or '',
            'carrier_status_at': order['carrier_status_at'] or '',
            'source': order['source'].replace('https://www.', '').replace('https://', ''),
            'manager': order['manager'] or '',
            'customer_name': f"{addr.get('first_name', '')} {addr.get('last_name', '')}".strip(),
            'customer_email': billing.get('email', ''),
            'customer_phone': addr.get('phone') or billing.get('phone', ''),
            'customer_address': customer_address,
            'state_mismatch': _au_state_mismatch(addr),
            'shipping_total': float(order['shipping_total'] or 0),
            'products': [{'name': item.get('name', ''), 'quantity': item.get('quantity', 1)} for item in (line_items or [])],
            'tracking_number': order['tracking_number'] or '',
            'carrier_slug': order['carrier_slug'] or '',
            'warehouse_name': order['warehouse_name'] or '',
            'customer_risk': _assess_customer_risk(billing, shipping_info, risk_idx, current_order_id=order['id']),
            'latest_note': order['latest_note'] or '',
            'latest_note_date': order['latest_note_date'] or '',
            'latest_note_author': order['latest_note_author'] or '',
        })

    return jsonify(result)


@app.route('/api/shipping/pending-outcome/ids-before')
@login_required
@shipper_required
def pending_outcome_ids_before():
    """Admin-only. List 待确认结局 candidate order IDs created strictly BEFORE a
    cutoff date — powers the 「一刀切」 bulk historical close-out: confirm all
    unconfirmed COD orders before <date> as delivered.

    Uses the EXACT same candidate definition as the 待确认结局 queue, so it
    inherently skips orders already marked undelivered / problem-return /
    confirmed, plus non-COD and failed/cancelled orders. The result therefore
    includes carrier_status 'returned'/'unknown'/'in_transit' rows (detected but
    not human-actioned) — which the operator explicitly wants closed out too."""
    if current_user.username != 'admin':
        return jsonify({'error': '只有超级管理员可使用一刀切'}), 403
    import re as _re
    date = (request.args.get('date') or '').strip()
    if not _re.match(r'^\d{4}-\d{2}-\d{2}$', date):
        return jsonify({'error': '日期格式应为 YYYY-MM-DD'}), 400

    conn = get_db_connection()
    conds = ["payment_method = 'cod'",
             "status IN ('on-hold', 'shipped', 'partial-shipped')",
             "COALESCE(is_undelivered, 0) = 0",
             "COALESCE(is_problem_return, 0) = 0",
             "COALESCE(delivery_confirmed, 0) = 0",
             "date_created < ?"]   # 'YYYY-MM-DD' < 'YYYY-MM-DDT..' → excludes the day itself
    params = [date]
    allowed_sources = get_user_allowed_sources(current_user.id, current_user.is_admin(), current_user.is_viewer())
    if allowed_sources is not None:
        if not allowed_sources:
            conn.close()
            return jsonify({'date': date, 'count': 0, 'ids': []})
        ph = ','.join(['?'] * len(allowed_sources))
        conds.append(f'source IN ({ph})')
        params.extend(allowed_sources)
    rows = conn.execute(f"SELECT id FROM orders WHERE {' AND '.join(conds)} ORDER BY date_created", params).fetchall()
    conn.close()
    ids = [r['id'] for r in rows]
    return jsonify({'date': date, 'count': len(ids), 'ids': ids})


@app.route('/api/shipping/shipped')
@login_required
@shipping_view_required
def get_shipped_orders():
    """Get shipped orders (status=on-hold) with tracking info"""
    conn = get_db_connection()
    
    # Get filter parameters
    source_filter = request.args.get('source', '')
    country_filter = request.args.get('country', '')
    
    query = '''
        SELECT o.id, o.number, o.status, o.total, o.currency, o.date_created, o.date_modified,
               o.source, o.billing, o.shipping, o.line_items, o.meta_data, o.shipping_lines, o.shipping_total,
               o.customer_note, o.warehouse_id,
               o.is_undelivered, o.shipping_loss_amount, o.undelivered_at, o.undelivered_note,
               o.is_problem_return, o.carrier_status, o.carrier_status_at,
               s.manager,
               w.name AS warehouse_name,
               sl.tracking_number, sl.carrier_slug, sl.shipped_at,
               u.name AS undelivered_by_name,
               n.note as latest_note, n.date_created as latest_note_date, n.author as latest_note_author
        FROM orders o
        LEFT JOIN sites s ON o.source = s.url
        LEFT JOIN warehouses w ON o.warehouse_id = w.id
        -- Join only the LATEST parcel as the representative tracking row, so a
        -- split-shipment order (multiple shipping_logs rows) appears ONCE here.
        -- The full parcel list is attached separately below.
        LEFT JOIN shipping_logs sl ON sl.id = (
            SELECT id FROM shipping_logs WHERE order_id = o.id ORDER BY id DESC LIMIT 1
        )
        LEFT JOIN users u ON o.undelivered_by = u.id
        LEFT JOIN (
            SELECT order_id, note, date_created, author
            FROM order_notes
            WHERE customer_note = 0
            GROUP BY order_id
            HAVING date_created = MAX(date_created)
        ) n ON o.id = n.order_id
        WHERE o.status IN ('on-hold', 'shipped', 'partial-shipped')
    '''
    params = []
    
    allowed_sources = get_user_allowed_sources(current_user.id, current_user.is_admin(), current_user.is_viewer())
    if allowed_sources is not None:
        placeholders = ','.join(['?' for _ in allowed_sources])
        query += f' AND o.source IN ({placeholders})'
        params.extend(allowed_sources)
    
    if source_filter:
        query += ' AND o.source = ?'
        params.append(source_filter)
        
    if country_filter:
        query += ' AND s.country = ?'
        params.append(country_filter)

    manager_filter = request.args.get('manager', '')
    if manager_filter:
        query += ' AND s.manager = ?'
        params.append(manager_filter)
        
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    search = request.args.get('search')
    
    if start_date:
        query += ' AND o.date_created >= ?'
        params.append(start_date + ' 00:00:00')
    
    if end_date:
        query += ' AND o.date_created <= ?'
        params.append(end_date + ' 23:59:59')
    
    if search:
        search_term = f'%{search}%'
        # Also match tracking numbers stored by external plugins (AST / VillaTheme / custom):
        # they live in meta_data, shipping_lines or line_items as JSON. LIKE on the raw JSON
        # is good enough for the typical case where tracking numbers are 10+ alphanumeric chars.
        query += ''' AND (
            o.number LIKE ?
            OR o.billing LIKE ?
            OR o.shipping LIKE ?
            OR EXISTS (SELECT 1 FROM shipping_logs slx WHERE slx.order_id = o.id AND slx.tracking_number LIKE ?)
            OR o.meta_data LIKE ?
            OR o.shipping_lines LIKE ?
            OR o.line_items LIKE ?
        )'''
        params.extend([search_term] * 7)

    query += ' ORDER BY sl.shipped_at DESC, o.date_modified DESC, o.date_created DESC'
    
    orders = conn.execute(query, params).fetchall()

    # All parcels per order (split shipment / 分批发货) so a completed multi-
    # parcel order can show every tracking number, not just the latest.
    shipped_ids = [o['id'] for o in orders]
    shipped_parcels_map = {}
    if shipped_ids:
        ph = ','.join(['?'] * len(shipped_ids))
        for r in conn.execute(
            f'''SELECT sl.order_id, sl.tracking_number, sl.carrier_slug, sl.shipped_at,
                       sl.is_reship, sl.reship_reason,
                       sc.name AS carrier_name, sc.tracking_url
                FROM shipping_logs sl
                LEFT JOIN shipping_carriers sc ON sc.slug = sl.carrier_slug
                WHERE sl.order_id IN ({ph}) ORDER BY sl.id''', shipped_ids).fetchall():
            tn = r['tracking_number'] or ''
            tmpl = r['tracking_url'] or ''
            rk = r.keys()
            shipped_parcels_map.setdefault(r['order_id'], []).append({
                'tracking_number': tn,
                'carrier_slug': r['carrier_slug'],
                'carrier_name': r['carrier_name'] or r['carrier_slug'],
                'tracking_url': (tmpl.replace('{tracking}', tn).replace('{tracking_number}', tn) if tmpl and tn else ''),
                'shipped_at': r['shipped_at'],
                'is_reship': (r['is_reship'] if 'is_reship' in rk else 0) or 0,
                'reship_reason': (r['reship_reason'] if 'reship_reason' in rk else '') or '',
            })

    # Get carriers for tracking URL
    carriers = {c['slug']: c for c in conn.execute('SELECT * FROM shipping_carriers').fetchall()}
    conn.close()
    
    # Mapping for Advanced Shipment Tracking Pro provider slugs
    ast_provider_mapping = {
        'inpost-paczkomaty': ('inpost', 'InPost'),
        'inpost': ('inpost', 'InPost'),
        'dpd': ('dpd', 'DPD'),
        'dpd-pl': ('dpd', 'DPD'),
    }
    
    # Build risk index once and pass into each row (cheaper than rebuilding
    # per-order; same orders' billing/shipping JSON is parsed once inside
    # process_shipped_order, so we can't easily share that — but the index
    # itself is the expensive part).
    risk_idx = _build_risk_index(conn)

    result = []
    for order in orders:
        try:
            payload = process_shipped_order(order, conn, carriers, ast_provider_mapping)
            payload['date_created'] = order['date_created']  # 下单日期 (shown in the 订单 column)
            _ps = shipped_parcels_map.get(order['id'], [])
            payload['parcels'] = _ps
            payload['parcels_shipped'] = len(_ps)
            payload['carrier_status'] = order['carrier_status']
            payload['carrier_status_at'] = order['carrier_status_at']
            payload['customer_risk'] = _assess_customer_risk(
                parse_json_field(order['billing']),
                parse_json_field(order['shipping']),
                risk_idx,
                current_order_id=order['id'],
            )
            result.append(payload)
        except Exception as e:
            print(f"Error processing shipped order {order['number']}: {e}")
            continue

    return jsonify(result)


@app.route('/api/shipping/find-by-tracking')
@login_required
@shipping_view_required
def find_orders_by_tracking():
    """Status-agnostic lookup by tracking number.

    Tracking numbers can live in many places depending on which WooCommerce plugin
    the site uses (Advanced Shipment Tracking Pro, Orders Tracking for WooCommerce,
    custom shipping_lines meta, our own shipping_logs, etc.). This endpoint:
      1. uses LIKE on number + every JSON column we know holds tracking data
         to get a candidate list across ALL statuses,
      2. runs each candidate through process_shipped_order() to extract the actual
         tracking number, then keeps only rows where the extracted number matches
         the user's query (substring, case-insensitive).
    """
    q = (request.args.get('q') or '').strip()
    if not q or len(q) < 4:
        return jsonify({'success': False, 'error': '请输入至少 4 位运单号', 'orders': []}), 400

    conn = get_db_connection()

    base_query = '''
        SELECT o.id, o.number, o.status, o.total, o.currency, o.date_created, o.date_modified,
               o.source, o.billing, o.shipping, o.line_items, o.meta_data, o.shipping_lines, o.shipping_total,
               o.customer_note,
               o.is_undelivered, o.shipping_loss_amount, o.undelivered_at, o.undelivered_note,
               o.is_problem_return,
               s.manager,
               sl.tracking_number, sl.carrier_slug, sl.shipped_at,
               u.name AS undelivered_by_name,
               n.note as latest_note, n.date_created as latest_note_date, n.author as latest_note_author
        FROM orders o
        LEFT JOIN sites s ON o.source = s.url
        LEFT JOIN shipping_logs sl ON sl.id = (
            SELECT id FROM shipping_logs WHERE order_id = o.id ORDER BY id DESC LIMIT 1
        )
        LEFT JOIN users u ON o.undelivered_by = u.id
        LEFT JOIN (
            SELECT order_id, note, date_created, author
            FROM order_notes
            WHERE customer_note = 0
            GROUP BY order_id
            HAVING date_created = MAX(date_created)
        ) n ON o.id = n.order_id
        WHERE (
            o.number LIKE ?
            OR sl.tracking_number LIKE ?
            OR o.meta_data LIKE ?
            OR o.line_items LIKE ?
            OR o.shipping_lines LIKE ?
        )
    '''
    like_term = f'%{q}%'
    params = [like_term] * 5

    allowed_sources = get_user_allowed_sources(current_user.id, current_user.is_admin(), current_user.is_viewer())
    if allowed_sources is not None:
        if not allowed_sources:
            conn.close()
            return jsonify({'success': True, 'orders': [], 'count': 0})
        placeholders = ','.join(['?' for _ in allowed_sources])
        base_query += f' AND o.source IN ({placeholders})'
        params.extend(allowed_sources)

    # Limit candidates to keep the JSON-parsing post-filter cheap
    base_query += ' ORDER BY o.date_modified DESC, o.date_created DESC LIMIT 200'

    candidates = conn.execute(base_query, params).fetchall()

    carriers = {c['slug']: c for c in conn.execute('SELECT * FROM shipping_carriers').fetchall()}
    ast_provider_mapping = {
        'inpost-paczkomaty': ('inpost', 'InPost'),
        'inpost': ('inpost', 'InPost'),
        'dpd': ('dpd', 'DPD'),
        'dpd-pl': ('dpd', 'DPD'),
    }

    q_lower = q.lower()
    results = []
    for row in candidates:
        try:
            payload = process_shipped_order(row, conn, carriers, ast_provider_mapping)
        except Exception as e:
            print(f"Error parsing order {row['number']} for tracking lookup: {e}")
            continue
        # Confirm the LIKE match by checking either the order number or the
        # extracted tracking number — drops false positives where the query
        # string happens to appear elsewhere in the JSON blob.
        order_no = str(payload.get('number') or '').lower()
        tracking = str(payload.get('tracking_number') or '').lower()
        if q_lower in order_no or (tracking and q_lower in tracking):
            payload['status'] = row['status']
            results.append(payload)

    conn.close()
    return jsonify({'success': True, 'orders': results, 'count': len(results)})


# ============================================================================
# Tracking format detection + payload builders
#
# WordPress sites we ship to use one of three plugins for the order-edit
# tracking metabox. Each one stores the tracking in a different meta key, so
# the only way to make our "发货" button mirror manual entry in WP-admin is
# to detect which one a given site uses and write the correct shape.
#
#   ast              → meta_data['_wc_shipment_tracking_items']  (Advanced Shipment Tracking)
#   villatheme       → line_items[*].meta_data['_vi_wot_order_item_tracking_data']
#                       (Orders Tracking for WooCommerce by VillaTheme)
#   custom_lineitem  → line_items[*].meta_data['tracking_number'] (poland.php / others)
#
# We always also write the generic '_tracking_number' / '_tracking_provider'
# keys so legacy themes and `wc_get_order_tracking_number()` still find it.
# ============================================================================

def detect_site_tracking_format(conn, site_url):
    """Look at recent shipped orders for this site to figure out which plugin
    holds the tracking. Returns 'ast' | 'villatheme' | 'custom_lineitem' | 'unknown'."""
    rows = conn.execute("""
        SELECT meta_data, line_items FROM orders
        WHERE source = ? AND status IN ('on-hold','shipped','completed')
        ORDER BY date_modified DESC LIMIT 10
    """, (site_url,)).fetchall()

    ast = villa = custom = 0
    for r in rows:
        md = r['meta_data'] or ''
        li = r['line_items'] or ''
        if '_wc_shipment_tracking_items' in md:
            ast += 1
        if '_vi_wot_order_item_tracking_data' in li:
            villa += 1
        elif '"key":"tracking_number"' in li or '"key": "tracking_number"' in li:
            custom += 1

    if ast and ast >= max(villa, custom):
        return 'ast'
    if villa and villa >= custom:
        return 'villatheme'
    if custom:
        return 'custom_lineitem'
    return 'unknown'


def _ast_provider_for_carrier(carrier_slug):
    """Map our carrier slug to the AST plugin's tracking_provider value.

    Most slugs already match AST's expected provider name (australia-post,
    dhl-express, fedex, ups, ...). A few legacy short slugs need aliasing.
    """
    cs = (carrier_slug or '').lower().strip()
    aliases = {
        'inpost': 'inpost-paczkomaty',
        'dpd': 'dpd-pl',  # our 'dpd' carrier row stores the PL tracking URL
        'auspost': 'australia-post',
        'australia_post': 'australia-post',
    }
    return aliases.get(cs, cs or 'custom')


def build_ast_tracking_value(tracking_number, carrier_slug, line_items):
    """Build the value for _wc_shipment_tracking_items (a list of one tracking record)."""
    import time, hashlib
    products = []
    for it in line_items or []:
        if isinstance(it, dict) and it.get('id'):
            products.append({
                'product': str(it.get('product_id', '')),
                'item_id': str(it.get('id')),
                'qty': str(it.get('quantity', 1))
            })

    return [{
        'tracking_number': tracking_number,
        'shipping_note': '',
        'tracking_provider': _ast_provider_for_carrier(carrier_slug),
        'custom_tracking_link': '',
        'tracking_product_code': '',
        'date_shipped': str(int(time.time())),
        'products_list': products,
        'status_shipped': '1',
        'tracking_id': hashlib.md5(f"{tracking_number}{time.time()}".encode()).hexdigest(),
    }]


def build_villatheme_lineitem_payload(tracking_number, carrier_slug, carrier_name, tracking_url_template, line_items):
    """Build line_items payload that adds _vi_wot_order_item_tracking_data to each item.

    `tracking_url_template` is the carrier's URL with a `{tracking}` or
    `{tracking_number}` placeholder (as stored in the shipping_carriers DB).
    We normalize to VillaTheme's `{tracking_number}` placeholder.
    """
    import time, json as _json

    url_template = (tracking_url_template or '').replace('{tracking}', '{tracking_number}')

    carrier_data = {
        'carrier_slug': carrier_slug or 'custom',
        'carrier_name': carrier_name or (carrier_slug.title() if carrier_slug else 'Custom'),
        'carrier_url': url_template,
        'carrier_type': 'custom-carrier',
    }

    # VillaTheme stores the tracking record as a JSON-encoded string
    tracking_value = _json.dumps([{
        'tracking_number': tracking_number,
        **carrier_data,
        'time': int(time.time())
    }])

    out = []
    for it in line_items or []:
        if isinstance(it, dict) and it.get('id'):
            out.append({
                'id': it.get('id'),
                'meta_data': [{'key': '_vi_wot_order_item_tracking_data', 'value': tracking_value}]
            })
    return out


# ── Multi-parcel (split shipment / 分批发货) builders ──────────────────────────
# A "parcel" dict carries everything a tracking record needs:
#   {tracking_number, carrier_slug, carrier_name, tracking_url, date_shipped}
# date_shipped is a unix-timestamp string. shipping_logs is the source of truth
# for which parcels an order has — we always rebuild the FULL tracking value
# from the parcel list and PUT it (the PUT replaces the meta), so AST/VillaTheme
# end up holding every parcel's tracking number, not just the latest.

def build_ast_tracking_items(parcels, line_items):
    """Build _wc_shipment_tracking_items as a list with ONE record per parcel.

    AST natively renders each list entry as a separate shipment, so a split
    order shows all its tracking numbers. Single-parcel orders pass a 1-element
    list and behave exactly as before.
    """
    import time, hashlib
    products = []
    for it in line_items or []:
        if isinstance(it, dict) and it.get('id'):
            products.append({
                'product': str(it.get('product_id', '')),
                'item_id': str(it.get('id')),
                'qty': str(it.get('quantity', 1))
            })
    items = []
    for p in parcels or []:
        tn = (p.get('tracking_number') or '').strip()
        if not tn:
            continue
        ds = str(p.get('date_shipped') or int(time.time()))
        items.append({
            'tracking_number': tn,
            'shipping_note': '',
            'tracking_provider': _ast_provider_for_carrier(p.get('carrier_slug')),
            'custom_tracking_link': '',
            'tracking_product_code': '',
            'date_shipped': ds,
            'products_list': products,
            'status_shipped': '1',
            # Stable per (number, date) so re-PUTs don't churn the id.
            'tracking_id': hashlib.md5(f"{tn}{ds}".encode()).hexdigest(),
        })
    return items


def build_villatheme_lineitem_payload_multi(parcels, line_items):
    """VillaTheme stores _vi_wot_order_item_tracking_data as a JSON array per
    line item. We write the SAME array (one entry per parcel) onto every item,
    so a split order keeps all tracking numbers."""
    import json as _json
    records = []
    for p in parcels or []:
        tn = (p.get('tracking_number') or '').strip()
        if not tn:
            continue
        url_template = (p.get('tracking_url_template') or '').replace('{tracking}', '{tracking_number}')
        records.append({
            'tracking_number': tn,
            'carrier_slug': p.get('carrier_slug') or 'custom',
            'carrier_name': p.get('carrier_name') or (p.get('carrier_slug') or 'Custom'),
            'carrier_url': url_template,
            'carrier_type': 'custom-carrier',
            'time': int(p.get('date_shipped') or 0),
        })
    tracking_value = _json.dumps(records)
    out = []
    for it in line_items or []:
        if isinstance(it, dict) and it.get('id'):
            out.append({
                'id': it.get('id'),
                'meta_data': [{'key': '_vi_wot_order_item_tracking_data', 'value': tracking_value}]
            })
    return out


def _post_fallback_customer_note(req, site, order, carrier_name, tracking_number, tracking_url, api_headers, warnings):
    """Last-resort fallback when the trigger-shipment-email endpoint is missing
    on the site (mu-plugin not installed yet). Posts a customer_note=true so
    WooCommerce sends its built-in 'Customer Note' email — uglier than the
    plugin-native one but better than no notification at all."""
    try:
        note_url = f"{site['url']}/wp-json/wc/v3/orders/{woo_post_id(order['id'])}/notes"
        if tracking_url:
            note_content = (
                f"Order has been shipped via {carrier_name}. "
                f"Tracking Number: <a href='{tracking_url}'>{tracking_number}</a>"
                f"\n<br>Track your package: <a href='{tracking_url}'>{tracking_url}</a>"
            )
        else:
            note_content = f"Order has been shipped via {carrier_name}. Tracking Number: {tracking_number}"
        note_resp = req.post(
            note_url,
            json={'note': note_content, 'customer_note': True},
            auth=(site['consumer_key'], site['consumer_secret']),
            timeout=45,
            headers=api_headers,
        )
        if note_resp.status_code not in (200, 201):
            warnings.append(f"回退客户备注失败 HTTP {note_resp.status_code}")
    except Exception as e:
        warnings.append(f"回退客户备注异常: {e}")


def target_status_for_format(fmt):
    """Pick the order status that semantically means 'shipped' on this site.

    AST plugin registers a custom 'shipped' status and uses it as the post-ship
    state. VillaTheme/poland.php sites stay at the standard WC 'on-hold' status
    after shipping (the user confirmed this is the expected behavior on those
    sites). Mirroring what manual entry in WP-admin would do.
    """
    return 'shipped' if fmt == 'ast' else 'on-hold'


def build_custom_lineitem_payload(tracking_number, carrier_slug, tracking_url, line_items):
    """Build line_items payload for sites that store a bare 'tracking_number' meta per item."""
    out = []
    for it in line_items or []:
        if isinstance(it, dict) and it.get('id'):
            meta = [{'key': 'tracking_number', 'value': tracking_number}]
            if tracking_url:
                meta.append({'key': 'tracking_url', 'value': tracking_url})
            if carrier_slug:
                meta.append({'key': 'carrier_slug', 'value': carrier_slug})
            out.append({'id': it.get('id'), 'meta_data': meta})
    return out


@app.route('/api/shipping/ship', methods=['POST'])
@login_required
@shipper_required
@order_site_editable
def ship_order():
    """Mirror manual tracking entry in WP-admin.

    For the chosen site we detect which tracking plugin it uses, then PUT
    /orders/{id} with the right meta_data shape AND status=on-hold in a single
    request. After that we (optionally) post a customer-visible note so WC
    sends the shipment email — same path the WP admin order-edit screen takes.
    """
    import requests as req
    import time

    data = request.json or {}
    order_id = data.get('order_id')
    tracking_number = (data.get('tracking_number') or '').strip()
    carrier_slug = (data.get('carrier_slug') or '').strip()
    send_email = data.get('send_email', True)
    # 分批发货 (split shipment) flags:
    #   new_parcel   → this action adds ANOTHER parcel: append a shipping_logs
    #                  row and rebuild the WP tracking value from ALL parcels.
    #   more_batches → not the final batch: leave the order status untouched so
    #                  it stays in the 待发货 queue and survives WC re-sync
    #                  (sync overwrites status). Status only advances on the
    #                  final batch (more_batches=False).
    new_parcel = bool(data.get('new_parcel'))
    more_batches = bool(data.get('more_batches'))
    # 补发货 (re-shipment): re-send a NEW tracking after the first parcel was
    # lost / never sent. A non-empty reship_reason flags it. A reship writes ONLY
    # the new parcel to WP (replace — the customer sees just the live tracking),
    # but keeps the original parcel row locally for history and records the
    # reason in an audit note. It is never a split / partial batch.
    reship_reason = (data.get('reship_reason') or '').strip()
    is_reship = bool(reship_reason)
    if is_reship:
        new_parcel = False
        more_batches = False

    if not all([order_id, tracking_number, carrier_slug]):
        return jsonify({'success': False, 'error': '缺少必填字段'}), 400

    conn = get_db_connection()

    order = conn.execute(
        'SELECT id, number, source, status, line_items FROM orders WHERE id = ?',
        (order_id,)
    ).fetchone()
    if not order:
        conn.close()
        return jsonify({'success': False, 'error': '订单不存在'}), 404

    site = conn.execute('SELECT * FROM sites WHERE url = ?', (order['source'],)).fetchone()
    if not site:
        conn.close()
        return jsonify({'success': False, 'error': '站点配置不存在'}), 404

    carrier = conn.execute('SELECT name, tracking_url FROM shipping_carriers WHERE slug = ?', (carrier_slug,)).fetchone()
    carrier_name = carrier['name'] if carrier else carrier_slug
    tracking_url_template = (carrier['tracking_url'] if carrier else '') or ''

    # Customer-facing tracking URL — resolve placeholders from the DB template.
    # No carrier-specific hardcoding here; if a new carrier needs a URL, add a
    # row to shipping_carriers instead.
    tracking_url = ''
    if tracking_url_template:
        tracking_url = tracking_url_template.replace('{tracking}', tracking_number).replace('{tracking_number}', tracking_number)

    line_items = parse_json_field(order['line_items']) or []
    fmt = detect_site_tracking_format(conn, order['source'])
    fmt_label = {'ast': 'AST', 'villatheme': 'VillaTheme', 'custom_lineitem': '自定义', 'unknown': '默认(AST)'}.get(fmt, fmt)
    target_status = target_status_for_format(fmt)

    api_headers = {
        "User-Agent": "WooCommerce API Client-Python/3.0.0",
        "Content-Type": "application/json",
        "Accept": "application/json"
    }

    # Assemble the FULL parcel list to write to WP. shipping_logs is the source
    # of truth for already-shipped parcels; on a new_parcel action we append,
    # otherwise (normal ship / correction) we replace with just this one.
    new_ds = int(time.time())
    carriers_by_slug = {r['slug']: r for r in conn.execute(
        'SELECT slug, name, tracking_url FROM shipping_carriers').fetchall()}

    def _mk_parcel(slug, tn, ds):
        c = carriers_by_slug.get(slug)
        tmpl = ((c['tracking_url'] if c else '') or '')
        return {
            'tracking_number': tn,
            'carrier_slug': slug,
            'carrier_name': (c['name'] if c else slug),
            'tracking_url_template': tmpl,
            'date_shipped': ds,
        }

    def _shipped_at_unix(s):
        if not s:
            return new_ds
        for f in ('%Y-%m-%d %H:%M:%S', '%Y-%m-%dT%H:%M:%S'):
            try:
                return int(datetime.strptime(str(s)[:19], f).timestamp())
            except Exception:
                pass
        return new_ds

    if new_parcel:
        prior = conn.execute(
            'SELECT tracking_number, carrier_slug, shipped_at FROM shipping_logs WHERE order_id=? ORDER BY id',
            (order_id,)).fetchall()
        parcels = [_mk_parcel(r['carrier_slug'], r['tracking_number'], _shipped_at_unix(r['shipped_at']))
                   for r in prior if (r['tracking_number'] or '').strip()]
        parcels.append(_mk_parcel(carrier_slug, tracking_number, new_ds))
    else:
        parcels = [_mk_parcel(carrier_slug, tracking_number, new_ds)]

    # Build the PUT payload: tracking meta (+ status, unless more batches follow).
    # Status varies by site: AST uses its custom 'shipped' status; VillaTheme
    # and poland.php sites stay at the standard 'on-hold' status after shipping.
    # During a partial shipment (more_batches) we DON'T send status, so the
    # order stays in its pre-ship state (待发货) until the final batch.
    put_payload = {}
    if not more_batches:
        put_payload['status'] = target_status
    base_meta = [
        {'key': '_tracking_number', 'value': tracking_number},
        {'key': '_tracking_provider', 'value': carrier_slug},
        {'key': '_date_shipped', 'value': str(new_ds)},
    ]

    if fmt == 'villatheme':
        put_payload['meta_data'] = base_meta
        put_payload['line_items'] = build_villatheme_lineitem_payload_multi(parcels, line_items)
    elif fmt == 'custom_lineitem':
        put_payload['meta_data'] = base_meta
        # custom_lineitem stores a single tracking meta per item; multi-parcel
        # isn't representable there, so we write the latest parcel (the full
        # parcel history is always kept locally in shipping_logs).
        put_payload['line_items'] = build_custom_lineitem_payload(tracking_number, carrier_slug, tracking_url, line_items)
    else:
        # ast or unknown — write AST format. Even if AST plugin isn't installed,
        # the meta is harmless and other readers (legacy code, our own sync)
        # already understand _wc_shipment_tracking_items.
        put_payload['meta_data'] = base_meta + [
            {'key': '_wc_shipment_tracking_items',
             'value': build_ast_tracking_items(parcels, line_items)},
        ]

    status_url = f"{site['url']}/wp-json/wc/v3/orders/{woo_post_id(order['id'])}"
    warnings = []
    remote_success = False

    # PUT with retry. The same payload is idempotent so retrying after timeout is safe.
    for attempt in range(3):
        try:
            resp = req.put(
                status_url,
                json=put_payload,
                auth=(site['consumer_key'], site['consumer_secret']),
                timeout=60,
                headers=api_headers
            )
            print(f"[SHIP] {site['url']} order {order['id']} fmt={fmt} attempt={attempt+1} status={resp.status_code}")
            if resp.status_code in (200, 201):
                remote_success = True
                break
            body = (resp.text or '')[:300]
            if body.lstrip().startswith(('<!', '<html')):
                warnings.append(f"WP 返回 HTML（疑似 WAF 拦截 / 认证失败），HTTP {resp.status_code}")
            else:
                warnings.append(f"远程返回 {resp.status_code}: {body}")
        except (req.exceptions.ConnectionError, req.exceptions.Timeout) as e:
            print(f"[SHIP] {site['url']} order {order['id']} attempt {attempt+1} timed out: {e}")
            # Verify by GET — the PUT may have actually applied even if the
            # response never made it back.
            try:
                check = req.get(status_url, auth=(site['consumer_key'], site['consumer_secret']), timeout=30)
                # On a final batch we confirm via the status flip; during a
                # partial batch (more_batches) status is intentionally unchanged,
                # so a reachable 200 is the best confirmation we can get (the PUT
                # is idempotent, so retrying is harmless anyway).
                if check.status_code == 200 and (more_batches or check.json().get('status') == target_status):
                    remote_success = True
                    warnings.append("PUT 响应超时，但二次查询确认订单可达")
                    break
            except Exception as verify_err:
                print(f"[SHIP] verify after timeout failed: {verify_err}")
            if attempt < 2:
                time.sleep(2)
        except Exception as e:
            warnings.append(f"远程异常: {e}")
            break

    if not remote_success:
        # For a NEW parcel (split shipment) we don't stash anything: the parcel
        # didn't actually ship, and inserting a row would create a phantom
        # parcel. The user just retries. For a normal/first shipment we stash
        # the tracking number locally so it doesn't have to be retyped on retry.
        if not new_parcel and not is_reship:
            existing_log = conn.execute('SELECT id FROM shipping_logs WHERE order_id = ?', (order_id,)).fetchone()
            if existing_log:
                conn.execute(
                    "UPDATE shipping_logs SET tracking_number=?, carrier_slug=?, shipped_by=?, shipped_at=datetime('now') WHERE order_id=?",
                    (tracking_number, carrier_slug, current_user.id, order_id)
                )
            else:
                conn.execute(
                    '''INSERT INTO shipping_logs (order_id, woo_order_id, source, tracking_number, carrier_slug, shipped_by)
                       VALUES (?, ?, ?, ?, ?, ?)''',
                    (order_id, order['number'], order['source'], tracking_number, carrier_slug, current_user.id)
                )
            conn.commit()
        # A failed reship must NOT stash/overwrite the original tracking row.
        stash_note = '' if (new_parcel or is_reship) else '运单号已暂存本地，请稍后重试。'
        conn.close()
        return jsonify({
            'success': False,
            'error': f"发货失败（{fmt_label}）：{'; '.join(warnings) or '远程未返回成功'}。{stash_note}"
        }), 500

    # Build the re-shipment audit line once; reused for the local order_notes
    # row and the best-effort remote WC note below.
    reship_log_line = ''
    if is_reship:
        reship_log_line = (
            f"「补发货」由 {current_user.name} 操作：物流商 {carrier_name}，"
            f"新运单号 {tracking_number}。补发原因：{reship_reason}"
        )

    # Trigger the site's native shipment email.
    #
    # Why this is a separate call (and not just the meta_data PUT above):
    #   - AST hooks into the wc-shipped status transition, so its email is
    #     already on its way the moment the PUT lands. Our trigger endpoint
    #     is a no-op for AST.
    #   - VillaTheme (Orders Tracking for WooCommerce) does NOT hook into
    #     status changes — its admin save handler invokes
    #     VI_WOO_ORDERS_TRACKING_ADMIN_IMPORT_CSV::send_mail() directly.
    #     Setting meta via REST never triggers send_mail(), so VillaTheme
    #     sites stopped emailing customers after the ship_order rewrite.
    #     The trigger endpoint replicates the admin call.
    # The endpoint itself decides which path applies; we just always call it
    # and let it self-detect.
    email_trigger_info = None
    # Did a customer-facing notification actually go out? Only consumed by the
    # reship fallback below; harmless for normal ships.
    customer_notified = False
    if send_email:
        try:
            trig_url = f"{site['url']}/wp-json/woo-tracking/v1/orders/{woo_post_id(order['id'])}/trigger-shipment-email"
            trig_resp = req.post(
                trig_url,
                json={'tracking_number': tracking_number, 'carrier_slug': carrier_slug},
                params={'consumer_key': site['consumer_key'], 'consumer_secret': site['consumer_secret']},
                timeout=30,
                headers=api_headers,
            )
            if trig_resp.status_code == 200:
                td = trig_resp.json() if trig_resp.text else {}
                email_trigger_info = td
                # Surface a warning only if we KNOW the email failed; AST
                # returns email_sent=null because it's async, that's fine.
                if td.get('email_sent') is True:
                    customer_notified = True
                elif td.get('email_sent') is False and td.get('plugin') not in ('AST',):
                    warnings.append(f"邮件触发失败（{td.get('plugin', '?')}）: {td.get('note', '')}")
            elif trig_resp.status_code == 404:
                # Plugin not installed on this site — fall back to the
                # legacy customer_note approach so the buyer still gets
                # *some* notification.
                warnings.append("邮件触发端点未找到（请确认 woo-orders-tracking-rest-api mu-plugin 已安装）")
                _post_fallback_customer_note(req, site, order, carrier_name, tracking_number, tracking_url, api_headers, warnings)
                customer_notified = True
            else:
                warnings.append(f"邮件触发返回 {trig_resp.status_code}")
        except Exception as e:
            warnings.append(f"邮件触发异常: {e}")

    # 补发 MUST reach the customer with the NEW tracking. The plugin-native
    # shipment email only reliably fires for VillaTheme (direct send_mail); AST
    # emails on a wc-shipped status *transition* that a reship doesn't cause
    # (the order is already shipped), and custom/unknown sites have no native
    # trigger at all. So whenever nothing customer-facing went out above, fall
    # back to a WooCommerce customer note that emails the buyer the new tracking.
    if is_reship and send_email and not customer_notified:
        _post_fallback_customer_note(req, site, order, carrier_name, tracking_number, tracking_url, api_headers, warnings)

    # Re-shipment: leave an explicit audit note on the WC order so the WP admin
    # reflects WHY a second tracking went out (best-effort; mirrors the
    # mark-undelivered remote-note path).
    if is_reship and reship_log_line:
        try:
            req.post(
                f"{site['url']}/wp-json/wc/v3/orders/{woo_post_id(order['id'])}/notes",
                json={'note': reship_log_line, 'customer_note': False},
                auth=(site['consumer_key'], site['consumer_secret']),
                timeout=10,
                headers=api_headers,
            )
        except Exception as remote_err:
            app.logger.warning(f"Remote reship note for order {order_id} failed: {remote_err}")

    # Local DB: status + shipping_logs (mirror the remote state we just set).
    try:
        # Advance status only on the final batch. During a partial shipment the
        # order keeps its pre-ship status so it stays in the 待发货 queue.
        if not more_batches:
            conn.execute("UPDATE orders SET status=? WHERE id=?", (target_status, order_id))
        if new_parcel:
            # Split shipment: record this parcel as its own row.
            conn.execute(
                '''INSERT INTO shipping_logs (order_id, woo_order_id, source, tracking_number, carrier_slug, shipped_by)
                   VALUES (?, ?, ?, ?, ?, ?)''',
                (order_id, order['number'], order['source'], tracking_number, carrier_slug, current_user.id)
            )
        elif is_reship:
            # Re-shipment: keep the original parcel row(s) untouched and append a
            # NEW row flagged is_reship with the reason, so the full history is
            # preserved locally even though WP only carries the new (replacing)
            # tracking. Also write an order_notes audit row (local source of
            # truth, always succeeds even if the remote note above failed).
            conn.execute(
                '''INSERT INTO shipping_logs (order_id, woo_order_id, source, tracking_number, carrier_slug, shipped_by, reship_reason, is_reship)
                   VALUES (?, ?, ?, ?, ?, ?, ?, 1)''',
                (order_id, order['number'], order['source'], tracking_number, carrier_slug, current_user.id, reship_reason)
            )
            conn.execute(
                '''INSERT INTO order_notes (order_id, note, date_created, customer_note, author, added_by_user)
                   VALUES (?, ?, datetime('now'), 0, ?, 1)''',
                (order_id, reship_log_line, current_user.name)
            )
        else:
            existing_log = conn.execute('SELECT id FROM shipping_logs WHERE order_id=?', (order_id,)).fetchone()
            if existing_log:
                conn.execute(
                    "UPDATE shipping_logs SET tracking_number=?, carrier_slug=?, shipped_by=?, shipped_at=datetime('now') WHERE order_id=?",
                    (tracking_number, carrier_slug, current_user.id, order_id)
                )
            else:
                conn.execute(
                    '''INSERT INTO shipping_logs (order_id, woo_order_id, source, tracking_number, carrier_slug, shipped_by)
                       VALUES (?, ?, ?, ?, ?, ?)''',
                    (order_id, order['number'], order['source'], tracking_number, carrier_slug, current_user.id)
                )
        conn.commit()
    except Exception as e:
        conn.close()
        return jsonify({'success': False, 'error': f"本地数据库更新失败: {e}"}), 500
    conn.close()

    # Annotate the success message — split-shipment aware.
    parcel_count = len(parcels)
    if is_reship:
        msg = f"补发成功（{fmt_label} 格式），新运单 {tracking_number}"
    elif more_batches:
        msg = f"第 {parcel_count} 个包裹已发货（分批中，订单仍在待发货，可继续添加包裹）"
    elif new_parcel and parcel_count > 1:
        msg = f"最后一个包裹已发货，共 {parcel_count} 个包裹，订单已完成发货（{fmt_label} 格式）"
    else:
        msg = f"发货成功（{fmt_label} 格式）"
    if is_reship:
        if not send_email:
            msg += "，未通知客户（未勾选发邮件）"
        elif customer_notified:
            plugin = (email_trigger_info or {}).get('plugin', '')
            msg += f"，新运单邮件已发送（{plugin}）" if plugin else "，新运单邮件已发送"
        else:
            # AST / custom / unknown: the native shipment email can't fire on a
            # reship (no status transition), so we emailed the buyer via a
            # WooCommerce customer note instead.
            msg += "，已通过客户备注邮件通知客户新运单"
    elif email_trigger_info:
        plugin = email_trigger_info.get('plugin', '?')
        sent = email_trigger_info.get('email_sent')
        if sent is True:
            msg += f"，发货邮件已发送（{plugin}）"
        elif sent is None and plugin == 'AST':
            msg += "，AST 已自动安排发货邮件"
        elif sent is False:
            msg += f"，邮件未发送：{email_trigger_info.get('note', '未识别物流插件')}"
    resp_extra = {'format': fmt, 'email_trigger': email_trigger_info,
                  'parcel_count': parcel_count, 'more_batches': more_batches}
    if warnings:
        return jsonify({'success': True, 'warning': True,
                        'message': msg + ' — 警告: ' + '; '.join(warnings), **resp_extra})
    return jsonify({'success': True, 'message': msg, **resp_extra})


@app.route('/api/shipping/debug/<order_id>', methods=['POST'])
@login_required
@shipper_required
@order_site_editable
def debug_tracking_sync(order_id):
    """Debug endpoint to manually resync tracking number to WordPress"""
    import requests as req
    
    conn = get_db_connection()
    
    # Get order and tracking info
    order = conn.execute('''
        SELECT o.*, sl.tracking_number, sl.carrier_slug
        FROM orders o
        LEFT JOIN shipping_logs sl ON sl.id = (
            SELECT id FROM shipping_logs WHERE order_id = o.id ORDER BY id DESC LIMIT 1
        )
        WHERE o.id = ?
    ''', (order_id,)).fetchone()
    
    if not order:
        conn.close()
        return jsonify({'success': False, 'error': '订单不存在'}), 404
    
    if not order['tracking_number']:
        conn.close()
        return jsonify({'success': False, 'error': '订单没有运单号记录'}), 400
    
    # Get site credentials
    site = conn.execute('SELECT * FROM sites WHERE url = ?', (order['source'],)).fetchone()
    conn.close()
    
    if not site:
        return jsonify({'success': False, 'error': '站点配置不存在'}), 404
    
    # Prepare the tracking API request
    tracking_url_api = f"{site['url']}/wp-json/woo-orders-tracking/v1/tracking/set"
    tracking_payload = {
        'order_id': order['number'],
        'tracking_data': [{
            'tracking_number': order['tracking_number'],
            'carrier_slug': order['carrier_slug']
        }],
        'send_email': False
    }
    
    debug_info = {
        'order_id': order['id'],
        'order_number': order['number'],
        'tracking_number': order['tracking_number'],
        'carrier_slug': order['carrier_slug'],
        'api_endpoint': tracking_url_api,
        'payload': tracking_payload
    }
    
    # Attempt to sync
    try:
        # Headers to simulate official WooCommerce API client to bypass WAF
        headers = {
            "User-Agent": "WooCommerce API Client-Python/3.0.0",
            "Content-Type": "application/json",
            "Accept": "application/json"
        }
        
        tracking_resp = req.post(
            tracking_url_api,
            json=tracking_payload,
            auth=(site['consumer_key'], site['consumer_secret']),
            timeout=30,
            headers=headers
        )
        
        debug_info['response_status'] = tracking_resp.status_code
        debug_info['response_body'] = tracking_resp.text
        
        if tracking_resp.status_code in [200, 201]:
            return jsonify({
                'success': True,
                'message': f'运单号 {order["tracking_number"]} 已成功同步到WordPress',
                'debug_info': debug_info
            })
        else:
            return jsonify({
                'success': False,
                'error': f'同步失败: API返回状态 {tracking_resp.status_code}',
                'debug_info': debug_info
            }), 400
            
    except Exception as e:
        debug_info['exception'] = str(e)
        return jsonify({
            'success': False,
            'error': f'同步异常: {str(e)}',
            'debug_info': debug_info
        }), 500


@app.route('/api/shipping/complete/<order_id>', methods=['POST'])
@login_required
@shipper_required
@order_site_editable
def complete_order(order_id):
    """Mark order as completed"""
    import requests as req
    
    conn = get_db_connection()
    
    order = conn.execute('SELECT id, number, source FROM orders WHERE id = ?', (order_id,)).fetchone()
    if not order:
        conn.close()
        return jsonify({'success': False, 'error': '订单不存在'}), 404
    
    site = conn.execute('SELECT * FROM sites WHERE url = ?', (order['source'],)).fetchone()
    if not site:
        conn.close()
        return jsonify({'success': False, 'error': '站点配置不存在'}), 404
    
    try:
        # Update order status via WooCommerce API
        url = f"{site['url']}/wp-json/wc/v3/orders/{woo_post_id(order['id'])}"
        resp = req.put(
            url,
            json={'status': 'completed'},
            auth=(site['consumer_key'], site['consumer_secret']),
            timeout=30
        )
        
        if resp.status_code not in [200, 201]:
            raise Exception(f"API错误: {resp.text}")
        
        # Update local database
        conn.execute("UPDATE orders SET status = 'completed' WHERE id = ?", (order_id,))
        conn.execute("UPDATE shipping_logs SET status = 'completed', completed_at = datetime('now') WHERE order_id = ?", (order_id,))
        conn.commit()
        conn.close()
        
        return jsonify({'success': True, 'message': '订单已完成'})
        
    except Exception as e:
        conn.close()
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/shipping/carriers')
@login_required
@shipping_view_required
def get_carriers():
    """Carriers visible to the current user.

    Admins / viewers see every active row in `shipping_carriers`.
    Restricted users see only the carriers actually in use on sites they
    have permission to ship from — so a PL-only user never sees Australia
    Post in the dropdown, and an AU-only user never sees InPost / DPD.
    Used by the ship modal as the fallback when a site has no shipping
    history yet.
    """
    conn = get_db_connection()
    allowed_sources = get_user_allowed_sources(current_user.id, current_user.is_admin(), current_user.is_viewer())

    if allowed_sources is None:
        # No restrictions
        carriers = conn.execute(
            'SELECT id, slug, name, tracking_url FROM shipping_carriers WHERE is_active = 1'
        ).fetchall()
        conn.close()
        return jsonify([dict(c) for c in carriers])

    if not allowed_sources:
        conn.close()
        return jsonify([])

    # Union of carriers detected across the user's authorized sites
    by_slug = {}
    for site_url in allowed_sources:
        for c in detect_carriers_for_site(conn, site_url):
            slug = c['slug']
            if slug in by_slug:
                by_slug[slug]['usage_count'] = by_slug[slug].get('usage_count', 0) + c.get('usage_count', 0)
            else:
                by_slug[slug] = dict(c)
    conn.close()
    # Sort by usage frequency so the most-used carrier surfaces first
    return jsonify(sorted(by_slug.values(), key=lambda x: -x.get('usage_count', 0)))


def detect_carriers_for_site(conn, site_url, lookback_orders=80):
    """Return the carriers a specific site has actually used, by inspecting
    recent shipped orders' tracking meta_data. Lets the ship modal show only
    options that make sense for *this* site (PL site → InPost/DPD; AU site →
    Australia Post) instead of one global cluttered dropdown.

    Slugs are enriched with name + tracking_url from the shipping_carriers
    table when available; otherwise we fall back to the slug itself as the
    display name and an empty URL.
    """
    rows = conn.execute("""
        SELECT meta_data, line_items FROM orders
        WHERE source = ? AND status IN ('on-hold','shipped','partial-shipped','completed','delivered')
        ORDER BY date_modified DESC LIMIT ?
    """, (site_url, lookback_orders)).fetchall()

    import re
    # VillaTheme generates per-entry slugs like 'custom_1775723085' when the
    # admin types a name freehand. Skip those and use carrier_name instead.
    custom_slug_re = re.compile(r'^custom_\d+$')
    # Lowercased carrier_name → canonical slug we want to surface.
    name_to_slug = {
        'inpost': 'inpost', 'inpost paczkomaty': 'inpost', 'paczkomaty': 'inpost',
        'dpd': 'dpd', 'dpd polska': 'dpd', 'dpd poland': 'dpd', 'dpd-pl': 'dpd',
        'australia post': 'australia-post', 'auspost': 'australia-post', 'australia-post': 'australia-post',
        # EMS shipments out of China to overseas (mostly used on AU sites for
        # the 国内发 warehouse fulfillment). Cover common ways WC plugins or
        # shippers label these so detection picks them up regardless of region.
        'ems': 'ems', 'ems china': 'ems', 'china ems': 'ems', 'china post ems': 'ems',
        'china post': 'ems', 'ems express': 'ems', 'epacket': 'ems',
        '中国邮政': 'ems', '中国邮政ems': 'ems', '邮政ems': 'ems',
    }
    # Slug aliases — collapse AST's regional variants down to our shipping_carriers
    # canonical slug so the URL lookup works in ship_order.
    slug_aliases = {
        'inpost-paczkomaty': 'inpost',
        'inpost-pl': 'inpost',
        'dpd-pl': 'dpd',
        'dpd-polska': 'dpd',
        'auspost': 'australia-post',
        'china-ems': 'ems',
        'china-post': 'ems',
        'china-post-ems': 'ems',
        'ems-china': 'ems',
        'epacket': 'ems',
    }

    def _record_slug(slug, fallback_name):
        """Normalize one tracking record's slug. Drop noise; map custom_NNN
        to a real carrier via its display name; collapse regional aliases."""
        s = (slug or '').strip().lower()
        n = (fallback_name or '').strip().lower()
        if not s and not n:
            return None
        if s and not custom_slug_re.match(s):
            return slug_aliases.get(s, s)
        # Custom_TIMESTAMP — recover a real carrier from carrier_name
        return name_to_slug.get(n)

    seen = {}  # slug → count
    for r in rows:
        # AST format: meta_data._wc_shipment_tracking_items[].tracking_provider
        try:
            md = json.loads(r['meta_data'] or '[]')
            for m in md:
                if isinstance(m, dict) and m.get('key') == '_wc_shipment_tracking_items':
                    items = m.get('value', [])
                    if isinstance(items, list):
                        for it in items:
                            if isinstance(it, dict):
                                slug = _record_slug(it.get('tracking_provider'), '')
                                if slug:
                                    seen[slug] = seen.get(slug, 0) + 1
        except Exception:
            pass

        # VillaTheme: line_items[*].meta_data._vi_wot_order_item_tracking_data
        try:
            li = json.loads(r['line_items'] or '[]')
            for item in li:
                if not isinstance(item, dict):
                    continue
                for m in item.get('meta_data', []):
                    if not isinstance(m, dict):
                        continue
                    if m.get('key') == '_vi_wot_order_item_tracking_data':
                        v = m.get('value')
                        try:
                            recs = json.loads(v) if isinstance(v, str) else v
                        except Exception:
                            recs = None
                        if isinstance(recs, list):
                            for rec in recs:
                                if isinstance(rec, dict):
                                    slug = _record_slug(rec.get('carrier_slug'), rec.get('carrier_name'))
                                    if slug:
                                        seen[slug] = seen.get(slug, 0) + 1
        except Exception:
            pass

    if not seen:
        return []

    # Pull names + URLs for the slugs we found
    db_carriers = {c['slug']: c for c in conn.execute(
        'SELECT slug, name, tracking_url FROM shipping_carriers'
    ).fetchall()}

    out = []
    for slug, count in sorted(seen.items(), key=lambda kv: -kv[1]):
        c = db_carriers.get(slug)
        # AST aliases: 'inpost-paczkomaty' → our DB row 'inpost'; 'dpd-pl' → 'dpd'
        if not c:
            alias = {'inpost-paczkomaty': 'inpost', 'dpd-pl': 'dpd', 'auspost': 'australia-post'}.get(slug)
            if alias:
                c = db_carriers.get(alias)
        out.append({
            'slug': slug,
            'name': (c['name'] if c else slug.replace('-', ' ').title()),
            'tracking_url': (c['tracking_url'] if c else ''),
            'usage_count': count,
        })
    return out


@app.route('/api/shipping/carriers-for-site')
@login_required
@shipping_view_required
def get_carriers_for_site():
    """Return the carriers actually in use on a given site (auto-detected
    from recent shipped orders). Used by the ship modal so each site only
    shows its own relevant carriers.

    For AU sites we always make sure both Australia Post and EMS are present
    even when one of them has zero history — orders fulfilled from the China
    warehouse use EMS and must be selectable from the very first shipment.
    """
    source = request.args.get('source', '').strip()
    if not source:
        return jsonify({'error': 'missing source'}), 400

    # Block probing of sites this user isn't authorized for — otherwise a
    # restricted user could pass any site URL here and learn its carriers.
    allowed_sources = get_user_allowed_sources(current_user.id, current_user.is_admin(), current_user.is_viewer())
    if allowed_sources is not None and source not in allowed_sources:
        return jsonify({'error': '无权访问该站点'}), 403

    conn = get_db_connection()
    carriers = detect_carriers_for_site(conn, source)

    # Augment with country-specific must-have carriers. Detected carriers keep
    # their usage_count and original order; required ones are appended at the
    # end (usage_count = 0) so the dropdown still highlights what the site
    # actually uses most.
    site = conn.execute('SELECT country FROM sites WHERE url = ?', (source,)).fetchone()
    required_slugs_by_country = {
        'AU': ('australia-post', 'ems'),
    }
    if site and site['country'] in required_slugs_by_country:
        present = {c['slug'] for c in carriers}
        for slug in required_slugs_by_country[site['country']]:
            if slug in present:
                continue
            row = conn.execute(
                'SELECT slug, name, tracking_url FROM shipping_carriers WHERE slug = ? AND is_active = 1',
                (slug,)
            ).fetchone()
            if row:
                carriers.append({
                    'slug': row['slug'],
                    'name': row['name'],
                    'tracking_url': row['tracking_url'] or '',
                    'usage_count': 0,
                })

    conn.close()
    return jsonify(carriers)


@app.route('/api/order/<order_id>/emails', methods=['GET'])
@login_required
def get_order_emails(order_id):
    """Fetch emails sent to this order's customer, by calling the source
    site's woo-tracking plugin endpoint. The plugin auto-detects whichever
    email-logging plugin is installed (FluentSMTP / WP Mail SMTP / Email Log)
    and returns a unified shape, so we just relay it.
    """
    import requests as req

    conn = get_db_connection()
    order = conn.execute('SELECT id, number, source FROM orders WHERE id = ?', (order_id,)).fetchone()
    if not order:
        conn.close()
        return jsonify({'success': False, 'error': '订单不存在'}), 404
    site = conn.execute('SELECT * FROM sites WHERE url = ?', (order['source'],)).fetchone()
    conn.close()
    if not site:
        return jsonify({'success': False, 'error': '站点配置不存在'}), 404

    url = f"{site['url']}/wp-json/woo-tracking/v1/orders/{woo_post_id(order['id'])}/email-logs"
    # Pass-through ?debug=1 surfaces the WP-side candidate-table list, so when
    # nothing matches the user can see which logger plugin (and table name)
    # the site actually has.
    debug = '1' if request.args.get('debug') else None
    try:
        params = {'consumer_key': site['consumer_key'], 'consumer_secret': site['consumer_secret']}
        if debug:
            params['debug'] = '1'
        r = req.get(
            url,
            params=params,
            timeout=15,
            headers={"User-Agent": "WooCommerce API Client-Python/3.0.0", "Accept": "application/json"},
        )
        if r.status_code == 200:
            data = r.json() if r.text else {}
            return jsonify({'success': True, **data})
        if r.status_code == 404:
            return jsonify({
                'success': False,
                'error': '该站点未安装 woo-tracking 插件（无法读取邮件日志）',
                'plugin': 'none', 'logs': [], 'count': 0,
            })
        body = (r.text or '')[:200]
        return jsonify({
            'success': False, 'error': f'远程返回 {r.status_code}: {body}',
            'plugin': 'none', 'logs': [], 'count': 0,
        })
    except Exception as e:
        return jsonify({
            'success': False, 'error': f'请求失败: {e}',
            'plugin': 'none', 'logs': [], 'count': 0,
        })


@app.route('/api/customer/emails', methods=['GET'])
@login_required
def get_customer_emails():
    """All emails ever sent to this customer, aggregated across every site
    where they have an order. Each row is tagged with site_id/site_url so
    the frontend can fetch its detail (the log_id namespace is per-site).
    Sites are queried in parallel because each WP round-trip is 1-3 s.
    """
    import requests as req
    import concurrent.futures

    email = (request.args.get('email') or '').strip()
    if not email:
        return jsonify({'success': False, 'error': 'missing email'}), 400

    conn = get_db_connection()
    rows = conn.execute("""
        SELECT DISTINCT s.id AS site_id, s.url AS site_url, s.consumer_key, s.consumer_secret
        FROM orders o
        JOIN sites s ON s.url = o.source
        WHERE json_extract(o.billing, '$.email') = ?
          AND COALESCE(s.consumer_key, '') != '' AND COALESCE(s.consumer_secret, '') != ''
    """, (email,)).fetchall()
    conn.close()
    sites = [dict(r) for r in rows]

    def fetch(site):
        url = f"{site['site_url']}/wp-json/woo-tracking/v1/customer-email-logs"
        try:
            r = req.get(
                url,
                params={'email': email, 'limit': 50,
                        'consumer_key': site['consumer_key'],
                        'consumer_secret': site['consumer_secret']},
                timeout=15,
                headers={"User-Agent": "WooCommerce API Client-Python/3.0.0", "Accept": "application/json"},
            )
            if r.status_code != 200:
                return {'site_id': site['site_id'], 'site_url': site['site_url'], 'logs': [],
                        'error': f"HTTP {r.status_code}", 'plugin': 'none'}
            data = r.json() if r.text else {}
            logs = data.get('logs', []) or []
            for lg in logs:
                lg['site_id'] = site['site_id']
                lg['site_url'] = site['site_url']
            return {'site_id': site['site_id'], 'site_url': site['site_url'],
                    'logs': logs, 'plugin': data.get('plugin', 'none')}
        except Exception as e:
            return {'site_id': site['site_id'], 'site_url': site['site_url'], 'logs': [],
                    'error': str(e), 'plugin': 'none'}

    site_results = []
    if sites:
        with concurrent.futures.ThreadPoolExecutor(max_workers=min(4, len(sites))) as ex:
            site_results = list(ex.map(fetch, sites))

    all_logs = []
    for sr in site_results:
        all_logs.extend(sr.get('logs', []))
    # Sort newest first
    all_logs.sort(key=lambda x: str(x.get('sent_at') or ''), reverse=True)

    return jsonify({
        'success': True,
        'email': email,
        'sites_queried': len(sites),
        'site_summaries': [{'site_id': s['site_id'],
                            'site_url': s['site_url'].replace('https://www.', '').replace('https://', ''),
                            'count': len(s.get('logs', [])),
                            'plugin': s.get('plugin', 'none'),
                            'error': s.get('error', '')} for s in site_results],
        'count': len(all_logs),
        'logs': all_logs[:200],
    })


@app.route('/api/site/<int:site_id>/emails/<int:log_id>', methods=['GET'])
@login_required
def get_site_email_detail(site_id, log_id):
    """Fetch full email detail by (site, log_id). Used by the customer modal
    where there's no specific order to scope to."""
    import requests as req

    conn = get_db_connection()
    site = conn.execute('SELECT * FROM sites WHERE id = ?', (site_id,)).fetchone()
    conn.close()
    if not site:
        return jsonify({'success': False, 'error': '站点不存在'}), 404

    url = f"{site['url']}/wp-json/woo-tracking/v1/email-logs/{log_id}/detail"
    try:
        r = req.get(
            url,
            params={'consumer_key': site['consumer_key'], 'consumer_secret': site['consumer_secret']},
            timeout=20,
            headers={"User-Agent": "WooCommerce API Client-Python/3.0.0", "Accept": "application/json"},
        )
        if r.status_code == 200:
            return jsonify(r.json())
        if r.status_code == 404:
            return jsonify({'success': False, 'error': '邮件记录未找到'}), 404
        return jsonify({'success': False, 'error': f'远程返回 {r.status_code}'}), 502
    except Exception as e:
        return jsonify({'success': False, 'error': f'请求失败: {e}'}), 502


@app.route('/api/order/<order_id>/emails/<int:log_id>', methods=['GET'])
@login_required
def get_order_email_detail(order_id, log_id):
    """Fetch full body / headers for one email log entry. Backed by the WP
    plugin's /orders/{id}/email-logs/{log_id} route, which auto-detects
    the site's logger plugin (FluentSMTP / WP Mail SMTP / etc)."""
    import requests as req

    conn = get_db_connection()
    order = conn.execute('SELECT id, source FROM orders WHERE id = ?', (order_id,)).fetchone()
    if not order:
        conn.close()
        return jsonify({'success': False, 'error': '订单不存在'}), 404
    site = conn.execute('SELECT * FROM sites WHERE url = ?', (order['source'],)).fetchone()
    conn.close()
    if not site:
        return jsonify({'success': False, 'error': '站点配置不存在'}), 404

    url = f"{site['url']}/wp-json/woo-tracking/v1/orders/{woo_post_id(order['id'])}/email-logs/{log_id}"
    try:
        r = req.get(
            url,
            params={'consumer_key': site['consumer_key'], 'consumer_secret': site['consumer_secret']},
            timeout=20,
            headers={"User-Agent": "WooCommerce API Client-Python/3.0.0", "Accept": "application/json"},
        )
        if r.status_code == 200:
            return jsonify(r.json())
        if r.status_code == 404:
            return jsonify({'success': False, 'error': '邮件记录未找到'}), 404
        return jsonify({'success': False, 'error': f'远程返回 {r.status_code}: {(r.text or "")[:200]}'}), 502
    except Exception as e:
        return jsonify({'success': False, 'error': f'请求失败: {e}'}), 502


@app.route('/api/order/<order_id>/notes', methods=['GET'])
@login_required
def get_order_notes(order_id):
    """Get all notes for an order from local database"""
    conn = get_db_connection()
    notes = conn.execute('''
        SELECT id, note, date_created, customer_note, author, added_by_user
        FROM order_notes
        WHERE order_id = ?
        ORDER BY date_created DESC
    ''', (order_id,)).fetchall()
    conn.close()
    
    return jsonify([dict(n) for n in notes])


@app.route('/api/order/<order_id>/note', methods=['POST'])
@login_required
@shipper_required
@order_site_editable
def add_order_note(order_id):
    """Add a note to an order, optionally notifying the customer.

    When notify_customer is True, WP runs the email send synchronously through
    whatever SMTP plugin is installed (FluentSMTP, etc) BEFORE returning the
    response. A slow SMTP relay can easily push that past the default 30 s
    timeout. So we:
      - bump the read timeout to 90 s when an email is involved
      - on timeout, do a verify GET to see if the note actually landed
        (the POST likely succeeded server-side; only the response was lost)
    """
    import requests as req

    data = request.json
    note = data.get('note', '')
    notify_customer = data.get('notify_customer', False)

    if not note:
        return jsonify({'success': False, 'error': '备注内容不能为空'}), 400

    conn = get_db_connection()
    order = conn.execute('SELECT number, source FROM orders WHERE id = ?', (order_id,)).fetchone()
    if not order:
        conn.close()
        return jsonify({'success': False, 'error': '订单不存在'}), 404

    site = conn.execute('SELECT * FROM sites WHERE url = ?', (order['source'],)).fetchone()
    conn.close()

    if not site:
        return jsonify({'success': False, 'error': '站点配置不存在'}), 404

    url = f"{site['url']}/wp-json/wc/v3/orders/{woo_post_id(order_id)}/notes"
    headers = {
        "User-Agent": "WooCommerce API Client-Python/3.0.0",
        "Content-Type": "application/json",
        "Accept": "application/json",
    }
    auth = (site['consumer_key'], site['consumer_secret'])
    # (connect, read) — connect should be fast; read needs to absorb SMTP latency
    timeout = (10, 90 if notify_customer else 30)

    try:
        try:
            resp = req.post(
                url,
                json={'note': note, 'customer_note': notify_customer},
                auth=auth, timeout=timeout, headers=headers,
            )
        except req.exceptions.ReadTimeout:
            # The note POST likely succeeded but the response never arrived.
            # Fetch the latest notes and look for ours by exact content match.
            app.logger.warning(f"add_order_note: read timeout for order {order_id}, verifying via GET...")
            try:
                check = req.get(url, auth=auth, timeout=(10, 30), headers=headers)
                if check.status_code == 200:
                    for n in check.json() or []:
                        if str(n.get('note', '')).strip() == note.strip():
                            return jsonify({
                                'success': True,
                                'message': '备注已添加（远端响应超时，但已确认入库）'
                                           + ('，邮件可能仍在发送中' if notify_customer else ''),
                                'verified_by_get': True,
                            })
            except Exception as verify_err:
                app.logger.warning(f"add_order_note verify GET failed: {verify_err}")
            return jsonify({
                'success': False,
                'error': '请求超时（90s）。如果勾选了「发送邮件」，邮件服务器响应可能慢；'
                         '请稍后到 WordPress 后台确认备注是否已添加。',
            }), 504

        if resp.status_code not in (200, 201):
            return jsonify({'success': False, 'error': f"API错误 {resp.status_code}: {resp.text[:200]}"}), 502

        # Save the new note locally so it shows up immediately without waiting for the next sync
        try:
            note_data = resp.json()
            if note_data and 'id' in note_data:
                from datetime import datetime
                conn = get_db_connection()
                conn.execute('''
                    INSERT OR REPLACE INTO order_notes (
                        wc_note_id, order_id, note, date_created, customer_note, author, added_by_user
                    ) VALUES (?, ?, ?, ?, ?, ?, ?)
                ''', (
                    note_data.get('id'),
                    order_id,
                    note_data.get('note', note),
                    note_data.get('date_created', datetime.now().isoformat()),
                    1 if notify_customer else 0,
                    note_data.get('author', current_user.username if current_user.is_authenticated else ''),
                    1,
                ))
                conn.commit()
                conn.close()
        except Exception as db_err:
            app.logger.error(f"Failed to save note to local DB: {db_err}")

        return jsonify({'success': True, 'message': '备注已添加' + ('，已通知客户' if notify_customer else '')})

    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/order/<order_id>/status', methods=['POST'])
@login_required
@editor_required
@order_site_editable
def update_order_status(order_id):
    """Update order status manually"""
    import requests as req
    
    data = request.json
    new_status = data.get('status', '').strip()
    
    # WooCommerce standard statuses
    valid_statuses = ['pending', 'processing', 'on-hold', 'completed', 'cancelled', 'refunded', 'failed', 'checkout-draft']
    
    if not new_status:
        return jsonify({'success': False, 'error': '请选择新状态'}), 400
    
    if new_status not in valid_statuses:
        return jsonify({'success': False, 'error': f'无效的订单状态: {new_status}'}), 400
    
    conn = get_db_connection()
    
    # Get order info
    order = conn.execute('SELECT id, number, source, status FROM orders WHERE id = ?', (order_id,)).fetchone()
    if not order:
        conn.close()
        return jsonify({'success': False, 'error': '订单不存在'}), 404
    
    old_status = order['status']
    
    # Check if status is actually changing
    if old_status == new_status:
        conn.close()
        return jsonify({'success': True, 'message': '状态未变化'})
    
    # Get site credentials
    site = conn.execute('SELECT * FROM sites WHERE url = ?', (order['source'],)).fetchone()
    if not site:
        conn.close()
        return jsonify({'success': False, 'error': '站点配置不存在'}), 404
    
    # Check if site has write permission
    api_write_status = site['api_write_status'] if 'api_write_status' in site.keys() else None
    if api_write_status == 'error':
        conn.close()
        return jsonify({'success': False, 'error': '该站点没有API写入权限，无法修改订单状态'}), 403
    
    # Headers for WooCommerce API
    api_headers = {
        "User-Agent": "WooCommerce API Client-Python/3.0.0",
        "Content-Type": "application/json",
        "Accept": "application/json"
    }
    
    # Status labels for note
    status_labels = {
        'pending': '待付款',
        'processing': '处理中',
        'on-hold': '已发货',
        'completed': '已完成',
        'cancelled': '已取消',
        'refunded': '已退款',
        'failed': '失败',
        'checkout-draft': '草稿'
    }
    
    try:
        # 1. Update order status via WooCommerce API.
        # Use orders.id (WC internal post ID), not order['number'], because sites
        # using Sequential Order Numbers have number != id and the REST API only
        # accepts the post ID.
        status_url = f"{site['url']}/wp-json/wc/v3/orders/{woo_post_id(order['id'])}"

        print(f"[DEBUG] Update Status - URL: {status_url}")
        print(f"[DEBUG] Update Status - Order ID: {order['id']}, Number: {order['number']}, New Status: {new_status}")

        status_resp = req.put(
            status_url,
            json={'status': new_status},
            auth=(site['consumer_key'], site['consumer_secret']),
            timeout=60,
            headers=api_headers
        )

        print(f"[DEBUG] Update Status - Response Code: {status_resp.status_code}")

        # Check if response is HTML instead of JSON
        response_text = status_resp.text or ''
        if response_text.strip().startswith('<!') or response_text.strip().startswith('<html'):
            print(f"[DEBUG] API returned HTML: {response_text[:200]}")
            raise Exception(f"WordPress返回HTML而非JSON，可能是WAF阻止或认证问题")

        if status_resp.status_code not in [200, 201]:
            raise Exception(f"远程API错误: {status_resp.status_code} - {response_text[:200]}")

        # 2. Add order note documenting the change
        note_url = f"{site['url']}/wp-json/wc/v3/orders/{woo_post_id(order['id'])}/notes"
        note_content = f"订单状态由 {current_user.name} 从 {status_labels.get(old_status, old_status)} 手动修改为 {status_labels.get(new_status, new_status)}"
        
        try:
            note_resp = req.post(
                note_url,
                json={'note': note_content, 'customer_note': False},
                auth=(site['consumer_key'], site['consumer_secret']),
                timeout=30,
                headers=api_headers
            )
            # Note failure is not critical, just log it
            if note_resp.status_code not in [200, 201]:
                print(f"Warning: Failed to add status change note: {note_resp.status_code}")
        except Exception as note_e:
            print(f"Warning: Failed to add status change note: {note_e}")
        
        # 3. Update local database
        conn.execute("UPDATE orders SET status = ? WHERE id = ?", (new_status, order_id))
        conn.commit()
        conn.close()
        
        return jsonify({
            'success': True, 
            'message': f'订单状态已从 {status_labels.get(old_status, old_status)} 修改为 {status_labels.get(new_status, new_status)}',
            'old_status': old_status,
            'new_status': new_status
        })
        
    except Exception as e:
        conn.close()
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/order/<order_id>/mark-undelivered', methods=['POST'])
@login_required
@shipper_required
@order_site_editable
def mark_order_undelivered(order_id):
    """Mark an order as undelivered (refused / returned to sender).
    Stores the lost shipping fee and an audit trail. Adds a remote WooCommerce
    note when possible so the WC admin reflects the same fact."""
    import requests as req

    data = request.get_json(silent=True) or {}
    raw_amount = data.get('shipping_loss_amount', 0)
    note_text = (data.get('note') or '').strip()

    try:
        loss_amount = float(raw_amount or 0)
    except (TypeError, ValueError):
        return jsonify({'success': False, 'error': '运费损失金额格式不正确'}), 400
    if loss_amount < 0:
        return jsonify({'success': False, 'error': '运费损失金额不能为负数'}), 400

    conn = get_db_connection()
    order = conn.execute(
        'SELECT id, number, source, is_undelivered FROM orders WHERE id = ?',
        (order_id,)
    ).fetchone()
    if not order:
        conn.close()
        return jsonify({'success': False, 'error': '订单不存在'}), 404

    if order['is_undelivered']:
        conn.close()
        return jsonify({'success': False, 'error': '该订单已被标记为未送达'}), 409

    try:
        conn.execute('''
            UPDATE orders
               SET is_undelivered = 1,
                   shipping_loss_amount = ?,
                   undelivered_at = datetime('now'),
                   undelivered_by = ?,
                   undelivered_note = ?
             WHERE id = ?
        ''', (loss_amount, int(current_user.id), note_text or None, order_id))

        # Local order_notes audit row (always succeeds even if remote API is down)
        log_line = f"订单被 {current_user.name} 标记为「未送达/退回」，运费损失 {loss_amount:.2f}"
        if note_text:
            log_line += f"。原因：{note_text}"
        conn.execute('''
            INSERT INTO order_notes (order_id, note, date_created, customer_note, author, added_by_user)
            VALUES (?, ?, datetime('now'), 0, ?, 1)
        ''', (order_id, log_line, current_user.name))

        conn.commit()
    except Exception as e:
        conn.close()
        return jsonify({'success': False, 'error': f'本地写入失败: {e}'}), 500

    # Best-effort remote note — failure is non-fatal
    site = conn.execute('SELECT * FROM sites WHERE url = ?', (order['source'],)).fetchone()
    conn.close()
    if site and site['consumer_key'] and site['consumer_secret']:
        try:
            # Use orders.id (WC post ID) — sites with Sequential Order Numbers
            # have number != id and the REST API only accepts the post ID.
            req.post(
                f"{site['url']}/wp-json/wc/v3/orders/{woo_post_id(order['id'])}/notes",
                json={'note': log_line, 'customer_note': False},
                auth=(site['consumer_key'], site['consumer_secret']),
                timeout=10,
                headers={
                    'User-Agent': 'WooCommerce API Client-Python/3.0.0',
                    'Content-Type': 'application/json',
                },
            )
        except Exception as remote_err:
            app.logger.warning(f'Remote note for undelivered order {order_id} failed: {remote_err}')

    return jsonify({
        'success': True,
        'message': f'已标记为未送达，运费损失 {loss_amount:.2f}',
        'shipping_loss_amount': loss_amount,
    })


@app.route('/api/order/<order_id>/confirm-delivery', methods=['POST'])
@login_required
@shipper_required
@order_site_editable
def confirm_order_delivery(order_id):
    """Approve a shipped order as delivered & accepted (COD collected).

    This is the human approval step in the 待确认结局 workflow. It:
      1. sets the local delivery_confirmed flag (clears it from the queue;
         a local-only column, so it survives every sync), then
      2. pushes the WooCommerce status to 'completed' so the store reflects
         reality. That fires WooCommerce's 'completed' customer email — the
         operator accepted this trade-off. WC sync is best-effort: if it fails
         the local confirm still stands and the order still leaves the queue."""
    import requests as req
    conn = get_db_connection()
    order = conn.execute(
        'SELECT id, number, source, status, is_undelivered, is_problem_return, delivery_confirmed FROM orders WHERE id = ?',
        (order_id,)
    ).fetchone()
    if not order:
        conn.close()
        return jsonify({'success': False, 'error': '订单不存在'}), 404

    if order['is_undelivered'] or order['is_problem_return']:
        conn.close()
        return jsonify({'success': False, 'error': '该订单已标记为未送达/问题退货，不能再标记签收'}), 409

    # 1. Local confirm — source of truth for the queue; always first.
    if not order['delivery_confirmed']:
        try:
            conn.execute('''UPDATE orders SET delivery_confirmed = 1,
                            delivery_confirmed_at = datetime('now'), delivery_confirmed_by = ?
                            WHERE id = ?''', (int(current_user.id), order_id))
            conn.execute('''INSERT INTO order_notes (order_id, note, date_created, customer_note, author, added_by_user)
                            VALUES (?, ?, datetime('now'), 0, ?, 1)''',
                         (order_id, f"订单被 {current_user.name} 确认「已签收」", current_user.name))
            conn.commit()
        except Exception as e:
            conn.close()
            return jsonify({'success': False, 'error': f'本地写入失败: {e}'}), 500

    # 2. Sync WooCommerce -> completed (best-effort; fires WC completed email).
    site = conn.execute('SELECT url, consumer_key, consumer_secret, api_write_status FROM sites WHERE url = ?',
                        (order['source'],)).fetchone()
    sync_msg = ''
    if order['status'] == 'completed':
        sync_msg = '；站点已是已完成'
    elif site and site['consumer_key'] and site['consumer_secret'] and (
            'api_write_status' not in site.keys() or site['api_write_status'] != 'error'):
        try:
            r = req.put(f"{site['url']}/wp-json/wc/v3/orders/{woo_post_id(order['id'])}",
                        json={'status': 'completed'},
                        auth=(site['consumer_key'], site['consumer_secret']), timeout=60,
                        headers={'User-Agent': 'WooCommerce API Client-Python/3.0.0', 'Content-Type': 'application/json'})
            if r.status_code in (200, 201):
                conn.execute("UPDATE orders SET status='completed' WHERE id=?", (order_id,))
                conn.commit()
                sync_msg = '；已同步站点为「已完成」'
            else:
                sync_msg = f'；站点同步失败({r.status_code})，本地已标记签收'
                app.logger.warning(f"confirm-delivery WC sync {order_id} -> {r.status_code}: {(r.text or '')[:200]}")
        except Exception as e:
            sync_msg = '；站点同步异常，本地已标记签收'
            app.logger.warning(f"confirm-delivery WC sync {order_id} error: {e}")
    else:
        sync_msg = '；该站点无写权限，仅本地标记签收'

    conn.close()
    return jsonify({'success': True, 'message': '已确认签收' + sync_msg})


def _persist_carrier_status(order_id, outcome):
    """Cache a freshly-detected carrier outcome onto the order — same write the
    nightly auto-resolver does — so an on-demand 查物流 immediately updates the
    待确认结局 badge and makes a delivered parcel batch-confirmable, without
    waiting for cron. Detection only: never confirms delivery or touches
    WooCommerce. No-ops for non-terminal / unknown outcomes."""
    if outcome not in ('delivered', 'returned', 'attention', 'in_transit'):
        return
    try:
        c = get_db_connection()
        c.execute("UPDATE orders SET carrier_status=?, carrier_status_at=datetime('now') WHERE id=?",
                  (outcome, order_id))
        c.commit()
        c.close()
    except Exception as e:
        app.logger.warning(f"persist carrier_status for {order_id} failed: {e}")


@app.route('/api/order/<order_id>/carrier-status')
@login_required
@shipping_view_required
def order_carrier_status(order_id):
    """On-demand 查物流: live carrier lookup for ONE order. Read-only w.r.t.
    delivery confirmation / WooCommerce, but it DOES cache the detected
    carrier_status (like the resolver) so the queue badge updates immediately.
    InPost = ShipX (+ Track718 'in-post' fallback). DPD/others = Track718."""
    import requests as req
    import carrier_tracking as ct
    conn = get_db_connection()
    order = conn.execute('SELECT id, number, meta_data, line_items, shipping_lines FROM orders WHERE id=?',
                         (order_id,)).fetchone()
    if not order:
        conn.close()
        return jsonify({'success': False, 'error': '订单不存在'}), 404
    krow = conn.execute("SELECT value FROM settings WHERE key='track718_api_key'").fetchone()
    key718 = krow['value'] if krow else None
    conn.close()

    # Extract tracking number + provider (AST -> VillaTheme -> _tracking_number)
    number, provider = '', ''
    md = parse_json_field(order['meta_data']) or []
    li = parse_json_field(order['line_items']) or []
    for m in md:
        if isinstance(m, dict) and m.get('key') == '_wc_shipment_tracking_items':
            v = m.get('value') or []
            if v and isinstance(v[0], dict) and v[0].get('tracking_number'):
                number = str(v[0]['tracking_number']).strip()
                provider = str(v[0].get('tracking_provider', ''))
                break
    if not number:
        for it in li:
            for m in (it.get('meta_data', []) if isinstance(it, dict) else []):
                if isinstance(m, dict) and m.get('key') == '_vi_wot_order_item_tracking_data':
                    try:
                        td = m.get('value')
                        td = json.loads(td) if isinstance(td, str) else td
                        if td and td[0].get('tracking_number'):
                            number = str(td[0]['tracking_number']).strip()
                            provider = str(td[0].get('carrier_slug') or td[0].get('carrier_name') or '')
                    except Exception:
                        pass
                if number:
                    break
            if number:
                break
    if not number:
        for m in md:
            if isinstance(m, dict) and m.get('key') == '_tracking_provider':
                provider = provider or str(m.get('value', ''))
            if isinstance(m, dict) and m.get('key') == '_tracking_number' and str(m.get('value', '')).strip():
                number = str(m['value']).strip()
    if not number:
        return jsonify({'success': False, 'error': '该订单没有运单号'}), 404

    carrier = ct.classify_carrier(provider, number)
    if carrier == 'inpost':
        try:
            r = req.get(f"https://api-shipx-pl.easypack24.net/v1/tracking/{number}",
                        headers={'User-Agent': 'woo-analysis-tracker/1.0'}, timeout=12)
        except Exception as e:
            return jsonify({'success': False, 'error': f'InPost 查询失败: {e}'}), 502
        if r.status_code == 404:
            # ShipX (InPost business API) has no record — common for parcels sent
            # via Paczkomat self-service. Fall back to Track718 under the InPost-PL
            # courier code ('in-post'), which tracks the consumer-side data the
            # official inpost.pl site shows.
            if key718:
                # Longer poll: a fresh number's crawl can take ~10-20s; break early
                # once a status lands (see track718_detail) so most clicks resolve.
                res = ct.track718_detail(number, key718, code=ct.TRACK718_INPOST_PL, poll=8, poll_wait=3)
                _persist_carrier_status(order_id, res.get('outcome'))
                if res.get('events'):
                    return jsonify({'success': True, 'carrier': 'InPost', 'tracking_number': number,
                                    'outcome': res.get('outcome', 'unknown'), 'events': res['events'],
                                    'note': '来自 Track718（InPost 商家接口无此单）'})
                if res.get('outcome') and res.get('outcome') != 'unknown':
                    return jsonify({'success': True, 'carrier': 'InPost', 'tracking_number': number,
                                    'outcome': res.get('outcome'), 'events': [],
                                    'note': 'Track718 已识别状态，详细轨迹抓取中，请稍后再查'})
                return jsonify({'success': True, 'carrier': 'InPost', 'tracking_number': number,
                                'outcome': 'unknown', 'events': [],
                                'note': 'InPost 商家接口查无；Track718 正在抓取，请稍后再查或点「官网」核对'})
            return jsonify({'success': True, 'carrier': 'InPost', 'tracking_number': number,
                            'outcome': 'unknown', 'events': [], 'note': 'InPost 查无此单（可能太老已清除）'})
        if r.status_code != 200:
            return jsonify({'success': False, 'error': f'InPost API {r.status_code}'}), 502
        d = r.json()
        raw = d.get('status', '')
        outcome = ct.INPOST_STATUS_MAP.get(raw, 'in_transit')
        _persist_carrier_status(order_id, outcome)
        events = sorted([{'time': ev.get('datetime', ''), 'status': ev.get('status', '')}
                         for ev in (d.get('tracking_details') or [])],
                        key=lambda e: e['time'], reverse=True)
        return jsonify({'success': True, 'carrier': 'InPost', 'tracking_number': number,
                        'raw': raw, 'outcome': outcome, 'events': events})
    # Everything that isn't InPost goes through Track718: DPD with its code,
    # any other carrier (EMS/中国邮政, Australia Post, GLS, …) via auto-detect.
    if not key718:
        return jsonify({'success': False, 'error': '未配置 Track718 key'}), 400
    res = ct.track718_detail(number, key718, code=('dpd-pl' if carrier == 'dpd' else None), poll=8, poll_wait=3)
    _persist_carrier_status(order_id, res.get('outcome'))
    name_map = {'dpd-pl': 'DPD', 'china-post': '中国邮政/EMS', 'australia-post': 'Australia Post',
                'inpost-paczkomaty': 'InPost', 'gls': 'GLS', 'poczta-polska': 'Poczta Polska'}
    cname = 'DPD' if carrier == 'dpd' else name_map.get((res.get('carrier') or '').lower(),
                                                         (res.get('carrier') or '物流').upper())
    if res.get('events'):
        return jsonify({'success': True, 'carrier': cname, 'tracking_number': number,
                        'outcome': res.get('outcome', 'unknown'), 'events': res['events']})
    return jsonify({'success': True, 'carrier': cname, 'tracking_number': number,
                    'outcome': res.get('outcome', 'unknown'), 'events': [],
                    'note': '物流商正在抓取轨迹或暂无记录（新登记的单需等几分钟），请稍后再查'})


@app.route('/api/order/<order_id>/unmark-undelivered', methods=['POST'])
@login_required
@shipper_required
@order_site_editable
def unmark_order_undelivered(order_id):
    """Reverse a previous mark-undelivered (e.g., the shipper made a mistake)."""
    conn = get_db_connection()
    order = conn.execute(
        'SELECT id, is_undelivered FROM orders WHERE id = ?',
        (order_id,)
    ).fetchone()
    if not order:
        conn.close()
        return jsonify({'success': False, 'error': '订单不存在'}), 404
    if not order['is_undelivered']:
        conn.close()
        return jsonify({'success': False, 'error': '该订单未被标记为未送达'}), 409

    try:
        conn.execute('''
            UPDATE orders
               SET is_undelivered = 0,
                   shipping_loss_amount = 0,
                   undelivered_at = NULL,
                   undelivered_by = NULL,
                   undelivered_note = NULL
             WHERE id = ?
        ''', (order_id,))
        conn.execute('''
            INSERT INTO order_notes (order_id, note, date_created, customer_note, author, added_by_user)
            VALUES (?, ?, datetime('now'), 0, ?, 1)
        ''', (order_id, f"{current_user.name} 撤销了「未送达」标记", current_user.name))
        conn.commit()
    except Exception as e:
        conn.close()
        return jsonify({'success': False, 'error': f'撤销失败: {e}'}), 500
    conn.close()
    return jsonify({'success': True, 'message': '已撤销未送达标记'})


PROBLEM_RETURN_TYPES = {
    'swap': '调包 / 假货',
    'short': '少件 / 空盒',
    'damaged': '实物损坏',
    'other': '其他',
}


@app.route('/api/order/<order_id>/mark-problem-return', methods=['POST'])
@login_required
@shipper_required
@order_site_editable
def mark_order_problem_return(order_id):
    """Mark an order as a problem return — the package came back but the
    contents were wrong / missing / damaged (e.g. the brick-swap scam).
    Records the lost product value and an audit trail. Distinct from
    mark-undelivered, which only covers a lost shipping fee."""
    import requests as req

    data = request.get_json(silent=True) or {}
    return_type = (data.get('type') or '').strip()
    raw_amount = data.get('product_loss_amount', 0)
    raw_shipping_loss = data.get('shipping_loss_amount', 0)
    note_text = (data.get('note') or '').strip()
    evidence_text = (data.get('evidence') or '').strip()

    if return_type not in PROBLEM_RETURN_TYPES:
        return jsonify({'success': False, 'error': '问题类型不正确'}), 400
    try:
        loss_amount = float(raw_amount or 0)
    except (TypeError, ValueError):
        return jsonify({'success': False, 'error': '货值损失金额格式不正确'}), 400
    if loss_amount < 0:
        return jsonify({'success': False, 'error': '货值损失金额不能为负数'}), 400
    try:
        shipping_loss = float(raw_shipping_loss or 0)
    except (TypeError, ValueError):
        return jsonify({'success': False, 'error': '运费损失金额格式不正确'}), 400
    if shipping_loss < 0:
        return jsonify({'success': False, 'error': '运费损失金额不能为负数'}), 400

    conn = get_db_connection()
    order = conn.execute(
        'SELECT id, number, source, is_problem_return FROM orders WHERE id = ?',
        (order_id,)
    ).fetchone()
    if not order:
        conn.close()
        return jsonify({'success': False, 'error': '订单不存在'}), 404
    if order['is_problem_return']:
        conn.close()
        return jsonify({'success': False, 'error': '该订单已被标记为问题退货'}), 409

    type_label = PROBLEM_RETURN_TYPES[return_type]
    try:
        conn.execute('''
            UPDATE orders
               SET is_problem_return = 1,
                   problem_return_type = ?,
                   product_loss_amount = ?,
                   shipping_loss_amount = ?,
                   problem_return_at = datetime('now'),
                   problem_return_by = ?,
                   problem_return_note = ?,
                   problem_return_evidence = ?
             WHERE id = ?
        ''', (return_type, loss_amount, shipping_loss, int(current_user.id),
              note_text or None, evidence_text or None, order_id))

        log_line = (f"订单被 {current_user.name} 标记为「问题退货 · {type_label}」，"
                    f"货值损失 {loss_amount:.2f}，运费损失 {shipping_loss:.2f}")
        if evidence_text:
            log_line += f"。证据：{evidence_text}"
        if note_text:
            log_line += f"。说明：{note_text}"
        conn.execute('''
            INSERT INTO order_notes (order_id, note, date_created, customer_note, author, added_by_user)
            VALUES (?, ?, datetime('now'), 0, ?, 1)
        ''', (order_id, log_line, current_user.name))

        conn.commit()
    except Exception as e:
        conn.close()
        return jsonify({'success': False, 'error': f'本地写入失败: {e}'}), 500

    # Best-effort remote note — failure is non-fatal
    site = conn.execute('SELECT * FROM sites WHERE url = ?', (order['source'],)).fetchone()
    conn.close()
    if site and site['consumer_key'] and site['consumer_secret']:
        try:
            req.post(
                f"{site['url']}/wp-json/wc/v3/orders/{woo_post_id(order['id'])}/notes",
                json={'note': log_line, 'customer_note': False},
                auth=(site['consumer_key'], site['consumer_secret']),
                timeout=10,
                headers={
                    'User-Agent': 'WooCommerce API Client-Python/3.0.0',
                    'Content-Type': 'application/json',
                },
            )
        except Exception as remote_err:
            app.logger.warning(f'Remote note for problem-return order {order_id} failed: {remote_err}')

    return jsonify({
        'success': True,
        'message': f'已标记为问题退货（{type_label}），货值损失 {loss_amount:.2f}，运费损失 {shipping_loss:.2f}',
        'product_loss_amount': loss_amount,
        'shipping_loss_amount': shipping_loss,
    })


@app.route('/api/order/<order_id>/unmark-problem-return', methods=['POST'])
@login_required
@shipper_required
@order_site_editable
def unmark_order_problem_return(order_id):
    """Reverse a previous mark-problem-return (e.g., marked by mistake)."""
    conn = get_db_connection()
    order = conn.execute(
        'SELECT id, is_problem_return, is_undelivered FROM orders WHERE id = ?',
        (order_id,)
    ).fetchone()
    if not order:
        conn.close()
        return jsonify({'success': False, 'error': '订单不存在'}), 404
    if not order['is_problem_return']:
        conn.close()
        return jsonify({'success': False, 'error': '该订单未被标记为问题退货'}), 409

    # Only clear shipping_loss_amount when no other marker still owns it
    # (the same column is shared with mark-undelivered).
    also_undelivered = bool(order['is_undelivered'])
    try:
        if also_undelivered:
            conn.execute('''
                UPDATE orders
                   SET is_problem_return = 0,
                       problem_return_type = NULL,
                       product_loss_amount = 0,
                       problem_return_at = NULL,
                       problem_return_by = NULL,
                       problem_return_note = NULL,
                       problem_return_evidence = NULL
                 WHERE id = ?
            ''', (order_id,))
        else:
            conn.execute('''
                UPDATE orders
                   SET is_problem_return = 0,
                       problem_return_type = NULL,
                       product_loss_amount = 0,
                       shipping_loss_amount = 0,
                       problem_return_at = NULL,
                       problem_return_by = NULL,
                       problem_return_note = NULL,
                       problem_return_evidence = NULL
                 WHERE id = ?
            ''', (order_id,))
        conn.execute('''
            INSERT INTO order_notes (order_id, note, date_created, customer_note, author, added_by_user)
            VALUES (?, ?, datetime('now'), 0, ?, 1)
        ''', (order_id, f"{current_user.name} 撤销了「问题退货」标记", current_user.name))
        conn.commit()
    except Exception as e:
        conn.close()
        return jsonify({'success': False, 'error': f'撤销失败: {e}'}), 500
    conn.close()
    return jsonify({'success': True, 'message': '已撤销问题退货标记'})


@app.route('/api/shipping/print/label/<order_id>')
@login_required
@shipping_view_required
def print_shipping_label(order_id):
    """Get order data for printing shipping label"""
    conn = get_db_connection()
    
    order = conn.execute('''
        SELECT o.*, sl.tracking_number, sl.carrier_slug, sc.name as carrier_name
        FROM orders o
        LEFT JOIN shipping_logs sl ON sl.id = (
            SELECT id FROM shipping_logs WHERE order_id = o.id ORDER BY id DESC LIMIT 1
        )
        LEFT JOIN shipping_carriers sc ON sl.carrier_slug = sc.slug
        WHERE o.id = ?
    ''', (order_id,)).fetchone()
    
    if not order:
        conn.close()
        return jsonify({'error': '订单不存在'}), 404
    
    # Get site manager
    site = conn.execute('SELECT manager FROM sites WHERE url = ?', (order['source'],)).fetchone()
    conn.close()
    
    billing = parse_json_field(order['billing'])
    shipping_info = parse_json_field(order['shipping'])
    line_items = parse_json_field(order['line_items'])
    shipping_lines = parse_json_field(order['shipping_lines'])
    shipping_method = shipping_lines[0].get('method_title', 'Unknown') if shipping_lines and len(shipping_lines) > 0 else ''
    
    addr = shipping_info if shipping_info and shipping_info.get('address_1') else billing
    
    return jsonify({
        'order_number': order['number'],
        'date': order['date_created'][:10] if order['date_created'] else '',
        'source': order['source'].replace('https://www.', '').replace('https://', ''),
        'manager': site['manager'] if site else '',
        'customer_name': f"{addr.get('first_name', '')} {addr.get('last_name', '')}".strip(),
        'customer_phone': addr.get('phone') or billing.get('phone', ''),
        'customer_email': billing.get('email', ''),
        'customer_address': _compose_address(addr, sep='\n'),
        'customer_inpost_id': extract_custom_billing_fields(parse_json_field(order['meta_data'])).get('customer_inpost_id', ''),
        'customer_social': extract_custom_billing_fields(parse_json_field(order['meta_data'])).get('customer_social', ''),
        'customer_note': order['customer_note'] or '',
        'customer_address_2': addr.get('address_2', ''),
        'shipping_method': shipping_method,
        'products': [{'name': item.get('name', ''), 'qty': item.get('quantity', 1), 'total': float(item.get('total', 0))} for item in (line_items or [])],
        'currency': order['currency'],
        'total': f"{float(order['total'] or 0):.2f} {order['currency']}",
        'shipping_total': f"{float(order['shipping_total'] or 0):.2f}",
        'tracking_number': order['tracking_number'] or '',
        'carrier_name': order['carrier_name'] or ''
    })


@app.route('/api/shipping/print/list')
@login_required
@shipping_view_required
def print_shipping_list():
    """Get last 24 hours pending orders for printing - reuses get_pending_orders logic"""
    from datetime import datetime, timedelta
    
    # Call the existing pending orders function
    with app.test_request_context('/api/shipping/pending'):
        response = get_pending_orders()
        all_orders = response.get_json()
    
    # Filter for last 24 hours
    now = datetime.now()
    cutoff = now - timedelta(hours=24)
    
    result = []
    for order in all_orders:
        # Parse order date
        try:
            order_date = datetime.fromisoformat(order['date_created'].replace('Z', '+00:00'))
            if order_date >= cutoff:
                method_lower = (order.get('shipping_method') or '').lower()
                is_inpost = 'inpost' in method_lower or 'paczkomat' in method_lower
                result.append({
                    'order_number': order['number'],
                    'source': order['source'],
                    'manager': order.get('manager', ''),
                    'customer_name': order['customer_name'],
                    'customer_phone': order.get('customer_phone', ''),
                    'customer_email': order.get('customer_email', ''),
                    'customer_address': order['customer_address'],
                    'customer_address_2': order.get('customer_address_2', ''),
                    'customer_inpost_id': order.get('customer_inpost_id') or ('送货上门: ' + order['customer_address'] if is_inpost else ''),
                    'customer_social': order.get('customer_social', ''),
                    'customer_note': order.get('customer_note', ''),
                    'shipping_method': order.get('shipping_method', ''),
                    'total': f"{order['total']:.2f} {order['currency']}",
                    'products': [{'name': p['name'], 'qty': p['quantity']} for p in order['products']],
                    'shipping_total': order.get('shipping_total', 0)
                })
        except:
            # If date parsing fails, skip this order
            continue

    today = datetime.now().strftime('%Y-%m-%d')
    return jsonify({'date': today, 'orders': result, 'count': len(result)})


@app.route('/api/shipping/print/pending')
@login_required
@shipping_view_required
def print_pending_list():
    """Get all pending orders for printing - reuses get_pending_orders logic"""
    from datetime import datetime
    
    # Temporarily store the request context
    original_args = dict(request.args)
    
    # Clear filters to get all pending orders
    with app.test_request_context('/api/shipping/pending'):
        # Call the existing pending orders function
        response = get_pending_orders()
        orders_data = response.get_json()
    
    # Format for print
    result = []
    for order in orders_data:
        method_lower = (order.get('shipping_method') or '').lower()
        is_inpost = 'inpost' in method_lower or 'paczkomat' in method_lower
        result.append({
            'order_number': order['number'],
            'source': order['source'],
            'manager': order.get('manager', ''),
            'customer_name': order['customer_name'],
            'customer_phone': order.get('customer_phone', ''),
            'customer_email': order.get('customer_email', ''),
            'customer_address': order['customer_address'],
            'customer_address_2': order.get('customer_address_2', ''),
            'customer_inpost_id': order.get('customer_inpost_id') or ('送货上门: ' + order['customer_address'] if is_inpost else ''),
            'customer_social': order.get('customer_social', ''),
            'customer_note': order.get('customer_note', ''),
            'shipping_method': order.get('shipping_method', ''),
            'total': f"{order['total']:.2f} {order['currency']}",
            'products': [{'name': p['name'], 'qty': p['quantity']} for p in order['products']],
            'shipping_total': order.get('shipping_total', 0)
        })

    today = datetime.now().strftime('%Y-%m-%d')
    return jsonify({'date': '截止 ' + today, 'orders': result, 'count': len(result)})


@app.route('/api/shipping/export/list')
@login_required
@shipping_view_required
def export_shipping_list():
    """Export last 24 hours pending orders to CSV"""
    import csv
    import io
    from flask import make_response
    from datetime import datetime, timedelta
    
    with app.test_request_context('/api/shipping/pending'):
        response = get_pending_orders()
        all_orders = response.get_json()
    
    now = datetime.now()
    cutoff = now - timedelta(hours=24)
    
    si = io.StringIO()
    cw = csv.writer(si)
    si.write('\ufeff')
    cw.writerow(['序号', '订单信息', '来源', '客户姓名', '电话', '邮箱', '配送方式', 'InPost ID/取货点', '收货地址', '客户留言', '商品明细', '运费', '总金额'])

    count = 1
    for order in all_orders:
        try:
            order_date = datetime.fromisoformat(order['date_created'].replace('Z', '+00:00'))
            if order_date >= cutoff:
                products_str = " | ".join([f"{p['name']} x{p['quantity']}" for p in order['products']])
                method = (order.get('shipping_method') or '').lower()
                isInPost = order.get('customer_inpost_id') or 'inpost' in method or 'paczkomat' in method
                isDPD = 'dpd' in method
                shipping_method = 'InPost' if isInPost else ('DPD' if isDPD else (order.get('shipping_method') or 'Standard'))

                inpost_details = order.get('customer_inpost_id', '')
                if not inpost_details and order.get('customer_address_2'):
                     inpost_details = order.get('customer_address_2')
                # Fallback to customer address if it's an InPost method but has no ID (typically courier/home delivery)
                if not inpost_details and isInPost:
                     inpost_details = "送货上门: " + order.get('customer_address', '')

                phone_social = " ".join(filter(None, [order.get('customer_phone', ''), order.get('customer_social', '')]))

                cw.writerow([
                    count,
                    f"#{order['number']}",
                    order['source'].replace('https://www.', '').replace('https://', ''),
                    order['customer_name'],
                    phone_social,
                    order.get('customer_email', ''),
                    shipping_method,
                    inpost_details,
                    order['customer_address'],
                    order.get('customer_note', ''),
                    products_str,
                    f"{order.get('shipping_total', 0)}",
                    f"{order['total']} {order['currency']}"
                ])
                count += 1
        except Exception:
             continue
             
    output = make_response(si.getvalue())
    output.headers["Content-Disposition"] = f"attachment; filename=shipping_today_{now.strftime('%Y%m%d')}.csv"
    output.headers["Content-type"] = "text/csv"
    return output


@app.route('/api/shipping/export/pending')
@login_required
@shipping_view_required
def export_pending_list():
    """Export all pending orders to CSV"""
    import csv
    import io
    from flask import make_response
    from datetime import datetime
    
    with app.test_request_context('/api/shipping/pending'):
        response = get_pending_orders()
        orders_data = response.get_json()
        
    si = io.StringIO()
    cw = csv.writer(si)
    si.write('\ufeff')
    cw.writerow(['序号', '订单信息', '来源', '客户姓名', '电话', '邮箱', '配送方式', 'InPost ID/取货点', '收货地址', '客户留言', '商品明细', '运费', '总金额'])

    count = 1
    for order in orders_data:
        try:
            products_str = " | ".join([f"{p['name']} x{p['quantity']}" for p in order['products']])
            method = (order.get('shipping_method') or '').lower()
            isInPost = order.get('customer_inpost_id') or 'inpost' in method or 'paczkomat' in method
            isDPD = 'dpd' in method
            shipping_method = 'InPost' if isInPost else ('DPD' if isDPD else (order.get('shipping_method') or 'Standard'))

            inpost_details = order.get('customer_inpost_id', '')
            if not inpost_details and order.get('customer_address_2'):
                 inpost_details = order.get('customer_address_2')
            # Fallback to customer address if it's an InPost method but has no ID
            if not inpost_details and isInPost:
                 inpost_details = "送货上门: " + order.get('customer_address', '')

            phone_social = " ".join(filter(None, [order.get('customer_phone', ''), order.get('customer_social', '')]))

            cw.writerow([
                count,
                f"#{order['number']}",
                order['source'].replace('https://www.', '').replace('https://', ''),
                order['customer_name'],
                phone_social,
                order.get('customer_email', ''),
                shipping_method,
                inpost_details,
                order['customer_address'],
                order.get('customer_note', ''),
                products_str,
                f"{order.get('shipping_total', 0)}",
                f"{order['total']} {order['currency']}"
            ])
            count += 1
        except Exception:
            continue
        
    now = datetime.now()
    output = make_response(si.getvalue())
    output.headers["Content-Disposition"] = f"attachment; filename=shipping_pending_{now.strftime('%Y%m%d')}.csv"
    output.headers["Content-type"] = "text/csv"
    return output


def process_shipped_order(order, conn, carriers, ast_provider_mapping):
    billing = parse_json_field(order['billing'])
    shipping_info = parse_json_field(order['shipping'])
    meta_data = parse_json_field(order['meta_data'])
    
    # Get shipping address for display
    addr = shipping_info if shipping_info and shipping_info.get('address_1') else billing
    custom_fields = extract_custom_billing_fields(meta_data)
    
    # Calculate customer address (Standard)
    customer_address = _compose_address(addr)

    # DPD Fallback: If standard address is empty but DPD fields exist
    if not addr.get('address_1') and (custom_fields.get('dpd_street') or custom_fields.get('dpd_city')):
        dpd_parts = [
            f"{custom_fields.get('dpd_street', '')} {custom_fields.get('dpd_house', '')}".strip(),
            custom_fields.get('dpd_zip', ''),
            custom_fields.get('dpd_city', '')
        ]
        customer_address = ', '.join(filter(None, dpd_parts))


    # Get tracking info - first from shipping_logs, then from order meta_data
    tracking_number = order['tracking_number'] or ''
    carrier_slug = order['carrier_slug'] or ''
    shipped_at = order['shipped_at']
    carrier_name = ''
    tracking_url = ''
    
    # If no tracking in shipping_logs, try to get from wc_shipment_tracking_items (Advanced Shipment Tracking Pro)
    if not tracking_number and meta_data:
        for meta in meta_data:
            if isinstance(meta, dict) and meta.get('key') == '_wc_shipment_tracking_items':
                tracking_items = meta.get('value', [])
                if isinstance(tracking_items, list) and len(tracking_items) > 0:
                    first_item = tracking_items[0]
                    if isinstance(first_item, dict):
                        tracking_number = first_item.get('tracking_number', '')
                        ast_provider = first_item.get('tracking_provider', '')
                        
                        # Map AST provider to our carrier slugs
                        if ast_provider in ast_provider_mapping:
                            carrier_slug, carrier_name = ast_provider_mapping[ast_provider]
                        else:
                            carrier_slug = ast_provider
                            carrier_name = ast_provider.replace('-', ' ').title()
                        
                        # Parse date_shipped (Unix timestamp)
                        date_shipped = first_item.get('date_shipped')
                        if date_shipped:
                            try:
                                from datetime import datetime
                                shipped_at = datetime.fromtimestamp(int(date_shipped)).strftime('%Y-%m-%d %H:%M:%S')
                            except:
                                pass
                break
    
    # If still no tracking, try line_items meta_data for Orders Tracking for WooCommerce (VillaTheme)
    if not tracking_number:
        line_items = parse_json_field(order['line_items'])
        if line_items:
            for item in line_items:
                if isinstance(item, dict):
                    item_meta = item.get('meta_data', [])
                    for meta in item_meta:
                        if isinstance(meta, dict) and meta.get('key') == '_vi_wot_order_item_tracking_data':
                            try:
                                tracking_data_str = meta.get('value', '')
                                if isinstance(tracking_data_str, str):
                                    import json
                                    tracking_data = json.loads(tracking_data_str)
                                else:
                                    tracking_data = tracking_data_str
                                
                                if isinstance(tracking_data, list) and len(tracking_data) > 0:
                                    first_track = tracking_data[0]
                                    if isinstance(first_track, dict):
                                        tracking_number = first_track.get('tracking_number', '')
                                        # Get carrier name and slug
                                        raw_carrier_name = first_track.get('carrier_name', '')
                                        raw_carrier_slug = first_track.get('carrier_slug', '')
                                        
                                        carrier_name = raw_carrier_name.title() if raw_carrier_name else raw_carrier_slug.title()
                                        carrier_slug = raw_carrier_slug
                                        
                                        # Normalize carrier slug for known carriers to ensure we use our verified DB URLs
                                        cn_lower = carrier_name.lower()
                                        plugin_tracking_url = first_track.get('carrier_url', '')

                                        if 'dpd' in cn_lower:
                                            carrier_slug = 'dpd'
                                            carrier_name = 'DPD'
                                            plugin_tracking_url = '' # Force use of DB configured URL
                                        elif 'inpost' in cn_lower:
                                            carrier_slug = 'inpost'
                                            carrier_name = 'InPost' 
                                            plugin_tracking_url = '' # Force use of DB configured URL

                                        # Get tracking URL from plugin data only if we didn't clear it
                                        if plugin_tracking_url and tracking_number:
                                            tracking_url = plugin_tracking_url.replace('{tracking_number}', tracking_number)
                                            # Fix common placeholder issues from plugins
                                            tracking_url = tracking_url.replace('{Tracking_number}', tracking_number)
                                        
                                        # Parse time (Unix timestamp)
                                        track_time = first_track.get('time')
                                        if track_time:
                                            try:
                                                from datetime import datetime
                                                shipped_at = datetime.fromtimestamp(int(track_time)).strftime('%Y-%m-%d %H:%M:%S')
                                            except:
                                                pass
                            except:
                                pass
                        elif isinstance(meta, dict) and meta.get('key') == 'tracking_number':
                            tracking_number = str(meta.get('value', '')).strip()
                        
                        if tracking_number: break
                if tracking_number: break

    # If still no tracking, try shipping_lines meta_data 'tracking_number'
    shipping_lines = parse_json_field(order['shipping_lines'])
    if not tracking_number and shipping_lines:
        for item in shipping_lines:
            if isinstance(item, dict):
                for meta in item.get('meta_data', []):
                    if isinstance(meta, dict) and meta.get('key') == 'tracking_number':
                        tracking_number = meta.get('value', '').strip()
                        if tracking_number:
                            break
            if tracking_number:
                break

    # If still no tracking, try order meta_data '_tracking_number'
    if not tracking_number and meta_data:
            for meta in meta_data:
                if isinstance(meta, dict) and meta.get('key') == '_tracking_number':
                    tracking_number = meta.get('value', '').strip()
                    if tracking_number:
                        break

    # Determine carrier if we found a tracking number via custom methods but no slug
    if tracking_number and not carrier_slug:
        carrier_name = 'Custom' # Default
        
        # Helper to check string format
        def is_alnum(s):
            return s.replace(' ', '').isalnum()

        # Logic from poland.php: InPost (24 digits), DPD (10-14 alphanumeric)
        if len(tracking_number) == 24 and tracking_number.isdigit():
            carrier_slug = 'inpost'
            carrier_name = 'InPost'
        elif 10 <= len(tracking_number) <= 14 and is_alnum(tracking_number):
            carrier_slug = 'dpd'
            carrier_name = 'DPD'
        
        # If still unsure, check shipping method title
        if not carrier_slug and shipping_lines and len(shipping_lines) > 0:
            method_title = str(shipping_lines[0].get('method_title', '')).lower()
            if 'inpost' in method_title:
                carrier_slug = 'inpost'
                carrier_name = 'InPost'
            elif 'dpd' in method_title:
                carrier_slug = 'dpd'
                carrier_name = 'DPD'

    # Build tracking URL (if not already set by plugin data)
    if tracking_number and carrier_slug and not tracking_url:
        carrier = carriers.get(carrier_slug)
        if carrier and carrier['tracking_url']:
            tracking_url = carrier['tracking_url'].replace('{tracking}', tracking_number)
    
    # Use date_modified as fallback if still no shipped_at
    if not shipped_at and order['date_modified']:
        shipped_at = order['date_modified']
    
    # Get carrier name from our database if not already set
    if not carrier_name and carrier_slug and carrier_slug in carriers:
        carrier_name = carriers[carrier_slug]['name']

    keys = order.keys() if hasattr(order, 'keys') else []
    is_undelivered = bool(order['is_undelivered']) if 'is_undelivered' in keys else False
    shipping_method = shipping_lines[0].get('method_title', '') if shipping_lines and len(shipping_lines) > 0 else ''

    return {
        'id': order['id'],
        'number': order['number'],
        'total': float(order['total'] or 0),
        'currency': order['currency'],
        'source': order['source'].replace('https://www.', '').replace('https://', ''),
        'manager': order['manager'] or '',
        'warehouse_id': order['warehouse_id'] if 'warehouse_id' in keys else None,
        'warehouse_name': (order['warehouse_name'] if 'warehouse_name' in keys else '') or '',
        'customer_name': f"{addr.get('first_name', '')} {addr.get('last_name', '')}".strip(),
        'customer_email': billing.get('email', ''),
        'customer_phone': addr.get('phone') or billing.get('phone', ''),
        'customer_address': customer_address,
        'customer_address_2': addr.get('address_2', ''),
        'customer_inpost_id': custom_fields['customer_inpost_id'],
        'customer_social': custom_fields['customer_social'],
        'customer_note': (order['customer_note'] if 'customer_note' in keys else '') or '',
        'shipping_method': shipping_method,
        'tracking_number': tracking_number,
        'carrier_slug': carrier_slug,
        'carrier_name': carrier_name,
        'tracking_url': tracking_url,
        'shipped_at': shipped_at,
        'has_tracking': bool(tracking_number),
        'is_undelivered': is_undelivered,
        'shipping_loss_amount': float(order['shipping_loss_amount'] or 0) if 'shipping_loss_amount' in keys else 0,
        'undelivered_at': order['undelivered_at'] if 'undelivered_at' in keys else None,
        'undelivered_note': order['undelivered_note'] if 'undelivered_note' in keys else None,
        'undelivered_by_name': order['undelivered_by_name'] if 'undelivered_by_name' in keys else None,
        'is_problem_return': bool(order['is_problem_return']) if 'is_problem_return' in keys else False,
        'products': [{'name': item.get('name', ''), 'quantity': item.get('quantity', 1), 'total': float(item.get('total', 0))} for item in (parse_json_field(order['line_items']) or [])],
        'shipping_total': float(order['shipping_total'] or 0),
        'product_count': sum(item.get('quantity', 1) for item in (parse_json_field(order['line_items']) or [])),
        'latest_note': (order['latest_note'] if 'latest_note' in keys else '') or '',
        'latest_note_date': (order['latest_note_date'] if 'latest_note_date' in keys else '') or '',
        'latest_note_author': (order['latest_note_author'] if 'latest_note_author' in keys else '') or ''
    }


@app.route('/api/export_orders')
@login_required
@admin_required
def export_orders():
    """Export all orders to CSV"""
    import csv
    import io
    from flask import make_response

    conn = get_db_connection()
    
    # Get user's allowed sources
    allowed_sources = get_user_allowed_sources(current_user.id, current_user.is_admin(), current_user.is_viewer())
    
    # Build query
    query = 'SELECT o.*, s.manager FROM orders o LEFT JOIN sites s ON o.source = s.url'
    conditions = []
    params = []
    
    if allowed_sources is not None:
        if allowed_sources:
            placeholders = ','.join(['?' for _ in allowed_sources])
            conditions.append(f'o.source IN ({placeholders})')
            params.extend(allowed_sources)
        else:
            conditions.append('1=0')
            
    if conditions:
        query += ' WHERE ' + ' AND '.join(conditions)
        
    query += ' ORDER BY o.date_created DESC'
    
    orders = conn.execute(query, params).fetchall()
    conn.close()
    
    # Prepare CSV
    si = io.StringIO()
    cw = csv.writer(si)
    
    # Write BOM for Excel support with UTF-8
    si.write('\ufeff')
    
    # Header
    cw.writerow(['订单号', '创建日期', '状态', '来源', '负责人', '客户姓名', '客户邮箱', '产品明细', '总数量', '运费', '订单金额', '净额', '汇率', '货币', '¥净额'])
    
    for order in orders:
        # Calculate CNY amount
        currency = order['currency']
        total = float(order['total'] or 0)
        shipping = float(order['shipping_total'] or 0)
        net_total = total - shipping
        
        month = order['date_created'][:7] if order['date_created'] else None
        rate, _ = get_cny_rate(currency, month)
        net_total_cny = round(net_total * rate, 2) if rate else 0
        
        status_text = STATUS_LABELS.get(order['status'], order['status'])
        source = order['source'].replace('https://www.', '').replace('https://', '')
        
        billing = parse_json_field(order['billing'])
        customer_name = f"{billing.get('first_name', '')} {billing.get('last_name', '')}".strip()
        
        # Process product details
        line_items = parse_json_field(order['line_items'])
        product_details = []
        total_quantity = 0
        if isinstance(line_items, list):
            for item in line_items:
                qty = item.get('quantity', 0)
                name = item.get('name', 'Unknown')
                # Extract simplified name if possible or use full name ?? 
                # For now use the name from line_item
                product_details.append(f"{name} x{qty}")
                total_quantity += qty
        
        products_str = " | ".join(product_details)
        
        cw.writerow([
            f"#{order['number']}",
            order['date_created'],
            status_text,
            source,
            order['manager'] or '',
            customer_name,
            billing.get('email', ''),
            products_str,
            total_quantity,
            f"{shipping:.2f}",
            f"{total:.2f}",
            f"{net_total:.2f}",
            rate or '',
            currency,
            f"{net_total_cny:.2f}"
        ])
        
    output = make_response(si.getvalue())
    output.headers["Content-Disposition"] = f"attachment; filename=orders_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
    output.headers["Content-type"] = "text/csv"
    return output




# ============== 51.LA API INTEGRATION ==============

# 51.la API Configuration
LA_API_CONFIG = {
    'access_key': 'j4dwJjSlJ30MUnhEdfRDqVDcbDuB6UTz',
    'secret_key': 'F3r4EIJpqhCl9hYZ5MI6m667WdgRX14c',
    'api_url': 'https://v6-open.51.la/open/trend/day',
    'sites': {
        'strefajednorazowek.pl': 'KqaoIzZaCgNcnxBY',
        'buchmistrz.pl': 'KqmtdhxdOsN9a7i7',
        'vapepolska.pl': 'L2T2pX4NmvjHYcyO',
        'vapeprime.pl': 'L1loh4Allytc0m4m',
        'vapico.pl': 'L4Nl5AtWBgRcaVlT',
    }
}


def fetch_51la_traffic(mask_id, start_day, end_day):
    """
    Fetch traffic data from 51.la API
    Low security mode: sign = accessKey
    """
    import requests
    import time
    import random
    import string
    
    timestamp = str(int(time.time() * 1000))
    nonce = ''.join(random.choices(string.ascii_letters + string.digits, k=4))
    
    payload = {
        'maskId': mask_id,
        'startDay': start_day,
        'endDay': end_day,
        'accessKey': LA_API_CONFIG['access_key'],
        'nonce': nonce,
        'timestamp': timestamp,
        'sign': LA_API_CONFIG['access_key']  # Low security mode
    }
    
    headers = {'Content-Type': 'application/json'}
    
    try:
        response = requests.post(LA_API_CONFIG['api_url'], json=payload, headers=headers, timeout=5)
        result = response.json()
        
        if result.get('success') and result.get('code') == '0000':
            return result.get('data', [])
        else:
            app.logger.error(f"51.la API error: {result.get('message')}")
            return []
    except Exception as e:
        app.logger.error(f"51.la API request failed: {e}")
        return []


def get_all_traffic_data(start_day, end_day):
    """Fetch traffic data for all configured sites"""
    all_traffic = {}
    
    for site_name, mask_id in LA_API_CONFIG['sites'].items():
        traffic = fetch_51la_traffic(mask_id, start_day, end_day)
        all_traffic[site_name] = traffic
    
    return all_traffic


def aggregate_traffic_by_month(traffic_data):
    """Aggregate daily traffic data by month"""
    from collections import defaultdict
    
    monthly = defaultdict(lambda: {
        'uv': 0, 'pv': 0, 'newUserCount': 0, 'sv': 0, 'ip': 0,
        'bounceRate_sum': 0, 'bounceRate_count': 0
    })
    
    for day_data in traffic_data:
        time_str = day_data.get('time', '')
        if len(time_str) >= 7:
            month = time_str[:7]  # YYYY-MM
            monthly[month]['uv'] += day_data.get('uv', 0)
            monthly[month]['pv'] += day_data.get('pv', 0)
            monthly[month]['newUserCount'] += day_data.get('newUserCount', 0)
            monthly[month]['sv'] += day_data.get('sv', 0)
            monthly[month]['ip'] += day_data.get('ip', 0)
            monthly[month]['bounceRate_sum'] += day_data.get('bounceRate', 0)
            monthly[month]['bounceRate_count'] += 1
    
    # Calculate average bounce rate
    result = {}
    for month, data in monthly.items():
        result[month] = {
            'uv': data['uv'],
            'pv': data['pv'],
            'newUserCount': data['newUserCount'],
            'sv': data['sv'],
            'ip': data['ip'],
            'bounceRate': round(data['bounceRate_sum'] / data['bounceRate_count'], 4) if data['bounceRate_count'] > 0 else 0
        }
    
    return result


def aggregate_traffic_by_day(traffic_data):
    """Keep daily traffic data (no aggregation needed, just restructure)"""
    result = {}
    for day_data in traffic_data:
        time_str = day_data.get('time', '')
        if len(time_str) >= 10:
            day = time_str[:10]  # YYYY-MM-DD
            result[day] = {
                'uv': day_data.get('uv', 0),
                'pv': day_data.get('pv', 0),
                'newUserCount': day_data.get('newUserCount', 0),
                'sv': day_data.get('sv', 0),
                'ip': day_data.get('ip', 0),
                'bounceRate': day_data.get('bounceRate', 0)
            }
    return result


def aggregate_traffic_by_week(traffic_data):
    """Aggregate daily traffic data by week (ISO week number)"""
    from collections import defaultdict
    from datetime import datetime
    
    weekly = defaultdict(lambda: {
        'uv': 0, 'pv': 0, 'newUserCount': 0, 'sv': 0, 'ip': 0,
        'bounceRate_sum': 0, 'bounceRate_count': 0
    })
    
    for day_data in traffic_data:
        time_str = day_data.get('time', '')
        if len(time_str) >= 10:
            try:
                dt = datetime.strptime(time_str[:10], '%Y-%m-%d')
                # ISO week format: YYYY-Www
                week_key = f"{dt.isocalendar()[0]}-W{dt.isocalendar()[1]:02d}"
                
                weekly[week_key]['uv'] += day_data.get('uv', 0)
                weekly[week_key]['pv'] += day_data.get('pv', 0)
                weekly[week_key]['newUserCount'] += day_data.get('newUserCount', 0)
                weekly[week_key]['sv'] += day_data.get('sv', 0)
                weekly[week_key]['ip'] += day_data.get('ip', 0)
                weekly[week_key]['bounceRate_sum'] += day_data.get('bounceRate', 0)
                weekly[week_key]['bounceRate_count'] += 1
            except:
                continue
    
    # Calculate average bounce rate
    result = {}
    for week, data in weekly.items():
        result[week] = {
            'uv': data['uv'],
            'pv': data['pv'],
            'newUserCount': data['newUserCount'],
            'sv': data['sv'],
            'ip': data['ip'],
            'bounceRate': round(data['bounceRate_sum'] / data['bounceRate_count'], 4) if data['bounceRate_count'] > 0 else 0
        }
    
    return result


def get_orders_for_report(start_date, end_date, granularity='month', source=None, country=None, manager=None):
    """Get orders aggregated by source and time period for report
    
    Args:
        granularity: 'day', 'week', or 'month'
    """
    conn = get_db_connection()
    
    # Build time grouping expression based on granularity
    if granularity == 'day':
        time_expr = "strftime('%Y-%m-%d', date_created)"
    elif granularity == 'week':
        # ISO week format: YYYY-Www
        time_expr = "strftime('%Y-W%W', date_created)"
    else:  # month
        time_expr = "strftime('%Y-%m', date_created)"
    
    # success_net excludes undelivered (via _revenue_status_cond) AND deducts
    # the shipping_loss that those undelivered orders cost us — keeps the
    # report's net column consistent with /monthly and /orders summary views.
    query = f'''
        SELECT
            source,
            {time_expr} as period,
            COUNT(*) as order_count,
            SUM(total) as total_amount,
            SUM(total - shipping_total) as net_amount,
            currency,
            SUM(CASE WHEN {_revenue_status_cond()}
                THEN total - shipping_total ELSE 0 END)
            - SUM(CASE WHEN COALESCE(is_undelivered, 0) = 1
                THEN COALESCE(shipping_loss_amount, 0) ELSE 0 END) as success_net
        FROM orders
        WHERE date_created >= ? AND date_created <= ?
          AND status NOT IN ('checkout-draft', 'trash')
    '''
    params = [start_date, end_date + 'T23:59:59']
    
    site_filters = []
    site_filter_params = []
    if country:
        site_filters.append('country = ?')
        site_filter_params.append(country)
    if manager:
        site_filters.append('manager = ?')
        site_filter_params.append(manager)

    if site_filters:
        filtered_sites = conn.execute(
            f'SELECT url FROM sites WHERE {" AND ".join(site_filters)}',
            site_filter_params
        ).fetchall()
        filtered_urls = [s['url'] for s in filtered_sites]
        if filtered_urls:
            placeholders = ', '.join(['?' for _ in filtered_urls])
            query += f' AND source IN ({placeholders})'
            params.extend(filtered_urls)
        else:
            query += ' AND 1=0'
    
    if source:
        query += ' AND source = ?'
        params.append(source)
    
    query += f' GROUP BY source, period, currency ORDER BY source, period'
    
    orders = conn.execute(query, params).fetchall()
    conn.close()
    
    result = {}
    for row in orders:
        source_url = row['source']
        period = row['period']
        currency = row['currency']
        
        if source_url not in result:
            result[source_url] = {'currency': currency, 'periods': {}}
        
        if period not in result[source_url]['periods']:
            result[source_url]['periods'][period] = {
                'order_count': 0,
                'total_amount': 0,
                'net_amount': 0,
                'net_cny': 0
            }
        
        # Get exchange rate (use month from period for rate lookup)
        rate_month = period[:7] if len(period) >= 7 else period
        rate, _ = get_cny_rate(currency, rate_month)
        rate = rate or 1
        
        result[source_url]['periods'][period]['order_count'] += row['order_count']
        result[source_url]['periods'][period]['total_amount'] += row['total_amount'] or 0
        result[source_url]['periods'][period]['net_amount'] += row['success_net'] or 0
        result[source_url]['periods'][period]['net_cny'] += (row['success_net'] or 0) * rate
    
    return result


@app.route('/report')
@login_required
@report_viewer_required
def report():
    """Combined traffic and orders report page"""
    conn = get_db_connection()
    all_countries = conn.execute('SELECT DISTINCT country FROM sites WHERE country IS NOT NULL AND country != "" ORDER BY country').fetchall()
    all_countries = [c['country'] for c in all_countries]
    all_managers = conn.execute('SELECT DISTINCT manager FROM sites WHERE manager IS NOT NULL AND manager != "" ORDER BY manager').fetchall()
    all_managers = [m['manager'] for m in all_managers]
    conn.close()
    resp = make_response(render_template('report.html', all_countries=all_countries, all_managers=all_managers))
    resp.headers['Cache-Control'] = 'no-store, no-cache, must-revalidate'
    resp.headers['Pragma'] = 'no-cache'
    return resp


@app.route('/api/report/data')
@login_required
@report_viewer_required
def api_report_data():
    """API to get combined report data with granularity support"""
    from datetime import datetime, timedelta
    
    start_date = request.args.get('start_date', '')
    end_date = request.args.get('end_date', '')
    granularity = request.args.get('granularity', 'month')  # day, week, month
    country = request.args.get('country', '')
    manager = request.args.get('manager', '')
    
    # Validate granularity
    if granularity not in ('day', 'week', 'month'):
        granularity = 'month'
    
    if not end_date:
        end_date = datetime.now().strftime('%Y-%m-%d')
    if not start_date:
        start_dt = datetime.now() - timedelta(days=180)
        start_date = start_dt.strftime('%Y-%m-%d')
    
    # Cache Logic - include granularity, country, and manager in key
    cache_key = f"traffic_{start_date}_{end_date}_{granularity}_{country}_{manager}"
    force_refresh = request.args.get('force', 'false') == 'true'
    
    # Try load from server cache first (if not forced)
    conn = get_db_connection()
    if not force_refresh:
        try:
            cached_row = conn.execute('SELECT cache_value FROM report_cache WHERE cache_key = ?', (cache_key,)).fetchone()
            if cached_row and cached_row['cache_value']:
                print(f"DEBUG: Hit server cache for {cache_key}")
                cached_data = json.loads(cached_row['cache_value'])
                conn.close()
                return jsonify(cached_data)
        except Exception as e:
            print(f"Cache read error: {e}")
    
    all_traffic = {}
    # Use the new function with granularity support
    orders_data = get_orders_for_report(start_date, end_date, granularity, country=country, manager=manager)
    
    # Get sites config from database
    db_sites = conn.execute('SELECT url, mask_id, manager, country FROM sites').fetchall()
    
    def normalize_domain(url):
        if not url: return "unknown"
        url = url.lower().strip()
        if 'https://' in url:
            url = url.replace('https://', '')
        if 'http://' in url:
            url = url.replace('http://', '')
        if 'www.' in url:
            url = url.replace('www.', '')
        return url.split('/')[0]

    # Create mask_id mapping (normalized)
    site_mask_map = {}
    for site in db_sites:
        domain = normalize_domain(site['url'])
        if site['mask_id']:
            site_mask_map[domain] = site['mask_id']
            
    # Filter sites by country / manager if selected
    if country or manager:
        filtered_domains = set()
        for site in db_sites:
            if country and (site['country'] or '') != country:
                continue
            if manager and (site['manager'] or '') != manager:
                continue
            filtered_domains.add(normalize_domain(site['url']))

        site_mask_map = {k: v for k, v in site_mask_map.items() if k in filtered_domains}
            
    # Add hardcoded sites if not in DB (backward compatibility)
    if not country and not manager:
        for site, mask_id in LA_API_CONFIG['sites'].items():
            norm_site = normalize_domain(site)
            if norm_site not in site_mask_map:
                site_mask_map[norm_site] = mask_id
            
    # Fetch traffic data for each site
    for site_name, mask_id in site_mask_map.items():
        traffic = fetch_51la_traffic(mask_id, start_date, end_date)
        if traffic:
            all_traffic[site_name] = traffic

    combined = {}
    all_periods = set()
    
    # Select aggregation function based on granularity
    for site_name, traffic_list in all_traffic.items():
        if site_name not in combined:
            combined[site_name] = {}
        
        # Aggregate traffic based on granularity
        if granularity == 'day':
            aggregated_traffic = aggregate_traffic_by_day(traffic_list)
        elif granularity == 'week':
            aggregated_traffic = aggregate_traffic_by_week(traffic_list)
        else:
            aggregated_traffic = aggregate_traffic_by_month(traffic_list)
        
        for period, data in aggregated_traffic.items():
            all_periods.add(period)
            if period not in combined[site_name]:
                combined[site_name][period] = {}
            combined[site_name][period]['uv'] = data['uv']
            combined[site_name][period]['pv'] = data['pv']
            combined[site_name][period]['newUserCount'] = data['newUserCount']
            combined[site_name][period]['bounceRate'] = data['bounceRate']
    
    # Process orders data (new structure with currency)
    site_currencies = {}  # Store currency per site
    for source, source_data in orders_data.items():
        site_name = None
        norm_source = normalize_domain(source)
        
        if norm_source in site_mask_map:
            site_name = norm_source
        else:
            for name in site_mask_map.keys():
                if name in norm_source:
                    site_name = name
                    break
        
        if not site_name:
            site_name = norm_source
        
        # Save currency for this site
        if 'currency' in source_data:
            site_currencies[site_name] = source_data['currency']
        
        if site_name not in combined:
            combined[site_name] = {}
        
        # Get periods from the new structure
        periods_data = source_data.get('periods', source_data)  # Fallback for old structure
        
        for period, data in periods_data.items():
            if period == 'currency':  # Skip the currency field
                continue
            all_periods.add(period)
            if period not in combined[site_name]:
                combined[site_name][period] = {}
            
            current = combined[site_name][period]
            current['order_count'] = current.get('order_count', 0) + data['order_count']
            current['total_amount'] = current.get('total_amount', 0) + data['total_amount']
            current['net_amount'] = current.get('net_amount', 0) + data['net_amount']
            current['net_cny'] = current.get('net_cny', 0) + data['net_cny']
    
    # Calculate derived metrics
    for site_name, periods in combined.items():
        for period, data in periods.items():
            uv = data.get('uv', 0)
            orders = data.get('order_count', 0)
            net_cny = data.get('net_cny', 0)
            
            data['conversion_rate'] = round(orders / uv * 100, 2) if uv > 0 else None
            data['aov_cny'] = round(net_cny / orders, 2) if orders > 0 else None
    
    sorted_periods = sorted(all_periods)
    
    final_result = {
        'success': True,
        'data': combined,
        'periods': sorted_periods,
        'months': sorted_periods,   # Keep for backward compatibility
        'sites': list(site_mask_map.keys()),
        'site_currencies': site_currencies,  # Currency per site
        'date_range': {'start': start_date, 'end': end_date},
        'granularity': granularity,
        'filters': {'country': country, 'manager': manager}
    }

    
    # Save to server cache if we have traffic data (to avoid caching empty/failed API calls)
    # Only cache if at least one site has traffic data
    has_traffic = any('uv' in m for s in combined.values() for m in s.values())
    if has_traffic:
        try:
            conn.execute('INSERT OR REPLACE INTO report_cache (cache_key, cache_value) VALUES (?, ?)', 
                        (cache_key, json.dumps(final_result)))
            conn.commit()
            print(f"DEBUG: Updated server cache for {cache_key}")
        except Exception as e:
            print(f"Cache write error: {e}")
            
    conn.close()
    return jsonify(final_result)


@app.route('/api/report/cache/sync', methods=['POST'])
@login_required
def report_cache_sync():
    """Sync client-side traffic data cache to server"""
    # Only admin should start this (or anyone with report view permission)
    if not current_user.can_view_report():
         return jsonify({'success': False, 'error': '无权限'}), 403
         
    try:
        data = request.json
        if not data:
             return jsonify({'success': False, 'error': '无数据'}), 400
             
        start_date = data.get('start_date')
        end_date = data.get('end_date')
        granularity = data.get('granularity', 'month')
        country = data.get('country', '')
        manager = data.get('manager', '')
        traffic_data = data.get('data') # Expecting the full result object structure
        
        if not start_date or not end_date or not traffic_data:
            return jsonify({'success': False, 'error': '数据不完整'}), 400
            
        cache_key = f"traffic_{start_date}_{end_date}_{granularity}_{country}_{manager}"
        
        # Verify structure roughly
        if 'data' not in traffic_data or 'months' not in traffic_data:
             return jsonify({'success': False, 'error': '数据格式错误'}), 400
             
        # Save to DB
        conn = get_db_connection()
        conn.execute('INSERT OR REPLACE INTO report_cache (cache_key, cache_value) VALUES (?, ?)', 
                    (cache_key, json.dumps(traffic_data)))
        conn.commit()
        conn.close()
        
        return jsonify({'success': True, 'message': '缓存已同步到服务器'})
        
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


## ==================== Sales Board (销售看板) ====================

def sales_board_required(f):
    """Decorator to require sales board viewing permission"""
    from functools import wraps
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not current_user.is_authenticated:
            return redirect(url_for('login'))
        # Full team view OR self-only view both grant access here; data scoping
        # for self-only happens in the route (others' rows/salary never sent).
        if not (current_user.can_view_sales_board() or current_user.can_view_own_sales_board()):
            return render_template_string('''
<!DOCTYPE html>
<html lang="zh"><head><meta charset="UTF-8"><title>访问受限</title>
<link href="https://cdn.jsdelivr.net/npm/bootstrap@5.3.0/dist/css/bootstrap.min.css" rel="stylesheet">
<style>body{min-height:100vh;display:flex;align-items:center;justify-content:center;background:#1a1a2e;color:#e0e0e0;}</style>
</head><body><div class="text-center"><h1 class="display-1">🔒</h1><h3>无权访问销售看板</h3><p class="text-muted">请联系管理员开通权限</p><a href="/" class="btn btn-primary mt-3">返回首页</a></div></body></html>
'''), 403
        return f(*args, **kwargs)
    return decorated_function


def _get_sales_board_rate_overrides(year_month):
    """Return dict of {currency_upper: rate} for the given month from sales_board_exchange_rates."""
    conn = get_db_connection()
    try:
        rows = conn.execute(
            'SELECT currency, rate_to_cny FROM sales_board_exchange_rates WHERE year_month = ?',
            (year_month,)
        ).fetchall()
        return {(r['currency'] or '').upper(): float(r['rate_to_cny']) for r in rows}
    finally:
        conn.close()


def _get_board_cny_rate(currency, year_month, overrides=None):
    """CNY rate for sales-board calculations only.

    overrides: optional pre-fetched dict for the month, to avoid repeated DB hits.
    Falls back to the global get_cny_rate when no override is set.
    """
    if not currency or currency.upper() == 'CNY':
        return 1.0, 'override' if overrides and 'CNY' in overrides else 'system'
    cur_u = currency.upper()
    if overrides is None:
        overrides = _get_sales_board_rate_overrides(year_month)
    if cur_u in overrides:
        return overrides[cur_u], 'override'
    rate, _ = get_cny_rate(currency, year_month)
    return (rate or 0), 'system'


def _compute_sales_board_data(selected_month, restrict_manager=None):
    """Compute sales board data (shared by view and export).

    restrict_manager: when set (self-only view), the board is limited to just
    that one manager's sites — no other people's rows/salary are computed, so a
    self-view user never receives anyone else's data.

    Returns a dict with all the data needed to render the sales board.
    """
    import datetime
    from collections import defaultdict

    # Get previous month for growth calculation
    ym_parts = selected_month.split('-')
    sel_year, sel_mon = int(ym_parts[0]), int(ym_parts[1])
    if sel_mon == 1:
        prev_month = f"{sel_year - 1}-12"
    else:
        prev_month = f"{sel_year}-{sel_mon - 1:02d}"

    conn = get_db_connection()

    # Sales-board-only exchange rate overrides
    _board_overrides_cur = _get_sales_board_rate_overrides(selected_month)
    _board_overrides_prev = _get_sales_board_rate_overrides(prev_month)

    # Get all managers and their sites
    sites = conn.execute('SELECT url, manager FROM sites WHERE manager IS NOT NULL AND manager != ""').fetchall()
    manager_sites = defaultdict(list)
    for s in sites:
        manager_sites[s['manager']].append(s['url'])

    # Self-only view: keep just this manager (drop everyone else entirely).
    if restrict_manager is not None:
        manager_sites = defaultdict(list, {restrict_manager: manager_sites.get(restrict_manager, [])})

    managers = sorted(manager_sites.keys())

    # Get sales targets for selected month
    targets = {}
    target_rows = conn.execute('SELECT * FROM sales_targets WHERE year_month = ?', (selected_month,)).fetchall()
    for t in target_rows:
        targets[t['manager']] = dict(t)

    # Get no-commission brand names
    no_comm_rows = conn.execute('SELECT brand_name FROM no_commission_brands').fetchall()
    no_commission_brands_display = [r['brand_name'] for r in no_comm_rows]
    no_commission_brands = [b.upper() for b in no_commission_brands_display]

    # Get all available brands for the checkbox picker
    all_brands_rows = conn.execute('SELECT name FROM brands ORDER BY name').fetchall()
    all_brands_list = [r['name'] for r in all_brands_rows]

    # Get brands cache for product name parsing
    brands_rows = conn.execute('SELECT id, name, aliases FROM brands').fetchall()
    brands_cache = []
    for row in brands_rows:
        brand_name = row['name']
        try:
            aliases = json.loads(row['aliases']) if row['aliases'] else []
        except:
            aliases = []
        brands_cache.append({
            'id': row['id'],
            'name': brand_name,
            'aliases': aliases,
            'patterns': [brand_name.upper()] + [a.upper() for a in aliases]
        })

    # Load profit settings for this month
    profit_row = conn.execute(
        'SELECT profit_mode, profit_percentage, country_percentages FROM sales_board_profit_settings WHERE year_month = ?',
        (selected_month,)
    ).fetchone()
    profit_mode = profit_row['profit_mode'] if profit_row else 'percentage'
    profit_percentage = profit_row['profit_percentage'] if profit_row else 50.0
    try:
        country_percentages = json.loads(profit_row['country_percentages']) if profit_row and profit_row['country_percentages'] else {}
    except:
        country_percentages = {}

    # Build site->country map (for per-country profit tracking)
    site_country_map = {}
    sc_rows = conn.execute('SELECT url, country FROM sites').fetchall()
    for sc in sc_rows:
        site_country_map[sc['url']] = sc['country'] or 'PL'

    # Load product costs lookup (for actual cost mode). The cost index is
    # date-aware: a product can have multiple cost entries with different
    # effective_dates (e.g. 2/28 @ 28 PLN, 3/15 @ 30 PLN). Lookup at order time
    # picks the latest cost <= order_date so historical orders use historical
    # costs.
    cost_idx = None  # None when not in actual mode
    country_default_wh_ids = {}
    if profit_mode == 'actual':
        cost_idx = _build_dated_cost_index(conn=conn)

        # Country -> ordered list of warehouse_ids (smallest id first), used as a
        # fallback when an order has no warehouse_id assigned (legacy NULL data).
        for w in conn.execute('SELECT id, country FROM warehouses ORDER BY country, id'):
            country_default_wh_ids.setdefault(w['country'], []).append(w['id'])

        # Load product_mappings for item->brand resolution. Index by canonical
        # (entity-decoded) raw_name so HTML-entity vs decoded variants both hit.
        pm_rows = conn.execute('SELECT raw_name, source, brand_id, series_id, puff_count, flavor FROM product_mappings').fetchall()
        product_mappings_cache = {}
        for pm in pm_rows:
            product_mappings_cache[(normalize_raw_name(pm['raw_name']), pm['source'])] = pm

    # Calculate current week range (Monday to Sunday)
    today = datetime.date.today()
    week_start = today - datetime.timedelta(days=today.weekday())
    week_end = week_start + datetime.timedelta(days=6)
    # Determine which week number in the month (1-based)
    current_week_num = (today.day - 1) // 7 + 1

    # Build manager performance data
    board_data = []
    for manager in managers:
        site_urls = manager_sites[manager]
        placeholders = ', '.join(['?' for _ in site_urls])

        # Current month successful orders
        month_orders = conn.execute(f'''
            SELECT id, total, shipping_total, currency, line_items, source, date_created, warehouse_id
            FROM orders
            WHERE source IN ({placeholders})
            AND strftime('%Y-%m', date_created) = ?
            AND {_revenue_status_cond()}
        ''', site_urls + [selected_month]).fetchall()

        # Previous month successful orders (for growth)
        prev_orders = conn.execute(f'''
            SELECT total, shipping_total, currency, line_items
            FROM orders
            WHERE source IN ({placeholders})
            AND strftime('%Y-%m', date_created) = ?
            AND {_revenue_status_cond()}
        ''', site_urls + [prev_month]).fetchall()

        # Current week orders
        week_orders = conn.execute(f'''
            SELECT total, shipping_total, currency, line_items, source
            FROM orders
            WHERE source IN ({placeholders})
            AND date(date_created) >= ? AND date(date_created) <= ?
            AND {_revenue_status_cond()}
        ''', site_urls + [week_start.isoformat(), week_end.isoformat()]).fetchall()

        # Undelivered + 问题退货 orders for this manager (shipping/product loss
        # visibility — separate from revenue). Both flows live on shared
        # shipping_loss_amount column; product_loss_amount is problem-return only.
        undelivered_rows = conn.execute(f'''
            SELECT shipping_loss_amount, product_loss_amount, currency, source,
                   is_undelivered, is_problem_return
            FROM orders
            WHERE source IN ({placeholders})
            AND strftime('%Y-%m', date_created) = ?
            AND (is_undelivered = 1 OR is_problem_return = 1)
        ''', site_urls + [selected_month]).fetchall()

        # Previous-month shipping + product loss — needed so the 上月销售 card on
        # the board uses the same net definition as the current month (and as
        # /monthly / dashboard).
        prev_undelivered_rows = conn.execute(f'''
            SELECT shipping_loss_amount, product_loss_amount, currency,
                   is_undelivered, is_problem_return
            FROM orders
            WHERE source IN ({placeholders})
            AND strftime('%Y-%m', date_created) = ?
            AND (is_undelivered = 1 OR is_problem_return = 1)
        ''', site_urls + [prev_month]).fetchall()

        # Calculate month amounts by currency and total products
        month_currency_amounts = defaultdict(lambda: {'net_amount': 0, 'amount': 0, 'shipping': 0})
        month_total_products = 0
        month_net_cny = 0
        month_commission_base_cny = 0  # Sales eligible for commission (in CNY)
        month_shipping_cny = 0           # Shipping not counted toward commission (in CNY)
        month_excluded_cny = 0           # Excluded-brand product revenue (in CNY)
        month_cost_cny = 0               # Total product cost (in CNY, actual mode)
        month_unmapped_revenue_cny = 0   # Revenue of products without cost mapping
        month_unmapped_count = 0         # Number of unmapped product types
        # Per-country profit tracking (for reconciliation)
        country_profit = defaultdict(lambda: {'net_cny': 0, 'cost_cny': 0, 'unmapped_revenue_cny': 0, 'unmapped_count': 0, 'shipping_loss_cny': 0, 'undelivered_count': 0})
        # Per-brand exclusion breakdown: {brand: {qty, revenue_by_currency, revenue_cny}}
        excluded_brand_breakdown = defaultdict(lambda: {
            'qty': 0, 'revenue_by_currency': defaultdict(float), 'revenue_cny': 0.0
        })
        order_count = len(month_orders)

        # Per-currency exchange rate cache (rate for the selected month, with sales-board overrides)
        _month_rates = {}
        _month_rate_sources = {}  # 'override' or 'system'
        def _rate_for(cur):
            if cur not in _month_rates:
                r, src = _get_board_cny_rate(cur, selected_month, _board_overrides_cur)
                _month_rates[cur] = r or 0
                _month_rate_sources[cur] = src
            return _month_rates[cur]

        for order in month_orders:
            currency = order['currency'] or 'PLN'
            total = float(order['total'] or 0)
            shipping = float(order['shipping_total'] or 0)
            net = total - shipping
            month_currency_amounts[currency]['amount'] += total
            month_currency_amounts[currency]['net_amount'] += net
            month_currency_amounts[currency]['shipping'] += shipping

            rate = _rate_for(currency)
            order_country = site_country_map.get(order['source'], 'PL')
            order_wh_id = order['warehouse_id'] if profit_mode == 'actual' else None

            # Products count
            items = parse_json_field(order['line_items'])
            if isinstance(items, list):
                for item in items:
                    qty = item.get('quantity', 0) or 0
                    month_total_products += qty

                    # Check if product brand is excluded from commission
                    product_name_raw = item.get('name', '') or ''
                    product_name = product_name_raw.upper()
                    excluded_brand = None
                    for nc_brand in no_commission_brands:
                        if nc_brand in product_name:
                            excluded_brand = nc_brand
                            break
                    # Also check against brands_cache for better matching
                    if not excluded_brand:
                        parsed = parse_product_name(product_name_raw, brands_cache)
                        if parsed.get('brand') and parsed['brand'].upper() in no_commission_brands:
                            excluded_brand = parsed['brand'].upper()

                    item_total = float(item.get('total', 0) or 0)
                    if excluded_brand:
                        bucket = excluded_brand_breakdown[excluded_brand]
                        bucket['qty'] += qty
                        bucket['revenue_by_currency'][currency] += item_total
                        if rate:
                            bucket['revenue_cny'] += item_total * rate
                            month_excluded_cny += item_total * rate
                    else:
                        if rate:
                            month_commission_base_cny += item_total * rate

                    # Actual cost calculation — uses date-aware lookup so that
                    # historical orders use cost prices that were effective on
                    # their order date (not whatever the current price is).
                    if profit_mode == 'actual' and rate and qty > 0 and cost_idx is not None:
                        source = order['source']
                        order_date = (order['date_created'] or '')[:10]  # YYYY-MM-DD
                        wh_id = order_wh_id
                        b_id, s_id, p_cnt, flav = _resolve_product_to_brand(
                            product_name_raw, source, brands_cache, product_mappings_cache
                        )
                        # Find cost entry matching this warehouse (priority fallback).
                        # If the order has no warehouse_id, fall back to country
                        # defaults rather than declaring the product unmapped.
                        effective_wh_ids = [wh_id] if wh_id else country_default_wh_ids.get(order_country, [])
                        cost_entry = None
                        if b_id and effective_wh_ids:
                            for ewh in effective_wh_ids:
                                cost_entry = _cost_at_date(cost_idx, b_id, s_id, p_cnt, flav, ewh, order_date)
                                if cost_entry:
                                    break

                        if cost_entry:
                            cost_rate = _rate_for(cost_entry['currency'])
                            item_cost = cost_entry['price'] * qty * cost_rate
                            month_cost_cny += item_cost
                            country_profit[order_country]['cost_cny'] += item_cost
                        else:
                            month_unmapped_revenue_cny += item_total * rate
                            month_unmapped_count += 1
                            country_profit[order_country]['unmapped_revenue_cny'] += item_total * rate
                            country_profit[order_country]['unmapped_count'] += 1

            # CNY conversion for net amount & shipping
            if rate:
                month_net_cny += net * rate
                month_shipping_cny += shipping * rate
                country_profit[order_country]['net_cny'] += net * rate

        # Previous month CNY (uses prev-month board overrides if set).
        # Net definition mirrors current month: gross product revenue minus
        # collected shipping minus shipping_loss from undelivered orders.
        prev_net_cny = 0
        _prev_rate_cache = {}
        def _prev_rate(cur):
            if cur not in _prev_rate_cache:
                r, _src = _get_board_cny_rate(cur, prev_month, _board_overrides_prev)
                _prev_rate_cache[cur] = r or 0
            return _prev_rate_cache[cur]

        for order in prev_orders:
            currency = order['currency'] or 'PLN'
            total = float(order['total'] or 0)
            shipping = float(order['shipping_total'] or 0)
            net = total - shipping
            rate = _prev_rate(currency)
            if rate:
                prev_net_cny += net * rate
        # Subtract prev-month shipping_loss AND product_loss so 上月销售 card
        # matches /monthly.
        for u in prev_undelivered_rows:
            cur = u['currency'] or 'PLN'
            amt = float(u['shipping_loss_amount'] or 0)
            p_amt = float(u['product_loss_amount'] or 0) if bool(u['is_problem_return']) else 0.0
            r = _prev_rate(cur)
            if r:
                prev_net_cny -= (amt + p_amt) * r

        # Current week amounts by currency
        week_currency_amounts = defaultdict(lambda: {'net_amount': 0})
        week_net_cny = 0
        week_total_products = 0
        for order in week_orders:
            currency = order['currency'] or 'PLN'
            total = float(order['total'] or 0)
            shipping = float(order['shipping_total'] or 0)
            net = total - shipping
            week_currency_amounts[currency]['net_amount'] += net
            items = parse_json_field(order['line_items'])
            if isinstance(items, list):
                week_total_products += sum(i.get('quantity', 0) for i in items)
            rate = _rate_for(currency)
            if rate:
                week_net_cny += net * rate

        # Aggregate undelivered + 问题退货 losses (count + by-currency + CNY-equivalent).
        # shipping_loss_amount is populated by either flow; product_loss_amount
        # only by 问题退货.
        undelivered_count = sum(1 for u in undelivered_rows if bool(u['is_undelivered']))
        problem_return_count = sum(1 for u in undelivered_rows if bool(u['is_problem_return']))
        shipping_loss_by_currency = defaultdict(float)
        product_loss_by_currency = defaultdict(float)
        shipping_loss_cny = 0
        product_loss_cny = 0
        for u in undelivered_rows:
            cur = u['currency'] or 'PLN'
            amt = float(u['shipping_loss_amount'] or 0)
            p_amt = float(u['product_loss_amount'] or 0) if bool(u['is_problem_return']) else 0.0
            shipping_loss_by_currency[cur] += amt
            product_loss_by_currency[cur] += p_amt
            r = _rate_for(cur)
            if r:
                shipping_loss_cny += amt * r
                product_loss_cny += p_amt * r
                u_country = site_country_map.get(u['source'], 'PL')
                country_profit[u_country]['shipping_loss_cny'] += amt * r
                if bool(u['is_undelivered']):
                    country_profit[u_country]['undelivered_count'] += 1
                if bool(u['is_problem_return']):
                    country_profit[u_country].setdefault('product_loss_cny', 0)
                    country_profit[u_country].setdefault('problem_return_count', 0)
                    country_profit[u_country]['product_loss_cny'] += p_amt * r
                    country_profit[u_country]['problem_return_count'] += 1

        # Subtract shipping + product loss from net so 完成率 / 利润 / per-country
        # net match the rest of the system. Per-currency totals get the losses
        # baked in too — same pattern as /monthly.
        month_net_cny = month_net_cny - shipping_loss_cny - product_loss_cny
        for cur, loss_amt in shipping_loss_by_currency.items():
            month_currency_amounts[cur]['net_amount'] -= loss_amt
        for cur, loss_amt in product_loss_by_currency.items():
            month_currency_amounts[cur]['net_amount'] -= loss_amt
        # Per-country net was incremented above (line ~13757) without the
        # loss; now subtract each country's loss to get the true net.
        for c, v in country_profit.items():
            v['net_cny'] = v['net_cny'] - v['shipping_loss_cny'] - v.get('product_loss_cny', 0)

        # Get target info
        target = targets.get(manager, {})
        _mt = target.get('monthly_target'); monthly_target = _mt if _mt is not None else 0
        _bs = target.get('base_salary');   base_salary   = _bs if _bs is not None else 7000
        _cr = target.get('commission_rate'); commission_rate = _cr if _cr is not None else 0.05
        weekly_targets_json = target.get('weekly_targets', '{}') or '{}'
        try:
            weekly_targets = json.loads(weekly_targets_json)
        except:
            weekly_targets = {}
        weekly_target = weekly_targets.get(f'w{current_week_num}', 0) or 0

        # Achievement rates
        month_achievement = (month_net_cny / monthly_target * 100) if monthly_target > 0 else 0
        week_achievement = (week_net_cny / weekly_target * 100) if weekly_target > 0 else 0

        # Growth rate
        if prev_net_cny > 0:
            growth_rate = (month_net_cny - prev_net_cny) / prev_net_cny * 100
        else:
            growth_rate = 100 if month_net_cny > 0 else 0

        # Salary calculation
        growth_met = growth_rate >= 20
        units_met = month_total_products >= 500
        salary_protected = growth_met or units_met
        salary_deduction = 0 if salary_protected else round(base_salary * 0.2, 2)
        actual_salary = base_salary - salary_deduction

        # Commission
        commission = round(month_commission_base_cny * commission_rate, 2)

        # Total income
        total_income = round(actual_salary + commission, 2)

        board_data.append({
            'manager': manager,
            'sites': [u.replace('https://www.', '').replace('https://', '') for u in site_urls],
            'sites_raw': site_urls,
            'monthly_target': monthly_target,
            'weekly_target': weekly_target,
            'weekly_targets': weekly_targets,
            'current_week_num': current_week_num,
            'month_currency_amounts': dict(month_currency_amounts),
            'week_currency_amounts': dict(week_currency_amounts),
            'month_net_cny': round(month_net_cny, 2),
            'week_net_cny': round(week_net_cny, 2),
            'month_total_products': month_total_products,
            'week_total_products': week_total_products,
            'month_achievement': round(month_achievement, 1),
            'week_achievement': round(week_achievement, 1),
            'prev_net_cny': round(prev_net_cny, 2),
            'growth_rate': round(growth_rate, 1),
            'growth_met': growth_met,
            'units_met': units_met,
            'salary_protected': salary_protected,
            'base_salary': base_salary,
            'salary_deduction': salary_deduction,
            'actual_salary': actual_salary,
            'commission_rate': commission_rate,
            'commission_base_cny': round(month_commission_base_cny, 2),
            'commission': commission,
            'total_income': total_income,
            'notes': target.get('notes', ''),
            # Extra details for export / audit
            'order_count': order_count,
            'month_shipping_cny': round(month_shipping_cny, 2),
            'month_excluded_cny': round(month_excluded_cny, 2),
            'undelivered_count': undelivered_count,
            'shipping_loss_by_currency': {c: round(v, 2) for c, v in shipping_loss_by_currency.items()},
            'shipping_loss_cny': round(shipping_loss_cny, 2),
            'problem_return_count': problem_return_count,
            'product_loss_by_currency': {c: round(v, 2) for c, v in product_loss_by_currency.items()},
            'product_loss_cny': round(product_loss_cny, 2),
            'excluded_brand_breakdown': {
                brand: {
                    'qty': v['qty'],
                    'revenue_by_currency': {c: round(amt, 2) for c, amt in v['revenue_by_currency'].items()},
                    'revenue_cny': round(v['revenue_cny'], 2),
                } for brand, v in excluded_brand_breakdown.items()
            },
            'month_rates': {c: r for c, r in _month_rates.items() if r},
            'month_rate_sources': dict(_month_rate_sources),
            # Profit calculation fields
            'month_cost_cny': round(month_cost_cny, 2),
            'month_unmapped_revenue_cny': round(month_unmapped_revenue_cny, 2),
            'month_unmapped_count': month_unmapped_count,
            'month_profit_actual_cny': round(month_net_cny - month_cost_cny - month_unmapped_revenue_cny * (1 - profit_percentage / 100), 2) if profit_mode == 'actual' else 0,
            'month_profit_pct_cny': round(month_net_cny * profit_percentage / 100, 2),
            'country_profit': {c: {
                'net_cny': round(v['net_cny'], 2),
                'cost_cny': round(v['cost_cny'], 2),
                'unmapped_revenue_cny': round(v['unmapped_revenue_cny'], 2),
                'unmapped_count': v['unmapped_count'],
                'shipping_loss_cny': round(v['shipping_loss_cny'], 2),
                'undelivered_count': v['undelivered_count'],
                'product_loss_cny': round(v.get('product_loss_cny', 0), 2),
                'problem_return_count': v.get('problem_return_count', 0),
            } for c, v in country_profit.items()},
        })

    # Group summaries
    group_summaries = []
    groups_raw = conn.execute('SELECT * FROM sales_groups ORDER BY id').fetchall()
    for group in groups_raw:
        members_rows = conn.execute(
            'SELECT manager FROM sales_group_members WHERE group_id = ?',
            (group['id'],)
        ).fetchall()
        member_names = [r['manager'] for r in members_rows]
        leader = group['leader_manager']

        # All member data (including leader) for group display totals
        member_data = [d for d in board_data if d['manager'] in member_names]
        # Exclude leader for bonus calculation
        non_leader_data = [d for d in member_data if d['manager'] != leader]

        group_month_net_cny = sum(d['month_net_cny'] for d in member_data)
        group_prev_net_cny = sum(d['prev_net_cny'] for d in member_data)
        group_week_net_cny = sum(d['week_net_cny'] for d in member_data)
        group_products = sum(d['month_total_products'] for d in member_data)
        group_monthly_target = sum(d['monthly_target'] for d in member_data)
        group_commission = sum(d['commission'] for d in member_data)
        group_total_income = sum(d['total_income'] for d in member_data)

        # Bonus base = commission base of non-leader members only
        bonus_base_cny = sum(d['commission_base_cny'] for d in non_leader_data)
        bonus_rate = group['bonus_rate'] or 0.02
        leader_bonus = round(bonus_base_cny * bonus_rate, 2)

        group_summaries.append({
            'id': group['id'],
            'name': group['name'],
            'leader_manager': leader,
            'bonus_rate': bonus_rate,
            'members': member_names,
            'month_net_cny': round(group_month_net_cny, 2),
            'bonus_base_cny': round(bonus_base_cny, 2),
            'prev_net_cny': round(group_prev_net_cny, 2),
            'week_net_cny': round(group_week_net_cny, 2),
            'month_total_products': group_products,
            'monthly_target': group_monthly_target,
            'month_achievement': round(group_month_net_cny / group_monthly_target * 100, 1) if group_monthly_target > 0 else 0,
            'leader_bonus': leader_bonus,
            'total_commission': round(group_commission, 2),
            'total_income': round(group_total_income, 2),
        })

    conn.close()

    # Team totals (include leader bonuses so the total matches the sum of displayed rows)
    total_leader_bonuses = sum(g['leader_bonus'] for g in group_summaries)
    team_totals = {
        'monthly_target': sum(d['monthly_target'] for d in board_data),
        'month_net_cny': sum(d['month_net_cny'] for d in board_data),
        'prev_net_cny': sum(d['prev_net_cny'] for d in board_data),
        'week_net_cny': sum(d['week_net_cny'] for d in board_data),
        'month_total_products': sum(d['month_total_products'] for d in board_data),
        'total_commission': sum(d['commission'] for d in board_data),
        'total_income': round(sum(d['total_income'] for d in board_data) + total_leader_bonuses, 2),
        'undelivered_count': sum(d.get('undelivered_count', 0) for d in board_data),
        'shipping_loss_cny': round(sum(d.get('shipping_loss_cny', 0) for d in board_data), 2),
        'problem_return_count': sum(d.get('problem_return_count', 0) for d in board_data),
        'product_loss_cny': round(sum(d.get('product_loss_cny', 0) for d in board_data), 2),
    }
    # Compute masked values for "hide leader commission" display mode.
    # Leaders' base salary is shown as the default (7000) so they look like regular members.
    MASK_BASE_SALARY = 7000
    leader_names_set = {g['leader_manager'] for g in group_summaries}
    total_leader_base_reduction = 0
    for d in board_data:
        if d['manager'] in leader_names_set:
            reduction = d['base_salary'] - MASK_BASE_SALARY
            total_leader_base_reduction += reduction
            d['masked_base_salary'] = MASK_BASE_SALARY
            d['masked_total_income'] = round(MASK_BASE_SALARY - d['salary_deduction'] + d['commission'], 2)
        else:
            d['masked_base_salary'] = d['base_salary']
            d['masked_total_income'] = d['total_income']
    team_totals['total_income_no_bonus'] = round(
        sum(d['total_income'] for d in board_data) - total_leader_base_reduction, 2
    )
    if team_totals['monthly_target'] > 0:
        team_totals['month_achievement'] = round(team_totals['month_net_cny'] / team_totals['monthly_target'] * 100, 1)
    else:
        team_totals['month_achievement'] = 0

    is_current_month = (selected_month == datetime.date.today().strftime('%Y-%m'))

    # Build leader bonus map: {leader_manager: leader_bonus}
    leader_bonus_map = {g['leader_manager']: g['leader_bonus'] for g in group_summaries}

    # Aggregate the unique exchange rates used (for display in UI)
    rates_in_use = {}
    for d in board_data:
        for cur, rate in (d.get('month_rates') or {}).items():
            if cur not in rates_in_use:
                rates_in_use[cur] = {
                    'rate': rate,
                    'source': (d.get('month_rate_sources') or {}).get(cur, 'system'),
                }

    # Profit summary for team totals
    team_totals['month_cost_cny'] = round(sum(d.get('month_cost_cny', 0) for d in board_data), 2)
    team_totals['month_unmapped_revenue_cny'] = round(sum(d.get('month_unmapped_revenue_cny', 0) for d in board_data), 2)
    team_totals['month_unmapped_count'] = sum(d.get('month_unmapped_count', 0) for d in board_data)
    team_totals['month_profit_actual_cny'] = round(sum(d.get('month_profit_actual_cny', 0) for d in board_data), 2)
    team_totals['month_profit_pct_cny'] = round(sum(d.get('month_profit_pct_cny', 0) for d in board_data), 2)

    # Per-country profit totals
    country_profit_totals = defaultdict(lambda: {'net_cny': 0, 'cost_cny': 0, 'unmapped_revenue_cny': 0, 'unmapped_count': 0, 'shipping_loss_cny': 0, 'undelivered_count': 0, 'product_loss_cny': 0, 'problem_return_count': 0})
    for d in board_data:
        for c, v in d.get('country_profit', {}).items():
            for k in country_profit_totals[c]:
                country_profit_totals[c][k] += v.get(k, 0)
    team_totals['country_profit'] = {}
    for c, v in sorted(country_profit_totals.items()):
        cpct = country_percentages.get(c, profit_percentage)
        team_totals['country_profit'][c] = {
            'net_cny': round(v['net_cny'], 2),
            'cost_cny': round(v['cost_cny'], 2),
            'unmapped_revenue_cny': round(v['unmapped_revenue_cny'], 2),
            'unmapped_count': v['unmapped_count'],
            'shipping_loss_cny': round(v['shipping_loss_cny'], 2),
            'undelivered_count': v['undelivered_count'],
            'product_loss_cny': round(v.get('product_loss_cny', 0), 2),
            'problem_return_count': v.get('problem_return_count', 0),
            'profit_pct': cpct,
            'profit_actual_cny': round(v['net_cny'] - v['cost_cny'] - v['unmapped_revenue_cny'] * (1 - cpct / 100), 2),
            'profit_pct_cny': round(v['net_cny'] * cpct / 100, 2),
        }
    # Recalculate total pct profit from per-country values when country percentages differ
    if country_percentages:
        team_totals['month_profit_pct_cny'] = round(sum(cp['profit_pct_cny'] for cp in team_totals['country_profit'].values()), 2)

    return {
        'board_data': board_data,
        'team_totals': team_totals,
        'group_summaries': group_summaries,
        'leader_bonus_map': leader_bonus_map,
        'managers': managers,
        'selected_month': selected_month,
        'prev_month': prev_month,
        'current_week_num': current_week_num,
        'is_current_month': is_current_month,
        'no_commission_brands': no_commission_brands_display,
        'all_brands_list': all_brands_list,
        'rates_in_use': rates_in_use,
        'rate_overrides': _board_overrides_cur,
        'profit_mode': profit_mode,
        'profit_percentage': profit_percentage,
        'country_percentages': country_percentages,
    }


@app.route('/sales-board')
@login_required
@sales_board_required
def sales_board():
    """Sales board dashboard"""
    import datetime

    selected_month = request.args.get('month', '')
    if not selected_month:
        today = datetime.date.today()
        selected_month = today.strftime('%Y-%m')

    # 仅本人站点 takes PRECEDENCE: if a user is granted the self-only permission
    # they are scoped to their own sites — even if they also hold the full-board
    # permission or an admin role. Full team view is for users WITHOUT this flag.
    self_view = current_user.can_view_own_sales_board()
    own_manager = current_user.sales_board_own_manager() if self_view else None
    # When self-only but the user has no sites (e.g. sites reassigned after the
    # grant), restrict to a sentinel that matches no manager → empty board, NOT
    # the full team (never fall back to None here, that would show everyone).
    restrict_manager = (own_manager or '\x00__no_sites__') if self_view else None

    data = _compute_sales_board_data(selected_month, restrict_manager=restrict_manager)
    data['self_view'] = self_view
    data['own_manager'] = own_manager or ''
    return render_template('sales_board.html', **data)


@app.route('/api/sales-targets', methods=['GET'])
@login_required
@sales_board_required
def get_sales_targets():
    """Get sales targets for a month"""
    year_month = request.args.get('month', '')
    conn = get_db_connection()
    if year_month:
        rows = conn.execute('SELECT * FROM sales_targets WHERE year_month = ?', (year_month,)).fetchall()
    else:
        rows = conn.execute('SELECT * FROM sales_targets ORDER BY year_month DESC').fetchall()
    conn.close()
    return jsonify([dict(r) for r in rows])


@app.route('/api/sales-targets', methods=['POST'])
@login_required
@admin_required
def save_sales_target():
    """Create or update a sales target"""
    data = request.json
    year_month = data.get('year_month', '').strip()
    manager = data.get('manager', '').strip()
    monthly_target = data.get('monthly_target', 0)
    weekly_targets = data.get('weekly_targets', {})
    base_salary = data.get('base_salary', 7000)
    commission_rate = data.get('commission_rate', 0.05)
    notes = data.get('notes', '')

    if not year_month or not manager:
        return jsonify({'error': '月份和负责人为必填项'}), 400

    conn = get_db_connection()
    try:
        conn.execute('''
            INSERT INTO sales_targets (year_month, manager, monthly_target, weekly_targets, base_salary, commission_rate, notes, updated_at)
            VALUES (?, ?, ?, ?, ?, ?, ?, datetime('now'))
            ON CONFLICT(year_month, manager) DO UPDATE SET
                monthly_target = excluded.monthly_target,
                weekly_targets = excluded.weekly_targets,
                base_salary = excluded.base_salary,
                commission_rate = excluded.commission_rate,
                notes = excluded.notes,
                updated_at = datetime('now')
        ''', (year_month, manager, monthly_target, json.dumps(weekly_targets), base_salary, commission_rate, notes))
        conn.commit()
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'error': str(e)}), 500
    finally:
        conn.close()


@app.route('/api/sales-targets/batch', methods=['POST'])
@login_required
@admin_required
def batch_save_sales_targets():
    """Batch save sales targets for all managers in a month"""
    data = request.json
    targets = data.get('targets', [])
    if not targets:
        return jsonify({'error': '无数据'}), 400

    conn = get_db_connection()
    try:
        for t in targets:
            year_month = t.get('year_month', '').strip()
            manager = t.get('manager', '').strip()
            if not year_month or not manager:
                continue
            conn.execute('''
                INSERT INTO sales_targets (year_month, manager, monthly_target, weekly_targets, base_salary, commission_rate, notes, updated_at)
                VALUES (?, ?, ?, ?, ?, ?, ?, datetime('now'))
                ON CONFLICT(year_month, manager) DO UPDATE SET
                    monthly_target = excluded.monthly_target,
                    weekly_targets = excluded.weekly_targets,
                    base_salary = excluded.base_salary,
                    commission_rate = excluded.commission_rate,
                    notes = excluded.notes,
                    updated_at = datetime('now')
            ''', (year_month, manager,
                  t.get('monthly_target', 0),
                  json.dumps(t.get('weekly_targets', {})),
                  t.get('base_salary', 7000),
                  t.get('commission_rate', 0.05),
                  t.get('notes', '')))
        conn.commit()
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'error': str(e)}), 500
    finally:
        conn.close()


@app.route('/api/no-commission-brands', methods=['GET'])
@login_required
@sales_board_required
def get_no_commission_brands():
    """Get list of no-commission brands"""
    conn = get_db_connection()
    rows = conn.execute('SELECT * FROM no_commission_brands ORDER BY brand_name').fetchall()
    conn.close()
    return jsonify([dict(r) for r in rows])


@app.route('/api/no-commission-brands', methods=['POST'])
@login_required
@admin_required
def update_no_commission_brands():
    """Update no-commission brands list"""
    data = request.json
    brands = data.get('brands', [])
    conn = get_db_connection()
    try:
        conn.execute('DELETE FROM no_commission_brands')
        for brand in brands:
            brand = brand.strip()
            if brand:
                conn.execute('INSERT INTO no_commission_brands (brand_name) VALUES (?)', (brand,))
        conn.commit()
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'error': str(e)}), 500
    finally:
        conn.close()


@app.route('/api/sales-groups', methods=['GET'])
@login_required
@sales_board_required
def get_sales_groups():
    """Get all sales groups with members"""
    conn = get_db_connection()
    groups = conn.execute('SELECT * FROM sales_groups ORDER BY id').fetchall()
    result = []
    for g in groups:
        members = conn.execute(
            'SELECT manager FROM sales_group_members WHERE group_id = ? ORDER BY manager',
            (g['id'],)
        ).fetchall()
        result.append({
            'id': g['id'],
            'name': g['name'],
            'leader_manager': g['leader_manager'],
            'bonus_rate': g['bonus_rate'],
            'members': [m['manager'] for m in members]
        })
    conn.close()
    return jsonify(result)


@app.route('/api/sales-groups', methods=['POST'])
@login_required
@admin_required
def save_sales_group():
    """Create or update a sales group"""
    data = request.json
    group_id = data.get('id')
    name = (data.get('name') or '').strip()
    leader_manager = (data.get('leader_manager') or '').strip()
    bonus_rate = data.get('bonus_rate', 0.02)
    members = data.get('members', [])

    if not name or not leader_manager:
        return jsonify({'error': '小组名称和组长为必填项'}), 400

    conn = get_db_connection()
    try:
        if group_id:
            conn.execute('''
                UPDATE sales_groups SET name = ?, leader_manager = ?, bonus_rate = ?, updated_at = datetime('now')
                WHERE id = ?
            ''', (name, leader_manager, bonus_rate, group_id))
        else:
            cursor = conn.execute('''
                INSERT INTO sales_groups (name, leader_manager, bonus_rate) VALUES (?, ?, ?)
            ''', (name, leader_manager, bonus_rate))
            group_id = cursor.lastrowid

        # Update members: delete all then reinsert
        conn.execute('DELETE FROM sales_group_members WHERE group_id = ?', (group_id,))
        for m in members:
            m = m.strip()
            if m:
                conn.execute('INSERT INTO sales_group_members (group_id, manager) VALUES (?, ?)', (group_id, m))

        conn.commit()
        return jsonify({'success': True, 'id': group_id})
    except Exception as e:
        return jsonify({'error': str(e)}), 500
    finally:
        conn.close()


@app.route('/api/sales-groups/<int:group_id>', methods=['DELETE'])
@login_required
@admin_required
def delete_sales_group(group_id):
    """Delete a sales group"""
    conn = get_db_connection()
    try:
        conn.execute('DELETE FROM sales_group_members WHERE group_id = ?', (group_id,))
        conn.execute('DELETE FROM sales_groups WHERE id = ?', (group_id,))
        conn.commit()
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'error': str(e)}), 500
    finally:
        conn.close()


# ----------------- Sales Board Custom Exchange Rates -----------------
@app.route('/api/sales-board/exchange-rates', methods=['GET'])
@login_required
@admin_required
def get_sales_board_exchange_rates():
    """List custom rates + applicable system rates for a given month.

    For each currency, returns {currency, system_rate, override_rate, in_use, source}.
    """
    month = (request.args.get('month') or '').strip()
    if not month:
        import datetime
        month = datetime.date.today().strftime('%Y-%m')

    overrides = _get_sales_board_rate_overrides(month)

    # Determine which currencies are in use this month
    conn = get_db_connection()
    try:
        rows = conn.execute('''
            SELECT DISTINCT o.currency
            FROM orders o
            INNER JOIN sites s ON o.source = s.url
            WHERE s.manager IS NOT NULL AND s.manager != ''
              AND strftime('%Y-%m', o.date_created) = ?
              AND o.currency IS NOT NULL AND o.currency != ''
        ''', (month,)).fetchall()
        currencies = sorted({(r['currency'] or '').upper() for r in rows if r['currency']})
        # Always include any currency that has an override even if no orders this month
        for c in overrides.keys():
            if c not in currencies:
                currencies.append(c)
        currencies = sorted(set(currencies))

        result = []
        for cur in currencies:
            if cur == 'CNY':
                continue
            sys_rate, _ = get_cny_rate(cur, month)
            override = overrides.get(cur)
            result.append({
                'currency': cur,
                'system_rate': round(sys_rate, 6) if sys_rate else None,
                'override_rate': override,
                'in_use': override if override is not None else (sys_rate or 0),
                'source': 'override' if override is not None else 'system',
            })
        return jsonify({'month': month, 'rates': result})
    finally:
        conn.close()


@app.route('/api/sales-board/exchange-rates', methods=['POST'])
@login_required
@admin_required
def save_sales_board_exchange_rates():
    """Save custom exchange rate overrides for a month.

    Payload: { "month": "YYYY-MM", "rates": [{"currency": "PLN", "rate": 1.95}, ...] }
    A rate of null/empty/0 means "remove override (use system rate)".
    """
    data = request.get_json(silent=True) or {}
    month = (data.get('month') or '').strip()
    rates = data.get('rates') or []
    if not month:
        return jsonify({'error': '缺少月份参数'}), 400

    conn = get_db_connection()
    try:
        for entry in rates:
            cur = (entry.get('currency') or '').strip().upper()
            if not cur or cur == 'CNY':
                continue
            raw = entry.get('rate', None)
            try:
                rate = float(raw) if raw not in (None, '', 0, '0') else None
            except (TypeError, ValueError):
                rate = None
            if rate is None or rate <= 0:
                conn.execute(
                    'DELETE FROM sales_board_exchange_rates WHERE year_month = ? AND currency = ?',
                    (month, cur)
                )
            else:
                conn.execute('''
                    INSERT INTO sales_board_exchange_rates (year_month, currency, rate_to_cny, updated_by)
                    VALUES (?, ?, ?, ?)
                    ON CONFLICT(year_month, currency) DO UPDATE SET
                        rate_to_cny = excluded.rate_to_cny,
                        updated_at = CURRENT_TIMESTAMP,
                        updated_by = excluded.updated_by
                ''', (month, cur, rate, getattr(current_user, 'username', '') or ''))
        conn.commit()
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500
    finally:
        conn.close()


# ----------------- Sales Board Export -----------------
def _get_sales_board_exports_dir():
    import os
    base = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'exports', 'sales_board')
    os.makedirs(base, exist_ok=True)
    return base


def _generate_sales_board_excel(data, hide_leader=False):
    """Generate a multi-sheet Excel workbook from sales board data.

    Sheets:
      1. 销售汇总 - Main per-person summary with totals
      2. 币种明细 - Per-person currency breakdown (gross/net/shipping/rate/CNY)
      3. 免提成产品 - Excluded-brand product revenue per person
      4. 小组奖金 - Group leader bonus calculation (only if not hide_leader)
      5. 规则说明 - Rules and notes
    """
    from io import BytesIO
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    board_data = data['board_data']
    team_totals = data['team_totals']
    group_summaries = data['group_summaries']
    leader_bonus_map = data['leader_bonus_map']
    selected_month = data['selected_month']
    prev_month = data['prev_month']
    no_commission_brands = data['no_commission_brands']

    leader_names = {g['leader_manager'] for g in group_summaries}

    # Currency → country classification
    CURRENCY_COUNTRY = {
        'PLN': '波兰', 'EUR': '欧元',
        'AUD': '澳洲',
        'AED': '阿联酋',
        'USD': '美元', 'GBP': '英镑', 'CNY': '中国',
    }
    def country_of(cur):
        return CURRENCY_COUNTRY.get((cur or '').upper(), '其他')

    def country_totals_for(d):
        """Return (pl_cny, au_cny, other_cny) net sales by country for one manager."""
        pl = au = other = 0.0
        for cur, amt in (d.get('month_currency_amounts') or {}).items():
            rate = (d.get('month_rates') or {}).get(cur, 0)
            if not rate: continue
            cny = amt.get('net_amount', 0) * rate
            c = country_of(cur)
            if c == '波兰': pl += cny
            elif c == '澳洲': au += cny
            else: other += cny
        return round(pl, 2), round(au, 2), round(other, 2)

    # ---------- Styles ----------
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill("solid", fgColor="305496")
    group_fill = PatternFill("solid", fgColor="D9E1F2")
    total_fill = PatternFill("solid", fgColor="FFF2CC")
    sub_fill = PatternFill("solid", fgColor="F2F2F2")
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    right = Alignment(horizontal="right", vertical="center")
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    thin = Side(border_style="thin", color="999999")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    money_fmt = '"¥"#,##0.00'
    pct_fmt = '0.0%'

    def write_header_row(ws, row, headers, col_widths=None):
        for i, h in enumerate(headers, 1):
            c = ws.cell(row=row, column=i, value=h)
            c.font = header_font; c.fill = header_fill; c.alignment = center; c.border = border
        ws.row_dimensions[row].height = 24
        if col_widths:
            for i, w in enumerate(col_widths, 1):
                ws.column_dimensions[get_column_letter(i)].width = w

    def set_default_row_height(ws, h=20):
        ws.sheet_format.defaultRowHeight = h
        ws.sheet_format.customHeight = True

    def enforce_row_heights(ws, min_h=20):
        """Explicitly set every used row's height to at least min_h.

        Excel viewers often ignore sheet_format.defaultRowHeight and fall back
        to their own default (~15). Setting an explicit height per row ensures
        the desired minimum is honored.
        """
        for r in range(1, (ws.max_row or 0) + 1):
            cur = ws.row_dimensions[r].height
            if cur is None or cur < min_h:
                ws.row_dimensions[r].height = min_h

    def autofit_columns(ws, fixed_widths=None, min_width=8, max_width=50, padding=2):
        """Auto-fit column widths based on cell content.

        Treats CJK characters as 2 units wide; ASCII as 1 unit.
        ``fixed_widths`` is a {column_letter: width} mapping that overrides auto-fit.
        """
        fixed_widths = fixed_widths or {}
        for col_idx in range(1, (ws.max_column or 0) + 1):
            letter = get_column_letter(col_idx)
            if letter in fixed_widths:
                ws.column_dimensions[letter].width = fixed_widths[letter]
                continue
            max_w = min_width
            for row_idx in range(1, (ws.max_row or 0) + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                v = cell.value
                if v is None:
                    continue
                # If this cell is part of a merged range that starts elsewhere, skip
                text = str(v)
                # Account for line breaks: take longest line
                line_widths = []
                for line in text.splitlines() or [text]:
                    line_widths.append(sum(2 if ord(c) > 127 else 1 for c in line))
                w = max(line_widths or [0])
                if w > max_w:
                    max_w = w
            ws.column_dimensions[letter].width = min(max(max_w + padding, min_width), max_width)

    def center_all_cells(ws, wrap_text=True):
        """Set every cell's alignment to horizontal+vertical center."""
        from openpyxl.styles import Alignment
        align = Alignment(horizontal="center", vertical="center", wrap_text=wrap_text)
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row or 0):
            for cell in row:
                cell.alignment = align

    wb = Workbook()

    # =============== Sheet 1: 销售汇总 ===============
    ws = wb.active
    ws.title = "销售汇总"
    set_default_row_height(ws)

    ws.merge_cells('A1:P1')
    ws['A1'] = f"销售看板 — {selected_month}" + ("（演示模式：已隐藏组长）" if hide_leader else "")
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = center
    ws.row_dimensions[1].height = 26

    # Compute team country totals
    team_pl = round(sum(country_totals_for(d)[0] for d in board_data), 2)
    team_au = round(sum(country_totals_for(d)[1] for d in board_data), 2)
    team_other = round(sum(country_totals_for(d)[2] for d in board_data), 2)

    # Top summary cards
    summary_pairs = [
        ("团队月目标", team_totals['monthly_target']),
        ("上月销售(¥)", team_totals['prev_net_cny']),
        ("团队月实际(¥)", team_totals['month_net_cny']),
        ("团队达成率", f"{team_totals['month_achievement']}%"),
        ("团队总产品数", team_totals['month_total_products']),
        ("波兰销售(¥)", team_pl),
        ("澳洲销售(¥)", team_au),
    ]
    col = 1
    for label, val in summary_pairs:
        lc = ws.cell(row=2, column=col, value=label)
        vc = ws.cell(row=2, column=col + 1, value=val)
        lc.font = Font(bold=True); lc.fill = sub_fill; lc.alignment = right
        vc.alignment = left
        if isinstance(val, (int, float)) and '¥' in label:
            vc.number_format = money_fmt
        col += 2
    ws.row_dimensions[2].height = 20

    headers = [
        "姓名", "网站", "月目标¥", "本月实际¥",
        "波兰¥", "澳洲¥", "其他¥",
        "达成率",
        "订单数", "产品数", "运费(不计¥)", "免提成(不计¥)",
        "提成基数¥", "提成率", "提成¥",
        "底薪¥", "扣除¥", "达标情况", "预计薪资¥",
    ]
    col_widths = [10, 36, 12, 14, 12, 12, 12, 10, 10, 10, 14, 16, 16, 10, 12, 12, 12, 22, 16]
    header_row = 4
    write_header_row(ws, header_row, headers, col_widths)

    r = header_row + 1
    for d in board_data:
        is_leader = d['manager'] in leader_names
        group_bonus = leader_bonus_map.get(d['manager'], 0)

        # Decide displayed salary/income
        if hide_leader and is_leader:
            base_salary_disp = d.get('masked_base_salary', d['base_salary'])
            total_income_disp = d.get('masked_total_income', d['total_income'])
        elif is_leader and group_bonus > 0:
            base_salary_disp = d['base_salary']
            total_income_disp = d['total_income'] + group_bonus
        else:
            base_salary_disp = d['base_salary']
            total_income_disp = d['total_income']

        if d['salary_protected']:
            reason_parts = []
            if d['growth_met']: reason_parts.append(f"环比+{d['growth_rate']}%")
            if d['units_met']: reason_parts.append(f"{d['month_total_products']}支")
            met_text = "达标（" + " / ".join(reason_parts) + "）"
        else:
            met_text = "未达标（扣底薪20%）"

        pl_cny, au_cny, other_cny = country_totals_for(d)
        row_vals = [
            d['manager'],
            ", ".join(d['sites']),
            d['monthly_target'],
            d['month_net_cny'],
            pl_cny,
            au_cny,
            other_cny,
            d['month_achievement'] / 100 if d['month_achievement'] else 0,
            d['order_count'],
            d['month_total_products'],
            d['month_shipping_cny'],
            d['month_excluded_cny'],
            d['commission_base_cny'],
            d['commission_rate'],
            d['commission'],
            base_salary_disp,
            -d['salary_deduction'] if d['salary_deduction'] > 0 else 0,
            met_text,
            total_income_disp,
        ]
        # column index map (1-based):
        # 1姓名 2网站 3月目标 4本月实际 5波兰 6澳洲 7其他 8达成率
        # 9订单数 10产品数 11运费 12免提成 13提成基数 14提成率 15提成
        # 16底薪 17扣除 18达标情况 19预计薪资
        money_cols = {3, 4, 5, 6, 7, 11, 12, 13, 15, 16, 17, 19}
        pct_cols = {8, 14}
        center_cols = {9, 10}
        for c_idx, val in enumerate(row_vals, 1):
            cell = ws.cell(row=r, column=c_idx, value=val)
            cell.border = border
            if c_idx == 1:
                cell.font = Font(bold=True); cell.alignment = center
            elif c_idx == 2:
                cell.alignment = left; cell.font = Font(size=9)
            elif c_idx in money_cols:
                cell.alignment = right; cell.number_format = money_fmt
            elif c_idx in pct_cols:
                cell.alignment = center; cell.number_format = pct_fmt
            elif c_idx in center_cols:
                cell.alignment = center
            elif c_idx == 18:  # 达标情况
                cell.alignment = center
                if d['salary_protected']:
                    cell.font = Font(color="008000")
                else:
                    cell.font = Font(color="C00000")
        # Tint country cells subtly to help scanning
        if pl_cny > 0:
            ws.cell(row=r, column=5).fill = PatternFill("solid", fgColor="FFF4E5")  # PL 浅橙
        if au_cny > 0:
            ws.cell(row=r, column=6).fill = PatternFill("solid", fgColor="E8F4FC")  # AU 浅蓝
        r += 1

    # Group summaries (only if not hide_leader)
    if not hide_leader and group_summaries:
        # Blank separator
        r += 1
        ws.cell(row=r, column=1, value="— 小组汇总 —").font = Font(bold=True, italic=True, color="305496")
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=19)
        r += 1
        board_by_mgr = {d['manager']: d for d in board_data}
        for g in group_summaries:
            # Group-level country split
            g_pl = g_au = g_other = 0.0
            for m in g.get('members', []):
                md = board_by_mgr.get(m)
                if md:
                    a, b, c = country_totals_for(md)
                    g_pl += a; g_au += b; g_other += c
            row_vals = [
                g['name'] + f"（组长：{g['leader_manager']}）", "",
                g['monthly_target'], g['month_net_cny'],
                round(g_pl, 2), round(g_au, 2), round(g_other, 2),
                g['month_achievement'] / 100 if g['month_achievement'] else 0,
                "", g['month_total_products'], "", "",
                g['bonus_base_cny'], g['bonus_rate'], g['leader_bonus'],
                "", "", f"带团奖金 = 成员基数 × {g['bonus_rate']*100:.1f}%", "",
            ]
            money_cols_g = {3, 4, 5, 6, 7, 13, 15}
            pct_cols_g = {8, 14}
            for c_idx, val in enumerate(row_vals, 1):
                cell = ws.cell(row=r, column=c_idx, value=val)
                cell.border = border; cell.fill = group_fill
                if c_idx in money_cols_g:
                    cell.alignment = right; cell.number_format = money_fmt
                elif c_idx in pct_cols_g:
                    cell.alignment = center; cell.number_format = pct_fmt
                elif c_idx == 1:
                    cell.font = Font(bold=True); cell.alignment = left
                elif c_idx == 18:
                    cell.alignment = left; cell.font = Font(size=9, color="595959")
                else:
                    cell.alignment = center
            r += 1

    # Team total
    r += 1
    total_income = team_totals['total_income_no_bonus'] if hide_leader else team_totals['total_income']
    total_row = [
        "团队合计", "",
        team_totals['monthly_target'], team_totals['month_net_cny'],
        team_pl, team_au, team_other,
        team_totals['month_achievement'] / 100 if team_totals['month_achievement'] else 0,
        sum(d['order_count'] for d in board_data),
        team_totals['month_total_products'],
        round(sum(d['month_shipping_cny'] for d in board_data), 2),
        round(sum(d['month_excluded_cny'] for d in board_data), 2),
        round(sum(d['commission_base_cny'] for d in board_data), 2),
        "", team_totals['total_commission'],
        "", "", "", total_income,
    ]
    money_cols_t = {3, 4, 5, 6, 7, 11, 12, 13, 15, 19}
    pct_cols_t = {8}
    for c_idx, val in enumerate(total_row, 1):
        cell = ws.cell(row=r, column=c_idx, value=val)
        cell.border = border; cell.fill = total_fill; cell.font = Font(bold=True)
        if c_idx in money_cols_t:
            cell.alignment = right; cell.number_format = money_fmt
        elif c_idx in pct_cols_t:
            cell.alignment = center; cell.number_format = pct_fmt
        else:
            cell.alignment = center

    ws.freeze_panes = 'C5'

    # Finalize summary sheet styling (per request):
    #  - All cells centered (horizontal + vertical), wrap text
    #  - Column B (网站) fixed at width 30
    #  - Other columns auto-fit by content
    #  - Minimum row height = 50
    center_all_cells(ws, wrap_text=True)
    autofit_columns(ws, fixed_widths={'B': 30}, min_width=8, max_width=60, padding=3)
    enforce_row_heights(ws, min_h=50)

    # =============== Sheet 2: 币种明细 ===============
    ws2 = wb.create_sheet("币种明细")
    set_default_row_height(ws2)
    ws2.merge_cells('A1:H1')
    ws2['A1'] = f"每人每币种销售明细 — {selected_month}"
    ws2['A1'].font = Font(bold=True, size=13); ws2['A1'].alignment = center
    ws2.row_dimensions[1].height = 24

    note = ws2.cell(row=2, column=1, value="说明：净额 = 订单总额 − 运费；运费不计入销售额，也不计入提成基数。")
    note.font = Font(italic=True, color="808080", size=9); note.alignment = left
    ws2.merge_cells('A2:H2')

    hdr2 = ["姓名", "国家/地区", "币种", "订单毛额", "运费", "净额", "汇率", "净额折¥", "运费折¥"]
    widths2 = [10, 10, 8, 14, 12, 14, 10, 14, 14]
    write_header_row(ws2, 4, hdr2, widths2)

    # Country-colored fills
    pl_fill = PatternFill("solid", fgColor="FFF4E5")
    au_fill = PatternFill("solid", fgColor="E8F4FC")

    r = 5
    for d in board_data:
        # Sort currencies so PLN → AUD → others cluster by country
        def _cur_order(cur):
            c = country_of(cur)
            return (0 if c == '波兰' else 1 if c == '澳洲' else 2, cur)
        currencies = sorted(d['month_currency_amounts'].keys(), key=_cur_order)
        if not currencies:
            continue
        first = True
        subtotal_net_cny = 0
        subtotal_ship_cny = 0
        for cur in currencies:
            amt = d['month_currency_amounts'][cur]
            rate = d['month_rates'].get(cur, 0)
            net_cny = amt['net_amount'] * rate if rate else 0
            ship_cny = amt['shipping'] * rate if rate else 0
            subtotal_net_cny += net_cny
            subtotal_ship_cny += ship_cny
            ctry = country_of(cur)
            row_vals = [
                d['manager'] if first else "",
                ctry,
                cur,
                round(amt['amount'], 2),
                round(amt['shipping'], 2),
                round(amt['net_amount'], 2),
                round(rate, 4) if rate else "无汇率",
                round(net_cny, 2) if rate else "",
                round(ship_cny, 2) if rate else "",
            ]
            row_fill = pl_fill if ctry == '波兰' else au_fill if ctry == '澳洲' else None
            for c_idx, val in enumerate(row_vals, 1):
                cell = ws2.cell(row=r, column=c_idx, value=val)
                cell.border = border
                if row_fill and c_idx in (2, 3):
                    cell.fill = row_fill
                if c_idx == 1:
                    cell.font = Font(bold=True); cell.alignment = center
                elif c_idx == 2:
                    cell.font = Font(bold=True, color="305496")
                    cell.alignment = center
                elif c_idx == 3:
                    cell.alignment = center
                elif c_idx in (4, 5, 6):
                    cell.alignment = right; cell.number_format = '#,##0.00'
                elif c_idx == 7:
                    cell.alignment = center; cell.number_format = '0.0000'
                else:
                    cell.alignment = right; cell.number_format = money_fmt
            r += 1
            first = False

        # Subtotal per person
        sub_row = ["", "", "合计¥", "", "", "", "", round(subtotal_net_cny, 2), round(subtotal_ship_cny, 2)]
        for c_idx, val in enumerate(sub_row, 1):
            cell = ws2.cell(row=r, column=c_idx, value=val)
            cell.border = border; cell.fill = sub_fill; cell.font = Font(bold=True)
            if c_idx in (8, 9): cell.alignment = right; cell.number_format = money_fmt
            else: cell.alignment = center
        r += 1

    ws2.freeze_panes = 'A5'

    # =============== Sheet 3: 免提成产品 ===============
    ws3 = wb.create_sheet("免提成产品")
    set_default_row_height(ws3)
    ws3.merge_cells('A1:F1')
    ws3['A1'] = f"免提成品牌产品明细 — {selected_month}"
    ws3['A1'].font = Font(bold=True, size=13); ws3['A1'].alignment = center
    ws3.row_dimensions[1].height = 24

    note = ws3.cell(row=2, column=1, value=f"免提成品牌：{', '.join(no_commission_brands) or '（无）'}。这些品牌的销售额不计入提成基数。")
    note.font = Font(italic=True, color="808080", size=9); note.alignment = left
    ws3.merge_cells('A2:F2')

    hdr3 = ["姓名", "品牌", "数量", "国家/地区", "币种", "原币金额", "折¥金额"]
    widths3 = [10, 14, 10, 10, 10, 16, 16]
    write_header_row(ws3, 4, hdr3, widths3)

    r = 5
    any_excluded = False
    for d in board_data:
        breakdown = d.get('excluded_brand_breakdown') or {}
        if not breakdown:
            continue
        any_excluded = True
        first = True
        person_total_cny = 0
        for brand in sorted(breakdown.keys()):
            info = breakdown[brand]
            person_total_cny += info['revenue_cny']
            def _cur_order2(cur):
                c = country_of(cur)
                return (0 if c == '波兰' else 1 if c == '澳洲' else 2, cur)
            curs = sorted(info['revenue_by_currency'].keys(), key=_cur_order2)
            brand_first = True
            for cur in curs:
                ctry = country_of(cur)
                row_vals = [
                    d['manager'] if first else "",
                    brand if brand_first else "",
                    info['qty'] if brand_first else "",
                    ctry,
                    cur,
                    info['revenue_by_currency'][cur],
                    info['revenue_cny'] if brand_first else "",
                ]
                row_fill = pl_fill if ctry == '波兰' else au_fill if ctry == '澳洲' else None
                for c_idx, val in enumerate(row_vals, 1):
                    cell = ws3.cell(row=r, column=c_idx, value=val)
                    cell.border = border
                    if row_fill and c_idx in (4, 5):
                        cell.fill = row_fill
                    if c_idx == 1:
                        cell.font = Font(bold=True); cell.alignment = center
                    elif c_idx == 2:
                        cell.font = Font(bold=True); cell.alignment = center
                    elif c_idx == 4:
                        cell.font = Font(bold=True, color="305496"); cell.alignment = center
                    elif c_idx == 6:
                        cell.alignment = right; cell.number_format = '#,##0.00'
                    elif c_idx == 7:
                        cell.alignment = right; cell.number_format = money_fmt
                    else:
                        cell.alignment = center
                r += 1
                first = False
                brand_first = False
        # Per-person subtotal
        sub_row = ["", "合计¥", "", "", "", "", round(person_total_cny, 2)]
        for c_idx, val in enumerate(sub_row, 1):
            cell = ws3.cell(row=r, column=c_idx, value=val)
            cell.border = border; cell.fill = sub_fill; cell.font = Font(bold=True)
            if c_idx == 7: cell.alignment = right; cell.number_format = money_fmt
            else: cell.alignment = center
        r += 1

    if not any_excluded:
        ws3.cell(row=5, column=1, value="本月未记录到免提成品牌销售。").alignment = left
        ws3.merge_cells(start_row=5, start_column=1, end_row=5, end_column=7)

    ws3.freeze_panes = 'A5'

    # =============== Sheet 4: 小组奖金 ===============
    if not hide_leader and group_summaries:
        ws4 = wb.create_sheet("小组奖金")
        set_default_row_height(ws4)
        ws4.merge_cells('A1:E1')
        ws4['A1'] = f"小组带团奖金计算明细 — {selected_month}"
        ws4['A1'].font = Font(bold=True, size=13); ws4['A1'].alignment = center
        ws4.row_dimensions[1].height = 24

        note = ws4.cell(row=2, column=1,
            value="说明：组长的带团奖金 = 组内非组长成员的【提成基数（¥）】之和 × 带团奖金比例。")
        note.font = Font(italic=True, color="808080", size=9); note.alignment = left
        ws4.merge_cells('A2:E2')

        hdr4 = ["小组", "成员", "角色", "提成基数¥", "计入奖金基数"]
        widths4 = [14, 12, 14, 16, 14]
        write_header_row(ws4, 4, hdr4, widths4)

        board_by_mgr = {d['manager']: d for d in board_data}
        r = 5
        for g in group_summaries:
            members = g['members'] or []
            leader = g['leader_manager']
            bonus_rate = g['bonus_rate']
            first = True
            for m in members:
                mdata = board_by_mgr.get(m)
                base = mdata['commission_base_cny'] if mdata else 0
                is_grp_leader = (m == leader)
                row_vals = [
                    g['name'] if first else "",
                    m,
                    "组长" if is_grp_leader else "组员",
                    base,
                    "否（组长不计入）" if is_grp_leader else "是",
                ]
                for c_idx, val in enumerate(row_vals, 1):
                    cell = ws4.cell(row=r, column=c_idx, value=val)
                    cell.border = border
                    if c_idx == 4:
                        cell.alignment = right; cell.number_format = money_fmt
                    elif c_idx == 3:
                        cell.alignment = center
                        if is_grp_leader:
                            cell.font = Font(bold=True, color="305496")
                    elif c_idx == 1:
                        cell.font = Font(bold=True); cell.alignment = center
                    else:
                        cell.alignment = center
                r += 1
                first = False
            # Summary row for this group
            for c_idx, val in enumerate([
                f"{g['name']} 奖金合计",
                f"组长：{leader}",
                f"比例 {bonus_rate*100:.1f}%",
                g['bonus_base_cny'],
                g['leader_bonus'],
            ], 1):
                cell = ws4.cell(row=r, column=c_idx, value=val)
                cell.border = border; cell.fill = total_fill; cell.font = Font(bold=True)
                if c_idx in (4, 5):
                    cell.alignment = right; cell.number_format = money_fmt
                else:
                    cell.alignment = center
            ws4.cell(row=r, column=5).font = Font(bold=True, color="C00000")
            r += 2

        ws4.freeze_panes = 'A5'

    # =============== Sheet 5: 规则说明 ===============
    ws5 = wb.create_sheet("规则说明")
    set_default_row_height(ws5)
    ws5.column_dimensions['A'].width = 100

    rates_in_use = data.get('rates_in_use') or {}
    rate_lines = []
    if rates_in_use:
        for cur in sorted(rates_in_use.keys()):
            info = rates_in_use[cur]
            tag = "（自定义）" if info.get('source') == 'override' else "（系统）"
            rate_lines.append(f"    - {cur} → 1 {cur} = ¥{info['rate']:.4f} {tag}")

    lines = [
        (f"销售看板 — {selected_month}", True, 14),
        ("", False, 11),
        ("【统计范围】", True, 12),
        (f"• 当月订单：{selected_month} 内状态为成功的订单（排除 failed / cancelled / checkout-draft / trash / cheat；排除未付款的 pending、未到账的 on-hold bacs）。", False, 10),
        (f"• 环比对比月份：{prev_month}", False, 10),
        ("", False, 10),
        ("【本月使用汇率】", True, 12),
        *([(line, False, 10) for line in rate_lines] if rate_lines else [("• （无）", False, 10)]),
        ("• 自定义汇率仅作用于销售看板和本导出文件，不影响系统其它模块。", False, 10),
        ("", False, 10),
        ("【底薪保护（满足其一即可）】", True, 12),
        ("• 环比增长 ≥ 20%", False, 10),
        ("• 当月销售 ≥ 500 支", False, 10),
        ("• 未达标则扣除底薪的 20%。", False, 10),
        ("", False, 10),
        ("【提成规则】", True, 12),
        ("• 提成 = 提成基数(¥) × 提成率（在「设置目标」里按人配置）", False, 10),
        ("• 提成基数 = 成功订单中 非免提成品牌 的产品小计（item.total），按本月汇率换算为人民币。", False, 10),
        ("• 不计入提成：", False, 10),
        ("    ① 运费（shipping_total）", False, 10),
        (f"    ② 免提成品牌的产品收入（当前名单：{', '.join(no_commission_brands) or '无'}）", False, 10),
        ("", False, 10),
        ("【小组带团奖金】", True, 12),
        ("• 组长带团奖金 = 组内非组长成员的提成基数(¥) 之和 × 带团奖金比例", False, 10),
        ("• 组长个人仍按其自己的提成基数和提成率计算个人提成，与带团奖金独立。", False, 10),
        ("", False, 10),
        ("【演示/完整版说明】", True, 12),
        ("• 完整版：包含小组汇总和组长带团奖金。", False, 10),
        ("• 演示版：不包含「小组奖金」Sheet；汇总页中组长的底薪按标准 7000 展示，预计薪资也按此重新计算；团队合计的预计薪资会相应扣减，避免从合计倒推出带团奖金。", False, 10),
        ("• 两份数字的差额 = 带团奖金之和。", False, 10),
    ]
    for i, (text, bold, size) in enumerate(lines, 1):
        c = ws5.cell(row=i, column=1, value=text)
        c.font = Font(bold=bold, size=size)
        c.alignment = left
        if bold and size >= 12:
            c.fill = sub_fill
        ws5.row_dimensions[i].height = 18 if size >= 12 else 16

    # Force minimum row height on every used row in every sheet
    for _sheet in wb.worksheets:
        enforce_row_heights(_sheet, min_h=20)

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio


@app.route('/api/sales-board/export', methods=['POST'])
@login_required
@super_admin_required
def export_sales_board():
    """Generate an Excel export for the sales board and save to history."""
    import datetime, os
    payload = request.get_json(silent=True) or {}
    month = payload.get('month') or datetime.date.today().strftime('%Y-%m')
    hide_leader = bool(payload.get('hide_leader', False))

    try:
        data = _compute_sales_board_data(month)
        bio = _generate_sales_board_excel(data, hide_leader=hide_leader)

        ts = datetime.datetime.now().strftime('%Y%m%d_%H%M%S')
        suffix = '_演示版' if hide_leader else ''
        filename = f"销售看板_{month}{suffix}_{ts}.xlsx"
        file_path = os.path.join(_get_sales_board_exports_dir(), filename)
        with open(file_path, 'wb') as f:
            f.write(bio.getvalue())
        file_size = os.path.getsize(file_path)

        conn = get_db_connection()
        cur = conn.execute(
            '''INSERT INTO sales_board_exports
               (year_month, filename, file_path, file_size, hide_leader, created_by)
               VALUES (?, ?, ?, ?, ?, ?)''',
            (month, filename, file_path, file_size, 1 if hide_leader else 0,
             getattr(current_user, 'username', '') or '')
        )
        conn.commit()
        export_id = cur.lastrowid
        conn.close()

        return jsonify({
            'success': True,
            'id': export_id,
            'filename': filename,
            'file_size': file_size,
        })
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/sales-board/exports', methods=['GET'])
@login_required
@super_admin_required
def list_sales_board_exports():
    """List export history (optionally filter by month)."""
    month = request.args.get('month', '').strip()
    conn = get_db_connection()
    try:
        if month:
            rows = conn.execute(
                '''SELECT id, year_month, filename, file_size, hide_leader, created_by, created_at
                   FROM sales_board_exports WHERE year_month = ?
                   ORDER BY created_at DESC''',
                (month,)
            ).fetchall()
        else:
            rows = conn.execute(
                '''SELECT id, year_month, filename, file_size, hide_leader, created_by, created_at
                   FROM sales_board_exports ORDER BY created_at DESC LIMIT 200'''
            ).fetchall()
        return jsonify([dict(r) for r in rows])
    finally:
        conn.close()


@app.route('/api/sales-board/exports/<int:export_id>/download', methods=['GET'])
@login_required
@super_admin_required
def download_sales_board_export(export_id):
    """Download a specific export file."""
    import os
    from flask import send_file
    conn = get_db_connection()
    try:
        row = conn.execute(
            'SELECT filename, file_path FROM sales_board_exports WHERE id = ?',
            (export_id,)
        ).fetchone()
    finally:
        conn.close()
    if not row:
        return jsonify({'error': '导出记录不存在'}), 404
    if not os.path.exists(row['file_path']):
        return jsonify({'error': '文件不存在或已被删除'}), 404
    return send_file(
        row['file_path'],
        as_attachment=True,
        download_name=row['filename'],
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )


@app.route('/api/sales-board/exports/<int:export_id>', methods=['DELETE'])
@login_required
@super_admin_required
def delete_sales_board_export(export_id):
    """Delete an export (both file and DB record)."""
    import os
    conn = get_db_connection()
    try:
        row = conn.execute(
            'SELECT file_path FROM sales_board_exports WHERE id = ?',
            (export_id,)
        ).fetchone()
        if not row:
            return jsonify({'error': '导出记录不存在'}), 404
        try:
            if os.path.exists(row['file_path']):
                os.remove(row['file_path'])
        except Exception:
            pass
        conn.execute('DELETE FROM sales_board_exports WHERE id = ?', (export_id,))
        conn.commit()
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500
    finally:
        conn.close()


# ===== Product Costs Management APIs =====

@app.route('/product-costs')
@login_required
@costs_view_required
def product_costs_page():
    """Product costs management page"""
    conn = get_db_connection()
    countries = conn.execute('SELECT DISTINCT country FROM sites WHERE country IS NOT NULL AND country != "" ORDER BY country').fetchall()
    countries = [r['country'] for r in countries]
    warehouses = [dict(r) for r in conn.execute('SELECT * FROM warehouses WHERE is_active=1 ORDER BY country, name').fetchall()]
    conn.close()
    # Pass scope info. Two distinct scopes:
    #   allowed_warehouse_ids → EDIT scope: which rows the user can edit/delete.
    #   view_warehouse_ids     → VIEW scope: which rows/countries the user can see.
    # A read-only viewer (can_view_costs, no partner binding) has edit scope []
    # but view scope None (sees all costs). None = unrestricted.
    allowed_wh = _user_allowed_warehouse_ids()
    view_wh = _user_allowed_warehouse_ids(for_view=True)
    return render_template('product_costs.html',
                           countries=countries,
                           warehouses=warehouses,
                           can_edit_costs=current_user.can_edit_costs(),
                           allowed_warehouse_ids=allowed_wh,
                           view_warehouse_ids=view_wh)


@app.route('/api/product-costs/options', methods=['GET'])
@login_required
@costs_view_required
def get_product_cost_options():
    """Get brand-puffs-flavor combinations for cascading selectors"""
    conn = get_db_connection()
    try:
        rows = conn.execute('''
            SELECT pm.brand_id, b.name as brand_name, pm.puff_count, pm.flavor
            FROM product_mappings pm
            LEFT JOIN brands b ON pm.brand_id = b.id
            WHERE pm.brand_id IS NOT NULL
        ''').fetchall()
        combos = []
        for r in rows:
            combos.append({
                'brand_id': r['brand_id'],
                'brand_name': r['brand_name'] or '',
                'puff_count': r['puff_count'],
                'flavor': r['flavor'] or ''
            })
        return jsonify(combos)
    finally:
        conn.close()


@app.route('/api/product-costs', methods=['GET'])
@login_required
@costs_view_required
def get_product_costs():
    """Get product costs.

    By default, returns the LATEST cost per (brand, series, puff, flavor,
    warehouse) tuple — that's the price effective right now. Each row also
    carries `history_count` so the UI can show "N 个历史价格" badge.

    Pass ?include_history=1 to get every row instead — useful for the
    "import from CSV" / "raw view" workflow.
    """
    include_history = request.args.get('include_history') in ('1', 'true', 'yes')
    conn = get_db_connection()
    try:
        if include_history:
            rows = conn.execute('''
                SELECT pc.*, b.name as brand_name,
                       s.name as series_name,
                       w.name as warehouse_name, w.code as warehouse_code,
                       1 as history_count
                FROM product_costs pc
                LEFT JOIN brands b ON pc.brand_id = b.id
                LEFT JOIN series s ON pc.series_id = s.id
                LEFT JOIN warehouses w ON pc.warehouse_id = w.id
                ORDER BY b.name, s.name, pc.puff_count, pc.effective_date DESC
            ''').fetchall()
            return jsonify([dict(r) for r in rows])

        # Latest per product-key (using window-function style via subquery —
        # SQLite supports ROW_NUMBER() since 3.25, but we keep it simple).
        rows = conn.execute('''
            WITH ranked AS (
                SELECT
                    pc.*,
                    b.name AS brand_name,
                    s.name AS series_name,
                    w.name AS warehouse_name, w.code AS warehouse_code,
                    ROW_NUMBER() OVER (
                        PARTITION BY pc.brand_id, COALESCE(pc.series_id, 0), COALESCE(pc.puff_count, 0),
                                     COALESCE(pc.flavor, ''), COALESCE(pc.warehouse_id, 0)
                        ORDER BY pc.effective_date DESC, pc.id DESC
                    ) AS rn,
                    COUNT(*) OVER (
                        PARTITION BY pc.brand_id, COALESCE(pc.series_id, 0), COALESCE(pc.puff_count, 0),
                                     COALESCE(pc.flavor, ''), COALESCE(pc.warehouse_id, 0)
                    ) AS history_count
                FROM product_costs pc
                LEFT JOIN brands b ON pc.brand_id = b.id
                LEFT JOIN series s ON pc.series_id = s.id
                LEFT JOIN warehouses w ON pc.warehouse_id = w.id
            )
            SELECT * FROM ranked WHERE rn = 1
            ORDER BY brand_name, series_name, puff_count
        ''').fetchall()
        result = []
        for r in rows:
            d = dict(r)
            d.pop('rn', None)
            result.append(d)
        return jsonify(result)
    finally:
        conn.close()


@app.route('/api/product-costs/history', methods=['GET'])
@login_required
@costs_view_required
def get_product_cost_history():
    """List ALL price versions for a single product (brand+series+puffs+flavor+warehouse).

    Query params: brand_id (required), series_id, puff_count, flavor, warehouse_id (required)
    Returns rows sorted by effective_date DESC.
    """
    brand_id = request.args.get('brand_id', type=int)
    warehouse_id = request.args.get('warehouse_id', type=int)
    if not brand_id or not warehouse_id:
        return jsonify({'error': 'brand_id 和 warehouse_id 必填'}), 400

    series_id = request.args.get('series_id', type=int)
    puff_count = request.args.get('puff_count', type=int)
    flavor = request.args.get('flavor', '').strip() or None

    conn = get_db_connection()
    try:
        sql = '''SELECT pc.*, b.name as brand_name, s.name as series_name,
                        w.name as warehouse_name, w.code as warehouse_code
                 FROM product_costs pc
                 LEFT JOIN brands b ON pc.brand_id = b.id
                 LEFT JOIN series s ON pc.series_id = s.id
                 LEFT JOIN warehouses w ON pc.warehouse_id = w.id
                 WHERE pc.brand_id = ? AND pc.warehouse_id = ?'''
        params = [brand_id, warehouse_id]
        if series_id is None:
            sql += ' AND pc.series_id IS NULL'
        else:
            sql += ' AND pc.series_id = ?'
            params.append(series_id)
        if puff_count is None:
            sql += ' AND pc.puff_count IS NULL'
        else:
            sql += ' AND pc.puff_count = ?'
            params.append(puff_count)
        if flavor is None:
            sql += " AND (pc.flavor IS NULL OR pc.flavor = '')"
        else:
            sql += ' AND pc.flavor = ?'
            params.append(flavor)
        sql += ' ORDER BY pc.effective_date DESC, pc.id DESC'
        rows = conn.execute(sql, params).fetchall()
        return jsonify([dict(r) for r in rows])
    finally:
        conn.close()


@app.route('/api/product-costs', methods=['POST'])
@login_required
@costs_edit_required
def create_product_cost():
    """Create a new product cost entry.

    Multiple rows for the same (brand, series, puff, flavor, warehouse) tuple
    are allowed AS LONG AS they have different effective_date — that's how a
    cost change over time is recorded (e.g. Feb @ 28 PLN, Mar @ 30 PLN).
    Same product + same date → conflict (uniqueness violation).
    """
    data = request.get_json()
    brand_id = data.get('brand_id')
    series_id = data.get('series_id') or None
    puff_count = data.get('puff_count') or None
    flavor = data.get('flavor') or None
    warehouse_id = data.get('warehouse_id')
    cost_price = data.get('cost_price')
    cost_currency = data.get('cost_currency', 'PLN')
    effective_date = data.get('effective_date', '2024-01-01')
    notes = data.get('notes', '')

    if not brand_id or cost_price is None or not warehouse_id:
        return jsonify({'error': '品牌、仓库和成本价为必填项'}), 400

    # Scope check — partner-bound users may only edit their own country's
    # warehouses. Super admin and admin-role users bypass this.
    if not _check_warehouse_scope(warehouse_id):
        return jsonify({'error': '无权管理此仓库的成本（仅限您所属合伙人国家的仓库）',
                        'permission': 'warehouse_scope'}), 403

    conn = get_db_connection()
    try:
        # Get country from warehouse for backward compat
        wh = conn.execute('SELECT country FROM warehouses WHERE id=?', (warehouse_id,)).fetchone()
        country = wh['country'] if wh else 'PL'
        try:
            conn.execute('''
                INSERT INTO product_costs
                (brand_id, series_id, puff_count, flavor, country, warehouse_id, cost_price, cost_currency, effective_date, notes)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            ''', (brand_id, series_id, puff_count, flavor, country, warehouse_id, cost_price, cost_currency, effective_date, notes))
        except sqlite3.IntegrityError:
            # Same product + same effective_date already exists. Tell the user
            # to either pick a different date or edit the existing row.
            return jsonify({
                'error': '该产品在此生效日期已有成本记录，如需修改请编辑现有记录，或选择不同的生效日期来新增一个版本。'
            }), 409
        conn.commit()
        return jsonify({'success': True, 'id': conn.execute('SELECT last_insert_rowid()').fetchone()[0]})
    except Exception as e:
        return jsonify({'error': str(e)}), 500
    finally:
        conn.close()


@app.route('/api/product-costs/<int:cost_id>', methods=['PUT'])
@login_required
@costs_edit_required
def update_product_cost(cost_id):
    """Update a product cost entry"""
    data = request.get_json()
    conn = get_db_connection()
    try:
        existing = conn.execute('SELECT id, warehouse_id FROM product_costs WHERE id = ?', (cost_id,)).fetchone()
        if not existing:
            return jsonify({'error': '记录不存在'}), 404

        # Scope check on BOTH the existing warehouse (otherwise a partner could
        # "edit" foreign rows) AND the target warehouse (otherwise they could
        # move a row out of their scope).
        if not _check_warehouse_scope(existing['warehouse_id']):
            return jsonify({'error': '无权修改此仓库的成本',
                            'permission': 'warehouse_scope'}), 403

        wh_id = data.get('warehouse_id')
        if wh_id and not _check_warehouse_scope(wh_id):
            return jsonify({'error': '无权将成本移到该仓库',
                            'permission': 'warehouse_scope'}), 403
        wh = conn.execute('SELECT country FROM warehouses WHERE id=?', (wh_id,)).fetchone() if wh_id else None
        country = wh['country'] if wh else data.get('country', 'PL')
        conn.execute('''
            UPDATE product_costs SET
                brand_id = ?, series_id = ?, puff_count = ?, flavor = ?,
                country = ?, warehouse_id = ?, cost_price = ?, cost_currency = ?,
                effective_date = ?, notes = ?, updated_at = CURRENT_TIMESTAMP
            WHERE id = ?
        ''', (
            data.get('brand_id'), data.get('series_id') or None,
            data.get('puff_count') or None, data.get('flavor') or None,
            country, wh_id, data.get('cost_price'),
            data.get('cost_currency', 'PLN'), data.get('effective_date', '2024-01-01'),
            data.get('notes', ''), cost_id
        ))
        conn.commit()
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'error': str(e)}), 500
    finally:
        conn.close()


@app.route('/api/product-costs/<int:cost_id>', methods=['DELETE'])
@login_required
@costs_edit_required
def delete_product_cost(cost_id):
    """Delete a product cost entry"""
    conn = get_db_connection()
    try:
        existing = conn.execute('SELECT warehouse_id FROM product_costs WHERE id = ?', (cost_id,)).fetchone()
        if not existing:
            return jsonify({'error': '记录不存在'}), 404
        if not _check_warehouse_scope(existing['warehouse_id']):
            return jsonify({'error': '无权删除此仓库的成本',
                            'permission': 'warehouse_scope'}), 403
        conn.execute('DELETE FROM product_costs WHERE id = ?', (cost_id,))
        conn.commit()
        return jsonify({'success': True})
    finally:
        conn.close()


@app.route('/api/product-costs/unmapped', methods=['GET'])
@login_required
@costs_view_required
def get_unmapped_products():
    """Get products sold in a given month that have no cost mapping"""
    import datetime
    year_month = request.args.get('year_month')
    country = request.args.get('country', 'PL')
    match_level = request.args.get('match_level', 'brand_puffs_series')
    if not year_month:
        year_month = datetime.date.today().strftime('%Y-%m')

    conn = get_db_connection()
    try:
        country_sites = conn.execute('SELECT url FROM sites WHERE country = ?', (country,)).fetchall()
        country_urls = [s['url'] for s in country_sites]
        if not country_urls:
            return jsonify([])

        placeholders = ', '.join(['?' for _ in country_urls])
        orders = conn.execute(f'''
            SELECT line_items, source FROM orders
            WHERE source IN ({placeholders})
            AND strftime('%Y-%m', date_created) = ?
            AND {_revenue_status_cond()}
        ''', country_urls + [year_month]).fetchall()

        brands_rows = conn.execute('SELECT id, name, aliases FROM brands').fetchall()
        brands_cache = []
        brand_names = {}
        for row in brands_rows:
            brand_name = row['name']
            brand_names[row['id']] = brand_name
            try:
                aliases = json.loads(row['aliases']) if row['aliases'] else []
            except:
                aliases = []
            brands_cache.append({
                'id': row['id'], 'name': brand_name, 'aliases': aliases,
                'patterns': [brand_name.upper()] + [a.upper() for a in aliases]
            })

        series_rows = conn.execute('SELECT id, name FROM series').fetchall()
        series_names = {r['id']: r['name'] for r in series_rows}

        country_wh_ids = [r['id'] for r in conn.execute('SELECT id FROM warehouses WHERE country=?', (country,)).fetchall()]
        if country_wh_ids:
            wh_placeholders = ', '.join(['?' for _ in country_wh_ids])
            costs = conn.execute(f'''
                SELECT brand_id, series_id, puff_count, flavor FROM product_costs WHERE warehouse_id IN ({wh_placeholders})
            ''', country_wh_ids).fetchall()
        else:
            costs = conn.execute('SELECT brand_id, series_id, puff_count, flavor FROM product_costs WHERE country=?', (country,)).fetchall()

        cost_keys_full = set()
        cost_keys_bps = set()
        cost_keys_bp = set()
        for c in costs:
            cost_keys_full.add((c['brand_id'], c['series_id'], c['puff_count'], c['flavor']))
            cost_keys_bps.add((c['brand_id'], c['series_id'], c['puff_count']))
            cost_keys_bp.add((c['brand_id'], c['puff_count']))

        product_mappings = {}
        pm_rows = conn.execute('SELECT raw_name, source, brand_id, series_id, puff_count, flavor FROM product_mappings').fetchall()
        for pm in pm_rows:
            product_mappings[(normalize_raw_name(pm['raw_name']), pm['source'])] = pm

        unmapped = {}
        for order in orders:
            items = parse_json_field(order['line_items'])
            if not isinstance(items, list):
                continue
            for item in items:
                name = item.get('name', '') or ''
                qty = item.get('quantity', 0) or 0
                item_total = float(item.get('total', 0) or 0)
                source = order['source']
                name_key = normalize_raw_name(name)

                pm = product_mappings.get((name_key, source)) or product_mappings.get((name_key, '')) or product_mappings.get((name_key, None))
                if pm:
                    brand_id = pm['brand_id']
                    series_id = pm['series_id']
                    puffs = pm['puff_count']
                    flav = pm['flavor']
                else:
                    parsed = parse_product_name(name, brands_cache)
                    brand_id = None
                    if parsed.get('brand'):
                        for bc in brands_cache:
                            if bc['name'].upper() == parsed['brand'].upper():
                                brand_id = bc['id']
                                break
                    series_id = None
                    puffs = parsed.get('puffs')
                    flav = parsed.get('flavor')

                if match_level == 'brand_puffs':
                    matched = brand_id and (brand_id, puffs) in cost_keys_bp
                elif match_level == 'brand_puffs_series':
                    matched = brand_id and (
                        (brand_id, series_id, puffs) in cost_keys_bps or
                        (brand_id, None, puffs) in cost_keys_bps
                    )
                else:
                    matched = _match_cost_key(brand_id, series_id, puffs, flav, cost_keys_full)

                if not matched:
                    if match_level == 'brand_puffs':
                        key = (brand_id, puffs)
                    elif match_level == 'brand_puffs_series':
                        key = (brand_id, series_id, puffs)
                    else:
                        key = name.upper().strip()

                    if key not in unmapped:
                        unmapped[key] = {
                            'raw_name': name,
                            'brand_id': brand_id,
                            'brand_name': brand_names.get(brand_id),
                            'series_id': series_id,
                            'series_name': series_names.get(series_id),
                            'puff_count': puffs,
                            'flavor': flav,
                            'total_qty': 0,
                            'total_revenue': 0,
                            'sources': set(),
                            'product_count': 0,
                            'raw_names': {}
                        }
                    unmapped[key]['total_qty'] += qty
                    unmapped[key]['total_revenue'] += item_total
                    unmapped[key]['sources'].add(source)
                    unmapped[key]['product_count'] += 1
                    rn_key = name.upper().strip()
                    if rn_key not in unmapped[key]['raw_names']:
                        unmapped[key]['raw_names'][rn_key] = {'name': name, 'qty': 0, 'revenue': 0}
                    unmapped[key]['raw_names'][rn_key]['qty'] += qty
                    unmapped[key]['raw_names'][rn_key]['revenue'] += item_total

        result = []
        for k, v in sorted(unmapped.items(), key=lambda x: -x[1]['total_revenue']):
            v['sources'] = list(v['sources'])
            v['raw_names'] = sorted(v['raw_names'].values(), key=lambda x: -x['qty'])
            result.append(v)
        return jsonify(result)
    finally:
        conn.close()


def _match_cost_key(brand_id, series_id, puff_count, flavor, cost_keys):
    """Match a product to a cost entry using priority fallback.

    Priority (most specific first):
    1. brand + series + puffs + flavor
    2. brand + series + puffs
    3. brand + puffs
    4. brand only
    """
    if not brand_id:
        return False
    if (brand_id, series_id, puff_count, flavor) in cost_keys:
        return True
    if (brand_id, series_id, puff_count, None) in cost_keys:
        return True
    if (brand_id, None, puff_count, None) in cost_keys:
        return True
    if (brand_id, None, None, None) in cost_keys:
        return True
    return False


def _find_cost_price(brand_id, series_id, puff_count, flavor, cost_lookup):
    """Find the best matching cost price using priority fallback.

    cost_lookup: dict keyed by (brand_id, series_id, puff_count, flavor) -> cost_price
    Returns cost_price or None.
    """
    if not brand_id:
        return None
    key = (brand_id, series_id, puff_count, flavor)
    if key in cost_lookup:
        return cost_lookup[key]
    key = (brand_id, series_id, puff_count, None)
    if key in cost_lookup:
        return cost_lookup[key]
    key = (brand_id, None, puff_count, None)
    if key in cost_lookup:
        return cost_lookup[key]
    key = (brand_id, None, None, None)
    if key in cost_lookup:
        return cost_lookup[key]
    return None


# ===== Sales Board Profit Settings API =====

@app.route('/api/sales-board/profit-settings', methods=['GET'])
@login_required
def get_profit_settings():
    """Get profit settings for a given month"""
    import datetime
    year_month = request.args.get('year_month')
    if not year_month:
        year_month = datetime.date.today().strftime('%Y-%m')

    conn = get_db_connection()
    try:
        row = conn.execute(
            'SELECT * FROM sales_board_profit_settings WHERE year_month = ?',
            (year_month,)
        ).fetchone()
        if row:
            return jsonify(dict(row))
        return jsonify({
            'year_month': year_month,
            'profit_mode': 'percentage',
            'profit_percentage': 50.0
        })
    finally:
        conn.close()


@app.route('/api/sales-board/profit-settings', methods=['POST'])
@login_required
@admin_required
def save_profit_settings():
    """Save profit settings for a given month"""
    data = request.get_json()
    year_month = data.get('year_month')
    profit_mode = data.get('profit_mode', 'percentage')
    profit_percentage = data.get('profit_percentage', 50.0)
    cp = data.get('country_percentages', {})
    country_percentages_json = json.dumps(cp) if cp else '{}'

    if not year_month:
        return jsonify({'error': '缺少月份参数'}), 400
    if profit_mode not in ('percentage', 'actual'):
        return jsonify({'error': '无效的利润模式'}), 400

    conn = get_db_connection()
    try:
        conn.execute('''
            INSERT INTO sales_board_profit_settings (year_month, profit_mode, profit_percentage, country_percentages)
            VALUES (?, ?, ?, ?)
            ON CONFLICT(year_month) DO UPDATE SET
                profit_mode = excluded.profit_mode,
                profit_percentage = excluded.profit_percentage,
                country_percentages = excluded.country_percentages,
                updated_at = CURRENT_TIMESTAMP
        ''', (year_month, profit_mode, profit_percentage, country_percentages_json))
        conn.commit()
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'error': str(e)}), 500
    finally:
        conn.close()


# ===== Warehouse Management API =====

@app.route('/api/warehouses', methods=['GET'])
@login_required
def get_warehouses():
    conn = get_db_connection()
    rows = conn.execute('SELECT * FROM warehouses ORDER BY country, name').fetchall()
    conn.close()
    return jsonify([dict(r) for r in rows])


@app.route('/api/warehouses', methods=['POST'])
@login_required
@admin_required
def create_warehouse():
    data = request.get_json()
    name = (data.get('name') or '').strip()
    code = (data.get('code') or '').strip()
    country = (data.get('country') or '').strip()
    default_currency = data.get('default_currency', 'PLN')
    notes = data.get('notes', '')
    if not name or not code or not country:
        return jsonify({'error': '名称、编码和国家为必填项'}), 400
    conn = get_db_connection()
    try:
        conn.execute('INSERT INTO warehouses (name, code, country, default_currency, notes) VALUES (?, ?, ?, ?, ?)',
                     (name, code, country, default_currency, notes))
        conn.commit()
        return jsonify({'success': True, 'id': conn.execute('SELECT last_insert_rowid()').fetchone()[0]})
    except Exception as e:
        if 'UNIQUE' in str(e):
            return jsonify({'error': '编码已存在'}), 400
        return jsonify({'error': str(e)}), 500
    finally:
        conn.close()


@app.route('/api/warehouses/<int:wid>', methods=['PUT'])
@login_required
@admin_required
def update_warehouse(wid):
    data = request.get_json()
    conn = get_db_connection()
    try:
        conn.execute('''UPDATE warehouses SET name=?, code=?, country=?, default_currency=?, notes=?, is_active=?,
                        updated_at=CURRENT_TIMESTAMP WHERE id=?''',
                     (data.get('name'), data.get('code'), data.get('country'),
                      data.get('default_currency', 'PLN'), data.get('notes', ''),
                      1 if data.get('is_active', True) else 0, wid))
        conn.commit()
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'error': str(e)}), 500
    finally:
        conn.close()


@app.route('/api/warehouses/<int:wid>', methods=['DELETE'])
@login_required
@admin_required
def delete_warehouse(wid):
    conn = get_db_connection()
    try:
        cost_count = conn.execute('SELECT COUNT(*) FROM product_costs WHERE warehouse_id=?', (wid,)).fetchone()[0]
        order_count = conn.execute('SELECT COUNT(*) FROM orders WHERE warehouse_id=?', (wid,)).fetchone()[0]
        if cost_count > 0 or order_count > 0:
            return jsonify({'error': f'无法删除：有 {cost_count} 条成本记录和 {order_count} 个订单关联此仓库'}), 400
        conn.execute('DELETE FROM warehouses WHERE id=?', (wid,))
        conn.commit()
        return jsonify({'success': True})
    finally:
        conn.close()


@app.route('/api/sales-board/unmapped', methods=['GET'])
@login_required
def get_sales_board_unmapped():
    """Return unmapped products for the sales board (actual cost mode)."""
    import datetime as _dt
    year_month = request.args.get('year_month') or _dt.date.today().strftime('%Y-%m')
    filter_country = request.args.get('country', '')

    conn = get_db_connection()
    try:
        cost_rows = conn.execute(
            'SELECT brand_id, series_id, puff_count, flavor, warehouse_id, cost_price, cost_currency FROM product_costs'
        ).fetchall()
        cost_lookup = {}
        for cr in cost_rows:
            key = (cr['brand_id'], cr['series_id'], cr['puff_count'], cr['flavor'], cr['warehouse_id'])
            cost_lookup[key] = {'price': cr['cost_price'], 'currency': cr['cost_currency']}

        pm_rows = conn.execute('SELECT raw_name, source, brand_id, series_id, puff_count, flavor FROM product_mappings').fetchall()
        product_mappings_cache = {}
        for pm in pm_rows:
            product_mappings_cache[(normalize_raw_name(pm['raw_name']), pm['source'])] = pm

        site_rows = conn.execute('SELECT url, country, manager FROM sites').fetchall()
        site_country_map = {s['url']: s['country'] or 'PL' for s in site_rows}
        site_manager_map = {s['url']: s['manager'] or '' for s in site_rows}

        # Country -> ordered list of warehouse_ids (smallest id first).
        # Used as a fallback when an order has no warehouse_id assigned.
        country_default_wh_ids = {}
        for w in conn.execute('SELECT id, country FROM warehouses ORDER BY country, id'):
            country_default_wh_ids.setdefault(w['country'], []).append(w['id'])

        # Per-country (brand_id, series_id, puff_count) → list of cost entries.
        # Mirrors cost-management's "品牌+口数+系列" match level: a product is
        # considered "mapped" if ANY cost row in the same country has matching
        # brand+series+puffs (or brand+puffs with series=NULL as a generic row).
        country_bps_costs = {}
        cost_with_country = conn.execute('''
            SELECT pc.brand_id, pc.series_id, pc.puff_count, pc.flavor, pc.cost_price, pc.cost_currency,
                   pc.warehouse_id, w.country
            FROM product_costs pc LEFT JOIN warehouses w ON pc.warehouse_id = w.id
        ''').fetchall()
        for cr in cost_with_country:
            ckey = (cr['country'], cr['brand_id'], cr['series_id'], cr['puff_count'])
            country_bps_costs.setdefault(ckey, []).append({
                'price': cr['cost_price'], 'currency': cr['cost_currency'],
                'warehouse_id': cr['warehouse_id'], 'flavor': cr['flavor'],
            })

        brands_rows = conn.execute('SELECT id, name, aliases FROM brands').fetchall()
        brands_cache = []
        brand_names = {}
        for row in brands_rows:
            brand_name = row['name']
            brand_names[row['id']] = brand_name
            try:
                aliases = json.loads(row['aliases']) if row['aliases'] else []
            except:
                aliases = []
            brands_cache.append({
                'id': row['id'], 'name': brand_name, 'aliases': aliases,
                'patterns': [brand_name.upper()] + [a.upper() for a in aliases]
            })

        orders = conn.execute(f'''
            SELECT id, line_items, source, currency, total, shipping_total, warehouse_id
            FROM orders
            WHERE strftime('%Y-%m', date_created) = ?
            AND {_revenue_status_cond()}
        ''', (year_month,)).fetchall()

        # Get exchange rates for the month
        from collections import defaultdict
        profit_row = conn.execute(
            'SELECT profit_mode, profit_percentage FROM sales_board_profit_settings WHERE year_month = ?',
            (year_month,)
        ).fetchone()
        _board_overrides = {}
        override_rows = conn.execute(
            'SELECT currency, rate_to_cny FROM sales_board_exchange_rates WHERE year_month = ?',
            (year_month,)
        ).fetchall()
        for ov in override_rows:
            _board_overrides[ov['currency']] = ov['rate_to_cny']

        rate_cache = {}
        def _rate_for(cur):
            if cur not in rate_cache:
                r, _ = _get_board_cny_rate(cur, year_month, _board_overrides)
                rate_cache[cur] = r or 0
            return rate_cache[cur]

        unmapped = {}
        parsed_cache = {}
        for order in orders:
            items = parse_json_field(order['line_items'])
            if not isinstance(items, list):
                continue
            if filter_country and site_country_map.get(order['source'], 'PL') != filter_country:
                continue
            source = order['source']
            currency = order['currency'] or 'PLN'
            country = site_country_map.get(source, 'PL')
            wh_id = order['warehouse_id']
            manager = site_manager_map.get(source, '')
            rate = _rate_for(currency)

            for item in items:
                name = item.get('name', '') or ''
                qty = item.get('quantity', 0) or 0
                item_total = float(item.get('total', 0) or 0)
                name_key = normalize_raw_name(name)

                pm = product_mappings_cache.get((name_key, source)) or product_mappings_cache.get((name_key, '')) or product_mappings_cache.get((name_key, None))
                if pm:
                    b_id = pm['brand_id']
                    s_id = pm['series_id']
                    p_cnt = pm['puff_count']
                    flav = pm['flavor']
                else:
                    if name not in parsed_cache:
                        parsed_cache[name] = parse_product_name(name, brands_cache)
                    parsed_item = parsed_cache[name]
                    b_id = None
                    if parsed_item.get('brand'):
                        for bc in brands_cache:
                            if bc['name'].upper() == parsed_item['brand'].upper():
                                b_id = bc['id']
                                break
                    s_id = None
                    p_cnt = parsed_item.get('puffs')
                    flav = parsed_item.get('flavor')

                # Defense in depth: if the order is missing warehouse_id we still
                # want to attempt cost matching using the country's default warehouse.
                # (Old data had a broken backfill that left ~38% of orders NULL.)
                effective_wh_ids = [wh_id] if wh_id else []
                if not effective_wh_ids:
                    effective_wh_ids = country_default_wh_ids.get(country, [])

                cost_entry = None
                if b_id and effective_wh_ids:
                    for ewh in effective_wh_ids:
                        for try_key in [
                            (b_id, s_id, p_cnt, flav, ewh),
                            (b_id, s_id, p_cnt, None, ewh),
                            (b_id, None, p_cnt, None, ewh),
                            (b_id, None, None, None, ewh),
                        ]:
                            if try_key in cost_lookup:
                                cost_entry = cost_lookup[try_key]
                                break
                        if cost_entry:
                            break

                # Brand+series+puffs cross-warehouse fallback to align with
                # cost-mgmt's "品牌+口数+系列" view: a cost row with matching
                # brand+series+puffs anywhere in this country (or generic
                # series=NULL) counts as mapped.
                if not cost_entry and b_id:
                    for try_skey in [(country, b_id, s_id, p_cnt),
                                     (country, b_id, None, p_cnt)]:
                        bps_candidates = country_bps_costs.get(try_skey, [])
                        if bps_candidates:
                            cost_entry = bps_candidates[0]
                            break

                if not cost_entry:
                    key = (brand_names.get(b_id, 'Unknown'), p_cnt, country)
                    if key not in unmapped:
                        unmapped[key] = {
                            'brand': brand_names.get(b_id, 'Unknown'),
                            'puffs': p_cnt,
                            'country': country,
                            'qty': 0,
                            'revenue_cny': 0,
                            'products': {},
                            'managers': set(),
                        }
                    unmapped[key]['qty'] += qty
                    unmapped[key]['revenue_cny'] += item_total * rate
                    unmapped[key]['managers'].add(manager)
                    pname_key = name.upper().strip()
                    if pname_key not in unmapped[key]['products']:
                        unmapped[key]['products'][pname_key] = {'name': name, 'qty': 0, 'revenue_cny': 0}
                    unmapped[key]['products'][pname_key]['qty'] += qty
                    unmapped[key]['products'][pname_key]['revenue_cny'] += item_total * rate

        result = []
        for v in sorted(unmapped.values(), key=lambda x: -x['revenue_cny']):
            v['revenue_cny'] = round(v['revenue_cny'], 2)
            v['managers'] = list(v['managers'])
            prods = sorted(v['products'].values(), key=lambda x: -x['qty'])
            for p in prods:
                p['revenue_cny'] = round(p['revenue_cny'], 2)
            v['products'] = prods
            result.append(v)
        return jsonify(result)
    finally:
        conn.close()


@app.route('/api/order/<order_id>/warehouse', methods=['PUT'])
@login_required
@admin_required
def set_order_warehouse(order_id):
    data = request.get_json()
    warehouse_id = data.get('warehouse_id')
    conn = get_db_connection()
    try:
        order = conn.execute('SELECT id, source FROM orders WHERE id=?', (order_id,)).fetchone()
        if not order:
            return jsonify({'error': '订单不存在'}), 404
        if warehouse_id:
            wh = conn.execute('SELECT id, name, country FROM warehouses WHERE id=?', (warehouse_id,)).fetchone()
            if not wh:
                return jsonify({'error': '仓库不存在'}), 404
            # Country-consistency guard: an order must ship from a warehouse in its
            # own site's country. Cross-country assignment (e.g. a PL order → 澳洲仓库)
            # is always a mis-click — it silently breaks cost lookup and puts the
            # order under the wrong warehouse in shipping/inventory views.
            site = conn.execute('SELECT country FROM sites WHERE url=?', (order['source'],)).fetchone()
            site_country = site['country'] if site else None
            if site_country and wh['country'] and wh['country'] != site_country:
                return jsonify({'error': f'订单来自 {site_country} 站点，不能指派到 {wh["country"]} 仓库「{wh["name"]}」。请选择 {site_country} 的仓库。'}), 400
        conn.execute('UPDATE orders SET warehouse_id=? WHERE id=?', (warehouse_id, order_id))
        conn.commit()
        return jsonify({'success': True})
    finally:
        conn.close()


# ─────────────────────── 进销存(库存)模块 ───────────────────────
# 库存功能拆到独立的 inv_*.py 模块(Blueprint),避免继续膨胀 app.py。
# 只读关联现有表,全部走 inv_migrations.py 的可回滚迁移建表。
try:
    from inv_warehouses import inv_wh_bp
    app.register_blueprint(inv_wh_bp)
    from inv_skus import inv_sku_bp
    app.register_blueprint(inv_sku_bp)
    from inv_inventory import inv_inv_bp
    app.register_blueprint(inv_inv_bp)
    from inv_batches import inv_batch_bp
    app.register_blueprint(inv_batch_bp)
    from inv_orders import inv_ord_bp
    app.register_blueprint(inv_ord_bp)
    from inv_push import inv_push_bp
    app.register_blueprint(inv_push_bp)
    from inv_notify import inv_notify_bp
    app.register_blueprint(inv_notify_bp)
    from inv_reports import inv_report_bp
    app.register_blueprint(inv_report_bp)
except Exception as _e:
    import logging as _logging
    _logging.getLogger(__name__).warning('库存模块未加载: %s', _e)


@app.context_processor
def inject_inventory_perms():
    """把库存权限标志暴露给所有模板(导航栏与库存页都用)。

    inv_can_view: 能否看到库存菜单/页面;
    is_inv_admin: 是否可管理仓库主数据(超管)。
    """
    try:
        from inv_common import can_manage_inventory, can_view_inventory_any
        is_admin = (getattr(current_user, 'username', None) == 'admin'
                    or (getattr(current_user, 'is_authenticated', False)
                        and current_user.is_admin()))
        return {
            'inv_can_view': bool(can_view_inventory_any()),
            'can_manage_inv': bool(is_admin or can_manage_inventory()),
            'is_inv_admin': bool(is_admin),
        }
    except Exception:
        return {'inv_can_view': False, 'can_manage_inv': False, 'is_inv_admin': False}


if __name__ == '__main__':
    app.run(debug=True, host='0.0.0.0', port=5000)
