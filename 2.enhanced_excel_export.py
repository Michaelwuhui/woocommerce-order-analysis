import sqlite3
import json
import pandas as pd
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

def export_enhanced_excel():
    """导出包含产品数量统计的增强版Excel文件"""
    
    print("正在连接数据库并提取数据...")
    
    # 连接数据库
    conn = sqlite3.connect('woocommerce_orders.db')
    
    # 获取所有订单数据
    query = """
    SELECT id, number, status, date_created, total, shipping_total, source, 
           customer_id, payment_method, line_items, billing, shipping
    FROM orders 
    ORDER BY date_created DESC
    """
    
    df = pd.read_sql_query(query, conn)
    conn.close()
    
    print(f"获取到 {len(df)} 条订单数据")
    
    # 添加产品数量统计列
    print("正在分析产品数量...")
    
    product_types_count = []
    total_quantity = []
    product_names_list = []
    product_skus_list = []
    
    for index, row in df.iterrows():
        line_items = row['line_items']
        
        if pd.isna(line_items) or line_items == '':
            product_types_count.append(0)
            total_quantity.append(0)
            product_names_list.append('')
            product_skus_list.append('')
            continue
            
        try:
            items = json.loads(line_items)
            
            # 计算产品种类数和总数量
            types_count = len(items)
            total_qty = sum(item.get('quantity', 0) for item in items)
            
            # 提取产品名称和SKU
            names = [item.get('name', '未知产品') for item in items]
            skus = [item.get('sku', '无SKU') for item in items]
            
            product_types_count.append(types_count)
            total_quantity.append(total_qty)
            product_names_list.append(' | '.join(names))
            product_skus_list.append(' | '.join(skus))
            
        except json.JSONDecodeError:
            product_types_count.append(0)
            total_quantity.append(0)
            product_names_list.append('JSON解析错误')
            product_skus_list.append('JSON解析错误')
    
    # 添加新列到DataFrame
    df['产品种类数'] = product_types_count
    df['产品数量'] = total_quantity
    df['产品名称'] = product_names_list
    df['产品SKU'] = product_skus_list
    
    # 解析billing和shipping信息
    print("正在解析客户信息...")
    
    billing_names = []
    billing_emails = []
    billing_phones = []
    shipping_addresses = []
    
    for index, row in df.iterrows():
        # 解析billing信息
        billing = row['billing']
        if pd.isna(billing) or billing == '':
            billing_names.append('')
            billing_emails.append('')
            billing_phones.append('')
        else:
            try:
                billing_data = json.loads(billing)
                billing_names.append(f"{billing_data.get('first_name', '')} {billing_data.get('last_name', '')}".strip())
                billing_emails.append(billing_data.get('email', ''))
                billing_phones.append(billing_data.get('phone', ''))
            except json.JSONDecodeError:
                billing_names.append('解析错误')
                billing_emails.append('解析错误')
                billing_phones.append('解析错误')
        
        # 解析shipping信息
        shipping = row['shipping']
        if pd.isna(shipping) or shipping == '':
            shipping_addresses.append('')
        else:
            try:
                shipping_data = json.loads(shipping)
                address_parts = [
                    shipping_data.get('address_1', ''),
                    shipping_data.get('address_2', ''),
                    shipping_data.get('city', ''),
                    shipping_data.get('postcode', ''),
                    shipping_data.get('country', '')
                ]
                shipping_addresses.append(', '.join([part for part in address_parts if part]))
            except json.JSONDecodeError:
                shipping_addresses.append('解析错误')
    
    # 添加客户信息列
    df['客户姓名'] = billing_names
    df['客户邮箱'] = billing_emails
    df['客户电话'] = billing_phones
    df['配送地址'] = shipping_addresses
    
    # 重新排列列的顺序
    columns_order = [
        'id', 'number', 'status', 'date_created', 'source',
        '产品种类数', '产品数量', 'total', 'shipping_total',
        '客户姓名', '客户邮箱', '客户电话', '配送地址',
        'customer_id', 'payment_method',
        '产品名称', '产品SKU'
    ]
    
    df = df[columns_order]
    
    # 重命名列
    df.rename(columns={
        'id': '订单ID',
        'number': '订单号',
        'status': '订单状态',
        'date_created': '创建日期',
        'source': '来源网站',
        'total': '订单总额',
        'shipping_total': '运费金额',
        'customer_id': '客户ID',
        'payment_method': '支付方式'
    }, inplace=True)
    
    # 生成文件名
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    filename = f'enhanced_orders_export_{timestamp}.xlsx'
    
    print(f"正在导出到Excel文件: {filename}")
    
    # 创建Excel文件
    with pd.ExcelWriter(filename, engine='openpyxl') as writer:
        # 导出主数据表
        df.to_excel(writer, sheet_name='订单详情', index=False)
        
        # 创建统计汇总表
        # create_summary_sheet(writer, df)
        
        # 创建月度统计表
        create_monthly_stats_sheet(writer, df)
        
        # 格式化主数据表
        format_main_sheet(writer.sheets['订单详情'], df)
    
    print(f"Excel文件已成功导出: {filename}")
    
    # 显示统计信息
    print("\n导出统计:")
    print(f"总订单数: {len(df)}")
    print(f"总产品销售数量: {df['产品数量'].sum()} 支")
    print(f"平均每订单产品数量: {df['产品数量'].mean():.2f} 支")
    
    # 按来源网站统计
    source_stats = df.groupby('来源网站').agg({
        '订单ID': 'count',
        '产品数量': 'sum',
        '订单总额': 'sum'
    }).rename(columns={'订单ID': '订单数'})
    
    print("\n按网站统计:")
    for source, stats in source_stats.iterrows():
        print(f"  {source}: {stats['订单数']} 订单, {stats['产品数量']} 支产品, 总额: {stats['订单总额']:.2f}")
    
    return filename

def create_summary_sheet(writer, df):
    """创建统计汇总表"""
    
    # 按来源网站统计
    source_stats = df.groupby('来源网站').agg({
        '订单ID': 'count',
        '产品数量': 'sum',
        '产品种类数': 'sum',
        '订单总额': 'sum'
    }).rename(columns={'订单ID': '订单数'})
    
    # 按订单状态统计
    status_stats = df.groupby('订单状态').agg({
        '订单ID': 'count',
        '产品数量': 'sum',
        '订单总额': 'sum'
    }).rename(columns={'订单ID': '订单数'})
    
    # 创建汇总数据
    summary_data = []
    
    # 总体统计
    summary_data.append(['总体统计', '', '', '', ''])
    summary_data.append(['总订单数', len(df), '', '', ''])
    summary_data.append(['总产品销售数量', df['产品数量'].sum(), '支', '', ''])
    summary_data.append(['总销售额', df['订单总额'].sum(), '', '', ''])
    summary_data.append(['平均订单金额', df['订单总额'].mean(), '', '', ''])
    summary_data.append(['平均每订单产品数量', df['产品数量'].mean(), '支', '', ''])
    summary_data.append(['', '', '', '', ''])
    
    # 按网站统计
    summary_data.append(['按网站统计', '订单数', '产品数量', '产品种类数', '销售额'])
    for source, stats in source_stats.iterrows():
        summary_data.append([source, stats['订单数'], stats['产品数量'], stats['产品种类数'], stats['订单总额']])
    
    summary_data.append(['', '', '', '', ''])
    
    # 按状态统计
    summary_data.append(['按订单状态统计', '订单数', '产品数量', '', '销售额'])
    for status, stats in status_stats.iterrows():
        summary_data.append([status, stats['订单数'], stats['产品数量'], '', stats['订单总额']])
    
    # 写入汇总表
    summary_df = pd.DataFrame(summary_data, columns=['项目', '数值1', '数值2', '数值3', '数值4'])
    summary_df.to_excel(writer, sheet_name='统计汇总', index=False)

def create_monthly_stats_sheet(writer, df):
    """创建月度统计表"""

    df['月份'] = pd.to_datetime(df['创建日期']).dt.to_period('M')

    rows = []
    for (month, source), gdf in df.groupby(['月份', '来源网站']):
        total_orders = len(gdf)
        total_products = gdf['产品数量'].sum()
        total_amount = gdf['订单总额'].sum()

        success_mask = ~gdf['订单状态'].isin(['failed', 'cancelled'])
        success_orders = int(success_mask.sum())
        success_amount = gdf.loc[success_mask, '订单总额'].sum()
        success_products = gdf.loc[success_mask, '产品数量'].sum()
        success_shipping_amount = gdf.loc[success_mask, '运费金额'].sum() if '运费金额' in gdf.columns else 0
        success_net_amount = success_amount - success_shipping_amount

        failed_mask = gdf['订单状态'] == 'failed'
        failed_orders = int(failed_mask.sum())
        failed_amount = gdf.loc[failed_mask, '订单总额'].sum()
        failed_products = gdf.loc[failed_mask, '产品数量'].sum()

        cancelled_mask = gdf['订单状态'] == 'cancelled'
        cancelled_orders = int(cancelled_mask.sum())
        cancelled_amount = gdf.loc[cancelled_mask, '订单总额'].sum()
        cancelled_products = gdf.loc[cancelled_mask, '产品数量'].sum()

        completion = round((total_products / 2000) * 100, 2)

        rows.append({
            '月份': str(month),
            '来源网站': source,
            '总订单数': total_orders,
            '总产品数量': total_products,
            '总订单总额': total_amount,
            '目标完成度(%)': completion,
            '失败订单总数': failed_orders,
            '失败产品数量': failed_products,
            '失败订单总金额': failed_amount,
            '取消订单总数': cancelled_orders,
            '取消产品数量': cancelled_products,
            '取消订单总金额': cancelled_amount,
            '成功订单数': success_orders,
            '成功产品数量': success_products,
            '成功销售金额': success_amount,
            '扣除运费成功销售金额': success_net_amount,
        })

    monthly_stats = pd.DataFrame(rows)[[
        '月份', '来源网站',
        '总订单数', '总产品数量', '总订单总额', '目标完成度(%)',
        '失败订单总数', '失败产品数量', '失败订单总金额',
        '取消订单总数', '取消产品数量', '取消订单总金额',
        '成功订单数', '成功产品数量', '成功销售金额', '扣除运费成功销售金额'
    ]]
    monthly_stats.to_excel(writer, sheet_name='月度统计', index=False)

    # 格式化月度统计表头为蓝色背景（与订单详情一致）
    ws = writer.sheets['月度统计']
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.freeze_panes = 'A2'
    def _text_width(s):
        if s is None:
            return 0
        t = str(s)
        w = 0
        for ch in t:
            w += 2 if ord(ch) > 127 else 1
        return w
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(horizontal='center', vertical='center')
    for column in ws.columns:
        max_w = 0
        col_letter = column[0].column_letter
        for cell in column:
            try:
                tw = _text_width(cell.value)
                if tw > max_w:
                    max_w = tw
            except:
                pass
        adjusted_width = min(max_w + 4, 100)
        ws.column_dimensions[col_letter].width = adjusted_width

def format_main_sheet(worksheet, df):
    """格式化主数据表"""
    
    # 设置标题行样式
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    
    for cell in worksheet[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # 自动调整列宽
    def _text_width(s):
        if s is None:
            return 0
        t = str(s)
        w = 0
        for ch in t:
            w += 2 if ord(ch) > 127 else 1
        return w
    for row in worksheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(horizontal='center', vertical='center')
    for column in worksheet.columns:
        max_w = 0
        col_letter = column[0].column_letter
        for cell in column:
            try:
                tw = _text_width(cell.value)
                if tw > max_w:
                    max_w = tw
            except:
                pass
        adjusted_width = min(max_w + 4, 100)
        worksheet.column_dimensions[col_letter].width = adjusted_width
    
    # 冻结首行
    worksheet.freeze_panes = 'A2'

if __name__ == "__main__":
    filename = export_enhanced_excel()
    print(f"\n增强版Excel导出完成: {filename}")
