import sqlite3
import json
import pandas as pd
from datetime import datetime, timedelta
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import BarChart, Reference
import matplotlib.pyplot as plt
import seaborn as sns

def create_monthly_sales_report():
    """创建月度产品销量统计报告，跟踪2000支产品销售目标"""
    
    print("正在生成月度产品销量统计报告...")
    
    # 连接数据库
    conn = sqlite3.connect('woocommerce_orders.db')
    
    # 获取订单数据
    query = """
    SELECT id, date_created, source, line_items, total, status
    FROM orders 
    WHERE line_items IS NOT NULL
    ORDER BY date_created
    """
    
    df = pd.read_sql_query(query, conn)
    conn.close()
    
    print(f"获取到 {len(df)} 条订单数据")
    
    # 解析产品数量
    print("正在解析产品数量...")
    
    product_quantities = []
    for index, row in df.iterrows():
        try:
            items = json.loads(row['line_items'])
            total_qty = sum(item.get('quantity', 0) for item in items)
            product_quantities.append(total_qty)
        except json.JSONDecodeError:
            product_quantities.append(0)
    
    df['产品数量'] = product_quantities
    
    # 转换日期格式
    df['date_created'] = pd.to_datetime(df['date_created'])
    df['年月'] = df['date_created'].dt.to_period('M')
    df['年'] = df['date_created'].dt.year
    df['月'] = df['date_created'].dt.month
    
    # 月度统计
    monthly_stats = df.groupby(['年月', 'source']).agg({
        'id': 'count',
        '产品数量': 'sum',
        'total': 'sum'
    }).rename(columns={'id': '订单数', 'total': '销售额'}).reset_index()
    
    # 总体月度统计
    monthly_total = df.groupby('年月').agg({
        'id': 'count',
        '产品数量': 'sum',
        'total': 'sum'
    }).rename(columns={'id': '订单数', 'total': '销售额'}).reset_index()
    
    # 添加目标完成度
    monthly_total['目标完成度(%)'] = (monthly_total['产品数量'] / 2000 * 100).round(2)
    monthly_total['距离目标'] = 2000 - monthly_total['产品数量']
    monthly_total['是否达标'] = monthly_total['产品数量'] >= 2000
    
    # 生成文件名
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    filename = f'monthly_sales_report_{timestamp}.xlsx'
    
    print(f"正在导出月度报告到: {filename}")
    
    # 创建Excel文件
    with pd.ExcelWriter(filename, engine='openpyxl') as writer:
        # 月度总体统计
        monthly_total.to_excel(writer, sheet_name='月度总体统计', index=False)
        
        # 月度分网站统计
        monthly_stats.to_excel(writer, sheet_name='月度分网站统计', index=False)
        
        # 创建目标跟踪表
        create_target_tracking_sheet(writer, monthly_total)
        
        # 创建详细分析表
        create_detailed_analysis_sheet(writer, df)
        
        # 格式化表格
        format_monthly_sheets(writer, monthly_total, monthly_stats)
    
    # 显示报告摘要
    print_report_summary(monthly_total, monthly_stats)
    
    # 生成图表
    create_charts(monthly_total, monthly_stats)
    
    return filename

def create_target_tracking_sheet(writer, monthly_total):
    """创建目标跟踪表"""
    
    # 创建目标跟踪数据
    target_data = []
    
    # 添加标题
    target_data.append(['月度销售目标跟踪报告', '', '', '', '', ''])
    target_data.append(['销售目标: 每月2000支电子烟产品', '', '', '', '', ''])
    target_data.append(['', '', '', '', '', ''])
    
    # 添加表头
    target_data.append(['月份', '实际销量', '目标销量', '完成度(%)', '差距', '状态'])
    
    # 添加数据
    for _, row in monthly_total.iterrows():
        status = '✅ 达标' if row['是否达标'] else '❌ 未达标'
        target_data.append([
            str(row['年月']),
            row['产品数量'],
            2000,
            row['目标完成度(%)'],
            row['距离目标'],
            status
        ])
    
    # 添加汇总信息
    total_sold = monthly_total['产品数量'].sum()
    avg_monthly = monthly_total['产品数量'].mean()
    months_count = len(monthly_total)
    target_data.append(['', '', '', '', '', ''])
    target_data.append(['汇总统计', '', '', '', '', ''])
    target_data.append(['总销量', total_sold, '', '', '', ''])
    target_data.append(['月均销量', f'{avg_monthly:.0f}', '', '', '', ''])
    target_data.append(['统计月数', months_count, '', '', '', ''])
    target_data.append(['达标月数', sum(monthly_total['是否达标']), '', '', '', ''])
    target_data.append(['达标率', f'{sum(monthly_total["是否达标"])/months_count*100:.1f}%', '', '', '', ''])
    
    # 写入Excel
    target_df = pd.DataFrame(target_data)
    target_df.to_excel(writer, sheet_name='目标跟踪', index=False, header=False)

def create_detailed_analysis_sheet(writer, df):
    """创建详细分析表"""
    
    # 按日统计
    daily_stats = df.groupby([df['date_created'].dt.date, 'source']).agg({
        'id': 'count',
        '产品数量': 'sum',
        'total': 'sum'
    }).rename(columns={'id': '订单数', 'total': '销售额'}).reset_index()
    
    daily_stats.to_excel(writer, sheet_name='每日详细统计', index=False)
    
    # 按周统计
    df['周'] = df['date_created'].dt.to_period('W')
    weekly_stats = df.groupby(['周', 'source']).agg({
        'id': 'count',
        '产品数量': 'sum',
        'total': 'sum'
    }).rename(columns={'id': '订单数', 'total': '销售额'}).reset_index()
    
    weekly_stats.to_excel(writer, sheet_name='每周统计', index=False)

def format_monthly_sheets(writer, monthly_total, monthly_stats):
    """格式化月度统计表格"""
    
    # 格式化月度总体统计表
    ws1 = writer.sheets['月度总体统计']
    format_sheet_header(ws1)
    
    # 为达标状态添加颜色
    for row_num in range(2, len(monthly_total) + 2):
        is_达标 = ws1.cell(row=row_num, column=6).value  # 是否达标列
        if is_达标:
            # 绿色背景表示达标
            fill = PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')
        else:
            # 红色背景表示未达标
            fill = PatternFill(start_color='FFB6C1', end_color='FFB6C1', fill_type='solid')
        
        for col in range(1, 7):
            ws1.cell(row=row_num, column=col).fill = fill
    
    # 格式化月度分网站统计表
    ws2 = writer.sheets['月度分网站统计']
    format_sheet_header(ws2)
    
    # 格式化目标跟踪表
    ws3 = writer.sheets['目标跟踪']
    
    # 设置标题样式
    title_font = Font(size=16, bold=True, color='000080')
    ws3['A1'].font = title_font
    ws3['A2'].font = Font(size=12, bold=True, color='008000')
    
    # 合并标题单元格
    ws3.merge_cells('A1:F1')
    ws3.merge_cells('A2:F2')

def format_sheet_header(worksheet):
    """格式化表格标题行"""
    
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    
    for cell in worksheet[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # 自动调整列宽
    for column in worksheet.columns:
        max_length = 0
        column_letter = column[0].column_letter
        
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        
        adjusted_width = min(max_length + 2, 30)
        worksheet.column_dimensions[column_letter].width = adjusted_width

def print_report_summary(monthly_total, monthly_stats):
    """打印报告摘要"""
    
    print("\n" + "="*60)
    print("月度销售目标跟踪报告摘要")
    print("="*60)
    
    total_sold = monthly_total['产品数量'].sum()
    avg_monthly = monthly_total['产品数量'].mean()
    达标_months = sum(monthly_total['是否达标'])
    total_months = len(monthly_total)
    
    print(f"📊 总体统计:")
    print(f"   总销量: {total_sold} 支")
    print(f"   月均销量: {avg_monthly:.0f} 支")
    print(f"   统计月数: {total_months} 个月")
    print(f"   达标月数: {达标_months} 个月")
    print(f"   达标率: {达标_months/total_months*100:.1f}%")
    
    print(f"\n🎯 目标分析:")
    print(f"   月度目标: 2000 支")
    if avg_monthly >= 2000:
        print(f"   ✅ 平均销量已达标")
    else:
        print(f"   ❌ 平均销量未达标，需提升 {2000-avg_monthly:.0f} 支")
    
    print(f"\n📈 月度表现:")
    for _, row in monthly_total.iterrows():
        status = "✅" if row['是否达标'] else "❌"
        print(f"   {row['年月']}: {row['产品数量']} 支 ({row['目标完成度(%)']}%) {status}")
    
    print(f"\n🏪 网站表现:")
    site_totals = monthly_stats.groupby('source')['产品数量'].sum()
    for site, total in site_totals.items():
        print(f"   {site}: {total} 支")

def create_charts(monthly_total, monthly_stats):
    """生成销量图表"""
    
    try:
        # 设置中文字体
        plt.rcParams['font.sans-serif'] = ['SimHei', 'Microsoft YaHei']
        plt.rcParams['axes.unicode_minus'] = False
        
        # 创建图表
        fig, ((ax1, ax2), (ax3, ax4)) = plt.subplots(2, 2, figsize=(15, 12))
        
        # 1. 月度销量趋势图
        months = [str(m) for m in monthly_total['年月']]
        quantities = monthly_total['产品数量']
        
        ax1.plot(months, quantities, marker='o', linewidth=2, markersize=8)
        ax1.axhline(y=2000, color='r', linestyle='--', label='目标线 (2000支)')
        ax1.set_title('月度销量趋势', fontsize=14, fontweight='bold')
        ax1.set_xlabel('月份')
        ax1.set_ylabel('销量 (支)')
        ax1.legend()
        ax1.grid(True, alpha=0.3)
        plt.setp(ax1.xaxis.get_majorticklabels(), rotation=45)
        
        # 2. 目标完成度
        colors = ['green' if x >= 100 else 'red' for x in monthly_total['目标完成度(%)']]
        ax2.bar(months, monthly_total['目标完成度(%)'], color=colors, alpha=0.7)
        ax2.axhline(y=100, color='black', linestyle='-', label='达标线 (100%)')
        ax2.set_title('月度目标完成度', fontsize=14, fontweight='bold')
        ax2.set_xlabel('月份')
        ax2.set_ylabel('完成度 (%)')
        ax2.legend()
        plt.setp(ax2.xaxis.get_majorticklabels(), rotation=45)
        
        # 3. 网站销量对比
        site_data = monthly_stats.groupby('source')['产品数量'].sum()
        ax3.pie(site_data.values, labels=site_data.index, autopct='%1.1f%%', startangle=90)
        ax3.set_title('各网站销量占比', fontsize=14, fontweight='bold')
        
        # 4. 月度订单数vs销量
        ax4_twin = ax4.twinx()
        
        bars1 = ax4.bar([str(m) for m in monthly_total['年月']], monthly_total['订单数'], 
                       alpha=0.7, color='skyblue', label='订单数')
        line1 = ax4_twin.plot([str(m) for m in monthly_total['年月']], monthly_total['产品数量'], 
                             color='red', marker='o', linewidth=2, label='销量')
        
        ax4.set_title('月度订单数与销量关系', fontsize=14, fontweight='bold')
        ax4.set_xlabel('月份')
        ax4.set_ylabel('订单数', color='blue')
        ax4_twin.set_ylabel('销量 (支)', color='red')
        
        # 合并图例
        lines1, labels1 = ax4.get_legend_handles_labels()
        lines2, labels2 = ax4_twin.get_legend_handles_labels()
        ax4.legend(lines1 + lines2, labels1 + labels2, loc='upper left')
        
        plt.setp(ax4.xaxis.get_majorticklabels(), rotation=45)
        
        plt.tight_layout()
        
        # 保存图表
        chart_filename = f'sales_charts_{datetime.now().strftime("%Y%m%d_%H%M%S")}.png'
        plt.savefig(chart_filename, dpi=300, bbox_inches='tight')
        print(f"\n📊 销量图表已保存: {chart_filename}")
        
        plt.close()
        
    except Exception as e:
        print(f"图表生成失败: {e}")

if __name__ == "__main__":
    filename = create_monthly_sales_report()
    print(f"\n✅ 月度销售报告已完成: {filename}")