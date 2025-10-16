import sqlite3
import json
import pandas as pd
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

def create_improved_sales_analysis():
    """创建改进的销售分析报告，按照用户指定的字段要求"""
    
    print("正在生成改进的销售分析报告...")
    
    # 连接数据库
    conn = sqlite3.connect('woocommerce_orders.db')
    
    # 获取订单数据
    query = """
    SELECT id, date_created, source, line_items, total, status, 
           payment_method, shipping_lines
    FROM orders 
    WHERE line_items IS NOT NULL
    ORDER BY date_created
    """
    
    df = pd.read_sql_query(query, conn)
    conn.close()
    
    print(f"获取到 {len(df)} 条订单数据")
    
    # 解析产品数量
    print("正在解析产品数量...")
    
    product_quantities = []
    
    for index, row in df.iterrows():
        try:
            items = json.loads(row['line_items'])
            total_qty = sum(item.get('quantity', 0) for item in items)
            product_quantities.append(total_qty)
        except json.JSONDecodeError:
            product_quantities.append(0)
    
    df['产品数量'] = product_quantities
    
    # 转换日期格式
    df['date_created'] = pd.to_datetime(df['date_created'])
    df['年月'] = df['date_created'].dt.to_period('M')
    
    # 按月份和网站分组统计
    print("正在生成按月份和网站的统计数据...")
    
    # 创建用户指定格式的数据
    analysis_data = []
    
    # 获取所有月份
    months = sorted(df['年月'].unique())
    
    for month in months:
        month_data = df[df['年月'] == month]
        
        # 按网站分组
        for source in ['https://www.strefajednorazowek.pl', 'https://www.buchmistrz.pl']:
            site_data = month_data[month_data['source'] == source]
            
            if len(site_data) == 0:
                continue
                
            # 网站名称简化
            site_name = 'strefajednorazowek.pl' if 'strefajednorazowek' in source else 'buchmistrz.pl'
            
            # 总体统计
            total_sales = site_data['total'].sum()
            total_orders = len(site_data)
            total_products = site_data['产品数量'].sum()
            
            # 按订单状态分类统计
            completed_data = site_data[site_data['status'] == 'completed']
            onhold_data = site_data[site_data['status'] == 'on-hold']
            processing_data = site_data[site_data['status'] == 'processing']
            cancelled_data = site_data[site_data['status'] == 'cancelled']
            failed_data = site_data[site_data['status'] == 'failed']
            
            # 成功签收统计 (completed状态)
            completed_sales = completed_data['total'].sum()
            completed_orders = len(completed_data)
            
            # 发货未签收统计 (on-hold状态)
            onhold_sales = onhold_data['total'].sum()
            onhold_orders = len(onhold_data)
            
            # 缺货统计 (processing状态)
            processing_sales = processing_data['total'].sum()
            processing_orders = len(processing_data)
            
            # 取消订单统计 (cancelled状态)
            cancelled_sales = cancelled_data['total'].sum()
            cancelled_orders = len(cancelled_data)
            
            # 失败订单统计 (failed状态)
            failed_sales = failed_data['total'].sum()
            failed_orders = len(failed_data)
            
            analysis_data.append({
                '月份': str(month),
                '网站': site_name,
                '总销售金额': total_sales,
                '总订单数量': total_orders,
                '总卖出支数': total_products,
                '成功签收销售额': completed_sales,
                '成功签收订单数量': completed_orders,
                '发货未签收金额': onhold_sales,
                '发货未签收订单数量': onhold_orders,
                '缺货订单金额': processing_sales,
                '缺货订单数量': processing_orders,
                '取消订单金额': cancelled_sales,
                '取消订单数量': cancelled_orders,
                '失败订单金额': failed_sales,
                '失败订单数量': failed_orders
            })
    
    # 创建DataFrame
    analysis_df = pd.DataFrame(analysis_data)
    
    # 生成Excel文件
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"improved_sales_analysis_{timestamp}.xlsx"
    
    print(f"正在生成Excel文件: {filename}")
    
    with pd.ExcelWriter(filename, engine='openpyxl') as writer:
        # 创建主要分析表
        create_main_analysis_sheet(writer, analysis_df)
        
        # 创建汇总表
        create_summary_sheet(writer, analysis_df)
        
        # 格式化工作表
        format_workbook(writer)
    
    # 打印分析摘要
    print_analysis_summary(analysis_df)
    
    return filename

def create_main_analysis_sheet(writer, analysis_df):
    """创建主要分析工作表"""
    
    # 写入数据到工作表
    analysis_df.to_excel(writer, sheet_name='销售分析报告', index=False)
    
    # 获取工作表对象
    worksheet = writer.sheets['销售分析报告']
    
    # 设置列宽
    column_widths = {
        'A': 12,  # 月份
        'B': 20,  # 网站
        'C': 15,  # 总销售金额
        'D': 15,  # 总订单数量
        'E': 15,  # 总卖出支数
        'F': 18,  # 成功签收销售额
        'G': 18,  # 成功签收订单数量
        'H': 18,  # 发货未签收金额
        'I': 18,  # 发货未签收订单数量
        'J': 15,  # 缺货订单金额
        'K': 15,  # 缺货订单数量
        'L': 15,  # 取消订单金额
        'M': 15,  # 取消订单数量
        'N': 15,  # 失败订单金额
        'O': 15   # 失败订单数量
    }
    
    for col, width in column_widths.items():
        worksheet.column_dimensions[col].width = width
    
    # 设置标题行样式
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    for col in range(1, len(analysis_df.columns) + 1):
        cell = worksheet.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # 设置数据行样式
    for row in range(2, len(analysis_df) + 2):
        for col in range(1, len(analysis_df.columns) + 1):
            cell = worksheet.cell(row=row, column=col)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # 金额列格式化
            if col in [3, 6, 8, 10, 12, 14]:  # 金额列
                if cell.value and isinstance(cell.value, (int, float)):
                    cell.number_format = '#,##0.00'

def create_summary_sheet(writer, analysis_df):
    """创建汇总工作表"""
    
    # 按网站汇总
    site_summary = analysis_df.groupby('网站').agg({
        '总销售金额': 'sum',
        '总订单数量': 'sum',
        '总卖出支数': 'sum',
        '成功签收销售额': 'sum',
        '成功签收订单数量': 'sum',
        '发货未签收金额': 'sum',
        '发货未签收订单数量': 'sum',
        '缺货订单金额': 'sum',
        '缺货订单数量': 'sum',
        '取消订单金额': 'sum',
        '取消订单数量': 'sum',
        '失败订单金额': 'sum',
        '失败订单数量': 'sum'
    }).reset_index()
    
    # 添加总计行
    total_row = {
        '网站': '总计',
        '总销售金额': site_summary['总销售金额'].sum(),
        '总订单数量': site_summary['总订单数量'].sum(),
        '总卖出支数': site_summary['总卖出支数'].sum(),
        '成功签收销售额': site_summary['成功签收销售额'].sum(),
        '成功签收订单数量': site_summary['成功签收订单数量'].sum(),
        '发货未签收金额': site_summary['发货未签收金额'].sum(),
        '发货未签收订单数量': site_summary['发货未签收订单数量'].sum(),
        '缺货订单金额': site_summary['缺货订单金额'].sum(),
        '缺货订单数量': site_summary['缺货订单数量'].sum(),
        '取消订单金额': site_summary['取消订单金额'].sum(),
        '取消订单数量': site_summary['取消订单数量'].sum(),
        '失败订单金额': site_summary['失败订单金额'].sum(),
        '失败订单数量': site_summary['失败订单数量'].sum()
    }
    
    site_summary = pd.concat([site_summary, pd.DataFrame([total_row])], ignore_index=True)
    
    # 写入汇总表
    site_summary.to_excel(writer, sheet_name='网站汇总', index=False)
    
    # 格式化汇总表
    worksheet = writer.sheets['网站汇总']
    
    # 设置列宽
    for col in range(1, len(site_summary.columns) + 1):
        worksheet.column_dimensions[chr(64 + col)].width = 18
    
    # 设置标题行样式
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    for col in range(1, len(site_summary.columns) + 1):
        cell = worksheet.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # 突出显示总计行
    total_row_num = len(site_summary) + 1
    total_fill = PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid")
    total_font = Font(bold=True)
    
    for col in range(1, len(site_summary.columns) + 1):
        cell = worksheet.cell(row=total_row_num, column=col)
        cell.fill = total_fill
        cell.font = total_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # 金额列格式化
        if col in [2, 5, 7, 9, 11, 13]:  # 金额列
            if cell.value and isinstance(cell.value, (int, float)):
                cell.number_format = '#,##0.00'

def format_workbook(writer):
    """格式化整个工作簿"""
    
    workbook = writer.book
    
    # 为所有工作表添加边框
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for sheet_name in workbook.sheetnames:
        worksheet = workbook[sheet_name]
        
        # 添加边框
        for row in worksheet.iter_rows():
            for cell in row:
                if cell.value is not None:
                    cell.border = thin_border

def print_analysis_summary(analysis_df):
    """打印分析摘要"""
    
    print("\n" + "="*60)
    print("📊 销售分析摘要")
    print("="*60)
    
    # 总体统计
    total_sales = analysis_df['总销售金额'].sum()
    total_orders = analysis_df['总订单数量'].sum()
    total_products = analysis_df['总卖出支数'].sum()
    
    print(f"💰 总销售金额: {total_sales:,.2f}")
    print(f"📦 总订单数量: {total_orders:,}")
    print(f"🎯 总卖出支数: {total_products:,}")
    
    print("\n📈 按订单状态分类:")
    completed_sales = analysis_df['成功签收销售额'].sum()
    completed_orders = analysis_df['成功签收订单数量'].sum()
    onhold_sales = analysis_df['发货未签收金额'].sum()
    onhold_orders = analysis_df['发货未签收订单数量'].sum()
    processing_sales = analysis_df['缺货订单金额'].sum()
    processing_orders = analysis_df['缺货订单数量'].sum()
    cancelled_sales = analysis_df['取消订单金额'].sum()
    cancelled_orders = analysis_df['取消订单数量'].sum()
    failed_sales = analysis_df['失败订单金额'].sum()
    failed_orders = analysis_df['失败订单数量'].sum()
    
    print(f"✅ 成功签收: {completed_sales:,.2f} ({completed_orders:,} 订单)")
    print(f"🚚 发货未签收: {onhold_sales:,.2f} ({onhold_orders:,} 订单)")
    print(f"⏳ 缺货: {processing_sales:,.2f} ({processing_orders:,} 订单)")
    print(f"❌ 取消订单: {cancelled_sales:,.2f} ({cancelled_orders:,} 订单)")
    print(f"💥 失败订单: {failed_sales:,.2f} ({failed_orders:,} 订单)")
    
    # 按网站统计
    print("\n🌐 按网站统计:")
    site_stats = analysis_df.groupby('网站').agg({
        '总销售金额': 'sum',
        '总订单数量': 'sum'
    })
    
    for site, stats in site_stats.iterrows():
        print(f"   {site}: {stats['总销售金额']:,.2f} ({stats['总订单数量']:,} 订单)")
    
    print("="*60)

if __name__ == "__main__":
    filename = create_improved_sales_analysis()
    print(f"\n✅ 改进的销售分析报告已完成: {filename}")