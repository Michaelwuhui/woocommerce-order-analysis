import os
import sys
import unittest

from openpyxl import load_workbook

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from shipping_export import (
    DPD_HEADERS,
    PENDING_HEADERS,
    build_australia_pending_workbook,
    build_australia_shipping_workbook,
    build_dpd_shipping_workbook,
    prepare_australia_pending_items,
)


class AustraliaShippingWorkbookTests(unittest.TestCase):
    def test_template_contains_no_customer_rows(self):
        template = os.path.join(
            os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
            'assets', 'export_templates', 'australia_shipping_template.xlsx',
        )
        workbook = load_workbook(template, data_only=False)
        worksheet = workbook['地址拆分']
        self.assertEqual(
            ['序号', '订单信息', '客户姓名', '电话', '城市', '州', '收货地址', '正式运输单号'],
            [worksheet.cell(1, column).value for column in range(1, 9)],
        )
        self.assertIn(worksheet['I1'].value, (None, ''))
        for row in worksheet.iter_rows(min_row=2, max_col=8):
            self.assertTrue(all(cell.value in (None, '') for cell in row))

    def test_values_are_text_safe_and_keep_leading_zeroes(self):
        rows = [{
            'order_number': '#TEST-001',
            'customer_name': 'Sample Customer',
            'phone': '0412345678',
            'city': 'Perth',
            'state': 'WA',
            'address': '=HYPERLINK("https://invalid.test")',
            'tracking_number': '00000000000000000002',
        }]
        output = build_australia_shipping_workbook(rows)
        workbook = load_workbook(output, data_only=False)
        worksheet = workbook['地址拆分']

        self.assertEqual('0412345678', worksheet['D2'].value)
        self.assertEqual('00000000000000000002', worksheet['H2'].value)
        self.assertEqual('=HYPERLINK("https://invalid.test")', worksheet['G2'].value)
        self.assertEqual(8, worksheet.max_column)
        for address in ('B2', 'C2', 'D2', 'E2', 'F2', 'G2', 'H2'):
            self.assertEqual('s', worksheet[address].data_type)
            self.assertEqual('@', worksheet[address].number_format)

        self.assertEqual('A1:H2', worksheet.auto_filter.ref)
        self.assertEqual('A2', worksheet.freeze_panes)

    def test_dpd_workbook_matches_partner_columns_and_types(self):
        rows = [{
            'customer_number': 'vsklep-00123',
            'recipient_name': 'Jan Kowalski',
            'province': 'Warszawa',
            'city': 'Warszawa',
            'address': '=HYPERLINK("https://invalid.test")',
            'phone': '0123456789',
            'email': 'jan@example.com',
            'postcode': '00-001',
            'cod_amount': 249.9,
        }]

        output = build_dpd_shipping_workbook(rows)
        workbook = load_workbook(output, data_only=False)
        worksheet = workbook['下单模板']

        self.assertEqual(DPD_HEADERS, [
            worksheet.cell(1, column).value for column in range(1, 24)
        ])
        self.assertEqual([
            'vsklep-00123', None, '2581', 'PL', 'Jan Kowalski',
            'Warszawa', 'Warszawa', '=HYPERLINK("https://invalid.test")',
            '0123456789', 'jan@example.com', '00-001', 1,
            'wanju', 'wanju', 'wanju', 1, 'usd', 1, 1,
            249.9, 'PLN', '是', 'wanju',
        ], [worksheet.cell(2, column).value for column in range(1, 24)])
        self.assertEqual('s', worksheet['A2'].data_type)
        self.assertEqual('s', worksheet['H2'].data_type)
        self.assertEqual('s', worksheet['I2'].data_type)
        self.assertEqual('s', worksheet['K2'].data_type)
        self.assertEqual('@', worksheet['K2'].number_format)
        self.assertEqual('0.00', worksheet['T2'].number_format)
        self.assertEqual('A1:W2', worksheet.auto_filter.ref)
        self.assertEqual('A2', worksheet.freeze_panes)

    def test_pending_items_split_product_flavor_and_group_products(self):
        items = prepare_australia_pending_items([
            {
                'name': 'Alibarbar INGOT 9000 - PEACH ICE',
                'parent_name': 'Alibarbar INGOT 9000',
                'quantity': 1,
                'meta_data': [{
                    'key': 'pa_flavor', 'value': 'peach-ice',
                    'display_value': 'PEACH ICE',
                }],
            },
            {
                'name': 'UMIN SLICK PLUS 10000 - TOBACCO',
                'quantity': '2',
                'meta_data': [{
                    'key': 'flavour', 'value': 'tobacco',
                    'display_value': 'TOBACCO',
                }],
            },
            {
                'name': 'Alibarbar INGOT 9000 - BLACKBERRY ICE',
                'parent_name': 'Alibarbar INGOT 9000',
                'quantity': 3,
                'meta_data': [{
                    'key': 'pa_flavour', 'value': 'blackberry-ice',
                    'display_value': 'BLACKBERRY ICE',
                }],
            },
        ])

        self.assertEqual(
            ['Alibarbar INGOT 9000', 'Alibarbar INGOT 9000', 'UMIN SLICK PLUS 10000'],
            [item['product'] for item in items],
        )
        self.assertEqual(['PEACH ICE', 'BLACKBERRY ICE', 'TOBACCO'], [
            item['flavor'] for item in items
        ])
        self.assertEqual([1, 3, 2], [item['quantity'] for item in items])

    def test_pending_workbook_uses_requested_headers_and_merges(self):
        rows = [{
            'order_date': '2026-07-31T08:30:00',
            'order_number': '#EOL74127',
            'items': [
                {'product': 'Alibarbar INGOT 9000', 'flavor': 'STRAWBERRY', 'quantity': 1},
                {'product': 'Alibarbar INGOT 9000', 'flavor': 'BLACKBERRY ICE', 'quantity': 2},
                {'product': 'UMIN SLICK PLUS 10000', 'flavor': 'TOBACCO', 'quantity': 1},
            ],
            'customer_name': 'Sample Customer',
            'phone': '0412345678',
            'state': 'WA',
            'city': 'Perth',
            'street_address': '=HYPERLINK("https://invalid.test")',
            'postcode': '06000',
        }]

        output = build_australia_pending_workbook(rows)
        workbook = load_workbook(output, data_only=False)
        worksheet = workbook['澳洲未发货']

        self.assertEqual(PENDING_HEADERS, [
            worksheet.cell(1, column).value for column in range(1, 13)
        ])
        self.assertEqual(12, worksheet.max_column)
        self.assertEqual(4, worksheet.max_row)
        self.assertEqual('EOL74127', worksheet['C2'].value)
        self.assertEqual('STRAWBERRY', worksheet['E2'].value)
        self.assertEqual('BLACKBERRY ICE', worksheet['E3'].value)
        self.assertEqual('TOBACCO', worksheet['E4'].value)
        self.assertEqual([1, 2, 1], [worksheet[f'F{row}'].value for row in range(2, 5)])
        self.assertEqual('0412345678', worksheet['H2'].value)
        self.assertEqual('06000', worksheet['L2'].value)
        self.assertEqual('=HYPERLINK("https://invalid.test")', worksheet['K2'].value)
        self.assertEqual('s', worksheet['K2'].data_type)
        self.assertEqual('@', worksheet['L2'].number_format)

        merged = {str(cell_range) for cell_range in worksheet.merged_cells.ranges}
        expected_order_merges = {
            f'{column}2:{column}4' for column in ('A', 'B', 'C', 'G', 'H', 'I', 'J', 'K', 'L')
        }
        self.assertTrue(expected_order_merges.issubset(merged))
        self.assertIn('D2:D3', merged)
        self.assertNotIn('E2:E4', merged)
        self.assertNotIn('F2:F4', merged)
        self.assertEqual('A2', worksheet.freeze_panes)


if __name__ == '__main__':
    unittest.main()
